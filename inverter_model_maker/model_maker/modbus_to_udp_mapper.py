# -*- coding: utf-8 -*-
"""
인버터 Modbus PDF → RTU UDP 프로토콜 레지스터맵 변환 GUI

다양한 인버터 제조사의 Modbus 프로토콜 PDF 파일을 읽어
RTU UDP 패킷 BODY 필드에 매핑되는 레지스터 목록을 자동 생성.
- 주기 데이터 (P): 인버터 실시간 데이터, MPPT, STRING
- 비주기 데이터 (A): H3 제어 (DER-AVM), IV Scan
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import csv
import re
import os
import threading

try:
    from model_maker import ai_generator as _ai_gen
except ImportError:
    try:
        import ai_generator as _ai_gen
    except ImportError:
        _ai_gen = None

try:
    from model_maker import stage_pipeline as _sp
except ImportError:
    try:
        import stage_pipeline as _sp
    except ImportError:
        _sp = None

# ──────────────────────────────────────────────────────────────────────
# PDF → Modbus Register 파서
# ──────────────────────────────────────────────────────────────────────

def parse_modbus_pdf(filepath):
    """
    Modbus 프로토콜 PDF를 읽어 레지스터 테이블을 추출한다.
    반환값: dict with keys:
      'manufacturer', 'version',
      'device_info', 'realtime_data', 'der_avm_realtime',
      'der_avm_params', 'inverter_params', 'iv_scan_regs',
      'all_registers'
    각 리스트 원소: {index, definition, address, regs, type, unit, rw, comment}
    """
    try:
        import fitz  # PyMuPDF
        _has_fitz = True
    except ImportError:
        _has_fitz = False

    # 텍스트 추출 (fitz 우선, 없으면 pdfminer fallback)
    full_text = ""
    pages_text = []
    if _has_fitz:
        doc = fitz.open(filepath)
        try:
            text_chunks = []
            for page in doc:
                txt = page.get_text()
                text_chunks.append(txt)
                pages_text.append(txt)
            full_text = "\n".join(text_chunks)
        finally:
            doc.close()
    else:
        # fitz 없을 때: pdfminer.six 시도
        try:
            from pdfminer.high_level import extract_text_by_page, extract_text
            full_text = extract_text(filepath)
            from pdfminer.high_level import extract_pages
            from pdfminer.layout import LTTextContainer
            for page_layout in extract_pages(filepath):
                page_parts = []
                for element in page_layout:
                    if isinstance(element, LTTextContainer):
                        page_parts.append(element.get_text())
                pages_text.append("".join(page_parts))
        except ImportError:
            pass
        except Exception as e:
            raise RuntimeError(f"PDF 파싱 오류: {e}") from e

    if not full_text:
        raise ImportError("PDF 텍스트 추출을 위해 PyMuPDF 또는 pdfminer.six가 필요합니다.\n"
                          "'pip install PyMuPDF' 또는 'pip install pdfminer.six' 를 실행하세요.")

    result = {
        'filepath': filepath,
        'filename': os.path.basename(filepath),
        'manufacturer': _detect_manufacturer(full_text),
        'version': _detect_version(full_text),
        'device_info': [],
        'realtime_data': [],
        'der_avm_realtime': [],
        'der_avm_params': [],
        'inverter_params': [],
        'iv_scan_regs': [],
        'all_registers': [],
    }

    # 제조사별 파서 선택
    manufacturer = result['manufacturer']
    if manufacturer == 'Kstar':
        all_regs = _parse_kstar_registers(pages_text, full_text)
    elif manufacturer == 'Huawei':
        if not _has_fitz:
            raise ImportError("Huawei PDF 파싱에는 PyMuPDF가 필요합니다.\n"
                              "'pip install PyMuPDF' 를 실행하세요.")
        all_regs = _parse_huawei_registers_from_tables(filepath)
    else:
        # Synergy(VerterKing) 및 기타
        all_regs = _parse_register_tables(pages_text, full_text)

    # Solarize PDF: 표준 주소 테이블로 누락 레지스터 보충
    # (PDF 파서가 high-word 등 일부 레지스터를 추출하지 못하는 경우 대비)
    if result['manufacturer'] == 'Solarize':
        all_regs = _fill_missing_solarize_regs(all_regs)

    result['all_registers'] = all_regs

    # 섹션 분류
    for reg in all_regs:
        section = reg.get('section', '')
        sl = section.lower()
        if 'device' in sl or 'system info' in sl:
            result['device_info'].append(reg)
        elif 'dea-avm' in sl or 'der-avm real' in sl or 'control monitor' in sl:
            result['der_avm_realtime'].append(reg)
        elif 'der-avm param' in sl or 'control' in sl or 'instruction' in sl:
            result['der_avm_params'].append(reg)
        elif 'iv' in sl and ('scan' in sl or 'curve' in sl):
            result['iv_scan_regs'].append(reg)
        elif 'param' in sl or 'setup' in sl or 'setting' in sl:
            result['inverter_params'].append(reg)
        else:
            result['realtime_data'].append(reg)

    return result


def _detect_manufacturer(text):
    """텍스트에서 제조사명 추정"""
    text_lower = text.lower()
    if 'senergy' in text_lower or 'solarize' in text_lower or 'verterking' in text_lower or 'apd' in text_lower:
        return 'Solarize'
    elif 'huawei' in text_lower or 'sun2000' in text_lower:
        return 'Huawei'
    elif 'kstar' in text_lower:
        return 'Kstar'
    elif 'sungrow' in text_lower:
        return 'Sungrow'
    elif 'ekos' in text_lower or 'modbus map-ek' in text_lower:
        return 'EKOS'
    elif 'growatt' in text_lower:
        return 'Growatt'
    elif 'goodwe' in text_lower:
        return 'GoodWe'
    elif 'sma' in text_lower:
        return 'SMA'
    elif 'fronius' in text_lower:
        return 'Fronius'
    return 'Unknown'


def _detect_version(text):
    """버전 번호 추출"""
    m = re.search(r'[Vv]ersion\s*[:\s]*(\d+\.\d+[\.\d]*)', text)
    if m:
        return m.group(1)
    m = re.search(r'(\d+\.\d+\.\d+)', text[:500])
    if m:
        return m.group(1)
    return 'Unknown'


def _parse_register_tables(pages_text, full_text):
    """
    PDF 텍스트에서 Modbus 레지스터 테이블을 파싱.
    테이블 패턴: Index/Definition/Address/Register Number/Type/Comment/R-W/Unit
    """
    registers = []
    current_section = ""

    # 섹션 헤더 패턴
    section_patterns = [
        (r'Device\s+information', 'Device Information'),
        (r'Inverter\s+real\s*time\s+data', 'Inverter Realtime Data'),
        (r'DEA-AVM\s+real\s*time\s+data', 'DEA-AVM Realtime Data'),
        (r'DER-AVM\s+parameters', 'DER-AVM Parameters'),
        (r'Inverter\s+parameters', 'Inverter Parameters'),
        (r'I-V\s+curve', 'IV Scan'),
        (r'IV\s+scan', 'IV Scan'),
        (r'REMS\s+register', 'REMS Register'),
        # Generic: "Register Address" section header (e.g., Goodwe "8. Register Address of Device")
        (r'Register\s+Address', 'Registers'),
    ]

    # 주소 패턴: 0xHHHH 또는 단순 십진수(큰 수)
    addr_hex_pattern = re.compile(r'0[xX]([0-9A-Fa-f]{2,6})')
    addr_dec_pattern = re.compile(r'^(\d{3,5})$')

    # 전체 텍스트를 줄 단위로 처리
    all_lines = full_text.split('\n')
    all_lines = [l.strip() for l in all_lines if l.strip()]

    # 헤더/푸터 필터
    skip_patterns = [
        r'^Project\s+No', r'^NA$', r'^MODBUS\s+PROTOCOL',
        r'^Security', r'^Confidential', r'^Project\s+Name',
        r'^Version$', r'^\d+\.\d+\.\d+$', r'^\d+\s*/\s*\d+$',
        r'^APD\s+CONFIDENTIAL', r'^Index$', r'^Definition$',
        r'^Address$', r'^Register$', r'^Number$', r'^Type$',
        r'^Comment$', r'^R/W$', r'^Unit$', r'^Note:', r'^Note\s+\d',
    ]
    skip_re = [re.compile(p, re.IGNORECASE) for p in skip_patterns]

    i = 0
    while i < len(all_lines):
        line = all_lines[i]

        # 섹션 탐지
        for pat, sec_name in section_patterns:
            if re.search(pat, line, re.IGNORECASE):
                current_section = sec_name
                break

        # 레지스터 항목 탐지: 숫자(index)로 시작하는 줄
        # Pattern 1: 1-3 digit index (Solarize/Senergy style)
        # Pattern 2: 3-5 digit address (Goodwe/generic style - address-first format)
        idx_match = re.match(r'^(\d{1,3})\s*$', line)
        addr_match = re.match(r'^(\d{3,5})\s*$', line) if not idx_match else None
        # For addr_match: verify next line looks like a register definition (not error table)
        if addr_match and i + 1 < len(all_lines):
            next_ln = all_lines[i + 1].strip()
            # Skip if next line is also a bare number (page number, bit value, etc.)
            if re.match(r'^\d{1,5}\s*$', next_ln) or re.match(r'^0[xX]', next_ln):
                addr_match = None
            # Skip hex bit values in error tables (0x0001, 0x0002, etc.)
            if re.match(r'^[0-9A-Fa-f]{8}$', next_ln):
                addr_match = None
        if (idx_match and current_section) or addr_match:
            if addr_match:
                index = int(addr_match.group(1))
                section = current_section or 'Registers'
            else:
                index = int(idx_match.group(1))
                section = current_section
            # 다음 줄들에서 definition, address, regs, type 등 수집
            reg = _collect_register_entry(all_lines, i, index, section)
            if reg and reg.get('address') is not None:
                defn = reg.get('definition', '').lower()
                addr_val = reg.get('address', 0)
                # Skip noise: range values as definitions, invalid addresses
                if re.match(r'^\[[\d,\s\.\-]+\]', defn):
                    i += 1; continue
                if addr_val > 65535 or addr_val < 0:
                    i += 1; continue
                # Skip empty definitions (likely page numbers or noise)
                if not defn or len(defn) < 2:
                    i += 1; continue
                _skip_defs = [
                    'function code', 'appendix', 'rems register',
                    'error code table', 'table of contents',
                    'unsigned integer of', 'signed integer of',
                    'read register address', 'write a single register',
                    'write multiple register', 'rems read register',
                    'register start address', 'register address',
                    'byte number', 'illegal code', '1~247',
                ]
                sect = reg.get('section', '').lower()
                _skip_sections = ['rems register']
                is_protocol_desc = ('der-avm' in sect and addr_val < 0x0100)
                if (not any(sd in defn for sd in _skip_defs)
                        and not any(ss in sect for ss in _skip_sections)
                        and not is_protocol_desc):
                    registers.append(reg)

        i += 1

    return registers


def _collect_register_entry(lines, start_idx, index, section):
    """
    Index 줄 이후의 연속 줄들에서 레지스터 정보를 수집.
    PDF 테이블은 각 셀이 별도 줄로 나오는 경우가 많다.
    Supports both index-first (Solarize) and address-first (Goodwe) formats.
    """
    reg = {
        'index': index,
        'section': section,
        'definition': '',
        'address': None,
        'address_hex': '',
        'regs': 1,
        'type': '',
        'unit': '',
        'rw': 'RO',
        'comment': '',
        'scale_factor': '',
    }

    _TYPE_SET = {'U16', 'S16', 'U32', 'S32', 'U8', 'S8', 'STR', 'STRING', 'FLOAT', 'FLOAT32'}
    _RW_SET = {'RO', 'RW', 'WO'}
    _UNIT_SET = {'V', 'A', 'W', 'kW', 'kWh', 'KWH', 'Wh', 'Hz', '%', 'Var', 'VA',
                 'NA', 'N/A', 'Hr', 's', 'degC', 'degree C', 'mA', '1mA'}
    _SKIP_RE = [
        re.compile(p, re.IGNORECASE) for p in [
            r'^Project\s+No', r'^MODBUS', r'^Security', r'^Confidential',
            r'^\d+\s*/\s*\d+$', r'^APD\s+CONF',
            r'^Index$', r'^Definition$', r'^Address$', r'^Register$',
            r'^Number$', r'^Type$', r'^Comment$', r'^R/W$', r'^Unit$',
            r'^#$', r'^SF\s+Gain', r'^Length', r'^Range$', r'^Note:?\s*$',
        ]]

    # Collect raw lines after the index/address line
    collected = []
    j = start_idx + 1
    limit = min(j + 15, len(lines))
    while j < limit:
        ln = lines[j].strip()
        if not ln:
            j += 1
            continue
        # Stop: next register address/index
        next_num = re.match(r'^(\d{1,5})\s*$', ln)
        if next_num and int(next_num.group(1)) != index:
            break
        # Stop: section headers
        if any(re.search(p, ln, re.IGNORECASE)
               for p, _ in [
                   (r'Device\s+information', ''), (r'Inverter\s+real\s*time', ''),
                   (r'DEA-AVM', ''), (r'DER-AVM', ''), (r'Inverter\s+param', ''),
                   (r'Register\s+Address', ''),
                   (r'The following registers', ''), (r'Address\s+\d+-\d+\s+for', ''),
               ]):
            break
        # Skip headers/footers
        if any(sp.match(ln) for sp in _SKIP_RE):
            j += 1
            continue
        collected.append(ln)
        j += 1

    if not collected:
        return None

    # ── Address extraction ──
    addr_hex = None
    addr_val = None
    for ci, c in enumerate(collected):
        m = re.search(r'0[xX]([0-9A-Fa-f]{2,6})', c)
        if m:
            addr_val = int(m.group(1), 16)
            addr_hex = f"0x{addr_val:04X}"
            collected[ci] = ''
            break
    if addr_val is None:
        for ci, c in enumerate(collected):
            if re.match(r'^(\d{3,5})$', c.strip()):
                addr_val = int(c.strip())
                addr_hex = f"0x{addr_val:04X}"
                collected[ci] = ''
                break
    if addr_val is None:
        if index >= 100:
            addr_val = index
            addr_hex = f"0x{addr_val:04X}"
        else:
            return None
    reg['address'] = addr_val
    reg['address_hex'] = addr_hex

    # ── Structured parsing: classify each collected line ──
    defn_parts = []
    comment_parts = []
    found_rw = False
    found_type = False
    phase = 'definition'  # definition → rw → type → regs → scale → unit → comment

    for ci, c in enumerate(collected):
        c_stripped = c.strip()
        if not c_stripped:
            continue

        c_upper = c_stripped.upper()

        # R/W
        if c_stripped in _RW_SET and not found_rw:
            reg['rw'] = c_stripped
            found_rw = True
            phase = 'type'
            continue

        # Type
        if c_upper in _TYPE_SET and not found_type:
            reg['type'] = c_stripped
            found_type = True
            phase = 'regs'
            continue

        # After type: register count (single digit 1-8)
        if phase == 'regs' and re.match(r'^[1-8]$', c_stripped):
            reg['regs'] = int(c_stripped)
            phase = 'scale'
            continue

        # Scale factor (e.g., "10", "100", "1")
        if phase == 'scale' and re.match(r'^\d{1,4}$', c_stripped):
            val = int(c_stripped)
            if val in (1, 10, 100, 1000, 10000):
                reg['scale_factor'] = str(val)
                phase = 'unit'
                continue

        # Unit
        if phase in ('scale', 'unit') and c_stripped in _UNIT_SET:
            reg['unit'] = c_stripped
            phase = 'comment'
            continue

        # Range values like [0, 1200] → comment
        if re.match(r'^\[[\d,\s\.\-]+\]', c_stripped):
            comment_parts.append(c_stripped)
            phase = 'comment'
            continue

        # Before R/W found: accumulate as definition
        if not found_rw:
            defn_parts.append(c_stripped)
        else:
            # After unit/scale: remaining is comment
            comment_parts.append(c_stripped)

    # Build definition (merge multi-line)
    reg['definition'] = ' '.join(defn_parts).strip()
    if comment_parts:
        reg['comment'] = ' | '.join(comment_parts).strip()

    return reg


# ──────────────────────────────────────────────────────────────────────
# Kstar 전용 파서
# ──────────────────────────────────────────────────────────────────────

def _parse_kstar_registers(pages_text, full_text):
    """
    Kstar PDF (KSG1-250K 등)의 Modbus 레지스터 테이블 파싱.
    테이블 형식: Register Address | Item | Byte | Byte No. | Unit | Data Type | Remark | FUNC
    주소는 십진수 (2964, 3000, 4000 등)
    """
    registers = []
    all_lines = full_text.split('\n')
    all_lines = [l.strip() for l in all_lines if l.strip()]

    # 섹션 매핑
    current_section = ""
    section_map = {
        'basic inverter information': 'Inverter Realtime Data',
        '04h telemetry': 'Inverter Realtime Data',
        'inverter system information': 'Device Information (System Info)',
        'inverter setup': 'Inverter Parameters (Setup)',
        'instruction execution': 'DER-AVM Parameters (Control)',
        '06h telemetry': 'DER-AVM Parameters (Control)',
    }

    # Kstar 헤더/푸터 필터 패턴
    skip_patterns_kstar = [
        r'^Shenzhen Kstar', r'^Doc Code', r'^Release',
        r'^Doc\s*$', r'^Name$', r'^Page\s+\d+', r'^\d+\s+of\s+\d+$',
        r'^KSG1', r'^Table\s+\d', r'^Note', r'^Register$',
        r'^Address$', r'^Item$', r'^Byte$', r'^No\.$',
        r'^Unit$', r'^Data$', r'^Type$', r'^Remark', r'^FUNC$',
        r'^Content$', r'^Value$', r'^Meaning$', r'^Description$',
    ]

    # 레지스터 주소 패턴: 4자리 십진수 (2964~4099 등)
    addr_pattern = re.compile(r'^(\d{4})$')
    # 또는 범위 형태: 3200-3204
    addr_range_pattern = re.compile(r'^(\d{4})\s*[-–]\s*(\d{4})$')

    # 타입 패턴
    type_names = {'U16', 'S16', 'U32', 'S32', 'U8', 'S8', 'ASCII'}

    i = 0
    while i < len(all_lines):
        line = all_lines[i]

        # 섹션 탐지
        for key, sec_name in section_map.items():
            if key in line.lower():
                current_section = sec_name
                break

        # 테이블에서 섹션 번호로 섹션 탐지
        sec_match = re.match(r'^3\.(\d)\.?\s', line)
        if sec_match:
            sec_num = sec_match.group(1)
            if sec_num == '1':
                current_section = 'Inverter Realtime Data'
            elif sec_num == '2':
                current_section = 'Device Information (System Info)'
            elif sec_num == '3':
                current_section = 'Inverter Parameters (Setup)'
            elif sec_num == '4':
                current_section = 'DER-AVM Parameters (Control)'

        # 레지스터 주소 탐지 (4자리 숫자 단독)
        addr_match = addr_pattern.match(line)
        if addr_match and current_section:
            addr_val = int(addr_match.group(1))
            if 2900 <= addr_val <= 5000:
                # U32의 두 번째 레지스터인지 확인 (이전 줄이 같은 주소-1이면 스킵)
                if registers and registers[-1]['address'] == addr_val - 1:
                    # 이전 레지스터가 U32/S32면 두 번째 바이트이므로 스킵
                    if registers[-1].get('type', '') in ('S32', 'U32'):
                        i += 1
                        continue
                    # 또는 다음 줄이 바로 FUNC 코드뿐이면 스킵
                    if i + 1 < len(all_lines) and re.match(r'^0[346]H$', all_lines[i+1].strip(), re.IGNORECASE):
                        i += 1
                        continue
                reg = _collect_kstar_entry(all_lines, i, addr_val, current_section)
                if reg:
                    registers.append(reg)

        # 범위 형태 (3200-3204) → 시작 주소만
        range_match = addr_range_pattern.match(line)
        if range_match and current_section:
            addr_start = int(range_match.group(1))
            addr_end = int(range_match.group(2))
            if 2900 <= addr_start <= 5000:
                reg = _collect_kstar_entry(all_lines, i, addr_start, current_section,
                                            regs_count=(addr_end - addr_start + 1))
                if reg:
                    registers.append(reg)

        i += 1

    return registers


def _collect_kstar_entry(lines, start_idx, addr_val, section, regs_count=None):
    """
    Kstar PDF에서 레지스터 주소 줄 다음의 항목 정보 수집.
    형식: Address 줄 → Item(이름) → Byte(크기) → Byte No. → Unit → Type → Remark → FUNC
    """
    reg = {
        'index': addr_val,
        'section': section,
        'definition': '',
        'address': addr_val,
        'address_hex': f"0x{addr_val:04X}",
        'regs': 1,
        'type': '',
        'unit': '',
        'rw': 'RO',
        'comment': '',
    }

    collected = []
    j = start_idx + 1
    limit = min(j + 10, len(lines))
    func_found = False

    while j < limit:
        ln = lines[j].strip()
        if not ln:
            j += 1
            continue
        # FUNC 코드 (04H, 03H, 06H, 10H)를 만나면 수집 종료
        if re.match(r'^0[346]H$', ln, re.IGNORECASE) or re.match(r'^10H$', ln, re.IGNORECASE):
            collected.append(ln)
            func_found = True
            j += 1
            break
        # 다른 레지스터 주소면 중단
        if re.match(r'^\d{4}$', ln) or re.match(r'^\d{4}\s*[-–]\s*\d{4}$', ln):
            try:
                next_addr = int(re.match(r'^(\d{4})', ln).group(1))
                # U32의 두 번째 레지스터 (addr+1)는 스킵
                if next_addr == addr_val + 1:
                    j += 1
                    # 다음 줄이 FUNC코드이면 같이 스킵
                    if j < limit and re.match(r'^0[346]H$', lines[j].strip(), re.IGNORECASE):
                        collected.append(lines[j].strip())
                        func_found = True
                        j += 1
                    break
                else:
                    break
            except (ValueError, AttributeError):
                break
        # 섹션 헤더면 중단
        if re.match(r'^3\.\d\.', ln) or re.search(r'Table\s+\d+\.\d+', ln):
            break
        # 헤더/푸터 스킵
        if any(re.match(p, ln, re.IGNORECASE) for p in [
            r'^Shenzhen Kstar', r'^Doc Code', r'^Release',
            r'^Page\s+\d+', r'^\d+\s+of\s+\d+',
            r'^Register$', r'^Address$', r'^Item$', r'^Byte$',
        ]):
            j += 1
            continue
        collected.append(ln)
        j += 1

    if not collected:
        return None

    # FUNC (기능 코드)에서 R/W 판별
    func_code = ''
    for ci, c in enumerate(collected):
        if re.match(r'^0[346]H$', c.strip(), re.IGNORECASE):
            func_code = c.strip().upper()
            collected[ci] = ''
        elif re.match(r'^10H$', c.strip(), re.IGNORECASE):
            func_code = '10H'
            collected[ci] = ''

    if func_code in ('06H', '10H'):
        reg['rw'] = 'RW'
    elif func_code == '04H':
        reg['rw'] = 'RO'
    elif func_code == '03H':
        reg['rw'] = 'RO'

    # Type 추출
    for ci, c in enumerate(collected):
        ct = c.strip().upper()
        if ct in ('U16', 'S16', 'U32', 'S32', 'U8', 'S8'):
            reg['type'] = c.strip()
            collected[ci] = ''
            break
        elif ct == 'ASCII':
            reg['type'] = 'string'
            collected[ci] = ''
            break

    # Byte 수 / 레지스터 수 추출
    for ci, c in enumerate(collected):
        if re.match(r'^[12468]$', c.strip()) or re.match(r'^1[0246]$', c.strip()) or re.match(r'^2[024]$', c.strip()):
            try:
                byte_count = int(c.strip())
                if 1 <= byte_count <= 32:
                    reg['regs'] = max(1, byte_count // 2)
                    collected[ci] = ''
                    break
            except ValueError:
                pass

    if regs_count is not None:
        reg['regs'] = regs_count

    # Unit 추출
    unit_patterns_kstar = [
        r'^0\.\d+[VAWHz%℃]+', r'^\d+[VAWHz%]+$',
        r'^[VAWHz%°℃]+$', r'^[Kk][Ww]h$', r'^[Ww]h$',
        r'^0\.1V$', r'^0\.01A$', r'^0\.1W$', r'^0\.01Hz$',
        r'^1W$', r'^1V$', r'^1A$', r'^1VA$', r'^1Var$',
        r'^1KVar$', r'^r/min$', r'^Kohm$', r'^Hour$',
        r'^%/min$', r'^0\.\s*1', r'^S$',
    ]
    for ci, c in enumerate(collected):
        for up in unit_patterns_kstar:
            if re.match(up, c.strip(), re.IGNORECASE):
                reg['unit'] = c.strip()
                collected[ci] = ''
                break

    # Byte No. (정수값) 제거
    for ci, c in enumerate(collected):
        if re.match(r'^\d{1,3}$', c.strip()):
            try:
                val = int(c.strip())
                if 0 <= val <= 400:
                    collected[ci] = ''
            except ValueError:
                pass

    # 나머지를 Definition과 Comment로
    remaining = [c for c in collected if c.strip()]
    if remaining:
        reg['definition'] = remaining[0]
        if len(remaining) > 1:
            reg['comment'] = ' | '.join(remaining[1:])

    return reg


# ──────────────────────────────────────────────────────────────────────
# Huawei 전용 파서
# ──────────────────────────────────────────────────────────────────────

def _parse_huawei_registers(pages_text, full_text):
    """
    Huawei PDF → _parse_huawei_registers_from_tables()로 위임.
    pages_text는 사용하지 않으나 인터페이스 호환성 유지.
    (parse_modbus_pdf에서 filepath를 직접 전달하도록 수정됨)
    """
    # 이 함수는 parse_modbus_pdf에서 직접 호출되지 않음
    # (filepath 기반의 _parse_huawei_registers_from_tables 사용)
    return []


def _parse_huawei_registers_from_tables(filepath):
    """
    Huawei PDF의 테이블을 PyMuPDF find_tables()로 직접 추출.
    각 테이블 행: [No, SignalName, RW, Type, Unit, Gain, Address, NumRegs, Scope]
    """
    import fitz

    doc = fitz.open(filepath)
    try:
        return _parse_huawei_tables(doc)
    finally:
        doc.close()

def _parse_huawei_tables(doc):
    registers = []
    seen_addrs = set()

    type_map = {
        'U16': 'U16', 'U32': 'U32',
        'I16': 'S16', 'I32': 'S32',
        'STR': 'string', 'E16': 'U16',
        'Bitfield16': 'U16', 'Bitfield32': 'U32',
        'MLD': 'string',
    }

    for page_idx in range(len(doc)):
        page = doc[page_idx]
        tables = page.find_tables()
        if not tables.tables:
            continue

        for table in tables.tables:
            rows = table.extract()
            for row in rows:
                if len(row) < 8:
                    continue

                # 헤더 행 스킵
                no_str = (row[0] or '').strip()
                if not no_str or not re.match(r'^\d{1,4}$', no_str):
                    continue

                entry_no = int(no_str)

                # 신호명: 줄바꿈 합치기
                sig_name = (row[1] or '').replace('\n', ' ').strip()
                sig_name = _merge_huawei_signal_name(sig_name)

                rw = (row[2] or '').replace('\n', '').strip()
                if rw not in ('RO', 'RW', 'WO'):
                    rw = 'RO'

                # 타입: 줄바꿈 제거
                raw_type = (row[3] or '').replace('\n', '').strip()
                reg_type = type_map.get(raw_type, raw_type)

                # 단위
                unit = (row[4] or '').replace('\n', '').strip()

                # 주소: 줄바꿈 제거 후 숫자 합치기
                addr_str = (row[6] or '').replace('\n', '').replace(' ', '').strip()
                try:
                    addr_val = int(addr_str)
                except (ValueError, TypeError):
                    continue

                # 중복 방지
                if addr_val in seen_addrs:
                    continue
                seen_addrs.add(addr_val)

                # 레지스터 수
                regs_str = (row[7] or '').replace('\n', '').strip()
                try:
                    num_regs = int(regs_str)
                except (ValueError, TypeError):
                    num_regs = 1

                # Scope/Description
                scope = ''
                if len(row) > 8:
                    scope = (row[8] or '').replace('\n', ' ').strip()

                # 섹션 결정 (주소 범위로)
                section = 'Inverter Realtime Data'
                if 30000 <= addr_val < 32000:
                    section = 'Device Information (System Info)'
                elif 32000 <= addr_val < 33000:
                    section = 'Inverter Realtime Data'
                elif 34000 <= addr_val < 40000:
                    section = 'Inverter Realtime Data'
                elif 40000 <= addr_val < 42000:
                    section = 'Inverter Parameters (Setup)'
                elif 42000 <= addr_val < 50000:
                    section = 'Inverter Parameters (Setup)'

                reg = {
                    'index': entry_no,
                    'section': section,
                    'definition': sig_name,
                    'address': addr_val,
                    'address_hex': f"0x{addr_val:04X}",
                    'regs': num_regs,
                    'type': reg_type,
                    'unit': unit,
                    'rw': rw,
                    'comment': scope[:120] if scope else '',
                }
                registers.append(reg)

    return registers


def _merge_huawei_signal_name(name):
    """
    Huawei PDF에서 분리된 신호명 합치기.
    예: 'Mod el' → 'Model', 'PV1 volta ge' → 'PV1 voltage'
    """
    parts = name.split()
    if len(parts) <= 1:
        return name

    merged = [parts[0]]
    for p in parts[1:]:
        prev = merged[-1] if merged else ''
        # 소문자로 시작하면서 이전 단어가 소문자로 끝나면 붙임 (단어 분리)
        # 예: "volta" + "ge" → "voltage", "curre" + "nt" → "current"
        if p and p[0].islower() and prev and prev[-1].islower():
            merged[-1] = prev + p
        # 숫자 뒤에 소문자가 오면 붙임 (예: "PV1" + "voltage" → "PV1voltage"는 안됨)
        # 대신 이전 단어가 대문자/숫자로 끝나면 공백 유지
        else:
            merged.append(p)

    return ' '.join(merged)


# ──────────────────────────────────────────────────────────────────────
# UDP 프로토콜 정의
# ──────────────────────────────────────────────────────────────────────

# (P) BODY(인버터) 기본 데이터
UDP_BODY_INVERTER_BASIC = [
    # (UDP Field Name, Size(bytes), Type, Unit, Scale)
    ("PV 전압",        2, "unsigned", "V",  1),
    ("PV 전류",        2, "unsigned", "A",  1),
    ("PV 출력",        4, "unsigned", "W",  1),
    ("R(RS)상 전압",   2, "unsigned", "V",  1),
    ("S(ST)상 전압",   2, "unsigned", "V",  1),
    ("T(TR)상 전압",   2, "unsigned", "V",  1),
    ("R상 전류",       2, "unsigned", "A",  1),
    ("S상 전류",       2, "unsigned", "A",  1),
    ("T상 전류",       2, "unsigned", "A",  1),
    ("인버터 출력",    4, "unsigned", "W",  1),
    ("역률",           2, "signed",   "%",  10),
    ("주파수",         2, "unsigned", "Hz", 10),
    ("누적 발전량",    8, "unsigned", "Wh", 1),
    ("상태 정보 1",    2, "unsigned", "",   0),
    ("상태 정보 2",    2, "unsigned", "",   0),
    ("상태 정보 3",    2, "unsigned", "",   0),
    ("상태 정보 4",    2, "unsigned", "",   0),
]

# (A) BODY 13: 인버터 제어 값 (H3 DER-AVM 제어)
UDP_BODY_CONTROL_VALUES = [
    ("인버터 기동/정지",  2, "unsigned", "",   0,  "0=기동, 1=정지"),
    ("역률 설정 값",      2, "signed",   "",   1000, "Power Factor"),
    ("동작 모드",         2, "unsigned", "",   0,  "0=self, 2=DER-AVM, 5=Q(V)"),
    ("무효전력 설정 값",  2, "unsigned", "%",  10, "Reactive Power %"),
    ("유효전력 설정 값",  2, "unsigned", "%",  10, "Active Power %"),
]

# (A) BODY 14: 인버터 제어 결과 모니터링 (DER-AVM 실시간)
UDP_BODY_CONTROL_MONITOR = [
    ("R상 전류",     4, "unsigned", "A",  10, ""),
    ("S상 전류",     4, "unsigned", "A",  10, ""),
    ("T상 전류",     4, "unsigned", "A",  10, ""),
    ("R(RS)상 전압", 4, "unsigned", "V",  10, ""),
    ("S(ST)상 전압", 4, "unsigned", "V",  10, ""),
    ("T(TR)상 전압", 4, "unsigned", "V",  10, ""),
    ("유효전력",     4, "signed",   "kW", 10, ""),
    ("무효전력",     4, "signed",   "Var", 0, ""),
    ("역률",         4, "signed",   "",   1000, ""),
    ("주파수",       4, "unsigned", "Hz", 10, ""),
    ("상태 플래그",  4, "bit",      "",   0,  ""),
]

# (A) BODY 15: IV SCAN 데이터
UDP_BODY_IV_SCAN = [
    ("STRING 전체 번호",  1, "unsigned", "", 0, "총 STRING 수"),
    ("STRING 현재 데이터", 1, "unsigned", "", 0, "현재 STRING 번호"),
    ("STRING 포인트 수",  1, "unsigned", "", 0, "IV 커브 포인트 수"),
    # 아래는 포인트 수만큼 반복
    ("STRING 전압",       2, "unsigned", "V", 10, "IV 포인트 전압"),
    ("STRING 전류",       2, "unsigned", "A", 10, "IV 포인트 전류"),
]

UDP_HEADER_FIELDS = [
    ("헤더 버전",    1, "unsigned", "버전 1"),
    ("데이터 순서",  2, "unsigned", "랜덤 시작 ~ 0xFFFF"),
    ("아이디",       4, "unsigned", "unique id (전화번호)"),
    ("시간",         8, "unsigned", "unix timestamp"),
    ("설비 종류",    1, "unsigned", "0=RTU, 1=인버터, 2=환경센서, 3=전력량계, 4=보호계전기"),
    ("설비 번호",    1, "unsigned", "1번부터 시작"),
    ("설비 모델",    1, "unsigned", "(C) 설비 모델 참고"),
    ("백업 데이터",  1, "unsigned", "0=신규, 1=백업"),
    ("바디 타입",    1, "signed",   "바디 타입 참고"),
]

# 비주기(A) REQUEST 요청 종류
APERIODIC_REQUEST_TYPES = [
    (1,  "RTU 재부팅"),
    (2,  "RTU 정보 요청"),
    (11, "인버터 모델 정보"),
    (12, "인버터 IV SCAN 데이터"),
    (13, "인버터 제어 값 조회"),
    (14, "인버터 제어 초기화"),
    (15, "인버터 ON/OFF 제어"),
    (16, "인버터 유효전력 제어"),
    (17, "인버터 역률 제어"),
    (18, "인버터 무효전력 제어"),
]


# ──────────────────────────────────────────────────────────────────────
# 자동 매핑 엔진
# ──────────────────────────────────────────────────────────────────────

def auto_map_to_udp(parsed_regs, mppt_count=4, string_count=8, body_type=4):
    """
    파싱된 Modbus 레지스터를 UDP BODY 필드에 자동 매핑.
    키워드 매칭 기반. 제조사별 키워드 패턴 적용.
    """
    rt = parsed_regs.get('realtime_data', [])
    der_rt = parsed_regs.get('der_avm_realtime', [])
    der_params = parsed_regs.get('der_avm_params', [])
    dev_info = parsed_regs.get('device_info', [])
    inv_params = parsed_regs.get('inverter_params', [])
    all_regs = parsed_regs.get('all_registers', [])
    manufacturer = parsed_regs.get('manufacturer', 'Unknown')

    mapping = {
        'periodic_basic': [],     # (P) 기본 데이터 매핑
        'periodic_mppt': [],      # (P) MPPT 매핑
        'periodic_string': [],    # (P) STRING 매핑
        'control_values': [],     # (A) 13 제어 값 (DER-AVM)
        'control_monitor': [],    # (A) 14 제어 모니터링 (DER-AVM)
        'iv_scan': [],            # (A) 15 IV SCAN
    }

    # ── 제조사별 키워드 프로파일 선택 ──
    basic_keywords = _get_basic_keywords(manufacturer)

    for udp_field, keywords in basic_keywords:
        found = _find_best_match(rt + all_regs, keywords)
        if found:
            mapping['periodic_basic'].append({
                'udp_field': udp_field,
                'mb_addr': found['address'],
                'mb_addr_hex': found['address_hex'],
                'mb_regs': found['regs'],
                'mb_type': found['type'],
                'mb_unit': found['unit'],
                'mb_definition': found['definition'],
                'mb_rw': found['rw'],
            })
        else:
            mapping['periodic_basic'].append({
                'udp_field': udp_field,
                'mb_addr': None, 'mb_addr_hex': '-',
                'mb_regs': 0, 'mb_type': '', 'mb_unit': '',
                'mb_definition': '(미매핑)', 'mb_rw': '',
            })

    # ── MPPT 매핑 ──
    if body_type in (2, 4):
        for ch in range(1, mppt_count + 1):
            v_kw, i_kw = _get_mppt_keywords(manufacturer, ch)
            v_found = _find_best_match(rt + all_regs, v_kw)
            i_found = _find_best_match(rt + all_regs, i_kw)
            if v_found:
                mapping['periodic_mppt'].append({
                    'udp_field': f'MPPT{ch} 전압',
                    'mb_addr': v_found['address'],
                    'mb_addr_hex': v_found['address_hex'],
                    'mb_regs': v_found['regs'],
                    'mb_type': v_found['type'],
                    'mb_unit': v_found['unit'],
                    'mb_definition': v_found['definition'],
                    'mb_rw': v_found['rw'],
                })
            if i_found:
                mapping['periodic_mppt'].append({
                    'udp_field': f'MPPT{ch} 전류',
                    'mb_addr': i_found['address'],
                    'mb_addr_hex': i_found['address_hex'],
                    'mb_regs': i_found['regs'],
                    'mb_type': i_found['type'],
                    'mb_unit': i_found['unit'],
                    'mb_definition': i_found['definition'],
                    'mb_rw': i_found['rw'],
                })

    # ── STRING 매핑 ──
    if body_type in (3, 4):
        for ch in range(1, string_count + 1):
            i_kw = _get_string_keywords(manufacturer, ch)
            i_found = _find_best_match(rt + all_regs, i_kw)
            if i_found:
                mapping['periodic_string'].append({
                    'udp_field': f'STRING{ch} 전류',
                    'mb_addr': i_found['address'],
                    'mb_addr_hex': i_found['address_hex'],
                    'mb_regs': i_found['regs'],
                    'mb_type': i_found['type'],
                    'mb_unit': i_found['unit'],
                    'mb_definition': i_found['definition'],
                    'mb_rw': i_found['rw'],
                })

    # ── DER-AVM 제어 값 매핑 (H3 제어) ──
    # Huawei는 DER-AVM 제어 미지원
    if manufacturer != 'Huawei':
        ctrl_keywords = _get_control_keywords(manufacturer)
        search_pool = der_params + inv_params + all_regs
        for udp_field, keywords in ctrl_keywords:
            found = _find_best_match(search_pool, keywords, prefer_rw=True)
            if found:
                mapping['control_values'].append({
                    'udp_field': udp_field,
                    'mb_addr': found['address'],
                    'mb_addr_hex': found['address_hex'],
                    'mb_regs': found['regs'],
                    'mb_type': found['type'],
                    'mb_unit': found['unit'],
                    'mb_definition': found['definition'],
                    'mb_rw': found['rw'],
                })

    # ── DER-AVM 제어 모니터링 매핑 ──
    # Huawei는 DER-AVM 제어 미지원
    if manufacturer != 'Huawei':
        monitor_keywords = _get_monitor_keywords(manufacturer)
        der_pool = der_rt + all_regs
        for udp_field, keywords in monitor_keywords:
            found = _find_best_match(der_pool, keywords)
            if found:
                mapping['control_monitor'].append({
                    'udp_field': udp_field,
                    'mb_addr': found['address'],
                    'mb_addr_hex': found['address_hex'],
                    'mb_regs': found['regs'],
                    'mb_type': found['type'],
                    'mb_unit': found['unit'],
                    'mb_definition': found['definition'],
                    'mb_rw': found['rw'],
                })

    # ── IV SCAN 매핑 ──
    # Kstar, Huawei는 IV scan 미지원
    if manufacturer not in ('Kstar', 'Huawei'):
        iv_kw_list = [
            ("IV Scan 트리거", ['i-v curve scan', 'iv.*scan', 'iv.*curve',
                               'i/v scan', 'iv scan.*status']),
        ]
        for udp_field, keywords in iv_kw_list:
            found = _find_best_match(inv_params + all_regs, keywords, prefer_rw=True)
            if found:
                mapping['iv_scan'].append({
                    'udp_field': udp_field,
                    'mb_addr': found['address'],
                    'mb_addr_hex': found['address_hex'],
                    'mb_regs': found['regs'],
                    'mb_type': found['type'],
                    'mb_unit': found['unit'],
                    'mb_definition': found['definition'],
                    'mb_rw': found['rw'],
                })

        # Tracker/String IV 데이터 레지스터 (0x8000~)
        for reg in all_regs:
            if reg.get('address', 0) >= 0x8000 and reg.get('address', 0) <= 0x8B3F:
                mapping['iv_scan'].append({
                    'udp_field': f"IV {reg['definition']}",
                    'mb_addr': reg['address'],
                    'mb_addr_hex': reg['address_hex'],
                    'mb_regs': reg['regs'],
                    'mb_type': reg['type'],
                    'mb_unit': reg['unit'],
                    'mb_definition': reg['definition'],
                    'mb_rw': reg['rw'],
                })

    return mapping


def _get_basic_keywords(manufacturer):
    """제조사별 주기 기본 데이터 매핑 키워드"""
    if manufacturer == 'Kstar':
        return [
            ("PV 전압",       [r'\bpv1 input voltage\b']),
            ("PV 전류",       [r'\bpv1 input current\b']),
            ("PV 출력",       [r'\bpv1 input power\b']),
            ("R(RS)상 전압",  [r'\brs.phase grid voltage\b']),
            ("S(ST)상 전압",  [r'\bst.phase grid voltage\b']),
            ("T(TR)상 전압",  [r'\btr.phase grid voltage\b']),
            ("R상 전류",      [r'\br.phase grid.tied current\b']),
            ("S상 전류",      [r'\bs.phase grid.tied current\b']),
            ("T상 전류",      [r'\bt.phase grid.tied current\b']),
            ("인버터 출력",   [r'\bgrid.tied power\b']),
            ("역률",          [r'\bpower factor\b']),
            ("주파수",        [r'\brs.phase grid frequency\b', r'\bgrid frequency\b']),
            ("누적 발전량",   [r'\btotal energy yield\b', r'\btotal energy\b']),
            ("상태 정보 1",   [r'\boperating mode\b']),
            ("상태 정보 2",   [r'\bdsp alarm\b']),
            ("상태 정보 3",   [r'\bdsp error\b']),
            ("상태 정보 4",   [r'\barm alarm\b', r'\barm error\b']),
        ]
    elif manufacturer == 'Huawei':
        return [
            ("PV 전압",       ['pv1.*volt', 'pv1 volt']),
            ("PV 전류",       ['pv1.*curr', 'pv1 curr']),
            ("PV 출력",       [r'^dc\s*power$', 'dc.*power']),
            ("R(RS)상 전압",  ['phase.*a.*volt', 'grid.*phase.*a.*volt']),
            ("S(ST)상 전압",  ['phase.*b.*volt', 'grid.*phase.*b.*volt']),
            ("T(TR)상 전압",  ['phase.*c.*volt', 'grid.*phase.*c.*volt']),
            ("R상 전류",      ['phase.*a.*curr', 'grid.*phase.*a.*curr']),
            ("S상 전류",      ['phase.*b.*curr', 'grid.*phase.*b.*curr']),
            ("T상 전류",      ['phase.*c.*curr', 'grid.*phase.*c.*curr']),
            ("인버터 출력",   [r'^active\s*power$', r'^active\s*power\b']),
            ("역률",          [r'power\s*factor', r'powerfactor']),
            ("주파수",        [r'grid\s*frequency', r'gridfrequency']),
            ("누적 발전량",   ['accumulated.*power.*gen', 'accumulated.*energy',
                              'total.*energy.*yield']),
            ("상태 정보 1",   ['device.*status']),
            ("상태 정보 2",   ['fault.*code']),
            ("상태 정보 3",   ['alarm.*1']),
            ("상태 정보 4",   ['alarm.*2']),
        ]
    else:
        # Synergy(VerterKing) 기본 키워드
        return [
            ("PV 전압",       ['mppt1 voltage', 'mppt1.*volt', 'pv.*voltage']),
            ("PV 전류",       ['mppt1 current', 'mppt1.*curr', 'pv.*current']),
            ("PV 출력",       ['pv total input power', 'pv.*power', 'total.*pv']),
            ("R(RS)상 전압",  ['l1 voltage', 'l1.*volt', 'r.*phase.*volt']),
            ("S(ST)상 전압",  ['l2 voltage', 'l2.*volt', 's.*phase.*volt']),
            ("T(TR)상 전압",  ['l3 voltage', 'l3.*volt', 't.*phase.*volt']),
            ("R상 전류",      ['l1 current', 'l1.*curr', 'r.*phase.*curr']),
            ("S상 전류",      ['l2 current', 'l2.*curr', 's.*phase.*curr']),
            ("T상 전류",      ['l3 current', 'l3.*curr', 't.*phase.*curr']),
            ("인버터 출력",   ['grid total active power', 'total.*active.*power', 'output.*power']),
            ("역률",          ['power factor']),
            ("주파수",        ['l1 frequency', 'frequency']),
            ("누적 발전량",   ['total energy']),
            ("상태 정보 1",   ['inverter mode', 'device status', 'operating mode']),
            ("상태 정보 2",   ['error code.*1', 'alarm.*1', 'alarm code']),
            ("상태 정보 3",   ['error code.*2', 'alarm.*2', 'error code']),
            ("상태 정보 4",   ['error code.*3', 'alarm.*3']),
        ]


def _get_mppt_keywords(manufacturer, ch):
    """제조사별 MPPT 매핑 키워드"""
    if manufacturer == 'Kstar':
        # Kstar: PV1~PV12 (MPPT 아닌 PV 채널 번호 사용)
        return (
            [rf'\bpv{ch} input voltage\b'],
            [rf'\bpv{ch} input current\b'],
        )
    elif manufacturer == 'Huawei':
        return (
            [f'pv{ch}.*volt', f'pv{ch} volt', f'mppt{ch}.*volt'],
            [f'pv{ch}.*curr', f'pv{ch} curr', f'mppt{ch}.*curr'],
        )
    else:
        return (
            [f'mppt{ch} voltage', f'mppt{ch}.*volt'],
            [f'mppt{ch} current', f'mppt{ch}.*curr'],
        )


def _get_string_keywords(manufacturer, ch):
    """제조사별 STRING 매핑 키워드"""
    if manufacturer == 'Kstar':
        # Kstar: PV1 String current 1, PV1 String current 2, ...
        # ch=1 → PV1 String current 1, ch=5 → PV2 String current 1
        pv_num = (ch - 1) // 4 + 1
        str_num = (ch - 1) % 4 + 1
        return [
            rf'\bpv{pv_num} string current {str_num}\b',
        ]
    elif manufacturer == 'Huawei':
        # Huawei has no individual string current registers (only string access status)
        # PV current is per MPPT, not per string
        return []
    else:
        return [
            f'string{ch} input current', f'string{ch}.*current',
            f'string.*{ch}.*curr',
        ]


def _get_control_keywords(manufacturer):
    """제조사별 DER-AVM 제어 값 키워드"""
    if manufacturer == 'Kstar':
        return [
            ("인버터 기동/정지", ['remote.*power.off']),
            ("역률 설정 값",     ['set.*power.*factor', 'set the power']),
            ("동작 모드",        ['reactive.*control.*mode', 'reactive.*mode']),
            ("무효전력 설정 값", ['set.*reactive.*power', 'set the reactive']),
            ("유효전력 설정 값", ['set.*active.*power', 'set the active']),
        ]
    elif manufacturer == 'Huawei':
        return [
            ("인버터 기동/정지", ['shutdown', 'start.*stop', 'on.*off',
                               'power on', 'power off']),
            ("역률 설정 값",     ['power factor.*set', 'power factor']),
            ("동작 모드",        ['active.*adjust.*mode', 'reactive.*adjust.*mode',
                                'scheduling.*mode']),
            ("무효전력 설정 값", ['reactive.*power.*set', 'reactive.*power.*percent',
                                'qmax']),
            ("유효전력 설정 값", ['active.*power.*set', 'active.*power.*percent',
                                'active.*power.*derating', 'pmax']),
        ]
    else:
        return [
            ("인버터 기동/정지", ['inverter off', 'inverter on', 'on.*off']),
            ("역률 설정 값",     ['power factor setting', 'power factor']),
            ("동작 모드",        ['action mode', 'der-avm mode']),
            ("무효전력 설정 값", [r'\breactive power percent\b']),
            ("유효전력 설정 값", [r'\bactive power percent\b', r'power derating percent']),
        ]


def _get_monitor_keywords(manufacturer):
    """제조사별 DER-AVM 모니터링 키워드"""
    if manufacturer == 'Kstar':
        return [
            ("R상 전류",     [r'\br.phase grid.tied current\b']),
            ("S상 전류",     [r'\bs.phase grid.tied current\b']),
            ("T상 전류",     [r'\bt.phase grid.tied current\b']),
            ("R(RS)상 전압", [r'\brs.phase grid voltage\b']),
            ("S(ST)상 전압", [r'\bst.phase grid voltage\b']),
            ("T(TR)상 전압", [r'\btr.phase grid voltage\b']),
            ("유효전력",     [r'\bgrid.tied power\b']),
            ("무효전력",     [r'\breactive power\b']),
            ("역률",         [r'\bpower factor\b']),
            ("주파수",       [r'\brs.phase grid frequency\b', r'\bgrid.*freq\b']),
            ("상태 플래그",  [r'\boperating mode\b']),
        ]
    elif manufacturer == 'Huawei':
        return [
            ("R상 전류",     ['phase.*a.*curr', 'grid.*phase.*a.*curr']),
            ("S상 전류",     ['phase.*b.*curr', 'grid.*phase.*b.*curr']),
            ("T상 전류",     ['phase.*c.*curr', 'grid.*phase.*c.*curr']),
            ("R(RS)상 전압", ['phase.*a.*volt', 'grid.*phase.*a.*volt']),
            ("S(ST)상 전압", ['phase.*b.*volt', 'grid.*phase.*b.*volt']),
            ("T(TR)상 전압", ['phase.*c.*volt', 'grid.*phase.*c.*volt']),
            ("유효전력",     [r'^active\s*power$']),
            ("무효전력",     [r'^reactive\s*power$']),
            ("역률",         [r'power\s*factor']),
            ("주파수",       [r'grid\s*frequency', r'gridfrequency']),
            ("상태 플래그",  [r'device\s*status']),
        ]
    else:
        return [
            ("R상 전류",     ['l1 current']),
            ("S상 전류",     ['l2 current']),
            ("T상 전류",     ['l3 current']),
            ("R(RS)상 전압", ['l1 voltage']),
            ("S(ST)상 전압", ['l2 voltage']),
            ("T(TR)상 전압", ['l3 voltage']),
            ("유효전력",     ['grid total active power', 'total.*active']),
            ("무효전력",     ['grid total reactive power', 'total.*reactive']),
            ("역률",         ['power factor']),
            ("주파수",       ['frequency']),
            ("상태 플래그",  ['status flag']),
        ]


def _find_best_match(regs, keywords, prefer_rw=False):
    """키워드 리스트로 가장 적합한 레지스터 찾기 (우선순위: 키워드 순서)
    매칭 우선순위: definition만 → definition+comment 합산 → comment만
    """
    for kw in keywords:
        pattern = re.compile(kw, re.IGNORECASE)
        def_candidates = []    # definition에서 매칭된 후보
        full_candidates = []   # 합산 텍스트에서 매칭된 후보
        cmt_candidates = []    # comment에서만 매칭된 후보

        for reg in regs:
            defn = reg.get('definition', '')
            cmt = reg.get('comment', '')
            full = defn + ' ' + cmt
            if reg.get('address') is None:
                continue
            # 노이즈 필터: 목차 줄 등 제외
            if '...' in full or len(full) > 200:
                continue
            if pattern.search(defn):
                def_candidates.append(reg)
            elif pattern.search(full):
                full_candidates.append(reg)
            elif pattern.search(cmt):
                cmt_candidates.append(reg)

        # 우선순위: definition → full → comment
        candidates = def_candidates or full_candidates or cmt_candidates
        if not candidates:
            continue

        if prefer_rw:
            rw_cands = [c for c in candidates if c.get('rw') in ('RW', 'WO')]
            if rw_cands:
                return rw_cands[0]

        return candidates[0]

    return None


# ──────────────────────────────────────────────────────────────────────
# Korean field name → English constant name lookup tables
# ──────────────────────────────────────────────────────────────────────

_KO_EXACT_MAP = {
    "PV 전압":          "PV_VOLTAGE",
    "PV 전류":          "PV_CURRENT",
    "PV 출력":          "PV_POWER",
    "R(RS)상 전압":     "R_PHASE_VOLTAGE",
    "S(ST)상 전압":     "S_PHASE_VOLTAGE",
    "T(TR)상 전압":     "T_PHASE_VOLTAGE",
    "R상 전류":         "R_PHASE_CURRENT",
    "S상 전류":         "S_PHASE_CURRENT",
    "T상 전류":         "T_PHASE_CURRENT",
    "인버터 출력":      "AC_POWER",
    "역률":             "POWER_FACTOR",
    "주파수":           "FREQUENCY",
    "누적 발전량":      "TOTAL_ENERGY",
    "상태 정보 1":      "INVERTER_MODE",
    "상태 정보 2":      "ERROR_CODE1",
    "상태 정보 3":      "ERROR_CODE2",
    "상태 정보 4":      "INNER_TEMP",
    "인버터 기동/정지": "ON_OFF",
    "역률 설정 값":     "POWER_FACTOR_SETPOINT",
    "동작 모드":        "OPERATING_MODE",
    "무효전력 설정 값": "REACTIVE_POWER_SETPOINT",
    "유효전력 설정 값": "ACTIVE_POWER_SETPOINT",
    "유효전력":         "ACTIVE_POWER",
    "무효전력":         "REACTIVE_POWER",
    "상태 플래그":      "STATUS_FLAG",
}

# Suffix map for pattern-based names: "MPPT3 전압" → "MPPT3_VOLTAGE"
_KO_SUFFIX_MAP = [
    (" 전압",   "_VOLTAGE"),
    (" 전류",   "_CURRENT"),
    (" 출력",   "_POWER"),
    (" 주파수", "_FREQUENCY"),
    (" 역률",   "_POWER_FACTOR"),
]

# ──────────────────────────────────────────────────────────────────────
# Solarize 프로토콜 표준 레지스터 네이밍
# solarize_registers.py와 동일한 출력을 보장하기 위한 주소→이름 매핑
# ──────────────────────────────────────────────────────────────────────

def _fill_missing_solarize_regs(all_regs):
    """
    Solarize 표준 주소 테이블(_SOLARIZE_ADDR_TO_NAME)을 기준으로
    PDF 파서가 누락한 high-word 레지스터를 자동 보충한다.

    보충 범위: 표준 4-MPPT / 8-string 구성의 핵심 레지스터 high-word만
    (DEA-AVM, AC Phase/MPPT1-4/Energy 32비트 쌍)
    MPPT5+ 등 확장 레지스터는 제외 — PDF 파서가 포함 시 자연스럽게 편입됨.
    """
    # 보충 대상 주소: addr-1이 원본에 있고, addr이 _SOLARIZE_ADDR_TO_NAME에 있으며,
    # 주소 범위가 표준 구성(0x03E8-0x03FD, 0x1000-0x104F) 이내인 것만
    _STANDARD_HIGH_RANGES = (
        (0x03E8, 0x03FD),   # DEA-AVM Realtime Data
        (0x1000, 0x104F),   # Inverter Realtime Data (MPPT1-4 포함)
        (0x1021, 0x1049),   # Energy / Power summary
    )

    existing_addrs = {r.get('address') for r in all_regs if r.get('address') is not None}
    original_addrs = frozenset(existing_addrs)

    def _in_standard_range(addr):
        return any(lo <= addr <= hi for lo, hi in _STANDARD_HIGH_RANGES)

    def _guess_section(addr):
        if 0x03E8 <= addr <= 0x03FD:
            return 'DEA-AVM Realtime Data'
        if 0x07D0 <= addr <= 0x0835 or addr == 0x3005:
            return 'DER-AVM Parameters'
        if 0x1A00 <= addr <= 0x1AFF:
            return 'Device Information'
        if 0x6000 <= addr <= 0x6FFF:
            return 'DER-AVM Parameters'
        if 0x8000 <= addr <= 0x9FFF:
            return 'IV Scan'
        return 'Inverter Realtime Data'

    # solarize_registers.py 표준 출력에 없는 진단용 레지스터 high-word 제외
    _EXCLUDE_HIGHWORD = frozenset({0x1035, 0x103C})

    # PDF 파서가 멀티라인 셀로 인해 누락하는 것으로 알려진 레지스터 — 직접 보충
    # (addr, section, rw, type) — _SOLARIZE_ADDR_TO_NAME에 이름/주석이 정의되어 있어야 함
    _KNOWN_MISSED = [
        (0x1210, 'Inverter Realtime Data',  'RO', 'U16'),  # DER-AVM digital meter connect status
        (0x1211, 'Inverter Parameters',     'RW', 'U16'),  # DER-AVM mode (self/DER-AVM control)
    ]

    extras = []

    # Pass 1: high-word補充 (기존 로직)
    for addr, (name, comment) in sorted(_SOLARIZE_ADDR_TO_NAME.items()):
        if (addr not in existing_addrs
                and (addr - 1) in original_addrs
                and _in_standard_range(addr)
                and addr not in _EXCLUDE_HIGHWORD):
            extras.append({
                'section':     _guess_section(addr),
                'definition':  name,
                'address':     addr,
                'address_hex': f'0x{addr:04X}',
                'regs':        1,
                'type':        'U16',
                'unit':        '',
                'rw':          'RO',
                'comment':     comment,
            })
            existing_addrs.add(addr)

    # Pass 2: 알려진 누락 레지스터 보충
    for addr, section, rw, dtype in _KNOWN_MISSED:
        if addr not in existing_addrs and addr in _SOLARIZE_ADDR_TO_NAME:
            name, comment = _SOLARIZE_ADDR_TO_NAME[addr]
            extras.append({
                'section':     section,
                'definition':  name,
                'address':     addr,
                'address_hex': f'0x{addr:04X}',
                'regs':        1,
                'type':        dtype,
                'unit':        '',
                'rw':          rw,
                'comment':     comment,
            })
            existing_addrs.add(addr)

    if not extras:
        return all_regs

    combined = all_regs + extras
    combined.sort(key=lambda r: r.get('address', 0))
    return combined


_SOLARIZE_ADDR_TO_NAME = {
    # Device Information (0x1A00-0x1A90)
    0x1A00: ('DEVICE_MODEL',              '8 regs, ASCII string'),
    0x1A10: ('SERIAL_NUMBER',             '8 regs, ASCII string'),
    0x1A1C: ('MASTER_FIRMWARE_VERSION',   '3 regs, ASCII string'),
    0x1A26: ('SLAVE_FIRMWARE_VERSION',    '3 regs, ASCII string'),
    0x1A3B: ('MPPT_COUNT',               '1 reg, U16'),
    0x1A44: ('NOMINAL_VOLTAGE',          '1 reg, U16, scale 0.1V'),
    0x1A45: ('NOMINAL_FREQUENCY',        '1 reg, U16, scale 0.01Hz'),
    0x1A46: ('NOMINAL_POWER_LOW',        '1 reg, U16, 1W'),
    0x1A4E: ('NOMINAL_POWER_HIGH',       '1 reg, U16, 1W (high word)'),
    0x1A48: ('GRID_PHASE_NUMBER',        '1 reg, U16: 1=single, 2=split, 3=three'),
    0x1A60: ('EMS_FIRMWARE_VERSION',     '3 regs, ASCII string'),
    0x1A8E: ('LCD_FIRMWARE_VERSION',     '3 regs, ASCII string'),
    # Inverter Real-time Data - AC Side (0x1001-0x100F)
    0x1001: ('L1_VOLTAGE',    'U16, scale 0.1V'),
    0x1002: ('L1_CURRENT',    'U16, scale 0.01A'),
    0x1003: ('L1_POWER_LOW',  'S32, scale 0.1W'),
    0x1004: ('L1_POWER_HIGH', ''),
    0x1005: ('L1_FREQUENCY',  'U16, scale 0.01Hz'),
    0x1006: ('L2_VOLTAGE',    'U16, scale 0.1V'),
    0x1007: ('L2_CURRENT',    'U16, scale 0.01A'),
    0x1008: ('L2_POWER_LOW',  'S32, scale 0.1W'),
    0x1009: ('L2_POWER_HIGH', ''),
    0x100A: ('L2_FREQUENCY',  'U16, scale 0.01Hz'),
    0x100B: ('L3_VOLTAGE',    'U16, scale 0.1V'),
    0x100C: ('L3_CURRENT',    'U16, scale 0.01A'),
    0x100D: ('L3_POWER_LOW',  'S32, scale 0.1W'),
    0x100E: ('L3_POWER_HIGH', ''),
    0x100F: ('L3_FREQUENCY',  'U16, scale 0.01Hz'),
    # MPPT Data (0x1010-0x1093)
    0x1010: ('MPPT1_VOLTAGE',    'U16, scale 0.1V'),
    0x1011: ('MPPT1_CURRENT',    'S16, scale 0.01A'),
    0x1012: ('MPPT1_POWER_LOW',  'U32, scale 0.1W'),
    0x1013: ('MPPT1_POWER_HIGH', ''),
    0x1014: ('MPPT2_VOLTAGE',    'U16, scale 0.1V'),
    0x1015: ('MPPT2_CURRENT',    'S16, scale 0.01A'),
    0x1016: ('MPPT2_POWER_LOW',  'U32, scale 0.1W'),
    0x1017: ('MPPT2_POWER_HIGH', ''),
    0x1018: ('MPPT3_VOLTAGE',    'U16, scale 0.1V'),
    0x1019: ('MPPT3_CURRENT',    'S16, scale 0.01A'),
    0x101A: ('MPPT3_POWER_LOW',  'U32, scale 0.1W'),
    0x101B: ('MPPT3_POWER_HIGH', ''),
    0x103E: ('MPPT4_VOLTAGE',    'U16, scale 0.1V'),
    0x103F: ('MPPT4_CURRENT',    'S16, scale 0.01A'),
    0x1040: ('MPPT4_POWER_LOW',  'U32, scale 0.1W'),
    0x1041: ('MPPT4_POWER_HIGH', ''),
    0x1080: ('MPPT5_VOLTAGE',    'U16, scale 0.1V'),
    0x1081: ('MPPT5_CURRENT',    'S16, scale 0.01A'),
    0x1082: ('MPPT5_POWER_LOW',  'U32, scale 0.1W'),
    0x1083: ('MPPT5_POWER_HIGH', ''),
    0x1084: ('MPPT6_VOLTAGE',    'U16, scale 0.1V'),
    0x1085: ('MPPT6_CURRENT',    'S16, scale 0.01A'),
    0x1086: ('MPPT6_POWER_LOW',  'U32, scale 0.1W'),
    0x1087: ('MPPT6_POWER_HIGH', ''),
    0x1088: ('MPPT7_VOLTAGE',    'U16, scale 0.1V'),
    0x1089: ('MPPT7_CURRENT',    'S16, scale 0.01A'),
    0x108A: ('MPPT7_POWER_LOW',  'U32, scale 0.1W'),
    0x108B: ('MPPT7_POWER_HIGH', ''),
    0x108C: ('MPPT8_VOLTAGE',    'U16, scale 0.1V'),
    0x108D: ('MPPT8_CURRENT',    'S16, scale 0.01A'),
    0x108E: ('MPPT8_POWER_LOW',  'U32, scale 0.1W'),
    0x108F: ('MPPT8_POWER_HIGH', ''),
    0x1090: ('MPPT9_VOLTAGE',    'U16, scale 0.1V'),
    0x1091: ('MPPT9_CURRENT',    'S16, scale 0.01A'),
    0x1092: ('MPPT9_POWER_LOW',  'U32, scale 0.1W'),
    0x1093: ('MPPT9_POWER_HIGH', ''),
    # DER-AVM Status / Mode (0x120C-0x1211)
    0x120C: ('FAN_SPEED_DUTY_PCT',        'U16, 1%, external fan speed duty'),
    0x1210: ('DER_AVM_METER_STATUS',      'U16, 0=disconnected, 1=connected'),
    0x1211: ('DER_AVM_MODE',              'U16, 0=self control, 1=DER-AVM control'),
    # Status & Error (0x101C-0x1020)
    0x101C: ('INNER_TEMP',    'S16, 1C'),
    0x101D: ('INVERTER_MODE', 'U16, see InverterMode class'),
    0x101E: ('ERROR_CODE1',   'U16, bit field'),
    0x101F: ('ERROR_CODE2',   'U16, bit field'),
    0x1020: ('ERROR_CODE3',   'U16, bit field'),
    # Energy Data (0x1021-0x1028)
    0x1021: ('TOTAL_ENERGY_LOW',             'U32, kWh'),
    0x1022: ('TOTAL_ENERGY_HIGH',            ''),
    0x1023: ('TOTAL_GENERATION_TIME_LOW',    'U32, Hour'),
    0x1024: ('TOTAL_GENERATION_TIME_HIGH',   ''),
    0x1027: ('TODAY_ENERGY_LOW',             'U32, Wh'),
    0x1028: ('TODAY_ENERGY_HIGH',            ''),
    # Grid Power Data (0x1034-0x103D)
    0x1034: ('FUSE_OPEN_DATA_LOW',              'U16, bit field'),
    0x1035: ('FUSE_OPEN_DATA_HIGH',             'U16, bit field'),
    0x1037: ('GRID_TOTAL_ACTIVE_POWER_LOW',     'S32, scale 0.1W (L1+L2+L3)'),
    0x1038: ('GRID_TOTAL_ACTIVE_POWER_HIGH',    ''),
    0x1039: ('GRID_TOTAL_REACTIVE_POWER_LOW',   'S32, scale 0.1Var (L1+L2+L3)'),
    0x103A: ('GRID_TOTAL_REACTIVE_POWER_HIGH',  ''),
    0x103B: ('PV_TODAY_PEAK_POWER_LOW',         'S32, scale 0.1W'),
    0x103C: ('PV_TODAY_PEAK_POWER_HIGH',        ''),
    0x103D: ('POWER_FACTOR',                    'S16, scale 0.001'),
    # PV Total Power (0x1048-0x104C)
    0x1048: ('PV_TOTAL_INPUT_POWER_LOW',   'U32, scale 0.1W (all MPPT sum)'),
    0x1049: ('PV_TOTAL_INPUT_POWER_HIGH',  ''),
    0x104C: ('TOTAL_ENERGY_DECIMALS',      'U16, Wh (decimals of 0x1021)'),
    # String Input Data (0x1050-0x107F)
    0x1050: ('STRING1_VOLTAGE',  'U16, scale 0.1V'),
    0x1051: ('STRING1_CURRENT',  'S16, scale 0.01A'),
    0x1052: ('STRING2_VOLTAGE',  'U16, scale 0.1V'),
    0x1053: ('STRING2_CURRENT',  'S16, scale 0.01A'),
    0x1054: ('STRING3_VOLTAGE',  'U16, scale 0.1V'),
    0x1055: ('STRING3_CURRENT',  'S16, scale 0.01A'),
    0x1056: ('STRING4_VOLTAGE',  'U16, scale 0.1V'),
    0x1057: ('STRING4_CURRENT',  'S16, scale 0.01A'),
    0x1058: ('STRING5_VOLTAGE',  'U16, scale 0.1V'),
    0x1059: ('STRING5_CURRENT',  'S16, scale 0.01A'),
    0x105A: ('STRING6_VOLTAGE',  'U16, scale 0.1V'),
    0x105B: ('STRING6_CURRENT',  'S16, scale 0.01A'),
    0x105C: ('STRING7_VOLTAGE',  'U16, scale 0.1V'),
    0x105D: ('STRING7_CURRENT',  'S16, scale 0.01A'),
    0x105E: ('STRING8_VOLTAGE',  'U16, scale 0.1V'),
    0x105F: ('STRING8_CURRENT',  'S16, scale 0.01A'),
    0x1060: ('STRING9_VOLTAGE',  'U16, scale 0.1V'),
    0x1061: ('STRING9_CURRENT',  'S16, scale 0.01A'),
    0x1062: ('STRING10_VOLTAGE', 'U16, scale 0.1V'),
    0x1063: ('STRING10_CURRENT', 'S16, scale 0.01A'),
    0x1064: ('STRING11_VOLTAGE', 'U16, scale 0.1V'),
    0x1065: ('STRING11_CURRENT', 'S16, scale 0.01A'),
    0x1066: ('STRING12_VOLTAGE', 'U16, scale 0.1V'),
    0x1067: ('STRING12_CURRENT', 'S16, scale 0.01A'),
    0x1068: ('STRING13_VOLTAGE', 'U16, scale 0.1V'),
    0x1069: ('STRING13_CURRENT', 'S16, scale 0.01A'),
    0x106A: ('STRING14_VOLTAGE', 'U16, scale 0.1V'),
    0x106B: ('STRING14_CURRENT', 'S16, scale 0.01A'),
    0x106C: ('STRING15_VOLTAGE', 'U16, scale 0.1V'),
    0x106D: ('STRING15_CURRENT', 'S16, scale 0.01A'),
    0x106E: ('STRING16_VOLTAGE', 'U16, scale 0.1V'),
    0x106F: ('STRING16_CURRENT', 'S16, scale 0.01A'),
    0x1070: ('STRING17_VOLTAGE', 'U16, scale 0.1V'),
    0x1071: ('STRING17_CURRENT', 'S16, scale 0.01A'),
    0x1072: ('STRING18_VOLTAGE', 'U16, scale 0.1V'),
    0x1073: ('STRING18_CURRENT', 'S16, scale 0.01A'),
    0x1074: ('STRING19_VOLTAGE', 'U16, scale 0.1V'),
    0x1075: ('STRING19_CURRENT', 'S16, scale 0.01A'),
    0x1076: ('STRING20_VOLTAGE', 'U16, scale 0.1V'),
    0x1077: ('STRING20_CURRENT', 'S16, scale 0.01A'),
    0x1078: ('STRING21_VOLTAGE', 'U16, scale 0.1V'),
    0x1079: ('STRING21_CURRENT', 'S16, scale 0.01A'),
    0x107A: ('STRING22_VOLTAGE', 'U16, scale 0.1V'),
    0x107B: ('STRING22_CURRENT', 'S16, scale 0.01A'),
    0x107C: ('STRING23_VOLTAGE', 'U16, scale 0.1V'),
    0x107D: ('STRING23_CURRENT', 'S16, scale 0.01A'),
    0x107E: ('STRING24_VOLTAGE', 'U16, scale 0.1V'),
    0x107F: ('STRING24_CURRENT', 'S16, scale 0.01A'),
    # DEA-AVM Real-time Data (0x03E8-0x03FD)
    0x03E8: ('DEA_L1_CURRENT_LOW',             'S32, scale 0.1A'),
    0x03E9: ('DEA_L1_CURRENT_HIGH',            ''),
    0x03EA: ('DEA_L2_CURRENT_LOW',             'S32, scale 0.1A'),
    0x03EB: ('DEA_L2_CURRENT_HIGH',            ''),
    0x03EC: ('DEA_L3_CURRENT_LOW',             'S32, scale 0.1A'),
    0x03ED: ('DEA_L3_CURRENT_HIGH',            ''),
    0x03EE: ('DEA_L1_VOLTAGE_LOW',             'S32, scale 0.1V'),
    0x03EF: ('DEA_L1_VOLTAGE_HIGH',            ''),
    0x03F0: ('DEA_L2_VOLTAGE_LOW',             'S32, scale 0.1V'),
    0x03F1: ('DEA_L2_VOLTAGE_HIGH',            ''),
    0x03F2: ('DEA_L3_VOLTAGE_LOW',             'S32, scale 0.1V'),
    0x03F3: ('DEA_L3_VOLTAGE_HIGH',            ''),
    0x03F4: ('DEA_TOTAL_ACTIVE_POWER_LOW',     'S32, scale 0.1kW'),
    0x03F5: ('DEA_TOTAL_ACTIVE_POWER_HIGH',    ''),
    0x03F6: ('DEA_TOTAL_REACTIVE_POWER_LOW',   'S32, scale 1 Var'),
    0x03F7: ('DEA_TOTAL_REACTIVE_POWER_HIGH',  ''),
    0x03F8: ('DEA_POWER_FACTOR_LOW',           'S32, scale 0.001'),
    0x03F9: ('DEA_POWER_FACTOR_HIGH',          ''),
    0x03FA: ('DEA_FREQUENCY_LOW',              'S32, scale 0.1Hz'),
    0x03FB: ('DEA_FREQUENCY_HIGH',             ''),
    0x03FC: ('DEA_STATUS_FLAG_LOW',            'S32, bit field'),
    0x03FD: ('DEA_STATUS_FLAG_HIGH',           ''),
    # DER-AVM Control Parameters (0x07D0-0x0835)
    0x07D0: ('DER_POWER_FACTOR_SET',    'S16, scale 0.001, [-1000,-800],[800,1000]'),
    0x07D1: ('DER_ACTION_MODE',         'U16: 0=self, 2=DER-AVM, 5=Q(V)'),
    0x07D2: ('DER_REACTIVE_POWER_PCT',  'S16, scale 0.1%, [-484, 484]'),
    0x07D3: ('DER_ACTIVE_POWER_PCT',    'U16, scale 0.1%, [0, 1100]'),
    0x0834: ('INVERTER_ON_OFF',         'U16: 0=ON, 1=OFF'),
    0x0835: ('CLEAR_PV_INSULATION_WARNING', 'U16: 0=Non-active, 1=Clear'),
    # Inverter Control (0x6001-0x6010)
    0x6001: ('INVERTER_CONTROL',     'U16: 0=power on, 1=shut down'),
    0x6006: ('EEPROM_DEFAULT',       'U16: 0=not default, 1=default'),
    0x6009: ('CLEAR_ENERGY_RECORD',  'U16: 0=not clear, 1=clear'),
    0x600D: ('IV_CURVE_SCAN',        'U16: Write 0=Non-active, 1=Active; Read 0=Idle, 1=Running, 2=Finished'),
    0x600F: ('POWER_FACTOR_DYNAMIC', 'S16, scale 0.001 (dynamic control)'),
    0x6010: ('REACTIVE_POWER_DYNAMIC', 'S16, scale 0.01% (dynamic control)'),
    # Power Derating (0x3005)
    0x3005: ('POWER_DERATING_PCT',   'U16, [0-110]%, dynamic active power'),
    # IV Scan Data Registers (0x8000-0x847F)
    0x8000: ('IV_TRACKER1_VOLTAGE_BASE',   '64 regs (0x8000-0x803F)'),
    0x8040: ('IV_STRING1_1_CURRENT_BASE',  '64 regs (0x8040-0x807F)'),
    0x8080: ('IV_STRING1_2_CURRENT_BASE',  '64 regs (0x8080-0x80BF)'),
    0x8140: ('IV_TRACKER2_VOLTAGE_BASE',   '64 regs (0x8140-0x817F)'),
    0x8180: ('IV_STRING2_1_CURRENT_BASE',  '64 regs (0x8180-0x81BF)'),
    0x81C0: ('IV_STRING2_2_CURRENT_BASE',  '64 regs (0x81C0-0x81FF)'),
    0x8280: ('IV_TRACKER3_VOLTAGE_BASE',   '64 regs (0x8280-0x82BF)'),
    0x82C0: ('IV_STRING3_1_CURRENT_BASE',  '64 regs (0x82C0-0x82FF)'),
    0x8300: ('IV_STRING3_2_CURRENT_BASE',  '64 regs (0x8300-0x833F)'),
    0x83C0: ('IV_TRACKER4_VOLTAGE_BASE',   '64 regs (0x83C0-0x83FF)'),
    0x8400: ('IV_STRING4_1_CURRENT_BASE',  '64 regs (0x8400-0x843F)'),
    0x8440: ('IV_STRING4_2_CURRENT_BASE',  '64 regs (0x8440-0x847F)'),
}

# Solarize 표준 SCALE (카테고리별 일반 키)
_SOLARIZE_STANDARD_SCALE = {
    'voltage': 0.1,
    'current': 0.01,
    'power': 0.1,
    'frequency': 0.01,
    'power_factor': 0.001,
    'dea_current': 0.1,
    'dea_voltage': 0.1,
    'dea_active_power': 0.1,
    'dea_reactive_power': 1,
    'dea_frequency': 0.1,
    'iv_voltage': 0.1,
    'iv_current': 0.1,
}

# ──────────────────────────────────────────────────────────────────────
# Kstar 레지스터 주소→이름 매핑 (kstar_registers.py 기반, decimal addresses)
# ──────────────────────────────────────────────────────────────────────

_KSTAR_ADDR_TO_NAME = {
    # PV (DC) 입력 — FC04
    3000: ('PV1_VOLTAGE',           'U16, x0.1V — MPPT1 전압'),
    3001: ('PV2_VOLTAGE',           'U16, x0.1V — MPPT2 전압'),
    3002: ('PV3_VOLTAGE',           'U16, x0.1V — MPPT3 전압'),
    3003: ('PV4_VOLTAGE',           'U16, x0.1V — MPPT4 전압'),
    3004: ('PV5_VOLTAGE',           'U16, x0.1V — MPPT5 전압'),
    3005: ('PV6_VOLTAGE',           'U16, x0.1V — MPPT6 전압'),
    3012: ('PV1_CURRENT',           'U16, x0.01A — MPPT1 전류'),
    3013: ('PV2_CURRENT',           'U16, x0.01A — MPPT2 전류'),
    3014: ('PV3_CURRENT',           'U16, x0.01A — MPPT3 전류'),
    3015: ('PV4_CURRENT',           'U16, x0.01A — MPPT4 전류'),
    3024: ('PV1_POWER',             'U16, 1W — MPPT1 전력'),
    3025: ('PV2_POWER',             'U16, 1W — MPPT2 전력'),
    3026: ('PV3_POWER',             'U16, 1W — MPPT3 전력'),
    3027: ('PV4_POWER',             'U16, 1W — MPPT4 전력'),
    # 에너지 데이터 — FC04
    3036: ('DAILY_PRODUCTION',      'U16, x0.1kWh — 금일 발전량'),
    3037: ('MONTHLY_PRODUCTION_L',  'U32 하위, 1kWh — 월간 발전량'),
    3038: ('MONTHLY_PRODUCTION_H',  'U32 상위'),
    3039: ('YEARLY_PRODUCTION_L',   'U32 하위, 1kWh — 연간 발전량'),
    3040: ('YEARLY_PRODUCTION_H',   'U32 상위'),
    3041: ('CUMULATIVE_PRODUCTION_L', 'U32 하위, x0.1kWh — 누적 발전량'),
    3042: ('CUMULATIVE_PRODUCTION_H', 'U32 상위'),
    # 상태 / 에러 — FC04
    3044: ('WORKING_MODE',          'U16 — 동작 모드'),
    3045: ('MODEL_CODE',            'U16 — 모델 코드'),
    3046: ('SYSTEM_STATUS',         'U16 — 시스템 상태 (KstarSystemStatus)'),
    3047: ('INVERTER_STATUS',       'U16 — 인버터 상태 (KstarInverterStatus)'),
    3049: ('DSP_ALARM_CODE_L',      'U32 하위 — DSP 알람 코드'),
    3050: ('DSP_ALARM_CODE_H',      'U32 상위'),
    3051: ('DSP_ERROR_CODE_L',      'U32 하위 — DSP 에러 코드'),
    3052: ('DSP_ERROR_CODE_H',      'U32 상위'),
    # 온도 / 버스 전압 — FC04
    3053: ('BUS_VOLTAGE',           'U16, x0.1V — 버스 전압'),
    3054: ('DC_BUS_VOLTAGE',        'U16, x0.1V — DC 버스 전압'),
    3055: ('RADIATOR_TEMP',         'S16, x0.1℃ — 방열판 온도'),
    3057: ('CHASSIS_TEMP',          'S16, x0.1℃ — 내부 온도'),
    # AC 출력 — R상 (FC04)
    3097: ('GRID_R_VOLTAGE',        'U16, x0.1V — 계통 R상 전압'),
    3098: ('GRID_FREQUENCY',        'U16, x0.01Hz — 계통 주파수'),
    3099: ('METER_R_CURRENT',       'S16, x0.001A — 미터 R상 전류'),
    3100: ('GRID_R_POWER',          'S16, 1W — 계통 R상 전력'),
    3123: ('INV_R_VOLTAGE',         'U16, x0.1V — 인버터 R상 출력 전압'),
    3124: ('INV_R_CURRENT',         'U16, x0.01A — 인버터 R상 출력 전류'),
    # AC 출력 — S상 (FC04)
    3125: ('INV_S_FREQUENCY',       'U16, x0.01Hz — 인버터 주파수'),
    3126: ('INV_R_POWER',           'S16, 1W — 인버터 R상 전력'),
    3127: ('GRID_S_VOLTAGE',        'U16, x0.1V — 계통 S상 전압'),
    3128: ('GRID_S_FREQUENCY',      'U16, x0.01Hz — 계통 S상 주파수'),
    3129: ('METER_S_CURRENT',       'S16, x0.001A — 미터 S상 전류'),
    3130: ('GRID_S_POWER',          'S16, 1W — 계통 S상 전력'),
    3131: ('INV_S_VOLTAGE',         'U16, x0.1V — 인버터 S상 출력 전압'),
    3132: ('INV_S_CURRENT',         'U16, x0.01A — 인버터 S상 출력 전류'),
    # AC 출력 — T상 (FC04)
    3133: ('INV_S_POWER',           'S16, 1W — 인버터 S상 전력'),
    3134: ('GRID_T_VOLTAGE',        'U16, x0.1V — 계통 T상 전압'),
    3135: ('GRID_T_FREQUENCY',      'U16, x0.01Hz — 계통 T상 주파수'),
    3136: ('METER_T_CURRENT',       'S16, x0.001A — 미터 T상 전류'),
    3137: ('GRID_T_POWER',          'S16, 1W — 계통 T상 전력'),
    3138: ('INV_T_VOLTAGE',         'U16, x0.1V — 인버터 T상 출력 전압'),
    3139: ('INV_T_CURRENT',         'U16, x0.01A — 인버터 T상 출력 전류'),
    3140: ('INV_T_POWER',           'S16, 1W — 인버터 T상 전력'),
    # 부하 전력 — FC04
    3144: ('TOTAL_LOAD_POWER',      'U16, 1W — 전체 부하 전력'),
    # 추가 에너지 데이터 — FC04
    3109: ('DAILY_ENERGY_PURCHASED', 'U16, x0.1kWh — 금일 수전량'),
    3116: ('DAILY_ENERGY_FEEDIN',   'U16, x0.1kWh — 금일 송전량'),
    3121: ('CUMULATIVE_FEEDIN_L',   'U32 하위, x0.1kWh — 누적 송전량'),
    3122: ('CUMULATIVE_FEEDIN_H',   'U32 상위'),
    3147: ('DAILY_CONSUMPTION',     'U16, x0.1kWh — 금일 소비량'),
    # DER-AVM 제어 레지스터 (hex addresses)
    0x07D0: ('DER_POWER_FACTOR_SET',    'S16, scale 0.001 — 역률 설정'),
    0x07D1: ('DER_ACTION_MODE',         'U16: 0=자립, 2=DER-AVM, 5=Q(V)'),
    0x07D2: ('DER_REACTIVE_POWER_PCT',  'S16, scale 0.1% — 무효전력 제한 설정'),
    0x07D3: ('DER_ACTIVE_POWER_PCT',    'U16, scale 0.1% — 유효전력 제한 설정'),
    0x0834: ('INVERTER_ON_OFF',         'U16: 0=운전(ON), 1=정지(OFF)'),
    # 계통 표준 — FC03
    3193: ('GRID_STANDARD',         'U16, 1 reg — 계통 표준 코드'),
    # 장비 정보 — FC03
    3200: ('MODEL_NAME_BASE',       'ASCII, 8 regs (3200~3207) — 모델명'),
    3216: ('ARM_VERSION',           'U16, 1 reg — ARM FW 버전'),
    3217: ('DSP_VERSION',           'U16, 1 reg — DSP FW 버전'),
    # 시리얼 번호 — FC04
    3228: ('SERIAL_NUMBER_BASE',    'ASCII, 11 regs (3228~3238) — 시리얼번호'),
}

# ──────────────────────────────────────────────────────────────────────
# Huawei 레지스터 주소→이름 매핑 (huawei_registers.py 기반, decimal addresses)
# ──────────────────────────────────────────────────────────────────────

_HUAWEI_ADDR_TO_NAME = {
    # Device Status
    32000: ('RUNNING_STATUS',   'U16 — Running status'),
    32002: ('FAULT_CODE_1',     'U32 — Fault code 1 (2 regs: 32002-32003)'),
    32004: ('FAULT_CODE_2',     'U32 — Fault code 2 (2 regs: 32004-32005)'),
    # PV String Inputs (DC side)
    32016: ('PV1_VOLTAGE',      'U16, 0.1V'),
    32017: ('PV1_CURRENT',      'S16, 0.01A'),
    32018: ('PV2_VOLTAGE',      'U16, 0.1V'),
    32019: ('PV2_CURRENT',      'S16, 0.01A'),
    32020: ('PV3_VOLTAGE',      'U16, 0.1V'),
    32021: ('PV3_CURRENT',      'S16, 0.01A'),
    32022: ('PV4_VOLTAGE',      'U16, 0.1V'),
    32023: ('PV4_CURRENT',      'S16, 0.01A'),
    32024: ('PV5_VOLTAGE',      'U16, 0.1V'),
    32025: ('PV5_CURRENT',      'S16, 0.01A'),
    32026: ('PV6_VOLTAGE',      'U16, 0.1V'),
    32027: ('PV6_CURRENT',      'S16, 0.01A'),
    32028: ('PV7_VOLTAGE',      'U16, 0.1V'),
    32029: ('PV7_CURRENT',      'S16, 0.01A'),
    32030: ('PV8_VOLTAGE',      'U16, 0.1V'),
    32031: ('PV8_CURRENT',      'S16, 0.01A'),
    # DC Input Power
    32064: ('INPUT_POWER',      'S32, 1W (2 regs: 32064-32065)'),
    # AC Grid (3-phase)
    32069: ('PHASE_A_VOLTAGE',  'U16, 1V'),
    32070: ('PHASE_B_VOLTAGE',  'U16, 1V'),
    32071: ('PHASE_C_VOLTAGE',  'U16, 1V'),
    32072: ('PHASE_A_CURRENT',  'S32, 0.001A (2 regs: 32072-32073)'),
    32074: ('PHASE_B_CURRENT',  'S32, 0.001A (2 regs: 32074-32075)'),
    32076: ('PHASE_C_CURRENT',  'S32, 0.001A (2 regs: 32076-32077)'),
    32080: ('ACTIVE_POWER',     'S32, 1W (2 regs: 32080-32081)'),
    32082: ('REACTIVE_POWER',   'S32, 1var (2 regs: 32082-32083)'),
    32084: ('POWER_FACTOR',     'S16, 0.001'),
    32085: ('GRID_FREQUENCY',   'U16, 0.01Hz'),
    # Temperature
    32087: ('INTERNAL_TEMP',    'S16, 0.1C'),
    # Energy
    32106: ('ACCUMULATED_ENERGY', 'U32, 1kWh (2 regs: 32106-32107)'),
}

# ──────────────────────────────────────────────────────────────────────
# 제조사별 SCALE 템플릿
# ──────────────────────────────────────────────────────────────────────

_KSTAR_STANDARD_SCALE = {
    'voltage': 0.1, 'current': 0.01, 'power': 1.0, 'frequency': 0.01, 'power_factor': 0.001,
    'pv_voltage': 0.1, 'pv_current': 0.01, 'pv_power': 1.0,
    'ac_voltage': 0.1, 'ac_current': 0.01, 'ac_power': 1.0,
    'meter_current': 0.001, 'temperature': 0.1, 'bus_voltage': 0.1,
    'daily_energy': 0.1, 'cum_energy': 0.1,
}

_HUAWEI_STANDARD_SCALE = {
    'voltage': 0.1, 'current': 0.01, 'power': 1.0, 'frequency': 0.01, 'power_factor': 0.001,
    'pv_voltage': 0.1, 'pv_current': 0.01,
    'ac_voltage': 1.0, 'ac_current': 0.001, 'temperature': 0.1,
}

# Modbus type string → struct format character
_MB_TYPE_FMT = {
    'U8':      'B',  'UINT8':   'B',
    'U16':     'H',  'UINT16':  'H',
    'U32':     'I',  'UINT32':  'I',
    'U64':     'Q',  'UINT64':  'Q',
    'S8':      'b',  'INT8':    'b',
    'S16':     'h',  'INT16':   'h',
    'S32':     'i',  'INT32':   'i',
    'S64':     'q',  'INT64':   'q',
    'FLOAT':   'f',  'FLOAT32': 'f',  'F32': 'f',
    'DOUBLE':  'd',  'FLOAT64': 'd',
}

# Per-field default (type, unit) when CSV/PDF source has no type/unit info.
# Key = const_name from _KO_EXACT_MAP / _KO_SUFFIX_MAP.
# Values: (mb_type, mb_unit)  — applied only when both are empty in source.
_FIELD_DEFAULTS = {
    # ── Device Information (solarize 표준 네이밍) ─────────────────────────
    'DEVICE_MODEL':             ('ASCII', ''),
    'DEVICE_MODEL_NAME':        ('ASCII', ''),     # Model Maker 호환
    'SERIAL_NUMBER':            ('ASCII', ''),
    'DEVICE_SERIAL_NUMBER':     ('ASCII', ''),     # Model Maker 호환
    'MASTER_FIRMWARE_VERSION':  ('ASCII', ''),
    'SLAVE_FIRMWARE_VERSION':   ('ASCII', ''),
    'EMS_FIRMWARE_VERSION':     ('ASCII', ''),
    'LCD_FIRMWARE_VERSION':     ('ASCII', ''),
    'MPPT_COUNT':               ('U16', ''),
    'MPPT_NUMBER':              ('U16', ''),        # Model Maker 호환
    'NOMINAL_VOLTAGE':          ('U16', '0.1V'),
    'NOMINAL_FREQUENCY':        ('U16', '0.01Hz'),
    'NOMINAL_POWER_LOW':        ('U16', '1W'),
    'NOMINAL_POWER_HIGH':       ('U16', '1W'),
    'NOMINAL_ACTIVE_POWER':     ('U32', '1W'),     # Model Maker 호환
    'GRID_PHASE_NUMBER':        ('U16', ''),
    # ── AC / Grid (solarize 표준: L1/L2/L3 네이밍) ───────────────────────
    'L1_VOLTAGE':               ('U16', '0.1V'),
    'L1_CURRENT':               ('U16', '0.01A'),
    'L1_POWER_LOW':             ('S32', '0.1W'),
    'L1_FREQUENCY':             ('U16', '0.01Hz'),
    'L2_VOLTAGE':               ('U16', '0.1V'),
    'L2_CURRENT':               ('U16', '0.01A'),
    'L2_POWER_LOW':             ('S32', '0.1W'),
    'L2_FREQUENCY':             ('U16', '0.01Hz'),
    'L3_VOLTAGE':               ('U16', '0.1V'),
    'L3_CURRENT':               ('U16', '0.01A'),
    'L3_POWER_LOW':             ('S32', '0.1W'),
    'L3_FREQUENCY':             ('U16', '0.01Hz'),
    'INNER_TEMP':               ('S16', '1C'),
    'INVERTER_MODE':            ('U16', ''),
    'ERROR_CODE1':              ('U16', ''),
    'ERROR_CODE2':              ('U16', ''),
    'ERROR_CODE3':              ('U16', ''),
    'POWER_FACTOR':             ('S16', '0.001'),
    'PV_TOTAL_INPUT_POWER_LOW': ('U32', '0.1W'),
    'GRID_TOTAL_ACTIVE_POWER_LOW':    ('S32', '0.1W'),
    'GRID_TOTAL_REACTIVE_POWER_LOW':  ('S32', '0.1Var'),
    # ── MPPT / STRING channels (suffix fallback) ─────────────────────────
    '_VOLTAGE':                 ('U16', '0.1V'),
    '_CURRENT':                 ('U16', '0.01A'),
    '_POWER':                   ('S32', '1W'),
    '_FREQUENCY':               ('U16', '0.01Hz'),
    # ── DER-AVM control (solarize 표준 네이밍) ────────────────────────────
    'INVERTER_ON_OFF':          ('U16', ''),
    'DER_POWER_FACTOR_SET':     ('S16', '0.001'),
    'DER_ACTION_MODE':          ('U16', ''),
    'DER_REACTIVE_POWER_PCT':   ('S16', '0.1%'),
    'DER_ACTIVE_POWER_PCT':     ('U16', '0.1%'),
    'ON_OFF':                   ('U16', ''),        # Model Maker 호환
    'OPERATING_MODE':           ('U16', ''),        # Model Maker 호환
    'ACTIVE_POWER_SETPOINT':    ('U16', '0.1%'),   # Model Maker 호환
    'REACTIVE_POWER_SETPOINT':  ('U16', '0.1%'),   # Model Maker 호환
    'POWER_FACTOR_SETPOINT':    ('S16', '0.001'),  # Model Maker 호환
    # ── DEA-AVM ───────────────────────────────────────────────────────────
    'DEA_TOTAL_ACTIVE_POWER_LOW':    ('S32', '0.1kW'),
    'DEA_TOTAL_REACTIVE_POWER_LOW':  ('S32', '1 Var'),
    'DEA_POWER_FACTOR_LOW':          ('S32', '0.001'),
    'DEA_FREQUENCY_LOW':             ('S32', '0.1Hz'),
    'DEA_STATUS_FLAG_LOW':           ('S32', ''),
    # ── IV Scan ───────────────────────────────────────────────────────────
    'IV_CURVE_SCAN':            ('U16', ''),
    'IV_SCAN':                  ('U16', ''),        # Model Maker 호환
}


# Known register file → device type / model mapping
# (device_type, model_number, model_name, filename, protocol, class_name)
_REGISTER_FILE_MAP_DEFAULT = [
    ('Inverter',  '1', 'Solarize Verterking', 'solarize_registers.py', 'solarize', 'RegisterMap'),
    ('Inverter',  '2', 'Huawei SUN2000',      'huawei_registers.py',   'huawei',   'HuaweiRegisters'),
    ('Inverter',  '3', 'Kstar KSG-60KT',      'kstar_registers.py',    'kstar',    'KstarRegisters'),
    ('Inverter',  '4', 'Solarize VK50K',       'verterking_registers.py','solarize', 'SolarizeRegisters'),
    ('Relay',     '1', 'KDU-300',              'relay_registers.py',    'modbus',   'KDU300RegisterMap'),
    ('Weather',   '1', 'SEM5046',              'weather_registers.py',  'modbus',   'SEM5046RegisterMap'),
]

# JSON 파일로 영구 저장/로드 — 재시작해도 유지
_REGISTER_MAP_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'register_file_map.json')

def _load_register_file_map():
    """JSON에서 등록 맵 로드, 없으면 기본값 사용"""
    if os.path.isfile(_REGISTER_MAP_JSON):
        try:
            with open(_REGISTER_MAP_JSON, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return [tuple(entry) for entry in data]
        except (json.JSONDecodeError, IOError, OSError) as e:
            print(f"[WARNING] Failed to load register map: {e}, using defaults")
    return list(_REGISTER_FILE_MAP_DEFAULT)

def _save_register_file_map():
    """현재 등록 맵을 JSON으로 저장"""
    try:
        with open(_REGISTER_MAP_JSON, 'w', encoding='utf-8') as f:
            json.dump(_REGISTER_FILE_MAP, f, indent=2, ensure_ascii=False)
    except (IOError, OSError) as e:
        print(f"[ERROR] Failed to save register map: {e}")

_REGISTER_FILE_MAP = _load_register_file_map()

# Protected register files — tested & in production, must not be overwritten by Save
_PROTECTED_FILES = {
    'solarize_registers.py',
    'huawei_registers.py',
    'kstar_registers.py',
    'relay_registers.py',
    'weather_registers.py',
}


# ──────────────────────────────────────────────────────────────────────
# Output directory helper
# ──────────────────────────────────────────────────────────────────────

def _ensure_output_dir(manufacturer):
    """Create and return model_maker/output/{manufacturer}/ directory."""
    base = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(base, 'output', manufacturer.lower().replace(' ', '_'))
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


# ──────────────────────────────────────────────────────────────────────
# EditableTreeview — reusable inline-edit treeview widget
# ──────────────────────────────────────────────────────────────────────

class EditableTreeview(ttk.Frame):
    """Treeview with double-click inline cell editing."""

    def __init__(self, parent, columns, headings, col_widths=None,
                 editable_columns=None, height=25, **kw):
        super().__init__(parent, **kw)
        self.columns = columns
        self.editable_columns = set(editable_columns) if editable_columns else set(columns)
        self._edit_widget = None

        self.tree = ttk.Treeview(self, columns=columns, show='headings', height=height)
        vsb = ttk.Scrollbar(self, orient='vertical', command=self.tree.yview)
        hsb = ttk.Scrollbar(self, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        for i, col in enumerate(columns):
            w = col_widths[i] if col_widths and i < len(col_widths) else 100
            h = headings[i] if i < len(headings) else col
            self.tree.heading(col, text=h)
            self.tree.column(col, width=w, minwidth=40)

        hsb.pack(side='bottom', fill='x')
        vsb.pack(side='right', fill='y')
        self.tree.pack(side='left', fill='both', expand=True)

        # Tags for color coding
        self.tree.tag_configure('auto',     background='#E8F5E9')
        self.tree.tag_configure('unmapped', background='#FFEBEE')
        self.tree.tag_configure('modified', background='#FFF9C4')
        self.tree.tag_configure('section',  background='#E3F2FD', font=('맑은 고딕', 9, 'bold'))

        self.tree.bind('<Double-1>', self._on_double_click)

    def _on_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != 'cell':
            return
        item = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not item or not col_id:
            return

        col_index = int(col_id.replace('#', '')) - 1
        col_name = self.columns[col_index]
        if col_name not in self.editable_columns:
            return

        # Cancel any existing edit
        self._cancel_edit()

        bbox = self.tree.bbox(item, col_id)
        if not bbox:
            return
        x, y, w, h = bbox

        current_val = self.tree.set(item, col_name)
        entry = tk.Entry(self.tree, font=('맑은 고딕', 9))
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, current_val)
        entry.select_range(0, 'end')
        entry.focus_set()

        entry.bind('<Return>', lambda e: self._commit_edit(item, col_name, entry))
        entry.bind('<Escape>', lambda e: self._cancel_edit())
        entry.bind('<FocusOut>', lambda e: self._commit_edit(item, col_name, entry))
        self._edit_widget = entry

    def _commit_edit(self, item, col_name, entry):
        if entry.winfo_exists():
            new_val = entry.get()
            old_val = self.tree.set(item, col_name)
            entry.destroy()
            self._edit_widget = None
            if new_val != old_val:
                self.tree.set(item, col_name, new_val)
                # Add modified tag while preserving existing tags
                tags = list(self.tree.item(item, 'tags'))
                if 'modified' not in tags:
                    tags.append('modified')
                    self.tree.item(item, tags=tags)

    def _cancel_edit(self):
        if self._edit_widget and self._edit_widget.winfo_exists():
            self._edit_widget.destroy()
            self._edit_widget = None

    def load_data(self, rows, tags_list=None):
        """Load rows (list of tuples). Optional tags_list for per-row tags."""
        self.tree.delete(*self.tree.get_children())
        for i, row in enumerate(rows):
            tag = tags_list[i] if tags_list and i < len(tags_list) else ()
            if isinstance(tag, str):
                tag = (tag,)
            self.tree.insert('', 'end', values=row, tags=tag)

    def get_data(self):
        """Return all rows as list of tuples."""
        rows = []
        for item in self.tree.get_children():
            rows.append(tuple(self.tree.item(item, 'values')))
        return rows

    def get_row_count(self):
        return len(self.tree.get_children())


# ──────────────────────────────────────────────────────────────────────
# GUI Application
# ──────────────────────────────────────────────────────────────────────

class ModbusToUdpMapper:
    VERSION = "2.1.0"

    def __init__(self, root):
        self.root = root
        self.root.title("Model Maker — 인버터 레지스터맵 생성기  v2.1.0")
        self.root.geometry("1500x900")
        self.root.resizable(True, True)

        self.parsed_data = None   # 파싱된 PDF 데이터
        self.mapping = None       # 자동 매핑 결과

        # ── Stage Pipeline state ──
        self._sp_pdf_path         = None   # PDF path used for Stage 1
        self._sp_stage1_path      = None   # Stage 1 Excel output path
        self._sp_stage2_path      = None   # Stage 2 Mapping Excel output path
        self._sp_generated_code   = ''     # Stage 3 generated code
        self._sp_manufacturer     = ''     # Detected manufacturer name

        # ── Stage 1/2 sub-tab treeviews ──
        self._s1_sheet_views = {}  # sheet_name -> EditableTreeview
        self._s2_sheet_views = {}

        style = ttk.Style()
        style.configure('Title.TLabel', font=('맑은 고딕', 13, 'bold'))
        style.configure('Section.TLabelframe.Label', font=('맑은 고딕', 10, 'bold'))
        style.configure('TButton', font=('맑은 고딕', 9))
        style.configure('TLabel', font=('맑은 고딕', 9))
        style.configure('Treeview', font=('맑은 고딕', 9), rowheight=22)
        style.configure('Treeview.Heading', font=('맑은 고딕', 9, 'bold'))
        style.configure('Info.TLabel', font=('맑은 고딕', 10), foreground='#0066CC')
        style.configure('Status.TLabel', font=('맑은 고딕', 9), foreground='#444444')

        self._build_ui()

    def _build_ui(self):
        # ── 상단 타이틀 바 ──
        top = ttk.Frame(self.root, padding=(10, 8, 10, 4))
        top.pack(side='top', fill='x')
        ttk.Label(top, text="Model Maker — 인버터 레지스터맵 생성기",
                  style='Title.TLabel').pack(side='left')
        ttk.Button(top, text="  닫기  ",
                   command=lambda: (self.root.destroy(), __import__('sys').exit(0)),
                   width=8).pack(side='right', padx=10)

        # 파일 정보
        info_frame = ttk.Frame(self.root, padding=(10, 0, 10, 5))
        info_frame.pack(fill='x')
        ttk.Label(info_frame, text="PDF:").pack(side='left')
        self.file_var = tk.StringVar(value="(PDF 파일을 열어주세요)")
        ttk.Label(info_frame, textvariable=self.file_var, style='Info.TLabel').pack(side='left', padx=5)
        self.info_var = tk.StringVar(value="")
        ttk.Label(info_frame, textvariable=self.info_var).pack(side='left', padx=20)

        # ── Legacy vars for old handlers still referenced ──
        self.body_type_var = tk.StringVar(value='4')
        self.mppt_var = tk.StringVar(value='4')
        self.string_var = tk.StringVar(value='8')

        # ── 4-Tab Notebook ──
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=(0, 5))
        self.notebook.bind('<<NotebookTabChanged>>', self._on_tab_changed)
        self._json_loaded_for = None
        self._auto_loading = False

        self._build_stage1_tab()       # Tab1: Stage 1 — PDF → Excel
        self._build_stage2_tab()       # Tab2: Stage 2 — UDP Mapping
        self._build_stage3_tab()       # Tab3: Stage 3 — Register .py + File Management

    # ══════════════════════════════════════════════════════════════════
    # Tab1: Stage 1 — PDF → Excel
    # ══════════════════════════════════════════════════════════════════

    def _build_stage1_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  Stage 1: PDF → Excel  ')

        # ── Top: PDF selector + mode + buttons ──
        top = ttk.LabelFrame(tab, text=" PDF 파일 및 생성 모드 ", padding=6)
        top.pack(fill='x', pady=(0, 4))

        row1 = ttk.Frame(top)
        row1.pack(fill='x', pady=(0, 4))

        ttk.Label(row1, text="PDF:").pack(side='left')
        self._sp_pdf_var = tk.StringVar(value="(PDF 파일을 선택하세요)")
        ttk.Label(row1, textvariable=self._sp_pdf_var,
                  foreground='#555555', width=50).pack(side='left', padx=(4, 8))

        ttk.Button(row1, text="  Open PDF  ",
                   command=self._s1_open_pdf).pack(side='left', padx=3)
        ttk.Button(row1, text="  Load Existing Excel  ",
                   command=self._s1_load_existing).pack(side='left', padx=3)

        # Mode selector
        row2 = ttk.Frame(top)
        row2.pack(fill='x')

        ttk.Label(row2, text="Mode:", font=('맑은 고딕', 9, 'bold')).pack(side='left', padx=(0, 6))
        self._sp_mode_var = tk.StringVar(value='offline')
        ttk.Radiobutton(row2, text="Offline (PyMuPDF + Reference)",
                        variable=self._sp_mode_var, value='offline').pack(side='left', padx=(0, 12))
        ttk.Radiobutton(row2, text="AI (PyMuPDF + Claude API)",
                        variable=self._sp_mode_var, value='ai').pack(side='left', padx=(0, 12))

        # Reference library status
        ref_count = 0
        if _sp is not None:
            try:
                rm = _sp._get_ref_manager()
                if rm:
                    ref_count = rm.count()
            except Exception:
                pass
        self._sp_ref_status_var = tk.StringVar(value=f"Reference: {ref_count} sets")
        ttk.Label(row2, textvariable=self._sp_ref_status_var,
                  foreground='#0066CC').pack(side='left', padx=(8, 0))

        ttk.Button(row2, text=" Manage References ",
                   command=self._sp_manage_references).pack(side='right', padx=3)
        ttk.Button(row2, text=" API Settings ",
                   command=self._on_api_settings).pack(side='right', padx=3)

        # Action buttons
        row3 = ttk.Frame(top)
        row3.pack(fill='x', pady=(4, 0))

        ttk.Button(row3, text="  자동생성 (Offline)  ",
                   command=lambda: self._s1_run_generate('offline')).pack(side='left', padx=3)
        ttk.Button(row3, text="  AI 생성  ",
                   command=lambda: self._s1_run_generate('ai')).pack(side='left', padx=3)

        self._sp_s1_status_var = tk.StringVar(value="Ready")
        ttk.Label(row3, textvariable=self._sp_s1_status_var,
                  foreground='#444444').pack(side='left', padx=(12, 0))

        self._s1_output_var = tk.StringVar(value="")
        ttk.Label(row3, textvariable=self._s1_output_var,
                  foreground='#555555').pack(side='left', padx=(8, 0))

        ttk.Button(row3, text="  Save Excel  ",
                   command=self._s1_save_excel).pack(side='right', padx=3)
        ttk.Button(row3, text="  Open in Excel  ",
                   command=self._s1_open_in_excel).pack(side='right', padx=3)

        # ── Middle: Sub-Notebook for Excel sheets ──
        self._s1_notebook = ttk.Notebook(tab)
        self._s1_notebook.pack(fill='both', expand=True, pady=(4, 2))

        # Create placeholder sub-tabs (will be populated when Excel is loaded)
        self._s1_sheet_names = [
            'Registers', 'Status_Definitions', 'Error_Fault_Codes', 'Grid_Codes',
            'Fuse_Open_Data', 'Status_Bits_03FC', 'Fuse_Check_Mask',
            'REMS_Error_Bits', 'REMS_Registers'
        ]
        for sheet_name in self._s1_sheet_names:
            frame = ttk.Frame(self._s1_notebook, padding=3)
            self._s1_notebook.add(frame, text=f'  {sheet_name}  ')
            lbl = ttk.Label(frame, text=f"(Excel 생성 후 데이터가 표시됩니다)",
                            foreground='#999999')
            lbl.pack(expand=True)

    def _s1_open_pdf(self):
        """Stage 1: Open PDF file and parse."""
        path = filedialog.askopenfilename(
            title="Select Modbus Protocol PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if not path:
            return
        self._sp_pdf_path = path
        fname = os.path.basename(path)
        self._sp_pdf_var.set(fname[:55])
        self.file_var.set(f"Loading: {fname} ...")
        self.root.update_idletasks()
        try:
            self.parsed_data = parse_modbus_pdf(path)
            self.file_var.set(fname)
            d = self.parsed_data
            mfr = d.get('manufacturer', '')
            cnt = len(d.get('all_registers', []))
            self._sp_manufacturer = mfr
            self.info_var.set(f"{mfr}  |  {cnt} registers")
            self._sp_s1_status_var.set("PDF loaded — click 자동생성 or AI생성")
        except Exception as e:
            self.file_var.set(fname)
            self._sp_s1_status_var.set(f"Parse error: {e}")

    def _s1_load_existing(self):
        """Load an existing Stage 1 Excel file directly."""
        path = filedialog.askopenfilename(
            title="Select Stage 1 Register Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        self._sp_stage1_path = path
        self._sp_s1_status_var.set(f"Loaded: {os.path.basename(path)}")
        self._s1_output_var.set(path)
        self._s1_load_excel_into_views(path)

    def _s1_run_generate(self, mode):
        """Run Stage 1: PDF → Excel in background thread."""
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        pdf_path = self._sp_pdf_path
        if not pdf_path:
            if self.parsed_data:
                pdf_path = self.parsed_data.get('filepath', '')
        if not pdf_path:
            messagebox.showwarning("Stage 1", "Open a PDF file first.")
            return

        # Force mode
        actual_mode = mode
        api_key = ''
        model_name = None
        if actual_mode == 'ai':
            if _ai_gen:
                api_key = _ai_gen.load_api_key()
                model_name = _ai_gen.load_model_name()
            if not api_key:
                messagebox.showwarning("AI Mode", "No API key configured.\n"
                                       "Set it via 'API Settings' or switch to Offline.")
                return

        # Create output directory and output path
        mfr = self._sp_manufacturer or 'unknown'
        out_dir = _ensure_output_dir(mfr)
        mfr_safe = mfr.lower().replace(' ', '_')
        out_path_target = os.path.join(out_dir, f'stage1_{mfr_safe}.xlsx')

        self._sp_s1_status_var.set(f"Running Stage 1 [{actual_mode}]...")
        self.root.update_idletasks()

        def _run():
            try:
                out_path = _sp.stage1_extract_to_excel(
                    parsed_data  = self.parsed_data,
                    pdf_path     = pdf_path,
                    output_path  = out_path_target,
                    mode         = actual_mode,
                    api_key      = api_key,
                    model        = model_name,
                    progress_cb  = lambda m: self.root.after(0, lambda: self._sp_s1_status_var.set(m[-50:]))
                )
                self._sp_stage1_path = out_path
                def _done():
                    self._sp_s1_status_var.set(f"Done: {os.path.basename(out_path)}")
                    self._s1_output_var.set(out_path)
                    self._s1_load_excel_into_views(out_path)
                self.root.after(0, _done)
            except Exception as e:
                self.root.after(0, lambda: self._sp_s1_status_var.set(f"Error: {e}"))
                self.root.after(0, lambda: messagebox.showerror("Stage 1 Error", str(e)))

        threading.Thread(target=_run, daemon=True).start()

    def _s1_load_excel_into_views(self, excel_path):
        """Read Excel file and populate sub-tab treeviews."""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error", "openpyxl required: pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel:\n{e}")
            return

        for i, sheet_name in enumerate(self._s1_sheet_names):
            # Get the sub-tab frame
            frame = self._s1_notebook.winfo_children()[i]
            # Clear existing widgets
            for w in frame.winfo_children():
                w.destroy()

            if sheet_name not in wb.sheetnames:
                ttk.Label(frame, text=f"(시트 '{sheet_name}' 없음)",
                          foreground='#999999').pack(expand=True)
                self._s1_sheet_views.pop(sheet_name, None)
                continue

            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            if not rows_data:
                ttk.Label(frame, text="(빈 시트)", foreground='#999999').pack(expand=True)
                continue

            # First row is header
            headers = [str(h) if h else '' for h in rows_data[0]]
            columns = [f'col{j}' for j in range(len(headers))]
            col_widths = [max(60, min(200, len(str(h)) * 10 + 20)) for h in headers]

            # Determine editable columns
            if sheet_name == 'Registers':
                editable = {f'col{j}' for j, h in enumerate(headers)
                            if h in ('Definition', 'Data_Type', 'Unit', 'Scale_Factor', 'R/W', 'Description')}
            else:
                editable = {f'col{j}' for j, h in enumerate(headers)
                            if h in ('Name', 'Description', 'Value', 'Notes')}

            tv = EditableTreeview(frame, columns=columns, headings=headers,
                                  col_widths=col_widths, editable_columns=editable, height=22)
            tv.pack(fill='both', expand=True)

            # Load data rows (skip header)
            data_rows = []
            tags_list = []
            for row in rows_data[1:]:
                vals = tuple(str(v) if v is not None else '' for v in row)
                data_rows.append(vals)
                # Section header rows (check if first col is section-like)
                if sheet_name == 'Registers' and len(vals) > 1:
                    if vals[0] and vals[0].startswith('─'):
                        tags_list.append('section')
                    else:
                        tags_list.append(())
                else:
                    tags_list.append(())
            tv.load_data(data_rows, tags_list)

            self._s1_sheet_views[sheet_name] = tv

            # Show count in tab title
            cnt = len(data_rows)
            self._s1_notebook.tab(i, text=f'  {sheet_name} ({cnt})  ')

        wb.close()
        self._sp_s1_status_var.set(f"Excel loaded: {len(wb.sheetnames)} sheets")

        # Auto-propagate to Stage 2
        self._sp_s1_path_var.set(os.path.basename(excel_path)[:50])

    def _s1_save_excel(self):
        """Save edited data back to Excel."""
        if not self._sp_stage1_path:
            messagebox.showwarning("Save", "No Excel file loaded.")
            return

        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error", "openpyxl required: pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(self._sp_stage1_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel:\n{e}")
            return

        for sheet_name, tv in self._s1_sheet_views.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            data = tv.get_data()
            # Overwrite data rows (keep row 1 header)
            for r_idx, row_vals in enumerate(data, start=2):
                for c_idx, val in enumerate(row_vals, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

        try:
            wb.save(self._sp_stage1_path)
            wb.close()
            self._sp_s1_status_var.set(f"Saved: {os.path.basename(self._sp_stage1_path)}")
            messagebox.showinfo("Saved", f"Excel saved:\n{self._sp_stage1_path}")
        except Exception as e:
            wb.close()
            messagebox.showerror("Save Error", str(e))

    def _s1_open_in_excel(self):
        """Open the Stage 1 Excel in the system default application."""
        path = self._sp_stage1_path
        if not path or not os.path.isfile(path):
            messagebox.showwarning("Open", "No Excel file to open.")
            return
        try:
            os.startfile(path)
        except AttributeError:
            import subprocess
            subprocess.Popen(['xdg-open', path])

    # ══════════════════════════════════════════════════════════════════
    # Tab2: Stage 2 — UDP Mapping
    # ══════════════════════════════════════════════════════════════════

    def _build_stage2_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  Stage 2: UDP 매핑  ')

        # ── Top: Source path + buttons ──
        top = ttk.LabelFrame(tab, text=" Stage 1 Excel → UDP 매핑 ", padding=6)
        top.pack(fill='x', pady=(0, 4))

        row1 = ttk.Frame(top)
        row1.pack(fill='x')

        ttk.Label(row1, text="Stage 1 Excel:").pack(side='left')
        self._sp_s1_path_var = tk.StringVar(value="(Stage 1 실행 후 자동 설정)")
        ttk.Label(row1, textvariable=self._sp_s1_path_var,
                  foreground='#555555', width=45).pack(side='left', padx=(4, 8))

        ttk.Button(row1, text="  Browse  ",
                   command=self._s2_browse_stage1).pack(side='left', padx=3)

        row2 = ttk.Frame(top)
        row2.pack(fill='x', pady=(4, 0))

        ttk.Button(row2, text="  UDP 매핑 실행 (Offline)  ",
                   command=lambda: self._s2_run_mapping('offline')).pack(side='left', padx=3)
        ttk.Button(row2, text="  UDP 매핑 실행 (AI)  ",
                   command=lambda: self._s2_run_mapping('ai')).pack(side='left', padx=3)
        ttk.Button(row2, text="  Open in Excel  ",
                   command=self._sp_open_stage2).pack(side='left', padx=3)

        self._sp_s2_status_var = tk.StringVar(value="Ready")
        ttk.Label(row2, textvariable=self._sp_s2_status_var,
                  foreground='#444444').pack(side='left', padx=(12, 0))

        self._s2_output_var = tk.StringVar(value="")
        ttk.Label(row2, textvariable=self._s2_output_var,
                  foreground='#555555').pack(side='left', padx=(8, 0))

        ttk.Button(row2, text="  Save Mapping Excel  ",
                   command=self._s2_save_mapping_excel).pack(side='right', padx=3)

        # ── Middle: Sub-Notebook for mapping sheets ──
        self._s2_notebook = ttk.Notebook(tab)
        self._s2_notebook.pack(fill='both', expand=True, pady=(4, 2))

        self._s2_sheet_names = [
            'Register_Mapping', 'Status_Mapping', 'DER_AVM_Control', 'DEA_AVM_Monitor'
        ]
        for sheet_name in self._s2_sheet_names:
            frame = ttk.Frame(self._s2_notebook, padding=3)
            self._s2_notebook.add(frame, text=f'  {sheet_name}  ')
            ttk.Label(frame, text=f"(매핑 실행 후 데이터가 표시됩니다)",
                      foreground='#999999').pack(expand=True)

    def _s2_browse_stage1(self):
        """Browse for Stage 1 Excel file."""
        path = filedialog.askopenfilename(
            title="Select Stage 1 Register Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        self._sp_stage1_path = path
        self._sp_s1_path_var.set(os.path.basename(path)[:50])

    def _s2_run_mapping(self, mode):
        """Run Stage 2: Stage 1 Excel → Mapping Excel in background thread."""
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        stage1_path = self._sp_stage1_path
        if not stage1_path or not os.path.isfile(stage1_path):
            messagebox.showwarning("Stage 2", "No Stage 1 Excel found.\nRun Stage 1 first or browse.")
            return

        api_key = ''
        model_name = None
        if mode == 'ai':
            if _ai_gen:
                api_key = _ai_gen.load_api_key()
                model_name = _ai_gen.load_model_name()
            if not api_key:
                messagebox.showwarning("AI Mode", "No API key. Use API Settings or Offline mode.")
                return

        try:
            solarize_map   = _SOLARIZE_ADDR_TO_NAME
            solarize_scale = _SOLARIZE_STANDARD_SCALE
        except NameError:
            solarize_map   = {}
            solarize_scale = {}

        # Output path
        mfr = self._sp_manufacturer or 'unknown'
        out_dir = _ensure_output_dir(mfr)
        mfr_safe = mfr.lower().replace(' ', '_')
        out_path_target = os.path.join(out_dir, f'stage2_{mfr_safe}_mapping.xlsx')

        self._sp_s2_status_var.set(f"Running Stage 2 [{mode}]...")
        self.root.update_idletasks()

        def _run():
            try:
                out_path = _sp.stage2_create_mapping_excel(
                    stage1_excel_path = stage1_path,
                    output_path       = out_path_target,
                    solarize_addr_map = solarize_map,
                    solarize_scale    = solarize_scale,
                    mode              = mode,
                    api_key           = api_key,
                    model             = model_name,
                    progress_cb       = lambda m: self.root.after(0, lambda: self._sp_s2_status_var.set(m[-50:]))
                )
                self._sp_stage2_path = out_path
                def _done():
                    self._sp_s2_status_var.set(f"Done: {os.path.basename(out_path)}")
                    self._s2_output_var.set(out_path)
                    self._sp_s2_path_var.set(os.path.basename(out_path)[:50])
                    self._s2_load_mapping_excel(out_path)
                self.root.after(0, _done)
            except Exception as e:
                self.root.after(0, lambda: self._sp_s2_status_var.set(f"Error: {e}"))
                self.root.after(0, lambda: messagebox.showerror("Stage 2 Error", str(e)))

        threading.Thread(target=_run, daemon=True).start()

    def _s2_load_mapping_excel(self, excel_path):
        """Read mapping Excel and populate sub-tab treeviews."""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error", "openpyxl required: pip install openpyxl")
            return

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel:\n{e}")
            return

        for i, sheet_name in enumerate(self._s2_sheet_names):
            frame = self._s2_notebook.winfo_children()[i]
            for w in frame.winfo_children():
                w.destroy()

            if sheet_name not in wb.sheetnames:
                ttk.Label(frame, text=f"(시트 '{sheet_name}' 없음)",
                          foreground='#999999').pack(expand=True)
                self._s2_sheet_views.pop(sheet_name, None)
                continue

            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            if not rows_data:
                ttk.Label(frame, text="(빈 시트)", foreground='#999999').pack(expand=True)
                continue

            headers = [str(h) if h else '' for h in rows_data[0]]
            columns = [f'col{j}' for j in range(len(headers))]
            col_widths = [max(60, min(200, len(str(h)) * 10 + 20)) for h in headers]

            # Editable columns
            if sheet_name == 'Register_Mapping':
                editable = {f'col{j}' for j, h in enumerate(headers)
                            if h in ('Solarize_Name', 'Notes')}
            elif sheet_name == 'Status_Mapping':
                editable = {f'col{j}' for j, h in enumerate(headers)
                            if h in ('InverterMode_State', 'Notes')}
            else:
                editable = set()  # DER_AVM, DEA_AVM are read-only

            tv = EditableTreeview(frame, columns=columns, headings=headers,
                                  col_widths=col_widths, editable_columns=editable, height=22)
            tv.pack(fill='both', expand=True)

            data_rows = []
            tags_list = []
            # Find column index for Match_Type if exists
            match_col = None
            for j, h in enumerate(headers):
                if h == 'Match_Type':
                    match_col = j
                    break

            for row in rows_data[1:]:
                vals = tuple(str(v) if v is not None else '' for v in row)
                data_rows.append(vals)
                # Color code: auto=green, unmapped=red
                if match_col is not None and len(vals) > match_col:
                    mt = vals[match_col].lower()
                    if mt in ('auto', 'ref', 'exact'):
                        tags_list.append('auto')
                    elif 'unmapped' in mt or mt == '':
                        tags_list.append('unmapped')
                    else:
                        tags_list.append(())
                else:
                    tags_list.append(())

            tv.load_data(data_rows, tags_list)
            self._s2_sheet_views[sheet_name] = tv

            cnt = len(data_rows)
            self._s2_notebook.tab(i, text=f'  {sheet_name} ({cnt})  ')

        wb.close()

    def _s2_save_mapping_excel(self):
        """Save edited mapping data back to Excel."""
        if not self._sp_stage2_path:
            messagebox.showwarning("Save", "No mapping Excel loaded.")
            return

        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error", "openpyxl required")
            return

        try:
            wb = openpyxl.load_workbook(self._sp_stage2_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open Excel:\n{e}")
            return

        for sheet_name, tv in self._s2_sheet_views.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            data = tv.get_data()
            for r_idx, row_vals in enumerate(data, start=2):
                for c_idx, val in enumerate(row_vals, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

        try:
            wb.save(self._sp_stage2_path)
            wb.close()
            self._sp_s2_status_var.set(f"Saved: {os.path.basename(self._sp_stage2_path)}")
            messagebox.showinfo("Saved", f"Mapping Excel saved:\n{self._sp_stage2_path}")
        except Exception as e:
            wb.close()
            messagebox.showerror("Save Error", str(e))

    # ══════════════════════════════════════════════════════════════════
    # Tab3: Stage 3 — Register .py Generation
    # ══════════════════════════════════════════════════════════════════

    def _build_stage3_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  Stage 3: 레지스터 생성  ')

        # ── Settings Panel ──
        cfg = ttk.LabelFrame(tab, text=" 생성 설정 ", padding=6)
        cfg.pack(fill='x', pady=(0, 4))

        # Row 0
        ttk.Label(cfg, text="Manufacturer:").grid(row=0, column=0, sticky='e', padx=(5, 2))
        self.cg_manufacturer_var = tk.StringVar(value='Solarize')
        ttk.Entry(cfg, textvariable=self.cg_manufacturer_var, width=18).grid(row=0, column=1, padx=2)

        ttk.Label(cfg, text="Protocol Name:").grid(row=0, column=2, sticky='e', padx=(10, 2))
        self.cg_protocol_var = tk.StringVar(value='solarize')
        self._s3_protocol_entry = ttk.Entry(cfg, textvariable=self.cg_protocol_var, width=14)
        self._s3_protocol_entry.grid(row=0, column=3, padx=2)
        self._s3_protocol_entry.bind('<FocusOut>', self._s3_check_duplicate)
        self._s3_protocol_entry.bind('<KeyRelease>', self._s3_check_duplicate)

        ttk.Label(cfg, text="Model #:").grid(row=0, column=4, sticky='e', padx=(10, 2))
        self.cg_model_var = tk.StringVar(value='1')
        self._s3_model_entry = ttk.Entry(cfg, textvariable=self.cg_model_var, width=6)
        self._s3_model_entry.grid(row=0, column=5, padx=2)
        self._s3_model_entry.bind('<FocusOut>', self._s3_check_duplicate)
        self._s3_model_entry.bind('<KeyRelease>', self._s3_check_duplicate)

        ttk.Label(cfg, text="Class:").grid(row=0, column=6, sticky='e', padx=(10, 2))
        self.cg_classname_var = tk.StringVar(value='RegisterMap')
        ttk.Entry(cfg, textvariable=self.cg_classname_var, width=16).grid(row=0, column=7, padx=2)

        # Row 1
        ttk.Label(cfg, text="FC:").grid(row=1, column=0, sticky='e', padx=(5, 2), pady=(4, 0))
        self.cg_fc_var = tk.StringVar(value='FC03')
        ttk.Combobox(cfg, textvariable=self.cg_fc_var, values=['FC03', 'FC04'],
                      state='readonly', width=8).grid(row=1, column=1, padx=2, pady=(4, 0), sticky='w')

        ttk.Label(cfg, text="MPPT:").grid(row=1, column=2, sticky='e', padx=(10, 2), pady=(4, 0))
        self._sp_mppt_var = tk.StringVar(value='4')
        ttk.Spinbox(cfg, from_=1, to=8, width=3,
                    textvariable=self._sp_mppt_var).grid(row=1, column=3, padx=2, pady=(4, 0), sticky='w')

        ttk.Label(cfg, text="Strings:").grid(row=1, column=4, sticky='e', padx=(10, 2), pady=(4, 0))
        self._sp_string_var = tk.StringVar(value='8')
        ttk.Spinbox(cfg, from_=1, to=24, width=3,
                    textvariable=self._sp_string_var).grid(row=1, column=5, padx=2, pady=(4, 0), sticky='w')

        self._sp_iv_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(cfg, text="IV Scan", variable=self._sp_iv_var).grid(
            row=1, column=6, padx=(10, 0), pady=(4, 0), sticky='w')

        self._sp_der_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(cfg, text="DER-AVM", variable=self._sp_der_var).grid(
            row=1, column=7, padx=(4, 0), pady=(4, 0), sticky='w')

        self.cg_ivscan_var = self._sp_iv_var
        self.cg_deravm_var = self._sp_der_var
        self._sp_dea_var = tk.BooleanVar(value=True)

        # ── Action Bar ──
        act = ttk.Frame(tab)
        act.pack(fill='x', pady=(0, 4))

        ttk.Label(act, text="Mapping Excel:").pack(side='left')
        self._sp_s2_path_var = tk.StringVar(value="(Stage 2 실행 후 자동 설정)")
        ttk.Label(act, textvariable=self._sp_s2_path_var,
                  foreground='#555555', width=42).pack(side='left', padx=(4, 8))
        ttk.Button(act, text="  Browse  ",
                   command=self._s3_browse_mapping).pack(side='left', padx=3)

        ttk.Button(act, text="  레지스터 생성  ",
                   command=self._s3_run_generate).pack(side='left', padx=5)
        ttk.Button(act, text="  Test  ",
                   command=self._s3_test).pack(side='left', padx=3)
        self._sp_s3_status_var = tk.StringVar(value="Ready")
        ttk.Label(act, textvariable=self._sp_s3_status_var,
                  foreground='#444444', width=20).pack(side='left', padx=(4, 0))
        ttk.Button(act, text="  Save to common/  ",
                   command=self._s3_save_to_common).pack(side='left', padx=3)
        ttk.Button(act, text="  Save Ref  ",
                   command=self._sp_save_to_reference).pack(side='left', padx=3)
        ttk.Button(act, text="  Manage Ref  ",
                   command=self._sp_manage_references).pack(side='left', padx=3)

        # ── 3-Pane PanedWindow: Register Files | Code Preview | Config Editor ──
        paned = tk.PanedWindow(tab, orient='horizontal', sashwidth=5, sashrelief='raised')
        paned.pack(fill='both', expand=True, pady=(0, 4))

        # ── Left Pane: Register Files ──
        left_frame = ttk.LabelFrame(paned, text=" Register Files ", padding=5)

        rf_inner = ttk.Frame(left_frame)
        rf_inner.pack(fill='both', expand=True)

        cols = ('device_type', 'model', 'model_name', 'filename', 'ref')
        self.rf_tree = ttk.Treeview(rf_inner, columns=cols, show='headings', height=10)
        self.rf_tree.heading('device_type', text='Type')
        self.rf_tree.heading('model', text='#')
        self.rf_tree.heading('model_name', text='Model Name')
        self.rf_tree.heading('filename', text='Filename')
        self.rf_tree.heading('ref', text='Ref')
        self.rf_tree.column('device_type', width=50, anchor='center')
        self.rf_tree.column('model', width=25, anchor='center')
        self.rf_tree.column('model_name', width=110)
        self.rf_tree.column('filename', width=120)
        self.rf_tree.column('ref', width=55, anchor='center')
        rf_vsb = ttk.Scrollbar(rf_inner, orient='vertical', command=self.rf_tree.yview)
        self.rf_tree.configure(yscrollcommand=rf_vsb.set)
        self.rf_tree.pack(side='left', fill='both', expand=True)
        rf_vsb.pack(side='right', fill='y')
        self.rf_tree.bind('<Double-1>', lambda e: self._on_open_selected())

        rf_btn = ttk.Frame(left_frame)
        rf_btn.pack(fill='x', pady=(4, 0))
        ttk.Button(rf_btn, text="Open", command=self._on_open_selected).pack(side='left', padx=2)
        ttk.Button(rf_btn, text="Delete", command=self._on_delete_selected).pack(side='left', padx=2)
        ttk.Button(rf_btn, text="Refresh", command=self._refresh_register_files).pack(side='left', padx=2)

        self._refresh_register_files()
        paned.add(left_frame, width=350, minsize=200)

        # ── Center Pane: Code Preview ──
        center_frame = ttk.LabelFrame(paned, text=" Code Preview ", padding=5)

        preview_inner = ttk.Frame(center_frame)
        preview_inner.pack(fill='both', expand=True)

        self.cg_text = tk.Text(preview_inner, wrap='none', font=('Consolas', 9),
                                state='disabled')
        vsb = ttk.Scrollbar(preview_inner, orient='vertical', command=self.cg_text.yview)
        hsb = ttk.Scrollbar(preview_inner, orient='horizontal', command=self.cg_text.xview)
        self.cg_text.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side='bottom', fill='x')
        vsb.pack(side='right', fill='y')
        self.cg_text.pack(side='left', fill='both', expand=True)

        paned.add(center_frame, minsize=350)

        # ── Right Pane: Config Editor ──
        config_frame = ttk.LabelFrame(paned, text=" Config Editor ", padding=5)

        cfg_sel = ttk.Frame(config_frame)
        cfg_sel.pack(fill='x', pady=(0, 4))

        base_dir = os.path.dirname(os.path.abspath(__file__))
        self._config_dir = os.path.normpath(os.path.join(base_dir, '..', 'config'))
        config_files = []
        if os.path.isdir(self._config_dir):
            config_files = sorted(
                f for f in os.listdir(self._config_dir)
                if f.endswith(('.ini', '.config', '.cfg'))
            )

        self.cfg_file_var = tk.StringVar()
        cfg_combo = ttk.Combobox(cfg_sel, textvariable=self.cfg_file_var,
                                  values=config_files, state='readonly', width=22)
        cfg_combo.pack(side='left', fill='x', expand=True)
        cfg_combo.bind('<<ComboboxSelected>>', self._on_config_selected)

        cfg_text_frame = ttk.Frame(config_frame)
        cfg_text_frame.pack(fill='both', expand=True)

        self.cfg_editor = tk.Text(cfg_text_frame, wrap='none', font=('Consolas', 9))
        cfg_vsb = ttk.Scrollbar(cfg_text_frame, orient='vertical', command=self.cfg_editor.yview)
        cfg_hsb = ttk.Scrollbar(cfg_text_frame, orient='horizontal', command=self.cfg_editor.xview)
        self.cfg_editor.configure(yscrollcommand=cfg_vsb.set, xscrollcommand=cfg_hsb.set)
        cfg_hsb.pack(side='bottom', fill='x')
        cfg_vsb.pack(side='right', fill='y')
        self.cfg_editor.pack(side='left', fill='both', expand=True)

        cfg_btn = ttk.Frame(config_frame)
        cfg_btn.pack(fill='x', pady=(4, 0))
        ttk.Button(cfg_btn, text="Add Model", command=self._on_add_model).pack(side='left', padx=3)
        ttk.Button(cfg_btn, text="Save Config", command=self._on_save_config).pack(side='left', padx=3)
        ttk.Button(cfg_btn, text="Build Package", command=self._on_build_package).pack(side='left', padx=3)

        paned.add(config_frame, width=300, minsize=200)

        # Auto-load first config
        if config_files:
            self.cfg_file_var.set(config_files[0])
            self._on_config_selected(None)

        # ── Bottom: Validation Results ──
        val_frame = ttk.LabelFrame(tab, text=" Validation Results ", padding=5)
        val_frame.pack(fill='x', pady=(0, 2))

        self.cg_guide_text = tk.Text(val_frame, wrap='word', font=('Consolas', 9),
                                      state='disabled', height=6, background='#F5F5F5')
        self.cg_guide_text.pack(fill='both', expand=True)

    def _s3_check_duplicate(self, event=None):
        """Protocol Name / Model # 중복 검사 — Entry 변경 시 호출."""
        protocol = self.cg_protocol_var.get().strip()
        model_num = self.cg_model_var.get().strip()

        warnings = []
        for entry in _REGISTER_FILE_MAP:
            if len(entry) >= 5 and entry[4] == protocol:
                warnings.append(f"Protocol '{protocol}' 사용중: {entry[3]}")
            if len(entry) >= 2 and entry[1] == model_num:
                warnings.append(f"Model #{model_num} 사용중: {entry[2]}")

        if warnings:
            self._sp_s3_status_var.set("! " + "; ".join(warnings[:2]))
        else:
            self._sp_s3_status_var.set("Ready")

    def _s3_browse_mapping(self):
        """Browse for mapping Excel file."""
        path = filedialog.askopenfilename(
            title="Select Stage 2 Mapping Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        self._sp_stage2_path = path
        self._sp_s2_path_var.set(os.path.basename(path)[:50])

    def _s3_run_generate(self):
        """Run Stage 3: Mapping Excel → register .py in background thread."""
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        stage2_path = self._sp_stage2_path
        if not stage2_path:
            self._s3_browse_mapping()
            stage2_path = self._sp_stage2_path
        if not stage2_path or not os.path.isfile(stage2_path):
            messagebox.showwarning("Stage 3", "Mapping Excel not found.\nRun Stage 2 first.")
            return

        settings = {
            'mppt_count':    self._sp_mppt_var.get(),
            'string_count':  self._sp_string_var.get(),
            'iv_scan':       self._sp_iv_var.get(),
            'der_avm':       self._sp_der_var.get(),
            'dea_avm':       self._sp_dea_var.get(),
            'manufacturer':  self.cg_manufacturer_var.get().strip() or 'Unknown',
            'class_name':    self.cg_classname_var.get().strip() or 'RegisterMap',
            'protocol_name': self.cg_protocol_var.get().strip() or 'custom',
            'fc_code':       self.cg_fc_var.get(),
        }

        mode    = self._sp_mode_var.get()
        api_key = ''
        model_name = None
        if mode == 'ai':
            if _ai_gen:
                api_key = _ai_gen.load_api_key()
                model_name = _ai_gen.load_model_name()
            if not api_key:
                messagebox.showwarning("AI Mode", "No API key configured.")
                return

        self._sp_s3_status_var.set(f"Running Stage 3 [{mode}]...")
        self.root.update_idletasks()

        def _run():
            try:
                result = _sp.stage3_generate_py(
                    mapping_excel_path = stage2_path,
                    settings           = settings,
                    output_path        = None,
                    mode               = mode,
                    api_key            = api_key,
                    model              = model_name,
                    progress_cb        = lambda m: self.root.after(0, lambda: self._sp_s3_status_var.set(m[-40:]))
                )
                code    = result.get('code', '')
                results = result.get('results', [])
                success = result.get('success', False)
                self._sp_generated_code = code

                def _update_ui():
                    self.cg_text.configure(state='normal')
                    self.cg_text.delete('1.0', 'end')
                    self.cg_text.insert('1.0', code)
                    self.cg_text.configure(state='disabled')

                    lines = [f'=== Stage 3 Results [Mode: {mode.upper()}] ===']
                    for status, msg in results:
                        lines.append(f'[{status}] {msg}')
                    fail_n = sum(1 for s, _ in results if s == 'FAIL')
                    warn_n = sum(1 for s, _ in results if s == 'WARN')
                    pass_n = sum(1 for s, _ in results if s == 'PASS')
                    lines.append('')
                    lines.append(f'Result: {pass_n} PASS  {warn_n} WARN  {fail_n} FAIL')
                    if success:
                        lines.append('>> VALIDATION PASSED')
                    else:
                        lines.append('>> VALIDATION FAILED')

                    self.cg_guide_text.configure(state='normal')
                    self.cg_guide_text.delete('1.0', 'end')
                    self.cg_guide_text.insert('1.0', '\n'.join(lines))
                    self.cg_guide_text.configure(state='disabled')

                    status_msg = f'{"PASS" if success else "FAIL"} — {pass_n}P {warn_n}W {fail_n}F'
                    self._sp_s3_status_var.set(status_msg)

                self.root.after(0, _update_ui)
            except Exception as e:
                self.root.after(0, lambda: self._sp_s3_status_var.set(f"Error: {e}"))
                self.root.after(0, lambda: messagebox.showerror("Stage 3 Error", str(e)))

        threading.Thread(target=_run, daemon=True).start()

    def _s3_test(self):
        """Test the generated Stage 3 code."""
        code = self._sp_generated_code or self.cg_text.get('1.0', 'end').strip()
        if not code:
            messagebox.showwarning("Test", "Generate code first.")
            return
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        class_name = self.cg_classname_var.get().strip() or 'RegisterMap'
        der_avm    = self._sp_der_var.get()
        results = _sp._validate_register_code(code, class_name, der_avm)
        fail_n  = sum(1 for s, _ in results if s == 'FAIL')
        warn_n  = sum(1 for s, _ in results if s == 'WARN')
        pass_n  = sum(1 for s, _ in results if s == 'PASS')

        lines = ['=== Validation Test Results ===']
        for status, msg in results:
            lines.append(f'[{status}] {msg}')
        lines += ['', f'Result: {pass_n} PASS  {warn_n} WARN  {fail_n} FAIL']

        self.cg_guide_text.configure(state='normal')
        self.cg_guide_text.delete('1.0', 'end')
        self.cg_guide_text.insert('1.0', '\n'.join(lines))
        self.cg_guide_text.configure(state='disabled')
        self._sp_s3_status_var.set(f'Test: {pass_n}P {warn_n}W {fail_n}F')

    def _s3_save_to_common(self):
        """Save generated code to common/{protocol}_registers.py."""
        code = self._sp_generated_code or self.cg_text.get('1.0', 'end').strip()
        if not code:
            messagebox.showwarning("Save", "Generate code first.")
            return

        protocol = self.cg_protocol_var.get().strip()
        if not protocol:
            messagebox.showwarning("Save", "Enter a Protocol Name.")
            return

        fname = f'{protocol}_registers.py'
        if fname in _PROTECTED_FILES:
            messagebox.showerror("Protected File",
                f"'{fname}' is protected.\nUse a different Protocol Name (e.g. '{protocol}_custom').")
            return

        base_dir   = os.path.dirname(os.path.abspath(__file__))
        target_dir = os.path.normpath(os.path.join(base_dir, '..', 'common'))
        os.makedirs(target_dir, exist_ok=True)
        filepath = os.path.join(target_dir, fname)

        if os.path.isfile(filepath):
            import shutil
            try:
                shutil.copy2(filepath, filepath + '.bak')
            except OSError:
                pass

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(code + '\n')

        self.cg_text.configure(state='normal')
        self.cg_text.delete('1.0', 'end')
        self.cg_text.insert('1.0', code)
        self.cg_text.configure(state='disabled')

        # Register in file map
        known = {e[3] for e in _REGISTER_FILE_MAP}
        if fname not in known:
            mfr       = self.cg_manufacturer_var.get().strip()
            model_num = self.cg_model_var.get().strip()
            cn        = self.cg_classname_var.get().strip()
            _REGISTER_FILE_MAP.append(('Inverter', model_num, mfr, fname, protocol, cn))
            _save_register_file_map()

        self._refresh_register_files()
        self._sp_s3_status_var.set(f"Saved: {fname}")
        messagebox.showinfo("Saved", f"Register file saved:\n{filepath}")

    # ══════════════════════════════════════════════════════════════════
    # Tab4: Register Files Management
    # ══════════════════════════════════════════════════════════════════

    def _build_management_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  레지스터 파일 관리  ')

        # ── Register File List ──
        rf_frame = ttk.LabelFrame(tab, text=" Register Files ", padding=5)
        rf_frame.pack(fill='x', pady=(0, 4))

        rf_inner = ttk.Frame(rf_frame)
        rf_inner.pack(fill='both', expand=True)

        cols = ('device_type', 'model', 'model_name', 'filename', 'ref')
        self.rf_tree = ttk.Treeview(rf_inner, columns=cols, show='headings', height=8)
        self.rf_tree.heading('device_type', text='Type')
        self.rf_tree.heading('model', text='#')
        self.rf_tree.heading('model_name', text='Model Name')
        self.rf_tree.heading('filename', text='Filename')
        self.rf_tree.heading('ref', text='Ref')
        self.rf_tree.column('device_type', width=65, anchor='center')
        self.rf_tree.column('model', width=25, anchor='center')
        self.rf_tree.column('model_name', width=150)
        self.rf_tree.column('filename', width=170)
        self.rf_tree.column('ref', width=55, anchor='center')
        rf_vsb = ttk.Scrollbar(rf_inner, orient='vertical', command=self.rf_tree.yview)
        self.rf_tree.configure(yscrollcommand=rf_vsb.set)
        self.rf_tree.pack(side='left', fill='both', expand=True)
        rf_vsb.pack(side='right', fill='y')
        self.rf_tree.bind('<Double-1>', lambda e: self._on_open_selected())

        rf_btn = ttk.Frame(rf_frame)
        rf_btn.pack(fill='x', pady=(5, 0))
        ttk.Button(rf_btn, text=" Open ", command=self._on_open_selected).pack(side='left', padx=2)
        ttk.Button(rf_btn, text=" Delete ", command=self._on_delete_selected).pack(side='left', padx=2)
        ttk.Button(rf_btn, text=" Refresh ", command=self._refresh_register_files).pack(side='left', padx=2)
        ttk.Button(rf_btn, text=" Add Model ", command=self._on_add_model).pack(side='left', padx=2)
        ttk.Button(rf_btn, text=" Manage Ref ", command=self._sp_manage_references).pack(side='left', padx=2)

        self._refresh_register_files()

        # ── Config Editor + Code Preview (PanedWindow) ──
        paned = tk.PanedWindow(tab, orient='horizontal', sashwidth=5, sashrelief='raised')
        paned.pack(fill='both', expand=True, pady=(4, 4))

        # Left: Config Editor
        config_frame = ttk.LabelFrame(paned, text=" Config Editor ", padding=5)

        cfg_sel = ttk.Frame(config_frame)
        cfg_sel.pack(fill='x', pady=(0, 5))

        base_dir = os.path.dirname(os.path.abspath(__file__))
        self._config_dir = os.path.normpath(os.path.join(base_dir, '..', 'config'))
        config_files = []
        if os.path.isdir(self._config_dir):
            config_files = sorted(
                f for f in os.listdir(self._config_dir)
                if f.endswith(('.ini', '.config', '.cfg'))
            )

        self.cfg_file_var = tk.StringVar()
        cfg_combo = ttk.Combobox(cfg_sel, textvariable=self.cfg_file_var,
                                  values=config_files, state='readonly', width=25)
        cfg_combo.pack(side='left', fill='x', expand=True)
        cfg_combo.bind('<<ComboboxSelected>>', self._on_config_selected)

        cfg_text_frame = ttk.Frame(config_frame)
        cfg_text_frame.pack(fill='both', expand=True)

        self.cfg_editor = tk.Text(cfg_text_frame, wrap='none', font=('Consolas', 9))
        cfg_vsb = ttk.Scrollbar(cfg_text_frame, orient='vertical', command=self.cfg_editor.yview)
        cfg_hsb = ttk.Scrollbar(cfg_text_frame, orient='horizontal', command=self.cfg_editor.xview)
        self.cfg_editor.configure(yscrollcommand=cfg_vsb.set, xscrollcommand=cfg_hsb.set)
        cfg_hsb.pack(side='bottom', fill='x')
        cfg_vsb.pack(side='right', fill='y')
        self.cfg_editor.pack(side='left', fill='both', expand=True)

        cfg_btn = ttk.Frame(config_frame)
        cfg_btn.pack(fill='x', pady=(5, 0))
        ttk.Button(cfg_btn, text="  Save Config  ",
                   command=self._on_save_config).pack(side='left', padx=5)
        ttk.Button(cfg_btn, text="  Build Package  ",
                   command=self._on_build_package).pack(side='left', padx=5)

        paned.add(config_frame, width=400, minsize=250)

        # Right: Code Preview for opened files
        preview_frame = ttk.LabelFrame(paned, text=" File Preview ", padding=5)

        self._mgmt_code_text = tk.Text(preview_frame, wrap='none', font=('Consolas', 9),
                                        state='disabled')
        p_vsb = ttk.Scrollbar(preview_frame, orient='vertical', command=self._mgmt_code_text.yview)
        p_hsb = ttk.Scrollbar(preview_frame, orient='horizontal', command=self._mgmt_code_text.xview)
        self._mgmt_code_text.configure(yscrollcommand=p_vsb.set, xscrollcommand=p_hsb.set)
        p_hsb.pack(side='bottom', fill='x')
        p_vsb.pack(side='right', fill='y')
        self._mgmt_code_text.pack(side='left', fill='both', expand=True)

        paned.add(preview_frame, minsize=300)

        # Auto-load first config
        if config_files:
            self.cfg_file_var.set(config_files[0])
            self._on_config_selected(None)

    # ══════════════════════════════════════════════════════════════════
    # Legacy Tab Builders (kept for backward compat, not used in new UI)
    # ══════════════════════════════════════════════════════════════════

    def _build_pdf_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  PDF 파싱 결과 (Modbus 레지스터)  ')

        frame = ttk.Frame(tab)
        frame.pack(fill='both', expand=True)

        cols = ("No", "섹션", "Definition", "Address", "Regs", "Type", "Unit", "R/W", "Comment")
        self.pdf_tree = ttk.Treeview(frame, columns=cols, show='headings', height=30)
        vsb = ttk.Scrollbar(frame, orient='vertical', command=self.pdf_tree.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=self.pdf_tree.xview)
        self.pdf_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        widths = [40, 160, 280, 80, 50, 60, 80, 40, 300]
        for i, c in enumerate(cols):
            self.pdf_tree.heading(c, text=c)
            self.pdf_tree.column(c, width=widths[i], anchor='center' if i < 8 else 'w')

        hsb.pack(side='bottom', fill='x')
        vsb.pack(side='right', fill='y')
        self.pdf_tree.pack(side='left', fill='both', expand=True)

    # ── 탭2: 주기 데이터 매핑 ────────────────────────────────────────

    def _build_periodic_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  (P) 주기 데이터 매핑  ')

        frame = ttk.Frame(tab)
        frame.pack(fill='both', expand=True)

        cols = ("No", "구분", "UDP 필드", "Modbus Definition",
                "Modbus Addr", "Regs", "Type", "Unit", "R/W")
        self.periodic_tree = ttk.Treeview(frame, columns=cols, show='headings', height=30)
        vsb = ttk.Scrollbar(frame, orient='vertical', command=self.periodic_tree.yview)
        self.periodic_tree.configure(yscrollcommand=vsb.set)

        widths = [40, 70, 150, 250, 90, 50, 60, 80, 40]
        for i, c in enumerate(cols):
            self.periodic_tree.heading(c, text=c)
            self.periodic_tree.column(c, width=widths[i], anchor='center' if i != 3 else 'w')

        vsb.pack(side='right', fill='y')
        self.periodic_tree.pack(side='left', fill='both', expand=True)

    # ── 탭3: H3 제어 / DER-AVM ───────────────────────────────────────

    def _build_control_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  (A) H3 제어 / DER-AVM  ')

        # 제어 값 (A13)
        lf1 = ttk.LabelFrame(tab, text=" (A) BODY 13: 인버터 제어 값 조회 - DER-AVM 파라미터 매핑 ", padding=5)
        lf1.pack(fill='x', pady=(0, 5))

        cols = ("No", "UDP 필드", "Modbus Definition", "Modbus Addr", "Regs", "Type", "Unit", "R/W")
        self.ctrl_tree = ttk.Treeview(lf1, columns=cols, show='headings', height=6)
        for i, c in enumerate(cols):
            self.ctrl_tree.heading(c, text=c)
        self.ctrl_tree.column("No", width=40, anchor='center')
        self.ctrl_tree.column("UDP 필드", width=160)
        self.ctrl_tree.column("Modbus Definition", width=280)
        self.ctrl_tree.column("Modbus Addr", width=100, anchor='center')
        self.ctrl_tree.column("Regs", width=50, anchor='center')
        self.ctrl_tree.column("Type", width=60, anchor='center')
        self.ctrl_tree.column("Unit", width=80, anchor='center')
        self.ctrl_tree.column("R/W", width=50, anchor='center')
        self.ctrl_tree.pack(fill='x')

        # 모니터링 (A14)
        lf2 = ttk.LabelFrame(tab, text=" (A) BODY 14: 인버터 제어 결과 모니터링 - DER-AVM 실시간 데이터 매핑 ", padding=5)
        lf2.pack(fill='both', expand=True, pady=5)

        frame2 = ttk.Frame(lf2)
        frame2.pack(fill='both', expand=True)
        self.monitor_tree = ttk.Treeview(frame2, columns=cols, show='headings', height=12)
        vsb2 = ttk.Scrollbar(frame2, orient='vertical', command=self.monitor_tree.yview)
        self.monitor_tree.configure(yscrollcommand=vsb2.set)
        for i, c in enumerate(cols):
            self.monitor_tree.heading(c, text=c)
        self.monitor_tree.column("No", width=40, anchor='center')
        self.monitor_tree.column("UDP 필드", width=160)
        self.monitor_tree.column("Modbus Definition", width=280)
        self.monitor_tree.column("Modbus Addr", width=100, anchor='center')
        self.monitor_tree.column("Regs", width=50, anchor='center')
        self.monitor_tree.column("Type", width=60, anchor='center')
        self.monitor_tree.column("Unit", width=80, anchor='center')
        self.monitor_tree.column("R/W", width=50, anchor='center')
        vsb2.pack(side='right', fill='y')
        self.monitor_tree.pack(side='left', fill='both', expand=True)

        # 요청 종류 참고
        lf3 = ttk.LabelFrame(tab, text=" (A) REQUEST 요청 종류 참고 ", padding=5)
        lf3.pack(fill='x', pady=5)
        req_cols = ("번호", "설명")
        req_tree = ttk.Treeview(lf3, columns=req_cols, show='headings', height=5)
        req_tree.heading("번호", text="요청 종류")
        req_tree.heading("설명", text="Description")
        req_tree.column("번호", width=80, anchor='center')
        req_tree.column("설명", width=400)
        for rt in APERIODIC_REQUEST_TYPES:
            req_tree.insert('', 'end', values=rt)
        req_tree.pack(fill='x')

    # ── 탭4: IV SCAN ─────────────────────────────────────────────────

    def _build_ivscan_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  (A) IV SCAN  ')

        # UDP IV SCAN 구조
        lf1 = ttk.LabelFrame(tab, text=" (A) BODY 15: IV SCAN UDP 데이터 구조 ", padding=5)
        lf1.pack(fill='x', pady=(0, 5))

        cols_s = ("필드", "크기(byte)", "타입", "비고")
        tree_s = ttk.Treeview(lf1, columns=cols_s, show='headings', height=6)
        for c in cols_s:
            tree_s.heading(c, text=c)
        tree_s.column("필드", width=180)
        tree_s.column("크기(byte)", width=100, anchor='center')
        tree_s.column("타입", width=100, anchor='center')
        tree_s.column("비고", width=400)
        for f in UDP_BODY_IV_SCAN:
            tree_s.insert('', 'end', values=(f[0], f[1], f[2], f[5]))
        tree_s.pack(fill='x')

        # Modbus IV 레지스터 매핑
        lf2 = ttk.LabelFrame(tab, text=" Modbus IV SCAN 레지스터 매핑 ", padding=5)
        lf2.pack(fill='both', expand=True, pady=5)

        frame2 = ttk.Frame(lf2)
        frame2.pack(fill='both', expand=True)

        cols = ("No", "UDP 필드", "Modbus Definition", "Modbus Addr", "Regs", "Type", "Unit", "R/W")
        self.iv_tree = ttk.Treeview(frame2, columns=cols, show='headings', height=20)
        vsb = ttk.Scrollbar(frame2, orient='vertical', command=self.iv_tree.yview)
        self.iv_tree.configure(yscrollcommand=vsb.set)
        widths = [40, 200, 280, 100, 50, 60, 80, 50]
        for i, c in enumerate(cols):
            self.iv_tree.heading(c, text=c)
            self.iv_tree.column(c, width=widths[i], anchor='center' if i != 2 else 'w')
        vsb.pack(side='right', fill='y')
        self.iv_tree.pack(side='left', fill='both', expand=True)

    # ── 탭5: 최종 레지스터맵 ─────────────────────────────────────────

    def _build_result_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  최종 레지스터맵 리스트  ')

        frame = ttk.Frame(tab)
        frame.pack(fill='both', expand=True)

        cols = ("No", "프로토콜", "구분", "UDP 필드",
                "UDP Offset", "UDP Size", "Modbus Func",
                "Modbus Addr", "Regs", "Type", "Unit", "R/W")
        self.result_tree = ttk.Treeview(frame, columns=cols, show='headings', height=30)
        vsb = ttk.Scrollbar(frame, orient='vertical', command=self.result_tree.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        widths = [35, 60, 80, 170, 80, 70, 80, 90, 45, 55, 70, 40]
        for i, c in enumerate(cols):
            self.result_tree.heading(c, text=c)
            self.result_tree.column(c, width=widths[i], anchor='center')

        hsb.pack(side='bottom', fill='x')
        vsb.pack(side='right', fill='y')
        self.result_tree.pack(side='left', fill='both', expand=True)

    # ── PDF 열기 & 파싱 ──────────────────────────────────────────────

    def _parse_excel(self, path):
        """Parse Modbus register map from Excel file (.xlsx/.xls)"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("오류", "openpyxl이 필요합니다.\npip install openpyxl")
            raise

        wb = openpyxl.load_workbook(path, data_only=True)
        if not wb.sheetnames:
            raise ValueError("Excel 파일에 시트가 없습니다")
        ws = wb[wb.sheetnames[0]]  # 첫 번째 시트

        all_registers = []
        current_fc = ''
        current_section = 'Unknown'

        for r in range(1, ws.max_row + 1):
            # Get cell values
            cells = []
            for c in range(1, min(ws.max_column + 1, 13)):
                v = ws.cell(r, c).value
                cells.append(str(v).strip() if v is not None else '')

            # Detect section headers (e.g., "1) write force single coil", "2) Read input register")
            row_text = ' '.join(cells).lower()
            if 'write' in row_text and ('coil' in row_text or 'register' in row_text or 'holding' in row_text):
                current_section = 'Inverter Parameters'
                continue
            if 'read' in row_text and ('input' in row_text or 'holding' in row_text):
                current_section = 'Inverter Realtime Data'
                continue

            # Detect FC (function code)
            for cell in cells:
                if cell.upper() in ('03H', '04H', '05H', '06H', '10H'):
                    current_fc = cell.upper().replace('H', '')
                    break

            # Skip header rows
            if any(h in cells for h in ['FC', 'reg.addr', 'ADDRESS', 'Parameter']):
                continue

            # Find register address column (looks for 30001, 40001, 1001, etc.)
            addr_str = ''
            addr_col = -1
            for ci, cell in enumerate(cells):
                if cell and any(c.isdigit() for c in cell):
                    # Match patterns: 30001, 30001~30002, 1001
                    import re
                    m = re.match(r'^(\d{4,5})(?:~.*)?$', cell.replace('`', ''))
                    if m:
                        addr_str = m.group(1)
                        addr_col = ci
                        break

            if not addr_str:
                continue

            # Parse address
            reg_num = int(addr_str)
            # Determine address offset based on register range
            if 30001 <= reg_num <= 39999:
                addr = reg_num - 30001  # FC04 input registers
            elif 40001 <= reg_num <= 49999:
                addr = reg_num - 40001  # FC03 holding registers
            elif 1001 <= reg_num <= 9999:
                addr = reg_num - 1001   # FC05 coils
            else:
                addr = reg_num - 1

            # Check for range (30033~30034 → 2 regs)
            range_match = re.search(r'~\D*(\d+)', cells[addr_col].replace('`', ''))
            if range_match:
                end_reg = int(range_match.group(1))
                regs = end_reg - reg_num + 1
            else:
                regs = 1

            # Find definition (parameter name) - usually after address columns
            definition = ''
            for ci in range(addr_col + 1, len(cells)):
                if cells[ci] and cells[ci] not in ('R', 'W', 'R/W', '-', ''):
                    c = cells[ci]
                    if not re.match(r'^[\d~]+$', c) and not re.match(r'^0x', c.lower()):
                        definition = c
                        break

            if not definition:
                continue

            # Find other fields
            unit = ''
            rw = ''
            data_type = ''
            for cell in cells:
                if cell in ('V', 'A', 'W', 'VA', 'Hz', 'kW', 'Wh', 'kWh', 'MWh', 'Sec', 'Ascii', '%'):
                    unit = cell
                if cell in ('R', 'W', 'R/W'):
                    rw = cell
                if cell.startswith(('Uint', 'Int', 'Float', 'uint', 'int', 'float')):
                    data_type = cell

            # Determine type from data_type string
            reg_type = 'U16'
            if 'float32' in data_type.lower() or 'Float32' in data_type:
                reg_type = 'Float32'
                regs = max(regs, 2)
            elif 'uint32' in data_type.lower() or 'Uint32' in data_type:
                reg_type = 'U32'
                regs = max(regs, 2)
            elif 'int16' in data_type.lower() or 'Int16' in data_type:
                reg_type = 'S16'
            elif 'int32' in data_type.lower() or 'Int32' in data_type:
                reg_type = 'S32'
                regs = max(regs, 2)

            all_registers.append({
                'section': current_section,
                'definition': definition,
                'address_hex': f'0x{addr:04X}',
                'address': addr,
                'regs': regs,
                'type': reg_type,
                'unit': unit,
                'rw': rw,
                'comment': '',
            })

        # Build parsed_data compatible with PDF parser output
        fname = os.path.basename(path)
        raw_mfr = fname.split('_')[0] if '_' in fname else fname.split('.')[0]
        manufacturer = _detect_manufacturer(raw_mfr + ' ' + fname)

        device_info = [r for r in all_registers if r['section'] == 'Unknown']
        realtime = [r for r in all_registers if r['section'] == 'Inverter Realtime Data']
        params = [r for r in all_registers if r['section'] == 'Inverter Parameters']

        return {
            'filename': fname,
            'manufacturer': manufacturer,
            'version': '',
            'all_registers': all_registers,
            'device_info': device_info,
            'realtime_data': realtime,
            'der_avm_realtime': [],
            'der_avm_params': [],
            'inverter_params': params,
            'iv_scan_regs': [],
        }

    def _open_file(self):
        path = filedialog.askopenfilename(
            title="인버터 Modbus 프로토콜 파일 열기",
            filetypes=[('Modbus Protocol', '*.pdf *.xlsx *.xls'),
                       ('PDF files', '*.pdf'),
                       ('Excel files', '*.xlsx *.xls'),
                       ('All files', '*.*')]
        )
        if not path:
            return

        # Loading 표시
        fname = os.path.basename(path)
        self.file_var.set(f"Loading: {fname} ...")
        self.info_var.set("")
        self.root.update_idletasks()

        ext = os.path.splitext(path)[1].lower()

        try:
            if ext in ('.xlsx', '.xls'):
                self.parsed_data = self._parse_excel(path)
            else:
                self.parsed_data = parse_modbus_pdf(path)
        except Exception as e:
            self.file_var.set("(파일을 열어주세요)")
            messagebox.showerror("파싱 오류", f"파일 파싱 중 오류:\n{e}")
            return

        d = self.parsed_data
        self.file_var.set(d['filename'])
        self.info_var.set(
            f"제조사: {d['manufacturer']}  |  버전: {d['version']}  |  "
            f"레지스터: {len(d['all_registers'])}개  "
            f"(Device: {len(d['device_info'])}, Realtime: {len(d['realtime_data'])}, "
            f"DER-AVM: {len(d['der_avm_realtime'])}+{len(d['der_avm_params'])}, "
            f"Params: {len(d['inverter_params'])}, IV: {len(d['iv_scan_regs'])})"
        )

        # PDF 트리 표시
        self.pdf_tree.delete(*self.pdf_tree.get_children())
        section_tags = {
            'Device Information': 'dev',
            'Inverter Realtime Data': 'rt',
            'DEA-AVM Realtime Data': 'der',
            'DER-AVM Parameters': 'derp',
            'Inverter Parameters': 'param',
            'IV Scan': 'iv',
        }
        self.pdf_tree.tag_configure('dev', background='#FFF3E0')
        self.pdf_tree.tag_configure('rt', background='#FFFFFF')
        self.pdf_tree.tag_configure('der', background='#E8F5E9')
        self.pdf_tree.tag_configure('derp', background='#C8E6C9')
        self.pdf_tree.tag_configure('param', background='#E3F2FD')
        self.pdf_tree.tag_configure('iv', background='#FCE4EC')

        for i, reg in enumerate(d['all_registers'], 1):
            tag = section_tags.get(reg.get('section', ''), '')
            self.pdf_tree.insert('', 'end', values=(
                i, reg.get('section', ''), reg.get('definition', ''),
                reg.get('address_hex', ''), reg.get('regs', ''),
                reg.get('type', ''), reg.get('unit', ''),
                reg.get('rw', ''), reg.get('comment', '')
            ), tags=(tag,))

        self.notebook.select(0)

        # ── 제조사별 Configuration 자동 설정 ──
        mfr = d.get('manufacturer', '')
        if 'Kstar' in mfr:
            self.cg_fc_var.set('FC04')
            self.cg_manufacturer_var.set('Kstar')
            self.cg_ivscan_var.set(False)
            self.cg_deravm_var.set(True)
        elif 'Huawei' in mfr:
            self.cg_fc_var.set('FC03')
            self.cg_manufacturer_var.set('Huawei')
            self.cg_ivscan_var.set(False)
            self.cg_deravm_var.set(False)
        else:
            # Solarize / VerterKing / GoodWe / Senergy / Unknown
            self.cg_fc_var.set('FC03')
            clean_mfr = mfr.split('(')[0].strip() if mfr else ''
            self.cg_manufacturer_var.set(clean_mfr or self.cg_manufacturer_var.get() or 'Solarize')
            # Only default IV/DER for Solarize-protocol inverters
            if not clean_mfr or 'Solarize' in clean_mfr or 'VerterKing' in clean_mfr:
                self.cg_ivscan_var.set(True)
                self.cg_deravm_var.set(True)

    # ── 자동 매핑 ────────────────────────────────────────────────────

    def _run_auto_mapping(self):
        if not self.parsed_data:
            messagebox.showwarning("경고", "먼저 Modbus 파일(PDF/Excel)을 열어주세요.")
            return

        body_type = int((self.body_type_var.get().split(' ')[0]) or '0')
        mppt_count = int(self.mppt_var.get())
        string_count = int(self.string_var.get())

        self.mapping = auto_map_to_udp(
            self.parsed_data, mppt_count, string_count, body_type)

        self._display_mapping()
        self.notebook.select(1)
        messagebox.showinfo("매핑 완료",
                            f"자동 매핑이 완료되었습니다.\n"
                            f"주기 기본: {len(self.mapping['periodic_basic'])}개\n"
                            f"MPPT: {len(self.mapping['periodic_mppt'])}개\n"
                            f"STRING: {len(self.mapping['periodic_string'])}개\n"
                            f"H3 제어값: {len(self.mapping['control_values'])}개\n"
                            f"제어 모니터링: {len(self.mapping['control_monitor'])}개\n"
                            f"IV SCAN: {len(self.mapping['iv_scan'])}개")

    def _display_mapping(self):
        if not self.mapping:
            return

        # 주기 데이터 탭
        self.periodic_tree.delete(*self.periodic_tree.get_children())
        self.periodic_tree.tag_configure('basic', background='#FFFFFF')
        self.periodic_tree.tag_configure('mppt', background='#E8F5E9')
        self.periodic_tree.tag_configure('string', background='#E3F2FD')
        self.periodic_tree.tag_configure('unmapped', background='#FFEBEE')

        no = 1
        for m in self.mapping['periodic_basic']:
            tag = 'unmapped' if not m.get('mb_addr') and not m.get('mb_addr_hex') else 'basic'
            self.periodic_tree.insert('', 'end', values=(
                no, "기본", m['udp_field'], m.get('mb_definition', ''),
                m.get('mb_addr_hex', ''), m.get('mb_regs', ''), m.get('mb_type', ''),
                m.get('mb_unit', ''), m.get('mb_rw', '')
            ), tags=(tag,))
            no += 1
        for m in self.mapping.get('periodic_mppt', []):
            self.periodic_tree.insert('', 'end', values=(
                no, "MPPT", m['udp_field'], m.get('mb_definition', ''),
                m.get('mb_addr_hex', ''), m.get('mb_regs', ''), m.get('mb_type', ''),
                m.get('mb_unit', ''), m.get('mb_rw', '')
            ), tags=('mppt',))
            no += 1
        for m in self.mapping.get('periodic_string', []):
            self.periodic_tree.insert('', 'end', values=(
                no, "STRING", m['udp_field'], m.get('mb_definition', ''),
                m.get('mb_addr_hex', ''), m.get('mb_regs', ''), m.get('mb_type', ''),
                m.get('mb_unit', ''), m.get('mb_rw', '')
            ), tags=('string',))
            no += 1

        # H3 제어 탭
        self.ctrl_tree.delete(*self.ctrl_tree.get_children())
        for i, m in enumerate(self.mapping.get('control_values', []), 1):
            self.ctrl_tree.insert('', 'end', values=(
                i, m['udp_field'], m.get('mb_definition', ''),
                m.get('mb_addr_hex', ''), m.get('mb_regs', ''), m.get('mb_type', ''),
                m.get('mb_unit', ''), m.get('mb_rw', '')
            ))

        self.monitor_tree.delete(*self.monitor_tree.get_children())
        for i, m in enumerate(self.mapping.get('control_monitor', []), 1):
            self.monitor_tree.insert('', 'end', values=(
                i, m['udp_field'], m.get('mb_definition', ''),
                m.get('mb_addr_hex', ''), m.get('mb_regs', ''), m.get('mb_type', ''),
                m.get('mb_unit', ''), m.get('mb_rw', '')
            ))

        # IV SCAN 탭
        self.iv_tree.delete(*self.iv_tree.get_children())
        for i, m in enumerate(self.mapping.get('iv_scan', []), 1):
            self.iv_tree.insert('', 'end', values=(
                i, m['udp_field'], m.get('mb_definition', ''),
                m.get('mb_addr_hex', ''), m.get('mb_regs', ''), m.get('mb_type', ''),
                m.get('mb_unit', ''), m.get('mb_rw', '')
            ))

    # ── 탭 변경 이벤트 ─────────────────────────────────────────────

    def _on_tab_changed(self, event):
        """탭 전환 시 Code Generator 탭이면 JSON 자동 로드."""
        try:
            tab_idx = self.notebook.index(self.notebook.select())
            if tab_idx == 5:  # Code Generator tab (0-indexed)
                self._auto_load_json_for_codegen()
        except Exception:
            pass

    def _auto_load_json_for_codegen(self):
        """현재 로드된 PDF의 제조사에 맞는 JSON을 자동 로드."""
        import os, glob

        # 제조사 감지
        manufacturer = ''
        if hasattr(self, 'parsed_data') and self.parsed_data:
            manufacturer = self.parsed_data.get('manufacturer', '')

        if not manufacturer:
            # PDF가 로드되지 않은 경우 → JSON 파일 직접 선택
            from tkinter import filedialog
            docs_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'docs')
            json_path = filedialog.askopenfilename(
                title="레지스터맵 JSON 선택",
                initialdir=docs_dir,
                filetypes=[('JSON files', '*_register_map.json'), ('All files', '*.*')])
            if json_path:
                try:
                    self._auto_loading = True
                    self._import_json_file(json_path)
                    self._json_loaded_for = 'manual'
                    self._auto_loading = False
                except Exception as e:
                    self._auto_loading = False
                    from tkinter import messagebox
                    messagebox.showerror("JSON 로드 실패", str(e))
            return

        # 중복 로드 방지
        if self._json_loaded_for == manufacturer:
            return

        # PDF 파일명으로 Solarize 구분 (레지스터맵 로드)
        pdf_filename = self.file_var.get().lower() if hasattr(self, 'file_var') else ''

        # 제조사별 JSON 파일 매칭
        mfr_lower = manufacturer.lower()
        json_name_map = {
            'huawei': 'huawei_register_map.json',
            'kstar': 'kstar_register_map.json',
            'sungrow': 'sungrow_register_map.json',
            'ekos': 'ekos_register_map.json',
            'modbus map-ek': 'ekos_register_map.json',
        }

        # Solarize 브랜드
        json_filename = None
        if 'solarize' in mfr_lower or 'verterking' in mfr_lower:
            json_filename = 'solarize_register_map.json'
        else:
            for key, fname in json_name_map.items():
                if key in mfr_lower:
                    json_filename = fname
                    break

        if not json_filename:
            # 와일드카드로 검색
            docs_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'docs')
            pattern = os.path.join(docs_dir, f'*{mfr_lower}*_register_map.json')
            matches = glob.glob(pattern)
            if matches:
                json_filename = os.path.basename(matches[0])

        if not json_filename:
            from tkinter import messagebox
            messagebox.showinfo("JSON 없음",
                f"'{manufacturer}' 제조사의 JSON 레지스터맵을 찾을 수 없습니다.\n\n"
                f"docs/ 폴더에 {{제조사}}_register_map.json 파일이 필요합니다.\n"
                f"add-inverter skill로 먼저 생성하세요.")
            return

        # JSON 파일 경로 확인
        docs_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'docs')
        json_path = os.path.join(docs_dir, json_filename)

        if not os.path.isfile(json_path):
            return

        # JSON 로드 (자동 로드 시 messagebox/탭전환 억제)
        try:
            self._auto_loading = True
            self._import_json_file(json_path)
            self._json_loaded_for = manufacturer
            self._auto_loading = False
        except Exception as e:
            self._auto_loading = False
            from tkinter import messagebox
            messagebox.showerror("JSON 로드 실패", f"JSON 파일 로드 중 오류:\n{e}")

    # ── 최종 레지스터맵 생성 ─────────────────────────────────────────

    def _generate_register_map(self):
        # mapping이 없거나 periodic_basic이 비어있으면 JSON 자동 로드 시도
        has_valid_mapping = (self.mapping
                            and self.mapping.get('periodic_basic')
                            and len(self.mapping['periodic_basic']) > 0)
        if not has_valid_mapping:
            # JSON 자동 검색: docs/ 폴더에서 {manufacturer}_register_map.json
            json_loaded = self._try_load_json_auto()
            if json_loaded:
                return  # _import_json_file이 _generate_register_map을 다시 호출
            messagebox.showwarning("경고",
                "레지스터맵을 생성할 수 없습니다.\n\n"
                "방법 1: PDF를 열고 Code Generator 탭에서 JSON 자동 로드\n"
                "방법 2: docs/ 폴더에 {제조사}_register_map.json 파일 준비\n"
                "  (add-inverter skill로 JSON 먼저 생성)")
            return

        self.result_tree.delete(*self.result_tree.get_children())
        no = 1

        # (P) HEADER
        offset = 0
        for h in UDP_HEADER_FIELDS:
            self.result_tree.insert('', 'end', values=(
                no, "(P)", "HEADER", h[0], offset, h[1],
                "-", "-", "-", h[2], "", ""
            ))
            offset += h[1]
            no += 1

        header_size = offset

        # Function Code 결정
        fc_str = getattr(self, 'cg_fc_var', None)
        fc_display = fc_str.get() if fc_str else 'FC03'
        fc_hex = '0x04' if '04' in fc_display else '0x03'

        # (P) BODY 기본
        body_offset = 0
        for i, m in enumerate(self.mapping['periodic_basic']):
            udp_def = UDP_BODY_INVERTER_BASIC[i] if i < len(UDP_BODY_INVERTER_BASIC) else None
            udp_size = udp_def[1] if udp_def else 2
            self.result_tree.insert('', 'end', values=(
                no, "(P)", "기본", m['udp_field'],
                f"{header_size}+{body_offset}", udp_size,
                fc_hex, m['mb_addr_hex'], m['mb_regs'],
                m['mb_type'], m['mb_unit'], m.get('mb_rw', '')
            ))
            body_offset += udp_size
            no += 1

        # MPPT
        for m in self.mapping['periodic_mppt']:
            self.result_tree.insert('', 'end', values=(
                no, "(P)", "MPPT", m['udp_field'],
                f"{header_size}+{body_offset}", 2,
                fc_hex, m['mb_addr_hex'], m['mb_regs'],
                m['mb_type'], m['mb_unit'], m.get('mb_rw', '')
            ))
            body_offset += 2
            no += 1

        # STRING
        for m in self.mapping['periodic_string']:
            self.result_tree.insert('', 'end', values=(
                no, "(P)", "STRING", m['udp_field'],
                f"{header_size}+{body_offset}", 2,
                fc_hex, m['mb_addr_hex'], m['mb_regs'],
                m['mb_type'], m['mb_unit'], m.get('mb_rw', '')
            ))
            body_offset += 2
            no += 1

        # 구분선
        self.result_tree.insert('', 'end', values=(
            "", "", "", f"── (P) 합계: HEADER={header_size}B + BODY={body_offset}B = {header_size+body_offset}B ──",
            "", "", "", "", "", "", "", ""
        ))

        # (A) H3 제어 값 (BODY 13)
        for m in self.mapping['control_values']:
            self.result_tree.insert('', 'end', values=(
                no, "(A)", "H3 제어", m['udp_field'],
                "-", 2, "0x03/0x06", m['mb_addr_hex'],
                m['mb_regs'], m['mb_type'], m['mb_unit'], m.get('mb_rw', '')
            ))
            no += 1

        # (A) 제어 모니터링 (BODY 14)
        for m in self.mapping['control_monitor']:
            self.result_tree.insert('', 'end', values=(
                no, "(A)", "DER-AVM", m['udp_field'],
                "-", 4, "0x03", m['mb_addr_hex'],
                m['mb_regs'], m['mb_type'], m['mb_unit'], m.get('mb_rw', '')
            ))
            no += 1

        # (A) IV SCAN (BODY 15)
        for m in self.mapping['iv_scan']:
            self.result_tree.insert('', 'end', values=(
                no, "(A)", "IV SCAN", m['udp_field'],
                "-", "-", "0x03/0x06", m['mb_addr_hex'],
                m['mb_regs'], m['mb_type'], m['mb_unit'], m.get('mb_rw', '')
            ))
            no += 1

        if not getattr(self, '_auto_loading', False):
            self.notebook.select(4)
            messagebox.showinfo("생성 완료",
                                f"레지스터맵 생성 완료.\n총 {no-1}개 항목")

    # ── 내보내기 / 불러오기 ──────────────────────────────────────────

    def _get_result_data(self):
        cols = [self.result_tree.heading(c)['text']
                for c in self.result_tree['columns']]
        rows = []
        for item in self.result_tree.get_children():
            vals = self.result_tree.item(item, 'values')
            rows.append(dict(zip(cols, vals)))
        return cols, rows

    def _get_export_filename(self, ext):
        """제조사별 내보내기 파일명 생성"""
        mfr = self.parsed_data.get('manufacturer', '') if self.parsed_data else ''
        name_map = {
            'Solarize': 'VK50K',
            'Huawei': 'HUAwei50K',
            'Kstar': 'KSTAR60K',
        }
        prefix = name_map.get(mfr, 'modbus_udp')
        return f'{prefix}_register_map.{ext}'

    def _export_csv(self):
        if not self.result_tree.get_children():
            messagebox.showwarning("경고", "먼저 레지스터맵을 생성해주세요.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension='.csv',
            filetypes=[('CSV', '*.csv')],
            initialfile=self._get_export_filename('csv')
        )
        if not path:
            return
        cols, rows = self._get_result_data()
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows(rows)
        messagebox.showinfo("완료", f"CSV 저장: {path}")

    def _export_json(self):
        if not self.mapping:
            messagebox.showwarning("경고", "먼저 매핑을 실행해주세요.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension='.json',
            filetypes=[('JSON', '*.json')],
            initialfile=self._get_export_filename('json')
        )
        if not path:
            return

        _, result_rows = self._get_result_data()
        export = {
            'source_pdf': self.parsed_data['filename'] if self.parsed_data else '',
            'manufacturer': self.parsed_data['manufacturer'] if self.parsed_data else '',
            'version': self.parsed_data['version'] if self.parsed_data else '',
            'body_type': int((self.body_type_var.get().split(' ')[0]) or '0'),
            'mppt_count': int(self.mppt_var.get()),
            'string_count': int(self.string_var.get()),
            'mapping': self.mapping,
            'register_map': result_rows,
        }
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(export, f, ensure_ascii=False, indent=2, default=str)
        messagebox.showinfo("완료", f"JSON 저장: {path}")

    def _try_load_json_auto(self):
        """docs/ 폴더에서 제조사 이름 기반 JSON 자동 검색 후 로드"""
        docs_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'docs')
        if not os.path.isdir(docs_dir):
            return False

        # 제조사명에서 검색 키워드 추출
        manufacturer = ''
        if hasattr(self, 'parsed_data') and self.parsed_data:
            manufacturer = self.parsed_data.get('manufacturer', '')
        if not manufacturer and hasattr(self, 'cg_manufacturer_var'):
            manufacturer = self.cg_manufacturer_var.get().strip()

        # docs/에서 *_register_map.json 파일 검색
        json_files = [f for f in os.listdir(docs_dir) if f.endswith('_register_map.json')]
        if not json_files:
            return False

        # 제조사명으로 매칭 시도
        matched = None
        if manufacturer:
            mfr_lower = manufacturer.lower().replace(' ', '')
            for jf in json_files:
                if mfr_lower in jf.lower().replace('_', '').replace('-', ''):
                    matched = jf
                    break

        if not matched:
            # 매칭 실패 → 파일 선택 다이얼로그
            if len(json_files) == 1:
                matched = json_files[0]
            else:
                path = filedialog.askopenfilename(
                    title="레지스터맵 JSON 선택",
                    initialdir=docs_dir,
                    filetypes=[('JSON', '*.json')]
                )
                if not path:
                    return False
                self._import_json_file(path)
                return True

        json_path = os.path.join(docs_dir, matched)
        self._import_json_file(json_path)
        return True

    def _import_json_file(self, path):
        """JSON 파일에서 레지스터맵 로드"""
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if 'mapping' not in data:
            messagebox.showerror("오류", "올바른 매핑 JSON이 아닙니다.\n'mapping' 키가 필요합니다.")
            return

        bt = data.get('body_type', 4)
        bt_map = {1: '1 - 기본', 2: '2 - 기본+MPPT',
                  3: '3 - 기본+STRING', 4: '4 - 기본+MPPT+STRING',
                  5: '5 - 단상 기본'}
        self.body_type_var.set(bt_map.get(bt, '4 - 기본+MPPT+STRING'))
        self.mppt_var.set(str(data.get('mppt_count', 4)))
        self.string_var.set(str(data.get('string_count', 8)))

        self.mapping = data['mapping']
        # Ensure all expected mapping keys exist
        for key in ('periodic_basic', 'periodic_mppt', 'periodic_string',
                     'control_values', 'control_monitor', 'iv_scan'):
            if key not in self.mapping:
                self.mapping[key] = []
        self._loaded_status_map = data.get('status_map', None)
        self._loaded_error_codes = data.get('error_codes', None)
        self._display_mapping()

        source = data.get('source_pdf', os.path.basename(path))
        manufacturer = data.get('manufacturer', '')
        version = data.get('version', '')
        self.file_var.set(source)
        self.info_var.set(f"제조사: {manufacturer} | 버전: {version}")

        if not self.parsed_data:
            self.parsed_data = {}
        self.parsed_data['manufacturer'] = manufacturer
        self.parsed_data['version'] = version
        self.parsed_data['device_info'] = data.get('device_info', [])

        if manufacturer:
            self.cg_manufacturer_var.set(manufacturer)
        fc = data.get('function_code', 'FC03')
        self.cg_fc_var.set(fc)
        if data.get('iv_scan') is not None:
            self.cg_ivscan_var.set(bool(data.get('iv_scan', False)))
        if data.get('der_avm') is not None:
            self.cg_deravm_var.set(bool(data.get('der_avm', False)))

        self._generate_register_map()

        if not getattr(self, '_auto_loading', False):
            self.notebook.select(4)
            messagebox.showinfo("완료",
                f"레지스터맵 로드 완료: {source}\n"
                f"제조사: {manufacturer}\n"
                f"Code Generator에서 Generate Code를 실행하세요.")

    def _import_json(self):
        """수동 JSON 열기 (내부 호출용, 버튼 제거됨)"""
        path = filedialog.askopenfilename(
            title="레지스터맵 JSON 열기",
            filetypes=[('JSON', '*.json'), ('All files', '*.*')]
        )
        if path:
            self._import_json_file(path)

    # ── 탭6: Code Generator ──────────────────────────────────────────

    def _build_codegen_tab(self):
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text='  Code Generator  ')

        # ── Top: Configuration ──
        cfg_frame = ttk.LabelFrame(tab, text=" Configuration ", padding=8)
        cfg_frame.pack(fill='x', pady=(0, 5))

        # Row 0
        ttk.Label(cfg_frame, text="Manufacturer Name:").grid(row=0, column=0, sticky='e', padx=(5, 2))
        self.cg_manufacturer_var = tk.StringVar(value='Solarize')
        ttk.Entry(cfg_frame, textvariable=self.cg_manufacturer_var, width=18).grid(row=0, column=1, padx=2)

        ttk.Label(cfg_frame, text="Class Name:").grid(row=0, column=2, sticky='e', padx=(10, 2))
        self.cg_classname_var = tk.StringVar(value='RegisterMap')
        ttk.Entry(cfg_frame, textvariable=self.cg_classname_var, width=22).grid(row=0, column=3, padx=2)

        ttk.Label(cfg_frame, text="Protocol Name:").grid(row=0, column=4, sticky='e', padx=(10, 2))
        self.cg_protocol_var = tk.StringVar(value='solarize')
        ttk.Entry(cfg_frame, textvariable=self.cg_protocol_var, width=14).grid(row=0, column=5, padx=2)

        ttk.Label(cfg_frame, text="Model Number:").grid(row=0, column=6, sticky='e', padx=(10, 2))
        self.cg_model_var = tk.StringVar(value='1')
        ttk.Entry(cfg_frame, textvariable=self.cg_model_var, width=6).grid(row=0, column=7, padx=2)

        # Row 1
        ttk.Label(cfg_frame, text="Function Code:").grid(row=1, column=0, sticky='e', padx=(5, 2), pady=(5, 0))
        self.cg_fc_var = tk.StringVar(value='FC03')
        ttk.Combobox(cfg_frame, textvariable=self.cg_fc_var, values=['FC03', 'FC04'],
                      state='readonly', width=8).grid(row=1, column=1, padx=2, pady=(5, 0), sticky='w')

        self.cg_ivscan_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(cfg_frame, text="IV Scan Support",
                         variable=self.cg_ivscan_var).grid(row=1, column=2, columnspan=2, sticky='w', padx=10, pady=(5, 0))

        self.cg_deravm_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(cfg_frame, text="DER-AVM Control Support",
                         variable=self.cg_deravm_var).grid(row=1, column=4, columnspan=2, sticky='w', padx=10, pady=(5, 0))

        # ── 3-Stage Pipeline ──
        self._build_pipeline_panel(tab)

        # ── Middle: PanedWindow (Register Files | Code Preview) ──
        paned = tk.PanedWindow(tab, orient='horizontal', sashwidth=5, sashrelief='raised')
        paned.pack(fill='both', expand=True, pady=5)

        # ── Left pane: Register Files ──
        left_frame = ttk.LabelFrame(paned, text=" Register Files ", padding=5)

        rf_inner = ttk.Frame(left_frame)
        rf_inner.pack(fill='both', expand=True)

        cols = ('device_type', 'model', 'model_name', 'filename', 'ref')
        self.rf_tree = ttk.Treeview(rf_inner, columns=cols, show='headings', height=8)
        self.rf_tree.heading('device_type', text='Type')
        self.rf_tree.heading('model', text='#')
        self.rf_tree.heading('model_name', text='Model Name')
        self.rf_tree.heading('filename', text='Filename')
        self.rf_tree.heading('ref', text='Ref')
        self.rf_tree.column('device_type', width=65, anchor='center')
        self.rf_tree.column('model', width=25, anchor='center')
        self.rf_tree.column('model_name', width=130)
        self.rf_tree.column('filename', width=130)
        self.rf_tree.column('ref', width=55, anchor='center')
        rf_vsb = ttk.Scrollbar(rf_inner, orient='vertical', command=self.rf_tree.yview)
        self.rf_tree.configure(yscrollcommand=rf_vsb.set)
        self.rf_tree.pack(side='left', fill='both', expand=True)
        rf_vsb.pack(side='right', fill='y')
        self.rf_tree.bind('<Double-1>', lambda e: self._on_open_selected())

        rf_btn = ttk.Frame(left_frame)
        rf_btn.pack(fill='x', pady=(5, 0))
        ttk.Button(rf_btn, text=" Open ",
                   command=self._on_open_selected).pack(side='left', padx=2)
        ttk.Button(rf_btn, text=" Delete ",
                   command=self._on_delete_selected).pack(side='left', padx=2)
        ttk.Button(rf_btn, text=" Refresh ",
                   command=self._refresh_register_files).pack(side='left', padx=2)
        ttk.Button(rf_btn, text=" Manage Ref ",
                   command=self._sp_manage_references).pack(side='left', padx=2)

        self._refresh_register_files()

        paned.add(left_frame, width=400, minsize=250)

        # ── Right pane: Code Preview ──
        right_frame = ttk.LabelFrame(paned, text=" Code Preview ", padding=5)

        preview_inner = ttk.Frame(right_frame)
        preview_inner.pack(fill='both', expand=True)

        self.cg_text = tk.Text(preview_inner, wrap='none', font=('Consolas', 9),
                                state='disabled', height=30)
        vsb = ttk.Scrollbar(preview_inner, orient='vertical', command=self.cg_text.yview)
        hsb = ttk.Scrollbar(preview_inner, orient='horizontal', command=self.cg_text.xview)
        self.cg_text.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side='bottom', fill='x')
        vsb.pack(side='right', fill='y')
        self.cg_text.pack(side='left', fill='both', expand=True)

        # Buttons inside Code Preview frame (below text)
        btn_frame = ttk.Frame(right_frame)
        btn_frame.pack(fill='x', pady=(5, 0))

        ttk.Button(btn_frame, text="  Generate Code  ",
                   command=self._on_generate_code).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="  Test Code  ",
                   command=self._on_test_code).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="  Save to common/  ",
                   command=self._on_save_code).pack(side='left', padx=5)
        ttk.Button(btn_frame, text="  Add Model  ",
                   command=self._on_add_model).pack(side='left', padx=5)

        # ── AI Generation buttons ──
        ttk.Separator(btn_frame, orient='vertical').pack(side='left', fill='y', padx=8)
        ai_btn = ttk.Button(btn_frame, text="  AI Generate  ",
                            command=self._on_ai_generate)
        ai_btn.pack(side='left', padx=5)
        ttk.Button(btn_frame, text="  API Settings  ",
                   command=self._on_api_settings).pack(side='left', padx=2)

        paned.add(right_frame, minsize=400)

        # ── Third pane: Config Editor ──
        config_frame = ttk.LabelFrame(paned, text=" Config Editor ", padding=5)

        # Config file selector
        cfg_sel_frame = ttk.Frame(config_frame)
        cfg_sel_frame.pack(fill='x', pady=(0, 5))

        base_dir = os.path.dirname(os.path.abspath(__file__))
        self._config_dir = os.path.normpath(os.path.join(base_dir, '..', 'config'))
        config_files = []
        if os.path.isdir(self._config_dir):
            config_files = sorted(
                f for f in os.listdir(self._config_dir)
                if f.endswith(('.ini', '.config', '.cfg'))
            )

        self.cfg_file_var = tk.StringVar()
        cfg_combo = ttk.Combobox(cfg_sel_frame, textvariable=self.cfg_file_var,
                                  values=config_files, state='readonly', width=25)
        cfg_combo.pack(side='left', fill='x', expand=True)
        cfg_combo.bind('<<ComboboxSelected>>', self._on_config_selected)

        # Config text editor
        cfg_text_frame = ttk.Frame(config_frame)
        cfg_text_frame.pack(fill='both', expand=True)

        self.cfg_editor = tk.Text(cfg_text_frame, wrap='none', font=('Consolas', 9))
        cfg_vsb = ttk.Scrollbar(cfg_text_frame, orient='vertical', command=self.cfg_editor.yview)
        cfg_hsb = ttk.Scrollbar(cfg_text_frame, orient='horizontal', command=self.cfg_editor.xview)
        self.cfg_editor.configure(yscrollcommand=cfg_vsb.set, xscrollcommand=cfg_hsb.set)
        cfg_hsb.pack(side='bottom', fill='x')
        cfg_vsb.pack(side='right', fill='y')
        self.cfg_editor.pack(side='left', fill='both', expand=True)

        # Save button
        cfg_btn_frame = ttk.Frame(config_frame)
        cfg_btn_frame.pack(fill='x', pady=(5, 0))
        ttk.Button(cfg_btn_frame, text="  Save Config  ",
                   command=self._on_save_config).pack(side='left', padx=5)
        ttk.Button(cfg_btn_frame, text="  Build Package  ",
                   command=self._on_build_package).pack(side='left', padx=5)

        paned.add(config_frame, width=350, minsize=200)

        # Auto-load first config file
        if config_files:
            self.cfg_file_var.set(config_files[0])
            self._on_config_selected(None)

        # ── Config Guide ──
        guide_frame = ttk.LabelFrame(tab, text=" Config Guide ", padding=5)
        guide_frame.pack(fill='both', expand=True, pady=(0, 5))

        self.cg_guide_text = tk.Text(guide_frame, wrap='word', font=('Consolas', 9),
                                      state='disabled', height=14, background='#F5F5F5')
        self.cg_guide_text.pack(fill='both', expand=True)

    # ── 3-Stage Pipeline UI ───────────────────────────────────────────

    def _build_pipeline_panel(self, parent):
        """Build the 3-Stage Pipeline LabelFrame inside the Code Generator tab."""
        pf = ttk.LabelFrame(parent,
                            text=" 3-Stage Register File Generation Pipeline ",
                            padding=8)
        pf.pack(fill='x', pady=(0, 5))

        # ── Mode selector row ──
        mode_row = ttk.Frame(pf)
        mode_row.pack(fill='x', pady=(0, 6))

        ttk.Label(mode_row, text="Mode:", font=('맑은 고딕', 9, 'bold')).pack(side='left', padx=(0, 6))

        self._sp_mode_var = tk.StringVar(value='offline')

        rb_off = ttk.Radiobutton(mode_row, text="Mode 1: Offline  (PyMuPDF + Reference Library)",
                                  variable=self._sp_mode_var, value='offline',
                                  command=self._sp_on_mode_change)
        rb_off.pack(side='left', padx=(0, 12))

        rb_ai = ttk.Radiobutton(mode_row, text="Mode 2: AI  (PyMuPDF + Claude API, auto-retry x3)",
                                 variable=self._sp_mode_var, value='ai',
                                 command=self._sp_on_mode_change)
        rb_ai.pack(side='left', padx=(0, 12))

        # Reference library status
        ref_count = 0
        if _sp is not None:
            try:
                rm = _sp._get_ref_manager()
                if rm:
                    ref_count = rm.count()
            except Exception:
                pass
        self._sp_ref_status_var = tk.StringVar(value=f"Reference library: {ref_count} sets")
        ttk.Label(mode_row, textvariable=self._sp_ref_status_var,
                  foreground='#0066CC').pack(side='left', padx=(8, 0))

        ttk.Separator(pf, orient='horizontal').pack(fill='x', pady=(0, 6))

        # ── Step 1 ──
        s1 = ttk.LabelFrame(pf, text=" Step 1: PDF → Excel (Register Extraction) ", padding=5)
        s1.pack(fill='x', pady=(0, 4))

        s1_row = ttk.Frame(s1)
        s1_row.pack(fill='x')

        ttk.Label(s1_row, text="PDF:").pack(side='left')
        self._sp_pdf_var = tk.StringVar(value="(no PDF selected)")
        ttk.Label(s1_row, textvariable=self._sp_pdf_var,
                  foreground='#555555', width=44).pack(side='left', padx=(4, 8))

        ttk.Button(s1_row, text="  Open PDF  ",
                   command=self._sp_open_pdf).pack(side='left', padx=3)
        ttk.Button(s1_row, text="  Generate Excel  ",
                   command=self._sp_run_stage1).pack(side='left', padx=3)
        ttk.Button(s1_row, text="  Load Existing  ",
                   command=self._sp_load_stage1).pack(side='left', padx=3)

        self._sp_s1_status_var = tk.StringVar(value="Ready")
        self._sp_s1_status_lbl = ttk.Label(s1_row, textvariable=self._sp_s1_status_var,
                  foreground='#444444', width=35)
        self._sp_s1_status_lbl.pack(side='left', padx=(8, 0))

        # ── Step 2 ──
        s2 = ttk.LabelFrame(pf, text=" Step 2: Excel → Mapping Excel (Solarize Standard Matching) ", padding=5)
        s2.pack(fill='x', pady=(0, 4))

        s2_row = ttk.Frame(s2)
        s2_row.pack(fill='x')

        ttk.Label(s2_row, text="Stage 1 Excel:").pack(side='left')
        self._sp_s1_path_var = tk.StringVar(value="(run Step 1 first)")
        ttk.Label(s2_row, textvariable=self._sp_s1_path_var,
                  foreground='#555555', width=42).pack(side='left', padx=(4, 8))

        ttk.Button(s2_row, text="  Generate Mapping  ",
                   command=self._sp_run_stage2).pack(side='left', padx=3)
        ttk.Button(s2_row, text="  Open Mapping Excel  ",
                   command=self._sp_open_stage2).pack(side='left', padx=3)

        self._sp_s2_status_var = tk.StringVar(value="Ready")
        ttk.Label(s2_row, textvariable=self._sp_s2_status_var,
                  foreground='#444444', width=35).pack(side='left', padx=(8, 0))

        # ── Step 3 ──
        s3 = ttk.LabelFrame(pf, text=" Step 3: Mapping Excel → Register .py File ", padding=5)
        s3.pack(fill='x')

        s3_cfg = ttk.Frame(s3)
        s3_cfg.pack(fill='x', pady=(0, 4))

        # Settings row
        ttk.Label(s3_cfg, text="MPPT:").pack(side='left')
        self._sp_mppt_var = tk.StringVar(value='4')
        ttk.Spinbox(s3_cfg, from_=1, to=8, width=3,
                    textvariable=self._sp_mppt_var).pack(side='left', padx=(2, 8))

        ttk.Label(s3_cfg, text="Strings:").pack(side='left')
        self._sp_string_var = tk.StringVar(value='8')
        ttk.Spinbox(s3_cfg, from_=1, to=24, width=3,
                    textvariable=self._sp_string_var).pack(side='left', padx=(2, 8))

        self._sp_iv_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(s3_cfg, text="IV Scan",
                        variable=self._sp_iv_var).pack(side='left', padx=(0, 6))

        self._sp_der_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(s3_cfg, text="DER-AVM",
                        variable=self._sp_der_var).pack(side='left', padx=(0, 6))

        self._sp_dea_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(s3_cfg, text="DEA-AVM Monitor",
                        variable=self._sp_dea_var).pack(side='left', padx=(0, 6))

        s3_btn = ttk.Frame(s3)
        s3_btn.pack(fill='x')

        ttk.Label(s3_btn, text="Mapping Excel:").pack(side='left')
        self._sp_s2_path_var = tk.StringVar(value="(run Step 2 first)")
        ttk.Label(s3_btn, textvariable=self._sp_s2_path_var,
                  foreground='#555555', width=38).pack(side='left', padx=(4, 8))

        ttk.Button(s3_btn, text="  Generate .py  ",
                   command=self._sp_run_stage3).pack(side='left', padx=3)
        ttk.Button(s3_btn, text="  Test  ",
                   command=self._sp_test_stage3).pack(side='left', padx=3)
        ttk.Button(s3_btn, text="  Save to common/  ",
                   command=self._sp_save_stage3).pack(side='left', padx=3)
        ttk.Button(s3_btn, text="  Save to Reference  ",
                   command=self._sp_save_to_reference).pack(side='left', padx=3)

        self._sp_s3_status_var = tk.StringVar(value="Ready")
        ttk.Label(s3_btn, textvariable=self._sp_s3_status_var,
                  foreground='#444444', width=30).pack(side='left', padx=(8, 0))

    # ── Stage Pipeline Handlers ───────────────────────────────────────

    def _sp_open_pdf(self):
        """Stage 1: Let user pick a PDF file."""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Select Modbus Protocol PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if not path:
            return
        self._sp_pdf_path = path
        fname = os.path.basename(path)
        self._sp_pdf_var.set(fname[:50])
        self._sp_s1_status_var.set("PDF loaded — click Generate Excel")
        # If PDF not yet parsed via main button, parse it now
        if self.parsed_data is None or self.parsed_data.get('filepath') != path:
            self.file_var.set(f"Loading: {os.path.basename(path)} ...")
            self.root.update_idletasks()
            try:
                self.parsed_data = parse_modbus_pdf(path)
                self.file_var.set(os.path.basename(path))
                d = self.parsed_data
                mfr = d.get('manufacturer', '')
                cnt = len(d.get('all_registers', []))
                self.info_var.set(f"{mfr}  |  {cnt} registers")
            except Exception as e:
                self.file_var.set(os.path.basename(path))
                self._sp_s1_status_var.set(f"Parse warning: {e}")

    def _sp_load_stage1(self):
        """Load an existing Stage 1 Excel file."""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Select Stage 1 Register Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        self._sp_stage1_path = path
        self._sp_s1_path_var.set(os.path.basename(path)[:50])
        self._sp_s1_status_var.set("Stage 1 Excel loaded")

    def _sp_run_stage1(self):
        """Run Stage 1: PDF → Excel in a background thread."""
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        pdf_path = self._sp_pdf_path
        if not pdf_path:
            # Try to get from the file_var
            fname = self.file_var.get()
            if fname and fname != "(PDF 파일을 열어주세요)":
                if self.parsed_data:
                    pdf_path = self.parsed_data.get('filepath', '')

        if not pdf_path:
            messagebox.showwarning("Stage 1", "Open a PDF file first (Open PDF button).")
            return

        mode    = self._sp_mode_var.get()
        api_key = ''
        model   = None
        if mode == 'ai':
            if _ai_gen:
                api_key = _ai_gen.load_api_key()
                model   = _ai_gen.load_model_name()
            if not api_key:
                messagebox.showwarning("AI Mode", "No API key configured.\n"
                                       "Set it via 'API Settings' or switch to Offline mode.")
                return

        self._sp_s1_status_var.set(f"Running Stage 1 [{mode}]...")
        self.root.update_idletasks()

        def _run():
            try:
                out_path = _sp.stage1_extract_to_excel(
                    parsed_data  = self.parsed_data,
                    pdf_path     = pdf_path,
                    mode         = mode,
                    api_key      = api_key,
                    model        = model,
                    progress_cb  = lambda m: self.root.after(0, lambda: self._sp_s1_status_var.set(m[-40:]))
                )
                self._sp_stage1_path = out_path
                self.root.after(0, lambda: self._sp_s1_path_var.set(os.path.basename(out_path)[:50]))
                self.root.after(0, lambda: self._sp_s1_status_var.set(f"Done: {os.path.basename(out_path)}"))
                self.root.after(0, lambda: messagebox.showinfo(
                    "Stage 1 Complete",
                    f"Register Excel saved:\n{out_path}\n\n"
                    "Open the Excel file to review and edit,\n"
                    "then run Step 2 (Generate Mapping)."))
            except Exception as e:
                self.root.after(0, lambda: self._sp_s1_status_var.set(f"Error: {e}"))
                self.root.after(0, lambda: messagebox.showerror("Stage 1 Error", str(e)))

        threading.Thread(target=_run, daemon=True).start()

    def _sp_run_stage2(self):
        """Run Stage 2: Stage 1 Excel → Mapping Excel in background thread."""
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        stage1_path = self._sp_stage1_path
        if not stage1_path or not os.path.isfile(stage1_path):
            messagebox.showwarning("Stage 2",
                                   "No Stage 1 Excel found.\nRun Step 1 first (or load existing).")
            return

        mode    = self._sp_mode_var.get()
        api_key = ''
        model   = None
        if mode == 'ai':
            if _ai_gen:
                api_key = _ai_gen.load_api_key()
                model   = _ai_gen.load_model_name()
            if not api_key:
                messagebox.showwarning("AI Mode", "No API key configured.\n"
                                       "Set it via 'API Settings' or switch to Offline mode.")
                return

        self._sp_s2_status_var.set(f"Running Stage 2 [{mode}]...")
        self.root.update_idletasks()

        # Get Solarize addr map from this module
        try:
            solarize_map   = _SOLARIZE_ADDR_TO_NAME
            solarize_scale = _SOLARIZE_STANDARD_SCALE
        except NameError:
            solarize_map   = {}
            solarize_scale = {}

        def _run():
            try:
                out_path = _sp.stage2_create_mapping_excel(
                    stage1_excel_path = stage1_path,
                    solarize_addr_map = solarize_map,
                    solarize_scale    = solarize_scale,
                    mode              = mode,
                    api_key           = api_key,
                    model             = model,
                    progress_cb       = lambda m: self.root.after(0, lambda: self._sp_s2_status_var.set(m[-40:]))
                )
                self._sp_stage2_path = out_path
                self.root.after(0, lambda: self._sp_s2_path_var.set(os.path.basename(out_path)[:50]))
                self.root.after(0, lambda: self._sp_s2_status_var.set(f"Done: {os.path.basename(out_path)}"))
                self.root.after(0, lambda: messagebox.showinfo(
                    "Stage 2 Complete",
                    f"Mapping Excel saved:\n{out_path}\n\n"
                    "Open the Excel and verify/edit 'Solarize_Name' column\n"
                    "for Unmapped rows (highlighted in red).\n"
                    "Then run Step 3 (Generate .py)."))
            except Exception as e:
                self.root.after(0, lambda: self._sp_s2_status_var.set(f"Error: {e}"))
                self.root.after(0, lambda: messagebox.showerror("Stage 2 Error", str(e)))

        threading.Thread(target=_run, daemon=True).start()

    def _sp_open_stage2(self):
        """Open Stage 2 mapping Excel in the default application."""
        import subprocess
        path = self._sp_stage2_path
        if not path:
            # Ask user to select
            from tkinter import filedialog
            path = filedialog.askopenfilename(
                title="Select Stage 2 Mapping Excel",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not path:
                return
            self._sp_stage2_path = path
            self._sp_s2_path_var.set(os.path.basename(path)[:50])

        if not os.path.isfile(path):
            messagebox.showwarning("Open Excel", f"File not found:\n{path}")
            return

        try:
            os.startfile(path)
        except AttributeError:
            subprocess.Popen(['xdg-open', path])

    def _sp_run_stage3(self):
        """Run Stage 3: Mapping Excel → register .py in background thread."""
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        stage2_path = self._sp_stage2_path
        if not stage2_path:
            # Ask user to select mapping Excel
            from tkinter import filedialog
            stage2_path = filedialog.askopenfilename(
                title="Select Stage 2 Mapping Excel",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            if not stage2_path:
                return
            self._sp_stage2_path = stage2_path
            self._sp_s2_path_var.set(os.path.basename(stage2_path)[:50])

        if not os.path.isfile(stage2_path):
            messagebox.showwarning("Stage 3", f"Mapping Excel not found:\n{stage2_path}")
            return

        settings = {
            'mppt_count':   self._sp_mppt_var.get(),
            'string_count': self._sp_string_var.get(),
            'iv_scan':      self._sp_iv_var.get(),
            'der_avm':      self._sp_der_var.get(),
            'dea_avm':      self._sp_dea_var.get(),
            'manufacturer': self.cg_manufacturer_var.get().strip() or 'Unknown',
            'class_name':   self.cg_classname_var.get().strip() or 'RegisterMap',
            'protocol_name': self.cg_protocol_var.get().strip() or 'custom',
            'fc_code':      self.cg_fc_var.get(),
        }

        mode    = self._sp_mode_var.get()
        api_key = ''
        model   = None
        if mode == 'ai':
            if _ai_gen:
                api_key = _ai_gen.load_api_key()
                model   = _ai_gen.load_model_name()
            if not api_key:
                messagebox.showwarning("AI Mode", "No API key configured.\n"
                                       "Set it via 'API Settings' or switch to Offline mode.")
                return

        self._sp_s3_status_var.set(f"Running Stage 3 [{mode}]...")
        self.root.update_idletasks()

        def _run():
            try:
                result = _sp.stage3_generate_py(
                    mapping_excel_path = stage2_path,
                    settings           = settings,
                    output_path        = None,   # No save yet — show in Code Preview first
                    mode               = mode,
                    api_key            = api_key,
                    model              = model,
                    progress_cb        = lambda m: self.root.after(0, lambda: self._sp_s3_status_var.set(m[-40:]))
                )
                code     = result.get('code', '')
                results  = result.get('results', [])
                success  = result.get('success', False)

                self._sp_generated_code = code

                # Show in Code Preview
                def _update_ui():
                    self.cg_text.configure(state='normal')
                    self.cg_text.delete('1.0', 'end')
                    self.cg_text.insert('1.0', code)
                    self.cg_text.configure(state='disabled')

                    # Show test results in guide area
                    lines = [f'=== Stage 3 Results [Mode: {mode.upper()}] ===']
                    for status, msg in results:
                        lines.append(f'[{status}] {msg}')
                    fail_n = sum(1 for s, _ in results if s == 'FAIL')
                    warn_n = sum(1 for s, _ in results if s == 'WARN')
                    pass_n = sum(1 for s, _ in results if s == 'PASS')
                    lines.append('')
                    lines.append(f'Result: {pass_n} PASS  {warn_n} WARN  {fail_n} FAIL')
                    if success:
                        lines.append('>> VALIDATION PASSED — click "Save to common/" to save.')
                    else:
                        lines.append('>> VALIDATION FAILED — edit mapping Excel and re-generate.')

                    self.cg_guide_text.configure(state='normal')
                    self.cg_guide_text.delete('1.0', 'end')
                    self.cg_guide_text.insert('1.0', '\n'.join(lines))
                    self.cg_guide_text.configure(state='disabled')

                    status_msg = f'{"PASS" if success else "FAIL"} — {pass_n}P {warn_n}W {fail_n}F'
                    self._sp_s3_status_var.set(status_msg)

                    if not success:
                        messagebox.showwarning(
                            "Stage 3: Validation Failed",
                            f"{fail_n} test(s) FAILED.\n"
                            "Review results in Config Guide area.\n\n"
                            "Fix the mapping Excel (Step 2) and re-run Stage 3.")

                self.root.after(0, _update_ui)

            except Exception as e:
                self.root.after(0, lambda: self._sp_s3_status_var.set(f"Error: {e}"))
                self.root.after(0, lambda: messagebox.showerror("Stage 3 Error", str(e)))

        threading.Thread(target=_run, daemon=True).start()

    def _sp_test_stage3(self):
        """Test the last generated Stage 3 code using the standard test suite."""
        code = self._sp_generated_code or self.cg_text.get('1.0', 'end').strip()
        if not code:
            messagebox.showwarning("Test", "Generate code first (Step 3 → Generate .py).")
            return

        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        class_name = self.cg_classname_var.get().strip() or 'RegisterMap'
        der_avm    = self._sp_der_var.get()

        results = _sp._validate_register_code(code, class_name, der_avm)
        fail_n  = sum(1 for s, _ in results if s == 'FAIL')
        warn_n  = sum(1 for s, _ in results if s == 'WARN')
        pass_n  = sum(1 for s, _ in results if s == 'PASS')

        lines = ['=== Stage 3 Validation Test Results ===']
        for status, msg in results:
            lines.append(f'[{status}] {msg}')
        lines += ['', f'Result: {pass_n} PASS  {warn_n} WARN  {fail_n} FAIL']

        self.cg_guide_text.configure(state='normal')
        self.cg_guide_text.delete('1.0', 'end')
        self.cg_guide_text.insert('1.0', '\n'.join(lines))
        self.cg_guide_text.configure(state='disabled')

        self._sp_s3_status_var.set(f'Test: {pass_n}P {warn_n}W {fail_n}F')

    def _sp_save_stage3(self):
        """Save Stage 3 generated code to common/{protocol}_registers.py."""
        code = self._sp_generated_code or self.cg_text.get('1.0', 'end').strip()
        if not code:
            messagebox.showwarning("Save", "Generate code first (Step 3 → Generate .py).")
            return

        protocol = self.cg_protocol_var.get().strip()
        if not protocol:
            messagebox.showwarning("Save", "Enter a Protocol Name in the Configuration panel.")
            return

        fname = f'{protocol}_registers.py'
        if fname in _PROTECTED_FILES:
            messagebox.showerror("Protected File",
                f"'{fname}' is protected and cannot be overwritten.\n"
                f"Use a different Protocol Name (e.g. '{protocol}_custom').")
            return

        base_dir   = os.path.dirname(os.path.abspath(__file__))
        target_dir = os.path.normpath(os.path.join(base_dir, '..', 'common'))
        os.makedirs(target_dir, exist_ok=True)
        filepath = os.path.join(target_dir, fname)

        if os.path.isfile(filepath):
            import shutil
            try:
                shutil.copy2(filepath, filepath + '.bak')
            except OSError:
                pass

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(code + '\n')

        # Also update Code Preview
        self.cg_text.configure(state='normal')
        self.cg_text.delete('1.0', 'end')
        self.cg_text.insert('1.0', code)
        self.cg_text.configure(state='disabled')

        # Register in file map
        known = {e[3] for e in _REGISTER_FILE_MAP}
        if fname not in known:
            mfr       = self.cg_manufacturer_var.get().strip()
            model_num = self.cg_model_var.get().strip()
            cn        = self.cg_classname_var.get().strip()
            _REGISTER_FILE_MAP.append(('Inverter', model_num, mfr, fname, protocol, cn))
            _save_register_file_map()

        self._refresh_register_files()
        self._sp_s3_status_var.set(f"Saved: {fname}")
        messagebox.showinfo("Saved", f"Register file saved:\n{filepath}")

    def _sp_save_to_reference(self):
        """Save Stage 3 generated code to the reference library."""
        code = self._sp_generated_code or self.cg_text.get('1.0', 'end').strip()
        if not code:
            messagebox.showwarning("Save to Reference",
                                   "Generate code first (Step 3 → Generate .py).")
            return
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        protocol = self.cg_protocol_var.get().strip() or 'custom'
        mfr      = self.cg_manufacturer_var.get().strip() or 'Unknown'

        meta = {
            'manufacturer': mfr,
            'protocol':     protocol,
            'mppt_count':   self._sp_mppt_var.get(),
            'string_count': self._sp_string_var.get(),
            'iv_scan':      self._sp_iv_var.get(),
            'der_avm':      self._sp_der_var.get(),
            'fc_code':      self.cg_fc_var.get(),
        }

        ok, msg = _sp.save_to_reference(
            protocol           = protocol,
            py_code            = code,
            mapping_excel_path = self._sp_stage2_path,
            meta               = meta,
        )

        if ok:
            # Update ref count display
            try:
                rm = _sp._get_ref_manager()
                if rm:
                    self._sp_ref_status_var.set(f"Reference library: {rm.count()} sets")
            except Exception:
                pass

            # Save to common/*_mm_registers.py and register in file map
            try:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                common_dir = os.path.normpath(os.path.join(base_dir, '..', 'common'))
                mm_fname = f'{protocol}_mm_registers.py'
                mm_path = os.path.join(common_dir, mm_fname)
                with open(mm_path, 'w', encoding='utf-8') as f:
                    f.write(code + '\n')

                # Add R# entry to register_file_map if not already present
                known_fnames = {e[3] for e in _REGISTER_FILE_MAP}
                if mm_fname not in known_fnames:
                    used_r_nums = []
                    for e in _REGISTER_FILE_MAP:
                        if isinstance(e[1], str) and e[1].startswith('R'):
                            try:
                                used_r_nums.append(int(e[1][1:]))
                            except ValueError:
                                pass
                    next_r = max(used_r_nums, default=0) + 1
                    classname = self.cg_classname_var.get().strip() or 'RegisterMap'
                    _REGISTER_FILE_MAP.append(
                        ('Inverter', f'R{next_r}', f'{mfr} (Ref)', mm_fname,
                         f'{protocol}_mm', classname))
                    _save_register_file_map()
            except Exception as e:
                messagebox.showwarning("Reference", f"Reference saved but _mm file failed:\n{e}")

            self._sp_s3_status_var.set(f"Saved to reference: {protocol}")
            self._refresh_register_files()
            messagebox.showinfo("Saved to Reference",
                                f"Reference '{protocol}' saved.\n"
                                f"common/{mm_fname} created.\n"
                                f"Future offline mappings will benefit from this reference.")
        else:
            messagebox.showerror("Save to Reference Failed", msg)

    def _sp_on_mode_change(self):
        """Called when the mode radio button changes."""
        mode = self._sp_mode_var.get()
        if mode == 'ai':
            # Check API key availability
            api_key = ''
            if _ai_gen:
                api_key = _ai_gen.load_api_key()
            if not api_key:
                messagebox.showwarning(
                    "AI Mode",
                    "Claude API key not configured.\n\n"
                    "Set the API key via the 'API Settings' button (existing Code Generator panel).\n"
                    "Switching to AI mode without a key will prompt again when running each step.")

    def _sp_manage_references(self):
        """Show a simple reference library management dialog."""
        if _sp is None:
            messagebox.showerror("Error", "stage_pipeline module not found.")
            return

        rm = _sp._get_ref_manager()
        if rm is None:
            messagebox.showerror("Error", "reference_manager module not found.")
            return

        # Build info text
        refs = rm.list_references()
        lines = [f"Reference Library — {len(refs)} sets\n"]
        for name, meta in refs:
            builtin = '(builtin)' if meta.get('builtin') else '(user)'
            mfr = meta.get('manufacturer', name)
            mppt = meta.get('mppt_count', '?')
            lines.append(f"  {name:<18} {mfr:<28} MPPT={mppt}  {builtin}")

        lines += [
            '',
            'Built-in references (from common/*_mm_registers.py):',
            '  solarize, huawei, kstar, sungrow, ekos, goodwe',
            '',
            'User references are saved in model_maker/reference/{name}/',
            'Use "Save to Reference" button to add new references.',
        ]

        info_win = tk.Toplevel(self.root)
        info_win.title("Reference Library")
        info_win.geometry("620x400")
        info_win.resizable(True, True)

        txt = tk.Text(info_win, font=('Consolas', 9), wrap='none')
        vsb = ttk.Scrollbar(info_win, orient='vertical', command=txt.yview)
        txt.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        txt.pack(fill='both', expand=True, padx=5, pady=5)
        txt.insert('1.0', '\n'.join(lines))
        txt.configure(state='disabled')

        btn_row = ttk.Frame(info_win)
        btn_row.pack(fill='x', padx=5, pady=5)
        ttk.Button(btn_row, text="  Close  ",
                   command=info_win.destroy).pack(side='right')

    # ── Code Generator helpers ───────────────────────────────────────

    @staticmethod
    def _to_upper_snake(name):
        """Convert a field name to UPPER_SNAKE_CASE."""
        # Remove parentheses content and special chars
        s = re.sub(r'[()（）\[\]【】]', '', name)
        # Replace Korean/spaces/hyphens/dots with underscores
        s = re.sub(r'[\s\-./·:]+', '_', s)
        # CamelCase boundaries
        s = re.sub(r'([a-z])([A-Z])', r'\1_\2', s)
        # Remove non-alphanumeric (keep underscore)
        s = re.sub(r'[^A-Za-z0-9_]', '', s)
        # Collapse multiple underscores
        s = re.sub(r'_+', '_', s).strip('_')
        result = s.upper()
        # Python 변수명은 숫자로 시작 불가 → STATUS_ 접두어 추가
        if result and result[0].isdigit():
            result = f'STATUS_{result}'
        return result

    @staticmethod
    def _sanitize_class_name(name):
        """Convert manufacturer name to valid Python CamelCase class name.
        e.g. 'ModBus map-EK' → 'ModbusMapEk', 'Sungrow' → 'Sungrow'
        """
        # Remove special chars, split by spaces/hyphens/underscores
        parts = re.sub(r'[^A-Za-z0-9\s_\-]', '', name).split()
        if not parts:
            return 'Device'
        # Each part → capitalize first letter
        result = ''.join(p.capitalize() for p in parts if p)
        # Remove remaining non-alphanumeric
        result = re.sub(r'[^A-Za-z0-9]', '', result)
        if not result or result[0].isdigit():
            result = 'Device' + result
        return result

    @staticmethod
    def _field_to_const_name(name):
        """Convert a UDP field name (possibly Korean) to UPPER_SNAKE_CASE.

        Lookup order:
        1. Exact match in _KO_EXACT_MAP (e.g. "역률" → "POWER_FACTOR")
        2. Suffix pattern in _KO_SUFFIX_MAP (e.g. "MPPT3 전압" → "MPPT3_VOLTAGE")
        3. Fallback to _to_upper_snake (works well for English names from PDF)
        """
        if name in _KO_EXACT_MAP:
            return _KO_EXACT_MAP[name]
        for ko_suffix, en_suffix in _KO_SUFFIX_MAP:
            if name.endswith(ko_suffix):
                prefix = name[:-len(ko_suffix)].strip()
                base = ModbusToUdpMapper._to_upper_snake(prefix)
                return (base + en_suffix) if base else en_suffix.lstrip('_')
        return ModbusToUdpMapper._to_upper_snake(name)

    def _refresh_register_files(self):
        """Scan common/ for *_registers.py and populate the Treeview.
        Grouped by device type, newest model number first within each group."""
        self.rf_tree.delete(*self.rf_tree.get_children())
        base_dir = os.path.dirname(os.path.abspath(__file__))
        common_dir = os.path.normpath(os.path.join(base_dir, '..', 'common'))

        # Load reference library for Ref column
        ref_dict = {}  # {name: meta}
        try:
            try:
                from model_maker import reference_manager as _rmmod
            except ImportError:
                import reference_manager as _rmmod
            rm = _rmmod.get_manager()
            if rm:
                ref_dict = dict(rm.list_references())
        except Exception:
            pass

        # Collect existing entries
        entries = []  # (dtype, model_num_int, model_str, name, fname, protocol)
        shown = set()

        for dtype, model, name, fname, proto, clsname in _REGISTER_FILE_MAP:
            fpath = os.path.join(common_dir, fname)
            if os.path.isfile(fpath):
                try:
                    model_int = int(model)
                except ValueError:
                    # R1, R2, ... → sort after numeric models
                    model_int = -1
                entries.append((dtype, model_int, model, name, fname, proto))
                shown.add(fname)

        # Scan for unknown register files not in the map — auto-register
        if os.path.isdir(common_dir):
            import glob as _glob
            newly_added = False
            for fpath in sorted(_glob.glob(os.path.join(common_dir, '*_registers.py'))):
                fname = os.path.basename(fpath)
                if fname not in shown and fname != '__init__.py':
                    # Auto-detect type/model from file content
                    dtype, model_name = 'Inverter', fname.replace('_registers.py', '')
                    protocol = fname.replace('_registers.py', '')
                    classname = 'RegisterMap'
                    try:
                        with open(fpath, 'r', encoding='utf-8') as f:
                            head = f.read(3000)
                        if 'Relay' in head or 'relay' in head:
                            dtype = 'Relay'
                        elif 'Weather' in head or 'SEM5046' in head:
                            dtype = 'Weather'
                        # Extract model name from docstring
                        import re as _re
                        m = _re.search(r'"""[^"]*?(\w[\w\s\-]+?)\s+(?:Inverter\s+)?Modbus\s+Register', head)
                        if m:
                            model_name = m.group(1).strip()
                        # Detect class name
                        cm = _re.search(r'^class\s+(\w+)', head, _re.MULTILINE)
                        if cm:
                            classname = cm.group(1)
                    except Exception:
                        pass
                    # Find next available model number
                    used_nums = {int(e[1]) for e in _REGISTER_FILE_MAP
                                 if e[0] == dtype and str(e[1]).isdigit()}
                    next_num = max(used_nums, default=0) + 1
                    model_str = str(next_num)
                    # Register permanently
                    _REGISTER_FILE_MAP.append(
                        (dtype, model_str, model_name, fname, protocol, classname))
                    newly_added = True
                    entries.append((dtype, next_num, model_str, model_name, fname, protocol))
                    shown.add(fname)
            if newly_added:
                _save_register_file_map()

        # Sort: by type order, then model number descending (newest first)
        type_order = {'Inverter': 0, 'Relay': 1, 'Weather': 2, 'Unknown': 9}
        entries.sort(key=lambda e: (type_order.get(e[0], 8), -e[1]))

        for dtype, _, model_str, name, fname, proto in entries:
            # Determine reference status
            ref_label = ''
            # Check exact protocol match first
            if proto in ref_dict:
                ref_label = 'built-in' if ref_dict[proto].get('builtin') else 'user'
            else:
                # Check without _mm suffix (e.g., 'solarize_mm' → 'solarize')
                base_proto = proto.replace('_mm', '') if proto.endswith('_mm') else None
                if base_proto and base_proto in ref_dict:
                    ref_label = 'built-in' if ref_dict[base_proto].get('builtin') else 'user'
            self.rf_tree.insert('', 'end', values=(dtype, model_str, name, fname, ref_label))

    def _on_open_selected(self):
        """Open the selected register file from the Treeview into Code Preview."""
        sel = self.rf_tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Select a register file first.")
            return
        values = self.rf_tree.item(sel[0], 'values')
        if not values or len(values) < 4:
            return
        fname = values[3]
        if fname.endswith(' (missing)'):
            messagebox.showwarning("Warning", "File does not exist.")
            return

        base_dir = os.path.dirname(os.path.abspath(__file__))
        filepath = os.path.normpath(os.path.join(base_dir, '..', 'common', fname))
        if not os.path.isfile(filepath):
            messagebox.showerror("Error", f"File not found:\n{filepath}")
            return

        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                code = f.read()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read file:\n{e}")
            return

        # Show in Code Preview pane (Stage 3 tab)
        self.cg_text.configure(state='normal')
        self.cg_text.delete('1.0', 'end')
        self.cg_text.insert('1.0', code)
        self.cg_text.configure(state='disabled')

        # Auto-fill config fields from known map
        known = {entry[3]: entry for entry in _REGISTER_FILE_MAP}
        if fname in known:
            _, model, _, _, proto, clsname = known[fname]
            self.cg_protocol_var.set(proto)
            self.cg_model_var.set(model)
            self.cg_classname_var.set(clsname)
        else:
            proto = fname.replace('_registers.py', '')
            self.cg_protocol_var.set(proto)

    def _on_delete_selected(self):
        """Delete the selected register file and clean up config files."""
        sel = self.rf_tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Select a register file first.")
            return
        values = self.rf_tree.item(sel[0], 'values')
        if not values or len(values) < 4:
            return
        fname = values[3]
        if fname.endswith(' (missing)'):
            messagebox.showinfo("Info", "File does not exist.")
            return

        base_dir = os.path.dirname(os.path.abspath(__file__))
        filepath = os.path.normpath(os.path.join(base_dir, '..', 'common', fname))
        if not os.path.isfile(filepath):
            messagebox.showerror("Error", f"File not found:\n{filepath}")
            return

        confirm = messagebox.askyesno(
            "Confirm Delete",
            f"Delete register file and remove from config?\n\n"
            f"{fname}\n\nThis cannot be undone.")
        if not confirm:
            return

        # Find model info before deleting
        deleted_entry = None
        for entry in _REGISTER_FILE_MAP:
            if entry[3] == fname:
                deleted_entry = entry
                break

        # 1. Delete the register file
        try:
            os.remove(filepath)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete:\n{e}")
            return

        # 2. Remove from _REGISTER_FILE_MAP
        _REGISTER_FILE_MAP[:] = [e for e in _REGISTER_FILE_MAP if e[3] != fname]
        _save_register_file_map()

        # 3. Clean up config files if model info is known
        if deleted_entry:
            d_type, d_model, d_name, d_fname, d_proto, d_cls = deleted_entry
            self._remove_model_from_configs(d_model, d_proto)

        # 4. Refresh UI
        self._refresh_register_files()
        current_cfg = self.cfg_file_var.get()
        if current_cfg in ('device_models.ini', 'rs485_ch1.ini'):
            self._on_config_selected(None)

        messagebox.showinfo("Deleted",
            f"Deleted: {fname}" +
            (f"\nRemoved model #{deleted_entry[1]} from configs." if deleted_entry else ""))

    def _remove_model_from_configs(self, model_num, protocol):
        """Remove a model from device_models.ini and rs485_ch1.ini."""
        import re as _re

        # ── device_models.ini: remove model_num lines ──
        models_path = os.path.join(self._config_dir, 'device_models.ini')
        if os.path.isfile(models_path):
            with open(models_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            filtered = [ln for ln in lines
                        if not _re.match(rf'^{model_num}\s*=', ln.strip())]
            if len(filtered) != len(lines):
                with open(models_path, 'w', encoding='utf-8') as f:
                    f.writelines(filtered)

        # ── rs485_ch1.ini: remove header comment + [device_N] sections ──
        rs485_path = os.path.join(self._config_dir, 'rs485_ch1.ini')
        if os.path.isfile(rs485_path):
            with open(rs485_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Remove header comment line: "#   N = ModelName (...)"
            comment_pattern = rf'^#\s+{_re.escape(model_num)}\s*=.*$'
            content = _re.sub(comment_pattern, '', content, flags=_re.MULTILINE)
            # Clean up double blank lines left behind
            content = _re.sub(r'\n{3,}', '\n\n', content)

            # Remove [device_N] sections with matching protocol+model
            sections = _re.split(r'(?=^# -{10,})', content, flags=_re.MULTILINE)
            kept = []
            for section in sections:
                has_proto = _re.search(rf'^protocol\s*=\s*{_re.escape(protocol)}\s*$',
                                       section, _re.MULTILINE)
                has_model = _re.search(rf'^model\s*=\s*{_re.escape(model_num)}\s*$',
                                       section, _re.MULTILINE)
                if has_proto and has_model:
                    continue
                kept.append(section)

            new_content = ''.join(kept).rstrip('\n') + '\n'
            if new_content != content:
                with open(rs485_path, 'w', encoding='utf-8') as f:
                    f.write(new_content)

    def _on_add_model(self):
        """Add the current model to device_models.ini and rs485_ch1.ini config files."""
        protocol = self.cg_protocol_var.get().strip()
        model_num = self.cg_model_var.get().strip()
        manufacturer = self.cg_manufacturer_var.get().strip()

        if not protocol or not model_num or not manufacturer:
            messagebox.showwarning("Warning",
                "Fill in Protocol Name, Model Number, and Manufacturer Name first.")
            return

        models_path = os.path.join(self._config_dir, 'device_models.ini')
        rs485_path = os.path.join(self._config_dir, 'rs485_ch1.ini')

        # ── 1. Add to device_models.ini ──
        if os.path.isfile(models_path):
            with open(models_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            def _insert_in_section(lines, section_header, new_line):
                """Insert new_line in sorted order (by model number) within a section."""
                import re as _re_inner
                in_section = False
                section_start = None
                section_end = None
                for i, ln in enumerate(lines):
                    stripped = ln.strip()
                    if stripped == section_header:
                        in_section = True
                        section_start = i
                        continue
                    if in_section:
                        # Already exists?
                        if stripped.startswith(f'{model_num} =') or stripped.startswith(f'{model_num}='):
                            return lines, False
                        # End of section
                        if (stripped.startswith('[') and stripped.endswith(']')) or \
                           stripped.startswith('# ---'):
                            section_end = i
                            break
                if section_start is None:
                    return lines, False
                if section_end is None:
                    section_end = len(lines)
                # Find correct sorted position by model number
                new_num = int(model_num) if model_num.isdigit() else 999
                insert_idx = section_end  # default: end of section
                for i in range(section_start + 1, section_end):
                    stripped = lines[i].strip()
                    if stripped and not stripped.startswith('#'):
                        m = _re_inner.match(r'^(\d+)\s*=', stripped)
                        if m and int(m.group(1)) > new_num:
                            insert_idx = i
                            break
                else:
                    # Insert after last data line
                    for i in range(section_end - 1, section_start, -1):
                        stripped = lines[i].strip()
                        if stripped and not stripped.startswith('#'):
                            insert_idx = i + 1
                            break
                lines.insert(insert_idx, new_line + '\n')
                return lines, True

            changed = False
            # Determine IV scan / KDN support from checkboxes
            iv_support = 'true' if self.cg_ivscan_var.get() else 'false'
            kdn_support = 'true' if self.cg_deravm_var.get() else 'false'

            lines, added = _insert_in_section(
                lines, '[inverter_models]', f'{model_num} = {manufacturer}')
            changed = changed or added
            lines, added = _insert_in_section(
                lines, '[inverter_features]', f'{model_num} = {iv_support}, {kdn_support}')
            changed = changed or added
            lines, added = _insert_in_section(
                lines, '[inverter_protocols]', f'{model_num} = {protocol}')
            changed = changed or added

            if changed:
                with open(models_path, 'w', encoding='utf-8') as f:
                    f.writelines(lines)

        # ── 2. Add to rs485_ch1.ini ──
        if os.path.isfile(rs485_path):
            import re as _re
            with open(rs485_path, 'r', encoding='utf-8') as f:
                rs485_content = f.read()

            # 2a. Add to header comment block
            iv_mark = 'O' if self.cg_ivscan_var.get() else 'X'
            kdn_mark = 'O' if self.cg_deravm_var.get() else 'X'
            comment_line = f'#   {model_num} = {manufacturer} (IV_SCAN: {iv_mark}, KDN: {kdn_mark}, protocol: {protocol})'
            # Insert after the last "# N = ..." line in Inverter Models comment block
            marker = '# Inverter Models:'
            if marker in rs485_content and comment_line not in rs485_content:
                lines_rs = rs485_content.split('\n')
                insert_at = None
                in_block = False
                for i, ln in enumerate(lines_rs):
                    if marker in ln:
                        in_block = True
                        continue
                    if in_block:
                        if _re.match(r'^#\s+\d+\s*=', ln):
                            insert_at = i  # track last model comment line
                        elif not ln.startswith('#') or ln.strip() == '#':
                            break
                if insert_at is not None:
                    lines_rs.insert(insert_at + 1, comment_line)
                    rs485_content = '\n'.join(lines_rs)

            # 2b. Add device section
            existing = _re.findall(r'\[device_(\d+)\]', rs485_content)
            next_num = max(int(n) for n in existing) + 1 if existing else 1

            new_section = (
                f"\n# ----------------------------------------------------------------------------\n"
                f"# Inverter {next_num} - {manufacturer} (Added by Model Maker)\n"
                f"# ----------------------------------------------------------------------------\n"
                f"[device_{next_num}]\n"
                f"slave_id = {next_num}\n"
                f"installed = NO\n"
                f"device_number = {next_num}\n"
                f"device_type = 1\n"
                f"protocol = {protocol}\n"
                f"model = {model_num}\n"
                f"mppt_count = 4\n"
                f"string_count = 8\n"
                f"iv_scan = {'true' if self.cg_ivscan_var.get() else 'false'}\n"
                f"control = {'DER_AVM' if self.cg_deravm_var.get() else 'NONE'}\n"
                f"simulation = false\n"
            )

            rs485_content += new_section
            with open(rs485_path, 'w', encoding='utf-8') as f:
                f.write(rs485_content)

        # Reload config editor if one of these files is open
        current_cfg = self.cfg_file_var.get()
        if current_cfg in ('device_models.ini', 'rs485_ch1.ini'):
            self._on_config_selected(None)

        messagebox.showinfo("Model Added",
            f"Model #{model_num} ({manufacturer}) added to:\n"
            f"- device_models.ini\n"
            f"- rs485_ch1.ini [device_{next_num}]\n\n"
            f"Edit the config to adjust settings.")

    def _on_config_selected(self, event):
        """Load the selected config file into the Config Editor."""
        fname = self.cfg_file_var.get()
        if not fname:
            return
        filepath = os.path.join(self._config_dir, fname)
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read config:\n{e}")
            return
        self.cfg_editor.delete('1.0', 'end')
        self.cfg_editor.insert('1.0', content)

    def _on_save_config(self):
        """Save the Config Editor content back to the selected config file."""
        fname = self.cfg_file_var.get()
        if not fname:
            messagebox.showwarning("Warning", "Select a config file first.")
            return
        filepath = os.path.join(self._config_dir, fname)
        content = self.cfg_editor.get('1.0', 'end').rstrip('\n')
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(content + '\n')
            messagebox.showinfo("Saved", f"Config saved:\n{fname}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save config:\n{e}")

    def _on_build_package(self):
        """Build a full .tar.gz firmware package (rtu_program + common + config)."""
        import tarfile
        from datetime import datetime

        base_dir = os.path.dirname(os.path.abspath(__file__))
        project_dir = os.path.normpath(os.path.join(base_dir, '..'))
        firmware_dir = os.path.join(project_dir, 'pc_programs', 'firmware')

        if not os.path.isdir(firmware_dir):
            os.makedirs(firmware_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        pkg_name = f'rtu_firmware_{timestamp}.tar.gz'
        pkg_path = os.path.join(firmware_dir, pkg_name)

        # Folders to include and their file filters
        # (folder_name, file extensions to include, recurse subdirs)
        pack_dirs = [
            ('rtu_program', ('.py',), True),
            ('common',      ('.py',), False),
            ('config',      ('.ini',), False),
        ]
        skip_dirs = {'__pycache__', '.claude', 'firmware'}

        try:
            file_count = 0
            with tarfile.open(pkg_path, 'w:gz') as tar:
                for folder, extensions, recurse in pack_dirs:
                    src_dir = os.path.join(project_dir, folder)
                    if not os.path.isdir(src_dir):
                        continue

                    if recurse:
                        for root, dirs, files in os.walk(src_dir):
                            # Skip unwanted directories
                            dirs[:] = [d for d in dirs if d not in skip_dirs]
                            for fname in sorted(files):
                                if not any(fname.endswith(ext) for ext in extensions):
                                    continue
                                fpath = os.path.join(root, fname)
                                relpath = os.path.relpath(fpath, project_dir)
                                arcname = relpath.replace('\\', '/')
                                tar.add(fpath, arcname=arcname)
                                file_count += 1
                    else:
                        for fname in sorted(os.listdir(src_dir)):
                            fpath = os.path.join(src_dir, fname)
                            if not os.path.isfile(fpath):
                                continue
                            if not any(fname.endswith(ext) for ext in extensions):
                                continue
                            tar.add(fpath, arcname=f'{folder}/{fname}')
                            file_count += 1

            pkg_size = os.path.getsize(pkg_path)
            result = (
                f"Firmware package built: {pkg_name}\n"
                f"Files: {file_count}, Size: {pkg_size:,} bytes\n"
                f"Includes: rtu_program/ + common/ + config/\n"
                f"Location: pc_programs/firmware/\n"
                f"\nDeploy via Web Dashboard -> Firmware tab"
            )
            self.cg_guide_text.configure(state='normal')
            self.cg_guide_text.delete('1.0', 'end')
            self.cg_guide_text.insert('1.0', result)
            self.cg_guide_text.configure(state='disabled')

            messagebox.showinfo("Package Built",
                f"Firmware package created:\n{pkg_name}\n\n"
                f"{file_count} files, {pkg_size:,} bytes\n"
                f"rtu_program + common + config")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to build package:\n{e}")

    def _on_open_code(self):
        """Open an existing register .py file into the Code Preview for editing."""
        base_dir = os.path.dirname(os.path.abspath(__file__))
        common_dir = os.path.normpath(os.path.join(base_dir, '..', 'common'))
        filepath = filedialog.askopenfilename(
            title="Open Register File",
            initialdir=common_dir if os.path.isdir(common_dir) else base_dir,
            filetypes=[("Python files", "*.py"), ("All files", "*.*")],
        )
        if not filepath:
            return
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                code = f.read()
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read file:\n{e}")
            return

        self.cg_text.configure(state='normal')
        self.cg_text.delete('1.0', 'end')
        self.cg_text.insert('1.0', code)

        # Set protocol name from filename for Save
        fname = os.path.splitext(os.path.basename(filepath))[0]
        if fname.endswith('_registers'):
            fname = fname[:-len('_registers')]
        self.cg_protocol_var.set(fname)

        self.cg_guide_text.configure(state='normal')
        self.cg_guide_text.delete('1.0', 'end')
        self.cg_guide_text.insert('1.0', f"Opened: {filepath}")
        self.cg_guide_text.configure(state='disabled')

    def _check_model_duplicate(self):
        """Check if the current model number conflicts with an existing register file.
        Only checks files that actually exist on disk. Returns warning or empty string."""
        model_num = self.cg_model_var.get().strip()
        protocol = self.cg_protocol_var.get().strip()
        fname = f'{protocol}_registers.py'
        base_dir = os.path.dirname(os.path.abspath(__file__))
        common_dir = os.path.normpath(os.path.join(base_dir, '..', 'common'))
        for dtype, m, name, fn, proto, cls in _REGISTER_FILE_MAP:
            if m == model_num and fn != fname:
                # Only conflict if the file actually exists
                if os.path.isfile(os.path.join(common_dir, fn)):
                    return (f"Model #{model_num} is already used by "
                            f"\"{name}\" ({fn}).\n"
                            f"Please use a different model number.")
        return ''

    def _on_generate_code(self):
        """Generate Python register code and display in preview."""
        if not self.mapping:
            messagebox.showwarning("Warning", "Please run auto-mapping first (Tab 2).")
            return

        # Check model number duplicate
        dup_msg = self._check_model_duplicate()
        if dup_msg:
            messagebox.showwarning("Model # Conflict", dup_msg)
            return

        code = self._generate_register_py()

        self.cg_text.configure(state='normal')
        self.cg_text.delete('1.0', 'end')
        self.cg_text.insert('1.0', code)

        # Update config guide
        protocol = self.cg_protocol_var.get().strip()
        model = self.cg_model_var.get().strip()
        guide = (f"# Add to rs485_ch1.ini:\n"
                 f"[device_N]\n"
                 f"protocol = {protocol}\n"
                 f"model = {model}\n")
        self.cg_guide_text.configure(state='normal')
        self.cg_guide_text.delete('1.0', 'end')
        self.cg_guide_text.insert('1.0', guide)
        self.cg_guide_text.configure(state='disabled')

    def _on_save_code(self):
        """Save generated code to ../../common/{protocol}_registers.py."""
        code = self.cg_text.get('1.0', 'end').strip()
        if not code:
            messagebox.showwarning("Warning", "Generate code first.")
            return
        protocol = self.cg_protocol_var.get().strip()
        if not protocol:
            messagebox.showwarning("Warning", "Enter a protocol name.")
            return

        # Check protected files
        fname = f'{protocol}_registers.py'
        if fname in _PROTECTED_FILES:
            messagebox.showerror("Protected File",
                f"'{fname}' is a production register file and cannot be overwritten.\n\n"
                f"Use a different Protocol Name (e.g. '{protocol}_custom').")
            return

        # Check model number duplicate
        dup_msg = self._check_model_duplicate()
        if dup_msg:
            messagebox.showwarning("Model # Conflict", dup_msg)
            return

        # Resolve path relative to this script's directory
        base_dir = os.path.dirname(os.path.abspath(__file__))
        target_dir = os.path.normpath(os.path.join(base_dir, '..', 'common'))
        if not os.path.isdir(target_dir):
            os.makedirs(target_dir, exist_ok=True)
        filepath = os.path.join(target_dir, f'{protocol}_registers.py')

        # Backup existing file before overwrite
        if os.path.isfile(filepath):
            import shutil
            bak_path = filepath + '.bak'
            try:
                shutil.copy2(filepath, bak_path)
            except OSError:
                pass

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(code + '\n')

        # Register the new file in the map if not already known
        fname = f'{protocol}_registers.py'
        known_fnames = {entry[3] for entry in _REGISTER_FILE_MAP}
        if fname not in known_fnames:
            manufacturer = self.cg_manufacturer_var.get().strip()
            model_num = self.cg_model_var.get().strip()
            classname = self.cg_classname_var.get().strip()
            # Infer device type from parsed data or default to Inverter
            dtype = 'Inverter'
            _REGISTER_FILE_MAP.append(
                (dtype, model_num, manufacturer, fname, protocol, classname))
            _save_register_file_map()

        self._refresh_register_files()
        messagebox.showinfo("Saved", f"Code saved to:\n{filepath}")

    def _on_copy_code(self):
        """Copy generated code to clipboard."""
        code = self.cg_text.get('1.0', 'end').strip()
        if not code:
            messagebox.showwarning("Warning", "Generate code first.")
            return
        self.root.clipboard_clear()
        self.root.clipboard_append(code)
        messagebox.showinfo("Copied", "Code copied to clipboard.")

    def _on_test_code(self):
        """Run validation tests on the generated code and display results."""
        code = self.cg_text.get('1.0', 'end').strip()
        if not code:
            messagebox.showwarning("Warning", "Generate code first.")
            return

        classname = self.cg_classname_var.get().strip()
        results = []  # (status, message)  status: 'PASS', 'WARN', 'FAIL'
        # Provide __file__ for exec() so generated code with sys.path.insert works
        base_dir = os.path.dirname(os.path.abspath(__file__))
        fake_file = os.path.join(base_dir, '..', 'common', 'test_registers.py')
        ns = {'__file__': fake_file}
        reg_cls = None

        # ── Test 1: Syntax ──
        try:
            compile(code, '<test>', 'exec')
            results.append(('PASS', '1. Syntax check'))
        except SyntaxError as e:
            results.append(('FAIL', f'1. Syntax error: {e}'))
            self._show_test_results(results)
            return

        # ── Test 2: Import & class ──
        try:
            exec(code, ns)
            reg_cls = ns.get(classname)
            if reg_cls is None:
                results.append(('FAIL', f'2. Class "{classname}" not found'))
            else:
                results.append(('PASS', f'2. Import & class found: {classname}'))
        except Exception as e:
            results.append(('FAIL', f'2. Exec error: {e}'))
            self._show_test_results(results)
            return

        # ── Test 3: Register constants ──
        if reg_cls:
            regs = {k: v for k, v in vars(reg_cls).items()
                    if not k.startswith('_') and isinstance(v, int)}
            if regs:
                neg = [k for k, v in regs.items() if v < 0]
                if neg:
                    results.append(('FAIL', f'3. Negative addresses: {", ".join(neg[:5])}'))
                else:
                    results.append(('PASS', f'3. Register constants: {len(regs)} registers'))
            else:
                results.append(('FAIL', '3. No register constants found'))
        else:
            results.append(('FAIL', '3. Skipped (no class)'))

        # ── Test 4: Address uniqueness ──
        if reg_cls:
            addr_map = {}
            for k, v in regs.items():
                addr_map.setdefault(v, []).append(k)
            dups = {hex(addr): names for addr, names in addr_map.items()
                    if len(names) > 1}
            if dups:
                dup_strs = [f'{addr}: {", ".join(names)}' for addr, names in
                            list(dups.items())[:3]]
                results.append(('WARN',
                    f'4. Duplicate addresses ({len(dups)}): '
                    + '; '.join(dup_strs)))
            else:
                results.append(('PASS', '4. Address uniqueness: all unique'))

        # ── Test 5: SCALE dict ──
        scale = ns.get('SCALE')
        if scale is None:
            results.append(('FAIL', '5. SCALE dict not found'))
        elif not isinstance(scale, dict):
            results.append(('FAIL', '5. SCALE is not a dict'))
        else:
            # 필수 일반 키 체크 (modbus_handler 호환)
            required_keys = ['voltage', 'current', 'power', 'frequency', 'power_factor']
            missing_keys = [k for k in required_keys if k not in scale]
            bad_vals = [k for k, v in scale.items()
                        if not isinstance(v, (int, float))]
            if missing_keys:
                results.append(('WARN',
                    f'5. SCALE missing standard keys: {", ".join(missing_keys)}'))
            elif bad_vals:
                results.append(('FAIL',
                    f'5. SCALE non-numeric values: {", ".join(bad_vals[:5])}'))
            else:
                results.append(('PASS',
                    f'5. SCALE dict: {len(scale)} entries, standard keys OK'))

        # ── Test 6: InverterMode class ──
        inv_mode = ns.get('InverterMode')
        if inv_mode is None:
            results.append(('FAIL', '6. InverterMode class not found'))
        else:
            required_attrs = ['INITIAL', 'STANDBY', 'ON_GRID', 'FAULT', 'SHUTDOWN']
            missing = [a for a in required_attrs if not hasattr(inv_mode, a)]
            if missing:
                results.append(('FAIL',
                    f'6. InverterMode missing: {", ".join(missing)}'))
            elif not hasattr(inv_mode, 'to_string'):
                results.append(('WARN', '6. InverterMode: to_string() not found'))
            else:
                results.append(('PASS',
                    f'6. InverterMode: all status codes + to_string() OK'))

        # ── Test 7: Helper functions ──
        fn_u32 = ns.get('registers_to_u32')
        fn_s32 = ns.get('registers_to_s32')
        if fn_u32 is None or fn_s32 is None:
            missing = []
            if fn_u32 is None:
                missing.append('registers_to_u32')
            if fn_s32 is None:
                missing.append('registers_to_s32')
            results.append(('FAIL', f'7. Missing functions: {", ".join(missing)}'))
        else:
            try:
                # Detect byte order: Huawei uses (hi, lo), others use (low, high)
                # Test both conventions and accept either
                u32_le = fn_u32(0x1234, 0x5678) == 0x56781234  # little-endian (low, high)
                u32_be = fn_u32(0x1234, 0x5678) == 0x12345678  # big-endian (hi, lo)
                u32_ok = u32_le or u32_be
                byte_order = 'big-endian' if u32_be else 'little-endian'
                s32_ok = (fn_s32(0, 0x8000) < 0) or (fn_s32(0x8000, 0) < 0)
                if u32_ok and s32_ok:
                    results.append(('PASS',
                        f'7. Helper functions: registers_to_u32, registers_to_s32 OK ({byte_order})'))
                else:
                    fails = []
                    if not u32_ok:
                        fails.append('u32')
                    if not s32_ok:
                        fails.append('s32')
                    results.append(('FAIL',
                        f'7. Wrong result: {", ".join(fails)}'))
            except Exception as e:
                results.append(('FAIL', f'7. Function error: {e}'))

        self._show_test_results(results)

    def _show_test_results(self, results):
        """Display test results in the Config Guide text area."""
        counts = {'PASS': 0, 'WARN': 0, 'FAIL': 0}
        lines = ['=== Code Test Results ===']
        for status, msg in results:
            counts[status] = counts.get(status, 0) + 1
            lines.append(f'[{status}] {msg}')
        lines.append('')
        lines.append(
            f'Result: {counts["PASS"]} PASS, '
            f'{counts["WARN"]} WARN, {counts["FAIL"]} FAIL')

        self.cg_guide_text.configure(state='normal')
        self.cg_guide_text.delete('1.0', 'end')
        self.cg_guide_text.insert('1.0', '\n'.join(lines))
        self.cg_guide_text.configure(state='disabled')

    # ── AI Generation ─────────────────────────────────────────────────────

    def _on_api_settings(self):
        """Open a dialog to configure Anthropic API key and model."""
        if _ai_gen is None:
            messagebox.showerror(
                "Module Missing",
                "ai_generator module not found.\n"
                "Ensure ai_generator.py is in the model_maker/ directory.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Claude API Settings")
        dlg.geometry("520x220")
        dlg.resizable(False, False)
        dlg.grab_set()

        pad = {'padx': 10, 'pady': 6}

        ttk.Label(dlg, text="Anthropic API Key:").grid(
            row=0, column=0, sticky='e', **pad)
        key_var = tk.StringVar(value=_ai_gen.load_api_key())
        key_entry = ttk.Entry(dlg, textvariable=key_var, width=46, show='*')
        key_entry.grid(row=0, column=1, columnspan=2, sticky='w', **pad)

        show_var = tk.BooleanVar(value=False)
        def _toggle_show():
            key_entry.configure(show='' if show_var.get() else '*')
        ttk.Checkbutton(dlg, text="Show", variable=show_var,
                        command=_toggle_show).grid(row=0, column=3, **pad)

        ttk.Label(dlg, text="Claude Model:").grid(
            row=1, column=0, sticky='e', **pad)
        model_var = tk.StringVar(value=_ai_gen.load_model_name())
        model_combo = ttk.Combobox(
            dlg, textvariable=model_var, width=32,
            values=['claude-opus-4-6', 'claude-sonnet-4-6',
                    'claude-haiku-4-5-20251001'],
            state='normal')
        model_combo.grid(row=1, column=1, columnspan=2, sticky='w', **pad)

        ttk.Label(dlg,
                  text="Get your key at: console.anthropic.com/keys",
                  foreground='#666').grid(
            row=2, column=0, columnspan=4, **pad)

        def _save():
            _ai_gen.save_api_key(key_var.get().strip())
            _ai_gen.save_model_name(model_var.get().strip())
            messagebox.showinfo("Saved", "API settings saved to config/ai_settings.ini",
                                parent=dlg)
            dlg.destroy()

        btn_frame = ttk.Frame(dlg)
        btn_frame.grid(row=3, column=0, columnspan=4, pady=10)
        ttk.Button(btn_frame, text="  Save  ", command=_save).pack(side='left', padx=8)
        ttk.Button(btn_frame, text="  Cancel  ",
                   command=dlg.destroy).pack(side='left', padx=8)

        dlg.transient(self.root)
        dlg.wait_window()

    def _on_ai_generate(self):
        """Generate register file from PDF via Claude API with 12-item validation."""
        if _ai_gen is None:
            messagebox.showerror(
                "Module Missing",
                "ai_generator module not found.\n"
                "Ensure ai_generator.py is in the model_maker/ directory.")
            return

        # Check API key
        api_key = _ai_gen.load_api_key()
        if not api_key:
            if messagebox.askyesno(
                    "API Key Required",
                    "Anthropic API key is not configured.\n\n"
                    "Open API Settings now?"):
                self._on_api_settings()
            return

        # Select PDF file
        base_dir = os.path.dirname(os.path.abspath(__file__))
        docs_dir = os.path.normpath(os.path.join(base_dir, '..', 'docs'))
        pdf_path = filedialog.askopenfilename(
            title="Select Inverter Modbus Protocol PDF",
            initialdir=docs_dir if os.path.isdir(docs_dir) else base_dir,
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if not pdf_path:
            return

        # Collect generation parameters from Tab 6 config fields
        manufacturer = self.cg_manufacturer_var.get().strip() or 'Unknown'
        protocol_name = self.cg_protocol_var.get().strip() or 'newbrand'
        class_name = self.cg_classname_var.get().strip() or 'RegisterMap'
        include_iv = self.cg_ivscan_var.get()
        include_deravm = self.cg_deravm_var.get()
        try:
            mppt_count = int(self.mppt_var.get())
        except ValueError:
            mppt_count = 4
        try:
            string_count = int(self.string_var.get())
        except ValueError:
            string_count = 8

        model = _ai_gen.load_model_name()

        # Confirm
        if not messagebox.askyesno(
                "AI Generate",
                f"Generate register file using Claude API?\n\n"
                f"PDF:      {os.path.basename(pdf_path)}\n"
                f"Maker:    {manufacturer}\n"
                f"Protocol: {protocol_name}\n"
                f"Class:    {class_name}\n"
                f"MPPT:     {mppt_count}   Strings: {string_count}\n"
                f"IV Scan:  {include_iv}   DER-AVM: {include_deravm}\n"
                f"Model:    {model}\n\n"
                f"Up to 3 retry attempts on validation failure.\n"
                f"This may take 30–90 seconds."):
            return

        # ── Progress Dialog ──
        prog_dlg = tk.Toplevel(self.root)
        prog_dlg.title("AI Generate — Progress")
        prog_dlg.geometry("560x340")
        prog_dlg.resizable(True, True)
        prog_dlg.grab_set()

        ttk.Label(prog_dlg,
                  text=f"Generating: {protocol_name}_registers.py",
                  font=('맑은 고딕', 10, 'bold')).pack(padx=10, pady=(10, 4))

        log_text = tk.Text(prog_dlg, wrap='word', font=('Consolas', 9),
                           height=14, state='disabled', background='#1E1E1E',
                           foreground='#D4D4D4', insertbackground='white')
        log_vsb = ttk.Scrollbar(prog_dlg, orient='vertical',
                                command=log_text.yview)
        log_text.configure(yscrollcommand=log_vsb.set)
        log_vsb.pack(side='right', fill='y', padx=(0, 5), pady=5)
        log_text.pack(fill='both', expand=True, padx=(10, 0), pady=5)

        status_var = tk.StringVar(value="Starting...")
        ttk.Label(prog_dlg, textvariable=status_var,
                  font=('맑은 고딕', 9)).pack(pady=(2, 0))

        close_btn = ttk.Button(prog_dlg, text="  Close  ",
                               command=prog_dlg.destroy, state='disabled')
        close_btn.pack(pady=8)

        _result_holder = [None]

        def _append_log(step, detail=''):
            """Thread-safe log append."""
            def _do():
                log_text.configure(state='normal')
                line = f">> {step}"
                if detail:
                    line += f"  [{detail}]"
                log_text.insert('end', line + '\n')
                log_text.see('end')
                log_text.configure(state='disabled')
                status_var.set(step)
            prog_dlg.after(0, _do)

        def _on_done(result_or_error):
            def _do():
                close_btn.configure(state='normal')
                if isinstance(result_or_error, Exception):
                    _append_log('ERROR', str(result_or_error))
                    messagebox.showerror(
                        "AI Generate Failed",
                        str(result_or_error),
                        parent=prog_dlg)
                    return

                result = result_or_error
                code = result.get('code', '')
                results = result.get('results', [])
                attempt = result.get('attempt', 1)
                success = result.get('success', False)

                # Display code in Code Preview
                self.cg_text.configure(state='normal')
                self.cg_text.delete('1.0', 'end')
                self.cg_text.insert('1.0', code)

                # Show test results in Config Guide
                counts = {'PASS': 0, 'WARN': 0, 'FAIL': 0}
                lines_out = [
                    f'=== AI Generate Results (attempt {attempt}/3) ===',
                    f'File: {protocol_name}_registers.py',
                    f'Model: {model}',
                    '']
                for status, msg in results:
                    counts[status] = counts.get(status, 0) + 1
                    lines_out.append(f'[{status}] {msg}')
                lines_out.append('')
                lines_out.append(
                    f'Summary: {counts["PASS"]} PASS  '
                    f'{counts["WARN"]} WARN  {counts["FAIL"]} FAIL')
                if success:
                    lines_out.append('')
                    lines_out.append(
                        '>> Code placed in Code Preview. '
                        'Click "Save to common/" to save.')

                self.cg_guide_text.configure(state='normal')
                self.cg_guide_text.delete('1.0', 'end')
                self.cg_guide_text.insert('1.0', '\n'.join(lines_out))
                self.cg_guide_text.configure(state='disabled')

                status_var.set(
                    f'Done — {counts["PASS"]} PASS  '
                    f'{counts["WARN"]} WARN  {counts["FAIL"]} FAIL')

                if success:
                    _append_log('Generation complete — no failures!')
                else:
                    _append_log(
                        f'Done with {counts["FAIL"]} FAIL item(s)',
                        'Review test results in Config Guide')

                # Switch to Code Generator tab
                try:
                    self.notebook.select(5)  # Tab 6 is index 5
                except Exception:
                    pass

                if not success:
                    messagebox.showwarning(
                        "Validation Issues",
                        f"Generated code has {counts['FAIL']} FAIL item(s).\n\n"
                        f"Review the test results in Config Guide, then\n"
                        f"manually fix the code in Code Preview before saving.",
                        parent=prog_dlg)

            prog_dlg.after(0, _do)

        def _worker():
            try:
                result = _ai_gen.generate_register_file(
                    pdf_path=pdf_path,
                    manufacturer=manufacturer,
                    mppt_count=mppt_count,
                    string_count=string_count,
                    include_iv=include_iv,
                    include_deravm=include_deravm,
                    protocol_name=protocol_name,
                    class_name=class_name,
                    api_key=api_key,
                    model=model,
                    max_retries=3,
                    progress_callback=_append_log,
                )
                _result_holder[0] = result
                _on_done(result)
            except Exception as exc:
                _on_done(exc)

        t = threading.Thread(target=_worker, daemon=True)
        t.start()

        prog_dlg.transient(self.root)
        prog_dlg.wait_window()

    def _generate_register_py(self):
        """Dispatcher: route to manufacturer-specific code generator."""
        manufacturer = self.parsed_data.get('manufacturer', '') if self.parsed_data else ''
        if 'Kstar' in manufacturer:
            return self._generate_kstar_py()
        elif 'Huawei' in manufacturer:
            return self._generate_huawei_py()
        elif 'Solarize' in manufacturer or 'VerterKing' in manufacturer:
            return self._generate_solarize_py()
        else:
            return self._generate_generic_py()

    def _generate_solarize_py(self):
        """Build a Python source string matching solarize_registers.py format.

        Uses _SOLARIZE_ADDR_TO_NAME for standard naming, generic SCALE keys,
        backward-compatibility aliases, and complete helper functions.
        """
        from datetime import datetime

        manufacturer = self.cg_manufacturer_var.get().strip()
        classname = self.cg_classname_var.get().strip()
        include_iv = self.cg_ivscan_var.get()
        include_deravm = self.cg_deravm_var.get()

        lines = []
        emitted_addrs = set()   # 이미 출력된 주소 (중복 방지)
        seen_names = set()      # 이미 사용된 이름

        def _addr_int(hex_str):
            """'0x1001' → 4097"""
            try:
                return int(hex_str, 16) if isinstance(hex_str, str) else int(hex_str)
            except (ValueError, TypeError):
                return None

        def _std_name(addr_int):
            """주소에 해당하는 solarize 표준 이름 반환"""
            entry = _SOLARIZE_ADDR_TO_NAME.get(addr_int)
            return entry[0] if entry else None

        def _std_comment(addr_int):
            """주소에 해당하는 solarize 표준 코멘트 반환"""
            entry = _SOLARIZE_ADDR_TO_NAME.get(addr_int)
            return entry[1] if entry else None

        def _emit_reg(addr_hex, name, comment):
            """RegisterMap에 레지스터 한 줄 추가"""
            addr = _addr_int(addr_hex)
            if addr is not None and addr in emitted_addrs:
                return
            if addr is not None:
                emitted_addrs.add(addr)
            seen_names.add(name)
            if comment:
                lines.append(f"    {name} = {addr_hex}      # {comment}")
            else:
                lines.append(f"    {name} = {addr_hex}")

        def _emit_std_reg(addr_int):
            """표준 테이블에서 레지스터 출력"""
            if addr_int in emitted_addrs:
                return
            entry = _SOLARIZE_ADDR_TO_NAME.get(addr_int)
            if entry:
                name, comment = entry
                hex_str = f"0x{addr_int:04X}"
                _emit_reg(hex_str, name, comment)

        def _resolve_name_and_comment(m):
            """매핑 항목에서 표준 이름/코멘트 결정"""
            addr_hex = m.get('mb_addr_hex', '-')
            addr = _addr_int(addr_hex)
            mb_type = m.get('mb_type', '')
            mb_unit = m.get('mb_unit', '')
            mb_regs = m.get('mb_regs', 0)

            # 표준 이름이 있으면 사용
            std = _std_name(addr) if addr else None
            if std:
                return addr_hex, std, _std_comment(addr)

            # 표준에 없으면 기존 방식으로 이름 생성
            udp_field = m.get('udp_field', '')
            definition = m.get('mb_definition', '')
            raw_name = udp_field if udp_field else definition
            name = self._field_to_const_name(raw_name)
            if not name:
                return addr_hex, None, None
            # 중복 방지
            base = name
            idx = 2
            while name in seen_names:
                name = f"{base}_{idx}"
                idx += 1

            # 코멘트 생성
            comment_parts = []
            if mb_type:
                comment_parts.append(mb_type)
            if mb_unit and mb_unit not in comment_parts:
                comment_parts.append(mb_unit)
            if mb_regs and int(mb_regs) > 1:
                comment_parts.append(f"{mb_regs} regs")
            return addr_hex, name, ', '.join(comment_parts)

        def _emit_mapping_section(section_key, section_name):
            """매핑 섹션의 레지스터들을 표준 이름으로 출력"""
            lines.append('    # =========================================================================')
            lines.append(f'    # {section_name}')
            lines.append('    # =========================================================================')
            has_any = False
            for m in self.mapping.get(section_key, []):
                if m.get('mb_addr_hex', '-') == '-':
                    continue
                addr_hex, name, comment = _resolve_name_and_comment(m)
                if not name:
                    continue
                addr = _addr_int(addr_hex)
                if addr is not None and addr in emitted_addrs:
                    # 주소가 이미 출력됨 → alias로 추가 (MPPT/STRING 중복 허용)
                    if name not in seen_names:
                        # 기존 출력된 이름 찾기
                        existing_name = None
                        for line in lines:
                            if f'= {addr_hex}' in line or f'= 0x{addr:04X}' in line:
                                parts = line.strip().split('=')
                                if len(parts) >= 2:
                                    existing_name = parts[0].strip()
                                    break
                        if existing_name:
                            lines.append(f"    {name} = {existing_name}      # Alias")
                        else:
                            lines.append(f"    {name} = {addr_hex}      # {comment}")
                        seen_names.add(name)
                        has_any = True
                else:
                    _emit_reg(addr_hex, name, comment)
                    has_any = True
            if not has_any:
                lines.append(f'    # (no {section_name.lower()} registers mapped)')
            lines.append('')

        def _fill_missing_standard(addr_range_start, addr_range_end):
            """표준 테이블에서 미출력된 레지스터 보충"""
            for addr, (name, comment) in sorted(_SOLARIZE_ADDR_TO_NAME.items()):
                if addr_range_start <= addr <= addr_range_end:
                    _emit_std_reg(addr)

        # ══════════════════════════════════════════════════════════════════
        # Header
        # ══════════════════════════════════════════════════════════════════
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        lines.append('"""')
        lines.append(f'{manufacturer} Inverter Modbus Register Map')
        lines.append(f'Based on Solarize Modbus Protocol')
        lines.append(f'Auto-generated by UDP RTU Model Maker')
        lines.append(f'Date: {now_str}')
        lines.append('"""')
        lines.append('')
        lines.append('')
        lines.append(f'class {classname}:')
        lines.append(f'    """{manufacturer} Modbus Register Map - Solarize Protocol Compliant"""')
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # Device Information (0x1A00-0x1A90)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # Device Information (0x1A00-0x1A90)')
        lines.append('    # =========================================================================')
        # 파싱된 device_info에서 주소 수집하되 표준 이름 사용
        if self.parsed_data and self.parsed_data.get('device_info'):
            for reg in self.parsed_data['device_info']:
                addr_hex = reg.get('address_hex', '')
                addr = _addr_int(addr_hex)
                if addr is None:
                    continue
                std = _std_name(addr)
                if std:
                    _emit_reg(addr_hex, std, _std_comment(addr))
                else:
                    definition = reg.get('definition', '')
                    name = self._field_to_const_name(definition)
                    if name:
                        mb_type = reg.get('type', '')
                        mb_unit = reg.get('unit', '')
                        comment = ', '.join(filter(None, [mb_type, mb_unit]))
                        _emit_reg(addr_hex, name, comment)
        # 표준 테이블에서 누락된 Device Info 보충
        _fill_missing_standard(0x1A00, 0x1A90)
        lines.append('    ')
        lines.append('    # Alias for backward compatibility')
        lines.append('    FIRMWARE_VERSION = MASTER_FIRMWARE_VERSION')
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # Inverter Real-time Data - AC Side (0x1001-0x100F)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # Inverter Real-time Data - AC Side (0x1001-0x100F)')
        lines.append('    # =========================================================================')
        # 파싱된 periodic_basic에서 AC 관련 주소 수집
        for m in self.mapping.get('periodic_basic', []):
            addr_hex = m.get('mb_addr_hex', '-')
            if addr_hex == '-':
                continue
            addr = _addr_int(addr_hex)
            std = _std_name(addr) if addr else None
            if std:
                _emit_reg(addr_hex, std, _std_comment(addr))
        # 표준 AC 레지스터 보충 (0x1001-0x100F)
        _fill_missing_standard(0x1001, 0x100F)
        lines.append('    ')
        # R/S/T Phase Aliases
        lines.append('    # R/S/T Phase Aliases')
        for phase, l_name in [('R', 'L1'), ('S', 'L2'), ('T', 'L3')]:
            for suffix in ['_VOLTAGE', '_CURRENT', '_POWER_LOW', '_POWER_HIGH', '_FREQUENCY']:
                lines.append(f'    {phase}{suffix} = {l_name}{suffix}')
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # MPPT Data (0x1010-0x1093)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # MPPT Data (0x1010-0x1093)')
        lines.append('    # =========================================================================')
        for m in self.mapping.get('periodic_mppt', []):
            addr_hex = m.get('mb_addr_hex', '-')
            if addr_hex == '-':
                continue
            addr = _addr_int(addr_hex)
            std = _std_name(addr) if addr else None
            if std:
                _emit_reg(addr_hex, std, _std_comment(addr))
        # MPPT 표준 레지스터 보충 (MPPT1-4: 0x1010-0x1041, MPPT5-9: 0x1080-0x1093)
        _fill_missing_standard(0x1010, 0x101B)
        _fill_missing_standard(0x103E, 0x1041)
        _fill_missing_standard(0x1080, 0x1093)
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # Status & Error (0x101C-0x1020)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # Status & Error (0x101C-0x1020)')
        lines.append('    # =========================================================================')
        _fill_missing_standard(0x101C, 0x1020)
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # Energy Data (0x1021-0x1028)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # Energy Data (0x1021-0x1028)')
        lines.append('    # =========================================================================')
        _fill_missing_standard(0x1021, 0x1028)
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # Grid Power Data (0x1034-0x103D)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # Grid Power Data (0x1034-0x103D)')
        lines.append('    # =========================================================================')
        _fill_missing_standard(0x1034, 0x103D)
        lines.append('    ')
        lines.append('    # Aliases')
        lines.append('    GRID_POWER_LOW = GRID_TOTAL_ACTIVE_POWER_LOW')
        lines.append('    GRID_POWER_HIGH = GRID_TOTAL_ACTIVE_POWER_HIGH')
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # PV Total Power (0x1048-0x104C)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # PV Total Power (0x1048-0x104C)')
        lines.append('    # =========================================================================')
        _fill_missing_standard(0x1048, 0x104C)
        lines.append('    ')
        lines.append('    # Aliases')
        lines.append('    PV_POWER_LOW = PV_TOTAL_INPUT_POWER_LOW')
        lines.append('    PV_POWER_HIGH = PV_TOTAL_INPUT_POWER_HIGH')
        lines.append('    ')
        lines.append('    # --- Standard handler compatibility aliases (T03 required) ---')
        lines.append('    R_PHASE_VOLTAGE = L1_VOLTAGE')
        lines.append('    S_PHASE_VOLTAGE = L2_VOLTAGE')
        lines.append('    T_PHASE_VOLTAGE = L3_VOLTAGE')
        lines.append('    R_PHASE_CURRENT = L1_CURRENT')
        lines.append('    S_PHASE_CURRENT = L2_CURRENT')
        lines.append('    T_PHASE_CURRENT = L3_CURRENT')
        lines.append('    FREQUENCY = L1_FREQUENCY')
        lines.append('    AC_POWER = GRID_TOTAL_ACTIVE_POWER_LOW')
        lines.append('    PV_POWER = PV_TOTAL_INPUT_POWER_LOW')
        lines.append('    TOTAL_ENERGY = TOTAL_ENERGY_LOW')
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # String Input Data (0x1050-0x107F)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # String Input Data (0x1050-0x107F) - Voltage/Current pairs')
        lines.append('    # =========================================================================')
        for m in self.mapping.get('periodic_string', []):
            addr_hex = m.get('mb_addr_hex', '-')
            if addr_hex == '-':
                continue
            addr = _addr_int(addr_hex)
            std = _std_name(addr) if addr else None
            if std:
                _emit_reg(addr_hex, std, _std_comment(addr))
        # 표준 String 레지스터 보충 (voltage+current 쌍)
        _fill_missing_standard(0x1050, 0x107F)
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # DEA-AVM Real-time Data (0x03E8-0x03FD)
        # ══════════════════════════════════════════════════════════════════
        if include_deravm:
            lines.append('    # =========================================================================')
            lines.append('    # DEA-AVM Real-time Data (0x03E8-0x03FD) - For H05 Body Type 14')
            lines.append('    # =========================================================================')
            for m in self.mapping.get('control_monitor', []):
                addr_hex = m.get('mb_addr_hex', '-')
                if addr_hex == '-':
                    continue
                addr = _addr_int(addr_hex)
                std = _std_name(addr) if addr else None
                if std:
                    _emit_reg(addr_hex, std, _std_comment(addr))
            _fill_missing_standard(0x03E8, 0x03FD)
            lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # DER-AVM Control Parameters (0x07D0-0x0835)
        # ══════════════════════════════════════════════════════════════════
        if include_deravm:
            lines.append('    # =========================================================================')
            lines.append('    # DER-AVM Control Parameters (0x07D0-0x0835) - For H05 Body Type 13')
            lines.append('    # =========================================================================')
            for m in self.mapping.get('control_values', []):
                addr_hex = m.get('mb_addr_hex', '-')
                if addr_hex == '-':
                    continue
                addr = _addr_int(addr_hex)
                std = _std_name(addr) if addr else None
                if std:
                    _emit_reg(addr_hex, std, _std_comment(addr))
            _fill_missing_standard(0x07D0, 0x07D3)
            _fill_missing_standard(0x0834, 0x0835)
            lines.append('    ')
            # Aliases for backward compatibility
            lines.append('    # Aliases for backward compatibility')
            lines.append('    POWER_FACTOR_SET = DER_POWER_FACTOR_SET')
            lines.append('    ACTION_MODE = DER_ACTION_MODE')
            lines.append('    REACTIVE_POWER_PCT = DER_REACTIVE_POWER_PCT')
            lines.append('    ACTIVE_POWER_PCT = DER_ACTIVE_POWER_PCT')
            lines.append('    OPERATION_MODE = DER_ACTION_MODE')
            lines.append('    REACTIVE_POWER_SET = DER_REACTIVE_POWER_PCT')
            lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # Inverter Control (0x6001-0x6010)
        # ══════════════════════════════════════════════════════════════════
        lines.append('    # =========================================================================')
        lines.append('    # Inverter Control (0x6001-0x6010)')
        lines.append('    # =========================================================================')
        _fill_missing_standard(0x6001, 0x6010)
        if include_iv:
            lines.append('    ')
            lines.append('    # Alias')
            lines.append('    IV_SCAN_COMMAND = IV_CURVE_SCAN')
            lines.append('    IV_SCAN_STATUS = IV_CURVE_SCAN')
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # Power Derating (0x3005)
        # ══════════════════════════════════════════════════════════════════
        _fill_missing_standard(0x3005, 0x3005)
        lines.append('    ')

        # ══════════════════════════════════════════════════════════════════
        # IV Scan Data Registers (0x8000-0x847F)
        # ══════════════════════════════════════════════════════════════════
        if include_iv:
            lines.append('    # =========================================================================')
            lines.append('    # IV Scan Data Registers (0x8000-0x847F)')
            lines.append('    # =========================================================================')
            # 파싱된 IV 레지스터 출력 (표준 이름 사용)
            for m in self.mapping.get('iv_scan', []):
                addr_hex = m.get('mb_addr_hex', '-')
                if addr_hex == '-':
                    continue
                addr = _addr_int(addr_hex)
                std = _std_name(addr) if addr else None
                if std:
                    _emit_reg(addr_hex, std, _std_comment(addr))
            # 표준 IV 레지스터 보충 (Tracker 1-4만)
            _fill_missing_standard(0x8000, 0x8440)
            lines.append('    ')
            lines.append('    # IV Scan data points count')
            lines.append('    IV_SCAN_DATA_POINTS = 64')
            lines.append('    ')
            lines.append('    # IV Scan tracker block size (0x140 = 320 registers)')
            lines.append('    IV_TRACKER_BLOCK_SIZE = 0x140')
        lines.append('')
        lines.append('')

        # ══════════════════════════════════════════════════════════════════
        # Classes: IVScanCommand, IVScanStatus
        # ══════════════════════════════════════════════════════════════════
        if include_iv:
            lines.append('class IVScanCommand:')
            lines.append('    """IV Scan Command values for writing to 0x600D"""')
            lines.append('    NON_ACTIVE = 0x0000  # Stop/Disable IV Scan')
            lines.append('    ACTIVE = 0x0001      # Start IV Scan')
            lines.append('')
            lines.append('')
            lines.append('class IVScanStatus:')
            lines.append('    """IV Scan Status values when reading from 0x600D"""')
            lines.append('    IDLE = 0x0000        # Idle or had been read')
            lines.append('    RUNNING = 0x0001     # IV Scan in progress')
            lines.append('    FINISHED = 0x0002    # IV Scan completed')
            lines.append('    ')
            lines.append('    @classmethod')
            lines.append('    def to_string(cls, status):')
            lines.append('        status_map = {')
            lines.append('            0x0000: "Idle",')
            lines.append('            0x0001: "Running",')
            lines.append('            0x0002: "Finished"')
            lines.append('        }')
            lines.append('        return status_map.get(status, f"Unknown({status})")')
            lines.append('')
            lines.append('')

        # ══════════════════════════════════════════════════════════════════
        # Class: InverterMode
        # ══════════════════════════════════════════════════════════════════
        lines.append('class InverterMode:')
        lines.append('    """Inverter Mode Table (0x101D) - Per Solarize Document"""')
        lines.append('    INITIAL = 0x00')
        lines.append('    STANDBY = 0x01')
        lines.append('    ON_GRID = 0x03')
        lines.append('    OFF_GRID = 0x04')
        lines.append('    FAULT = 0x05')
        lines.append('    SHUTDOWN = 0x09')
        lines.append('    ')
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, mode):')
        lines.append('        mode_map = {')
        lines.append('            0x00: "Initial",')
        lines.append('            0x01: "Standby",')
        lines.append('            0x03: "On-Grid",')
        lines.append('            0x04: "Off-Grid",')
        lines.append('            0x05: "Fault",')
        lines.append('            0x09: "Shutdown"')
        lines.append('        }')
        lines.append('        return mode_map.get(mode, f"Unknown({mode})")')
        lines.append('')
        lines.append('')

        # ══════════════════════════════════════════════════════════════════
        # Class: DerActionMode
        # ══════════════════════════════════════════════════════════════════
        if include_deravm:
            lines.append('class DerActionMode:')
            lines.append('    """DER-AVM Action Mode (0x07D1)"""')
            lines.append('    SELF_CONTROL = 0')
            lines.append('    DER_AVM_CONTROL = 2')
            lines.append('    QV_CONTROL = 5')
            lines.append('')
            lines.append('')

        # ══════════════════════════════════════════════════════════════════
        # Classes: DeviceType, ControlMode, IVScanBodyType
        # ══════════════════════════════════════════════════════════════════
        lines.append('class DeviceType:')
        lines.append('    """Device Type for config file"""')
        lines.append('    RTU = 0')
        lines.append('    INVERTER = 1')
        lines.append('    ENVIRONMENT_SENSOR = 2')
        lines.append('    POWER_METER = 3')
        lines.append('    PROTECTION_RELAY = 4')
        lines.append('    ')
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, dtype):')
        lines.append('        type_map = {')
        lines.append('            0: "RTU",')
        lines.append('            1: "Inverter",')
        lines.append('            2: "Environment Sensor",')
        lines.append('            3: "Power Meter",')
        lines.append('            4: "Protection Relay"')
        lines.append('        }')
        lines.append('        return type_map.get(dtype, f"Unknown({dtype})")')
        lines.append('')
        lines.append('')
        lines.append('class ControlMode:')
        lines.append('    """Control Mode for config file"""')
        lines.append('    NONE = "NONE"')
        lines.append('    DER_AVM = "DER_AVM"')
        lines.append('')
        lines.append('')

        if include_iv:
            lines.append('class IVScanBodyType:')
            lines.append('    """IV Scan Body Type mapping for H05 protocol"""')
            lines.append('    IV_SCAN_RESULT = 12')
            lines.append('    IV_SCAN_DATA = 15')
            lines.append('    TRACKER1_VOLTAGE = 134')
            lines.append('    TRACKER2_VOLTAGE = 139')
            lines.append('    TRACKER3_VOLTAGE = 144')
            lines.append('    TRACKER4_VOLTAGE = 149')
            lines.append('    STRING1_1_CURRENT = 135')
            lines.append('    STRING1_2_CURRENT = 136')
            lines.append('    STRING2_1_CURRENT = 140')
            lines.append('    STRING2_2_CURRENT = 141')
            lines.append('    STRING3_1_CURRENT = 145')
            lines.append('    STRING3_2_CURRENT = 146')
            lines.append('    STRING4_1_CURRENT = 150')
            lines.append('    STRING4_2_CURRENT = 151')
            lines.append('')
            lines.append('')

        # ══════════════════════════════════════════════════════════════════
        # ErrorCode1, ErrorCode2, ErrorCode3
        # ══════════════════════════════════════════════════════════════════
        error_defs = [
            ('ErrorCode1', 'Error Code Table1 (0x101E) - Bit field', [
                'Inverter over dc-bias current', 'Inverter relay abnormal',
                'Remote off', 'Inverter over temperature',
                'GFCI abnormal', 'PV string reverse',
                'System type error', 'Fan abnormal',
                'Dc-link unbalance or under voltage', 'Dc-link over voltage',
                'Internal communication error', 'Software incompatibility',
                'Internal storage error', 'Data inconsistency',
                'Inverter abnormal', 'Boost abnormal',
            ]),
            ('ErrorCode2', 'Error Code Table2 (0x101F) - Bit field', [
                'Grid over voltage', 'Grid under voltage',
                'Grid absent', 'Grid over frequency',
                'Grid under frequency', 'PV over voltage',
                'PV insulation abnormal', 'Leakage current abnormal',
                'Inverter in power limit state', 'Internal power supply abnormal',
                'PV string abnormal', 'PV under voltage',
                'PV irradiation weak', 'Grid abnormal',
                'Arc fault detection', 'AC moving average voltage high',
            ]),
            ('ErrorCode3', 'Error Code Table3 (0x1020) - Bit field', [
                'Reserved', 'Logger/E-Display EEPROM fail',
                'Reserved', 'Single tracker detect warning',
                'AFCI lost', 'Data logger lost',
                'Meter lost', 'Inverter lost',
                'Grid N abnormal', 'Surge Protection Devices (SPD) defective',
                'Parallel ID warning', 'Parallel SYN signal warning',
                'Parallel BAT abnormal', 'Parallel GRID abnormal',
                'Generator voltage abnormal', 'Reserved',
            ]),
        ]
        for ec_name, ec_doc, ec_bits in error_defs:
            lines.append(f'class {ec_name}:')
            lines.append(f'    """{ec_doc}"""')
            lines.append('    BITS = {')
            for i, desc in enumerate(ec_bits):
                lines.append(f'        {i}: "{desc}",')
            lines.append('    }')
            lines.append('')
            lines.append('    @classmethod')
            lines.append('    def decode(cls, value):')
            lines.append('        return [f"E{b}:{d}" for b, d in cls.BITS.items() if value & (1 << b)]')
            lines.append('')
            lines.append('    @classmethod')
            lines.append('    def to_string(cls, value):')
            lines.append('        return ", ".join(cls.decode(value)) if value else "OK"')
            lines.append('')
            lines.append('')

        # ══════════════════════════════════════════════════════════════════
        # SCALE dict (카테고리별 일반 키 — solarize 표준)
        # ══════════════════════════════════════════════════════════════════
        lines.append('# Scale factors - Per Solarize protocol document')
        lines.append('SCALE = {')
        for key, val in _SOLARIZE_STANDARD_SCALE.items():
            # dea/iv 관련은 조건부
            if key.startswith('dea_') and not include_deravm:
                continue
            if key.startswith('iv_') and not include_iv:
                continue
            lines.append(f"    '{key}': {val},")
        lines.append('}')
        lines.append('')
        lines.append('')

        # ══════════════════════════════════════════════════════════════════
        # Helper functions
        # ══════════════════════════════════════════════════════════════════
        lines.append('def registers_to_u32(low, high):')
        lines.append('    """Combine two U16 to U32"""')
        lines.append('    return (high << 16) | low')
        lines.append('')
        lines.append('')
        lines.append('def registers_to_s32(low, high):')
        lines.append('    """Combine two U16 to S32"""')
        lines.append('    value = (high << 16) | low')
        lines.append('    if value >= 0x80000000:')
        lines.append('        value -= 0x100000000')
        lines.append('    return value')
        lines.append('')
        lines.append('')
        lines.append('def get_string_registers(string_num):')
        lines.append('    """Get voltage and current register addresses for a string number (1-24)"""')
        lines.append('    if string_num < 1 or string_num > 24:')
        lines.append('        raise ValueError(f"String number must be 1-24, got {string_num}")')
        lines.append('    base = 0x1050 + (string_num - 1) * 2')
        lines.append("    return {'voltage': base, 'current': base + 1}")
        lines.append('')
        lines.append('')
        lines.append('def get_mppt_registers(mppt_num):')
        lines.append('    """Get voltage, current, power registers for MPPT number (1-9)"""')
        lines.append('    if mppt_num < 1 or mppt_num > 9:')
        lines.append('        raise ValueError(f"MPPT number must be 1-9, got {mppt_num}")')
        lines.append('    if mppt_num <= 3:')
        lines.append('        base = 0x1010 + (mppt_num - 1) * 4')
        lines.append('    elif mppt_num == 4:')
        lines.append('        base = 0x103E')
        lines.append('    else:')
        lines.append('        base = 0x1080 + (mppt_num - 5) * 4')
        lines.append("    return {'voltage': base, 'current': base + 1,")
        lines.append("            'power_low': base + 2, 'power_high': base + 3}")
        lines.append('')
        lines.append('')

        if include_iv:
            lines.append(f'def get_iv_tracker_voltage_registers(tracker_num, data_points=64):')
            lines.append('    """Get IV scan voltage register range for a tracker (1-4)"""')
            lines.append('    if tracker_num < 1 or tracker_num > 4:')
            lines.append('        raise ValueError(f"Tracker number must be 1-4, got {tracker_num}")')
            lines.append(f'    base_addresses = {{')
            lines.append(f'        1: {classname}.IV_TRACKER1_VOLTAGE_BASE,')
            lines.append(f'        2: {classname}.IV_TRACKER2_VOLTAGE_BASE,')
            lines.append(f'        3: {classname}.IV_TRACKER3_VOLTAGE_BASE,')
            lines.append(f'        4: {classname}.IV_TRACKER4_VOLTAGE_BASE,')
            lines.append('    }')
            lines.append('    base = base_addresses[tracker_num]')
            lines.append("    return {'base': base, 'count': data_points, 'end': base + data_points - 1}")
            lines.append('')
            lines.append('')
            lines.append('def get_iv_string_current_registers(mppt_num, string_num, data_points=64):')
            lines.append('    """Get IV scan current register range for a specific string"""')
            lines.append('    if mppt_num < 1 or mppt_num > 4:')
            lines.append('        raise ValueError(f"MPPT number must be 1-4, got {mppt_num}")')
            lines.append('    if string_num < 1 or string_num > 2:')
            lines.append('        raise ValueError(f"String number must be 1-2, got {string_num}")')
            lines.append('    base_addresses = {')
            lines.append(f'        (1, 1): {classname}.IV_STRING1_1_CURRENT_BASE,')
            lines.append(f'        (1, 2): {classname}.IV_STRING1_2_CURRENT_BASE,')
            lines.append(f'        (2, 1): {classname}.IV_STRING2_1_CURRENT_BASE,')
            lines.append(f'        (2, 2): {classname}.IV_STRING2_2_CURRENT_BASE,')
            lines.append(f'        (3, 1): {classname}.IV_STRING3_1_CURRENT_BASE,')
            lines.append(f'        (3, 2): {classname}.IV_STRING3_2_CURRENT_BASE,')
            lines.append(f'        (4, 1): {classname}.IV_STRING4_1_CURRENT_BASE,')
            lines.append(f'        (4, 2): {classname}.IV_STRING4_2_CURRENT_BASE,')
            lines.append('    }')
            lines.append('    base = base_addresses[(mppt_num, string_num)]')
            lines.append("    return {'base': base, 'count': data_points, 'end': base + data_points - 1}")
            lines.append('')
            lines.append('')
            lines.append('def get_iv_string_mapping(total_strings=8, strings_per_mppt=2):')
            lines.append('    """Get mapping of string numbers to voltage/current register addresses"""')
            lines.append('    mapping = []')
            lines.append(f'    data_points = {classname}.IV_SCAN_DATA_POINTS')
            lines.append('    for string_idx in range(total_strings):')
            lines.append('        mppt_num = (string_idx // strings_per_mppt) + 1')
            lines.append('        string_in_mppt = (string_idx % strings_per_mppt) + 1')
            lines.append('        v_regs = get_iv_tracker_voltage_registers(mppt_num, data_points)')
            lines.append('        i_regs = get_iv_string_current_registers(mppt_num, string_in_mppt, data_points)')
            lines.append('        mapping.append({')
            lines.append("            'string_num': string_idx + 1,")
            lines.append("            'total_strings': total_strings,")
            lines.append("            'mppt_num': mppt_num,")
            lines.append("            'string_in_mppt': string_in_mppt,")
            lines.append("            'voltage_base': v_regs['base'],")
            lines.append("            'current_base': i_regs['base'],")
            lines.append("            'data_points': data_points")
            lines.append('        })')
            lines.append('    return mapping')

        # ── SolarizeStatusConverter + StatusConverter alias ──
        lines.append('class SolarizeStatusConverter:')
        lines.append('    """Solarize INVERTER_MODE register already contains InverterMode values."""')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_inverter_mode(cls, raw):')
        lines.append('        """Solarize INVERTER_MODE is already InverterMode-compatible."""')
        lines.append('        return raw')
        lines.append('')
        lines.append('')
        lines.append('# Dynamic-loader alias')
        lines.append('StatusConverter = SolarizeStatusConverter')
        lines.append('')
        lines.append('')

        # ── DATA_TYPES dict ──
        self._emit_data_types_dict(lines)

        return '\n'.join(lines)

    # ══════════════════════════════════════════════════════════════════
    # Kstar code generator
    # ══════════════════════════════════════════════════════════════════

    def _generate_kstar_py(self):
        """Build a Python source string matching kstar_registers.py format.

        Generates code with KstarRegisters class (decimal addresses),
        Block definitions, KstarSystemStatus, KstarInverterStatus,
        KstarStatusConverter, InverterMode, SCALE, and helper functions.
        """
        from datetime import datetime

        manufacturer = self.cg_manufacturer_var.get().strip()
        include_deravm = self.cg_deravm_var.get()

        lines = []
        emitted_addrs = set()
        seen_names = set()

        def _emit_reg(addr, name, comment):
            if addr in emitted_addrs:
                return
            emitted_addrs.add(addr)
            seen_names.add(name)
            if comment:
                lines.append(f"    {name} = {addr}      # {comment}")
            else:
                lines.append(f"    {name} = {addr}")

        def _emit_hex_reg(addr_int, name, comment):
            if addr_int in emitted_addrs:
                return
            emitted_addrs.add(addr_int)
            seen_names.add(name)
            hex_str = f"0x{addr_int:04X}"
            if comment:
                lines.append(f"    {name} = {hex_str}   # {comment}")
            else:
                lines.append(f"    {name} = {hex_str}")

        def _fill_missing_kstar(addr_start, addr_end):
            for addr, (name, comment) in sorted(_KSTAR_ADDR_TO_NAME.items()):
                if addr_start <= addr <= addr_end:
                    if addr >= 0x0700:
                        _emit_hex_reg(addr, name, comment)
                    else:
                        _emit_reg(addr, name, comment)

        # ── Header ──
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        lines.append('#!/usr/bin/env python3')
        lines.append('# -*- coding: utf-8 -*-')
        lines.append('"""')
        lines.append(f'{manufacturer} Inverter Modbus Register Map')
        lines.append(f'Based on KSTAR Inverter Modbus Communication Protocol')
        lines.append(f'Auto-generated by UDP RTU Model Maker')
        lines.append(f'Date: {now_str}')
        lines.append('')
        lines.append('Protocol: Modbus RTU over RS485')
        lines.append('FC04 (Read Input Registers): real-time data')
        lines.append('FC03 (Read Holding Registers): device info')
        lines.append('FC06 (Write Single Register): control commands')
        lines.append('"""')
        lines.append('')
        lines.append('')

        # ── KstarRegisters class ──
        lines.append('class KstarRegisters:')
        lines.append(f'    """')
        lines.append(f'    {manufacturer} Modbus Register Addresses')
        lines.append('    """')
        lines.append('')

        # PV (DC) — limit to actual MPPT count
        mppt_count = int(self.mppt_var.get()) if self.mppt_var.get() else 4
        lines.append('    # =========================================================================')
        lines.append('    # PV (DC) Input — FC04 (Input Registers)')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3000, 3000 + mppt_count - 1)   # PV voltages
        lines.append('')
        _fill_missing_kstar(3012, 3012 + min(mppt_count, 4) - 1)   # PV currents (max 4)
        lines.append('')
        _fill_missing_kstar(3024, 3024 + min(mppt_count, 4) - 1)   # PV powers (max 4)
        lines.append('')

        # Energy
        lines.append('    # =========================================================================')
        lines.append('    # Energy Data — FC04')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3036, 3042)
        lines.append('')

        # Status / Error
        lines.append('    # =========================================================================')
        lines.append('    # Status / Error — FC04')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3044, 3052)
        lines.append('')

        # Temperature / Bus Voltage
        lines.append('    # =========================================================================')
        lines.append('    # Temperature / Bus Voltage — FC04')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3053, 3057)
        lines.append('')

        # AC R-phase
        lines.append('    # =========================================================================')
        lines.append('    # AC Output — R phase (FC04)')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3097, 3100)
        _fill_missing_kstar(3123, 3124)
        lines.append('')

        # AC S-phase
        lines.append('    # =========================================================================')
        lines.append('    # AC Output — S phase (FC04)')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3125, 3132)
        lines.append('')

        # AC T-phase
        lines.append('    # =========================================================================')
        lines.append('    # AC Output — T phase (FC04)')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3133, 3140)
        lines.append('')

        # Load Power
        lines.append('    # =========================================================================')
        lines.append('    # Load Power — FC04')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3144, 3144)
        lines.append('')

        # Additional Energy Data
        lines.append('    # =========================================================================')
        lines.append('    # Additional Energy Data — FC04')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3109, 3109)
        _fill_missing_kstar(3116, 3116)
        _fill_missing_kstar(3121, 3122)
        _fill_missing_kstar(3147, 3147)
        lines.append('')

        # DER-AVM Control Registers
        if include_deravm:
            lines.append('    # =========================================================================')
            lines.append('    # DER-AVM Control Registers — FC03 Read / FC06 Write')
            lines.append('    # =========================================================================')
            lines.append('')
            _fill_missing_kstar(0x07D0, 0x07D3)
            _fill_missing_kstar(0x0834, 0x0834)
            lines.append('')
            lines.append('    # Alias (Solarize compatible)')
            lines.append('    POWER_FACTOR_SET    = DER_POWER_FACTOR_SET')
            lines.append('    ACTION_MODE         = DER_ACTION_MODE')
            lines.append('    REACTIVE_POWER_PCT  = DER_REACTIVE_POWER_PCT')
            lines.append('    OPERATION_MODE      = DER_ACTION_MODE')
            lines.append('    REACTIVE_POWER_SET  = DER_REACTIVE_POWER_PCT')
            lines.append('    ACTIVE_POWER_PCT    = DER_ACTIVE_POWER_PCT')
            lines.append('')

        # DEA-AVM Real-time Monitoring registers
        if include_deravm:
            lines.append('    # =========================================================================')
            lines.append('    # DEA-AVM Real-time Monitoring Data (0x03E8-0x03FD) - For H05 Body Type 14')
            lines.append('    # =========================================================================')
            lines.append('')
            dea_regs = [
                (0x03E8, 'DEA_L1_CURRENT_LOW',            'S32 low, scale 0.1A'),
                (0x03E9, 'DEA_L1_CURRENT_HIGH',           'S32 high'),
                (0x03EA, 'DEA_L2_CURRENT_LOW',            'S32 low, scale 0.1A'),
                (0x03EB, 'DEA_L2_CURRENT_HIGH',           'S32 high'),
                (0x03EC, 'DEA_L3_CURRENT_LOW',            'S32 low, scale 0.1A'),
                (0x03ED, 'DEA_L3_CURRENT_HIGH',           'S32 high'),
                (0x03EE, 'DEA_L1_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                (0x03EF, 'DEA_L1_VOLTAGE_HIGH',           'S32 high'),
                (0x03F0, 'DEA_L2_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                (0x03F1, 'DEA_L2_VOLTAGE_HIGH',           'S32 high'),
                (0x03F2, 'DEA_L3_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                (0x03F3, 'DEA_L3_VOLTAGE_HIGH',           'S32 high'),
                (0x03F4, 'DEA_TOTAL_ACTIVE_POWER_LOW',    'S32 low, scale 0.1kW'),
                (0x03F5, 'DEA_TOTAL_ACTIVE_POWER_HIGH',   'S32 high'),
                (0x03F6, 'DEA_TOTAL_REACTIVE_POWER_LOW',  'S32 low, scale 1 Var'),
                (0x03F7, 'DEA_TOTAL_REACTIVE_POWER_HIGH', 'S32 high'),
                (0x03F8, 'DEA_POWER_FACTOR_LOW',          'S32 low, scale 0.001'),
                (0x03F9, 'DEA_POWER_FACTOR_HIGH',         'S32 high'),
                (0x03FA, 'DEA_FREQUENCY_LOW',             'S32 low, scale 0.1Hz'),
                (0x03FB, 'DEA_FREQUENCY_HIGH',            'S32 high'),
                (0x03FC, 'DEA_STATUS_FLAG_LOW',           'U32 low, bit field'),
                (0x03FD, 'DEA_STATUS_FLAG_HIGH',          'U32 high'),
            ]
            for addr_int, name, comment in dea_regs:
                _emit_hex_reg(addr_int, name, comment)
            lines.append('')

        # Grid Standard
        lines.append('    # =========================================================================')
        lines.append('    # Grid Standard — FC03 (Holding Register)')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3193, 3193)
        lines.append('')

        # Device Info
        lines.append('    # =========================================================================')
        lines.append('    # Device Info — FC03 (Holding Registers)')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3200, 3217)
        lines.append('')

        # Serial Number
        lines.append('    # =========================================================================')
        lines.append('    # Serial Number — FC04')
        lines.append('    # =========================================================================')
        lines.append('')
        _fill_missing_kstar(3228, 3228)
        lines.append('')

        # Block Definitions
        lines.append('    # =========================================================================')
        lines.append('    # Modbus Read Block Definitions')
        lines.append('    # =========================================================================')
        lines.append('')
        lines.append('    # Block1: PV data, energy, status')
        lines.append('    BLOCK1_START = 3000')
        lines.append('    BLOCK1_COUNT = 60           # 3000~3059')
        lines.append('')
        lines.append('    # Block2: temperature, AC data (R phase)')
        lines.append('    BLOCK2_START = 3060')
        lines.append('    BLOCK2_COUNT = 65           # 3060~3124')
        lines.append('')
        lines.append('    # Block3: AC data (S/T phase), inverter power')
        lines.append('    BLOCK3_START = 3125')
        lines.append('    BLOCK3_COUNT = 25           # 3125~3149')
        lines.append('')
        lines.append('    # Block4: device info (FC03)')
        lines.append('    BLOCK4_START = 3200')
        lines.append('    BLOCK4_COUNT = 18           # 3200~3217')
        lines.append('')
        lines.append('    # Block5: serial number, additional data (FC04)')
        lines.append('    BLOCK5_START = 3228')
        lines.append('    BLOCK5_COUNT = 22           # 3228~3249')
        lines.append('')

        # ── Standardized aliases (Solarize-compatible) ──
        lines.append('    # =========================================================================')
        lines.append('    # Standardized aliases (Solarize-compatible, for modbus_handler)')
        lines.append('    # =========================================================================')
        _kstar_std_aliases = {
            'SYSTEM_STATUS':       'INVERTER_MODE',
            'GRID_FREQUENCY':      'FREQUENCY',
            'RADIATOR_TEMP':       'INNER_TEMP',
            'DSP_ALARM_CODE_L':    'ERROR_CODE1',
            'DSP_ALARM_CODE_H':    'ERROR_CODE1_HIGH',
            'DSP_ERROR_CODE_L':    'ERROR_CODE2',
            'DSP_ERROR_CODE_H':    'ERROR_CODE2_HIGH',
            'CUMULATIVE_PRODUCTION_L': 'TOTAL_ENERGY_LOW',
            'CUMULATIVE_PRODUCTION_H': 'TOTAL_ENERGY_HIGH',
            'GRID_R_VOLTAGE':      'R_PHASE_VOLTAGE',
        }
        for native, std in _kstar_std_aliases.items():
            if native in seen_names and std not in seen_names:
                lines.append(f'    {std} = {native}')
                seen_names.add(std)
        lines.append('')
        # TOTAL_ENERGY alias (same as TOTAL_ENERGY_LOW for single-value access)
        if 'CUMULATIVE_PRODUCTION_L' in seen_names and 'TOTAL_ENERGY' not in seen_names:
            lines.append('    TOTAL_ENERGY = CUMULATIVE_PRODUCTION_L')
            seen_names.add('TOTAL_ENERGY')
        # MPPT aliases (PV1 -> MPPT1, etc.)
        for i in range(1, 7):
            pv_v = f'PV{i}_VOLTAGE'
            mppt_v = f'MPPT{i}_VOLTAGE'
            if pv_v in seen_names and mppt_v not in seen_names:
                lines.append(f'    {mppt_v} = {pv_v}')
                seen_names.add(mppt_v)
        for i in range(1, 5):
            pv_i = f'PV{i}_CURRENT'
            mppt_i = f'MPPT{i}_CURRENT'
            if pv_i in seen_names and mppt_i not in seen_names:
                lines.append(f'    {mppt_i} = {pv_i}')
                seen_names.add(mppt_i)
        lines.append('')
        # STRING aliases (Kstar has per-MPPT current, divide by 3 for strings;
        # map STRING{n}_CURRENT -> PV{mppt}_CURRENT as approximation)
        for mppt_idx in range(3):
            pv_i = f'PV{mppt_idx+1}_CURRENT'
            if pv_i in seen_names:
                for s in range(3):
                    str_num = mppt_idx * 3 + s + 1
                    str_name = f'STRING{str_num}_CURRENT'
                    if str_name not in seen_names:
                        lines.append(f'    {str_name} = {pv_i}')
                        seen_names.add(str_name)
        lines.append('')
        # POWER_FACTOR (register 0 = not available, use R-phase power factor area)
        if 'POWER_FACTOR' not in seen_names:
            lines.append('    POWER_FACTOR = 0      # Not available in Kstar protocol')
            seen_names.add('POWER_FACTOR')
        # AC_POWER (Kstar uses calculated AC power from R+S+T phase)
        if 'AC_POWER' not in seen_names:
            lines.append('    AC_POWER = 3150       # Virtual register for calculated total AC power')
            seen_names.add('AC_POWER')
        # PV_POWER (Kstar total PV = sum of PV1+PV2+PV3 power)
        if 'PV_POWER' not in seen_names:
            lines.append('    PV_POWER = 3032       # Virtual register for calculated total PV power')
            seen_names.add('PV_POWER')
        # R/S/T_PHASE_CURRENT and S/T_PHASE_VOLTAGE
        # These are in Block3 area but may not be in _KSTAR_ADDR_TO_NAME
        if 'R_PHASE_CURRENT' not in seen_names:
            lines.append('    R_PHASE_CURRENT = 3148')
            seen_names.add('R_PHASE_CURRENT')
        if 'S_PHASE_VOLTAGE' not in seen_names:
            lines.append('    S_PHASE_VOLTAGE = 3119')
            seen_names.add('S_PHASE_VOLTAGE')
        if 'S_PHASE_CURRENT' not in seen_names:
            lines.append('    S_PHASE_CURRENT = 3156')
            seen_names.add('S_PHASE_CURRENT')
        if 'T_PHASE_VOLTAGE' not in seen_names:
            lines.append('    T_PHASE_VOLTAGE = 3126')
            seen_names.add('T_PHASE_VOLTAGE')
        if 'T_PHASE_CURRENT' not in seen_names:
            lines.append('    T_PHASE_CURRENT = 3163')
            seen_names.add('T_PHASE_CURRENT')
        lines.append('')
        # Additional energy aliases
        _kstar_energy_aliases = {
            'CUMULATIVE_FEEDIN_L':   'CUMULATIVE_FEEDIN_LOW',
            'CUMULATIVE_FEEDIN_H':   'CUMULATIVE_FEEDIN_HIGH',
            'CUMULATIVE_PRODUCTION_L': 'CUMULATIVE_PRODUCTION_LOW',
            'CUMULATIVE_PRODUCTION_H': 'CUMULATIVE_PRODUCTION_HIGH',
            'DSP_ALARM_CODE_L':      'DSP_ALARM_CODE_LOW',
            'DSP_ALARM_CODE_H':      'DSP_ALARM_CODE_HIGH',
            'DSP_ERROR_CODE_L':      'DSP_ERROR_CODE_LOW',
            'DSP_ERROR_CODE_H':      'DSP_ERROR_CODE_HIGH',
            'MONTHLY_PRODUCTION_L':  'MONTHLY_PRODUCTION_LOW',
            'MONTHLY_PRODUCTION_H':  'MONTHLY_PRODUCTION_HIGH',
            'YEARLY_PRODUCTION_L':   'YEARLY_PRODUCTION_LOW',
            'YEARLY_PRODUCTION_H':   'YEARLY_PRODUCTION_HIGH',
        }
        for native, alias in _kstar_energy_aliases.items():
            if native in seen_names and alias not in seen_names:
                lines.append(f'    {alias} = {native}')
                seen_names.add(alias)
        lines.append('')
        lines.append('')

        # ── KstarSystemStatus class ──
        lines.append('class KstarSystemStatus:')
        lines.append('    """')
        lines.append('    System Status register (3046) value definitions')
        lines.append('    """')
        status_codes = [
            (0, 'INITIALIZE', 'Initialize'),
            (1, 'STAND_BY', 'Stand-by'),
            (2, 'STARTING', 'Starting'),
            (3, 'SELF_CHECK', 'Self-check'),
            (4, 'RUNNING', 'Running'),
            (5, 'RECOVERY_FAULT', 'Recovery Fault'),
            (6, 'PERMANENT_FAULT', 'Permanent Fault'),
            (7, 'UPGRADING', 'Upgrading'),
            (8, 'SELF_CHARGING', 'Self Charging'),
            (9, 'SELF_CHECK_TIMEOUT', 'Self Check Timeout'),
            (10, 'FAN_CHECK', 'Fan Check'),
            (11, 'S1_GROUND_DETECT', 'S1 Ground Detect'),
            (12, 'PRE_RUNNING', 'Pre Running'),
            (13, 'MCU_BURN', 'MCU Burn'),
        ]
        for val, name, _ in status_codes:
            lines.append(f'    {name:25s} = {val}')
        lines.append('')
        lines.append('    _STATUS_MAP = {')
        for val, _, desc in status_codes:
            lines.append(f'        {val}:  "{desc}",')
        lines.append('    }')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, status):')
        lines.append('        return cls._STATUS_MAP.get(status, f"Unknown({status})")')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def is_running(cls, status):')
        lines.append('        return status == cls.RUNNING')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def is_fault(cls, status):')
        lines.append('        return status in (cls.RECOVERY_FAULT, cls.PERMANENT_FAULT)')
        lines.append('')
        lines.append('')

        # ── KstarInverterStatus class ──
        lines.append('class KstarInverterStatus:')
        lines.append('    """Inverter Status register (3047) value definitions"""')
        inv_status_codes = [
            (0, 'STAND_BY', 'Stand-by'),
            (1, 'GRID_TIED', 'Grid-tied'),
            (2, 'GRID_CHARGING', 'Grid Charging'),
            (3, 'GRID_TIED_TO_OFF', 'Grid-tied to Off-Grid'),
        ]
        for val, name, _ in inv_status_codes:
            lines.append(f'    {name:25s} = {val}')
        lines.append('')
        lines.append('    _STATUS_MAP = {')
        for val, _, desc in inv_status_codes:
            lines.append(f'        {val}: "{desc}",')
        lines.append('    }')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, status):')
        lines.append('        return cls._STATUS_MAP.get(status, f"Unknown({status})")')
        lines.append('')
        lines.append('')

        # ── InverterMode class (must be before KstarStatusConverter) ──
        self._emit_inverter_mode_class(lines)

        # ── KstarStatusConverter class ──
        lines.append('class KstarStatusConverter:')
        lines.append('    """Kstar System Status (reg 3046) to InverterMode conversion.')
        lines.append('')
        lines.append('    Maps manufacturer-specific running state codes to standard')
        lines.append('    Solarize InverterMode values used by the RTU H01 packet.')
        lines.append('    """')
        lines.append('')
        lines.append('    _STATUS_NAMES = {')
        status_names = [
            (0x0000, 'Initialize'),
            (0x0001, 'Stand-by'),
            (0x0002, 'Starting'),
            (0x0003, 'Self-check'),
            (0x0004, 'Running'),
            (0x0005, 'Recovery Fault'),
            (0x0006, 'Permanent Fault'),
            (0x0007, 'Upgrading'),
            (0x0008, 'Self Charging'),
            (0x0009, 'Self Check Timeout'),
            (0x000A, 'Fan Check'),
            (0x000B, 'S1 Ground Detect'),
            (0x000C, 'Pre Running'),
            (0x000D, 'MCU Burn'),
        ]
        for val, desc in status_names:
            lines.append(f'        0x{val:04X}: "{desc}",')
        lines.append('    }')
        lines.append('')
        lines.append('    _CONVERSION_MAP = {')
        conversion_map = [
            (0x0000, 'InverterMode.INITIAL',  'Initialize'),
            (0x0001, 'InverterMode.STANDBY',  'Stand-by'),
            (0x0002, 'InverterMode.STANDBY',  'Starting'),
            (0x0003, 'InverterMode.STANDBY',  'Self-check'),
            (0x0004, 'InverterMode.ON_GRID',  'Running (On-Grid)'),
            (0x0005, 'InverterMode.FAULT',    'Recovery Fault'),
            (0x0006, 'InverterMode.FAULT',    'Permanent Fault'),
            (0x0007, 'InverterMode.SHUTDOWN', 'Upgrading'),
            (0x0008, 'InverterMode.STANDBY',  'Self Charging'),
            (0x0009, 'InverterMode.STANDBY',  'Self Check Timeout'),
            (0x000A, 'InverterMode.STANDBY',  'Fan Check'),
            (0x000B, 'InverterMode.STANDBY',  'S1 Ground Detect'),
            (0x000C, 'InverterMode.STANDBY',  'Pre Running'),
            (0x000D, 'InverterMode.STANDBY',  'MCU Burn'),
        ]
        for val, mode, comment in conversion_map:
            lines.append(f'        0x{val:04X}: {mode + ",":32s} # {comment}')
        lines.append('    }')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_inverter_mode(cls, raw_status):')
        lines.append('        """Convert Kstar system status to Solarize InverterMode code."""')
        lines.append('        return cls._CONVERSION_MAP.get(raw_status, InverterMode.STANDBY)')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, raw_status):')
        lines.append('        """Get human-readable name for Kstar system status."""')
        lines.append('        return cls._STATUS_NAMES.get(raw_status, f"Unknown({raw_status})")')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def is_running(cls, raw_status):')
        lines.append('        return raw_status == 0x0004')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def is_fault(cls, raw_status):')
        lines.append('        return raw_status in (0x0005, 0x0006)')
        lines.append('')
        lines.append('')

        # ── RegisterMap alias ──
        lines.append('# RegisterMap alias (modbus_handler dynamic loading)')
        lines.append('RegisterMap = KstarRegisters')
        lines.append('')
        lines.append('')

        # ── SCALE dict ──
        lines.append('# =============================================================================')
        lines.append('# Scale constants')
        lines.append('# =============================================================================')
        lines.append('')
        lines.append('SCALE = {')
        for key, val in _KSTAR_STANDARD_SCALE.items():
            lines.append(f"    '{key}': {val!r},")
        lines.append('}')
        lines.append('')
        lines.append('')

        # ── Helper functions ──
        lines.append('# =============================================================================')
        lines.append('# Helper functions')
        lines.append('# =============================================================================')
        lines.append('')
        lines.append('def registers_to_u32(low, high):')
        lines.append('    """Combine two U16 registers to U32 (low word first)"""')
        lines.append('    return ((high & 0xFFFF) << 16) | (low & 0xFFFF)')
        lines.append('')
        lines.append('')
        lines.append('def registers_to_s32(low, high):')
        lines.append('    """Combine two U16 registers to S32 (low word first)"""')
        lines.append('    value = ((high & 0xFFFF) << 16) | (low & 0xFFFF)')
        lines.append('    if value >= 0x80000000:')
        lines.append('        value -= 0x100000000')
        lines.append('    return value')
        lines.append('')
        lines.append('')
        lines.append('def decode_ascii_registers(registers):')
        lines.append('    """Convert U16 register array to ASCII string"""')
        lines.append('    result = []')
        lines.append('    for reg in registers:')
        lines.append('        high_byte = (reg >> 8) & 0xFF')
        lines.append('        low_byte = reg & 0xFF')
        lines.append('        if high_byte:')
        lines.append('            result.append(chr(high_byte))')
        lines.append('        if low_byte:')
        lines.append('            result.append(chr(low_byte))')
        lines.append("    return ''.join(result).rstrip('\\x00').strip()")
        lines.append('')
        lines.append('')
        lines.append('def calc_pv_total_power(block1_data):')
        lines.append('    """')
        lines.append('    Calculate PV total power from Block1 data (MPPT1~3)')
        lines.append('    """')
        lines.append('    base = KstarRegisters.BLOCK1_START')
        lines.append('    p1 = block1_data[KstarRegisters.PV1_POWER - base]')
        lines.append('    p2 = block1_data[KstarRegisters.PV2_POWER - base]')
        lines.append('    p3 = block1_data[KstarRegisters.PV3_POWER - base]')
        lines.append('    return p1 + p2 + p3')
        lines.append('')
        lines.append('')
        lines.append('def calc_ac_total_power(block3_data):')
        lines.append('    """')
        lines.append('    Calculate AC total power from Block3 data (R+S+T phase)')
        lines.append('    """')
        lines.append('    base = KstarRegisters.BLOCK3_START')
        lines.append('    r_raw = block3_data[KstarRegisters.INV_R_POWER - base]')
        lines.append('    s_raw = block3_data[KstarRegisters.INV_S_POWER - base]')
        lines.append('    t_raw = block3_data[KstarRegisters.INV_T_POWER - base]')
        lines.append('')
        lines.append('    def to_s16(v):')
        lines.append('        return v - 0x10000 if v >= 0x8000 else v')
        lines.append('')
        lines.append('    return to_s16(r_raw) + to_s16(s_raw) + to_s16(t_raw)')
        lines.append('')
        lines.append('')
        lines.append(f'def get_mppt_data(block1_data, mppt_num):')
        lines.append('    """')
        lines.append(f'    Get voltage/current/power for MPPT number (1~{mppt_count})')
        lines.append('    """')
        lines.append(f'    if mppt_num < 1 or mppt_num > {mppt_count}:')
        lines.append(f'        raise ValueError(f"MPPT number must be 1-{mppt_count}, got {{mppt_num}}")')
        lines.append('')
        lines.append('    base = KstarRegisters.BLOCK1_START')
        lines.append('    voltage_regs = [')
        for i in range(1, mppt_count + 1):
            lines.append(f'        KstarRegisters.PV{i}_VOLTAGE,')
        lines.append('    ]')
        lines.append('    current_regs = [')
        for i in range(1, min(mppt_count, 4) + 1):
            lines.append(f'        KstarRegisters.PV{i}_CURRENT,')
        lines.append('    ]')
        lines.append('    power_regs = [')
        for i in range(1, min(mppt_count, 4) + 1):
            lines.append(f'        KstarRegisters.PV{i}_POWER,')
        lines.append('    ]')
        lines.append('')
        lines.append('    idx = mppt_num - 1')
        lines.append('    raw_v = block1_data[voltage_regs[idx] - base]')
        lines.append('    raw_i = block1_data[current_regs[min(idx, len(current_regs) - 1)] - base]')
        lines.append('    raw_p = block1_data[power_regs[min(idx, len(power_regs) - 1)] - base]')
        lines.append('')
        lines.append('    return {')
        lines.append("        'voltage': raw_v * SCALE['pv_voltage'],")
        lines.append("        'current': raw_i * SCALE['pv_current'],")
        lines.append("        'power':   raw_p * SCALE['pv_power'],")
        lines.append("        'raw_voltage': raw_v,")
        lines.append("        'raw_current': raw_i,")
        lines.append('    }')
        lines.append('')
        lines.append('')
        cur_count = min(mppt_count, 4)
        lines.append(f'def get_string_currents(block1_data, strings_per_mppt=3):')
        lines.append('    """')
        lines.append('    Get string currents (MPPT current divided by strings_per_mppt)')
        lines.append('    """')
        lines.append('    base = KstarRegisters.BLOCK1_START')
        lines.append('    current_regs = [')
        for i in range(1, cur_count + 1):
            lines.append(f'        KstarRegisters.PV{i}_CURRENT,')
        lines.append('    ]')
        lines.append('    result = []')
        lines.append(f'    for mppt_idx in range({cur_count}):')
        lines.append('        mppt_num = mppt_idx + 1')
        lines.append('        raw_mppt_i = block1_data[current_regs[mppt_idx] - base]')
        lines.append('        divided_raw = raw_mppt_i // strings_per_mppt')
        lines.append('')
        lines.append('        for s in range(strings_per_mppt):')
        lines.append('            string_num = mppt_idx * strings_per_mppt + s + 1')
        lines.append('            result.append({')
        lines.append("                'string_num':  string_num,")
        lines.append("                'mppt_num':    mppt_num,")
        lines.append("                'current':     divided_raw * SCALE['pv_current'],")
        lines.append("                'raw_current': divided_raw,")
        lines.append('            })')
        lines.append('    return result')
        lines.append('')
        lines.append('')
        lines.append('def get_cumulative_energy_wh(block1_data):')
        lines.append('    """')
        lines.append('    Get cumulative energy in Wh from Block1 data')
        lines.append('    CUMULATIVE_PRODUCTION is U32, scale x0.1kWh -> Wh = raw * 100')
        lines.append('    """')
        lines.append('    base = KstarRegisters.BLOCK1_START')
        lines.append('    low  = block1_data[KstarRegisters.CUMULATIVE_PRODUCTION_L - base]')
        lines.append('    high = block1_data[KstarRegisters.CUMULATIVE_PRODUCTION_H - base]')
        lines.append('    raw_u32 = registers_to_u32(low, high)')
        lines.append('    return raw_u32 * 100')
        lines.append('')
        lines.append('')

        # ── StatusConverter class (from status_map in JSON) ──
        status_map = getattr(self, '_loaded_status_map', None)
        if status_map and 'values' in status_map:
            self._emit_status_converter_class(lines, manufacturer, status_map)

        # ── FaultCode class (from error_codes in JSON) ──
        error_codes = getattr(self, '_loaded_error_codes', None)
        if error_codes and 'values' in error_codes:
            self._emit_fault_code_class(lines, manufacturer, error_codes)

        # ── DATA_TYPES dict ──
        self._emit_data_types_dict(lines)

        # ── StatusConverter alias ──
        lines.append('# Dynamic-loader alias')
        lines.append('StatusConverter = KstarStatusConverter')
        lines.append('')
        lines.append('')

        return '\n'.join(lines)

    # ══════════════════════════════════════════════════════════════════
    # Huawei code generator
    # ══════════════════════════════════════════════════════════════════

    def _generate_huawei_py(self):
        """Build a Python source string matching huawei_registers.py format.

        Generates code with HuaweiRegisters class (decimal addresses),
        HuaweiStatusConverter, InverterMode, SCALE, and big-endian helpers.
        """
        from datetime import datetime

        manufacturer = self.cg_manufacturer_var.get().strip()
        include_deravm = self.cg_deravm_var.get()

        lines = []
        emitted_addrs = set()
        seen_names = set()

        def _emit_reg(addr, name, comment):
            if addr in emitted_addrs:
                return
            emitted_addrs.add(addr)
            seen_names.add(name)
            if comment:
                lines.append(f"    {name} = {addr}   # {comment}")
            else:
                lines.append(f"    {name} = {addr}")

        def _emit_hex_reg(addr_int, name, comment):
            if addr_int in emitted_addrs:
                return
            emitted_addrs.add(addr_int)
            seen_names.add(name)
            hex_str = f"0x{addr_int:04X}"
            if comment:
                lines.append(f"    {name} = {hex_str}   # {comment}")
            else:
                lines.append(f"    {name} = {hex_str}")

        def _fill_missing_huawei(addr_start, addr_end):
            for addr, (name, comment) in sorted(_HUAWEI_ADDR_TO_NAME.items()):
                if addr_start <= addr <= addr_end:
                    _emit_hex_reg(addr, name, comment)

        # ── Header ──
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        lines.append('# ============================================================================')
        lines.append(f'# {manufacturer} Modbus Register Map')
        lines.append(f'# Auto-generated by UDP RTU Model Maker')
        lines.append(f'# Date: {now_str}')
        lines.append('# Protocol: Modbus RTU / TCP, Function Code 03 (Read Holding Registers)')
        lines.append('# Byte order: Big-endian (MSB first)')
        lines.append('# ============================================================================')
        lines.append('')
        lines.append('')
        lines.append('')

        # ── HuaweiRegisters class ──
        lines.append('# ============================================================================')
        lines.append('# Register Address Constants')
        lines.append('# ============================================================================')
        lines.append('')
        lines.append('class HuaweiRegisters:')
        lines.append(f'    """Huawei SUN2000 Holding Register (FC03) address map"""')
        lines.append('')

        # Device Status
        lines.append('    # --- Device Status ---')
        _fill_missing_huawei(32000, 32005)
        lines.append('')

        # PV String Inputs
        lines.append('    # --- PV String Inputs (DC side) ---')
        lines.append('    PV_STRING_BASE = 0x7D10   # First register of PV string block')
        lines.append('    PV_STRING_COUNT = 16     # 16 registers total (8 strings x 2 regs each)')
        lines.append('')
        lines.append('    # Individual PV string registers')
        _fill_missing_huawei(32016, 32031)
        lines.append('')

        # DC Input Power
        lines.append('    # --- DC Input Power ---')
        _fill_missing_huawei(32064, 32064)
        lines.append('')

        # AC Grid
        lines.append('    # --- AC Grid (3-phase) ---')
        _fill_missing_huawei(32069, 32085)
        lines.append('')

        # Temperature
        lines.append('    # --- Temperature ---')
        _fill_missing_huawei(32087, 32087)
        lines.append('')

        # Energy
        lines.append('    # --- Energy ---')
        _fill_missing_huawei(32106, 32106)
        lines.append('')

        # DER-AVM Control Registers
        if include_deravm:
            lines.append('    # =========================================================================')
            lines.append('    # DER-AVM Control Registers — FC03 Read / FC06 Write')
            lines.append('    # =========================================================================')
            lines.append('')
            der_ctrl_regs = [
                (0x07D0, 'DER_POWER_FACTOR_SET',    'S16, scale 0.001'),
                (0x07D1, 'DER_ACTION_MODE',         'U16: 0=self, 2=DER-AVM, 5=Q(V)'),
                (0x07D2, 'DER_REACTIVE_POWER_PCT',  'S16, scale 0.1%'),
                (0x07D3, 'DER_ACTIVE_POWER_PCT',    'U16, scale 0.1%'),
                (0x0834, 'INVERTER_ON_OFF',         'U16: 0=ON, 1=OFF'),
            ]
            for addr_int, name, comment in der_ctrl_regs:
                _emit_hex_reg(addr_int, name, comment)
            lines.append('')
            lines.append('    # Alias (Solarize compatible)')
            lines.append('    POWER_FACTOR_SET    = DER_POWER_FACTOR_SET')
            lines.append('    ACTION_MODE         = DER_ACTION_MODE')
            lines.append('    REACTIVE_POWER_PCT  = DER_REACTIVE_POWER_PCT')
            lines.append('    OPERATION_MODE      = DER_ACTION_MODE')
            lines.append('    REACTIVE_POWER_SET  = DER_REACTIVE_POWER_PCT')
            lines.append('    ACTIVE_POWER_PCT    = DER_ACTIVE_POWER_PCT')
            lines.append('')

        # DEA-AVM Real-time Monitoring registers
        if include_deravm:
            lines.append('    # =========================================================================')
            lines.append('    # DEA-AVM Real-time Monitoring Data (0x03E8-0x03FD) - For H05 Body Type 14')
            lines.append('    # =========================================================================')
            lines.append('')
            dea_regs = [
                (0x03E8, 'DEA_L1_CURRENT_LOW',            'S32 low, scale 0.1A'),
                (0x03E9, 'DEA_L1_CURRENT_HIGH',           'S32 high'),
                (0x03EA, 'DEA_L2_CURRENT_LOW',            'S32 low, scale 0.1A'),
                (0x03EB, 'DEA_L2_CURRENT_HIGH',           'S32 high'),
                (0x03EC, 'DEA_L3_CURRENT_LOW',            'S32 low, scale 0.1A'),
                (0x03ED, 'DEA_L3_CURRENT_HIGH',           'S32 high'),
                (0x03EE, 'DEA_L1_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                (0x03EF, 'DEA_L1_VOLTAGE_HIGH',           'S32 high'),
                (0x03F0, 'DEA_L2_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                (0x03F1, 'DEA_L2_VOLTAGE_HIGH',           'S32 high'),
                (0x03F2, 'DEA_L3_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                (0x03F3, 'DEA_L3_VOLTAGE_HIGH',           'S32 high'),
                (0x03F4, 'DEA_TOTAL_ACTIVE_POWER_LOW',    'S32 low, scale 0.1kW'),
                (0x03F5, 'DEA_TOTAL_ACTIVE_POWER_HIGH',   'S32 high'),
                (0x03F6, 'DEA_TOTAL_REACTIVE_POWER_LOW',  'S32 low, scale 1 Var'),
                (0x03F7, 'DEA_TOTAL_REACTIVE_POWER_HIGH', 'S32 high'),
                (0x03F8, 'DEA_POWER_FACTOR_LOW',          'S32 low, scale 0.001'),
                (0x03F9, 'DEA_POWER_FACTOR_HIGH',         'S32 high'),
                (0x03FA, 'DEA_FREQUENCY_LOW',             'S32 low, scale 0.1Hz'),
                (0x03FB, 'DEA_FREQUENCY_HIGH',            'S32 high'),
                (0x03FC, 'DEA_STATUS_FLAG_LOW',           'U32 low, bit field'),
                (0x03FD, 'DEA_STATUS_FLAG_HIGH',          'U32 high'),
            ]
            for addr_int, name, comment in dea_regs:
                _emit_hex_reg(addr_int, name, comment)
            lines.append('')

        lines.append('')

        # ── Standardized aliases (Solarize-compatible names for modbus_handler) ──
        # _HUAWEI_ADDR_TO_NAME uses Huawei-native names; add standard aliases.
        _huawei_alias_map = {
            'ACTIVE_POWER':     'AC_POWER',
            'PHASE_A_VOLTAGE':  'R_PHASE_VOLTAGE',
            'PHASE_B_VOLTAGE':  'S_PHASE_VOLTAGE',
            'PHASE_C_VOLTAGE':  'T_PHASE_VOLTAGE',
            'PHASE_A_CURRENT':  'R_PHASE_CURRENT',
            'PHASE_B_CURRENT':  'S_PHASE_CURRENT',
            'PHASE_C_CURRENT':  'T_PHASE_CURRENT',
            'GRID_FREQUENCY':   'FREQUENCY',
            'INTERNAL_TEMP':    'INNER_TEMP',
            'RUNNING_STATUS':   'INVERTER_MODE',
            'FAULT_CODE_1':     'ERROR_CODE1',
            'FAULT_CODE_2':     'ERROR_CODE2',
            'ACCUMULATED_ENERGY': 'TOTAL_ENERGY',
            'INPUT_POWER':      'PV_POWER',
        }
        alias_lines = []
        for native_name, std_name in _huawei_alias_map.items():
            if native_name in seen_names and std_name not in seen_names:
                alias_lines.append(f'    {std_name} = {native_name}')
                seen_names.add(std_name)
        if alias_lines:
            lines.append('    # =========================================================================')
            lines.append('    # Standardized aliases (Solarize-compatible)')
            lines.append('    # =========================================================================')
            for al in alias_lines:
                lines.append(al)
            lines.append('')

        # ── MPPT aliases (derived from PV string pairs) ──
        # MPPT1 = PV1, MPPT2 = PV3, MPPT3 = PV5, MPPT4 = PV7
        mppt_alias_lines = []
        mppt_pv_map = [(1, 1), (2, 3), (3, 5), (4, 7)]
        for mppt_n, pv_n in mppt_pv_map:
            pv_v = f'PV{pv_n}_VOLTAGE'
            pv_i = f'PV{pv_n}_CURRENT'
            mppt_v = f'MPPT{mppt_n}_VOLTAGE'
            mppt_i = f'MPPT{mppt_n}_CURRENT'
            if pv_v in seen_names and mppt_v not in seen_names:
                mppt_alias_lines.append(f'    {mppt_v} = {pv_v}')
                seen_names.add(mppt_v)
            if pv_i in seen_names and mppt_i not in seen_names:
                mppt_alias_lines.append(f'    {mppt_i} = {pv_i}')
                seen_names.add(mppt_i)
        if mppt_alias_lines:
            lines.append('    # =========================================================================')
            lines.append('    # MPPT aliases (derived from PV string pairs)')
            lines.append('    # =========================================================================')
            for al in mppt_alias_lines:
                lines.append(al)
            lines.append('')

        # ── STRING aliases (PV string currents) ──
        string_alias_lines = []
        for i in range(1, 9):
            pv_i_name = f'PV{i}_CURRENT'
            str_name = f'STRING{i}_CURRENT'
            if pv_i_name in seen_names and str_name not in seen_names:
                string_alias_lines.append(f'    {str_name} = {pv_i_name}')
                seen_names.add(str_name)
        if string_alias_lines:
            lines.append('    # =========================================================================')
            lines.append('    # STRING aliases (PV string currents)')
            lines.append('    # =========================================================================')
            for al in string_alias_lines:
                lines.append(al)
            lines.append('')

        lines.append('')

        # ── SCALE (generic 5 keys only, for modbus_handler compatibility) ──
        _huawei_generic_scale = {
            'voltage': 0.1,
            'current': 0.01,
            'power': 1.0,
            'frequency': 0.01,
            'power_factor': 0.001,
        }
        lines.append('# ============================================================================')
        lines.append('# Scale Factors')
        lines.append('# ============================================================================')
        lines.append('')
        lines.append('SCALE = {')
        for key, val in _huawei_generic_scale.items():
            lines.append(f"    '{key}': {val!r},")
        lines.append('}')
        lines.append('')
        lines.append('')

        # ── InverterMode class (must be before HuaweiStatusConverter) ──
        lines.append('# ============================================================================')
        lines.append('# Solarize compatible (modbus_handler dynamic loading)')
        lines.append('# ============================================================================')
        lines.append('')
        self._emit_inverter_mode_class(lines)

        # ── HuaweiStatusConverter class ──
        lines.append('# ============================================================================')
        lines.append('# Running Status -> InverterMode Converter')
        lines.append('# ============================================================================')
        lines.append('')
        lines.append('class HuaweiStatusConverter:')
        lines.append('    """Convert Huawei running_status register value to InverterMode"""')
        lines.append('')
        huawei_statuses = [
            ('STATUS_STANDBY_INIT',       '0x0000', 'Standby: initializing'),
            ('STATUS_STANDBY_INSULATION', '0x0001', 'Standby: insulation resistance detecting'),
            ('STATUS_STANDBY_SUNLIGHT',   '0x0002', 'Standby: sunlight detecting'),
            ('STATUS_STANDBY_NETWORK',    '0x0003', 'Standby: power network detecting'),
            ('STATUS_STARTING',           '0x0100', 'Starting'),
            ('STATUS_ON_GRID',            '0x0200', 'On-grid (normal operation)'),
            ('STATUS_ON_GRID_DERATING',   '0x0201', 'On-grid: derating due to power'),
            ('STATUS_ON_GRID_TEMP',       '0x0202', 'On-grid: derating due to temperature'),
            ('STATUS_FAULT',              '0x0300', 'Shutdown: fault'),
            ('STATUS_SHUTDOWN_CMD',       '0x0301', 'Shutdown: command'),
            ('STATUS_SHUTDOWN_OVGR',      '0x0302', 'Shutdown: OVGR'),
            ('STATUS_SHUTDOWN_COMM',      '0x0303', 'Shutdown: communication disconnected'),
            ('STATUS_SHUTDOWN_POWER',     '0x0304', 'Shutdown: power limited'),
            ('STATUS_SHUTDOWN_MANUAL',    '0x0305', 'Shutdown: manual startup required'),
            ('STATUS_SHUTDOWN_DC',        '0x0306', 'Shutdown: DC switch disconnected'),
        ]
        for name, val, comment in huawei_statuses:
            lines.append(f'    {name:30s} = {val}  # {comment}')
        lines.append('')
        lines.append('    _STATUS_NAMES = {')
        for name, val, comment in huawei_statuses:
            lines.append(f'        {val}: "{comment}",')
        lines.append('    }')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_inverter_mode(cls, raw_status: int) -> int:')
        lines.append('        """Convert raw running_status to InverterMode code"""')
        lines.append('        if raw_status == cls.STATUS_ON_GRID or raw_status == cls.STATUS_ON_GRID_DERATING \\')
        lines.append('                or raw_status == cls.STATUS_ON_GRID_TEMP:')
        lines.append('            return InverterMode.ON_GRID')
        lines.append('        elif raw_status == cls.STATUS_FAULT:')
        lines.append('            return InverterMode.FAULT')
        lines.append('        elif cls.STATUS_STANDBY_INIT <= raw_status <= cls.STATUS_STARTING \\')
        lines.append('                or cls.STATUS_SHUTDOWN_CMD <= raw_status <= cls.STATUS_SHUTDOWN_DC:')
        lines.append('            return InverterMode.STANDBY')
        lines.append('        else:')
        lines.append('            return InverterMode.INITIAL')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, raw_status: int) -> str:')
        lines.append('        """Get human-readable name for Huawei running status."""')
        lines.append('        return cls._STATUS_NAMES.get(raw_status, f"Unknown({raw_status})")')
        lines.append('')
        lines.append('')

        # ── Helper Functions ──
        lines.append('# ============================================================================')
        lines.append('# Helper Functions')
        lines.append('# ============================================================================')
        lines.append('')
        lines.append('def registers_to_s32(hi: int, lo: int) -> int:')
        lines.append('    """Combine two U16 registers into a signed 32-bit integer (big-endian)"""')
        lines.append('    val = ((hi & 0xFFFF) << 16) | (lo & 0xFFFF)')
        lines.append('    if val >= 0x80000000:')
        lines.append('        val -= 0x100000000')
        lines.append('    return val')
        lines.append('')
        lines.append('')
        lines.append('def registers_to_u32(hi: int, lo: int) -> int:')
        lines.append('    """Combine two U16 registers into an unsigned 32-bit integer (big-endian)"""')
        lines.append('    return ((hi & 0xFFFF) << 16) | (lo & 0xFFFF)')
        lines.append('')
        lines.append('')
        lines.append('def s16(val: int) -> int:')
        lines.append('    """Convert unsigned U16 register value to signed S16"""')
        lines.append('    if val >= 0x8000:')
        lines.append('        val -= 0x10000')
        lines.append('    return val')
        lines.append('')
        lines.append('')
        lines.append('def get_pv_string_data(regs: list) -> list:')
        lines.append('    """')
        lines.append('    Parse 16 consecutive registers (PV_STRING_BASE..+15) into per-string dicts.')
        lines.append('    Layout: [V1, I1, V2, I2, ..., V8, I8]')
        lines.append("    Returns list of 8 dicts: {'voltage': raw_U16, 'current': raw_S16}")
        lines.append('    """')
        lines.append('    result = []')
        lines.append('    for i in range(8):')
        lines.append('        v = regs[i * 2]          # U16, 0.1V')
        lines.append('        c = s16(regs[i * 2 + 1]) # S16, 0.01A')
        lines.append('        result.append({')
        lines.append("            'voltage': max(0, v),")
        lines.append("            'current': max(0, c),")
        lines.append('        })')
        lines.append('    return result')
        lines.append('')
        lines.append('')
        lines.append('def get_mppt_from_strings(pv_data: list) -> list:')
        lines.append('    """')
        lines.append('    Derive 4 MPPT values from 8 PV string measurements.')
        lines.append('    MPPT grouping: MPPT_n uses PV(2n-1) and PV(2n).')
        lines.append('    """')
        lines.append('    mppt = []')
        lines.append('    for i in range(4):')
        lines.append('        s1 = pv_data[i * 2]')
        lines.append('        s2 = pv_data[i * 2 + 1]')
        lines.append('')
        lines.append('        # Voltage: average of two strings (0.1V unit)')
        lines.append("        if s1['voltage'] > 0 and s2['voltage'] > 0:")
        lines.append("            mppt_v = (s1['voltage'] + s2['voltage']) // 2")
        lines.append('        else:')
        lines.append("            mppt_v = max(s1['voltage'], s2['voltage'])")
        lines.append('')
        lines.append('        # Current: sum of two strings (0.01A unit)')
        lines.append("        mppt_c = s1['current'] + s2['current']")
        lines.append('')
        lines.append("        mppt.append({'voltage': mppt_v, 'current': mppt_c})")
        lines.append('    return mppt')
        lines.append('')
        lines.append('')
        lines.append('def get_string_currents(pv_data: list) -> list:')
        lines.append('    """')
        lines.append('    Extract string currents for H01 string array.')
        lines.append('    Returns list of 8 raw current values (0.01A unit).')
        lines.append('    """')
        lines.append("    return [p['current'] for p in pv_data]")
        lines.append('')
        lines.append('')
        lines.append('def get_cumulative_energy_wh(hi: int, lo: int) -> int:')
        lines.append('    """')
        lines.append('    Convert accumulated energy U32 register pair to Wh.')
        lines.append('    Register unit: 1 kWh -> multiply x 1000 for Wh.')
        lines.append('    """')
        lines.append('    kwh = registers_to_u32(hi, lo)')
        lines.append('    return kwh * 1000')
        lines.append('')
        lines.append('')

        # ── StatusConverter class (from status_map in JSON) ──
        # Skip if HuaweiStatusConverter was already emitted manually above
        status_map = getattr(self, '_loaded_status_map', None)
        if status_map and 'values' in status_map:
            if 'class HuaweiStatusConverter:' not in '\n'.join(lines):
                self._emit_status_converter_class(lines, manufacturer, status_map)

        # ── FaultCode class (from error_codes in JSON) ──
        # Skip if a FaultCode class was already emitted manually above
        error_codes = getattr(self, '_loaded_error_codes', None)
        if error_codes and 'values' in error_codes:
            fault_class_name = f'{manufacturer.capitalize()}FaultCode'
            if f'class {fault_class_name}:' not in '\n'.join(lines):
                self._emit_fault_code_class(lines, manufacturer, error_codes)

        # ── RegisterMap alias ──
        lines.append('# RegisterMap alias (modbus_handler dynamic loading)')
        lines.append('RegisterMap = HuaweiRegisters')
        lines.append('')
        lines.append('')

        # ── DATA_TYPES dict ──
        self._emit_data_types_dict(lines)

        # ── StatusConverter alias ──
        lines.append('# Dynamic-loader alias')
        lines.append('StatusConverter = HuaweiStatusConverter')
        lines.append('')
        lines.append('')

        return '\n'.join(lines)

    # ══════════════════════════════════════════════════════════════════
    # Generic code generator (unknown manufacturers)
    # ══════════════════════════════════════════════════════════════════

    def _generate_generic_py(self):
        """Build a minimal RTU-compatible Python source for unknown manufacturers.

        Uses parsed addresses directly without remapping.
        """
        from datetime import datetime

        manufacturer = self.cg_manufacturer_var.get().strip()
        classname = self.cg_classname_var.get().strip()
        include_iv = self.cg_ivscan_var.get()
        include_deravm = self.cg_deravm_var.get()

        lines = []
        emitted_addrs = set()
        seen_names = set()

        def _addr_int(hex_str):
            """'0x1001' → 4097"""
            try:
                return int(hex_str, 16) if isinstance(hex_str, str) else int(hex_str)
            except (ValueError, TypeError):
                return None

        def _emit_reg(addr_hex, name, comment):
            addr = _addr_int(addr_hex)
            if addr is None:
                return
            if addr in emitted_addrs:
                return
            emitted_addrs.add(addr)
            base = name
            idx = 2
            while name in seen_names:
                name = f"{base}_{idx}"
                idx += 1
            seen_names.add(name)
            if comment:
                lines.append(f"    {name} = {addr_hex}      # {comment}")
            else:
                lines.append(f"    {name} = {addr_hex}")

        # Check if any register uses Float32
        has_float32 = any(
            m.get('mb_type', '') == 'Float32'
            for section_key in ('periodic_basic', 'periodic_mppt', 'periodic_string')
            for m in self.mapping.get(section_key, [])
        )

        # ── Header ──
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        lines.append('"""')
        lines.append(f'{manufacturer} Inverter Modbus Register Map')
        lines.append(f'Auto-generated by UDP RTU Model Maker')
        lines.append(f'Date: {now_str}')
        lines.append('"""')
        if has_float32:
            lines.append('import struct')
        lines.append('')
        lines.append('')

        # ── RegisterMap class ──
        lines.append(f'class {classname}:')
        lines.append(f'    """{manufacturer} Modbus Register Map"""')
        lines.append('')

        # ── Device Information (from parsed_data) ──
        dev_info = (self.parsed_data or {}).get('device_info', [])
        if dev_info:
            lines.append('    # =========================================================================')
            lines.append('    # Device Information')
            lines.append('    # =========================================================================')
            for reg in dev_info:
                addr_hex = reg.get('address_hex', '')
                if not addr_hex or addr_hex == '-':
                    continue
                definition = reg.get('definition', '')
                name = self._field_to_const_name(definition)
                if not name:
                    continue
                mb_type = reg.get('type', '')
                mb_unit = reg.get('unit', '')
                mb_regs = reg.get('regs', '')
                comment_parts = []
                if mb_type:
                    comment_parts.append(mb_type)
                if mb_regs and str(mb_regs) != '1':
                    comment_parts.append(f'{mb_regs} regs')
                if mb_unit:
                    comment_parts.append(mb_unit)
                comment = ', '.join(comment_parts)
                _emit_reg(addr_hex, name, comment)
            lines.append('')

        # Emit all mapped registers
        section_keys = [
            ('periodic_basic', 'Periodic Basic Data'),
            ('periodic_mppt', 'MPPT Data'),
            ('periodic_string', 'STRING Data'),
        ]
        if include_deravm:
            # DER-AVM: JSON에 control_values가 비어있어도 고정 Solarize 주소 자동 추가
            cv = self.mapping.get('control_values', [])
            cm = self.mapping.get('control_monitor', [])
            if not cv:
                self.mapping['control_values'] = [
                    {'udp_field': 'DER_POWER_FACTOR_SET',    'mb_addr_hex': '0x07D0', 'mb_regs': 1, 'mb_type': 'S16', 'mb_unit': '0.001', 'mb_rw': 'RW'},
                    {'udp_field': 'DER_ACTION_MODE',         'mb_addr_hex': '0x07D1', 'mb_regs': 1, 'mb_type': 'U16', 'mb_unit': '',      'mb_rw': 'RW'},
                    {'udp_field': 'DER_REACTIVE_POWER_PCT',  'mb_addr_hex': '0x07D2', 'mb_regs': 1, 'mb_type': 'S16', 'mb_unit': '0.1%',  'mb_rw': 'RW'},
                    {'udp_field': 'DER_ACTIVE_POWER_PCT',    'mb_addr_hex': '0x07D3', 'mb_regs': 1, 'mb_type': 'U16', 'mb_unit': '0.1%',  'mb_rw': 'RW'},
                    {'udp_field': 'INVERTER_ON_OFF',         'mb_addr_hex': '0x0834', 'mb_regs': 1, 'mb_type': 'U16', 'mb_unit': '',      'mb_rw': 'RW'},
                ]
            if not cm:
                self.mapping['control_monitor'] = [
                    {'udp_field': 'DER_POWER_FACTOR_SET',    'mb_addr_hex': '0x07D0', 'mb_regs': 1, 'mb_type': 'S16', 'mb_unit': '0.001', 'mb_rw': 'R'},
                    {'udp_field': 'DER_ACTION_MODE',         'mb_addr_hex': '0x07D1', 'mb_regs': 1, 'mb_type': 'U16', 'mb_unit': '',      'mb_rw': 'R'},
                    {'udp_field': 'DER_REACTIVE_POWER_PCT',  'mb_addr_hex': '0x07D2', 'mb_regs': 1, 'mb_type': 'S16', 'mb_unit': '0.1%',  'mb_rw': 'R'},
                    {'udp_field': 'DER_ACTIVE_POWER_PCT',    'mb_addr_hex': '0x07D3', 'mb_regs': 1, 'mb_type': 'U16', 'mb_unit': '0.1%',  'mb_rw': 'R'},
                    {'udp_field': 'INVERTER_ON_OFF',         'mb_addr_hex': '0x0834', 'mb_regs': 1, 'mb_type': 'U16', 'mb_unit': '',      'mb_rw': 'R'},
                ]
            section_keys.append(('control_values', 'DER-AVM Control Values'))
            # control_monitor has the same addresses as control_values (read-only view),
            # so skip it to avoid duplicate _2 suffix aliases.
            # section_keys.append(('control_monitor', 'DER-AVM Control Monitor'))
        if include_iv:
            section_keys.append(('iv_scan', 'IV Scan Data'))

        for section_key, section_name in section_keys:
            lines.append('    # =========================================================================')
            lines.append(f'    # {section_name}')
            lines.append('    # =========================================================================')
            if section_key == 'periodic_basic':
                lines.append('    # PV_VOLTAGE: calculated by handler (MPPT voltage average, >= 100V)')
                lines.append('    # PV_CURRENT: calculated by handler (MPPT current sum)')
            has_any = False
            for m in self.mapping.get(section_key, []):
                addr_hex = m.get('mb_addr_hex', '-')
                if addr_hex == '-':
                    continue
                udp_field = m.get('udp_field', '')
                definition = m.get('mb_definition', '')
                raw_name = udp_field if udp_field else definition
                name = self._field_to_const_name(raw_name)
                # Skip PV_VOLTAGE/PV_CURRENT - calculated by handler
                if name in ('PV_VOLTAGE', 'PV_CURRENT') and section_key == 'periodic_basic':
                    has_any = True
                    continue
                if not name:
                    continue
                # 이름 중복 방지
                base = name
                idx = 2
                while name in seen_names:
                    name = f"{base}_{idx}"
                    idx += 1
                mb_type = m.get('mb_type', '')
                mb_unit = m.get('mb_unit', '')
                comment_parts = []
                if mb_type:
                    comment_parts.append(mb_type)
                if mb_unit:
                    comment_parts.append(mb_unit)
                comment = ', '.join(comment_parts)
                addr = _addr_int(addr_hex)
                # Skip address 0 (fault code bit definitions, not real Modbus addresses)
                if addr == 0:
                    continue
                if addr is not None and addr in emitted_addrs:
                    # 주소 중복 → alias로 추가
                    existing_name = None
                    for line in lines:
                        if f'= {addr_hex}' in line and '#' in line:
                            existing_name = line.strip().split('=')[0].strip()
                            break
                    if existing_name and name != existing_name:
                        lines.append(f"    {name} = {existing_name}      # Alias")
                        seen_names.add(name)
                        has_any = True
                else:
                    _emit_reg(addr_hex, name, comment)
                    has_any = True
            if not has_any:
                lines.append(f'    # (no {section_name.lower()} registers mapped)')
            lines.append('')

        # DEA-AVM Real-time Monitoring registers (fixed addresses for all Solarize-protocol inverters)
        if include_deravm:
            lines.append('    # =========================================================================')
            lines.append('    # DEA-AVM Real-time Monitoring Data (0x03E8-0x03FD) - For H05 Body Type 14')
            lines.append('    # =========================================================================')
            dea_regs = [
                ('0x03E8', 'DEA_L1_CURRENT_LOW',            'S32 low, scale 0.1A'),
                ('0x03E9', 'DEA_L1_CURRENT_HIGH',           'S32 high'),
                ('0x03EA', 'DEA_L2_CURRENT_LOW',            'S32 low, scale 0.1A'),
                ('0x03EB', 'DEA_L2_CURRENT_HIGH',           'S32 high'),
                ('0x03EC', 'DEA_L3_CURRENT_LOW',            'S32 low, scale 0.1A'),
                ('0x03ED', 'DEA_L3_CURRENT_HIGH',           'S32 high'),
                ('0x03EE', 'DEA_L1_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                ('0x03EF', 'DEA_L1_VOLTAGE_HIGH',           'S32 high'),
                ('0x03F0', 'DEA_L2_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                ('0x03F1', 'DEA_L2_VOLTAGE_HIGH',           'S32 high'),
                ('0x03F2', 'DEA_L3_VOLTAGE_LOW',            'S32 low, scale 0.1V'),
                ('0x03F3', 'DEA_L3_VOLTAGE_HIGH',           'S32 high'),
                ('0x03F4', 'DEA_TOTAL_ACTIVE_POWER_LOW',    'S32 low, scale 0.1kW'),
                ('0x03F5', 'DEA_TOTAL_ACTIVE_POWER_HIGH',   'S32 high'),
                ('0x03F6', 'DEA_TOTAL_REACTIVE_POWER_LOW',  'S32 low, scale 1 Var'),
                ('0x03F7', 'DEA_TOTAL_REACTIVE_POWER_HIGH', 'S32 high'),
                ('0x03F8', 'DEA_POWER_FACTOR_LOW',          'S32 low, scale 0.001'),
                ('0x03F9', 'DEA_POWER_FACTOR_HIGH',         'S32 high'),
                ('0x03FA', 'DEA_FREQUENCY_LOW',             'S32 low, scale 0.1Hz'),
                ('0x03FB', 'DEA_FREQUENCY_HIGH',            'S32 high'),
                ('0x03FC', 'DEA_STATUS_FLAG_LOW',           'U32 low, bit field'),
                ('0x03FD', 'DEA_STATUS_FLAG_HIGH',          'U32 high'),
            ]
            for addr_hex, name, comment in dea_regs:
                _emit_reg(addr_hex, name, comment)
            lines.append('')

        # ── Compatibility aliases for DER-AVM control registers ──
        if include_deravm:
            cv = self.mapping.get('control_values', [])
            der_names = [m.get('udp_field', '') for m in cv]
            if 'DER_POWER_FACTOR_SET' in der_names:
                lines.append('    # --- Compatibility aliases ---')
                lines.append('    POWER_FACTOR_SET = DER_POWER_FACTOR_SET')
                lines.append('    OPERATION_MODE = DER_ACTION_MODE')
                lines.append('    REACTIVE_POWER_PCT = DER_REACTIVE_POWER_PCT')
                lines.append('    REACTIVE_POWER_SET = DER_REACTIVE_POWER_PCT')
                lines.append('    ACTIVE_POWER_PCT = DER_ACTIVE_POWER_PCT')
                lines.append('')

        # ── Compatibility aliases (auto-generated for validation) ──
        compat_aliases = []
        phase_map = [
            ('L1_VOLTAGE', ['R_PHASE_VOLTAGE', 'R_VOLTAGE', 'PHASE_A_VOLTAGE',
                            'GRID_R_VOLTAGE', 'U_A', 'PHASE_1_VOLTAGE',
                            'S_PHASE_VOLTAGE', 'T_PHASE_VOLTAGE']),
            ('L2_VOLTAGE', ['S_PHASE_VOLTAGE', 'S_VOLTAGE', 'PHASE_B_VOLTAGE',
                            'GRID_S_VOLTAGE', 'U_B', 'PHASE_2_VOLTAGE',
                            'T_PHASE_VOLTAGE']),
            ('L3_VOLTAGE', ['T_PHASE_VOLTAGE', 'T_VOLTAGE', 'PHASE_C_VOLTAGE',
                            'GRID_T_VOLTAGE', 'U_C', 'PHASE_3_VOLTAGE']),
            ('L1_CURRENT', ['R_PHASE_CURRENT', 'R_CURRENT', 'PHASE_A_CURRENT',
                            'INV_R_CURRENT', 'I_A', 'PHASE_1_CURRENT',
                            'S_PHASE_CURRENT']),
            ('L2_CURRENT', ['S_PHASE_CURRENT', 'S_CURRENT', 'PHASE_B_CURRENT',
                            'INV_S_CURRENT', 'I_B', 'PHASE_2_CURRENT']),
            ('L3_CURRENT', ['T_PHASE_CURRENT', 'T_CURRENT', 'PHASE_C_CURRENT',
                            'INV_T_CURRENT', 'I_C', 'PHASE_3_CURRENT']),
        ]
        for alias, sources in phase_map:
            if alias not in seen_names:
                for src in sources:
                    if src in seen_names:
                        compat_aliases.append(f'    {alias} = {src}')
                        seen_names.add(alias)
                        break

        energy_map = [
            ('TOTAL_ENERGY_LOW', ['TOTAL_ENERGY', 'ACCUMULATED_ENERGY', 'CUMULATIVE_PRODUCTION_L',
                                   'TOTAL_ENERGY_L', 'TOTAL_ENERGY_KWH', 'TOTAL_YIELD_L',
                                   'TOTAL_GENERATED_ENERGY', 'CUMULATIVE_ENERGY',
                                   'HIGH_BYTE_OF_TOTAL_FEED_POWER_TO_GRID_TOTAL_POWER_GENERATION']),
            ('TODAY_ENERGY_LOW', ['TODAY_ENERGY', 'TODAY_ENERGY_L', 'DAILY_PRODUCTION',
                                   'DAILY_ENERGY', 'DAY_ENERGY', 'TODAY_GENERATION',
                                   'TODAY_GENERATED_ENERGY', 'DAILY_GENERATED_ENERGY',
                                   'E_DAY_DAILY_POWER_GENERATION']),
        ]
        for alias, sources in energy_map:
            if alias not in seen_names:
                for src in sources:
                    if src in seen_names:
                        compat_aliases.append(f'    {alias} = {src}')
                        seen_names.add(alias)
                        break

        mppt_alias_map = [
            ('MPPT1_VOLTAGE', ['PV1_VOLTAGE', 'PV_VOLTAGE', 'MPPT_1_VOLTAGE',
                                'VPV1_PV1_VOLTAGE', 'VPV1_MPPT1_PV1_VOLTAGE']),
            ('MPPT1_CURRENT', ['PV1_CURRENT', 'PV_CURRENT', 'MPPT_1_CURRENT',
                                'IPV1_PV1_CURRENT', 'IPV1_MPPT1_PV1_CURRENT']),
        ]
        for alias, sources in mppt_alias_map:
            if alias not in seen_names:
                for src in sources:
                    if src in seen_names:
                        compat_aliases.append(f'    {alias} = {src}')
                        seen_names.add(alias)
                        break

        if 'INVERTER_MODE' not in seen_names:
            for src in ['RUNNING_STATUS', 'DEVICE_STATUS', 'INVERTER_STATUS',
                        'SYSTEM_STATUS', 'WORK_MODE', 'OPERATING_MODE_STATUS',
                        'STATUS_1', 'WORKING_MODE']:
                if src in seen_names:
                    compat_aliases.append(f'    INVERTER_MODE = {src}')
                    seen_names.add('INVERTER_MODE')
                    break

        if 'ERROR_CODE1' not in seen_names:
            for src in ['FAULT_CODE_1', 'FAULT_CODE1', 'ERROR_MESSAGE_H',
                        'ALARM_CODE_1', 'FAULT_CODE', 'ERROR_CODE', 'STATUS_2']:
                if src in seen_names:
                    compat_aliases.append(f'    ERROR_CODE1 = {src}')
                    seen_names.add('ERROR_CODE1')
                    break

        if 'ERROR_CODE2' not in seen_names:
            for src in ['FAULT_CODE_2', 'FAULT_CODE2', 'ERROR_MESSAGE_L',
                        'ALARM_CODE_2', 'ERROR_2', 'STATUS_3']:
                if src in seen_names:
                    compat_aliases.append(f'    ERROR_CODE2 = {src}')
                    seen_names.add('ERROR_CODE2')
                    break

        if compat_aliases:
            lines.append('    # --- Compatibility aliases (auto-generated) ---')
            lines.extend(compat_aliases)
            lines.append('')

        lines.append('')
        lines.append('')

        # ── InverterMode class ──
        self._emit_inverter_mode_class(lines)

        # ── StatusConverter class (from status_map in JSON) ──
        status_map = getattr(self, '_loaded_status_map', None)
        if status_map and 'values' in status_map:
            self._emit_status_converter_class(lines, manufacturer, status_map)
            safe_name = self._sanitize_class_name(manufacturer) if manufacturer else 'Generic'
            sc_cls_name = f'{safe_name}StatusConverter'
            lines.append(f'StatusConverter = {sc_cls_name}')
            lines.append('')
            lines.append('')
        else:
            # Default StatusConverter when no status_map is available
            safe_name = self._sanitize_class_name(manufacturer) if manufacturer else 'Generic'
            sc_cls_name = f'{safe_name}StatusConverter'
            lines.append(f'class {sc_cls_name}:')
            lines.append(f'    """Default status converter for {manufacturer}."""')
            lines.append('')
            lines.append('    @classmethod')
            lines.append('    def to_inverter_mode(cls, raw):')
            lines.append('        """Map raw status to InverterMode (direct or STANDBY fallback)."""')
            lines.append('        valid = {0x00, 0x01, 0x03, 0x04, 0x05, 0x09}')
            lines.append('        return raw if raw in valid else 0x01')
            lines.append('')
            lines.append('')
            lines.append(f'StatusConverter = {sc_cls_name}')
            lines.append('')
            lines.append('')

        # ── FaultCode class (from error_codes in JSON) ──
        error_codes = getattr(self, '_loaded_error_codes', None)
        if error_codes and 'values' in error_codes:
            self._emit_fault_code_class(lines, manufacturer, error_codes)

        # ── RegisterMap alias (if classname != RegisterMap) ──
        if classname != 'RegisterMap':
            lines.append(f'RegisterMap = {classname}')
            lines.append('')
            lines.append('')

        # ── SCALE dict ──
        lines.append('# Scale factors')
        lines.append('SCALE = {')
        # Required 5 keys with defaults
        generic_scale = {
            'voltage': 0.1,
            'current': 0.01,
            'power': 1.0,
            'frequency': 0.01,
            'power_factor': 0.001,
        }
        # Try to detect scales from mapping data
        for m in self.mapping.get('periodic_basic', []):
            mb_unit = m.get('mb_unit', '')
            mb_type = m.get('mb_type', '')
            scale = self._detect_scale(mb_unit, mb_type)
            # Could enrich generic_scale here if desired
        for key, val in generic_scale.items():
            lines.append(f"    '{key}': {val!r},")
        lines.append('}')
        lines.append('')
        lines.append('')

        # ── Helper functions ──
        lines.append('def registers_to_u32(low, high):')
        lines.append('    """Combine two U16 to U32"""')
        lines.append('    return ((high & 0xFFFF) << 16) | (low & 0xFFFF)')
        lines.append('')
        lines.append('')
        lines.append('def registers_to_s32(low, high):')
        lines.append('    """Combine two U16 to S32"""')
        lines.append('    value = ((high & 0xFFFF) << 16) | (low & 0xFFFF)')
        lines.append('    if value >= 0x80000000:')
        lines.append('        value -= 0x100000000')
        lines.append('    return value')
        lines.append('')
        lines.append('')

        if has_float32:
            lines.append('def registers_to_float32(high, low):')
            lines.append('    """Convert two U16 registers to IEEE 754 Float32."""')
            lines.append('    import struct')
            lines.append("    raw = struct.pack('>HH', high & 0xFFFF, low & 0xFFFF)")
            lines.append("    val = struct.unpack('>f', raw)[0]")
            lines.append('    if val != val or abs(val) > 1e10:  # NaN or unreasonable')
            lines.append('        return 0.0')
            lines.append('    return val')
            lines.append('')
            lines.append('')

        # ── DATA_TYPES dict ──
        self._emit_data_types_dict(lines)

        return '\n'.join(lines)

    # ══════════════════════════════════════════════════════════════════
    # Shared helper for code generation
    # ══════════════════════════════════════════════════════════════════

    def _emit_inverter_mode_class(self, lines):
        """Emit the standard InverterMode class (shared across all manufacturers)."""
        lines.append('class InverterMode:')
        lines.append('    """Solarize compatible InverterMode"""')
        lines.append('    INITIAL = 0x00')
        lines.append('    STANDBY = 0x01')
        lines.append('    ON_GRID = 0x03')
        lines.append('    OFF_GRID = 0x04')
        lines.append('    FAULT = 0x05')
        lines.append('    SHUTDOWN = 0x09')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, mode):')
        lines.append('        mode_map = {')
        lines.append('            0x00: "Initial", 0x01: "Standby", 0x03: "On-Grid",')
        lines.append('            0x04: "Off-Grid", 0x05: "Fault", 0x09: "Shutdown"')
        lines.append('        }')
        lines.append('        return mode_map.get(mode, f"Unknown({mode})")')
        lines.append('')
        lines.append('')

    def _emit_data_types_dict(self, lines):
        """Emit DATA_TYPES and FLOAT32_FIELDS dictionaries for register type info."""
        type_map = {
            'Float32': 'float32',
            'U32': 'u32',
            'S32': 's32',
            'S16': 's16',
            'U16': 'u16',
        }
        entries = []
        has_float32 = False
        for section_key in ('periodic_basic', 'periodic_mppt', 'periodic_string'):
            for m in self.mapping.get(section_key, []):
                addr_hex = m.get('mb_addr_hex', '')
                udp_field = m.get('udp_field', '')
                if not addr_hex or addr_hex == '-' or not udp_field:
                    continue
                const_name = self._field_to_const_name(udp_field)
                if not const_name:
                    continue
                mb_type = m.get('mb_type', 'U16')
                py_type = type_map.get(mb_type, 'u16')
                if py_type == 'float32':
                    has_float32 = True
                entries.append((const_name, py_type))

        # Ensure T03/T07 required standard fields are always present
        _required_dt = [
            ('R_PHASE_VOLTAGE', 'u16'), ('S_PHASE_VOLTAGE', 'u16'), ('T_PHASE_VOLTAGE', 'u16'),
            ('R_PHASE_CURRENT', 'u16'), ('S_PHASE_CURRENT', 'u16'), ('T_PHASE_CURRENT', 'u16'),
            ('FREQUENCY', 'u16'), ('MPPT1_VOLTAGE', 'u16'), ('MPPT1_CURRENT', 'u16'),
            ('AC_POWER', 's32'), ('PV_POWER', 'u32'),
            ('POWER_FACTOR', 's16'), ('INVERTER_MODE', 'u16'),
            ('ERROR_CODE1', 'u16'), ('ERROR_CODE2', 'u16'),
            ('TOTAL_ENERGY', 'u32'),
        ]
        existing_names = {name for name, _ in entries}
        for name, dtype in _required_dt:
            if name not in existing_names:
                entries.append((name, dtype))

        lines.append('# Data type info per register')
        lines.append('DATA_TYPES = {')
        for const_name, py_type in entries:
            lines.append(f"    '{const_name}': '{py_type}',")
        lines.append('}')
        lines.append('')
        if has_float32:
            lines.append("FLOAT32_FIELDS = {k for k, v in DATA_TYPES.items() if v == 'float32'}")
        else:
            lines.append('FLOAT32_FIELDS = set()')
        lines.append('')
        lines.append('')

    def _emit_status_converter_class(self, lines, manufacturer, status_map):
        """Emit a StatusConverter class from JSON status_map."""
        safe_name = self._sanitize_class_name(manufacturer)
        cls_name = f'{safe_name}StatusConverter'
        inv_mode_map = {
            'INITIAL': 'InverterMode.INITIAL',
            'STANDBY': 'InverterMode.STANDBY',
            'ON_GRID': 'InverterMode.ON_GRID',
            'OFF_GRID': 'InverterMode.OFF_GRID',
            'FAULT': 'InverterMode.FAULT',
            'SHUTDOWN': 'InverterMode.SHUTDOWN',
        }

        lines.append(f'class {cls_name}:')
        desc = status_map.get('description', 'Status converter')
        lines.append(f'    """{desc}"""')
        lines.append('')

        # Status constants
        for hex_val, info in status_map['values'].items():
            if isinstance(info, dict):
                name = info.get('name', info.get('description', f'STATUS_{hex_val}'))
            else:
                name = str(info)
            # Sanitize name for Python constant
            safe_status = self._to_upper_snake(str(name)) if name else f'STATUS_{hex_val}'
            if not safe_status:
                safe_status = f'STATUS_{hex_val}'
            desc = info.get('description', name) if isinstance(info, dict) else str(info)
            lines.append(f'    STATUS_{safe_status} = {hex_val}      # {desc}')
        lines.append('')

        # to_string method
        lines.append('    _STATUS_MAP = {')
        for hex_val, info in status_map['values'].items():
            if isinstance(info, dict):
                display = info.get('description', info.get('name', f'Status {hex_val}'))
            else:
                display = str(info)
            lines.append(f'        {hex_val}: "{display}",')
        lines.append('    }')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, status):')
        lines.append('        return cls._STATUS_MAP.get(status, f"Unknown(0x{status:04X})")')
        lines.append('')

        # to_inverter_mode method
        lines.append('    _CONVERSION_MAP = {')
        for hex_val, info in status_map['values'].items():
            if isinstance(info, dict):
                mode_key = info.get('inverter_mode', 'STANDBY')
                desc = info.get('description', info.get('name', ''))
            else:
                mode_key = 'STANDBY'
                desc = str(info)
            mode_ref = inv_mode_map.get(mode_key, 'InverterMode.STANDBY')
            lines.append(f'        {hex_val}: {mode_ref},      # {desc}')
        lines.append('    }')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def to_inverter_mode(cls, raw_status):')
        lines.append(f'        """Convert {safe_name} status to InverterMode code"""')
        lines.append('        return cls._CONVERSION_MAP.get(raw_status, InverterMode.STANDBY)')
        lines.append('')
        lines.append('')

    def _emit_fault_code_class(self, lines, manufacturer, error_codes):
        """Emit a FaultCode class from JSON error_codes.

        Supports two formats:
        1. Flat format: values = {"0x0001": {"description": ..., "type": ...}, ...}
        2. Multi-table bit-field format (Solarize-style):
           values = {"table1_0xADDR": {"bit0": {...}, "bit1": {...}}, ...}
           Generates per-table ErrorCode1/2/3 classes + wrapper FaultCode class.
        """
        safe_name = self._sanitize_class_name(manufacturer)
        cls_name = f'{safe_name}FaultCode'
        desc = error_codes.get('description', 'Fault/Alarm codes')

        # Detect multi-table bit-field format
        values = error_codes['values']
        is_multi_table = any(k.startswith('table') for k in values.keys())

        if is_multi_table:
            self._emit_multi_table_fault_classes(lines, safe_name, cls_name, desc, error_codes)
        else:
            self._emit_flat_fault_code_class(lines, cls_name, desc, values)

    def _emit_flat_fault_code_class(self, lines, cls_name, desc, values):
        """Emit flat FaultCode class (original format)."""
        lines.append(f'class {cls_name}:')
        lines.append(f'    """{desc}"""')
        lines.append('')

        # Fault/Alarm code map
        lines.append('    _CODE_MAP = {')
        for hex_val, info in values.items():
            code_type = info.get('type', 'Fault')
            desc_str = info.get('description', 'Unknown')
            lines.append(f'        {hex_val}: ("{desc_str}", "{code_type}"),')
        lines.append('    }')
        lines.append('')

        # to_string method
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, code):')
        lines.append('        entry = cls._CODE_MAP.get(code)')
        lines.append('        if entry:')
        lines.append('            return f"{entry[0]} ({entry[1]})"')
        lines.append('        return f"Unknown(0x{code:04X})"')
        lines.append('')

        # is_fault / is_alarm
        lines.append('    @classmethod')
        lines.append('    def is_fault(cls, code):')
        lines.append('        entry = cls._CODE_MAP.get(code)')
        lines.append('        return entry[1] == "Fault" if entry else False')
        lines.append('')
        lines.append('    @classmethod')
        lines.append('    def is_alarm(cls, code):')
        lines.append('        entry = cls._CODE_MAP.get(code)')
        lines.append('        return entry[1] == "Alarm" if entry else False')
        lines.append('')
        lines.append('')

    def _emit_multi_table_fault_classes(self, lines, safe_name, wrapper_cls_name, desc, error_codes):
        """Emit per-table ErrorCode classes + wrapper FaultCode class for multi-table bit-field format."""
        values = error_codes['values']

        # Categorize tables by description keywords
        _TABLE_CATEGORY = {
            1: 'Hardware Faults',
            2: 'Grid/PV Faults',
            3: 'System Warnings',
        }

        # Sort table keys to ensure consistent ordering (table1, table2, table3, ...)
        table_keys = sorted(
            [k for k in values.keys() if k.startswith('table')],
            key=lambda k: int(re.match(r'table(\d+)', k).group(1))
        )

        table_class_names = []
        has_reserved_bits = {}  # track which tables have reserved bits

        for table_key in table_keys:
            # Extract table number and register address
            m = re.match(r'table(\d+)(?:_0x([0-9A-Fa-f]+))?', table_key)
            if not m:
                continue
            table_num = int(m.group(1))
            reg_addr = m.group(2) or ''
            addr_str = f'0x{reg_addr}' if reg_addr else ''

            class_name = f'ErrorCode{table_num}'
            table_class_names.append(class_name)
            bits_data = values[table_key]

            # Check if any bits are Reserved
            table_has_reserved = any(
                info.get('type', '') == 'Reserved'
                for info in bits_data.values()
                if isinstance(info, dict)
            )
            has_reserved_bits[table_num] = table_has_reserved

            # Determine category
            category = _TABLE_CATEGORY.get(table_num, f'Table {table_num}')

            doc_parts = [f'Error Code Table {table_num}']
            if addr_str:
                doc_parts.append(f'({addr_str})')
            doc_parts.append(f'- {category}, Bit field')

            lines.append(f'class {class_name}:')
            lines.append(f'    """{" ".join(doc_parts)}"""')

            # BITS dict
            lines.append('    BITS = {')
            for bit_key in sorted(bits_data.keys(), key=lambda k: int(k.replace('bit', ''))):
                bit_num = int(bit_key.replace('bit', ''))
                info = bits_data[bit_key]
                bit_desc = info.get('description', 'Unknown')
                bit_type = info.get('type', 'Fault')
                lines.append(f'        {bit_num}: ("{bit_desc}", "{bit_type}"),')
            lines.append('    }')
            lines.append('')

            # decode method
            lines.append('    @classmethod')
            lines.append('    def decode(cls, value):')
            lines.append('        """Decode error code to list of error descriptions"""')
            lines.append('        errors = []')
            lines.append('        for bit, (desc, severity) in cls.BITS.items():')
            lines.append('            if value & (1 << bit):')
            if table_has_reserved:
                lines.append('                if severity != "Reserved":')
                lines.append(f'                    errors.append(f"E{table_num}.{{bit}}:{{desc}} ({{severity}})")')
            else:
                lines.append(f'                errors.append(f"E{table_num}.{{bit}}:{{desc}} ({{severity}})")')
            lines.append('        return errors')
            lines.append('')

            # to_string method
            lines.append('    @classmethod')
            lines.append('    def to_string(cls, value):')
            lines.append('        """Get error string or \'OK\' if no errors"""')
            lines.append('        if value == 0:')
            lines.append('            return "OK"')
            lines.append('        errors = cls.decode(value)')
            if table_has_reserved:
                lines.append('        return ", ".join(errors) if errors else "OK"')
            else:
                lines.append('        return ", ".join(errors)')
            lines.append('')
            lines.append('')

        # Wrapper class
        num_tables = len(table_class_names)
        lines.append(f'class {wrapper_cls_name}:')

        # Build docstring
        lines.append(f'    """{desc}')
        lines.append('')
        for i, tc in enumerate(table_class_names, 1):
            table_key = table_keys[i - 1]
            m = re.match(r'table(\d+)(?:_0x([0-9A-Fa-f]+))?', table_key)
            addr_str = f'0x{m.group(2)}' if m and m.group(2) else ''
            category = _TABLE_CATEGORY.get(i, f'Table {i}')
            addr_part = f' ({addr_str})' if addr_str else ''
            lines.append(f'    Error Code {i}{addr_part}: {category} - 16 bits')
        lines.append(f'    Total: {num_tables * 16} fault codes')
        lines.append('    """')
        lines.append('')

        # Build parameter strings for variable number of tables
        params_required = ', '.join(f'error{i}' for i in range(1, num_tables + 1))
        params_optional = 'error1' + ''.join(f', error{i}=0' for i in range(2, num_tables + 1))

        # decode_all classmethod
        lines.append('    @classmethod')
        lines.append(f'    def decode_all(cls, {params_required}):')
        lines.append(f'        """Decode all {num_tables} error code registers at once.')
        lines.append('')
        lines.append('        Args:')
        for i, table_key in enumerate(table_keys, 1):
            m = re.match(r'table(\d+)(?:_0x([0-9A-Fa-f]+))?', table_key)
            addr_str = f'0x{m.group(2)}' if m and m.group(2) else ''
            addr_part = f' ({addr_str})' if addr_str else ''
            lines.append(f'            error{i}: Error Code {i} register value{addr_part}')
        lines.append('')
        lines.append('        Returns:')
        lines.append('            List of error description strings')
        lines.append('        """')
        lines.append('        errors = []')
        for i, tc in enumerate(table_class_names, 1):
            lines.append(f'        errors.extend({tc}.decode(error{i}))')
        lines.append('        return errors')
        lines.append('')

        # to_string classmethod
        lines.append('    @classmethod')
        lines.append('    def to_string(cls, code, table=1):')
        lines.append('        """Get error string for a single error code register.')
        lines.append('')
        lines.append('        Args:')
        lines.append('            code: Error code register value')
        lines.append(f'            table: Which table (1-{num_tables})')
        lines.append('        """')
        table_map_entries = ', '.join(f'{i}: {tc}' for i, tc in enumerate(table_class_names, 1))
        lines.append(f'        table_map = {{{table_map_entries}}}')
        lines.append('        handler = table_map.get(table)')
        lines.append('        if handler:')
        lines.append('            return handler.to_string(code)')
        lines.append(f'        return f"Unknown table {{table}}"')
        lines.append('')

        # has_fault classmethod
        lines.append('    @classmethod')
        lines.append(f'    def has_fault(cls, {params_optional}):')
        lines.append('        """Check if any fault (not warning) is active across all tables."""')
        for i, tc in enumerate(table_class_names, 1):
            lines.append(f'        for bit, (desc, severity) in {tc}.BITS.items():')
            lines.append(f'            if error{i} & (1 << bit) and severity == "Fault":')
            lines.append('                return True')
        lines.append('        return False')
        lines.append('')
        lines.append('')

    @staticmethod
    def _detect_scale(unit_str, type_str):
        """Try to infer a numeric scale factor from the unit or type string.

        Returns the multiplier to convert raw register value to physical value.
        e.g. unit "0.1V" → scale 0.1  (raw × 0.1 = volts)
             unit "0.01A" → scale 0.01
        """
        if not unit_str:
            return None
        u = unit_str.lower().strip()

        # Pattern: numeric prefix + unit  (e.g. "0.1V", "0.01A", "0.1kWh")
        # Note: kwh/kvarh must come before kw/kvar to avoid partial match
        m = re.match(
            r'^([0-9]*\.?[0-9]+)\s*'
            r'(kwh|kvarh|kvar|kva|kw|wh|var|va|v|a|w|hz|%)',
            u
        )
        if m:
            val = float(m.group(1))
            if val != 0 and val != 1:
                return val

        # x10, x100, ×0.1 patterns (multiplier)
        m2 = re.match(r'^[x×]\s*([0-9]*\.?[0-9]+)', u)
        if m2:
            val = float(m2.group(1))
            return val if val != 1 else None

        # /10, /100 patterns (divisor → scale = 1/N)
        m3 = re.match(r'^[/÷]\s*([0-9]+)', u)
        if m3:
            divisor = int(m3.group(1))
            return 1.0 / divisor if divisor > 1 else None

        # "Gain=100" or "gain=0.1" patterns
        m4 = re.search(r'gain\s*[=:]\s*([0-9]*\.?[0-9]+)', u)
        if m4:
            val = float(m4.group(1))
            if val != 0 and val != 1:
                return val

        return None


# ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import sys
    root = tk.Tk()
    app = ModbusToUdpMapper(root)
    root.protocol('WM_DELETE_WINDOW', lambda: (root.destroy(), sys.exit(0)))
    root.mainloop()
