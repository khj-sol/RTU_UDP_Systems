#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
All-Inverter Offline vs AI Comparison Script
5 inverters × 2 modes × 3 stages = full comparison
Inverters: Solarize, Kstar, Huawei, Sungrow, EKOS
"""
import sys, os, re, json, time, shutil, traceback

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from model_maker import stage_pipeline
from model_maker.modbus_to_udp_mapper import parse_modbus_pdf, ModbusToUdpMapper, _SOLARIZE_ADDR_TO_NAME

DOCS = os.path.join(ROOT, 'docs')
OUT_ROOT = os.path.join(ROOT, 'model_maker', 'comparison_output_all')
os.makedirs(OUT_ROOT, exist_ok=True)

API_KEY = (os.environ.get('ANTHROPIC_API_KEY', '') or
           os.environ.get('CLAUDE_CODE_OAUTH_TOKEN', ''))
AI_MODEL = 'claude-haiku-4-5-20251001'
print(f"API_KEY available: {bool(API_KEY)}")

SOLARIZE_ADDR_MAP = _SOLARIZE_ADDR_TO_NAME
SOLARIZE_SCALE = {'voltage': 0.1, 'current': 0.01, 'power': 0.1,
                  'frequency': 0.01, 'power_factor': 0.001}

# ─────────────────────────────────────────────────────────────────────
# Inverter configurations
# ─────────────────────────────────────────────────────────────────────
INVERTERS = [
    {
        'name': 'Solarize',
        'key': 'solarize',
        'pdf': os.path.join(DOCS, 'Solarize Modbus Protocol-Korea-V1.2.4.pdf'),
        'ref_py': os.path.join(ROOT, 'common', 'solarize_registers.py'),
        'settings': {
            'mppt_count': 4, 'string_count': 8,
            'iv_scan': True, 'der_avm': True, 'dea_avm': True,
            'manufacturer': 'Solarize VK50', 'class_name': 'RegisterMap',
            'fc_code': 'FC03',
        },
    },
    {
        'name': 'Kstar',
        'key': 'kstar',
        'pdf': os.path.join(DOCS, '1_KSG1.250K.Inverter.Modbus.Communication.Protocol.3.5.pdf'),
        'ref_py': os.path.join(ROOT, 'common', 'kstar_mm_registers.py'),
        'settings': {
            'mppt_count': 3, 'string_count': 6,
            'iv_scan': False, 'der_avm': True, 'dea_avm': True,
            'manufacturer': 'Kstar KSG-60KT', 'class_name': 'RegisterMap',
            'fc_code': 'FC04',
        },
    },
    {
        'name': 'Huawei',
        'key': 'huawei',
        'pdf': os.path.join(DOCS, 'SUN2000MC V200R023C00 Modbus Interface Definitions.pdf'),
        'ref_py': os.path.join(ROOT, 'common', 'huawei_mm_registers.py'),
        'settings': {
            'mppt_count': 4, 'string_count': 8,
            'iv_scan': False, 'der_avm': True, 'dea_avm': True,
            'manufacturer': 'Huawei SUN2000', 'class_name': 'RegisterMap',
            'fc_code': 'FC03',
        },
    },
    {
        'name': 'Sungrow',
        'key': 'sungrow',
        'pdf': os.path.join(DOCS, 'Communication Protocol of PV Grid-Connected String Inverters_V1.1.37_EN.pdf'),
        'ref_py': os.path.join(ROOT, 'common', 'sungrow_mm_registers.py'),
        'settings': {
            'mppt_count': 4, 'string_count': 8,
            'iv_scan': False, 'der_avm': True, 'dea_avm': True,
            'manufacturer': 'Sungrow SG', 'class_name': 'RegisterMap',
            'fc_code': 'FC03',
        },
    },
    {
        'name': 'EKOS',
        'key': 'ekos',
        'pdf': None,  # EKOS uses xlsx
        'xlsx': os.path.join(DOCS, 'ModBus map-EK_20220209_통신테스트용.xlsx'),
        'ref_py': os.path.join(ROOT, 'common', 'ekos_mm_registers.py'),
        'settings': {
            'mppt_count': 4, 'string_count': 8,
            'iv_scan': False, 'der_avm': True, 'dea_avm': True,
            'manufacturer': 'EKOS EK', 'class_name': 'RegisterMap',
            'fc_code': 'FC03',
        },
    },
]

# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────
def log_cb(logs, mode):
    def _cb(msg):
        logs.setdefault(mode, []).append(msg)
    return _cb


def count_mapped(wb_path):
    """Count Auto/Ref/AI mapped, manual, unmapped in Stage 2 Excel."""
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = wb['Register_Mapping']
        auto, manual, unmapped, ai_extra = 0, 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if isinstance(row[0], str) and row[0].startswith('---'):
                continue
            if not isinstance(row[0], (int, float)):
                continue
            mt = str(row[13] or '') if len(row) > 13 else ''
            if mt.startswith('AI('):
                ai_extra += 1
                auto += 1
            elif mt in ('Auto',) or mt.startswith('Ref('):
                auto += 1
            elif mt == 'Manual':
                manual += 1
            else:
                unmapped += 1
        der = sum(1 for r in wb['DER_AVM_Control'].iter_rows(min_row=2, values_only=True) if r[0] is not None)
        dea = sum(1 for r in wb['DEA_AVM_Monitor'].iter_rows(min_row=2, values_only=True) if r[0] is not None)
        return auto, manual, unmapped, ai_extra, der, dea
    except Exception as e:
        return 0, 0, 0, 0, 0, 0


def count_stage1(wb_path):
    """Count registers, status codes, error codes in Stage 1 Excel."""
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = wb['Registers']
        regs = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                   if isinstance(row[0], (int, float)))
        status = 0
        error = 0
        if 'Status_Definitions' in wb.sheetnames:
            status = sum(1 for row in wb['Status_Definitions'].iter_rows(min_row=2, values_only=True)
                         if row[0] is not None and not isinstance(row[0], str))
        if 'Error_Fault_Codes' in wb.sheetnames:
            error = sum(1 for row in wb['Error_Fault_Codes'].iter_rows(min_row=2, values_only=True)
                        if row[0] is not None)
        return regs, status, error, wb.sheetnames
    except Exception as e:
        return 0, 0, 0, []


def read_stage1_data(wb_path):
    """Read all register rows from Stage 1 excel for content comparison."""
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = wb['Registers']
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if isinstance(row[0], (int, float)):
                rows.append(row)
        return rows
    except:
        return []


def read_stage2_data(wb_path):
    """Read mapping rows from Stage 2 excel."""
    try:
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        ws = wb['Register_Mapping']
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and isinstance(row[0], (int, float)):
                rows.append(row)
        return rows
    except:
        return []


def run_12_tests(code, label=''):
    """12-item validation test suite."""
    if not code:
        return [('SKIP', 'No code')] * 12
    results = []

    # T1 Syntax
    try:
        compile(code, '<generated>', 'exec')
        results.append(('OK', 'Python syntax valid'))
    except SyntaxError as e:
        results.append(('FAIL', f'SyntaxError: {e}'))

    # T2 RegisterMap
    results.append(('OK', 'class RegisterMap exists') if re.search(r'\bclass\s+RegisterMap\b', code)
                   else ('FAIL', 'class RegisterMap missing'))

    # T3 Constants
    consts = re.findall(r'^\s+(\w+)\s*=\s*(0x[0-9A-Fa-f]+|\d+)', code, re.MULTILINE)
    results.append(('OK', f'{len(consts)} register constants') if consts
                   else ('FAIL', 'No register constants'))

    # T4 Essential aliases
    essential = ['L1_VOLTAGE', 'L1_CURRENT', 'L2_VOLTAGE', 'L3_VOLTAGE',
                 'MPPT1_VOLTAGE', 'MPPT1_CURRENT', 'INVERTER_MODE', 'ERROR_CODE1',
                 'TOTAL_ENERGY_LOW', 'TODAY_ENERGY_LOW', 'POWER_FACTOR',
                 'PV_TOTAL_INPUT_POWER_LOW', 'GRID_TOTAL_ACTIVE_POWER_LOW']
    missing = [a for a in essential if a not in code]
    results.append(('OK', f'All {len(essential)} essential aliases') if not missing
                   else ('FAIL', f'Missing: {missing[:4]}{"..." if len(missing)>4 else ""}'))

    # T5 Unique addresses
    addrs = [int(a, 16) for a in re.findall(r'=\s*(0x[0-9A-Fa-f]+)', code)]
    dups = len(addrs) - len(set(addrs))
    results.append(('OK', f'{len(addrs)} addrs, all unique') if dups == 0
                   else ('WARN', f'{dups} duplicate addresses'))

    # T6 SCALE dict
    scale_seg = code[code.find('SCALE'):code.find('SCALE')+500] if 'SCALE' in code else ''
    scale_keys = re.findall(r"'(\w+)':\s*[\d.]+", scale_seg)
    missing_scale = [k for k in ['voltage', 'current', 'power', 'frequency', 'power_factor'] if k not in scale_keys]
    results.append(('OK', f'SCALE: {len(scale_keys)} keys') if not missing_scale
                   else ('FAIL', f'SCALE missing: {missing_scale}'))

    # T7 InverterMode
    if re.search(r'\bclass\s+InverterMode\b', code):
        ok = all(m in code for m in ['INITIAL', 'STANDBY', 'ON_GRID', 'FAULT', 'SHUTDOWN'])
        results.append(('OK', 'InverterMode complete') if ok and 'to_string' in code
                       else ('WARN', 'InverterMode incomplete'))
    else:
        results.append(('FAIL', 'InverterMode missing'))

    # T8 Helpers
    missing_h = [h for h in ['registers_to_u32', 'registers_to_s32'] if h not in code]
    results.append(('OK', 'u32/s32 helpers present') if not missing_h
                   else ('FAIL', f'Missing helpers: {missing_h}'))

    # T9 get_mppt_registers
    results.append(('OK', 'get_mppt_registers()') if 'get_mppt_registers' in code
                   else ('FAIL', 'get_mppt_registers() missing'))

    # T10 get_string_registers
    results.append(('OK', 'get_string_registers()') if 'get_string_registers' in code
                   else ('FAIL', 'get_string_registers() missing'))

    # T11 DATA_TYPES
    results.append(('OK', 'DATA_TYPES present') if 'DATA_TYPES' in code
                   else ('FAIL', 'DATA_TYPES missing'))

    # T12 DER-AVM
    der_attrs = ['DER_POWER_FACTOR_SET', 'DER_ACTION_MODE',
                 'DER_REACTIVE_POWER_PCT', 'DER_ACTIVE_POWER_PCT', 'INVERTER_ON_OFF']
    missing_der = [a for a in der_attrs if a not in code]
    results.append(('OK', f'All {len(der_attrs)} DER-AVM attrs') if not missing_der
                   else ('FAIL', f'DER-AVM missing: {missing_der[:3]}'))

    return results


def extract_attrs(code):
    """Extract RegisterMap attribute name→address pairs."""
    attrs = {}
    in_class = False
    for line in code.split('\n'):
        s = line.strip()
        if re.match(r'^class\s+\w+.*:', s):
            in_class = 'RegisterMap' in s
            continue
        if not in_class:
            continue
        m = re.match(r'(\w+)\s*=\s*(0x[0-9A-Fa-f]+|\d+)', s)
        if m:
            v = m.group(2)
            attrs[m.group(1)] = int(v, 16) if v.startswith('0x') else int(v)
    return attrs


def parse_ekos_xlsx(xlsx_path):
    """Parse EKOS xlsx file into parsed_data format compatible with stage pipeline.

    v1.4.1: Handles string range addresses like '30035~30036' (Float32 registers)
    and multi-section sheets (FC05h control, FC04h realtime, FC10h write).
    All sections are extracted from sheet 0 (the single data sheet).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_regs = []
    reg_no = 1

    # EKOS xlsx has all data in sheet 0; sheets 1-3 contain format legends/fault maps
    # Section is inferred from the FC column header rows within sheet 0
    ws = wb.worksheets[0]

    _SKIP_UNITS = {'-', 'V', 'A', 'W', 'kWh', 'Hz', '%', 'kW', 'MWh', 'Wh',
                  'VA', 'Sec', 'Ascii', 'R', 'RW', '05h', '04h', '10h',
                  'FC', 'Bit', 'N/A'}
    _RANGE_RE = re.compile(r'^(\d+)\s*[~\-`]\s*\d+')

    current_section = 'control'
    section_keywords = {
        'read input': 'realtime_data',
        'write single': 'write_reg',
        'write force': 'control',
        'fault': 'fault',
    }

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not any(v is not None for v in row):
            continue

        # Detect section header rows (like "2) Read input register (measur...")
        row_text = ' '.join(str(v).lower() for v in row if v is not None)
        for kw, sec in section_keywords.items():
            if kw in row_text:
                current_section = sec
                break

        addr = None
        defn = None
        unit = ''
        dtype = 'U16'
        rw = 'RO'
        regs = 1
        found_range = False

        for i, v in enumerate(row):
            if addr is None:
                if isinstance(v, (int, float)) and 0 <= int(v) < 0xFFFF:
                    addr = int(v)
                elif isinstance(v, str):
                    m = _RANGE_RE.match(v.strip())
                    if m:
                        candidate = int(m.group(1))
                        if 0 <= candidate < 0xFFFF:
                            addr = candidate
                            found_range = True
            elif defn is None and isinstance(v, str):
                vs = v.strip()
                if len(vs) < 2:
                    continue
                # Skip range strings (e.g. '34~35'), FC codes, known units/attrs
                if _RANGE_RE.match(vs):
                    continue
                if vs in _SKIP_UNITS:
                    continue
                # Skip pure numeric / hex values
                if re.match(r'^[\d\.,\-~`]+$', vs) or re.match(r'^0x[0-9A-Fa-f]+', vs):
                    continue
                defn = vs
            # Data type column (col index 9 in EKOS sheet) sets dtype
            if i == 9 and isinstance(v, str):
                vl = v.lower()
                if 'float' in vl:
                    dtype = 'Float32'
                    regs = 2
                elif 'uint32' in vl or 'u32' in vl:
                    dtype = 'U32'
                    regs = 2
                elif 'int16' in vl or 's16' in vl:
                    dtype = 'S16'
            # Unit detection
            if isinstance(v, str) and v in ('V', 'A', 'W', 'kWh', 'Hz', '%', 'kW', 'MWh', 'Wh', 'VA'):
                unit = v

        if addr is None or defn is None:
            continue

        # Assign section: Float32 real-time measurements go to realtime_data
        section = current_section
        if found_range and section == 'control':
            section = 'realtime_data'

        all_regs.append({
            'section': section,
            'definition': defn,
            'address': addr,
            'address_hex': f'0x{addr:04X}',
            'regs': regs,
            'type': dtype,
            'unit': unit,
            'rw': rw,
            'comment': '',
        })
        reg_no += 1

    return {
        'manufacturer': 'EKOS',
        'version': '2022-02-09',
        'all_registers': all_regs,
    }


# ─────────────────────────────────────────────────────────────────────
# Main: run all inverters
# ─────────────────────────────────────────────────────────────────────
all_results = {}
TEST_LABELS = ['T1:Syntax', 'T2:RegisterMap', 'T3:Constants', 'T4:Aliases',
               'T5:UniqueAddr', 'T6:SCALE', 'T7:InverterMode', 'T8:Helpers',
               'T9:MPPT_fn', 'T10:String_fn', 'T11:DATA_TYPES', 'T12:DER-AVM']

for inv in INVERTERS:
    name = inv['name']
    key = inv['key']
    print(f"\n{'='*70}")
    print(f"  INVERTER: {name}")
    print(f"{'='*70}")

    out_dir = os.path.join(OUT_ROOT, key)
    os.makedirs(out_dir, exist_ok=True)

    result = {'name': name, 'key': key}
    logs = {}

    # ─── Parse source document ───────────────────────────────────────
    t0 = time.time()
    if inv.get('pdf') and os.path.isfile(inv['pdf']):
        print(f"  Parsing PDF: {os.path.basename(inv['pdf'])}")
        try:
            parsed = parse_modbus_pdf(inv['pdf'])
            result['parse_ok'] = True
        except Exception as e:
            print(f"  PDF parse error: {e}")
            parsed = {'manufacturer': name, 'all_registers': []}
            result['parse_ok'] = False
    elif inv.get('xlsx') and os.path.isfile(inv.get('xlsx', '')):
        print(f"  Parsing XLSX: {os.path.basename(inv['xlsx'])}")
        try:
            parsed = parse_ekos_xlsx(inv['xlsx'])
            result['parse_ok'] = True
        except Exception as e:
            print(f"  XLSX parse error: {e}")
            parsed = {'manufacturer': name, 'all_registers': []}
            result['parse_ok'] = False
    else:
        print(f"  No source document found — using empty parsed data")
        parsed = {'manufacturer': name, 'all_registers': []}
        result['parse_ok'] = False

    result['parse_time'] = time.time() - t0
    total_regs = len(parsed.get('all_registers', []))
    print(f"  Parsed: {total_regs} registers, manufacturer={parsed.get('manufacturer')}")
    result['parsed_regs'] = total_regs

    # ─── Stage 1: PDF → Excel ─────────────────────────────────────────
    print(f"\n  [STAGE 1: PDF → Excel]")

    # Offline
    s1_off = os.path.join(out_dir, f'{key}_stage1_offline.xlsx')
    t0 = time.time()
    try:
        stage_pipeline.stage1_extract_to_excel(
            parsed, inv.get('pdf', ''), s1_off,
            mode='offline', progress_cb=log_cb(logs, 'offline')
        )
        result['s1_off_ok'] = True
    except Exception as e:
        print(f"  Stage1 Offline ERROR: {e}")
        result['s1_off_ok'] = False
    result['s1_off_time'] = time.time() - t0
    result['s1_off_regs'], result['s1_off_status'], result['s1_off_errors'], result['s1_off_sheets'] = count_stage1(s1_off)
    result['s1_off_data'] = read_stage1_data(s1_off)[:5]  # first 5 rows for content comparison
    print(f"  Offline: {result['s1_off_regs']} regs, {result['s1_off_status']} status, {result['s1_off_errors']} errors, {result['s1_off_time']:.2f}s")
    print(f"  Sheets: {result['s1_off_sheets']}")

    # AI
    s1_ai = os.path.join(out_dir, f'{key}_stage1_ai.xlsx')
    result['s1_ai_real'] = False
    t0 = time.time()
    if API_KEY:
        try:
            stage_pipeline.stage1_extract_to_excel(
                parsed, inv.get('pdf', ''), s1_ai,
                mode='ai', api_key=API_KEY, model=AI_MODEL,
                progress_cb=log_cb(logs, 'ai')
            )
            result['s1_ai_real'] = True
            result['s1_ai_ok'] = True
        except Exception as e:
            print(f"  Stage1 AI ERROR: {e} — using offline copy")
            shutil.copy(s1_off, s1_ai)
            result['s1_ai_ok'] = False
    else:
        shutil.copy(s1_off, s1_ai)
        result['s1_ai_ok'] = False
    result['s1_ai_time'] = time.time() - t0
    result['s1_ai_regs'], result['s1_ai_status'], result['s1_ai_errors'], result['s1_ai_sheets'] = count_stage1(s1_ai)
    result['s1_ai_data'] = read_stage1_data(s1_ai)[:5]
    print(f"  AI:      {result['s1_ai_regs']} regs, real_api={result['s1_ai_real']}, {result['s1_ai_time']:.2f}s")

    # ─── Stage 2: Excel → Mapping Excel ──────────────────────────────
    print(f"\n  [STAGE 2: Excel → Mapping Excel]")

    s2_off = os.path.join(out_dir, f'{key}_stage2_offline.xlsx')
    t0 = time.time()
    try:
        stage_pipeline.stage2_create_mapping_excel(
            s1_off, s2_off,
            solarize_addr_map=SOLARIZE_ADDR_MAP,
            solarize_scale=SOLARIZE_SCALE,
            mode='offline',
            progress_cb=log_cb(logs, 'offline')
        )
        result['s2_off_ok'] = True
    except Exception as e:
        print(f"  Stage2 Offline ERROR: {e}")
        traceback.print_exc()
        result['s2_off_ok'] = False
    result['s2_off_time'] = time.time() - t0
    (result['s2_off_auto'], result['s2_off_manual'], result['s2_off_unmap'],
     result['s2_off_ai_extra'], result['s2_off_der'], result['s2_off_dea']) = count_mapped(s2_off)
    tot_o = result['s2_off_auto'] + result['s2_off_manual'] + result['s2_off_unmap']
    rate_o = result['s2_off_auto'] / tot_o * 100 if tot_o > 0 else 0
    print(f"  Offline: auto={result['s2_off_auto']}, unmap={result['s2_off_unmap']}, rate={rate_o:.1f}%, DER={result['s2_off_der']}, DEA={result['s2_off_dea']}, {result['s2_off_time']:.2f}s")
    result['s2_off_data'] = read_stage2_data(s2_off)[:5]

    s2_ai = os.path.join(out_dir, f'{key}_stage2_ai.xlsx')
    result['s2_ai_real'] = False
    t0 = time.time()
    if API_KEY:
        try:
            stage_pipeline.stage2_create_mapping_excel(
                s1_ai, s2_ai,
                solarize_addr_map=SOLARIZE_ADDR_MAP,
                solarize_scale=SOLARIZE_SCALE,
                mode='ai', api_key=API_KEY, model=AI_MODEL,
                progress_cb=log_cb(logs, 'ai')
            )
            result['s2_ai_real'] = True
            result['s2_ai_ok'] = True
        except Exception as e:
            print(f"  Stage2 AI ERROR: {e} — using offline copy")
            shutil.copy(s2_off, s2_ai)
            result['s2_ai_ok'] = False
    else:
        shutil.copy(s2_off, s2_ai)
        result['s2_ai_ok'] = False
    result['s2_ai_time'] = time.time() - t0
    (result['s2_ai_auto'], result['s2_ai_manual'], result['s2_ai_unmap'],
     result['s2_ai_extra'], result['s2_ai_der'], result['s2_ai_dea']) = count_mapped(s2_ai)
    tot_a = result['s2_ai_auto'] + result['s2_ai_manual'] + result['s2_ai_unmap']
    rate_a = result['s2_ai_auto'] / tot_a * 100 if tot_a > 0 else 0
    print(f"  AI:      auto={result['s2_ai_auto']}, unmap={result['s2_ai_unmap']}, rate={rate_a:.1f}%, real_api={result['s2_ai_real']}, {result['s2_ai_time']:.2f}s")
    result['s2_ai_data'] = read_stage2_data(s2_ai)[:5]

    # ─── Stage 3: Mapping Excel → .py ────────────────────────────────
    print(f"\n  [STAGE 3: Mapping Excel → .py]")

    settings_off = dict(inv['settings'])
    settings_off['protocol_name'] = f'{key}_offline_test'
    s3_off = os.path.join(out_dir, f'{key}_offline_registers.py')
    t0 = time.time()
    try:
        r3o = stage_pipeline.stage3_generate_py(
            s2_off, settings_off, s3_off,
            mode='offline', progress_cb=log_cb(logs, 'offline')
        )
        result['s3_off_ok'] = r3o.get('success', False)
        result['s3_off_code'] = r3o.get('code', '')
        result['s3_off_attempt'] = r3o.get('attempt', 1)
    except Exception as e:
        print(f"  Stage3 Offline ERROR: {e}")
        result['s3_off_ok'] = False
        result['s3_off_code'] = ''
        result['s3_off_attempt'] = 0
    result['s3_off_time'] = time.time() - t0
    result['s3_off_tests'] = run_12_tests(result['s3_off_code'])
    result['s3_off_score'] = sum(1 for s, _ in result['s3_off_tests'] if s == 'OK')
    result['s3_off_attrs'] = len(extract_attrs(result['s3_off_code']))
    print(f"  Offline: ok={result['s3_off_ok']}, score={result['s3_off_score']}/12, attrs={result['s3_off_attrs']}, {result['s3_off_time']:.2f}s")

    settings_ai = dict(inv['settings'])
    settings_ai['protocol_name'] = f'{key}_ai_test'
    s3_ai = os.path.join(out_dir, f'{key}_ai_registers.py')
    result['s3_ai_real'] = False
    t0 = time.time()
    if API_KEY:
        try:
            r3a = stage_pipeline.stage3_generate_py(
                s2_ai, settings_ai, s3_ai,
                mode='ai', api_key=API_KEY, model=AI_MODEL,
                progress_cb=log_cb(logs, 'ai')
            )
            result['s3_ai_real'] = True
            result['s3_ai_ok'] = r3a.get('success', False)
            result['s3_ai_code'] = r3a.get('code', '')
            result['s3_ai_attempt'] = r3a.get('attempt', 1)
        except Exception as e:
            print(f"  Stage3 AI ERROR: {e}")
            result['s3_ai_ok'] = False
            result['s3_ai_code'] = result['s3_off_code']
            result['s3_ai_attempt'] = 0
            shutil.copy(s3_off, s3_ai)
    else:
        result['s3_ai_ok'] = False
        result['s3_ai_code'] = result['s3_off_code']
        result['s3_ai_attempt'] = 0
        shutil.copy(s3_off, s3_ai)
    result['s3_ai_time'] = time.time() - t0
    result['s3_ai_tests'] = run_12_tests(result['s3_ai_code'])
    result['s3_ai_score'] = sum(1 for s, _ in result['s3_ai_tests'] if s == 'OK')
    result['s3_ai_attrs'] = len(extract_attrs(result['s3_ai_code']))
    print(f"  AI:      ok={result['s3_ai_ok']}, score={result['s3_ai_score']}/12, attrs={result['s3_ai_attrs']}, real_api={result['s3_ai_real']}, {result['s3_ai_time']:.2f}s")

    # ─── Reference .py comparison ─────────────────────────────────────
    ref_py = inv.get('ref_py', '')
    if ref_py and os.path.isfile(ref_py):
        with open(ref_py, 'r', encoding='utf-8', errors='replace') as f:
            result['ref_code'] = f.read()
        result['ref_tests'] = run_12_tests(result['ref_code'])
        result['ref_score'] = sum(1 for s, _ in result['ref_tests'] if s == 'OK')
        result['ref_attrs'] = len(extract_attrs(result['ref_code']))
        ref_attrs = extract_attrs(result['ref_code'])
        off_attrs = extract_attrs(result['s3_off_code'])
        ai_attrs  = extract_attrs(result['s3_ai_code'])
        result['off_addr_match'] = sum(1 for k, v in off_attrs.items() if ref_attrs.get(k) == v)
        result['ai_addr_match']  = sum(1 for k, v in ai_attrs.items()  if ref_attrs.get(k) == v)
        result['off_common_keys'] = sum(1 for k in off_attrs if k in ref_attrs)
        result['ai_common_keys']  = sum(1 for k in ai_attrs  if k in ref_attrs)
        print(f"  Ref:     score={result['ref_score']}/12, attrs={result['ref_attrs']}")
        print(f"  Off vs Ref: {result['off_addr_match']}/{result['off_common_keys']} exact matches")
    else:
        result['ref_code'] = ''
        result['ref_score'] = 0
        result['ref_attrs'] = 0
        result['ref_tests'] = [('SKIP', 'No ref')] * 12
        result['off_addr_match'] = 0
        result['ai_addr_match'] = 0
        result['off_common_keys'] = 0
        result['ai_common_keys'] = 0

    all_results[key] = result


# ─────────────────────────────────────────────────────────────────────
# Print comprehensive comparison tables
# ─────────────────────────────────────────────────────────────────────
print(f"\n\n{'='*90}")
print("  FINAL COMPARISON — ALL INVERTERS × ALL STAGES × OFFLINE vs AI")
print(f"{'='*90}")

# ─── Stage 1 Table ────────────────────────────────────────────────────
print(f"\n{'─'*90}")
print(f"  STAGE 1: PDF/XLSX → Excel  (Register extraction)")
print(f"{'─'*90}")
print(f"  {'Inverter':<12} {'Source':<12} {'Off.Regs':>9} {'AI.Regs':>9} {'Status':>7} {'Errors':>7} {'Off.t':>7} {'AI.t':>7} {'Sheets':>5} {'AI실제':>6}")
print(f"  {'-'*88}")
for key, r in all_results.items():
    src = 'PDF' if r.get('parse_ok') and key != 'ekos' else ('XLSX' if key == 'ekos' else 'FAIL')
    ai_real = 'YES' if r.get('s1_ai_real') else 'sim'
    sheets = len(r.get('s1_off_sheets', []))
    print(f"  {r['name']:<12} {src:<12} {r['s1_off_regs']:>9} {r['s1_ai_regs']:>9} "
          f"{r['s1_off_status']:>7} {r['s1_off_errors']:>7} "
          f"{r['s1_off_time']:>6.1f}s {r['s1_ai_time']:>6.1f}s "
          f"{sheets:>5} {ai_real:>6}")

# Stage 1 content sample comparison
print(f"\n  Stage 1 데이터 내용 비교 (첫 3행):")
for key, r in all_results.items():
    print(f"\n  [{r['name']}] Offline Stage1 첫 3행:")
    for row in r.get('s1_off_data', [])[:3]:
        clean = tuple(str(v)[:25] if v is not None else '' for v in row[:8])
        print(f"    {clean}")
    if r.get('s1_ai_real'):
        print(f"  [{r['name']}] AI Stage1 첫 3행:")
        for row in r.get('s1_ai_data', [])[:3]:
            clean = tuple(str(v)[:25] if v is not None else '' for v in row[:8])
            print(f"    {clean}")

# ─── Stage 2 Table ────────────────────────────────────────────────────
print(f"\n{'─'*90}")
print(f"  STAGE 2: Excel → Mapping Excel  (Solarize 표준 매핑)")
print(f"{'─'*90}")
print(f"  {'Inverter':<12} {'Off.Auto':>9} {'Off.Unmap':>10} {'Off.Rate%':>10} {'AI.Auto':>8} {'AI.Unmap':>9} {'AI.Rate%':>9} {'DER':>5} {'DEA':>5} {'AI실제':>6}")
print(f"  {'-'*88}")
for key, r in all_results.items():
    tot_o = r['s2_off_auto'] + r['s2_off_manual'] + r['s2_off_unmap']
    tot_a = r['s2_ai_auto'] + r['s2_ai_manual'] + r['s2_ai_unmap']
    rate_o = r['s2_off_auto'] / tot_o * 100 if tot_o > 0 else 0
    rate_a = r['s2_ai_auto'] / tot_a * 100 if tot_a > 0 else 0
    ai_real = 'YES' if r.get('s2_ai_real') else 'sim'
    print(f"  {r['name']:<12} {r['s2_off_auto']:>9} {r['s2_off_unmap']:>10} {rate_o:>9.1f}% "
          f"{r['s2_ai_auto']:>8} {r['s2_ai_unmap']:>9} {rate_a:>8.1f}% "
          f"{r['s2_off_der']:>5} {r['s2_off_dea']:>5} {ai_real:>6}")

# Stage 2 content sample
print(f"\n  Stage 2 매핑 내용 비교 (첫 3행, 컬럼: No/Addr/Name/→/SolName/MatchType):")
for key, r in all_results.items():
    print(f"\n  [{r['name']}] Offline Stage2 첫 3행:")
    for row in r.get('s2_off_data', [])[:3]:
        # row layout: No, Section, AddrHex, AddrDec, Name, Type, Unit, Scale, RW, Regs, →, sol_name, sol_addr, match_type, notes
        if len(row) >= 14:
            print(f"    No={row[0]}, Addr={row[2]}, Name={str(row[4])[:20]}, SolName={str(row[11] or '')[:20]}, Type={row[13]}")
        else:
            print(f"    {tuple(str(v)[:15] if v else '' for v in row[:8])}")

# ─── Stage 3 Table ────────────────────────────────────────────────────
print(f"\n{'─'*90}")
print(f"  STAGE 3: .py 파일 생성 및 12개 항목 검증")
print(f"{'─'*90}")
print(f"  {'Inverter':<12} {'Ref/12':>7} {'Off/12':>7} {'AI/12':>7} {'Off.Attr':>9} {'AI.Attr':>8} {'Ref.Attr':>9} {'Off.Match':>10} {'AI.Match':>9}")
print(f"  {'-'*88}")
for key, r in all_results.items():
    print(f"  {r['name']:<12} {r['ref_score']:>6}/12 {r['s3_off_score']:>6}/12 {r['s3_ai_score']:>6}/12 "
          f"{r['s3_off_attrs']:>9} {r['s3_ai_attrs']:>8} {r['ref_attrs']:>9} "
          f"{r['off_addr_match']:>4}/{r['off_common_keys']:<4} {r['ai_addr_match']:>4}/{r['ai_common_keys']:<4}")

# ─── 12-Test Detail per Inverter ──────────────────────────────────────
print(f"\n{'─'*90}")
print(f"  12-ITEM VALIDATION DETAIL")
print(f"{'─'*90}")

for key, r in all_results.items():
    print(f"\n  [{r['name']}]")
    print(f"  {'Test':<18} {'Reference':^10} {'Offline':^10} {'AI':^10}  {'Off_msg'}")
    print(f"  {'-'*75}")
    for i, lbl in enumerate(TEST_LABELS):
        ref_s = r['ref_tests'][i][0] if i < len(r['ref_tests']) else 'SKIP'
        off_s = r['s3_off_tests'][i][0] if i < len(r['s3_off_tests']) else 'SKIP'
        ai_s  = r['s3_ai_tests'][i][0] if i < len(r['s3_ai_tests']) else 'SKIP'
        off_msg = r['s3_off_tests'][i][1][:35] if i < len(r['s3_off_tests']) else ''
        icon_r = '✓' if ref_s=='OK' else ('△' if ref_s=='WARN' else '✗')
        icon_o = '✓' if off_s=='OK' else ('△' if off_s=='WARN' else '✗')
        icon_a = '✓' if ai_s=='OK' else ('△' if ai_s=='WARN' else '✗')
        print(f"  {lbl:<18} {icon_r:^10} {icon_o:^10} {icon_a:^10}  {off_msg}")

# ─── Stage 2 Unmapped items ───────────────────────────────────────────
print(f"\n{'─'*90}")
print(f"  UNMAPPED 레지스터 상세 (Offline 기준, 첫 10개)")
print(f"{'─'*90}")
for key, r in all_results.items():
    try:
        wb = openpyxl.load_workbook(
            os.path.join(OUT_ROOT, key, f'{key}_stage2_offline.xlsx'),
            data_only=True)
        ws = wb['Register_Mapping']
        unmapped_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or isinstance(row[0], str):
                continue
            if not isinstance(row[0], (int, float)):
                continue
            mt = str(row[13] or '') if len(row) > 13 else ''
            if mt == 'Unmapped' or mt == '':
                unmapped_rows.append(row)
        print(f"\n  [{r['name']}] Unmapped: {len(unmapped_rows)} rows")
        for row in unmapped_rows[:10]:
            addr = row[2] if len(row) > 2 else ''
            name = str(row[4])[:30] if len(row) > 4 and row[4] else ''
            print(f"    {addr} | {name}")
    except Exception as e:
        print(f"  [{r['name']}] unmapped read error: {e}")

# ─── Summary ──────────────────────────────────────────────────────────
print(f"\n{'='*90}")
print(f"  SUMMARY — 전체 5종 인버터 비교")
print(f"{'='*90}")
print(f"  API 실제 사용: {'YES' if API_KEY else 'NO — AI 모드는 Offline 복사본으로 시뮬레이션'}")
print(f"\n  {'Inverter':<12} {'S1:Regs':>8} {'S2:Rate%':>9} {'S2:DER':>7} {'S3:Score':>9} {'S3:Attrs':>9} {'ref_match':>10}")
print(f"  {'-'*70}")
for key, r in all_results.items():
    tot = r['s2_off_auto'] + r['s2_off_manual'] + r['s2_off_unmap']
    rate = r['s2_off_auto'] / tot * 100 if tot > 0 else 0
    match_str = f"{r['off_addr_match']}/{r['off_common_keys']}" if r['off_common_keys'] > 0 else 'N/A'
    print(f"  {r['name']:<12} {r['s1_off_regs']:>8} {rate:>8.1f}% {r['s2_off_der']:>7} "
          f"{r['s3_off_score']:>7}/12 {r['s3_off_attrs']:>9} {match_str:>10}")

print(f"\n  생성된 파일 목록:")
for key in all_results:
    d = os.path.join(OUT_ROOT, key)
    files = sorted(os.listdir(d))
    print(f"  {key}/: {', '.join(files)}")

print(f"\nDONE. Output dir: {OUT_ROOT}")
