#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Offline vs AI Mode Comparison Script — Solarize Inverter
Stage 1/2/3 pipeline both modes, 12-item validation, reference comparison.
"""
import sys, os, re, json, time, importlib, traceback

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

OUT_DIR = os.path.join(ROOT, 'model_maker', 'comparison_output')
os.makedirs(OUT_DIR, exist_ok=True)

PDF_PATH = os.path.join(ROOT, 'docs', 'Solarize Modbus Protocol-Korea-V1.2.4.pdf')

# ─────────────────────────────────────────────
# 0. API KEY
# ─────────────────────────────────────────────
API_KEY = os.environ.get('CLAUDE_CODE_OAUTH_TOKEN', '') or os.environ.get('ANTHROPIC_API_KEY', '')
AI_MODEL = 'claude-haiku-4-5-20251001'
print(f"API_KEY available: {bool(API_KEY)} (len={len(API_KEY)})")

# ─────────────────────────────────────────────
# 1. Import pipeline
# ─────────────────────────────────────────────
from model_maker import stage_pipeline
from model_maker.modbus_to_udp_mapper import parse_modbus_pdf, ModbusToUdpMapper

# Get Solarize addr map from mapper
mapper_instance = ModbusToUdpMapper.__new__(ModbusToUdpMapper)
SOLARIZE_ADDR_MAP = getattr(mapper_instance, '_SOLARIZE_ADDR_TO_NAME',
                             getattr(ModbusToUdpMapper, '_SOLARIZE_ADDR_TO_NAME', {}))
SOLARIZE_SCALE = {'voltage': 0.1, 'current': 0.01, 'power': 0.1,
                   'frequency': 0.01, 'power_factor': 0.001}

logs = {'offline': [], 'ai': []}
def log_cb(mode):
    def _cb(msg):
        logs[mode].append(msg)
        print(f"  [{mode.upper()}] {msg}")
    return _cb

# ─────────────────────────────────────────────
# 2. PARSE PDF (shared)
# ─────────────────────────────────────────────
print("\n" + "="*60)
print("PARSING PDF")
print("="*60)
t0 = time.time()
parsed = parse_modbus_pdf(PDF_PATH)
parse_time = time.time() - t0
total_regs = len(parsed.get('all_registers', []))
print(f"  Manufacturer : {parsed.get('manufacturer')}")
print(f"  Version      : {parsed.get('version', 'N/A')}")
print(f"  Total regs   : {total_regs}")
print(f"  Parse time   : {parse_time:.2f}s")

# Section breakdown
sections = {}
for r in parsed['all_registers']:
    s = r.get('section', 'unknown')
    sections[s] = sections.get(s, 0) + 1
print(f"  Sections     : {dict(sorted(sections.items()))}")

# ─────────────────────────────────────────────
# 3. STAGE 1
# ─────────────────────────────────────────────
print("\n" + "="*60)
print("STAGE 1: PDF → Excel")
print("="*60)

# --- Offline ---
s1_offline_path = os.path.join(OUT_DIR, 'solarize_stage1_offline.xlsx')
t0 = time.time()
stage_pipeline.stage1_extract_to_excel(
    parsed, PDF_PATH, s1_offline_path,
    mode='offline', progress_cb=log_cb('offline')
)
s1_offline_time = time.time() - t0

# Count rows in offline xlsx
import openpyxl
wb1o = openpyxl.load_workbook(s1_offline_path, data_only=True)
ws1o = wb1o['Registers']
s1_offline_rows = sum(1 for row in ws1o.iter_rows(min_row=2, values_only=True)
                      if isinstance(row[0], (int, float)))
s2_offline_statuses = sum(1 for row in wb1o['Status_Definitions'].iter_rows(min_row=2, values_only=True)
                           if row[0] is not None and not isinstance(row[0], str))
s2_offline_errors = sum(1 for row in wb1o['Error_Fault_Codes'].iter_rows(min_row=2, values_only=True)
                         if row[0] is not None)
print(f"\n[OFFLINE Stage 1 Result]")
print(f"  Register rows : {s1_offline_rows}")
print(f"  Status codes  : {s2_offline_statuses}")
print(f"  Error codes   : {s2_offline_errors}")
print(f"  Time          : {s1_offline_time:.2f}s")
print(f"  File          : {s1_offline_path}")

# --- AI ---
s1_ai_path = os.path.join(OUT_DIR, 'solarize_stage1_ai.xlsx')
s1_ai_rows = s1_offline_rows  # default to offline if AI fails
s1_ai_time = 0.0
s1_ai_success = False

if API_KEY:
    print(f"\n[AI Stage 1 - calling Claude API...]")
    try:
        t0 = time.time()
        stage_pipeline.stage1_extract_to_excel(
            parsed, PDF_PATH, s1_ai_path,
            mode='ai', api_key=API_KEY, model=AI_MODEL,
            progress_cb=log_cb('ai')
        )
        s1_ai_time = time.time() - t0

        wb1a = openpyxl.load_workbook(s1_ai_path, data_only=True)
        ws1a = wb1a['Registers']
        s1_ai_rows = sum(1 for row in ws1a.iter_rows(min_row=2, values_only=True)
                         if isinstance(row[0], (int, float)))
        s1_ai_success = True
        print(f"  Register rows : {s1_ai_rows}")
        print(f"  Time          : {s1_ai_time:.2f}s")
    except Exception as e:
        print(f"  AI Stage 1 FAILED: {e}")
        import shutil
        shutil.copy(s1_offline_path, s1_ai_path)
        s1_ai_rows = s1_offline_rows
else:
    import shutil
    shutil.copy(s1_offline_path, s1_ai_path)
    print(f"  No API key — using offline copy for AI path")

# ─────────────────────────────────────────────
# 4. STAGE 2
# ─────────────────────────────────────────────
print("\n" + "="*60)
print("STAGE 2: Excel → Mapping Excel")
print("="*60)

# --- Offline ---
s2_offline_path = os.path.join(OUT_DIR, 'solarize_stage2_offline.xlsx')
t0 = time.time()
stage_pipeline.stage2_create_mapping_excel(
    s1_offline_path, s2_offline_path,
    solarize_addr_map=SOLARIZE_ADDR_MAP,
    solarize_scale=SOLARIZE_SCALE,
    mode='offline',
    progress_cb=log_cb('offline')
)
s2_offline_time = time.time() - t0

wb2o = openpyxl.load_workbook(s2_offline_path, data_only=True)
ws2o = wb2o['Register_Mapping']
auto_o = man_o = unm_o = 0
for row in ws2o.iter_rows(min_row=2, values_only=True):
    if row[0] is None or (isinstance(row[0], str) and row[0].startswith('---')):
        continue
    if not isinstance(row[0], (int, float)):
        continue
    mt = str(row[13] or '') if len(row) > 13 else ''
    if mt == 'Auto' or mt.startswith('Ref('):
        auto_o += 1
    elif mt == 'Manual':
        man_o += 1
    else:
        unm_o += 1

der_o = len([r for r in wb2o['DER_AVM_Control'].iter_rows(min_row=2, values_only=True)
             if r[0] is not None])
dea_o = len([r for r in wb2o['DEA_AVM_Monitor'].iter_rows(min_row=2, values_only=True)
             if r[0] is not None])
total_mapped_o = auto_o + man_o

print(f"\n[OFFLINE Stage 2 Result]")
print(f"  Auto/Ref mapped : {auto_o}")
print(f"  Manual mapped   : {man_o}")
print(f"  Unmapped        : {unm_o}")
print(f"  Mapping rate    : {total_mapped_o/(auto_o+man_o+unm_o)*100:.1f}% ({total_mapped_o}/{auto_o+man_o+unm_o})" if (auto_o+man_o+unm_o) > 0 else "  Mapping rate    : N/A")
print(f"  DER-AVM regs    : {der_o}")
print(f"  DEA-AVM regs    : {dea_o}")
print(f"  Time            : {s2_offline_time:.2f}s")

# --- AI ---
s2_ai_path = os.path.join(OUT_DIR, 'solarize_stage2_ai.xlsx')
s2_ai_success = False
auto_a = auto_o; man_a = man_o; unm_a = unm_o

if API_KEY:
    print(f"\n[AI Stage 2 - calling Claude API for unmapped regs...]")
    try:
        t0 = time.time()
        stage_pipeline.stage2_create_mapping_excel(
            s1_ai_path, s2_ai_path,
            solarize_addr_map=SOLARIZE_ADDR_MAP,
            solarize_scale=SOLARIZE_SCALE,
            mode='ai', api_key=API_KEY, model=AI_MODEL,
            progress_cb=log_cb('ai')
        )
        s2_ai_time = time.time() - t0

        wb2a = openpyxl.load_workbook(s2_ai_path, data_only=True)
        ws2a = wb2a['Register_Mapping']
        auto_a = man_a = unm_a = ai_mapped_a = 0
        for row in ws2a.iter_rows(min_row=2, values_only=True):
            if row[0] is None or (isinstance(row[0], str) and row[0].startswith('---')):
                continue
            if not isinstance(row[0], (int, float)):
                continue
            mt = str(row[13] or '') if len(row) > 13 else ''
            if mt.startswith('AI('):
                ai_mapped_a += 1
                auto_a += 1
            elif mt == 'Auto' or mt.startswith('Ref('):
                auto_a += 1
            elif mt == 'Manual':
                man_a += 1
            else:
                unm_a += 1
        s2_ai_success = True
        total_mapped_a = auto_a + man_a
        print(f"  Auto/Ref+AI     : {auto_a}  (AI additional: {ai_mapped_a})")
        print(f"  Unmapped        : {unm_a}")
        print(f"  Mapping rate    : {total_mapped_a/(auto_a+man_a+unm_a)*100:.1f}%")
        print(f"  Time            : {s2_ai_time:.2f}s")
    except Exception as e:
        print(f"  AI Stage 2 FAILED: {e}")
        import shutil
        shutil.copy(s2_offline_path, s2_ai_path)
        s2_ai_time = 0
else:
    import shutil
    shutil.copy(s2_offline_path, s2_ai_path)
    s2_ai_time = 0
    print("  No API key — using offline copy for AI path")

# ─────────────────────────────────────────────
# 5. STAGE 3
# ─────────────────────────────────────────────
print("\n" + "="*60)
print("STAGE 3: Mapping Excel → .py")
print("="*60)

SETTINGS = {
    'mppt_count': 4, 'string_count': 8,
    'iv_scan': True, 'der_avm': True, 'dea_avm': True,
    'manufacturer': 'Solarize VK50', 'class_name': 'RegisterMap',
    'protocol_name': 'solarize_offline_test', 'fc_code': 'FC03',
}

# --- Offline Stage 3 ---
s3_offline_path = os.path.join(OUT_DIR, 'solarize_offline_registers.py')
t0 = time.time()
r3o = stage_pipeline.stage3_generate_py(
    s2_offline_path, SETTINGS, s3_offline_path,
    mode='offline', progress_cb=log_cb('offline')
)
s3_offline_time = time.time() - t0
code_offline = r3o.get('code', '')
print(f"\n[OFFLINE Stage 3 Result]")
print(f"  Success      : {r3o.get('success', False)}")
print(f"  Code length  : {len(code_offline)} chars")
print(f"  Time         : {s3_offline_time:.2f}s")
if r3o.get('results'):
    for status, msg in r3o['results'][:5]:
        print(f"  Validation   : [{status}] {msg[:80]}")

# --- AI Stage 3 ---
SETTINGS_AI = dict(SETTINGS)
SETTINGS_AI['protocol_name'] = 'solarize_ai_test'
s3_ai_path = os.path.join(OUT_DIR, 'solarize_ai_registers.py')
s3_ai_success = False

if API_KEY:
    print(f"\n[AI Stage 3 - code gen + validation + retry...]")
    try:
        t0 = time.time()
        r3a = stage_pipeline.stage3_generate_py(
            s2_ai_path, SETTINGS_AI, s3_ai_path,
            mode='ai', api_key=API_KEY, model=AI_MODEL,
            progress_cb=log_cb('ai')
        )
        s3_ai_time = time.time() - t0
        code_ai = r3a.get('code', '')
        s3_ai_success = r3a.get('success', False)
        print(f"  Success      : {s3_ai_success}")
        print(f"  Code length  : {len(code_ai)} chars")
        print(f"  Attempt      : {r3a.get('attempt', 1)}")
        print(f"  Time         : {s3_ai_time:.2f}s")
        if r3a.get('results'):
            for status, msg in r3a['results'][:8]:
                print(f"  Validation   : [{status}] {msg[:80]}")
    except Exception as e:
        print(f"  AI Stage 3 FAILED: {e}")
        traceback.print_exc()
        code_ai = code_offline
        s3_ai_time = 0
        r3a = {'success': False, 'code': code_offline, 'results': [], 'attempt': 0}
else:
    code_ai = code_offline
    s3_ai_time = 0
    r3a = {'success': False, 'code': code_offline, 'results': []}
    print("  No API key — AI Stage 3 skipped")

# ─────────────────────────────────────────────
# 6. 12-ITEM VALIDATION
# ─────────────────────────────────────────────
print("\n" + "="*60)
print("12-ITEM VALIDATION TESTS")
print("="*60)

def run_12_tests(code, label):
    results = []
    if not code:
        return [('SKIP', 'No code')] * 12

    # T1: Syntax
    try:
        compile(code, '<generated>', 'exec')
        results.append(('OK', 'Python syntax valid'))
    except SyntaxError as e:
        results.append(('FAIL', f'SyntaxError: {e}'))

    # T2: RegisterMap class
    if re.search(r'\bclass\s+RegisterMap\b', code):
        results.append(('OK', 'class RegisterMap exists'))
    else:
        results.append(('FAIL', 'class RegisterMap missing'))

    # T3: Register constants (non-negative ints)
    consts = re.findall(r'^\s+(\w+)\s*=\s*(0x[0-9A-Fa-f]+|\d+)', code, re.MULTILINE)
    if consts:
        results.append(('OK', f'{len(consts)} register constants found'))
    else:
        results.append(('FAIL', 'No register constants found'))

    # T4: Essential aliases
    essential = ['L1_VOLTAGE', 'L1_CURRENT', 'L2_VOLTAGE', 'L3_VOLTAGE',
                 'MPPT1_VOLTAGE', 'MPPT1_CURRENT', 'INVERTER_MODE', 'ERROR_CODE1',
                 'TOTAL_ENERGY_LOW', 'TODAY_ENERGY_LOW', 'POWER_FACTOR',
                 'PV_TOTAL_INPUT_POWER_LOW', 'GRID_TOTAL_ACTIVE_POWER_LOW']
    missing = [a for a in essential if a not in code]
    if not missing:
        results.append(('OK', f'All {len(essential)} essential aliases present'))
    else:
        results.append(('FAIL', f'Missing aliases: {missing}'))

    # T5: Address uniqueness
    addr_matches = re.findall(r'=\s*(0x[0-9A-Fa-f]+)', code)
    addrs = [int(a, 16) for a in addr_matches]
    dups = len(addrs) - len(set(addrs))
    if dups == 0:
        results.append(('OK', f'{len(addrs)} addresses, all unique'))
    else:
        results.append(('WARN', f'{dups} duplicate addresses (check aliases)'))

    # T6: SCALE dict
    scale_keys = re.findall(r"'(\w+)':\s*[\d.]+", code[code.find('SCALE'):code.find('SCALE')+500] if 'SCALE' in code else '')
    required_scale = ['voltage', 'current', 'power', 'frequency', 'power_factor']
    missing_scale = [k for k in required_scale if k not in scale_keys]
    if not missing_scale:
        results.append(('OK', f'SCALE has all {len(required_scale)} required keys'))
    else:
        results.append(('FAIL', f'SCALE missing: {missing_scale}'))

    # T7: InverterMode class
    if re.search(r'\bclass\s+InverterMode\b', code):
        modes_ok = all(m in code for m in ['INITIAL', 'STANDBY', 'ON_GRID', 'FAULT', 'SHUTDOWN'])
        if modes_ok and 'to_string' in code:
            results.append(('OK', 'InverterMode class complete (5 states + to_string)'))
        else:
            results.append(('WARN', 'InverterMode class found but incomplete'))
    else:
        results.append(('FAIL', 'InverterMode class missing'))

    # T8: Helper functions
    helpers = ['registers_to_u32', 'registers_to_s32']
    missing_h = [h for h in helpers if h not in code]
    if not missing_h:
        results.append(('OK', 'registers_to_u32/s32 present'))
    else:
        results.append(('FAIL', f'Missing helpers: {missing_h}'))

    # T9: get_mppt_registers
    if 'get_mppt_registers' in code:
        results.append(('OK', 'get_mppt_registers() present'))
    else:
        results.append(('FAIL', 'get_mppt_registers() missing'))

    # T10: get_string_registers
    if 'get_string_registers' in code:
        results.append(('OK', 'get_string_registers() present'))
    else:
        results.append(('FAIL', 'get_string_registers() missing'))

    # T11: DATA_TYPES dict
    if 'DATA_TYPES' in code:
        valid_types = ['U16', 'S16', 'U32', 'S32', 'ASCII', 'FLOAT']
        found_types = [t for t in valid_types if t in code[code.find('DATA_TYPES'):code.find('DATA_TYPES')+1000]]
        results.append(('OK', f'DATA_TYPES present, types found: {found_types}'))
    else:
        results.append(('FAIL', 'DATA_TYPES dict missing'))

    # T12: DER-AVM attributes
    der_attrs = ['DER_POWER_FACTOR_SET', 'DER_ACTION_MODE',
                 'DER_REACTIVE_POWER_PCT', 'DER_ACTIVE_POWER_PCT', 'INVERTER_ON_OFF']
    missing_der = [a for a in der_attrs if a not in code]
    if not missing_der:
        results.append(('OK', f'All {len(der_attrs)} DER-AVM attributes present'))
    else:
        results.append(('FAIL', f'DER-AVM missing: {missing_der}'))

    return results

test_labels = ['T1:Syntax', 'T2:RegisterMap', 'T3:Constants', 'T4:Aliases',
               'T5:UniqueAddr', 'T6:SCALE', 'T7:InverterMode', 'T8:Helpers',
               'T9:MPPT_fn', 'T10:String_fn', 'T11:DATA_TYPES', 'T12:DER-AVM']

results_offline = run_12_tests(code_offline, 'Offline')
results_ai      = run_12_tests(code_ai,      'AI')

# Load reference solarize_registers.py for comparison
ref_path = os.path.join(ROOT, 'common', 'solarize_registers.py')
with open(ref_path, 'r', encoding='utf-8') as f:
    code_ref = f.read()
results_ref = run_12_tests(code_ref, 'Reference')

print(f"\n{'Test':<18} {'Reference':^12} {'Offline':^12} {'AI':^12}")
print("-"*58)
offline_score = ai_score = ref_score = 0
for i, (label, ro, ra, rr) in enumerate(zip(test_labels, results_offline, results_ai, results_ref)):
    so = 'OK' if ro[0] == 'OK' else ('WARN' if ro[0] == 'WARN' else 'FAIL')
    sa = 'OK' if ra[0] == 'OK' else ('WARN' if ra[0] == 'WARN' else 'FAIL')
    sr = 'OK' if rr[0] == 'OK' else ('WARN' if rr[0] == 'WARN' else 'FAIL')
    if ro[0] == 'OK': offline_score += 1
    if ra[0] == 'OK': ai_score += 1
    if rr[0] == 'OK': ref_score += 1
    print(f"  {label:<16} {sr:^12} {so:^12} {sa:^12}")

print("-"*58)
print(f"  {'TOTAL SCORE':<16} {ref_score}/12{'':<7} {offline_score}/12{'':<7} {ai_score}/12")

# ─────────────────────────────────────────────
# 7. ADDRESS COMPARISON with solarize_registers.py
# ─────────────────────────────────────────────
print("\n" + "="*60)
print("ADDRESS COMPARISON vs solarize_registers.py")
print("="*60)

def extract_attrs(code):
    attrs = {}
    in_class = False
    for line in code.split('\n'):
        s = line.strip()
        if re.match(r'^class\s+\w+.*:', s):
            in_class = 'RegisterMap' in s
            continue
        if not in_class:
            continue
        m = re.match(r'(\w+)\s*=\s*(0x[0-9A-Fa-f]+|\d+)', s)
        if m:
            v = m.group(2)
            attrs[m.group(1)] = int(v, 16) if v.startswith('0x') else int(v)
    return attrs

ref_attrs = extract_attrs(code_ref)
off_attrs = extract_attrs(code_offline)
ai_attrs  = extract_attrs(code_ai)

print(f"\n  Reference attributes : {len(ref_attrs)}")
print(f"  Offline attributes   : {len(off_attrs)}")
print(f"  AI attributes        : {len(ai_attrs)}")

# Exact address matches
off_match = sum(1 for k, v in off_attrs.items() if ref_attrs.get(k) == v)
ai_match  = sum(1 for k, v in ai_attrs.items()  if ref_attrs.get(k) == v)
off_total = sum(1 for k in off_attrs if k in ref_attrs)
ai_total  = sum(1 for k in ai_attrs  if k in ref_attrs)

print(f"\n  Offline vs Reference: {off_match}/{off_total} exact address matches")
print(f"  AI vs Reference     : {ai_match}/{ai_total} exact address matches")

# Key register check
key_regs = {
    'R_PHASE_VOLTAGE/L1_VOLTAGE': [0x1001],
    'MPPT1_VOLTAGE':               [0x1010],
    'INVERTER_MODE':               [0x101D],
    'ERROR_CODE1':                 [0x101E],
    'TOTAL_ENERGY/LOW':            [0x1021],
    'AC_POWER/GRID_POWER':         [0x1037],
    'DER_POWER_FACTOR_SET':        [0x07D0],
    'INVERTER_ON_OFF':             [0x0834],
}
print(f"\n  Key register address check:")
print(f"  {'Register':<30} {'Ref':>8} {'Offline':>10} {'AI':>10}")
print(f"  {'-'*62}")
for name, expected in key_regs.items():
    ref_v   = ref_attrs.get(name.split('/')[0], ref_attrs.get(name.split('/')[-1] if '/' in name else name))
    off_v   = off_attrs.get(name.split('/')[0], off_attrs.get(name.split('/')[-1] if '/' in name else name))
    ai_v    = ai_attrs.get(name.split('/')[0], ai_attrs.get(name.split('/')[-1] if '/' in name else name))
    ref_s   = hex(ref_v) if ref_v is not None else 'N/A'
    off_s   = hex(off_v) if off_v is not None else 'N/A'
    ai_s    = hex(ai_v)  if ai_v is not None else 'N/A'
    off_ok  = 'OK' if off_v is not None and off_v in expected + ([ref_v] if ref_v else []) else '??'
    ai_ok   = 'OK' if ai_v is not None and ai_v in expected + ([ref_v] if ref_v else []) else '??'
    print(f"  {name:<30} {ref_s:>8}  {off_s:>8}{off_ok}  {ai_s:>8}{ai_ok}")

# ─────────────────────────────────────────────
# 8. FINAL SUMMARY TABLE
# ─────────────────────────────────────────────
print("\n" + "="*60)
print("FINAL COMPARISON SUMMARY")
print("="*60)

summary = f"""
+----------------------------------------------------------+
|       Solarize Offline vs AI Mode Comparison            |
+---------------------+---------------+-------------------+
| Item                |   Offline     |    AI Mode        |
+---------------------+---------------+-------------------+
| [Stage 1]           |               |                   |
|  Extracted regs     | {s1_offline_rows:>5}         | {s1_ai_rows:>5}             |
|  Status codes       | {s2_offline_statuses:>5}         | (same)            |
|  Error codes        | {s2_offline_errors:>5}         | (same)            |
|  Time               | {s1_offline_time:>4.1f}s         | {"API+"+str(round(s1_ai_time,1))+"s" if s1_ai_success else "N/A (copy)":>15} |
+---------------------+---------------+-------------------+
| [Stage 2]           |               |                   |
|  Auto/Ref mapped    | {auto_o:>5}         | {auto_a:>5}             |
|  Unmapped           | {unm_o:>5}         | {unm_a:>5}             |
|  Mapping rate       | {total_mapped_o/(auto_o+man_o+unm_o)*100 if (auto_o+man_o+unm_o)>0 else 0:>4.1f}%        | {auto_a/(auto_a+man_a+unm_a)*100 if (auto_a+man_a+unm_a)>0 else 0:>4.1f}%             |
|  DER-AVM included   |  {der_o} regs      | (same)            |
|  Time               | {s2_offline_time:>4.1f}s         | {"API+"+str(round(s2_ai_time,1))+"s" if s2_ai_success else "N/A (copy)":>15} |
+---------------------+---------------+-------------------+
| [Stage 3]           |               |                   |
|  Validation /12     | {offline_score:>5}         | {ai_score:>5}             |
|  Code length        | {len(code_offline):>5} chars     | {len(code_ai):>5} chars         |
|  RegisterMap attrs  | {len(off_attrs):>5}         | {len(ai_attrs):>5}             |
|  Addr match (ref)   | {off_match}/{off_total}          | {ai_match}/{ai_total}              |
|  Time               | {s3_offline_time:>4.1f}s         | {"API+"+str(round(s3_ai_time,1))+"s" if s3_ai_success else "N/A (copy)":>15} |
+---------------------+---------------+-------------------+
| [vs solarize_reg]   | Reference=12  |                   |
|  Validation score   | {offline_score}/12         | {ai_score}/12             |
+---------------------+---------------+-------------------+
"""
print(summary)

print(f"Output files:")
print(f"  Stage1: {s1_offline_path}")
print(f"  Stage1: {s1_ai_path}")
print(f"  Stage2: {s2_offline_path}")
print(f"  Stage2: {s2_ai_path}")
print(f"  Stage3: {s3_offline_path}")
print(f"  Stage3: {s3_ai_path}")

print("\nDONE.")
