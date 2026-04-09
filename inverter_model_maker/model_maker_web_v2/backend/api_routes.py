# -*- coding: utf-8 -*-
"""
REST API + WebSocket 라우트 — 파이프라인 연결
"""
from fastapi import APIRouter, UploadFile, File, Form, WebSocket, WebSocketDisconnect, HTTPException
from fastapi.responses import FileResponse
import os
import shutil
import asyncio
import traceback
import json

from .session_store import SessionStore
from .ws_manager import ws_manager

router = APIRouter()

import hashlib
import threading

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
COMMON_DIR = os.path.join(PROJECT_ROOT, 'common')
RESULTS_DIR = os.path.join(os.path.dirname(__file__), '..', 'results')
DEFINITIONS_DIR = os.path.join(os.path.dirname(__file__), 'pipeline', 'definitions')
SUCCESS_INDEX_PATH = os.path.join(RESULTS_DIR, '_success_index.json')
_success_index_lock = threading.Lock()


def _load_success_index() -> dict:
    try:
        with open(SUCCESS_INDEX_PATH, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _save_success_index(index: dict):
    os.makedirs(RESULTS_DIR, exist_ok=True)
    with open(SUCCESS_INDEX_PATH, 'w', encoding='utf-8') as f:
        json.dump(index, f, ensure_ascii=False, indent=2)


def _pdf_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def _record_success(pdf_path: str, manufacturer: str, h01_match: dict,
                    der_match: dict, info_match: dict, meta: dict):
    """H01+DER 완전 매칭 시 성공 인덱스에 기록"""
    try:
        sha = _pdf_sha256(pdf_path)
        record = {
            'manufacturer': manufacturer,
            'pdf_filename': os.path.basename(pdf_path),
            'h01': [h01_match.get('matched', 0), h01_match.get('total', 0)],
            'der': [der_match.get('matched', 0), der_match.get('total', 0)],
            'info_match': info_match,
            'max_mppt': meta.get('max_mppt', 0),
            'max_string': meta.get('max_string', 0),
            'date': __import__('datetime').date.today().isoformat(),
        }
        with _success_index_lock:
            index = _load_success_index()
            index[sha] = record
            _save_success_index(index)
    except Exception:
        pass


def _check_success(pdf_path: str) -> dict | None:
    """PDF 해시로 기존 성공 기록 조회. 없으면 None."""
    try:
        sha = _pdf_sha256(pdf_path)
        with _success_index_lock:
            return _load_success_index().get(sha)
    except Exception:
        return None


def _save_to_results(manufacturer: str, src_path: str, label: str) -> str:
    """Stage 결과 파일을 results/{제조사}/ 디렉토리에 저장. 저장된 경로 반환."""
    if not manufacturer or not os.path.exists(src_path):
        return ''
    mfr_dir = os.path.join(RESULTS_DIR, manufacturer)
    os.makedirs(mfr_dir, exist_ok=True)
    ext = os.path.splitext(src_path)[1]
    dest = os.path.join(mfr_dir, f'{manufacturer}_{label}{ext}')
    shutil.copy2(src_path, dest)
    return dest


def _copy_definitions_to_results(manufacturer: str):
    """definitions/{mfr}_definitions.json → results/{제조사}/"""
    mfr_key = manufacturer.lower()
    json_src = os.path.join(DEFINITIONS_DIR, f'{mfr_key}_definitions.json')
    if not os.path.exists(json_src):
        return
    mfr_dir = os.path.join(RESULTS_DIR, manufacturer)
    os.makedirs(mfr_dir, exist_ok=True)
    shutil.copy2(json_src, os.path.join(mfr_dir, f'{manufacturer}_definitions.json'))


def _copy_source_to_results(manufacturer: str, src_path: str):
    """업로드된 원본 프로토콜 문서(PDF/Excel) → results/{제조사}/"""
    if not manufacturer or not src_path or not os.path.exists(src_path):
        return
    mfr_dir = os.path.join(RESULTS_DIR, manufacturer)
    os.makedirs(mfr_dir, exist_ok=True)
    fname = os.path.basename(src_path)
    dest = os.path.join(mfr_dir, fname)
    if not os.path.exists(dest):  # 이미 있으면 덮어쓰지 않음
        shutil.copy2(src_path, dest)


# ─── 비동기 파이프라인 실행 헬퍼 ─────────────────────────────────────────────

async def _run_in_thread(func, *args, **kwargs):
    """동기 함수를 스레드 풀에서 실행"""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, lambda: func(*args, **kwargs))


def _make_progress_callback(sid: str, stage: str, loop: asyncio.AbstractEventLoop):
    """WebSocket 로그 전송 콜백 생성 — loop을 캡처하여 스레드 안전"""
    def callback(msg: str, level: str = 'info'):
        asyncio.run_coroutine_threadsafe(
            ws_manager.send_json(sid, {
                'stage': stage,
                'text': msg,
                'level': level,
            }),
            loop
        )
    return callback


# ─── 세션 ────────────────────────────────────────────────────────────────────

@router.post('/session/new')
def new_session():
    sid = SessionStore.create()
    return {'session_id': sid}


@router.get('/session/{sid}')
def get_session(sid: str):
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')
    return s


# ─── Stage 1 ─────────────────────────────────────────────────────────────────

@router.post('/stage1/upload')
async def stage1_upload(
    session_id: str = Form(...),
    file: UploadFile = File(...),
):
    """PDF 또는 Excel 업로드 → 작업 디렉토리에 저장"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    work_dir = SessionStore.get_work_dir(session_id)
    dest = os.path.join(work_dir, file.filename)
    with open(dest, 'wb') as f:
        shutil.copyfileobj(file.file, f)

    # 새 파일 업로드 → 진행 중 태스크 취소 + 세션 상태 완전 리셋
    SessionStore.cancel_running_task(session_id)
    SessionStore.update(
        session_id,
        uploaded_file=dest,
        stage=0,
        stage1_excel=None,
        stage2_excel=None,
        registers_py=None,
        meta={},
    )
    return {'saved': dest, 'filename': file.filename}


def _load_ai_settings() -> dict:
    """Load Claude API settings from config/ai_settings.ini.

    Searches two locations: PROJECT_ROOT/config/ (inverter_model_maker/) and
    the parent RTU project root (V2_0_0/config/) since the dashboard saves
    ai_settings.ini in V2_0_0/config/.
    """
    import configparser
    # Try MM2 project root first, then parent RTU project root
    rtu_root = os.path.abspath(os.path.join(PROJECT_ROOT, '..'))
    candidates = [
        os.path.join(PROJECT_ROOT, 'config', 'ai_settings.ini'),
        os.path.join(rtu_root, 'config', 'ai_settings.ini'),
    ]
    cfg_path = next((p for p in candidates if os.path.exists(p)), candidates[0])
    cp = configparser.ConfigParser()
    cp.read(cfg_path, encoding='utf-8')
    key = cp.get('claude_api', 'api_key', fallback='')
    model = cp.get('claude_api', 'model', fallback='claude-sonnet-4-6')
    if key and key != 'YOUR_ANTHROPIC_API_KEY_HERE':
        return {'api_key': key, 'model': model}
    return {}


@router.post('/stage1/run')
async def stage1_run(body: dict):
    """
    Stage 1 실행
    body: {session_id, filename, device_type, use_ai?}
    """
    sid = body.get('session_id')
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')

    device_type = body.get('device_type', 'inverter')
    use_ai = body.get('use_ai', False)
    uploaded = s.get('uploaded_file')
    if not uploaded or not os.path.exists(uploaded):
        raise HTTPException(400, 'no file uploaded')

    work_dir = SessionStore.get_work_dir(sid)

    # Load AI settings if AI mode requested
    ai_settings = _load_ai_settings() if use_ai else {}
    if use_ai and not ai_settings:
        raise HTTPException(400, 'AI mode requested but no API key configured. '
                            'Set your Claude API key in the dashboard Model Maker tab.')

    # 이전 태스크 취소 후 새 태스크 실행
    SessionStore.cancel_running_task(sid)

    async def _run():
        try:
            from .pipeline.stage1 import run_stage1, NotRegisterMapError

            # ── 성공 인덱스 조회 (PDF 해시 기반) ──
            cached = await _run_in_thread(_check_success, uploaded)
            if cached:
                h01c, der_c = cached['h01'], cached['der']
                await ws_manager.send_json(sid, {
                    'event': 'already_verified',
                    'stage': 's1',
                    'manufacturer': cached['manufacturer'],
                    'pdf_filename': cached['pdf_filename'],
                    'h01_matched': h01c[0], 'h01_total': h01c[1],
                    'der_matched': der_c[0], 'der_total': der_c[1],
                    'info_match': cached.get('info_match', {}),
                    'max_mppt': cached.get('max_mppt', 0),
                    'max_string': cached.get('max_string', 0),
                    'date': cached.get('date', ''),
                })
                # 성공 기록이 있어도 Stage 1을 계속 실행하여 Excel 생성
                # (다운로드/Stage2 진행 가능하도록)

            progress = _make_progress_callback(sid, 's1', asyncio.get_running_loop())
            result = await _run_in_thread(
                run_stage1, uploaded, work_dir, device_type, progress,
                ai_settings if use_ai else None)

            # 태스크 취소로 인해 세션이 리셋된 경우 이벤트 전송 생략
            if SessionStore.get(sid) is None:
                return

            SessionStore.update(sid,
                                stage=1,
                                stage1_excel=result['output_path'],
                                meta=result['meta'],
                                suggestions=result.get('suggestions', {}))

            # results/{제조사}/ 에 자동 저장
            mfr = result['meta'].get('manufacturer', '')
            if mfr:
                _save_to_results(mfr, result['output_path'], 'stage1')
                _copy_definitions_to_results(mfr)
                _copy_source_to_results(mfr, uploaded)

            # Stage 1에서 H01+DER 완전 매칭이면 성공 기록
            h01m = result.get('h01_match', {})
            derm = result.get('der_match', {})
            if (h01m.get('matched') == h01m.get('total') and h01m.get('total', 0) > 0
                    and derm.get('matched') == derm.get('total') and derm.get('total', 0) > 0
                    and not cached):
                await _run_in_thread(
                    _record_success, uploaded, mfr, h01m, derm,
                    result.get('info_match', {}), result['meta'])

            await ws_manager.send_json(sid, {
                'event': 'stage1_done',
                'stage': 's1',
                'text': f'Stage 1 완료: {result["output_name"]}',
                'level': 'ok',
                'counts': result['counts'],
                'review_count': result['review_count'],
                'iv_scan': result['meta'].get('iv_scan', False),
                'iv_data_points': result['meta'].get('iv_data_points', 0),
                'info_match': result.get('info_match', {'model': False, 'sn': False}),
                'suggestions': result.get('suggestions', {}),
                'max_mppt': result['meta'].get('max_mppt', 0),
                'max_string': result['meta'].get('max_string', 0),
                'phase_type': result['meta'].get('phase_type', 'unknown'),
            })
        except asyncio.CancelledError:
            pass  # 새 파일 업로드로 인해 취소됨 — 정상
        except NotRegisterMapError as e:
            await ws_manager.send_json(sid, {
                'event': 'invalid_file',
                'stage': 's1',
                'text': str(e),
                'level': 'err',
            })
        except Exception as e:
            await ws_manager.send_json(sid, {
                'stage': 's1',
                'text': f'오류: {str(e)}',
                'level': 'err',
            })
            traceback.print_exc()

    task = asyncio.ensure_future(_run(), loop=asyncio.get_event_loop())
    SessionStore.update(sid, _running_task=task)
    return {'status': 'running', 'session_id': sid}


@router.post('/stage1/apply-suggestion')
async def apply_suggestion(body: dict):
    """
    Stage 1 suggestion 적용 → synonym_db에 등록 → Stage 1 재실행
    body: {session_id, field, definition, addr, rerun}
    """
    sid = body.get('session_id')
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')

    field = body.get('field', '')
    definition = body.get('definition', '')
    addr = body.get('addr', '')
    rerun = body.get('rerun', True)

    # synonym_db에 (definition → field) 매핑 추가
    if field and definition:
        from .pipeline import load_synonym_db, save_synonym_db
        db = load_synonym_db()
        fields = db.setdefault('fields', {})
        # 기존 항목이 list이면 dict 형식으로 변환 (구버전 호환)
        entry = fields.get(field)
        if isinstance(entry, list):
            fields[field] = {'category': 'MONITORING', 'h01_field': field, 'synonyms': entry}
        elif entry is None:
            fields[field] = {'category': 'MONITORING', 'h01_field': field, 'synonyms': []}
        synonyms = fields[field].setdefault('synonyms', [])
        key = definition.upper().replace(' ', '_')
        if key not in synonyms:
            synonyms.append(key)
            save_synonym_db(db)

    if not rerun:
        return {'status': 'ok', 'field': field, 'added': definition}

    # Stage 1 재실행
    uploaded = s.get('uploaded_file')
    device_type = s.get('meta', {}).get('device_type', 'inverter')
    if not uploaded or not os.path.exists(uploaded):
        raise HTTPException(400, 'no file uploaded')

    work_dir = SessionStore.get_work_dir(sid)

    async def _rerun():
        try:
            from .pipeline.stage1 import run_stage1, NotRegisterMapError
            loop = asyncio.get_running_loop()
            progress = _make_progress_callback(sid, 's1', loop)
            progress(f'[제안 적용] {field} → {definition}', 'info')
            result = await _run_in_thread(run_stage1, uploaded, work_dir, device_type, progress)

            SessionStore.update(sid,
                                stage=1,
                                stage1_excel=result['output_path'],
                                meta=result['meta'],
                                suggestions=result.get('suggestions', {}))

            mfr = result['meta'].get('manufacturer', '')
            if mfr:
                _save_to_results(mfr, result['output_path'], 'stage1')
                _copy_definitions_to_results(mfr)
                _copy_source_to_results(mfr, uploaded)

            await ws_manager.send_json(sid, {
                'event': 'stage1_done',
                'stage': 's1',
                'text': f'Stage 1 재실행 완료: {result["output_name"]}',
                'level': 'ok',
                'counts': result['counts'],
                'review_count': result['review_count'],
                'iv_scan': result['meta'].get('iv_scan', False),
                'iv_data_points': result['meta'].get('iv_data_points', 0),
                'info_match': result.get('info_match', {'model': False, 'sn': False}),
                'suggestions': result.get('suggestions', {}),
                'max_mppt': result['meta'].get('max_mppt', 0),
                'max_string': result['meta'].get('max_string', 0),
                'phase_type': result['meta'].get('phase_type', 'unknown'),
            })
        except NotRegisterMapError as e:
            await ws_manager.send_json(sid, {
                'event': 'invalid_file',
                'stage': 's1',
                'text': str(e),
                'level': 'err',
            })
        except Exception as e:
            await ws_manager.send_json(sid, {
                'stage': 's1',
                'text': f'재실행 오류: {str(e)}',
                'level': 'err',
            })
            import traceback; traceback.print_exc()

    asyncio.ensure_future(_rerun(), loop=asyncio.get_event_loop())
    return {'status': 'rerunning', 'session_id': sid}


# ─── Stage 2 ─────────────────────────────────────────────────────────────────

@router.post('/stage2/upload')
async def stage2_upload(
    session_id: str = Form(...),
    file: UploadFile = File(...),
):
    """Stage 1 Excel 직접 업로드 → PDF 없이 Stage 2 시작"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    fname = file.filename or ''
    if not fname.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, 'Stage 1 Excel (.xlsx) 파일만 업로드 가능합니다')

    work_dir = SessionStore.get_work_dir(session_id)
    dest = os.path.join(work_dir, fname)
    with open(dest, 'wb') as f:
        shutil.copyfileobj(file.file, f)

    SessionStore.update(session_id,
                        stage1_excel=dest,
                        stage=1,
                        stage2_excel=None,
                        registers_py=None)
    return {'saved': dest, 'filename': fname}


@router.post('/stage2/run')
async def stage2_run(body: dict):
    """
    body: {session_id, mppt, total_strings, capacity, iv_data_points}
    backward-compat: strings_per_mppt = total_strings // mppt 형식도 허용
    """
    sid = body.get('session_id')
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')

    stage1_excel = s.get('stage1_excel')
    if not stage1_excel or not os.path.exists(stage1_excel):
        raise HTTPException(400, 'Stage 1 not completed')

    mppt = int(body.get('mppt', 4))
    # 신규: total_strings (사용자가 직접 총 스트링 수 입력 — Growatt 등 비대칭 지원)
    # 구버전 호환: strings_per_mppt가 들어오면 mppt × spm 으로 계산
    if 'total_strings' in body:
        total_strings = int(body.get('total_strings', 0))
    else:
        spm = int(body.get('strings_per_mppt', 2))
        total_strings = mppt * spm
    capacity = body.get('capacity', '')
    iv_data_points = int(body.get('iv_data_points', 64))
    work_dir = SessionStore.get_work_dir(sid)
    SessionStore.update(sid, iv_data_points=iv_data_points)

    async def _run():
        try:
            from .pipeline.stage2 import run_stage2
            progress = _make_progress_callback(sid, 's2', asyncio.get_running_loop())
            result = await _run_in_thread(
                run_stage2, stage1_excel, work_dir,
                mppt, total_strings, capacity, progress)

            SessionStore.update(sid,
                                stage=2,
                                stage2_excel=result['output_path'])

            # results/{제조사}/ 에 자동 저장
            sess = SessionStore.get(sid) or {}
            mfr = sess.get('meta', {}).get('manufacturer', '')
            if mfr:
                _save_to_results(mfr, result['output_path'], 'stage2')

            # Stage 2에서 H01+DER 완전 매칭이면 성공 기록
            h01_ok = result['h01_matched'] == result['h01_total'] and result['h01_total'] > 0
            der_ok = result['der_matched'] == result['der_total'] and result['der_total'] > 0
            if h01_ok and der_ok and mfr:
                uploaded = sess.get('uploaded_file', '')
                if uploaded and os.path.exists(uploaded):
                    h01m = {'matched': result['h01_matched'], 'total': result['h01_total']}
                    derm = {'matched': result['der_matched'], 'total': result['der_total']}
                    meta = sess.get('meta', {})
                    info_match = sess.get('suggestions', {}).get('_info_match',
                                         {'model': False, 'sn': False})
                    await _run_in_thread(
                        _record_success, uploaded, mfr, h01m, derm, info_match, meta)

            await ws_manager.send_json(sid, {
                'event': 'stage2_done',
                'stage': 's2',
                'text': f'Stage 2 완료: {result["output_name"]}',
                'level': 'ok',
                'h01_matched': result['h01_matched'],
                'h01_total': result['h01_total'],
                'der_matched': result['der_matched'],
                'der_total': result['der_total'],
                'review_count': result['review_count'],
                'review_items': result.get('review_items', []),
                'stage2_validation': result.get('stage2_validation', {}),
                'stage2_pass': result.get('stage2_pass', False),
                'fail_reasons': result.get('fail_reasons', []),
                'phase_type': result.get('phase_type', 'unknown'),
                'mppt_count': result.get('mppt_count', 0),
                'total_strings': result.get('total_strings', 0),
            })
        except Exception as e:
            await ws_manager.send_json(sid, {
                'stage': 's2',
                'text': f'오류: {str(e)}',
                'level': 'err',
            })
            traceback.print_exc()

    asyncio.ensure_future(_run(), loop=asyncio.get_event_loop())
    return {'status': 'running', 'session_id': sid}


# ─── Stage 3 ─────────────────────────────────────────────────────────────────

@router.post('/stage3/run')
async def stage3_run(body: dict):
    """
    body: {session_id}
    """
    sid = body.get('session_id')
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')

    stage2_excel = s.get('stage2_excel')
    if not stage2_excel or not os.path.exists(stage2_excel):
        raise HTTPException(400, 'Stage 2 not completed')

    work_dir = SessionStore.get_work_dir(sid)
    iv_data_points = s.get('iv_data_points', 64)

    async def _run():
        try:
            from .pipeline.stage3 import run_stage3
            progress = _make_progress_callback(sid, 's3', asyncio.get_running_loop())
            result = await _run_in_thread(
                run_stage3, stage2_excel, work_dir, progress, iv_data_points)

            SessionStore.update(sid,
                                stage=3,
                                registers_py=result['output_path'])

            await ws_manager.send_json(sid, {
                'event': 'stage3_done',
                'stage': 's3',
                'text': f'Stage 3 완료: {result["filename"]}',
                'level': 'ok',
                'filename': result['filename'],
                'validation': result['validation'],
                'stage4': result.get('stage4'),
                'stage4_pass': result.get('stage4_pass', False),
                'phase_type': result.get('phase_type', 'unknown'),
                'synonym_added': result['synonym_added'],
                'review_recorded': result['review_recorded'],
            })
        except Exception as e:
            await ws_manager.send_json(sid, {
                'stage': 's3',
                'text': f'오류: {str(e)}',
                'level': 'err',
            })
            traceback.print_exc()

    asyncio.ensure_future(_run(), loop=asyncio.get_event_loop())
    return {'status': 'running', 'session_id': sid}


# ─── H01 매핑 ────────────────────────────────────────────────────────────────

_RE_CLEAN_REG = __import__('re').compile(r'^([A-Za-z0-9_가-힣]+)')


def _parse_reg_display(s: str) -> str:
    """'NAME (0x0123); NAME2 (0x0456)' → 'NAME' (첫 항목의 깨끗한 이름)"""
    if not s:
        return ''
    first = s.split(';')[0].strip()
    m = _RE_CLEAN_REG.match(first)
    return m.group(1) if m else ''


@router.get('/h01-mapping/{session_id}')
def get_h01_mapping(session_id: str):
    """Stage2 Excel의 H01_MAPPING 시트를 읽어 JSON 반환"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    stage2_excel = s.get('stage2_excel')
    if not stage2_excel or not os.path.exists(stage2_excel):
        raise HTTPException(400, 'Stage 2 not completed')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(stage2_excel, data_only=True)
        if 'H01_MAPPING' not in wb.sheetnames:
            raise HTTPException(404, 'H01_MAPPING sheet not found')

        ws = wb['H01_MAPPING']
        DATA_START = 4

        # H열 전체에서 사용 가능한 레지스터 목록 수집 (드롭다운용) — 알파벳 정렬
        # J/K/L열(숨김 메타데이터)에서 레지스터별 FC + 주소 맵을 읽음
        #   J=name, K="3"/"4"/"3,4", L="0x0091|0x0091" (fcs 순서와 동일)
        seen = set()
        name_to_fcs: dict[str, list[int]] = {}
        name_fc_to_addr: dict[tuple, str] = {}
        for row_idx in range(DATA_START, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=8).value  # H열
            if val:
                v = str(val).strip()
                if v:
                    seen.add(v)
            jname = ws.cell(row=row_idx, column=10).value  # J열
            kfcs = ws.cell(row=row_idx, column=11).value   # K열
            laddr = ws.cell(row=row_idx, column=12).value  # L열
            if jname:
                nm = str(jname).strip()
                fcs: list[int] = []
                if kfcs:
                    for tok in str(kfcs).split(','):
                        t = tok.strip()
                        if t in ('3', '4'):
                            fcs.append(int(t))
                fcs = sorted(set(fcs))
                if nm and nm not in name_to_fcs:
                    name_to_fcs[nm] = fcs
                # 주소 매핑 — fcs 순서와 같은 길이의 '|' 구분 문자열
                if nm and laddr:
                    addrs = str(laddr).split('|')
                    for fc, addr in zip(fcs, addrs):
                        if addr:
                            name_fc_to_addr[(nm, fc)] = addr.strip()
        available_registers = sorted(seen)
        # 드롭다운에서 사용할 풍부한 형태 (FC + 주소 정보 포함)
        available_with_fc = []
        for nm in available_registers:
            fcs = name_to_fcs.get(nm, [])
            addrs = {fc: name_fc_to_addr.get((nm, fc), '') for fc in fcs}
            available_with_fc.append({'name': nm, 'fcs': fcs, 'addrs': addrs})

        # H01 필드 데이터 (B열에 값이 있는 행만 동적으로 읽기)
        fields = []
        for row in range(DATA_START, ws.max_row + 1):
            h01_field = str(ws.cell(row=row, column=2).value or '').strip()
            if not h01_field:
                break
            description = str(ws.cell(row=row, column=3).value or '').strip()
            current_raw = str(ws.cell(row=row, column=4).value or '').strip()
            # 깨끗한 register name 추출 (auto-fill 형식 "NAME (addr); ..." → NAME만)
            current_clean = _parse_reg_display(current_raw)
            # M열: 매칭 출처 (Stage2 가 기록) — 'pdf' / 'handler' / 'der' / ''
            match_source_raw = str(ws.cell(row=row, column=13).value or '').strip().lower()
            # 매칭 판정:
            #   pdf      → clean name 이 available 목록에 실제 존재해야 진짜 매칭
            #   handler  → 항상 매칭 (HANDLER 계산 — pv_voltage/current/power)
            #   der      → 항상 매칭 (DER-AVM 고정 주소)
            if match_source_raw == 'handler':
                is_matched = True
                match_source = 'handler'
            elif match_source_raw == 'der':
                is_matched = True
                match_source = 'der'
            else:
                is_matched = bool(current_clean) and current_clean in available_registers
                match_source = 'pdf' if is_matched else ''

            # 추천 후보 (E~G열) — 동일하게 정리
            suggestions = []
            for col in (5, 6, 7):
                sv = ws.cell(row=row, column=col).value
                if sv:
                    clean = _parse_reg_display(str(sv).strip())
                    if clean and clean in available_registers and clean not in suggestions:
                        suggestions.append(clean)

            # FC 코드 (I열, col 9)
            fc_val = ws.cell(row=row, column=9).value
            fc = int(fc_val) if fc_val else 3

            # 현재 매칭된 레지스터의 FC 정보로 lock/ambiguous 판단
            cur_fcs = name_to_fcs.get(current_clean, []) if current_clean else []
            fc_locked = len(cur_fcs) == 1
            fc_ambiguous = len(cur_fcs) >= 2
            if fc_locked:
                # 단일 FC만 갖는 레지스터면 그 값으로 강제
                fc = cur_fcs[0]
            # 현재 선택된 (name, fc) 조합의 실제 주소 — 사용자 검증용
            current_addr = name_fc_to_addr.get((current_clean, fc), '') if current_clean else ''

            fields.append({
                'h01_field': h01_field,
                'description': description,
                'current_register': current_clean,  # 깨끗한 이름만
                'current_raw': current_raw,  # 원본 (디버깅용)
                'is_matched': is_matched,
                'match_source': match_source,  # 'pdf' / 'handler' / 'der' / ''
                'suggestions': suggestions,
                'fc': fc,
                'fc_locked': fc_locked,
                'fc_ambiguous': fc_ambiguous,
                'current_fcs': cur_fcs,
                'current_addr': current_addr,
            })

        wb.close()
        return {
            'fields': fields,
            'available_registers': available_registers,
            'available_with_fc': available_with_fc,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f'H01_MAPPING 읽기 오류: {str(e)}')


@router.post('/h01-mapping/{session_id}')
def save_h01_mapping(session_id: str, body: dict):
    """사용자가 수정한 H01 매핑을 Stage2 Excel에 저장"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    stage2_excel = s.get('stage2_excel')
    if not stage2_excel or not os.path.exists(stage2_excel):
        raise HTTPException(400, 'Stage 2 not completed')

    mappings = body.get('mappings', {})
    fc_map = body.get('fc', {})  # {h01_field: 3 or 4}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(stage2_excel)
        if 'H01_MAPPING' not in wb.sheetnames:
            raise HTTPException(404, 'H01_MAPPING sheet not found')

        ws = wb['H01_MAPPING']
        DATA_START = 4

        updated = 0
        for row in range(DATA_START, ws.max_row + 1):
            h01_field = str(ws.cell(row=row, column=2).value or '').strip()
            if not h01_field:
                break
            if h01_field in mappings:
                ws.cell(row=row, column=4, value=mappings[h01_field] or None)
                updated += 1
            if h01_field in fc_map:
                ws.cell(row=row, column=9, value=int(fc_map[h01_field]))

        wb.save(stage2_excel)
        wb.close()
        return {'status': 'ok', 'updated': updated}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f'H01_MAPPING 저장 오류: {str(e)}')


# ─── MPPT_MAPPING (Stage 2) ──────────────────────────────────────────────────

@router.get('/mppt-mapping/{session_id}')
def get_mppt_mapping(session_id: str):
    """Stage2 Excel의 MPPT_MAPPING 시트를 읽어 JSON 반환"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    stage2_excel = s.get('stage2_excel')
    if not stage2_excel or not os.path.exists(stage2_excel):
        raise HTTPException(400, 'Stage 2 not completed')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(stage2_excel, data_only=True)
        if 'MPPT_MAPPING' not in wb.sheetnames:
            raise HTTPException(404, 'MPPT_MAPPING sheet not found')

        ws = wb['MPPT_MAPPING']
        DATA_START = 4

        seen = set()
        name_to_addr: dict[str, str] = {}
        name_to_fc: dict[str, str] = {}
        for row_idx in range(DATA_START, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=8).value
            if val:
                v = str(val).strip()
                if v:
                    seen.add(v)
            jname = ws.cell(row=row_idx, column=10).value
            kfc = ws.cell(row=row_idx, column=11).value
            laddr = ws.cell(row=row_idx, column=12).value
            if jname:
                nm = str(jname).strip()
                if nm and nm not in name_to_addr:
                    name_to_addr[nm] = str(laddr or '').strip()
                    name_to_fc[nm] = str(kfc or '3').strip()
        available_registers = sorted(seen)
        available_with_fc = [
            {'name': nm, 'addr': name_to_addr.get(nm, ''), 'fc': name_to_fc.get(nm, '3')}
            for nm in available_registers
        ]

        # B열에 값이 있는 행만 동적으로 읽기
        fields = []
        for row in range(DATA_START, ws.max_row + 1):
            field = str(ws.cell(row=row, column=2).value or '').strip()
            if not field:
                break
            description = str(ws.cell(row=row, column=3).value or '').strip()
            current_raw = str(ws.cell(row=row, column=4).value or '').strip()
            current_clean = _parse_reg_display(current_raw)

            match_source_raw = str(ws.cell(row=row, column=13).value or '').strip().lower()
            if match_source_raw == 'handler':
                is_matched = True
                match_source = 'handler'
            elif match_source_raw == 'der':
                is_matched = True
                match_source = 'der'
            else:
                is_matched = bool(current_clean) and current_clean in available_registers
                match_source = 'pdf' if is_matched else ''

            suggestions = []
            for col in (5, 6, 7):
                sv = ws.cell(row=row, column=col).value
                if sv:
                    clean = _parse_reg_display(str(sv).strip())
                    if clean and clean not in suggestions:
                        suggestions.append(clean)

            fc_val = ws.cell(row=row, column=9).value
            fc = int(fc_val) if fc_val else 3
            current_addr = name_to_addr.get(current_clean, '') if current_clean else ''

            fields.append({
                'field': field,
                'description': description,
                'current_register': current_clean,
                'current_raw': current_raw,
                'is_matched': is_matched,
                'match_source': match_source,
                'suggestions': suggestions,
                'fc': fc,
                'current_addr': current_addr,
            })

        wb.close()
        return {
            'fields': fields,
            'available_registers': available_registers,
            'available_with_fc': available_with_fc,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f'MPPT_MAPPING 읽기 오류: {str(e)}')


@router.post('/mppt-mapping/{session_id}')
def save_mppt_mapping(session_id: str, body: dict):
    """사용자가 수정한 MPPT 매핑을 Stage2 Excel에 저장"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    stage2_excel = s.get('stage2_excel')
    if not stage2_excel or not os.path.exists(stage2_excel):
        raise HTTPException(400, 'Stage 2 not completed')

    mappings = body.get('mappings', {})
    try:
        import openpyxl
        wb = openpyxl.load_workbook(stage2_excel)
        if 'MPPT_MAPPING' not in wb.sheetnames:
            raise HTTPException(404, 'MPPT_MAPPING sheet not found')
        ws = wb['MPPT_MAPPING']
        DATA_START = 4
        updated = 0
        for row in range(DATA_START, ws.max_row + 1):
            field = str(ws.cell(row=row, column=2).value or '').strip()
            if not field:
                break
            if field in mappings:
                ws.cell(row=row, column=4, value=mappings[field] or None)
                ws.cell(row=row, column=13, value='pdf' if mappings[field] else '')
                updated += 1
        wb.save(stage2_excel)
        wb.close()
        return {'status': 'ok', 'updated': updated}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f'MPPT_MAPPING 저장 오류: {str(e)}')


# ─── INFO_MAPPING (Stage 1) ──────────────────────────────────────────────────

@router.get('/info-mapping/{session_id}')
def get_info_mapping(session_id: str):
    """Stage1 Excel의 INFO_MAPPING 시트를 읽어 JSON 반환 (Model + SN)"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    stage1_excel = s.get('stage1_excel')
    if not stage1_excel or not os.path.exists(stage1_excel):
        raise HTTPException(400, 'Stage 1 not completed')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(stage1_excel, data_only=True)
        if 'INFO_MAPPING' not in wb.sheetnames:
            raise HTTPException(404, 'INFO_MAPPING sheet not found')

        ws = wb['INFO_MAPPING']
        DATA_START = 4

        # H열 전체 후보 + J/K/L 메타 수집
        seen = set()
        name_to_addr: dict[str, str] = {}
        name_to_fc: dict[str, str] = {}
        for row_idx in range(DATA_START, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=8).value
            if val:
                v = str(val).strip()
                if v:
                    seen.add(v)
            jname = ws.cell(row=row_idx, column=10).value
            kfc = ws.cell(row=row_idx, column=11).value
            laddr = ws.cell(row=row_idx, column=12).value
            if jname:
                nm = str(jname).strip()
                if nm and nm not in name_to_addr:
                    name_to_addr[nm] = str(laddr or '').strip()
                    name_to_fc[nm] = str(kfc or '3').strip()
        available_registers = sorted(seen)
        available_with_fc = [
            {'name': nm, 'addr': name_to_addr.get(nm, ''), 'fc': name_to_fc.get(nm, '3')}
            for nm in available_registers
        ]

        # 데이터 행 — model, sn
        fields = []
        for row in range(DATA_START, DATA_START + 2):
            field = str(ws.cell(row=row, column=2).value or '').strip()
            if not field:
                break
            description = str(ws.cell(row=row, column=3).value or '').strip()
            current_raw = str(ws.cell(row=row, column=4).value or '').strip()
            current_clean = _parse_reg_display(current_raw)

            match_source_raw = str(ws.cell(row=row, column=13).value or '').strip().lower()
            is_matched = bool(current_clean) and current_clean in available_registers
            match_source = match_source_raw if match_source_raw else ('pdf' if is_matched else '')

            suggestions = []
            for col in (5, 6, 7):
                sv = ws.cell(row=row, column=col).value
                if sv:
                    clean = _parse_reg_display(str(sv).strip())
                    if clean and clean not in suggestions:
                        suggestions.append(clean)

            fc_val = ws.cell(row=row, column=9).value
            fc = int(fc_val) if fc_val else 3

            current_addr = name_to_addr.get(current_clean, '') if current_clean else ''

            fields.append({
                'field': field,
                'description': description,
                'current_register': current_clean,
                'current_raw': current_raw,
                'is_matched': is_matched,
                'match_source': match_source,
                'suggestions': suggestions,
                'fc': fc,
                'current_addr': current_addr,
            })

        wb.close()
        return {
            'fields': fields,
            'available_registers': available_registers,
            'available_with_fc': available_with_fc,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f'INFO_MAPPING 읽기 오류: {str(e)}')


@router.post('/info-mapping/{session_id}')
def save_info_mapping(session_id: str, body: dict):
    """사용자가 수정한 INFO 매핑을 Stage1 Excel에 저장"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    stage1_excel = s.get('stage1_excel')
    if not stage1_excel or not os.path.exists(stage1_excel):
        raise HTTPException(400, 'Stage 1 not completed')

    mappings = body.get('mappings', {})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(stage1_excel)
        if 'INFO_MAPPING' not in wb.sheetnames:
            raise HTTPException(404, 'INFO_MAPPING sheet not found')

        ws = wb['INFO_MAPPING']
        DATA_START = 4
        updated = 0
        for row in range(DATA_START, DATA_START + 2):
            field = str(ws.cell(row=row, column=2).value or '').strip()
            if not field:
                break
            if field in mappings:
                ws.cell(row=row, column=4, value=mappings[field] or None)
                # 사용자 수동 매핑은 'pdf' 출처로 기록
                ws.cell(row=row, column=13, value='pdf' if mappings[field] else '')
                updated += 1

        wb.save(stage1_excel)
        wb.close()
        return {'status': 'ok', 'updated': updated}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f'INFO_MAPPING 저장 오류: {str(e)}')


# ─── 파일 다운로드 ───────────────────────────────────────────────────────────

@router.get('/download/{session_id}/{filename}')
def download_file(session_id: str, filename: str):
    work_dir = SessionStore.get_work_dir(session_id)
    path = os.path.join(work_dir, filename)
    if not os.path.exists(path):
        raise HTTPException(404, 'file not found')
    return FileResponse(path, filename=filename)


@router.get('/download-stage/{session_id}/{stage}')
def download_stage_file(session_id: str, stage: int):
    """Stage별 결과 파일 다운로드"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    if stage == 1:
        path = s.get('stage1_excel')
    elif stage == 2:
        path = s.get('stage2_excel')
    elif stage == 3:
        path = s.get('registers_py')
    else:
        raise HTTPException(400, 'invalid stage')

    if not path or not os.path.exists(path):
        raise HTTPException(404, 'file not found')
    return FileResponse(path, filename=os.path.basename(path))


# ─── WebSocket ───────────────────────────────────────────────────────────────

@router.websocket('/ws/{session_id}')
async def websocket_endpoint(ws: WebSocket, session_id: str):
    await ws_manager.connect(session_id, ws)
    try:
        while True:
            data = await ws.receive_text()
            # 클라이언트 메시지 처리 (REVIEW 판정 등)
            try:
                msg = json.loads(data)
                if msg.get('type') == 'review_verdicts':
                    # REVIEW 판정을 Stage 2 Excel에 반영
                    s = SessionStore.get(session_id)
                    if s and s.get('stage2_excel'):
                        _update_review_verdicts(s['stage2_excel'], msg.get('verdicts', []))
                        await ws.send_json({'event': 'verdicts_saved', 'count': len(msg.get('verdicts', []))})
            except (json.JSONDecodeError, KeyError):
                pass
    except WebSocketDisconnect:
        ws_manager.disconnect(session_id)


def _update_review_verdicts(excel_path: str, verdicts: list):
    """Stage 2 Excel REVIEW 시트에 사용자 판정 기록 + review_history 저장"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)

        # V2: 2_REVIEW 또는 REVIEW 시트
        review_sheet = '2_REVIEW' if '2_REVIEW' in wb.sheetnames else 'REVIEW'
        if review_sheet not in wb.sheetnames:
            wb.close()
            return

        ws = wb[review_sheet]
        header = [str(c.value).strip() if c.value else '' for c in ws[1]]

        # '선택' 또는 '판정' 컬럼 인덱스
        verdict_col = None
        for i, h in enumerate(header):
            if '판정' in h or '선택' in h:
                verdict_col = i + 1
                break
        if verdict_col is None:
            wb.close()
            return

        # Definition 컬럼
        def_col = None
        addr_col = None
        for i, h in enumerate(header):
            if 'definition' in h.lower() or h == 'Definition':
                def_col = i + 1
            if 'address' in h.lower() or h == 'Address':
                addr_col = i + 1

        # 판정 기록 + review_history 수집
        history_entries = []
        for v in verdicts:
            row_idx = v.get('row', 0) + 4  # V2: 데이터는 row 4부터 (title + empty + header + data)
            verdict_val = v.get('verdict', '')
            # ALT1/ALT2 → KEEP으로 매핑 (추천 채택)
            write_val = verdict_val
            if verdict_val in ('ALT1', 'ALT2'):
                write_val = 'KEEP'  # 추천 채택 = 포함
            ws.cell(row=row_idx, column=verdict_col, value=write_val)

            # review_history에 저장할 데이터
            defn = str(ws.cell(row=row_idx, column=def_col).value or '') if def_col else ''
            addr = str(ws.cell(row=row_idx, column=addr_col).value or '') if addr_col else ''
            if defn and verdict_val == 'DELETE':
                history_entries.append({
                    'definition': defn.upper().replace(' ', '_'),
                    'address': addr,
                    'verdict': 'DELETE',
                })

        wb.save(excel_path)
        wb.close()

        # review_history에 DELETE 항목 저장 (다음번 자동 제외)
        if history_entries:
            from .pipeline import load_review_history, save_review_history
            history = load_review_history()
            existing = {item.get('definition', '') for item in history.get('approved', [])}
            added = 0
            for entry in history_entries:
                if entry['definition'] not in existing:
                    history['approved'].append(entry)
                    added += 1
            if added:
                history['stats']['total_reviewed'] = history['stats'].get('total_reviewed', 0) + added
                save_review_history(history)

    except Exception:
        pass
