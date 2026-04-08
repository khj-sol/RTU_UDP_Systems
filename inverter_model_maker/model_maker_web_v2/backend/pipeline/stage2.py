# -*- coding: utf-8 -*-
"""
Stage 2 — MPPT/String 필터링 + REVIEW 추천 → Stage 2 Excel (MAPPING_RULES_V2)

입력: Stage 1 Excel (4시트: 1_INFO, 2_H01, 3_DER, 4_IV)
출력: Stage 2 Excel (3시트: 1_REGISTER_MAP, 2_REVIEW, 3_SUMMARY)

V2 핵심:
- 사용자가 MPPT/String/용량 지정 → 해당 채널만 필터링
- REVIEW 항목에 추천 대안 매핑 제안 (synonym 퍼지 + 주소 근접 레퍼런스)
"""
import os
import re
from datetime import datetime
from typing import List, Dict, Optional, Tuple

from . import (
    PROJECT_ROOT, RegisterRow, CATEGORY_COLORS, MATCH_COLORS,
    H01_DER_OVERLAP_FIELDS, H01_HANDLER_COMPUTED_FIELDS,
    DER_CONTROL_REGS, DER_MONITOR_REGS,
    to_upper_snake, parse_address,
    match_synonym, match_synonym_fuzzy, detect_channel_number,
    load_synonym_db, load_review_history, load_reference_patterns,
    get_openpyxl, ProgressCallback,
)
from .rules import filter_channels_stage2, distribute_alarms


# ─── Stage 1 Excel V2 파서 ─────────────────────────────────────────────────

def read_stage1_excel_v2(excel_path: str) -> dict:
    """
    Stage 1 Excel (4시트 구조) 읽기
    Returns: {
        'meta': dict,
        'info': [RegisterRow],
        'monitoring': [RegisterRow],
        'status': [RegisterRow],
        'alarm': [RegisterRow],
        'review': [RegisterRow],
        'h01_match': [dict],  # H01 매칭 테이블
        'iv_regs': [RegisterRow],
    }
    """
    openpyxl = get_openpyxl()
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    result = {
        'meta': {}, 'info': [], 'monitoring': [], 'status': [],
        'alarm': [], 'review': [], 'h01_match': [], 'iv_regs': [],
    }

    # 구 형식 호환 (SUMMARY/META/INFO/MONITORING/... 개별 시트)
    is_v2_format = '1_INFO' in wb.sheetnames
    if not is_v2_format:
        result = _read_stage1_legacy(wb)
        wb.close()
        return result

    # ── 1_INFO 시트: 메타 + INFO/STATUS/ALARM/REVIEW 섹션 파싱 ──
    if '1_INFO' in wb.sheetnames:
        ws = wb['1_INFO']
        current_section = 'META'
        header = None

        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            first = cells[0] if cells else ''

            # 섹션 감지
            if first.startswith('Stage 1'):
                continue
            if 'INFO 레지스터' in first:
                current_section = 'INFO'
                header = None
                continue
            if 'STATUS 레지스터' in first:
                current_section = 'STATUS'
                header = None
                continue
            if 'ALARM 레지스터' in first:
                current_section = 'ALARM'
                header = None
                continue
            if 'REVIEW' in first and '개' in first:
                current_section = 'REVIEW'
                header = None
                continue

            if current_section == 'META':
                if first and cells[1] if len(cells) > 1 else '':
                    key = first.strip()
                    val = cells[1].strip() if len(cells) > 1 else ''
                    # 메타 키 매핑
                    meta_map = {
                        '제조사': 'manufacturer', '프로토콜 버전': 'protocol_version',
                        '설비 타입': 'device_type', 'MPPT': 'max_mppt',
                        'String': 'max_string', 'IV Scan': 'iv_scan',
                        'IV Data Points': 'iv_data_points', 'IV Trackers': 'iv_trackers',
                        'H01 매칭': 'h01_match_str', 'DER 매칭': 'der_match_str',
                        '추출 레지스터': 'total_extracted', 'REVIEW': 'review_count',
                    }
                    if key in meta_map:
                        result['meta'][meta_map[key]] = val
            else:
                # 헤더 행 감지
                if first == 'No' or first == 'No.':
                    header = cells
                    continue
                if not header or not first:
                    continue

                # 레지스터 행 파싱
                reg = _parse_section_row(cells, header, current_section)
                if reg:
                    if current_section == 'INFO':
                        result['info'].append(reg)
                    elif current_section == 'STATUS':
                        result['status'].append(reg)
                    elif current_section == 'ALARM':
                        result['alarm'].append(reg)
                    elif current_section == 'REVIEW':
                        result['review'].append(reg)

    # ── 2_H01 시트: H01 매칭 테이블만 (MONITORING은 3_MPPT_STRING에) ──
    if '2_H01' in wb.sheetnames:
        ws = wb['2_H01']
        header = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            first = cells[0] if cells else ''
            if first.startswith('H01') or (first and not first.isdigit() and first != 'No'):
                if first == 'No':
                    header = cells
                continue
            if first == 'No':
                header = cells
                continue
            if header and first.isdigit():
                result['h01_match'].append({
                    'field': cells[1], 'source': cells[2] if len(cells) > 2 else '',
                    'status': cells[3] if len(cells) > 3 else '',
                    'address': cells[4] if len(cells) > 4 else '',
                    'definition': cells[5] if len(cells) > 5 else '',
                    'note': cells[9] if len(cells) > 9 else '',
                })

    # ── 3_MPPT_STRING 시트: MONITORING 전체 목록 ──
    ms_sheet = '3_MPPT_STRING'
    if ms_sheet in wb.sheetnames:
        ws = wb[ms_sheet]
        current_section = 'META'
        header = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            first = cells[0] if cells else ''
            if 'MONITORING 전체' in first:
                current_section = 'MONITORING'
                header = None
                continue
            if current_section != 'MONITORING':
                continue
            if first == 'No':
                header = cells
                continue
            if header and first.isdigit():
                reg = _parse_section_row(cells, header, 'MONITORING')
                if reg:
                    result['monitoring'].append(reg)

    # ── 5_IV 시트: IV 레지스터 (backward compat: 4_IV도 지원) ──
    # 두 섹션:
    #   1) IV Scan Command (헤더: Name/Address/Type/R/W/Description)
    #   2) IV Data 레지스터 매핑 (헤더: No/Type/Name/Address/Regs/Data Type/Scale)
    # 빈 행이나 새 섹션 제목 행에서 헤더 재감지
    iv_sheet = '5_IV' if '5_IV' in wb.sheetnames else ('4_IV' if '4_IV' in wb.sheetnames else None)
    if iv_sheet:
        ws = wb[iv_sheet]
        col_name = col_addr = col_type = None
        in_data_rows = False
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            if not any(cells):
                # 빈 행 → 다음 섹션 헤더 대기 모드
                in_data_rows = False
                col_name = col_addr = col_type = None
                continue
            upper = [c.upper() for c in cells]
            # 헤더 감지: NAME + ADDRESS 컬럼 있는 행
            if 'NAME' in upper and 'ADDRESS' in upper:
                col_name = upper.index('NAME')
                col_addr = upper.index('ADDRESS')
                if 'TYPE' in upper:
                    col_type = upper.index('TYPE')
                else:
                    col_type = None
                # 'DATA TYPE' 컬럼이 있으면 그것을 사용
                if 'DATA TYPE' in upper:
                    col_type = upper.index('DATA TYPE')
                in_data_rows = True
                continue
            if not in_data_rows or col_name is None:
                continue
            # 데이터 행
            try:
                name = cells[col_name] if col_name < len(cells) else ''
                addr_str = cells[col_addr] if col_addr < len(cells) else ''
                if not name or not addr_str:
                    continue
                addr = parse_address(addr_str)
                if addr is None:
                    continue
                dtype = cells[col_type] if col_type is not None and col_type < len(cells) else 'U16'
                result['iv_regs'].append(RegisterRow(
                    definition=name,
                    address=addr,
                    data_type=dtype or 'U16',
                    category='IV_SCAN',
                ))
            except (IndexError, ValueError):
                continue

    wb.close()
    return result


def _read_stage1_legacy(wb) -> dict:
    """구 형식 Stage 1 Excel (개별 카테고리 시트) 읽기"""
    result = {
        'meta': {}, 'info': [], 'monitoring': [], 'status': [],
        'alarm': [], 'review': [], 'h01_match': [], 'iv_regs': [],
    }

    # META 시트
    if 'META' in wb.sheetnames:
        ws = wb['META']
        for row in ws.iter_rows(values_only=True):
            if row[0] and row[1]:
                result['meta'][str(row[0]).strip()] = str(row[1]).strip()

    # 카테고리 시트 파싱
    cat_map = {
        'INFO': 'info', 'MONITORING': 'monitoring', 'STATUS': 'status',
        'ALARM': 'alarm', 'REVIEW': 'review', 'IV_SCAN': 'iv_regs',
    }
    for sheet_name, key in cat_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            if not header:
                if cells[0] in ('No', 'No.'):
                    header = cells
                continue
            if not cells[0]:
                continue
            reg = _parse_section_row(cells, header, sheet_name)
            if reg:
                result[key].append(reg)

    # H01_MATCH 시트 (있으면)
    if 'H01_MATCH' in wb.sheetnames:
        ws = wb['H01_MATCH']
        header = None
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            if not header:
                if cells[0] in ('No', 'No.'):
                    header = cells
                continue
            if cells[0] and cells[0].isdigit():
                result['h01_match'].append({
                    'field': cells[1], 'source': cells[2] if len(cells) > 2 else '',
                    'status': cells[3] if len(cells) > 3 else '',
                    'address': cells[4] if len(cells) > 4 else '',
                    'definition': cells[5] if len(cells) > 5 else '',
                    'note': cells[6] if len(cells) > 6 else '',
                })

    return result


def _parse_section_row(cells: list, header: list, category: str) -> Optional[RegisterRow]:
    """섹션 행 → RegisterRow"""
    col_map = {}
    for i, h in enumerate(header):
        hl = h.lower()
        if h == 'No' or h == 'No.':
            continue
        elif 'definition' in hl or 'name' in hl:
            col_map['definition'] = i
        elif 'address' in hl:
            col_map['address'] = i
        elif hl == 'fc':
            # Modbus Function Code (3=Holding, 4=Input)
            col_map['fc'] = i
        elif 'type' in hl and 'data' not in hl:
            col_map['type'] = i
        elif 'unit' in hl or 'scale' in hl:
            col_map['scale'] = i
        elif 'r/w' in hl or 'rw' in hl:
            col_map['rw'] = i
        elif 'h01' in hl:
            col_map['h01_field'] = i
        elif 'comment' in hl or '비고' in hl:
            col_map['comment'] = i
        elif '사유' in hl or 'reason' in hl:
            col_map['review_reason'] = i
        elif '제안' in hl or 'suggest' in hl:
            col_map['review_suggestion'] = i

    defn = cells[col_map['definition']] if 'definition' in col_map else ''
    if not defn:
        return None

    addr_raw = cells[col_map['address']] if 'address' in col_map else ''
    addr = parse_address(addr_raw)
    fc_val = cells[col_map['fc']] if 'fc' in col_map and col_map['fc'] < len(cells) else ''

    return RegisterRow(
        definition=defn,
        address=addr if addr is not None else 0,
        address_hex=addr_raw if addr_raw.startswith('0x') else (f'0x{addr:04X}' if addr else ''),
        data_type=cells[col_map.get('type', 99)] if col_map.get('type', 99) < len(cells) else 'U16',
        scale=cells[col_map.get('scale', 99)] if col_map.get('scale', 99) < len(cells) else '',
        rw=cells[col_map.get('rw', 99)] if col_map.get('rw', 99) < len(cells) else 'RO',
        fc=fc_val,
        h01_field=cells[col_map.get('h01_field', 99)] if col_map.get('h01_field', 99) < len(cells) else '',
        comment=cells[col_map.get('comment', 99)] if col_map.get('comment', 99) < len(cells) else '',
        category=category,
        review_reason=cells[col_map.get('review_reason', 99)] if col_map.get('review_reason', 99) < len(cells) else '',
        review_suggestion=cells[col_map.get('review_suggestion', 99)] if col_map.get('review_suggestion', 99) < len(cells) else '',
    )


# ─── REVIEW 추천 생성 ──────────────────────────────────────────────────────

def generate_review_alternatives(
    review_regs: List[RegisterRow],
    synonym_db: dict,
    ref_patterns: dict,
    manufacturer: str,
) -> List[dict]:
    """
    REVIEW 항목에 대해 추천 대안 매핑 생성
    Returns: [{
        'reg': RegisterRow,
        'alt1': {'definition': str, 'address': str, 'category': str, 'reason': str} or None,
        'alt2': {'definition': str, 'address': str, 'category': str, 'reason': str} or None,
    }]
    """
    mfr_lower = manufacturer.lower()
    mfr_ref = {k: v for k, v in ref_patterns.items() if mfr_lower in k.lower()}

    results = []
    for reg in review_regs:
        alts = []
        addr = reg.address if isinstance(reg.address, int) else parse_address(reg.address)

        # 1) synonym_db 퍼지 매칭 (Phase B: 0.4 → 0.75 false positive 방지)
        fuzzy = match_synonym_fuzzy(reg.definition, synonym_db, threshold=0.75)
        if fuzzy:
            alts.append({
                'definition': fuzzy['field'],
                'address': reg.address_hex,
                'category': fuzzy['category'],
                'reason': f'synonym 유사도 {fuzzy["score"]:.0%}',
            })

        # 2) 주소 근접 레퍼런스 (같은 제조사, ±10 주소)
        if addr is not None and mfr_ref:
            for proto, addr_map in mfr_ref.items():
                for ref_addr, ref_name in addr_map.items():
                    if abs(ref_addr - addr) <= 10 and ref_addr != addr:
                        # 이 ref_name의 카테고리 추측
                        ref_cat = _guess_category(ref_name)
                        alts.append({
                            'definition': ref_name,
                            'address': f'0x{ref_addr:04X}',
                            'category': ref_cat,
                            'reason': f'주소 근접 ({proto}, 0x{ref_addr:04X})',
                        })
                        break
                if len(alts) >= 2:
                    break

        # 3) 키워드 기반 카테고리 추측 (alts가 부족하면)
        if len(alts) < 2:
            guessed_cat = _guess_category(reg.definition)
            if guessed_cat and guessed_cat != 'REVIEW':
                alts.append({
                    'definition': reg.definition,
                    'address': reg.address_hex,
                    'category': guessed_cat,
                    'reason': f'키워드 추측 → {guessed_cat}',
                })

        results.append({
            'reg': reg,
            'alt1': alts[0] if len(alts) > 0 else None,
            'alt2': alts[1] if len(alts) > 1 else None,
        })

    return results


def _guess_category(definition: str) -> str:
    """키워드 기반 카테고리 추측"""
    dl = definition.lower()
    if any(k in dl for k in ['error', 'fault', 'alarm', 'warning']):
        return 'ALARM'
    if any(k in dl for k in ['status', 'state', 'mode', 'running']):
        return 'STATUS'
    if any(k in dl for k in ['model', 'serial', 'firmware', 'nominal', 'rated',
                              'version', 'mppt_count', 'phase']):
        return 'INFO'
    if any(k in dl for k in ['voltage', 'current', 'power', 'energy', 'frequency',
                              'temperature', 'factor', 'mppt', 'string', 'pv',
                              'grid', 'load', 'watt', 'fan', 'fuse']):
        return 'MONITORING'
    return 'MONITORING'


# H01 고정 필드 목록 (순서 고정) — (h01_field명, 설명)
_H01_MAPPING_FIELDS = [
    ('pv_voltage',        'HANDLER 계산 (평균 MPPT 전압)'),
    ('pv_current',        'HANDLER 계산 (합산 MPPT 전류)'),
    ('pv_power',          'HANDLER 계산 (합산 MPPT 전력)'),
    ('r_voltage',         'R상(L1) 전압'),
    ('s_voltage',         'S상(L2) 전압'),
    ('t_voltage',         'T상(L3) 전압'),
    ('r_current',         'R상(L1) 전류'),
    ('s_current',         'S상(L2) 전류'),
    ('t_current',         'T상(L3) 전류'),
    ('ac_power',          'AC 출력 전력'),
    ('power_factor',      '역률'),
    ('frequency',         '계통 주파수'),
    ('cumulative_energy', '누적 발전량'),
    ('status',            '인버터 상태'),
    ('alarm1',            '경보 코드 1'),
    ('alarm2',            '경보 코드 2'),
    ('alarm3',            '경보 코드 3'),
]


# H01 필드별 추천 후보 키워드 패턴 (소문자 매칭)
_H01_SUGGEST_PATTERNS: Dict[str, List[str]] = {
    'pv_voltage':        ['pv.*volt', 'dc.*volt', 'mppt.*volt'],
    'pv_current':        ['pv.*curr', 'dc.*curr', 'mppt.*curr', 'string.*curr'],
    'pv_power':          ['pv.*power', 'dc.*power', 'total.*dc'],
    'r_voltage':         ['l1.*volt', 'r.*phase.*volt', 'phase.*a.*volt', 'grid.*volt', 'ac.*volt'],
    's_voltage':         ['l2.*volt', 's.*phase.*volt', 'phase.*b.*volt'],
    't_voltage':         ['l3.*volt', 't.*phase.*volt', 'phase.*c.*volt'],
    'r_current':         ['l1.*curr', 'r.*phase.*curr', 'phase.*a.*curr', 'grid.*curr', 'ac.*curr'],
    's_current':         ['l2.*curr', 's.*phase.*curr', 'phase.*b.*curr'],
    't_current':         ['l3.*curr', 't.*phase.*curr', 'phase.*c.*curr'],
    'ac_power':          ['ac.*power', 'active.*power', 'output.*power', 'total.*active'],
    'power_factor':      ['power.*factor', 'pf$', 'cos.*phi'],
    'frequency':         ['frequen', 'freq$', 'grid.*freq'],
    'cumulative_energy': ['cumulative', 'total.*energy', 'lifetime.*energy', 'generation.*total'],
    'status':            ['inverter.*mode', 'inverter.*stat', 'device.*stat', 'operat.*mode'],
    'alarm1':            ['alarm', 'error.*code', 'fault.*code', 'warning'],
    'alarm2':            ['alarm', 'error.*code', 'fault.*code', 'warning'],
    'alarm3':            ['alarm', 'error.*code', 'fault.*code', 'warning'],
}


def _suggest_candidates(field: str, all_regs: list, max_n: int = 3) -> list:
    """H01 필드에 대해 키워드 매칭으로 추천 후보 반환. [{name, addr_hex}, ...]"""
    patterns = _H01_SUGGEST_PATTERNS.get(field, [])
    if not patterns:
        return []
    candidates = []
    seen = set()
    for reg in all_regs:
        name = to_upper_snake(reg.definition)
        if not name or name in seen:
            continue
        defn_lower = reg.definition.lower().replace(' ', '_')
        name_lower = name.lower()
        for pat in patterns:
            if re.search(pat, defn_lower) or re.search(pat, name_lower):
                candidates.append({
                    'name': name,
                    'addr_hex': getattr(reg, 'address_hex', '') or '',
                })
                seen.add(name)
                break
        if len(candidates) >= max_n:
            break
    return candidates


def _add_h01_mapping_sheet(wb, s1: dict, openpyxl_module) -> None:
    """H01_MAPPING 시트 추가 — H01 필드별 레지스터 수동 매핑 (주소+추천후보 포함)."""
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet('H01_MAPPING')

    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='333333')
    auto_fill  = PatternFill('solid', fgColor='D5F5E3')  # 연초록 — 자동 매칭
    empty_fill = PatternFill('solid', fgColor='FDECEA')  # 연빨강 — 미매칭
    ref_fill   = PatternFill('solid', fgColor='EBF5FB')  # 연파랑 — 참조 목록
    suggest_fill = PatternFill('solid', fgColor='FFF9C4')  # 연노랑 — 추천 후보

    # Row 1 제목
    ws['A1'] = 'H01_MAPPING — H01 필드별 레지스터 수동 매핑'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = ('※ D열에서 매핑할 레지스터명을 선택/입력하세요.'
                '  E~G열은 추천 후보, H열은 전체 목록, I열은 FC코드(3 or 4).')
    ws['A2'].font = Font(italic=True, color='555555')

    # Row 3 헤더
    headers = ['No', 'H01 Field', '설명', '매칭된 레지스터 (주소)',
               '추천 후보 1', '추천 후보 2', '추천 후보 3', '전체 레지스터 목록', 'FC']
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=j, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin

    # 자동 매칭 빌드: h01_match 테이블 우선 사용 (Stage1 H01 매칭 결과)
    h01_auto: Dict[str, List[Tuple[str, str]]] = {}
    h01_fc: Dict[str, str] = {}  # h01_field → FC ('3' or '4')
    all_cats = (s1.get('monitoring', []) + s1.get('info', []) +
                s1.get('status', []) + s1.get('alarm', []))

    # 1차: h01_match 테이블에서 매칭 정보 수집
    # Phase C: DER 가상 주소(DEA_*)는 사용자 드롭다운 H열에 없으므로 D열 auto-populate에서 제외
    # (Stage3의 alias generator가 자동 처리)
    for hm in s1.get('h01_match', []):
        field = hm.get('field', '').strip()
        status = hm.get('status', '')
        if not field or status not in ('O', '~'):
            continue
        source = hm.get('source', '')
        if source == 'DER':
            continue  # DEA_* 가상 주소 건너뛰기
        defn = hm.get('definition', '') or ''
        addr = hm.get('address', '') or ''
        name = to_upper_snake(defn) if defn and defn != '-' else ''
        if source == 'HANDLER':
            display_name = 'HANDLER'
            display_addr = ''
        else:
            display_name = name or defn
            display_addr = addr
        if display_name:
            entries = h01_auto.setdefault(field, [])
            if not any(e[0] == display_name for e in entries):
                entries.append((display_name, display_addr))

    # 2차: reg.h01_field에서 추가 매칭 보완 + FC 수집
    for reg in all_cats:
        h01 = (getattr(reg, 'h01_field', '') or '').strip()
        if h01 and h01 in {f for f, _ in _H01_MAPPING_FIELDS}:
            name = to_upper_snake(reg.definition)
            addr = getattr(reg, 'address_hex', '') or ''
            fc = str(getattr(reg, 'fc', '') or '').strip()
            if name:
                entries = h01_auto.setdefault(h01, [])
                if not any(e[0] == name for e in entries):
                    entries.append((name, addr))
            if fc and h01 not in h01_fc:
                h01_fc[h01] = fc

    # 전체 레지스터 목록 (H열, 드롭다운용) — 이름만, 알파벳 정렬
    # 동시에 name → {fc} 맵 수집 (FC3/FC4 구분용)
    def _norm_fc(v) -> str:
        s = str(v or '').strip()
        if not s:
            return ''
        if s in ('3', '03', 'FC03', 'fc03', 'Fc03'):
            return '3'
        if s in ('4', '04', 'FC04', 'fc04', 'Fc04'):
            return '4'
        return ''

    name_to_fcs: Dict[str, set] = {}
    # (name, fc) → address (Hex 문자열) 매핑 — 같은 이름이 FC3/FC4 양쪽에 있을 때
    # 사용자가 어느 FC인지 검증할 수 있도록 주소도 함께 노출
    name_fc_to_addr: Dict[tuple, str] = {}
    for reg in all_cats:
        name = to_upper_snake(reg.definition)
        if not name:
            continue
        fc_n = _norm_fc(getattr(reg, 'fc', ''))
        addr_hex = getattr(reg, 'address_hex', '') or ''
        if fc_n:
            name_to_fcs.setdefault(name, set()).add(fc_n)
            if (name, fc_n) not in name_fc_to_addr and addr_hex:
                name_fc_to_addr[(name, fc_n)] = addr_hex
        else:
            name_to_fcs.setdefault(name, set())
    avail_entries: list = sorted(name_to_fcs.keys())

    DATA_START = 4
    n_fields = len(_H01_MAPPING_FIELDS)

    # Phase C: Stage3와 동일한 semantic validator 공유
    try:
        from .stage1 import _h01_semantic_valid
    except ImportError:
        def _h01_semantic_valid(field, name):
            return True

    # 데이터 행
    for i, (field, desc) in enumerate(_H01_MAPPING_FIELDS):
        row = DATA_START + i
        matched_list = h01_auto.get(field, [])

        # Phase C: semantic 검증을 통과한 후보만 D열에 auto-populate
        # HANDLER/DEA_* 는 semantic valid 통과 (Stage3 에서 자동 alias 처리)
        valid_matched = [
            (name, addr) for name, addr in matched_list
            if _h01_semantic_valid(field, name)
        ]

        # Fallback: h01_match 에 없거나 HANDLER 만 있으면 all_cats 에서
        # _suggest_candidates 로 최고 점수 후보를 자동 채택 (top 1)
        # HANDLER 뿐이면 실제 PDF 레지스터도 함께 표시
        if (not valid_matched) or all(n == 'HANDLER' for n, _ in valid_matched):
            candidates = _suggest_candidates(field, all_cats)
            auto_candidates = [
                c for c in candidates
                if _h01_semantic_valid(field, c['name'])
            ]
            if auto_candidates:
                # 드롭다운 H 열에 존재하는 것만 자동 채택 (실제 추출 레지스터)
                for ac in auto_candidates[:1]:  # 최고 점수 1개만 auto
                    nm = ac['name']
                    if nm in name_to_fcs and not any(e[0] == nm for e in valid_matched):
                        valid_matched.append((nm, ac.get('addr_hex', '')))

        # D열: semantic-valid 매칭만 표시 (세미콜론 구분)
        if valid_matched:
            parts = []
            for name, addr in valid_matched:
                parts.append(f'{name} ({addr})' if addr else name)
            matched_display = '; '.join(parts)
        else:
            matched_display = ''

        ws.cell(row=row, column=1, value=i + 1).border = thin
        ws.cell(row=row, column=2, value=field).border = thin
        ws.cell(row=row, column=3, value=desc).border = thin

        cell_d = ws.cell(row=row, column=4, value=matched_display)
        cell_d.border = thin
        cell_d.fill = auto_fill if valid_matched else empty_fill

        # 추천 후보 (E~G열) — 이미 매칭된 것 제외 + semantic 검증 통과만
        matched_names = {e[0] for e in matched_list} | {n for n, _ in valid_matched}
        candidates = _suggest_candidates(field, all_cats)
        filtered = [
            c for c in candidates
            if c['name'] not in matched_names
            and _h01_semantic_valid(field, c['name'])
        ]
        for ci, cand in enumerate(filtered[:3]):
            val = f'{cand["name"]} ({cand["addr_hex"]})' if cand['addr_hex'] else cand['name']
            cell = ws.cell(row=row, column=5 + ci, value=val)
            cell.border = thin
            cell.fill = suggest_fill

    # H열: 전체 레지스터 목록 (드롭다운 소스)
    # J/K/L열(숨김): 레지스터 → FC + 주소 맵 — API에서 읽음
    #   J = name
    #   K = "3" / "4" / "3,4"
    #   L = "0x0091" (FC3 주소) / "|0x0091" (FC4 주소) / "0x0091|0x0091" (양쪽)
    #       — fcs 리스트 순서와 동일하게 '|' 구분
    for i, name in enumerate(avail_entries):
        ws.cell(row=DATA_START + i, column=8, value=name).fill = ref_fill
        ws.cell(row=DATA_START + i, column=10, value=name)
        fcs = sorted(name_to_fcs.get(name, set()))
        ws.cell(row=DATA_START + i, column=11, value=','.join(fcs) if fcs else '')
        addr_parts = [name_fc_to_addr.get((name, fc), '') for fc in fcs]
        ws.cell(row=DATA_START + i, column=12, value='|'.join(addr_parts) if addr_parts else '')

    # I열: FC 코드 (기본값 자동, 사용자 수정 가능)
    # 기본 FC 감지 (MONITORING 다수파)
    _fc_vals = [str(getattr(r, 'fc', '')).strip() for r in all_cats
                if str(getattr(r, 'fc', '')).strip()]
    if _fc_vals and all(v in ('4', '04', 'FC04') for v in _fc_vals):
        default_fc = '4'
    else:
        default_fc = '3'

    # 매칭된 레지스터의 FC를 우선 사용 (단일 FC만 갖는 레지스터면 자동확정)
    for i, (field, desc) in enumerate(_H01_MAPPING_FIELDS):
        row = DATA_START + i
        fc_val = h01_fc.get(field, '')
        if not fc_val:
            # D열의 첫 매칭 레지스터로부터 FC 자동 추론
            matched_list = h01_auto.get(field, [])
            for nm, _addr in matched_list:
                fcs = name_to_fcs.get(nm, set())
                if len(fcs) == 1:
                    fc_val = next(iter(fcs))
                    break
        if not fc_val:
            fc_val = default_fc
        cell_fc = ws.cell(row=row, column=9, value=int(fc_val))
        cell_fc.border = thin

    # DataValidation — D열에 H열 범위 드롭다운
    last_avail_row = DATA_START + max(len(avail_entries), n_fields) - 1
    last_data_row  = DATA_START + n_fields - 1

    if avail_entries:
        dv = DataValidation(
            type='list',
            formula1=f'$H${DATA_START}:$H${last_avail_row}',
            allow_blank=True,
            showDropDown=False,
        )
        dv.error      = '목록에 없는 값입니다. H열 전체 레지스터 목록을 확인하세요.'
        dv.errorTitle = '잘못된 입력'
        dv.prompt     = '드롭다운에서 선택하거나 직접 입력하세요.'
        dv.promptTitle = '레지스터 선택'
        dv.sqref = f'D{DATA_START}:D{last_data_row}'
        ws.add_data_validation(dv)

    # DataValidation — I열 FC 드롭다운 (3 or 4)
    dv_fc = DataValidation(type='list', formula1='"3,4"', allow_blank=False, showDropDown=False)
    dv_fc.prompt = 'FC03=Holding, FC04=Input'
    dv_fc.promptTitle = 'Function Code'
    dv_fc.sqref = f'I{DATA_START}:I{last_data_row}'
    ws.add_data_validation(dv_fc)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 5
    # J/K/L: 숨김 메타데이터 (FC + 주소 맵)
    ws.column_dimensions['J'].hidden = True
    ws.column_dimensions['K'].hidden = True
    ws.column_dimensions['L'].hidden = True


# ─── Stage 2 메인 함수 ───────────────────────────────────────────────────────

def run_stage2(
    stage1_path: str,
    output_dir: str,
    mppt_count: int,
    total_strings: int,
    capacity: str = '',
    progress: ProgressCallback = None,
) -> dict:
    """Stage 2: Stage 1 Excel + 파라미터 → Stage 2 Excel (3시트)

    total_strings: 인버터 전체의 총 String 개수 (MPPT 수와 독립).
                   Growatt 처럼 MPPT × N ≠ Total String 인 비대칭 인버터 지원.
                   균등 분배가 가능하면 strings_per_mppt = total // mppt 로 파생.
    """
    def log(msg, level='info'):
        if progress:
            progress(msg, level)

    openpyxl = get_openpyxl()
    from openpyxl.styles import Font, PatternFill, Border, Side

    # MPPT당 String 수는 균등 분배 가정으로 파생 (코드 생성용 — 비대칭이면 0 가능)
    strings_per_mppt = (total_strings // mppt_count) if mppt_count > 0 else 0
    log(f'MPPT: {mppt_count}, Total Strings: {total_strings} '
        f'({strings_per_mppt}/MPPT 균등분배 가정)')

    # ── Step 1: Stage 1 Excel 읽기 (V2 4시트) ──
    log('Stage 1 Excel 읽기...')
    s1 = read_stage1_excel_v2(stage1_path)
    meta = s1['meta']
    device_type = meta.get('device_type', 'inverter')
    manufacturer = meta.get('manufacturer', 'Unknown')

    # ── Step 2: REVIEW 자동 처리 ──
    review_history = load_review_history()
    review_regs = s1['review']
    remaining_review = []
    auto_processed = 0

    for reg in review_regs:
        auto_applied = False
        for item in review_history.get('approved', []):
            defn_match = item.get('definition', '').upper() == to_upper_snake(reg.definition)
            addr_match = item.get('address') == reg.address_hex
            if defn_match or addr_match:
                verdict = item.get('verdict', '')
                if verdict == 'DELETE':
                    auto_processed += 1
                    auto_applied = True
                    break
                elif verdict.startswith('MOVE:'):
                    target_cat = verdict.replace('MOVE:', '')
                    reg.category = target_cat
                    target_list = {'INFO': s1['info'], 'MONITORING': s1['monitoring'],
                                   'STATUS': s1['status'], 'ALARM': s1['alarm']}
                    if target_cat in target_list:
                        target_list[target_cat].append(reg)
                    auto_processed += 1
                    auto_applied = True
                    break
        if not auto_applied:
            remaining_review.append(reg)

    if auto_processed:
        log(f'  REVIEW 자동 처리: {auto_processed}개')
    log(f'  REVIEW 미결: {len(remaining_review)}개')

    # ── Step 3: MPPT/String 필터링 ──
    log('MPPT/String 필터링...')
    ref_patterns = load_reference_patterns()

    mon_before = len(s1['monitoring'])
    s1['monitoring'] = filter_channels_stage2(
        s1['monitoring'], mppt_count, total_strings, ref_patterns)
    mon_after = len(s1['monitoring'])
    if mon_before != mon_after:
        log(f'  MONITORING: {mon_before} → {mon_after}')

    # ── Step 4: REVIEW 추천 생성 ──
    synonym_db = load_synonym_db()
    review_alts = []
    if remaining_review:
        log('REVIEW 추천 생성...')
        review_alts = generate_review_alternatives(
            remaining_review, synonym_db, ref_patterns, manufacturer)
        alts_with_rec = sum(1 for a in review_alts if a['alt1'])
        log(f'  추천 있는 항목: {alts_with_rec}/{len(review_alts)}')

    # ── Step 5: 통계 ──
    counts = {
        'INFO': len(s1['info']),
        'MONITORING': len(s1['monitoring']),
        'STATUS': len(s1['status']),
        'ALARM': len(s1['alarm']),
        'REVIEW': len(remaining_review),
    }
    total_regs = sum(counts.values()) - counts['REVIEW']

    # H01 매칭 (Stage 1에서 가져옴)
    # status '-' = DER 겹침 필드 → 분모 제외 (PDF 매칭 대상 아님)
    # status 'O' 또는 '~' = 매칭 성공 (HANDLER 계산 포함)
    h01_matched = sum(1 for h in s1['h01_match'] if h['status'] in ('O', '~'))
    h01_total   = sum(1 for h in s1['h01_match'] if h['status'] != '-')

    # DER 매칭 (고정)
    der_total = len(DER_CONTROL_REGS) + len(DER_MONITOR_REGS)
    der_matched = der_total

    log(f'총 레지스터: {total_regs}개, H01: {h01_matched}/{h01_total}, DER: {der_matched}/{der_total}')

    # ── Stage 2 검증: MPPT 전압/전류 + String 전압/전류 추출 여부 ──
    # 인버터 타입:
    #   String type   : MPPT 1~N개, 각각 V/I 있음. 일부는 String 전류도 있음.
    #   Central type  : MPPT 1개. V는 PV_VOLTAGE, I는 PV1_CURRENT/PV2_CURRENT 등
    #                   "PV (n) 전류" 형식으로 표시될 수 있음.
    # 매핑 의미: MPPT/String V/I → H01 의 pv_voltage/pv_current/pv_power
    #           (pv_power 는 직접 또는 V*I 계산)
    import re as _re
    _H01_MPPT_V_RE = _re.compile(r'^mppt(\d+)_voltage$', _re.I)
    _H01_MPPT_I_RE = _re.compile(r'^mppt(\d+)_current$', _re.I)
    _H01_STR_V_RE  = _re.compile(r'^string(\d+)_voltage$', _re.I)
    _H01_STR_I_RE  = _re.compile(r'^string(\d+)_current$', _re.I)
    # central type fallback names
    _CENTRAL_PV_V = {'PV_VOLTAGE', 'PV_INPUT_VOLTAGE', 'DC_VOLTAGE', 'DC_INPUT_VOLTAGE',
                     'BUS_VOLTAGE', 'DC_BUS_VOLTAGE', 'VPV', 'VDC', 'V_DC', 'V_PV'}
    _CENTRAL_PV_I = {'PV_CURRENT', 'PV_INPUT_CURRENT', 'DC_CURRENT', 'DC_INPUT_CURRENT',
                     '태양전지_전류', 'IPV', 'IDC', 'I_DC', 'I_PV'}

    mppt_v_found = set()
    mppt_i_found = set()
    str_v_found = set()
    str_i_found = set()
    central_pv_v_found = False
    central_pv_i_count = set()  # PV(n) 전류 개수

    for reg in s1['monitoring'] + s1['info']:
        h01 = (getattr(reg, 'h01_field', '') or '').strip().lower()
        # h01_field 기반 (가장 신뢰 높음)
        if h01:
            m = _H01_MPPT_V_RE.match(h01)
            if m:
                mppt_v_found.add(int(m.group(1))); continue
            m = _H01_MPPT_I_RE.match(h01)
            if m:
                mppt_i_found.add(int(m.group(1))); continue
            m = _H01_STR_V_RE.match(h01)
            if m:
                str_v_found.add(int(m.group(1))); continue
            m = _H01_STR_I_RE.match(h01)
            if m:
                str_i_found.add(int(m.group(1))); continue
        # 이름 기반 보조 매칭
        name = to_upper_snake(reg.definition) if reg.definition else ''
        if not name:
            continue
        # 채널 번호 기반 (MPPTn/PVn/REGn 등)
        cm = _re.match(r'^(?:MPPT|PV|REG|MOD|MODULE|VPV)_?(\d+)_?(?:INPUT_?)?(VOLTAGE|VOLT|DCV)$', name)
        if cm:
            mppt_v_found.add(int(cm.group(1))); continue
        cm = _re.match(r'^(?:MPPT|PV|REG|MOD|MODULE|IPV)_?(\d+)_?(?:INPUT_?)?(CURRENT|CURR|DCA)$', name)
        if cm:
            mppt_i_found.add(int(cm.group(1)))
            central_pv_i_count.add(int(cm.group(1)))  # central type "PV(n)"도 같이 카운트
            continue
        # STRING n V/I
        sm = _re.match(r'^STRING_?(\d+)_?(?:INPUT_?)?(VOLTAGE|VOLT|CURRENT|CURR)$', name)
        if sm:
            idx = int(sm.group(1))
            if 'VOLT' in sm.group(2):
                str_v_found.add(idx)
            else:
                str_i_found.add(idx)
            continue
        # Central type fallback (PV_VOLTAGE 단독, 채널 번호 없음)
        if name in _CENTRAL_PV_V:
            central_pv_v_found = True
        if name in _CENTRAL_PV_I:
            mppt_i_found.add(1)  # 단일 PV 전류 → MPPT1로
        # 한글 "태양전지N_전류"
        km = _re.match(r'태양전지\s*(\d+)\s*_?\s*(전압|전류)', reg.definition or '')
        if km:
            idx = int(km.group(1))
            if km.group(2) == '전압':
                mppt_v_found.add(idx)
            else:
                mppt_i_found.add(idx)

    # Central type 보정: MPPT V는 없는데 PV_VOLTAGE 단독이면 mppt1_voltage로 카운트
    if not mppt_v_found and central_pv_v_found:
        mppt_v_found.add(1)

    # 검증 카운트
    stage2_validation = {
        'mppt_voltage': {
            'found': len([i for i in mppt_v_found if 1 <= i <= mppt_count]),
            'expected': mppt_count,
        },
        'mppt_current': {
            'found': len([i for i in mppt_i_found if 1 <= i <= mppt_count]),
            'expected': mppt_count,
        },
        'string_voltage': {
            'found': len([i for i in str_v_found if 1 <= i <= total_strings]),
            'expected': total_strings,
        },
        'string_current': {
            'found': len([i for i in str_i_found if 1 <= i <= total_strings]),
            'expected': total_strings,
        },
    }

    # ── IV Scan 검증: 지원 인버터에 한해 IV 명령/결과 레지스터 ──
    # meta['iv_scan'] 은 'Yes'/'No'/'-' 문자열 또는 bool 일 수 있음
    _iv_raw = meta.get('iv_scan', False)
    if isinstance(_iv_raw, bool):
        iv_supported = _iv_raw
    else:
        iv_supported = str(_iv_raw).strip().lower() in ('yes', 'true', '1', 'y')
    iv_data_points = 0
    try:
        iv_data_points = int(meta.get('iv_data_points', 0) or 0)
    except (ValueError, TypeError):
        pass
    iv_trackers = 0
    try:
        iv_trackers = int(meta.get('iv_trackers', 0) or 0)
    except (ValueError, TypeError):
        pass

    iv_command_found = False
    iv_result_found = 0
    if iv_supported:
        # IV 명령 레지스터 (IV_CURVE_SCAN, IV_SCAN_COMMAND, 0x600D 등)
        for reg in s1.get('iv_regs', []) + s1['monitoring']:
            n = (to_upper_snake(reg.definition) if reg.definition else '').upper()
            if n in ('IV_CURVE_SCAN', 'IV_SCAN_COMMAND', 'IV_SCAN_CONTROL',
                     'IV_SCAN_START', 'IV_COMMAND'):
                iv_command_found = True
                break
            # 주소 0x600D (Solarize 표준)
            try:
                addr = reg.address if isinstance(reg.address, int) else None
                if addr in (0x600D, 0x600E):
                    iv_command_found = True
                    break
            except Exception:
                pass

        # IV 결과 레지스터 (tracker별 voltage 블록 + string current 블록)
        # 5_IV 시트에서 읽은 모든 데이터 레지스터를 카운트
        # 단, IV_CURVE_SCAN/IV_SCAN_COMMAND 같은 명령 레지스터는 제외
        _IV_CMD_NAMES = {'IV_CURVE_SCAN', 'IV_SCAN_COMMAND', 'IV_SCAN_CONTROL',
                         'IV_SCAN_START', 'IV_COMMAND'}
        for reg in s1.get('iv_regs', []):
            n = (to_upper_snake(reg.definition) if reg.definition else '').upper()
            if not n or n in _IV_CMD_NAMES:
                continue
            iv_result_found += 1

    stage2_validation['iv_scan'] = {
        'supported': iv_supported,
        'command_found': iv_command_found,
        'result_count': iv_result_found,
        'expected_trackers': iv_trackers,
        'data_points': iv_data_points,
    }

    log(f'MPPT V: {stage2_validation["mppt_voltage"]["found"]}/{mppt_count}, '
        f'MPPT I: {stage2_validation["mppt_current"]["found"]}/{mppt_count}, '
        f'String V: {stage2_validation["string_voltage"]["found"]}/{total_strings}, '
        f'String I: {stage2_validation["string_current"]["found"]}/{total_strings}')
    if iv_supported:
        log(f'IV Scan: cmd {"✓" if iv_command_found else "✗"}, '
            f'result regs {iv_result_found}, trackers {iv_trackers}, '
            f'points {iv_data_points}')

    # ── Step 6: Excel 생성 (3시트) ──
    basename = os.path.splitext(os.path.basename(stage1_path))[0].replace('_stage1', '')
    cap_str = f'_{capacity}' if capacity else ''
    output_name = f'{basename}{cap_str}_MPPT{mppt_count}_STR{total_strings}_stage2.xlsx'
    output_path = os.path.join(output_dir, output_name)
    log(f'Excel 생성: {output_name}')

    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='333333')
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color='1F4E79')

    def _write_header(ws, columns, row=1):
        for j, cn in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=j, value=cn)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = thin_border

    def _write_regs(ws, regs, start_row, cols, cat_color=None):
        """레지스터 목록 기록, 다음 빈 행 반환"""
        sorted_regs = sorted(regs, key=lambda r: (r.address if isinstance(r.address, int) else 0))
        for i, reg in enumerate(sorted_regs, start=1):
            su = f'{reg.unit} {reg.scale}'.strip() if (reg.unit or reg.scale) else ''
            vals = [i, reg.definition, reg.address_hex, reg.data_type, su,
                    reg.rw, reg.h01_field or '', reg.comment]
            for j, val in enumerate(vals[:len(cols)], start=1):
                cell = ws.cell(row=start_row + i, column=j, value=val)
                cell.border = thin_border
                if cat_color:
                    cell.fill = PatternFill('solid', fgColor=cat_color)
        return start_row + len(sorted_regs) + 1

    reg_cols = ['No', 'Definition', 'Address', 'Type', 'Unit/Scale', 'R/W', 'H01 Field', 'Comment']

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 1: REGISTER_MAP — 확정 레지스터맵
    # ═══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = '1_REGISTER_MAP'

    # 메타
    ws['A1'] = f'Stage 2 — 레지스터맵 (MPPT={mppt_count}, STR={total_strings})'
    ws['A1'].font = title_font
    meta_items = [
        ('제조사', manufacturer), ('용량', capacity or '-'),
        ('MPPT', mppt_count),
        ('Total Strings', total_strings),
        ('String/MPPT (파생)', strings_per_mppt),
        ('H01 매칭', f'{h01_matched}/{h01_total}'),
        ('DER 매칭', f'{der_matched}/{der_total}'),
        ('총 레지스터', total_regs), ('REVIEW', len(remaining_review)),
    ]
    for i, (k, v) in enumerate(meta_items, start=3):
        ws[f'A{i}'] = k
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = str(v)

    row_n = len(meta_items) + 5

    # INFO 섹션
    ws.cell(row=row_n, column=1, value='INFO 레지스터').font = section_font
    _write_header(ws, reg_cols, row_n + 1)
    row_n = _write_regs(ws, s1['info'], row_n + 1, reg_cols, CATEGORY_COLORS['INFO'])
    row_n += 2

    # H01 MONITORING 섹션
    ws.cell(row=row_n, column=1, value=f'H01 MONITORING ({len(s1["monitoring"])}개)').font = section_font
    _write_header(ws, reg_cols, row_n + 1)
    row_n = _write_regs(ws, s1['monitoring'], row_n + 1, reg_cols, CATEGORY_COLORS['MONITORING'])
    row_n += 2

    # STATUS 섹션
    ws.cell(row=row_n, column=1, value='STATUS 레지스터').font = section_font
    _write_header(ws, reg_cols, row_n + 1)
    row_n = _write_regs(ws, s1['status'], row_n + 1, reg_cols, CATEGORY_COLORS['STATUS'])
    row_n += 2

    # ALARM 섹션
    ws.cell(row=row_n, column=1, value='ALARM 레지스터').font = section_font
    _write_header(ws, reg_cols, row_n + 1)
    row_n = _write_regs(ws, s1['alarm'], row_n + 1, reg_cols, CATEGORY_COLORS['ALARM'])

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 35

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 2: REVIEW — 미결 항목 + 추천 대안
    # ═══════════════════════════════════════════════════════════════════
    ws_rv = wb.create_sheet('2_REVIEW')
    ws_rv['A1'] = f'REVIEW — {len(remaining_review)}개 미결'
    ws_rv['A1'].font = title_font

    if remaining_review:
        rv_cols = ['No', 'Definition', 'Address', 'Type', 'R/W', '사유',
                   '추천1', '추천1_카테고리', '추천1_사유',
                   '추천2', '추천2_카테고리', '추천2_사유',
                   '선택']
        _write_header(ws_rv, rv_cols, 3)

        alt1_fill = PatternFill('solid', fgColor='D5F5E3')  # 연한 초록
        alt2_fill = PatternFill('solid', fgColor='D6EAF8')  # 연한 파랑
        review_fill = PatternFill('solid', fgColor=CATEGORY_COLORS['REVIEW'])

        for i, alt_item in enumerate(review_alts, start=1):
            reg = alt_item['reg']
            a1 = alt_item['alt1']
            a2 = alt_item['alt2']

            vals = [
                i, reg.definition, reg.address_hex, reg.data_type, reg.rw,
                reg.review_reason,
                a1['definition'] if a1 else '-',
                a1['category'] if a1 else '-',
                a1['reason'] if a1 else '-',
                a2['definition'] if a2 else '-',
                a2['category'] if a2 else '-',
                a2['reason'] if a2 else '-',
                '',  # 사용자 선택
            ]
            for j, val in enumerate(vals, start=1):
                cell = ws_rv.cell(row=3 + i, column=j, value=val)
                cell.border = thin_border
                if j <= 6:
                    cell.fill = review_fill
                elif j <= 9:
                    cell.fill = alt1_fill
                elif j <= 12:
                    cell.fill = alt2_fill

        ws_rv.column_dimensions['B'].width = 40
        ws_rv.column_dimensions['C'].width = 12
        ws_rv.column_dimensions['F'].width = 35
        ws_rv.column_dimensions['G'].width = 30
        ws_rv.column_dimensions['H'].width = 15
        ws_rv.column_dimensions['I'].width = 25
        ws_rv.column_dimensions['J'].width = 30
        ws_rv.column_dimensions['K'].width = 15
        ws_rv.column_dimensions['L'].width = 25
        ws_rv.column_dimensions['M'].width = 15
    else:
        ws_rv.cell(row=3, column=1, value='REVIEW 항목 없음 — 모든 레지스터 자동 분류 완료')
        ws_rv['A3'].font = Font(bold=True, color='228B22')

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 3: SUMMARY — 통계
    # ═══════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet('3_SUMMARY')
    ws_sum['A1'] = 'Stage 2 Summary'
    ws_sum['A1'].font = title_font

    summary_items = [
        ('제조사', manufacturer), ('용량', capacity or '-'),
        ('MPPT', f'{mppt_count} (PDF max {meta.get("max_mppt", "?")}개)'),
        ('String', f'총 {total_strings}개 (파생 {strings_per_mppt}/MPPT)'),
        ('H01 매칭', f'{h01_matched}/{h01_total}'),
        ('DER 매칭', f'{der_matched}/{der_total}'),
        ('총 레지스터', total_regs),
        ('REVIEW 미결', len(remaining_review)),
        ('REVIEW 자동 처리', auto_processed),
    ]
    for i, (k, v) in enumerate(summary_items, start=3):
        ws_sum[f'A{i}'] = k
        ws_sum[f'A{i}'].font = Font(bold=True)
        ws_sum[f'B{i}'] = str(v)

    # 카테고리별 수량
    cat_start = len(summary_items) + 5
    ws_sum.cell(row=cat_start, column=1, value='카테고리별 수량').font = Font(bold=True, size=12)
    for i, (cat, cnt) in enumerate(counts.items(), start=1):
        if cnt > 0:
            ws_sum.cell(row=cat_start + i, column=1, value=cat).fill = PatternFill(
                'solid', fgColor=CATEGORY_COLORS.get(cat, 'FFFFFF'))
            ws_sum.cell(row=cat_start + i, column=2, value=cnt)

    # H01 매칭 테이블 복사
    h01_start = cat_start + len([c for c in counts.values() if c > 0]) + 3
    ws_sum.cell(row=h01_start, column=1, value='H01 매칭 상태').font = Font(bold=True, size=12)
    h01_cols = ['No', 'H01 Field', 'Source', 'Status', 'Address', 'Definition']
    _write_header(ws_sum, h01_cols, h01_start + 1)
    for i, hm in enumerate(s1['h01_match'], start=1):
        vals = [i, hm['field'], hm['source'], hm['status'], hm['address'], hm['definition']]
        sc = MATCH_COLORS.get(hm['status'], 'FFFFFF')
        for j, val in enumerate(vals, start=1):
            cell = ws_sum.cell(row=h01_start + 1 + i, column=j, value=val)
            cell.border = thin_border
            if j == 4:
                cell.fill = PatternFill('solid', fgColor=sc)

    ws_sum.column_dimensions['B'].width = 25
    ws_sum.column_dimensions['E'].width = 20
    ws_sum.column_dimensions['F'].width = 35

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 4: H01_MAPPING — H01 필드별 레지스터 수동 매핑
    # ═══════════════════════════════════════════════════════════════════
    _add_h01_mapping_sheet(wb, s1, openpyxl)

    wb.save(output_path)
    wb.close()
    log(f'Stage 2 완료: {output_name}', 'ok')

    # REVIEW items for web UI
    review_items = []
    for alt_item in review_alts:
        reg = alt_item['reg']
        review_items.append({
            'definition': reg.definition,
            'address': reg.address_hex,
            'type': reg.data_type,
            'rw': reg.rw,
            'reason': reg.review_reason,
            'alt1': alt_item['alt1'],
            'alt2': alt_item['alt2'],
        })

    return {
        'output_path': output_path,
        'output_name': output_name,
        'h01_matched': h01_matched,
        'h01_total': h01_total,
        'der_matched': der_matched,
        'der_total': der_total,
        'review_count': len(remaining_review),
        'review_items': review_items,
        'register_count': total_regs,
        'counts': counts,
        'stage2_validation': stage2_validation,
        'mppt_count': mppt_count,
        'total_strings': total_strings,
    }
