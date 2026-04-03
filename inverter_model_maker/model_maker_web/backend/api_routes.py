# -*- coding: utf-8 -*-
"""
API Routes - Model Maker Web v1.0.0

All REST API endpoints for the Model Maker web application.
Reuses stage_pipeline.py, ai_generator.py, reference_manager.py without modification.
"""

import os
import sys
import asyncio
import logging
import threading
import traceback
import json

from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
from typing import Optional

logger = logging.getLogger(__name__)

# Project root on sys.path (set by main.py)
router = APIRouter()

# Injected by main.py
ws_manager = None
store = None

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _project_root():
    """Return the real project root, resolving out of .claude/worktrees/ if needed."""
    raw = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    # If running inside a worktree (.claude/worktrees/<name>), resolve to actual project root
    marker = os.sep.join([".claude", "worktrees"])
    if marker in raw:
        idx = raw.index(marker)
        return raw[:idx].rstrip(os.sep)
    return raw


def _import_pipeline():
    """Import stage_pipeline from model_maker/."""
    root = _project_root()
    if root not in sys.path:
        sys.path.insert(0, root)
    from model_maker import stage_pipeline
    return stage_pipeline


def _import_mapper():
    """Import parse_modbus_pdf from model_maker/modbus_to_udp_mapper."""
    root = _project_root()
    if root not in sys.path:
        sys.path.insert(0, root)
    from model_maker.modbus_to_udp_mapper import parse_modbus_pdf
    return parse_modbus_pdf


def _import_ai():
    """Import ai_generator from model_maker/."""
    root = _project_root()
    if root not in sys.path:
        sys.path.insert(0, root)
    from model_maker import ai_generator
    return ai_generator


def _import_ref():
    """Import reference_manager from model_maker/."""
    root = _project_root()
    if root not in sys.path:
        sys.path.insert(0, root)
    from model_maker import reference_manager
    return reference_manager


def _get_session(session_id: str):
    s = store.get(session_id)
    if s is None:
        raise HTTPException(status_code=404, detail=f"Session not found: {session_id}")
    return s


def _excel_to_json_rows(excel_path: str, sheet_name: str) -> list:
    """Read an Excel sheet and return rows as list of dicts."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            result.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
        return result
    except Exception as e:
        logger.error(f"Excel read error ({sheet_name}): {e}")
        return []


def _run_in_thread(fn, *args, **kwargs):
    """Run a blocking function in a thread and return a Future."""
    loop = asyncio.get_event_loop()
    future = loop.run_in_executor(None, lambda: fn(*args, **kwargs))
    return future


# ---------------------------------------------------------------------------
# Session Management
# ---------------------------------------------------------------------------

@router.post("/api/session/create")
async def create_session():
    """Create a new session."""
    s = store.create()
    return {"session_id": s.session_id}


# ---------------------------------------------------------------------------
# PDF Upload
# ---------------------------------------------------------------------------

@router.post("/api/upload-pdf")
async def upload_pdf(
    file: UploadFile = File(...),
    session_id: str = Query(...)
):
    """Upload a PDF or Excel file. Returns session_id and filename."""
    ext = file.filename.lower().rsplit(".", 1)[-1] if "." in file.filename else ""
    if ext not in ("pdf", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="PDF 또는 Excel(.xlsx) 파일만 업로드 가능합니다.")

    s = store.get_or_create(session_id)
    session_dir = s.session_dir()

    # Save uploaded file
    safe_name = os.path.basename(file.filename)
    pdf_path = os.path.join(session_dir, safe_name)
    contents = await file.read()
    with open(pdf_path, "wb") as f:
        f.write(contents)

    s.pdf_path = pdf_path
    s.pdf_filename = safe_name
    # Reset downstream data
    s.stage1_excel_path = None
    s.stage2_excel_path = None
    s.stage1_rows = []
    s.stage2_rows = []
    s.stage3_code = ""
    s.stage3_results = []

    logger.info(f"PDF uploaded: {safe_name} ({len(contents)} bytes) session={session_id}")
    return {
        "session_id": s.session_id,
        "filename": safe_name,
        "size": len(contents),
    }


# ---------------------------------------------------------------------------
# Stage 1
# ---------------------------------------------------------------------------

class Stage1Request(BaseModel):
    session_id: str
    mode: str = "offline"  # "offline" or "ai"


@router.post("/api/stage1/run")
async def stage1_run(req: Stage1Request):
    """Run Stage 1: PDF/Excel → Excel. Progress streamed via WebSocket."""
    s = _get_session(req.session_id)
    if s.pdf_path is None or not os.path.isfile(s.pdf_path):
        raise HTTPException(status_code=400, detail="파일이 업로드되지 않았습니다.")
    if s.running:
        raise HTTPException(status_code=409, detail="이미 실행 중입니다.")

    s.running = True
    s.stage1_rows = []
    s.stage2_rows = []
    s.stage3_code = ""
    s.stage3_results = []

    is_excel = s.pdf_filename.lower().endswith(('.xlsx', '.xls'))

    async def _do():
        try:
            pipeline = _import_pipeline()
            loop = asyncio.get_event_loop()

            async def progress(msg):
                await ws_manager.send_progress(s.session_id, "stage1", msg)

            if is_excel:
                # ── Excel direct import: skip PDF parsing ──
                await progress(f"Excel 직접 로드: {s.pdf_filename}")
                import openpyxl
                session_dir = s.session_dir()
                base = os.path.splitext(s.pdf_filename)[0]
                stage1_path = os.path.join(session_dir, base + "_registers.xlsx")

                # Read the uploaded Excel, detect header row and columns,
                # then convert to Stage 1 standard format.
                src_wb = openpyxl.load_workbook(s.pdf_path, data_only=True)
                src_ws = None
                src_sheet_name = ""

                # Try "Registers" sheet first, then first sheet
                if "Registers" in src_wb.sheetnames:
                    src_ws = src_wb["Registers"]
                    src_sheet_name = "Registers"
                else:
                    src_ws = src_wb[src_wb.sheetnames[0]]
                    src_sheet_name = src_wb.sheetnames[0]

                await progress(f"시트 '{src_sheet_name}' 분석 중...")

                # Auto-detect header row: find the row that has the most non-None cells
                all_rows = list(src_ws.iter_rows(values_only=True))
                header_idx = 0
                max_filled = 0
                for i, row in enumerate(all_rows[:10]):  # scan first 10 rows
                    filled = sum(1 for v in row if v is not None)
                    if filled > max_filled:
                        max_filled = filled
                        header_idx = i

                raw_headers = [str(v).strip().lower() if v else "" for v in all_rows[header_idx]]
                await progress(f"헤더 감지 (row {header_idx + 1}): {[h for h in raw_headers if h]}")

                # Column mapping: source header keywords → Stage 1 standard columns
                # Order matters: first match wins per standard column
                _COL_MAP = {
                    'Address_Dec': ['address', 'reg.addr', 'addr', 'register address', 'reg addr'],
                    'Definition': ['parame', 'parameter', 'name', 'definition', 'signal name',
                                   'english name', 'description'],
                    'Data_Type': ['data type', 'data_type', 'data\ud615', 'type'],
                    'FC_Code': ['fc', 'fc_code', 'function code'],
                    'Registers': ['length', 'registers', 'regs', 'reg count', 'size', 'word'],
                    'Unit': ['unit', 'units'],
                    'Scale_Factor': ['scale', 'scale_factor', 'sf gain', 'gain', 'step'],
                    'R/W': ['r/w', 'rw', 'read/write', 'access', '\uc18d\uc131', 'property'],
                    'Description': ['range', 'scope', 'note', 'comment', 'remark'],
                    'Format': ['format'],
                }

                col_indices = {}  # Stage1_col -> source column index
                for std_col, keywords in _COL_MAP.items():
                    for i, h in enumerate(raw_headers):
                        if h and any(kw in h for kw in keywords):
                            if std_col not in col_indices:
                                col_indices[std_col] = i
                                break

                await progress(f"컬럼 매핑: {col_indices}")

                # Build Stage 1 standard output Excel
                out_wb = openpyxl.Workbook()
                out_ws = out_wb.active
                out_ws.title = "Registers"
                s1_headers = ['No', 'Section', 'Address_Hex', 'Address_Dec',
                              'Definition', 'Data_Type', 'FC_Code', 'Registers',
                              'Unit', 'Scale_Factor', 'R/W', 'Description']
                out_ws.append(s1_headers)

                reg_no = 0
                for row in all_rows[header_idx + 1:]:
                    if all(v is None for v in row):
                        continue

                    def _get(col_name):
                        idx = col_indices.get(col_name)
                        if idx is not None and idx < len(row):
                            return row[idx]
                        return None

                    addr_dec = _get('Address_Dec')
                    definition = _get('Definition')
                    if addr_dec is None and definition is None:
                        continue

                    reg_no += 1
                    # Convert address to hex
                    addr_hex = ""
                    if addr_dec is not None:
                        try:
                            addr_int = int(addr_dec)
                            addr_hex = f"0x{addr_int:04X}"
                        except (ValueError, TypeError):
                            pass

                    out_ws.append([
                        reg_no,
                        "",  # Section
                        addr_hex,
                        addr_dec,
                        str(definition or ""),
                        str(_get('Data_Type') or ""),
                        str(_get('FC_Code') or ""),
                        _get('Registers') or 1,
                        str(_get('Unit') or ""),
                        str(_get('Scale_Factor') or ""),
                        str(_get('R/W') or ""),
                        str(_get('Description') or ""),
                    ])

                out_wb.save(stage1_path)
                await progress(f"Stage 1 Excel 생성 완료: {reg_no}개 레지스터")

                s.stage1_excel_path = stage1_path
                s.stage1_rows = _excel_to_json_rows(stage1_path, "Registers")
                # Try to guess manufacturer from filename
                fname = s.pdf_filename.lower()
                _brand_keywords = {
                    'solarize': 'Solarize', 'goodwe': 'GoodWe', 'huawei': 'Huawei',
                    'kstar': 'Kstar', 'sungrow': 'Sungrow', 'ekos': 'EKOS', 'ek_': 'EKOS',
                    'senergy': 'Senergy',
                }
                s.detected_manufacturer = ''
                for kw, brand in _brand_keywords.items():
                    if kw in fname:
                        s.detected_manufacturer = brand
                        break

                await ws_manager.send_done(s.session_id, "stage1", True,
                    f"Stage 1 완료 (Excel 변환): {len(s.stage1_rows)}개 레지스터")
            else:
                # ── PDF parsing flow ──
                parse_modbus_pdf = _import_mapper()
                ai_gen = _import_ai()

                session_dir = s.session_dir()
                pdf_path = s.pdf_path

                await progress(f"PDF 파싱 시작: {s.pdf_filename}")

                def _parse():
                    try:
                        return parse_modbus_pdf(pdf_path), None
                    except Exception as e:
                        return None, str(e)

                parsed_data, err = await loop.run_in_executor(None, _parse)
                if err:
                    await ws_manager.send_error(s.session_id, "stage1", f"PDF 파싱 오류: {err}")
                    s.running = False
                    return

                reg_count = len(parsed_data.get("all_registers", [])) if parsed_data else 0
                s.detected_manufacturer = parsed_data.get("manufacturer", "") if parsed_data else ""
                await progress(f"PDF 파싱 완료: {reg_count}개 레지스터 추출 (제조사: {s.detected_manufacturer or 'Unknown'})")

                base = os.path.splitext(s.pdf_filename)[0]
                stage1_path = os.path.join(session_dir, base + "_registers.xlsx")

                api_key = ""
                model = "claude-opus-4-6"
                if req.mode == "ai":
                    api_key = ai_gen.load_api_key()
                    model = ai_gen.load_model_name()
                    if not api_key:
                        await ws_manager.send_error(s.session_id, "stage1",
                            "AI 모드를 사용하려면 AI 설정에서 API 키를 먼저 설정하세요.")
                        s.running = False
                        return

                progress_msgs = []

                def _progress_cb(msg):
                    progress_msgs.append(msg)
                    asyncio.run_coroutine_threadsafe(
                        ws_manager.send_progress(s.session_id, "stage1", msg),
                        loop
                    )

                def _run_stage1():
                    return pipeline.stage1_extract_to_excel(
                        parsed_data=parsed_data,
                        pdf_path=pdf_path,
                        output_path=stage1_path,
                        mode=req.mode,
                        api_key=api_key if req.mode == "ai" else None,
                        model=model,
                        progress_cb=_progress_cb,
                    )

                result_path = await loop.run_in_executor(None, _run_stage1)

                s.stage1_excel_path = result_path
                s.stage1_rows = _excel_to_json_rows(result_path, "Registers")

                await ws_manager.send_done(s.session_id, "stage1", True,
                    f"Stage 1 완료: {len(s.stage1_rows)}개 레지스터")

        except Exception as e:
            tb = traceback.format_exc()
            logger.error(f"Stage 1 error: {e}\n{tb}")
            await ws_manager.send_error(s.session_id, "stage1", str(e))
        finally:
            s.running = False

    asyncio.create_task(_do())
    return {"status": "started", "session_id": s.session_id}


def _detect_counts(rows):
    """Auto-detect MPPT count, String count, IV scan from Stage 1 register names."""
    import re
    mppt_nums = set()
    string_nums = set()
    pv_string_map = {}  # {pv_num: set of string_nums} for "PVx String current Y" pattern
    _iv_has_command = False   # IV scan start command register exists
    _iv_has_data = False      # IV scan result data registers exist

    for r in rows:
        name = str(r.get('Definition', '') or '').lower()
        # MPPT detection: Vpv1, MPPT3, 태양전지1, 솔라셀2, etc.
        # "PV1 input voltage" = MPPT, but "PV1 String current" = String (not MPPT)
        for pat in [r'(?:mppt|vpv|ipv)[\s_-]*(\d+)',
                    r'\ud0dc\uc591\uc804\uc9c0(\d+)',  # 태양전지N
                    r'\uc194\ub77c\uc140(\d+)']:        # 솔라셀N
            for m in re.finditer(pat, name):
                mppt_nums.add(int(m.group(1)))
        # PV with input voltage/current/power = MPPT (not string)
        m_pv_input = re.search(r'pv[\s_-]*(\d+)\s*(?:input|voltage|current|power)', name)
        if m_pv_input and 'string' not in name:
            mppt_nums.add(int(m_pv_input.group(1)))
        # PV with abnormal/alarm/fault = MPPT indicator (but NOT overvoltage — too generic)
        m_pv_alarm = re.search(r'pv[\s_-]*(\d+)\s*(?:abnormal|alarm|fault)', name)
        if m_pv_alarm and 'overvoltage' not in name and 'string' not in name:
            mppt_nums.add(int(m_pv_alarm.group(1)))

        # String detection: Istr1, String2, "PV1 String current 1", 스트링3, etc.
        for pat in [r'(?:istr|string)[\s_-]*(\d+)',
                    r'\uc2a4\ud2b8\ub9c1(\d+)']:        # 스트링N
            for m in re.finditer(pat, name):
                string_nums.add(int(m.group(1)))
        # "PVx String current Y" pattern: track per-MPPT strings
        m_pv_str = re.search(r'pv(\d+)\s+string\s+current\s+(\d+)', name)
        if m_pv_str:
            pv_n = int(m_pv_str.group(1))
            str_n = int(m_pv_str.group(2))
            pv_string_map.setdefault(pv_n, set()).add(str_n)

        # IV scan command detection (start/status register)
        if any(kw in name for kw in ['iv curve', 'iv_curve', 'iv scan', 'iv_scan',
                                      'i-v curve', 'i-v scan', 'iv curve scan',
                                      'i/v scan', 'start i/v', 'start iv']):
            _iv_has_command = True

        # Detect IV command by known addresses (0x600D, 0xA71B, etc.)
        addr_raw = str(r.get('Address_Hex', '') or '').upper()
        addr_dec = r.get('Address_Dec', '')
        try:
            addr_int = int(addr_dec) if addr_dec else 0
        except (ValueError, TypeError):
            addr_int = 0
        if addr_int == 0x600D or '600D' in addr_raw:
            _iv_has_command = True

        # IV scan RESULT data detection by name patterns:
        # IV result registers have names like "Tracker N voltage", "String N-N current",
        # "IV tracker", "IV string", "IV data", "I-V result" etc.
        # These are array-type registers with voltage/current curves per tracker/string.
        if any(kw in name for kw in ['iv_tracker', 'iv_string', 'iv_data', 'iv data',
                                      'i-v data', 'i-v result', 'iv result',
                                      'iv curve data', 'i-v curve data',
                                      'iv\ub370\uc774\ud130', 'iv \ub370\uc774\ud130']):
            _iv_has_data = True
        # "Tracker N voltage" / "String N-N current" pattern (Solarize IV data)
        if re.search(r'tracker\s*\d+\s*voltage', name):
            _iv_has_data = True
        if re.search(r'string\s*\d+-\d+\s*current', name):
            _iv_has_data = True
        # "PV1 Voltage Point 1" / "PV1 Current Point 1" pattern (Kstar IV data)
        if re.search(r'pv\d+\s*(voltage|current)\s*point', name):
            _iv_has_data = True

    # IV scan supported only when BOTH command AND result data registers exist
    iv_scan = _iv_has_command and _iv_has_data

    # IV data points per set: analyze IV register block size
    # Solarize: 0x8000 base, each tracker block = 320 regs,
    #   voltage block = 64 regs (= 64 data points per curve)
    # IV data points per set
    iv_data_points = 0
    if iv_scan:
        # Method 1: Solarize-style — gap between tracker voltage bases
        # "Tracker 1 voltage" at 0x8000, "String 1-1 current" at 0x8040 → 64 pts
        iv_base_addrs = []
        for r in rows:
            name = str(r.get('Definition', '') or '').lower()
            try:
                a = int(r.get('Address_Dec', 0))
            except (ValueError, TypeError):
                continue
            if re.search(r'tracker\s*\d+\s*voltage', name) or \
               re.search(r'string\s*\d+-\d+\s*current', name) or \
               any(kw in name for kw in ['iv_tracker', 'iv_string']):
                iv_base_addrs.append(a)

        if len(iv_base_addrs) >= 2:
            iv_base_addrs.sort()
            iv_data_points = iv_base_addrs[1] - iv_base_addrs[0]

        # Method 2: Kstar-style — "PVx Voltage/Current Point N" pattern
        # V and I alternate: point = 2 registers, total_regs / strings / 2
        if iv_data_points <= 0:
            iv_point_addrs = []
            for r in rows:
                name = str(r.get('Definition', '') or '').lower()
                try:
                    a = int(r.get('Address_Dec', 0))
                except (ValueError, TypeError):
                    continue
                if re.search(r'pv\d+\s*(voltage|current)\s*point', name):
                    iv_point_addrs.append(a)
                # Also check Description for "occupying N registers" or "N registers"
                desc = str(r.get('Description', '') or '').lower()
                all_text = name + ' ' + desc
                m_occ = re.search(r'(?:occupying|iv.*?)\s*(\d{3,})\s*registers', all_text)
                if m_occ:
                    total_iv_regs = int(m_occ.group(1))
                    # total_regs / mppt_count / 2 (V+I pairs) = points per string
                    _mppt = max(mppt_nums) if mppt_nums else 4
                    if _mppt > 0:
                        iv_data_points = total_iv_regs // _mppt // 2

            # Fallback: if we found point addresses, estimate from address gap
            if iv_data_points <= 0 and len(iv_point_addrs) >= 2:
                iv_point_addrs.sort()
                # "PV1 Voltage Point 1"=5000, "PV1 Current Point 1"=5001,
                # "PV1 Voltage Point 2"=5002 → gap between V points = 2
                # Find gap between consecutive same-type points
                iv_data_points = iv_point_addrs[1] - iv_point_addrs[0]

        if iv_data_points <= 0:
            iv_data_points = 64  # default

    mppt_count = max(mppt_nums) if mppt_nums else 4

    # FC code detection: find the most common read FC (FC03 or FC04)
    fc_counts = {}
    for r in rows:
        fc = str(r.get('FC_Code', '') or '').strip().upper()
        if not fc:
            continue
        # Normalize: "04H" -> "FC04", "03H" -> "FC03"
        if fc in ('03H', '3'):
            fc = 'FC03'
        elif fc in ('04H', '4'):
            fc = 'FC04'
        # Only count pure read FCs (not FC06/FC10 write, not FC03/FC06 mixed)
        if fc in ('FC03', 'FC04'):
            fc_counts[fc] = fc_counts.get(fc, 0) + 1
    # Pick the one with more read-only registers
    detected_fc = 'FC03'
    if fc_counts:
        detected_fc = max(fc_counts, key=fc_counts.get)

    # String count: use explicit string registers, or PV String pattern
    if string_nums:
        string_count = max(string_nums)
    elif pv_string_map:
        # "PVx String current Y" → total strings = MPPT count × strings per MPPT
        strings_per_mppt = max(max(s) for s in pv_string_map.values())
        string_count = mppt_count * strings_per_mppt
    else:
        string_count = 8  # default

    # String count must be >= MPPT count
    if string_count < mppt_count:
        string_count = mppt_count

    return {
        "mppt_count": mppt_count,
        "string_count": string_count,
        "fc_code": detected_fc,
        "iv_scan": iv_scan,
        "iv_data_points": iv_data_points,
    }


@router.get("/api/stage1/result")
async def stage1_result(session_id: str = Query(...)):
    """Return Stage 1 result as JSON rows + auto-detected counts."""
    s = _get_session(session_id)
    if not s.stage1_excel_path:
        raise HTTPException(status_code=404, detail="Stage 1 결과가 없습니다.")
    if not s.stage1_rows:
        s.stage1_rows = _excel_to_json_rows(s.stage1_excel_path, "Registers")
    detected = _detect_counts(s.stage1_rows)
    detected["manufacturer"] = getattr(s, 'detected_manufacturer', '') or ''
    return {"rows": s.stage1_rows, "count": len(s.stage1_rows), "detected": detected}


@router.get("/api/stage1/download-excel")
async def stage1_download(session_id: str = Query(...)):
    """Download Stage 1 Excel file."""
    s = _get_session(session_id)
    if not s.stage1_excel_path or not os.path.isfile(s.stage1_excel_path):
        raise HTTPException(status_code=404, detail="Stage 1 Excel 파일이 없습니다.")
    return FileResponse(
        s.stage1_excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(s.stage1_excel_path),
    )


# ---------------------------------------------------------------------------
# Stage 2
# ---------------------------------------------------------------------------

class Stage2Request(BaseModel):
    session_id: str
    mode: str = "offline"
    mppt_count: int = 4
    string_count: int = 8


@router.post("/api/stage2/run")
async def stage2_run(req: Stage2Request):
    """Run Stage 2: Excel → Mapping Excel."""
    s = _get_session(req.session_id)
    if not s.stage1_excel_path or not os.path.isfile(s.stage1_excel_path):
        raise HTTPException(status_code=400, detail="Stage 1을 먼저 실행하세요.")
    if s.running:
        raise HTTPException(status_code=409, detail="이미 실행 중입니다.")

    s.running = True
    s.stage2_rows = []
    s.stage3_code = ""
    s.stage3_results = []

    async def _do():
        try:
            pipeline = _import_pipeline()
            ai_gen = _import_ai()

            session_dir = s.session_dir()
            loop = asyncio.get_event_loop()

            base = os.path.splitext(os.path.basename(s.stage1_excel_path))[0]
            stage2_path = os.path.join(session_dir, base + "_mapping.xlsx")

            api_key = ""
            model = "claude-opus-4-6"
            if req.mode == "ai":
                api_key = ai_gen.load_api_key()
                model = ai_gen.load_model_name()
                if not api_key:
                    await ws_manager.send_error(s.session_id, "stage2",
                        "AI 모드를 사용하려면 AI 설정에서 API 키를 먼저 설정하세요.")
                    s.running = False
                    return

            def _progress_cb(msg):
                asyncio.run_coroutine_threadsafe(
                    ws_manager.send_progress(s.session_id, "stage2", msg),
                    loop
                )

            def _run():
                return pipeline.stage2_create_mapping_excel(
                    stage1_excel_path=s.stage1_excel_path,
                    output_path=stage2_path,
                    mode=req.mode,
                    api_key=api_key if req.mode == "ai" else None,
                    model=model,
                    progress_cb=_progress_cb,
                )

            result_path = await loop.run_in_executor(None, _run)
            s.stage2_excel_path = result_path
            s.stage2_rows = _excel_to_json_rows(result_path, "Register_Mapping")

            await ws_manager.send_done(s.session_id, "stage2", True,
                f"Stage 2 완료: {len(s.stage2_rows)}개 매핑")

        except Exception as e:
            tb = traceback.format_exc()
            logger.error(f"Stage 2 error: {e}\n{tb}")
            await ws_manager.send_error(s.session_id, "stage2", str(e))
        finally:
            s.running = False

    asyncio.create_task(_do())
    return {"status": "started", "session_id": s.session_id}


@router.get("/api/stage2/result")
async def stage2_result(session_id: str = Query(...)):
    """Return Stage 2 mapping table as JSON."""
    s = _get_session(session_id)
    if not s.stage2_excel_path:
        raise HTTPException(status_code=404, detail="Stage 2 결과가 없습니다.")
    if not s.stage2_rows:
        s.stage2_rows = _excel_to_json_rows(s.stage2_excel_path, "Register_Mapping")
    return {"rows": s.stage2_rows, "count": len(s.stage2_rows)}


class Stage2UpdateRow(BaseModel):
    session_id: str
    row_index: int
    col_name: str
    value: str


@router.put("/api/stage2/update-row")
async def stage2_update_row(req: Stage2UpdateRow):
    """Inline edit: update a single cell in Stage 2 mapping Excel."""
    s = _get_session(req.session_id)
    if not s.stage2_excel_path or not os.path.isfile(s.stage2_excel_path):
        raise HTTPException(status_code=404, detail="Stage 2 Excel 파일이 없습니다.")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(s.stage2_excel_path)
        ws = wb["Register_Mapping"]
        # Find column index by header name (row 1)
        headers = [cell.value for cell in ws[1]]
        if req.col_name not in headers:
            raise HTTPException(status_code=400, detail=f"컬럼 없음: {req.col_name}")
        col_idx = headers.index(req.col_name) + 1
        excel_row = req.row_index + 2  # +1 for header, +1 for 0-based
        ws.cell(row=excel_row, column=col_idx, value=req.value)
        wb.save(s.stage2_excel_path)

        # Refresh cache
        s.stage2_rows = _excel_to_json_rows(s.stage2_excel_path, "Register_Mapping")
        return {"ok": True, "rows": len(s.stage2_rows)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/stage2/download-excel")
async def stage2_download(session_id: str = Query(...)):
    """Download Stage 2 mapping Excel."""
    s = _get_session(session_id)
    if not s.stage2_excel_path or not os.path.isfile(s.stage2_excel_path):
        raise HTTPException(status_code=404, detail="Stage 2 Excel 파일이 없습니다.")
    return FileResponse(
        s.stage2_excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(s.stage2_excel_path),
    )


# ---------------------------------------------------------------------------
# Stage 3
# ---------------------------------------------------------------------------

class Stage3Request(BaseModel):
    session_id: str
    mode: str = "offline"
    protocol_name: str = "custom"
    manufacturer: str = ""
    mppt_count: int = 4
    string_count: int = 8
    iv_scan: bool = False
    der_avm: bool = True
    dea_avm: bool = True
    class_name: str = "RegisterMap"
    fc_code: str = "FC03"


@router.post("/api/stage3/run")
async def stage3_run(req: Stage3Request):
    """Run Stage 3: Mapping Excel → registers.py with 12-item validation."""
    s = _get_session(req.session_id)
    if not s.stage2_excel_path or not os.path.isfile(s.stage2_excel_path):
        raise HTTPException(status_code=400, detail="Stage 2를 먼저 실행하세요.")
    if s.running:
        raise HTTPException(status_code=409, detail="이미 실행 중입니다.")

    s.running = True
    s.stage3_code = ""
    s.stage3_results = []

    async def _do():
        try:
            pipeline = _import_pipeline()
            ai_gen = _import_ai()

            session_dir = s.session_dir()
            loop = asyncio.get_event_loop()

            out_name = f"{req.protocol_name}_registers.py"
            output_path = os.path.join(session_dir, out_name)

            api_key = ""
            model = "claude-opus-4-6"
            if req.mode == "ai":
                api_key = ai_gen.load_api_key()
                model = ai_gen.load_model_name()
                if not api_key:
                    await ws_manager.send_error(s.session_id, "stage3",
                        "AI 모드를 사용하려면 AI 설정에서 API 키를 먼저 설정하세요.")
                    s.running = False
                    return

            settings = {
                "protocol_name": req.protocol_name,
                "manufacturer":  req.manufacturer,
                "mppt_count":    req.mppt_count,
                "string_count":  req.string_count,
                "iv_scan":       req.iv_scan,
                "der_avm":       req.der_avm,
                "dea_avm":       req.dea_avm,
                "class_name":    req.class_name,
                "fc_code":       req.fc_code,
            }

            def _progress_cb(msg):
                asyncio.run_coroutine_threadsafe(
                    ws_manager.send_progress(s.session_id, "stage3", msg),
                    loop
                )

            def _run():
                return pipeline.stage3_generate_py(
                    mapping_excel_path=s.stage2_excel_path,
                    settings=settings,
                    output_path=output_path,
                    mode=req.mode,
                    api_key=api_key if req.mode == "ai" else None,
                    model=model,
                    progress_cb=_progress_cb,
                )

            result = await loop.run_in_executor(None, _run)

            s.stage3_code = result.get("code", "")
            s.stage3_results = result.get("results", [])
            s.stage3_success = result.get("success", False)
            s.stage3_output_path = result.get("output_path")

            pass_count = sum(1 for r in s.stage3_results if r[0] == "PASS")
            total = len(s.stage3_results)
            await ws_manager.send_done(s.session_id, "stage3", s.stage3_success,
                f"검증 {pass_count}/{total} 통과")

        except Exception as e:
            tb = traceback.format_exc()
            logger.error(f"Stage 3 error: {e}\n{tb}")
            await ws_manager.send_error(s.session_id, "stage3", str(e))
        finally:
            s.running = False

    asyncio.create_task(_do())
    return {"status": "started", "session_id": s.session_id}


@router.get("/api/stage3/result")
async def stage3_result(session_id: str = Query(...)):
    """Return Stage 3 generated code and validation results."""
    s = _get_session(session_id)
    if not s.stage3_code:
        raise HTTPException(status_code=404, detail="Stage 3 결과가 없습니다.")
    return {
        "code": s.stage3_code,
        "results": s.stage3_results,
        "success": s.stage3_success,
    }


class Stage3SaveRequest(BaseModel):
    session_id: str
    protocol_name: str
    save_as_reference: bool = True
    manufacturer: str = ""
    description: str = ""
    capacity: str = ""
    mppt_count: int = 4
    string_count: int = 8
    fc_code: str = "FC03"
    iv_scan: bool = False
    der_avm: bool = True


@router.post("/api/stage3/save")
async def stage3_save(req: Stage3SaveRequest):
    """Save generated registers.py to common/ directory."""
    s = _get_session(req.session_id)
    if not s.stage3_code:
        raise HTTPException(status_code=400, detail="Stage 3 코드가 없습니다.")

    root = _project_root()

    # Check for protected files
    protected = {
        "solarize_registers.py", "kstar_registers.py", "huawei_registers.py",
        "sungrow_registers.py", "ekos_registers.py", "goodwe_registers.py",
        "senergy_registers.py", "relay_registers.py", "weather_registers.py",
    }
    out_name = f"{req.protocol_name}_registers.py"
    if out_name in protected:
        raise HTTPException(status_code=403,
            detail=f"{out_name}은 보호된 파일입니다. 다른 프로토콜 이름을 사용하세요.")

    out_path = os.path.join(root, "common", out_name)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(s.stage3_code + "\n")

    # Optionally save to reference library
    if req.save_as_reference:
        try:
            pipeline = _import_pipeline()
            pipeline.save_to_reference(
                protocol=req.protocol_name,
                py_code=s.stage3_code,
                mapping_excel_path=s.stage2_excel_path,
                meta={
                    "protocol": req.protocol_name,
                    "manufacturer": req.manufacturer,
                    "description": req.description,
                    "capacity": req.capacity,
                    "mppt_count": req.mppt_count,
                    "string_count": req.string_count,
                    "fc_code": req.fc_code,
                    "iv_scan": req.iv_scan,
                    "der_avm": req.der_avm,
                },
            )
        except Exception as e:
            logger.warning(f"Reference save failed: {e}")

    logger.info(f"Saved: {out_path}")
    return {"ok": True, "path": out_path, "filename": out_name}


@router.get("/api/stage3/download-py")
async def stage3_download(session_id: str = Query(...), protocol_name: str = Query("custom")):
    """Download generated registers.py."""
    s = _get_session(session_id)
    if not s.stage3_code:
        raise HTTPException(status_code=404, detail="Stage 3 코드가 없습니다.")

    # Write to temp file if output_path not set
    if s.stage3_output_path and os.path.isfile(s.stage3_output_path):
        path = s.stage3_output_path
    else:
        path = os.path.join(s.session_dir(), f"{protocol_name}_registers.py")
        with open(path, "w", encoding="utf-8") as f:
            f.write(s.stage3_code + "\n")

    return FileResponse(
        path,
        media_type="text/x-python",
        filename=os.path.basename(path),
    )


# ---------------------------------------------------------------------------
# Add Model to device_models.ini
# ---------------------------------------------------------------------------

class AddModelRequest(BaseModel):
    model_name: str          # e.g. "GoodWe"
    protocol_name: str       # e.g. "goodwe"
    device_type: str = "inverter"  # inverter, relay, weather
    iv_scan: bool = False
    kdn: bool = False


@router.post("/api/add-model")
async def add_model(req: AddModelRequest):
    """Add a new model entry to config/device_models.ini."""
    import configparser

    ini_path = os.path.join(_project_root(), "config", "device_models.ini")
    if not os.path.isfile(ini_path):
        raise HTTPException(status_code=404, detail="device_models.ini 파일을 찾을 수 없습니다.")

    cfg = configparser.ConfigParser()
    cfg.read(ini_path, encoding="utf-8")

    dt = req.device_type.lower()
    if dt == "inverter":
        model_sec = "inverter_models"
        proto_sec = "inverter_protocols"
        feat_sec = "inverter_features"
    elif dt == "relay":
        model_sec = "relay_models"
        proto_sec = "relay_protocols"
        feat_sec = None
    elif dt == "weather":
        model_sec = "weather_models"
        proto_sec = "weather_protocols"
        feat_sec = None
    else:
        raise HTTPException(status_code=400, detail=f"지원하지 않는 디바이스 타입: {dt}")

    # Check duplicate protocol
    if cfg.has_section(proto_sec):
        for mid, pname in cfg.items(proto_sec):
            if mid.startswith("#"):
                continue
            if pname.strip() == req.protocol_name.strip():
                raise HTTPException(status_code=409,
                    detail=f"프로토콜 '{req.protocol_name}'이(가) 이미 model_id={mid}에 등록되어 있습니다.")

    # Find next model_id
    existing_ids = []
    if cfg.has_section(model_sec):
        for mid, _ in cfg.items(model_sec):
            if mid.startswith("#"):
                continue
            try:
                existing_ids.append(int(mid))
            except ValueError:
                pass
    next_id = max(existing_ids) + 1 if existing_ids else 1

    # Add entries
    sid = str(next_id)
    if not cfg.has_section(model_sec):
        cfg.add_section(model_sec)
    cfg.set(model_sec, sid, req.model_name)

    if not cfg.has_section(proto_sec):
        cfg.add_section(proto_sec)
    cfg.set(proto_sec, sid, req.protocol_name)

    if feat_sec and dt == "inverter":
        if not cfg.has_section(feat_sec):
            cfg.add_section(feat_sec)
        feat_val = f"{'true' if req.iv_scan else 'false'}, {'true' if req.kdn else 'false'}"
        cfg.set(feat_sec, sid, feat_val)

    # Write back preserving comments — use raw file manipulation
    # configparser loses comments, so manually append instead
    lines = []
    with open(ini_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    def _append_after_section_last_entry(lines, section, key, value):
        """Append key=value after the last entry of [section]."""
        in_section = False
        last_entry_idx = -1
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped.lower() == f"[{section}]":
                in_section = True
                last_entry_idx = i
                continue
            if in_section:
                if stripped.startswith("["):
                    break  # next section
                if stripped and not stripped.startswith("#"):
                    last_entry_idx = i
        if last_entry_idx >= 0:
            lines.insert(last_entry_idx + 1, f"{key} = {value}\n")
        return lines

    lines = _append_after_section_last_entry(lines, model_sec, sid, req.model_name)
    lines = _append_after_section_last_entry(lines, proto_sec, sid, req.protocol_name)
    if feat_sec and dt == "inverter":
        feat_val = f"{'true' if req.iv_scan else 'false'}, {'true' if req.kdn else 'false'}"
        lines = _append_after_section_last_entry(lines, feat_sec, sid, feat_val)

    with open(ini_path, "w", encoding="utf-8") as f:
        f.writelines(lines)

    logger.info(f"Model added: {req.model_name} (id={sid}, protocol={req.protocol_name})")

    return {
        "ok": True,
        "model_id": int(sid),
        "model_name": req.model_name,
        "protocol_name": req.protocol_name,
        "device_type": dt,
    }


# ---------------------------------------------------------------------------
# References
# ---------------------------------------------------------------------------

@router.get("/api/references")
async def get_references():
    """List all references (builtin + user-created)."""
    ref_mgr = _import_ref()
    mgr = ref_mgr.get_manager()

    # Load protocol_id from device_models.ini
    import configparser
    _proto_ids = {}
    ini_path = os.path.join(_project_root(), "config", "device_models.ini")
    if os.path.isfile(ini_path):
        cfg = configparser.ConfigParser()
        cfg.read(ini_path, encoding="utf-8")
        if cfg.has_section("inverter_protocols"):
            for mid, pname in cfg.items("inverter_protocols"):
                if not mid.startswith("#"):
                    _proto_ids[pname.strip()] = int(mid)

    result = []
    for name, meta in mgr._all_index.items():
        result.append({
            "protocol_id": _proto_ids.get(name, _proto_ids.get(meta.get("protocol", ""), "")),
            "name": name,
            "manufacturer": meta.get("manufacturer", ""),
            "protocol": meta.get("protocol", name),
            "description": meta.get("description", ""),
            "capacity": meta.get("capacity", ""),
            "mppt_count": meta.get("mppt_count", 4),
            "string_count": meta.get("string_count", 8),
            "fc_code": meta.get("fc_code", "FC03"),
            "iv_scan": meta.get("iv_scan", False),
            "der_avm": meta.get("der_avm", False),
            "builtin": meta.get("builtin", False),
            "created": meta.get("created", ""),
        })
    return {"references": result}


class SaveReferenceRequest(BaseModel):
    session_id: str
    name: str
    manufacturer: str = ""
    description: str = ""


@router.post("/api/references")
async def save_reference(req: SaveReferenceRequest):
    """Save current Stage 3 result as a new reference."""
    s = _get_session(req.session_id)
    if not s.stage3_code:
        raise HTTPException(status_code=400, detail="Stage 3 코드가 없습니다.")

    ref_mgr = _import_ref()
    mgr = ref_mgr.get_manager()
    meta = {
        "manufacturer": req.manufacturer,
        "protocol": req.name,
        "description": req.description,
    }
    mgr.save_reference(req.name, s.stage3_code, meta, s.stage2_excel_path)
    return {"ok": True, "name": req.name}


@router.delete("/api/references/{name}")
async def delete_reference(name: str):
    """Delete a user-created reference, config entries, and common/ files."""
    ref_mgr = _import_ref()
    mgr = ref_mgr.get_manager()
    ok, msg = mgr.delete_reference(name)
    if not ok:
        raise HTTPException(status_code=403, detail=msg)

    details = [msg]

    # Remove from device_models.ini
    import configparser
    ini_path = os.path.join(_project_root(), "config", "device_models.ini")
    if os.path.isfile(ini_path):
        with open(ini_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        # Find model_id for this protocol
        model_id = None
        for line in lines:
            stripped = line.strip()
            if '=' in stripped and not stripped.startswith('#') and not stripped.startswith('['):
                parts = stripped.split('=', 1)
                if parts[1].strip() == name:
                    try:
                        model_id = parts[0].strip()
                        break
                    except ValueError:
                        pass

        if model_id:
            # Remove lines with this model_id from all inverter sections
            new_lines = []
            for line in lines:
                stripped = line.strip()
                if '=' in stripped and not stripped.startswith('#') and not stripped.startswith('['):
                    key = stripped.split('=', 1)[0].strip()
                    if key == model_id:
                        continue  # skip this line
                new_lines.append(line)
            with open(ini_path, 'w', encoding='utf-8') as f:
                f.writelines(new_lines)
            details.append(f"device_models.ini에서 id={model_id} 제거")

    # Remove common/ files
    root = _project_root()
    for suffix in ['_registers.py', '_mm_registers.py']:
        fpath = os.path.join(root, "common", f"{name}{suffix}")
        if os.path.isfile(fpath):
            os.remove(fpath)
            details.append(f"common/{name}{suffix} 삭제")

    logger.info(f"Reference deleted: {name} - {details}")
    return {"ok": True, "message": " | ".join(details)}


@router.post("/api/references/{name}/promote")
async def promote_reference(name: str):
    """Promote a user-created reference to builtin (verified) status."""
    ref_mgr = _import_ref()
    mgr = ref_mgr.get_manager()
    ok, msg = mgr.promote_to_builtin(name)
    if not ok:
        raise HTTPException(status_code=400, detail=msg)
    return {"ok": True, "message": msg}


# ---------------------------------------------------------------------------
# AI Settings
# ---------------------------------------------------------------------------

@router.get("/api/ai-settings")
async def get_ai_settings():
    """Get current AI settings (API key masked)."""
    ai_gen = _import_ai()
    key = ai_gen.load_api_key()
    model = ai_gen.load_model_name()
    return {
        "api_key_set": bool(key),
        "api_key_preview": (key[:8] + "..." + key[-4:]) if len(key) > 12 else ("***" if key else ""),
        "model": model,
    }


class AISettingsRequest(BaseModel):
    api_key: Optional[str] = None
    model: Optional[str] = None


@router.put("/api/ai-settings")
async def update_ai_settings(req: AISettingsRequest):
    """Update AI settings."""
    ai_gen = _import_ai()
    if req.api_key is not None:
        ai_gen.save_api_key(req.api_key)
    if req.model is not None:
        ai_gen.save_model_name(req.model)
    return {"ok": True}


# ---------------------------------------------------------------------------
# Config File Management
# ---------------------------------------------------------------------------

@router.get("/api/config/list")
async def config_list():
    """List available config files."""
    config_dir = os.path.join(_project_root(), "config")
    files = []
    if os.path.isdir(config_dir):
        for f in sorted(os.listdir(config_dir)):
            if f.endswith('.ini'):
                fpath = os.path.join(config_dir, f)
                files.append({
                    "name": f,
                    "size": os.path.getsize(fpath),
                })
    return {"files": files}


@router.get("/api/config/{filename}")
async def config_read(filename: str):
    """Read a config file content."""
    if not filename.endswith('.ini'):
        raise HTTPException(status_code=400, detail="INI 파일만 읽을 수 있습니다.")
    fpath = os.path.join(_project_root(), "config", filename)
    if not os.path.isfile(fpath):
        raise HTTPException(status_code=404, detail=f"파일을 찾을 수 없습니다: {filename}")
    with open(fpath, 'r', encoding='utf-8') as f:
        content = f.read()
    return {"name": filename, "content": content}


class ConfigSaveRequest(BaseModel):
    content: str


@router.put("/api/config/{filename}")
async def config_save(filename: str, req: ConfigSaveRequest):
    """Save a config file."""
    if not filename.endswith('.ini'):
        raise HTTPException(status_code=400, detail="INI 파일만 저장할 수 있습니다.")
    fpath = os.path.join(_project_root(), "config", filename)
    if not os.path.isfile(fpath):
        raise HTTPException(status_code=404, detail=f"파일을 찾을 수 없습니다: {filename}")
    # Backup before save
    import shutil
    backup = fpath + '.bak'
    shutil.copy2(fpath, backup)
    with open(fpath, 'w', encoding='utf-8') as f:
        f.write(req.content)
    logger.info(f"Config saved: {filename} ({len(req.content)} bytes)")
    return {"ok": True, "name": filename, "size": len(req.content)}
