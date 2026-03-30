# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
import openpyxl

ROOT = r'C:\CM4_4rs485\RTU_UDP_System_V1_0_0\model_maker\comparison_output_all'

INVERTERS = ['solarize', 'kstar', 'huawei', 'sungrow', 'ekos']

DER_AVM_12 = [
    'INVERTER_ON_OFF', 'ACTIVE_POWER_LIMIT', 'POWER_FACTOR_SET',
    'REACTIVE_POWER_SET', 'RAMP_RATE', 'VOLT_VAR_ENABLE',
    'VOLT_WATT_ENABLE', 'DER_ACTION_MODE', 'DER_ACTIVE_POWER_PCT',
    'DER_REACTIVE_POWER_PCT', 'DER_POWER_FACTOR_SET', 'DER_RAMP_RATE'
]

DEA_AVM_23 = [
    'DEA_L1_CURRENT_LOW', 'DEA_L1_CURRENT_HIGH', 'DEA_L2_CURRENT_LOW', 'DEA_L2_CURRENT_HIGH',
    'DEA_L3_CURRENT_LOW', 'DEA_L3_CURRENT_HIGH', 'DEA_L1_VOLTAGE', 'DEA_L2_VOLTAGE',
    'DEA_L3_VOLTAGE', 'DEA_FREQ', 'DEA_ACTIVE_POWER', 'DEA_REACTIVE_POWER',
    'DEA_APPARENT_POWER', 'DEA_POWER_FACTOR', 'DEA_TOTAL_ENERGY_LOW', 'DEA_TOTAL_ENERGY_HIGH',
    'DEA_TODAY_ENERGY_LOW', 'DEA_TODAY_ENERGY_HIGH', 'DEA_CB_STATUS', 'DEA_FAULT_STATUS',
    'DEA_OPERATION_MODE', 'DEA_INVERTER_MODE', 'DEA_ERROR_CODE'
]

print("=" * 90)
print("  STAGE 2 EXCEL 상세 검사 -- 5종 인버터 × 2모드")
print("=" * 90)

for key in INVERTERS:
    print("\n" + "-"*90)
    print("  [%s]" % key.upper())
    print("-"*90)

    for mode in ['offline', 'ai']:
        path = os.path.join(ROOT, key, '%s_stage2_%s.xlsx' % (key, mode))
        if not os.path.exists(path):
            print("  [%s] 파일 없음: %s" % (mode.upper(), path))
            continue

        wb = openpyxl.load_workbook(path, data_only=True)
        sheets = wb.sheetnames
        print("\n  [%s] 시트 목록: %s" % (mode.upper(), sheets))

        if 'Register_Mapping' not in sheets:
            print("  [%s] Register_Mapping 시트 없음!" % mode.upper())
            continue

        ws = wb['Register_Mapping']
        auto, manual, unmapped, ai_extra, ref_mapped = 0, 0, 0, 0, 0
        unmapped_list = []
        match_types = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if isinstance(row[0], str) and row[0].startswith('---'):
                continue
            if not isinstance(row[0], (int, float)):
                continue

            mt = str(row[13] or '').strip() if len(row) > 13 else ''
            match_types[mt] = match_types.get(mt, 0) + 1

            if mt.startswith('AI('):
                ai_extra += 1
                auto += 1
            elif mt == 'Auto' or mt.startswith('Ref('):
                auto += 1
                if mt.startswith('Ref('):
                    ref_mapped += 1
            elif mt == 'Manual':
                manual += 1
            else:
                unmapped += 1
                addr = str(row[2] or '') if len(row) > 2 else ''
                defn = str(row[4] or '')[:40] if len(row) > 4 else ''
                unmapped_list.append((addr, defn))

        total = auto + manual + unmapped
        rate = auto / total * 100 if total > 0 else 0
        auto_only = auto - ref_mapped - ai_extra
        print("  [%s] Register_Mapping: 총 %d개" % (mode.upper(), total))
        print("    - Auto (주소직접):  %d" % auto_only)
        print("    - Ref 매핑:        %d" % ref_mapped)
        print("    - AI 매핑:         %d" % ai_extra)
        print("    - Manual:          %d" % manual)
        print("    - Unmapped:        %d" % unmapped)
        print("    - 자동 매핑률:     %.1f%%" % rate)
        top_types = sorted(match_types.items(), key=lambda x: -x[1])[:6]
        print("    - Match 타입 분포: %s" % dict(top_types))

        if unmapped_list:
            print("    - Unmapped 목록 (첫 15개):")
            for a, d in unmapped_list[:15]:
                print("        %8s | %s" % (a, d))

        # InverterMode
        if 'InverterMode' in sheets:
            ws_m = wb['InverterMode']
            mode_rows = [(r[0], r[1]) for r in ws_m.iter_rows(min_row=2, values_only=True)
                         if r[0] is not None]
            required_modes = {'INITIAL', 'STANDBY', 'ON_GRID', 'FAULT', 'SHUTDOWN'}
            found_modes = {str(r[1] or '').strip() for r in mode_rows}
            has_all = required_modes.issubset(found_modes)
            print("  [%s] InverterMode: %d개, 필수5개=%s, found=%s" % (
                mode.upper(), len(mode_rows),
                'OK' if has_all else 'FAIL',
                str(found_modes & required_modes)))
        else:
            print("  [%s] InverterMode 시트 없음!" % mode.upper())

        # DER_AVM_Control
        if 'DER_AVM_Control' in sheets:
            ws_d = wb['DER_AVM_Control']
            der_rows = [(r[0], r[1]) for r in ws_d.iter_rows(min_row=2, values_only=True)
                        if r[0] is not None]
            found_der = {str(r[1] or '').strip() for r in der_rows}
            missing_der = [d for d in DER_AVM_12 if d not in found_der]
            print("  [%s] DER_AVM_Control: %d개 (필수12개 기준), 누락=%s" % (
                mode.upper(), len(der_rows),
                '없음' if not missing_der else str(missing_der)))
        else:
            print("  [%s] DER_AVM_Control 시트 없음!" % mode.upper())

        # DEA_AVM_Monitor
        if 'DEA_AVM_Monitor' in sheets:
            ws_dea = wb['DEA_AVM_Monitor']
            dea_rows = [(r[0], r[1]) for r in ws_dea.iter_rows(min_row=2, values_only=True)
                        if r[0] is not None]
            found_dea = {str(r[1] or '').strip() for r in dea_rows}
            missing_dea = [d for d in DEA_AVM_23 if d not in found_dea]
            print("  [%s] DEA_AVM_Monitor: %d개 (필수23개 기준), 누락=%s" % (
                mode.upper(), len(dea_rows),
                '없음' if not missing_dea else str(missing_dea)))
        else:
            print("  [%s] DEA_AVM_Monitor 시트 없음!" % mode.upper())

        # Status/Error
        for extra_sheet in ['Status_Definitions', 'Error_Fault_Codes']:
            if extra_sheet in sheets:
                ws_e = wb[extra_sheet]
                cnt = sum(1 for r in ws_e.iter_rows(min_row=2, values_only=True)
                          if r[0] is not None)
                print("  [%s] %s: %d개 항목" % (mode.upper(), extra_sheet, cnt))


print("\n" + "=" * 90)
print("  최종 요약 비교표 (이전 결과 대비)")
print("=" * 90)
print("  %-10s %10s %8s %10s %6s %6s" % ("인버터", "Offline%", "AI%", "이전대비", "DER12", "DEA23"))
print("  " + "-"*55)

prev_rates = {'solarize': 83.8, 'kstar': 63.9, 'huawei': 28.3, 'sungrow': 11.5, 'ekos': 56.0}

for key in INVERTERS:
    rates = {}
    der_ok = {}
    dea_ok = {}
    for mode in ['offline', 'ai']:
        path = os.path.join(ROOT, key, '%s_stage2_%s.xlsx' % (key, mode))
        if not os.path.exists(path):
            rates[mode] = 0
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb['Register_Mapping']
        auto_c, manual_c, unmapped_c = 0, 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None or not isinstance(row[0], (int, float)):
                continue
            mt = str(row[13] or '').strip() if len(row) > 13 else ''
            if mt.startswith('AI(') or mt in ('Auto',) or mt.startswith('Ref('):
                auto_c += 1
            elif mt == 'Manual':
                manual_c += 1
            else:
                unmapped_c += 1
        total = auto_c + manual_c + unmapped_c
        rates[mode] = auto_c / total * 100 if total > 0 else 0
        der_cnt = sum(1 for r in wb['DER_AVM_Control'].iter_rows(min_row=2, values_only=True)
                      if r[0] is not None) if 'DER_AVM_Control' in wb.sheetnames else 0
        dea_cnt = sum(1 for r in wb['DEA_AVM_Monitor'].iter_rows(min_row=2, values_only=True)
                      if r[0] is not None) if 'DEA_AVM_Monitor' in wb.sheetnames else 0
        der_ok[mode] = der_cnt
        dea_ok[mode] = dea_cnt

    prev = prev_rates.get(key, 0)
    curr = rates.get('offline', 0)
    ai_r = rates.get('ai', 0)
    delta = curr - prev
    delta_str = "+%.1f%%" % delta if delta >= 0 else "%.1f%%" % delta
    print("  %-10s %9.1f%% %7.1f%% %10s %6d %6d" % (
        key, curr, ai_r, delta_str,
        der_ok.get('offline', 0), dea_ok.get('offline', 0)))
