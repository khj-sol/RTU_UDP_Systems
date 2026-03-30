# -*- coding: utf-8 -*-
"""
Stage Pipeline - Model Maker v1.4.0

3-stage register file generation pipeline with two operating modes:

  Mode 1 - Offline (no API):
    Stage 1: PDF -> Excel  (PyMuPDF rule-based)
    Stage 2: Excel -> Mapping Excel  (Solarize addr-table + reference name similarity)
    Stage 3: Mapping Excel -> .py  (code generation + 12-item validation)

  Mode 2 - AI (Claude API):
    Stage 1: PDF -> Excel  (PyMuPDF + Claude for higher accuracy)
    Stage 2: Excel -> Mapping Excel  (Solarize addr-table + Claude few-shot mapping)
    Stage 3: Mapping Excel -> .py  (code generation + validation + auto-retry x3)

Both modes:
  - Reference library (model_maker/reference/) used for matching
  - DER-AVM registers always added
  - On success: new reference auto-saved to reference library

Usage:
  import stage_pipeline
  stage_pipeline.stage1_extract_to_excel(parsed_data, pdf_path, output_path,
      mode='offline')  # or mode='ai', api_key=..., model=...
  stage_pipeline.stage2_create_mapping_excel(stage1_path, output_path,
      solarize_addr_map, solarize_scale, mode='offline')
  stage_pipeline.stage3_generate_py(mapping_excel_path, settings, output_path,
      mode='offline')   # mode='ai' enables auto-retry
"""

import os
import re
from datetime import datetime

# Reference manager (lazy import to avoid circular)
def _get_ref_manager():
    try:
        from model_maker import reference_manager as _rm
    except ImportError:
        try:
            import reference_manager as _rm
        except ImportError:
            return None
    return _rm.get_manager()

# ---------------------------------------------------------------------------
# Standard Solarize register tables (always included in Stage 2/3)
# ---------------------------------------------------------------------------

# Standard InverterMode definitions — codes from Solarize Modbus Protocol V1.2.4 Table (0x101D)
_INVERTER_MODES = [
    (0x00, 'INITIAL',   'Initial mode'),
    (0x01, 'STANDBY',   'Standby mode'),
    (0x03, 'ON_GRID',   'On-Grid mode'),
    (0x04, 'OFF_GRID',  'Off-Grid mode'),
    (0x05, 'FAULT',     'Fault mode'),
    (0x09, 'SHUTDOWN',  'Shutdown mode'),
]

# Standard DER-AVM control registers (Sheet 3 in Stage 2 - always included)
# columns: addr_hex, addr_dec, name, data_type, fc, regs, unit, scale, rw, description
_DER_AVM_CONTROL_REGS = [
    ('0x07D0', 0x07D0, 'DER_POWER_FACTOR_SET',          'S16', 'FC06', 1, '',  '0.001', 'RW', 'Power factor setpoint [-1000,-800],[800,1000]'),
    ('0x07D1', 0x07D1, 'DER_ACTION_MODE',               'U16', 'FC06', 1, '',  '1',     'RW', 'Action mode: 0=Self, 2=DER-AVM, 5=Q(V)'),
    ('0x07D2', 0x07D2, 'DER_REACTIVE_POWER_PCT',        'S16', 'FC06', 1, '%', '0.1',   'RW', 'Reactive power % [-484,484]'),
    ('0x07D3', 0x07D3, 'DER_ACTIVE_POWER_PCT',          'U16', 'FC06', 1, '%', '0.1',   'RW', 'Active power % [0,1100]'),
    ('0x0834', 0x0834, 'INVERTER_ON_OFF',               'U16', 'FC06', 1, '',  '1',     'RW', 'On/Off control: 0=ON, 1=OFF'),
    ('0x0835', 0x0835, 'CLEAR_PV_INSULATION_WARNING',   'U16', 'FC06', 1, '',  '1',     'WO', 'Clear PV insulation warning: 1=Clear'),
    ('0x6001', 0x6001, 'INVERTER_CONTROL',              'U16', 'FC06', 1, '',  '1',     'WO', 'Inverter control: 0=Power on, 1=Shut down'),
    ('0x600D', 0x600D, 'IV_CURVE_SCAN',                 'U16', 'FC06', 1, '',  '1',     'RW', 'IV curve scan: W 0=Stop,1=Start; R 0=Idle,1=Running,2=Finished'),
    ('0x600F', 0x600F, 'POWER_FACTOR_DYNAMIC',          'S16', 'FC06', 1, '',  '0.001', 'RW', 'Dynamic power factor control'),
    ('0x6010', 0x6010, 'REACTIVE_POWER_DYNAMIC',        'S16', 'FC06', 1, '%', '0.01',  'RW', 'Dynamic reactive power control (%)'),
    ('0x3005', 0x3005, 'POWER_DERATING_PCT',            'U16', 'FC06', 1, '%', '1',     'RW', 'Active power derating [0-110]%'),
]

# Standard DEA-AVM monitoring registers (Sheet 4 in Stage 2 - always included)
_DEA_AVM_MONITOR_REGS = [
    ('0x03E8', 0x03E8, 'DEA_L1_CURRENT_LOW',            'S32', 'FC03', 2, 'A',   '0.1',   'RO', 'DEA L1 current (low word)'),
    ('0x03E9', 0x03E9, 'DEA_L1_CURRENT_HIGH',           'S32', 'FC03', 1, '',    '',       'RO', 'DEA L1 current (high word)'),
    ('0x03EA', 0x03EA, 'DEA_L2_CURRENT_LOW',            'S32', 'FC03', 2, 'A',   '0.1',   'RO', 'DEA L2 current (low word)'),
    ('0x03EB', 0x03EB, 'DEA_L2_CURRENT_HIGH',           'S32', 'FC03', 1, '',    '',       'RO', 'DEA L2 current (high word)'),
    ('0x03EC', 0x03EC, 'DEA_L3_CURRENT_LOW',            'S32', 'FC03', 2, 'A',   '0.1',   'RO', 'DEA L3 current (low word)'),
    ('0x03ED', 0x03ED, 'DEA_L3_CURRENT_HIGH',           'S32', 'FC03', 1, '',    '',       'RO', 'DEA L3 current (high word)'),
    ('0x03EE', 0x03EE, 'DEA_L1_VOLTAGE_LOW',            'S32', 'FC03', 2, 'V',   '0.1',   'RO', 'DEA L1 voltage (low word)'),
    ('0x03EF', 0x03EF, 'DEA_L1_VOLTAGE_HIGH',           'S32', 'FC03', 1, '',    '',       'RO', 'DEA L1 voltage (high word)'),
    ('0x03F0', 0x03F0, 'DEA_L2_VOLTAGE_LOW',            'S32', 'FC03', 2, 'V',   '0.1',   'RO', 'DEA L2 voltage (low word)'),
    ('0x03F1', 0x03F1, 'DEA_L2_VOLTAGE_HIGH',           'S32', 'FC03', 1, '',    '',       'RO', 'DEA L2 voltage (high word)'),
    ('0x03F2', 0x03F2, 'DEA_L3_VOLTAGE_LOW',            'S32', 'FC03', 2, 'V',   '0.1',   'RO', 'DEA L3 voltage (low word)'),
    ('0x03F3', 0x03F3, 'DEA_L3_VOLTAGE_HIGH',           'S32', 'FC03', 1, '',    '',       'RO', 'DEA L3 voltage (high word)'),
    ('0x03F4', 0x03F4, 'DEA_TOTAL_ACTIVE_POWER_LOW',    'S32', 'FC03', 2, 'kW',  '0.1',   'RO', 'DEA total active power (low word)'),
    ('0x03F5', 0x03F5, 'DEA_TOTAL_ACTIVE_POWER_HIGH',   'S32', 'FC03', 1, '',    '',       'RO', 'DEA total active power (high word)'),
    ('0x03F6', 0x03F6, 'DEA_TOTAL_REACTIVE_POWER_LOW',  'S32', 'FC03', 2, 'Var', '1',     'RO', 'DEA total reactive power (low word)'),
    ('0x03F7', 0x03F7, 'DEA_TOTAL_REACTIVE_POWER_HIGH', 'S32', 'FC03', 1, '',    '',       'RO', 'DEA total reactive power (high word)'),
    ('0x03F8', 0x03F8, 'DEA_POWER_FACTOR_LOW',          'S32', 'FC03', 2, '',    '0.001', 'RO', 'DEA power factor (low word)'),
    ('0x03F9', 0x03F9, 'DEA_POWER_FACTOR_HIGH',         'S32', 'FC03', 1, '',    '',       'RO', 'DEA power factor (high word)'),
    ('0x03FA', 0x03FA, 'DEA_FREQUENCY_LOW',             'S32', 'FC03', 2, 'Hz',  '0.1',   'RO', 'DEA frequency (low word)'),
    ('0x03FB', 0x03FB, 'DEA_FREQUENCY_HIGH',            'S32', 'FC03', 1, '',    '',       'RO', 'DEA frequency (high word)'),
    ('0x03FC', 0x03FC, 'DEA_STATUS_FLAG_LOW',           'S32', 'FC03', 2, '',    '1',     'RO', 'DEA status flag (low word)'),
    ('0x03FD', 0x03FD, 'DEA_STATUS_FLAG_HIGH',          'S32', 'FC03', 1, '',    '',       'RO', 'DEA status flag (high word)'),
]

# ---------------------------------------------------------------------------
# Supplement address map — manufacturer-specific known addresses (v1.4.1)
# Used in Stage 2 as Priority-4 fallback when Solarize/reference matching fails.
# Clusters: addresses are registered in manufacturer-specific blocks. Stage 2
# injects missing cluster members when ≥3 cluster addresses are detected in Stage 1.
# ---------------------------------------------------------------------------

_SUPPLEMENT_ADDR_MAP = {
    # Sungrow SG series measurement block (from sungrow_mm_registers.py)
    0x139A: 'MPPT1_VOLTAGE',
    0x139B: 'MPPT1_CURRENT',
    0x139C: 'MPPT2_VOLTAGE',
    0x139D: 'MPPT2_CURRENT',
    0x139E: 'MPPT3_VOLTAGE',
    0x139F: 'MPPT3_CURRENT',
    0x13A0: 'MPPT4_VOLTAGE',
    0x13A1: 'MPPT4_CURRENT',
    0x13A2: 'L1_VOLTAGE',
    0x13A3: 'L2_VOLTAGE',
    0x13A4: 'L3_VOLTAGE',
    0x13A5: 'L1_CURRENT',
    0x13A6: 'L2_CURRENT',
    0x13A7: 'L3_CURRENT',
    0x13A8: 'GRID_TOTAL_ACTIVE_POWER_LOW',
    0x13AC: 'POWER_FACTOR',
    0x13AD: 'FREQUENCY',
    0x13AE: 'INVERTER_MODE',
    0x13B6: 'TOTAL_ENERGY_LOW',
    0x13B8: 'TODAY_ENERGY_LOW',
    0x13BA: 'PV_TOTAL_INPUT_POWER_LOW',
    # EKOS EK series fault registers
    0x754C: 'ERROR_CODE1',
    0x754D: 'ERROR_CODE2',
}

# Cluster member sets — inject missing members when ≥3 from a cluster are detected
_SUPPLEMENT_CLUSTERS = [
    # Sungrow SG measurement cluster
    frozenset([0x139A, 0x139B, 0x139C, 0x139D, 0x139E, 0x139F,
               0x13A0, 0x13A1, 0x13A2, 0x13A3, 0x13A4, 0x13A5,
               0x13A6, 0x13A7, 0x13A8, 0x13AC, 0x13AD, 0x13AE,
               0x13B6, 0x13B8, 0x13BA]),
]

# Name-based status keyword matching → InverterMode state
_STATUS_KEYWORD_MAP = [
    (['initial', 'init', 'power on', 'starting'],               'INITIAL',  0x00),
    (['standby', 'wait', 'idle', 'ready'],                      'STANDBY',  0x01),
    (['on-grid', 'on grid', 'grid-connected', 'normal', 'grid'], 'ON_GRID',  0x03),
    (['off-grid', 'off grid', 'island'],                        'OFF_GRID', 0x04),
    (['fault', 'error', 'alarm', 'failure'],                    'FAULT',    0x05),
    (['shutdown', 'shut down', 'stop', 'sleep', 'night'],       'SHUTDOWN', 0x09),
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required.\n"
            "Install with: pip install openpyxl"
        )


def _to_upper_snake(name):
    """Convert a field name to UPPER_SNAKE_CASE."""
    s = re.sub(r'[()（）\[\]【】]', '', str(name))
    s = re.sub(r'[\s\-./·:]+', '_', s)
    s = re.sub(r'([a-z])([A-Z])', r'\1_\2', s)
    s = re.sub(r'[^A-Za-z0-9_]', '', s)
    s = re.sub(r'_+', '_', s).strip('_')
    result = s.upper()
    if result and result[0].isdigit():
        result = 'REG_' + result
    return result


# ---------------------------------------------------------------------------
# EKOS Korean → English definition name translation (v1.4.0)
# ---------------------------------------------------------------------------

_EKOS_KO_EN_MAP = {
    # Compound words first (longest match wins)
    '누적발전량': 'Total Generated Energy', '일일발전량': 'Daily Generated Energy',
    '금일발전량': 'Today Generated Energy', '현재시각': 'Current Time',
    '유효전력': 'Active Power', '무효전력': 'Reactive Power', '피상전력': 'Apparent Power',
    '계통전압': 'Grid Voltage', '계통전류': 'Grid Current', '계통주파수': 'Grid Frequency',
    '출력전압': 'Output Voltage', '출력전류': 'Output Current', '출력전력': 'Output Power',
    '입력전압': 'Input Voltage', '입력전류': 'Input Current', '입력전력': 'Input Power',
    '선간전압': 'Line Voltage', '상전압': 'Phase Voltage',
    '운전상태': 'Operating Status', '운전모드': 'Operating Mode', '운전중': 'Operating',
    '방열판온도': 'Heatsink Temperature', '모듈온도': 'Module Temperature',
    '내부온도': 'Internal Temperature', '외부온도': 'Ambient Temperature',
    '태양전지': 'PV', '태양광': 'PV', '일발전량': 'Today Generated Energy',
    '과전압': 'Overvoltage', '저전압': 'Undervoltage', '과전류': 'Overcurrent',
    '정격출력': 'Rated Output', '최대전력': 'Maximum Power', '최소전력': 'Minimum Power',
    '오류코드': 'Error Code', '고장코드': 'Fault Code',
    '발전유무': 'Generation Status', '발전량': 'Generated Energy',
    '시스템동작': 'System Operation',
    # Single-term / shorter (order matters — longer first already done above)
    '전압': 'Voltage', '전류': 'Current', '전력': 'Power', '역률': 'Power Factor',
    '주파수': 'Frequency', '에너지': 'Energy', '발전': 'Generation',
    # Phases / Lines
    '상': 'Phase', 'R상': 'L1', 'S상': 'L2', 'T상': 'L3',
    'A상': 'L1', 'B상': 'L2', 'C상': 'L3',
    # Temperature
    '온도': 'Temperature',
    # Status / Control
    '상태': 'Status', '모드': 'Mode', '제어': 'Control', '운전': 'Operation',
    '시작': 'Start', '정지': 'Stop', '중지': 'Stop', '이상': 'Fault',
    '경보': 'Alarm', '고장': 'Fault', '오류': 'Error', '알람': 'Alarm', '경고': 'Warning',
    '유무': 'Status', '동작': 'Operation', '시스템': 'System',
    # Device info
    '기기': 'Device', '인버터': 'Inverter', '모델': 'Model', '시리얼': 'Serial',
    '버전': 'Version', '펌웨어': 'Firmware', '정격': 'Rated', '용량': 'Capacity',
    # Grid / AC
    '계통': 'Grid',
    # Protection / Settings
    '최대': 'Maximum', '최소': 'Minimum', '설정': 'Setting', '한계': 'Limit',
    '보호': 'Protection', '차단': 'Trip', '과열': 'Overheat',
    # Time / Misc
    '시각': 'Time', '시간': 'Duration', '날짜': 'Date', '현재': 'Current',
    '통신': 'Communication', '주소': 'Address', '계수': 'Counter', '횟수': 'Count',
    '기동': 'Startup', '단위': 'Unit', '소수점': 'Decimal', '스케일': 'Scale',
    '누적': 'Total', '일일': 'Daily', '금일': 'Today',
    '가동': 'Running', '정상': 'Normal', '최근': 'Latest',
    '정지시각': 'Stop Time', '발전시간': 'Generation Duration',
    '최근정지': 'Latest Stop', '정상동작': 'Normal Operation',
    # PV-specific
    '일사량': 'Solar Irradiance', '수평면': 'Horizontal', '경사면': 'Inclined',
    '외기': 'Outdoor', '지락': 'Ground Fault', '잔류전류': 'Residual Current',
    '잔류': 'Residual', '불균형': 'Imbalance', '정전': 'Power Outage',
    '전체': 'Total All', '채널': 'Channel', '접속함': 'Junction Box',
    '일체형': 'Integrated', '컨버터': 'Converter', '센서': 'Sensor',
    '과주파수': 'Over Frequency', '저주파수': 'Under Frequency',
    '과전력': 'Over Power',
}


def _translate_korean_to_english(text):
    """Translate Korean terms in a definition string to English.

    Replaces known Korean inverter/electrical terms using _EKOS_KO_EN_MAP.
    Leaves non-Korean text (ASCII) unchanged.
    Falls back to a best-effort romanization for unknown Korean words.
    """
    if not text:
        return text
    # If text contains no Korean characters, return as-is
    if not any('\uAC00' <= ch <= '\uD7A3' or '\u1100' <= ch <= '\u11FF' for ch in text):
        return text

    result = text
    # Apply longest-match substitutions first; wrap replacement with spaces so
    # adjacent Korean words don't merge after translation (e.g. 발전시작 → Generation Start)
    for ko, en in sorted(_EKOS_KO_EN_MAP.items(), key=lambda x: -len(x[0])):
        result = result.replace(ko, f' {en} ')

    # Any remaining Korean (AC00-D7A3) → strip and keep ASCII parts
    import unicodedata
    cleaned_parts = []
    word = []
    for ch in result:
        if '\uAC00' <= ch <= '\uD7A3' or '\u1100' <= ch <= '\u11FF':
            word.append(ch)
        else:
            if word:
                cleaned_parts.append('KO_' + ''.join(word)[:6])
                word = []
            cleaned_parts.append(ch)
    if word:
        cleaned_parts.append('KO_' + ''.join(word)[:6])
    result = ''.join(cleaned_parts).strip()

    # Collapse excess whitespace
    result = re.sub(r'\s+', ' ', result).strip()
    return result if result else text


def _addr_to_int(addr_hex, addr_dec):
    """Convert hex string or decimal to int address."""
    if addr_hex:
        try:
            return int(str(addr_hex).strip(), 16)
        except ValueError:
            pass
    if addr_dec is not None:
        try:
            return int(addr_dec)
        except (ValueError, TypeError):
            pass
    return None


def _extract_pdf_text(pdf_path):
    """Extract full text from PDF using PyMuPDF or pdfminer.six."""
    try:
        import fitz
        doc = fitz.open(pdf_path)
        pages = []
        try:
            for page in doc:
                pages.append(page.get_text())
        finally:
            doc.close()
        return '\n'.join(pages)
    except ImportError:
        pass

    try:
        from pdfminer.high_level import extract_text
        return extract_text(pdf_path)
    except ImportError:
        pass

    return ''


def _extract_status_codes_from_text(full_text):
    """Best-effort extraction of InverterMode code table from PDF text.

    Specifically targets the Solarize Inverter Mode Table appendix (0x101D):
      Header: "Inverter Mode Table (0x101D)"
      Format: hex-code line then description line (two-column table split by PyMuPDF)
        0x00  /  Initial mode
        0x01  /  Standby mode  ...

    Falls back to a generic "code - description" format for non-Solarize PDFs.
    """
    statuses = []
    lines = full_text.split('\n')
    in_section = False
    just_added_hex = False   # True on the line immediately after a hex code

    for i, line in enumerate(lines):
        s = line.strip()

        # Detect the specific InverterMode Table appendix section
        # Must contain BOTH "Inverter Mode Table" and a register address reference,
        # or at minimum "Inverter Mode Table" at a reasonable length.
        # Avoid matching the register definition table rows that contain "Inverter Mode".
        is_inv_table_header = (
            re.search(r'Inverter\s+Mode\s+Table', s, re.IGNORECASE)
            and re.search(r'0x101D|Table', s, re.IGNORECASE)
            and not re.search(r'^\d+\s', s)   # not a numbered register row
            and len(s) < 80
        )
        if is_inv_table_header:
            in_section = True
            just_added_hex = False
            statuses = []  # reset — keep only the last (appendix) table
            continue

        if not in_section:
            continue

        # Skip known header/label lines
        if re.match(r'^(No\.?|Description|Index|Value|Code)$', s, re.IGNORECASE):
            just_added_hex = False
            continue

        # If the previous line was a hex code, this line is the description — skip end-check
        if just_added_hex:
            just_added_hex = False
            continue

        # Solarize format: hex code on its own line (e.g. "0x00", "0x09")
        # Range limited to 0x00-0xFF — inverter mode values never exceed one byte
        m_hex = re.match(r'^(0x[0-9A-Fa-f]{1,2})$', s)
        if m_hex:
            code_int = int(m_hex.group(1), 16)
            desc = ''
            for j in range(i + 1, min(i + 4, len(lines))):
                nxt = lines[j].strip()
                if nxt and not re.match(r'^0x[0-9A-Fa-f]', nxt) and not re.match(r'^(No\.?|Description)$', nxt, re.IGNORECASE):
                    desc = nxt[:80]
                    break
            if desc:
                statuses.append({'code': code_int, 'name': desc[:40],
                                 'description': desc, 'bit_def': ''})
                just_added_hex = True
            continue

        # Generic format: "0 - Initial state" or "0x00 : Initial mode"
        m_gen = re.match(r'^(0x[0-9A-Fa-f]{1,2}|\d{1,2})\s*[-:=\t]\s*(.+)', s)
        if m_gen:
            raw = m_gen.group(1)
            code_int = int(raw, 16) if raw.startswith('0x') else int(raw)
            desc = m_gen.group(2).strip()[:80]
            statuses.append({'code': code_int, 'name': desc[:40],
                             'description': desc, 'bit_def': ''})
            continue

        # End of section: non-empty line that doesn't look like table content
        if s and not re.match(r'^[\d\-\*#]', s) and len(statuses) >= 3:
            in_section = False

    return statuses


def _extract_error_codes_from_text(full_text, manufacturer=''):
    """Extract bit-field error code tables from PDF text (v1.4.0 multi-brand).

    Handles multiple manufacturer formats:

    1. Solarize (triplet format): each row split across 3 lines:
         Bit number line (e.g. "0"), No (global) line (e.g. "16"),
         Description line (e.g. "Inverter over dc-bias current")
       Header: "Error Code TableN (0xADDR)"

    2. Kstar / generic: "Fault Code N : Description" or "N - Description" tables
       Header: "Fault Code Table", "Alarm Code", "Error List"

    3. Huawei: alarm table with columns Code|Name|Cause|Suggestion
       Header: "Alarm/Warning List", "Fault Information"

    4. Sungrow: fault bit table with "Alarm ID | Alarm Name | Alarm Level"
       Header: "Alarm List", "Fault Code"

    Returns one entry per fault across all detected sections.
    """
    errors = []
    lines = full_text.split('\n')

    # ---- Pass 1: locate each "Error Code TableN (0xADDR)" section (Solarize) ----
    sections = []          # [(reg_addr_str, [content_lines])]
    cur_addr = None
    cur_lines = []

    for line in lines:
        s = line.strip()
        m_hdr = re.search(r'Error\s+Code\s+Table\d*\s*\(?(0x[0-9A-Fa-f]+)\)?',
                          s, re.IGNORECASE)
        if m_hdr:
            if cur_addr is not None:
                sections.append((cur_addr, cur_lines))
            cur_addr = m_hdr.group(1).upper()
            cur_lines = []
            continue
        if cur_addr is not None:
            # Stop collecting when next major section header starts
            if re.search(r'(Regulation\s+code|Fuse\s+Open|Inverter\s+Mode|Reference\s+doc)',
                         s, re.IGNORECASE):
                sections.append((cur_addr, cur_lines))
                cur_addr = None
                cur_lines = []
            else:
                cur_lines.append(s)

    if cur_addr is not None:
        sections.append((cur_addr, cur_lines))

    # ---- Pass 2: parse each section's lines into (bit, no, description) ----
    for reg_addr, sec_lines in sections:
        # Filter out blank lines and known header words
        data_lines = [l for l in sec_lines
                      if l and l not in ('Bit', 'No', 'Description', '?')]
        # Group into triplets: [bit_line, no_line, desc_line]
        # bit_line is a small integer 0-15; no_line is a larger integer; desc is text
        i = 0
        while i < len(data_lines):
            # Try to match a triplet
            if i + 2 < len(data_lines):
                a, b, c = data_lines[i], data_lines[i+1], data_lines[i+2]
                if (re.match(r'^\d{1,2}$', a) and re.match(r'^\d{1,3}$', b)
                        and not re.match(r'^\d+$', c)):
                    bit_num = int(a)
                    no_num  = int(b)
                    desc    = c[:80]
                    errors.append({
                        'code':        f'{reg_addr} Bit{bit_num}',
                        'name':        f'Bit{bit_num}_{no_num}',
                        'description': desc,
                        'cause':       f'{reg_addr} bit {bit_num} (Error No.{no_num})',
                        'solution':    '',
                    })
                    i += 3
                    continue
            # Single-line fallback: "0 - description" or "0x.. - description"
            m = re.match(r'^(0x[0-9A-Fa-f]+|\d{1,6})\s*[-:=\t]\s*(.+)', data_lines[i])
            if m:
                errors.append({
                    'code':        m.group(1),
                    'name':        m.group(2).strip()[:40],
                    'description': m.group(2).strip()[:80],
                    'cause':       reg_addr,
                    'solution':    '',
                })
            i += 1

    # ---- Pass 2: generic multi-brand fault table extraction (v1.4.0) ----
    # Covers Kstar, Huawei, Sungrow fault tables not matched by Solarize triplet parser

    _FAULT_SECTION_START = re.compile(
        r'(Fault\s+(Code|List|Information|Table)|Alarm\s+(List|Information|Code|Warning)|'
        r'Error\s+(List|Information|Table)|Warning\s+(List|Code)|'
        r'Protection\s+List|Abnormal\s+Table)',
        re.IGNORECASE
    )
    _FAULT_SECTION_END = re.compile(
        r'(Setting\s+Param|Device\s+Info|Communication\s+Param|'
        r'Running\s+Inform|Realtime\s+Data|Register\s+List|'
        r'Appendix|Reference\s+doc|Table\s+of\s+Content)',
        re.IGNORECASE
    )

    in_fault = False
    fault_lines = []
    for line in lines:
        s = line.strip()
        if not in_fault:
            if _FAULT_SECTION_START.search(s) and len(s) < 100:
                in_fault = True
                fault_lines = []
            continue
        if _FAULT_SECTION_END.search(s) and len(s) < 100:
            in_fault = False
            # Parse collected fault_lines
            _parse_generic_fault_lines(fault_lines, errors)
            fault_lines = []
            continue
        fault_lines.append(s)

    if fault_lines:
        _parse_generic_fault_lines(fault_lines, errors)

    # Deduplicate errors by code string
    seen = set()
    deduped = []
    for e in errors:
        key = str(e.get('code', '')) + str(e.get('description', ''))[:20]
        if key not in seen:
            seen.add(key)
            deduped.append(e)
    return deduped


def _parse_generic_fault_lines(fault_lines, errors):
    """Parse generic fault/alarm table lines into error entries (v1.4.0).

    Supports formats:
      "N : Description"         (Kstar, code-description)
      "N - Description"         (generic)
      "ID Name Level/Cause"     (Sungrow 3-column, aligned)
      "0xNNNN  AlarmName  ..."  (Huawei)
    """
    # Filter out header and empty lines
    _SKIP = re.compile(
        r'^(Fault\s+Code|Alarm\s+ID|Alarm\s+Name|Error\s+Code|Error\s+Name|'
        r'No\.?|Description|Level|Cause|Suggestion|Warning|Category|'
        r'Bit|Register|ID|Code|Name|Type|Solution|Remark|Note)$',
        re.IGNORECASE
    )
    data = [l for l in fault_lines if l and not _SKIP.match(l)
            and not re.search(r'\.{5,}\s*\d+\s*$', l)]

    i = 0
    while i < len(data):
        line = data[i]

        # Format A: "N : Description" or "N - Description" or "0xN : Desc"
        m = re.match(r'^(0x[0-9A-Fa-f]+|\d{1,6})\s*[-:]\s*(.{3,})', line)
        if m:
            code = m.group(1)
            desc = m.group(2).strip()[:80]
            errors.append({'code': code, 'name': desc[:40], 'description': desc,
                           'cause': 'Fault', 'solution': ''})
            i += 1
            continue

        # Format B: standalone number followed by name on next line
        if re.match(r'^(0x[0-9A-Fa-f]+|\d{1,5})$', line) and i + 1 < len(data):
            nxt = data[i + 1]
            # next line must be descriptive text (not another number or short code)
            if not re.match(r'^[\d\.]+$', nxt) and len(nxt) > 3:
                code = line
                desc = nxt[:80]
                errors.append({'code': code, 'name': desc[:40], 'description': desc,
                               'cause': 'Fault', 'solution': ''})
                i += 2
                continue

        # Format C: Huawei "0xNNNN  AlarmName" (hex code then text, same line)
        m2 = re.match(r'^(0x[0-9A-Fa-f]{4,})\s+(.{4,})', line)
        if m2:
            code = m2.group(1).upper()
            desc = m2.group(2).strip()[:80]
            errors.append({'code': code, 'name': desc[:40], 'description': desc,
                           'cause': 'Alarm', 'solution': ''})
            i += 1
            continue

        i += 1


def _extract_bit_table_from_text(full_text, section_header_pattern, end_patterns=None):
    """범용 비트필드 테이블 추출기.

    PDF에서 다음 형식의 테이블을 파싱한다 (PyMuPDF가 3행으로 분리):
      Bit_number
      Description text (possibly multi-line)
      Value definition (e.g. "0:normal, 1: Fuse open")

    section_header_pattern : 섹션 시작 헤더 regex (str)
    end_patterns           : 섹션 종료 트리거 regex 목록 (list[str]), None이면 기본값 사용

    Returns list of {'bit': int, 'description': str, 'values': str}
    """
    if end_patterns is None:
        end_patterns = [
            r'Inverter Mode Table', r'Fuse open Check Mask', r'Reference document',
            r'REMS register', r'REMS Error Code', r'Additional description',
            r'Fuse Open Data', r'Illegal code Table',
        ]

    rows = []
    lines = full_text.split('\n')
    in_section = False

    # Collect section lines (skip TOC entries and page header/footer noise)
    sec_lines = []
    for line in lines:
        s = line.strip()
        # Skip TOC entries
        if re.search(r'\.{5,}\s*\d+\s*$', s):
            continue
        if not in_section:
            if re.search(section_header_pattern, s, re.IGNORECASE) and len(s) < 120:
                in_section = True
                sec_lines = []
                continue
        else:
            # If the same header appears again, restart (take the later occurrence)
            if re.search(section_header_pattern, s, re.IGNORECASE) and len(s) < 120:
                sec_lines = []
                continue
            # End section
            ended = any(re.search(p, s, re.IGNORECASE) and len(s) < 100
                        for p in end_patterns)
            if ended:
                in_section = False
                continue
            sec_lines.append(s)

    # Filter out blank, page-noise lines
    _PAGE_NOISE = re.compile(
        r'^(Project No\.|NA|MODBUS PROTOCOL|Security|Confidential|'
        r'Project Name|Version|1\.\d\.\d|APD CONFIDENTIAL|Bit|Description|\d{1,2}\s*/\s*35)$',
        re.IGNORECASE
    )
    data = [l for l in sec_lines if l and not _PAGE_NOISE.match(l)]

    # Parse triplets: bit_line / description_line(s) / value_line
    i = 0
    while i < len(data):
        # Expect a bit number (1-2 digits)
        if re.match(r'^\d{1,2}$', data[i]):
            bit_num = int(data[i])
            desc_parts = []
            val_parts = []
            j = i + 1
            # Collect description and value lines until next bit number
            while j < len(data) and not re.match(r'^\d{1,2}$', data[j]):
                line = data[j]
                # Value definition lines contain "0:" or "1:"
                if re.match(r'^0\s*:', line) or re.match(r'^[01]\s*:', line):
                    val_parts.append(line)
                else:
                    desc_parts.append(line)
                j += 1
            desc = ' '.join(desc_parts).strip()
            val  = ' / '.join(val_parts).strip()
            if desc:
                rows.append({'bit': bit_num, 'description': desc, 'values': val})
            i = j
        else:
            i += 1

    return rows


def _extract_rems_registers_from_text(full_text):
    """REMS register value 테이블 추출 (No./Description/Length/Unit 4열).

    Format (PyMuPDF 분리):
      No_number / Description / Length_number / Unit
    """
    rows = []
    lines = full_text.split('\n')
    in_section = False
    sec_lines = []

    for line in lines:
        s = line.strip()
        if re.search(r'\.{5,}\s*\d+\s*$', s):
            continue
        if not in_section:
            if re.match(r'^REMS register value', s, re.IGNORECASE):
                in_section = True
                sec_lines = []
                continue
        else:
            if re.search(r'REMS Error Code Table', s, re.IGNORECASE) and len(s) < 60:
                in_section = False
                continue
            sec_lines.append(s)

    _PAGE_NOISE = re.compile(
        r'^(Project No\.|NA|MODBUS PROTOCOL|Security|Confidential|'
        r'Project Name|Version|1\.\d\.\d|APD CONFIDENTIAL|No\.?|Description|'
        r'Length\s*\(Byte.*\)|Unit|\d{1,2}\s*/\s*35)$',
        re.IGNORECASE
    )
    data = [l for l in sec_lines if l and not _PAGE_NOISE.match(l)]

    # Each REMS register entry is exactly 4 items (PyMuPDF renders columns as separate lines):
    #   No(digit)  /  Description(text)  /  Length_bytes(digit)  /  Unit(text/empty)
    # Strategy: find runs of [digit, text, digit, text-or-digit] groups.
    i = 0
    while i < len(data):
        # Must start with a sequential number (1-digit or 2-digit)
        if not re.match(r'^\d{1,2}$', data[i]):
            i += 1
            continue
        no_num = int(data[i])
        # Next line: description (must be non-pure-digit text)
        if i + 1 >= len(data) or re.match(r'^\d+$', data[i + 1]):
            i += 1
            continue
        desc = data[i + 1]
        # Next: length in bytes (should be a small number: 2, 4, or 8)
        length = ''
        unit = ''
        if i + 2 < len(data) and re.match(r'^\d{1,2}$', data[i + 2]):
            length = data[i + 2]
            # Next: unit (text, may be empty or absent for last entry)
            if i + 3 < len(data) and not re.match(r'^\d{1,2}$', data[i + 3]):
                unit = data[i + 3]
                i += 4
            else:
                i += 3
        else:
            i += 2
        rows.append({'no': no_num, 'description': desc,
                     'length_bytes': length, 'unit': unit})

    return rows


def _extract_sungrow_registers_from_text(full_text):
    """Sungrow-specific register table parser (v1.4.0).

    Sungrow PDFs use a 5-6 column table:
      Register Address | Register Name (definition) | Data Format | Range | Unit | Description

    The generic parser incorrectly picks up the Range column (e.g. "0-1100") as the
    definition name.  This parser identifies rows by:
      1. Leading hex or decimal register address (0x4901 or 4901 or 5001)
      2. Next non-empty token is the register NAME (text, not a pure number or range)
      3. Skip data format token (U16/U32/S16/S32/etc.)
      4. Skip range token (digits with ~/-/%)
      5. Optional unit token

    Returns: list of register dicts compatible with stage_pipeline format.
    """
    # Data format types used by Sungrow
    _DATA_TYPES = {'U16', 'S16', 'U32', 'S32', 'STRING', 'UTF8', 'FLOAT', 'BOOL'}
    _RANGE_PAT = re.compile(r'^[\d\.]+\s*[~\-]\s*[\d\.]+$|^0x[0-9A-Fa-f]+\s*[~\-]')
    _ADDR_PAT = re.compile(r'^(0x[0-9A-Fa-f]{4,5}|\d{4,5})$')

    regs = []
    lines = [l.strip() for l in full_text.split('\n')]

    # Table sections in Sungrow docs often have headers like:
    #   "Register Address", "Register Name", "Data Format", "Range", "Unit", "Description"
    # PyMuPDF renders each cell as a separate line, so we need to group by runs.

    # Strategy: scan all lines for address-like tokens, then collect the surrounding context
    section = 'Realtime Data'
    section_keywords = {
        'running information': 'Running Information',
        'alarm information': 'Alarm Information',
        'setting parameters': 'Setting Parameters',
        'device information': 'Device Information',
        'mppt': 'MPPT Data',
        'string': 'String Data',
        'control': 'Control',
        'power quality': 'Power Quality',
        'energy statistics': 'Energy Statistics',
        'grid': 'Grid Data',
        'dc input': 'DC Input',
        'output': 'Output Data',
    }

    i = 0
    while i < len(lines):
        line = lines[i]

        # Detect section header
        ll = line.lower()
        for kw, sec_name in section_keywords.items():
            if kw in ll and len(line) < 80 and not _ADDR_PAT.match(line):
                section = sec_name
                break

        # Look for a register address line
        if not _ADDR_PAT.match(line):
            i += 1
            continue

        # Parse address
        try:
            if line.startswith('0x') or line.startswith('0X'):
                addr = int(line, 16)
            else:
                addr_int = int(line)
                # Sungrow decimal addresses are typically 4-5 digit values >= 4865 (0x1301)
                if addr_int < 100:
                    i += 1
                    continue
                addr = addr_int
        except ValueError:
            i += 1
            continue

        addr_hex = f'0x{addr:04X}'

        # Gather next 5 lines for context
        ctx = []
        for j in range(i + 1, min(i + 8, len(lines))):
            if lines[j]:
                ctx.append(lines[j])
            if len(ctx) >= 5:
                break

        if not ctx:
            i += 1
            continue

        # Try to extract: definition, data_type, range, unit from ctx
        definition = ''
        dtype = 'U16'
        unit = ''
        rw = 'RO'

        for tok in ctx:
            tok_u = tok.upper().strip()
            # Data type token
            if tok_u in _DATA_TYPES or re.match(r'^(U|S)(16|32)$', tok_u) or re.match(r'^STRING\d*$', tok_u):
                if not dtype or dtype == 'U16':
                    dtype = tok_u
                continue
            # Range token (0~1100, 0-100%, 0x0000~0xFFFF)
            if _RANGE_PAT.match(tok):
                continue
            # Read/Write specifier
            if tok_u in ('R', 'W', 'R/W', 'RO', 'WO', 'RW', 'READ', 'WRITE', 'READ/WRITE'):
                rw_map = {'R': 'RO', 'W': 'WO', 'R/W': 'RW', 'READ': 'RO', 'WRITE': 'WO',
                          'READ/WRITE': 'RW', 'RO': 'RO', 'WO': 'WO', 'RW': 'RW'}
                rw = rw_map.get(tok_u, 'RO')
                continue
            # Unit (single/short token)
            if tok in ('W', 'kW', 'MW', 'V', 'A', 'Hz', '%', 'kWh', 'MWh', 'Wh',
                       'var', 'kvar', 'VA', 'kVA', 'Ω', 'C', '°C', 'min', 's', 'h',
                       'rpm', 'dBm', 'ms') and not definition:
                unit = tok
                continue
            # Unit after definition is set
            if tok in ('W', 'kW', 'MW', 'V', 'A', 'Hz', '%', 'kWh', 'MWh', 'Wh',
                       'var', 'kvar', 'VA', 'kVA', 'Ω', 'C', '°C', 'min', 's', 'h',
                       'rpm', 'dBm', 'ms') and definition:
                unit = tok
                continue
            # Pure number → skip (bit count, enum value, etc.)
            if re.match(r'^[\d\.]+$', tok):
                continue
            # Otherwise: likely the definition name
            if not definition and len(tok) > 1:
                definition = tok

        if not definition:
            i += 1
            continue

        # Skip if definition looks like a range (e.g. "0-1100", "0~9999")
        if _RANGE_PAT.match(definition) or re.match(r'^[\d\.]+$', definition):
            i += 1
            continue

        regs.append({
            'section':     section,
            'definition':  definition,
            'address':     addr,
            'address_hex': addr_hex,
            'regs':        2 if dtype in ('U32', 'S32') else 1,
            'type':        dtype,
            'unit':        unit,
            'rw':          rw,
            'comment':     '',
        })
        i += 1

    return regs


def _extract_regulation_codes_from_text(full_text):
    """Extract regulation/grid code enumeration table (0x5101) from PDF text.

    Solarize PDF format (two-column table split by PyMuPDF):
      "Regulation code Table (0x5101)" header line
      "No." / "Regulation code" column header lines
      Then alternating:  "0" line  /  "0xFFFF : UNKNOWN" line
                         "1" line  /  "0x0001 : AU (...)" line  ...

    Skips Table-of-Contents entries (lines with many dots "....." + page num).
    """
    regs = []
    lines = full_text.split('\n')
    in_section = False

    for i, line in enumerate(lines):
        s = line.strip()

        # Skip Table of Contents entries (contain 5+ dots followed by page number)
        if re.search(r'\.{5,}\s*\d+\s*$', s):
            continue

        # Must start with "Regulation code Table" — avoids matching note sentences
        if re.match(r'^Regulation\s+code\s+Table\b', s, re.IGNORECASE):
            in_section = True
            regs = []
            continue
        if not in_section:
            continue
        # End section when next appendix section header appears
        if re.search(r'(Fuse\s+Open|Inverter\s+Mode|Illegal\s+code|Reference\s+doc)',
                     s, re.IGNORECASE) and len(s) < 80:
            in_section = False
            continue
        # Skip header lines
        if re.match(r'^(No\.?|Regulation\s+code|Code)$', s, re.IGNORECASE):
            continue
        # Match "0x0001 : AU (...description...)" or "0xFFFF : UNKNOWN"
        m = re.match(r'^(0x[0-9A-Fa-f]{4})\s*[:\-]\s*(.+)', s)
        if m:
            code_hex = m.group(1).upper()
            # Strip non-ASCII display-only chars from description
            desc = m.group(2).strip().encode('ascii', 'replace').decode('ascii')
            regs.append({'code': code_hex, 'description': desc})
            continue
        # Skip pure-number lines (the "No." column rendered on a separate line)
        if re.match(r'^\d+$', s):
            continue

    return regs


def _extract_status_codes_enhanced(full_text, manufacturer=''):
    """Enhanced status code extraction for all manufacturers (v1.4.0).

    Supplements the Solarize-specific _extract_status_codes_from_text() with
    generic patterns used by Kstar, Huawei, Sungrow, and EKOS:

      Kstar:   "Operating Status" table with "N  Description" rows
      Huawei:  "Device Status" register code table (hex/decimal values)
      Sungrow: "Running Status" table with value-description rows
      Generic: Any table near keywords "state", "status", "mode" with code-desc pairs

    Returns list of status dicts: {code, name, description, bit_def}
    """
    statuses = []

    # --- Pattern 1: "N - Description" / "N : Description" near state/mode headers ---
    _STATUS_HDR = re.compile(
        r'(Operating\s+Stat|Running\s+Stat|Device\s+Stat|Inverter\s+Stat|'
        r'Work\s+Stat|System\s+Stat|State\s+Code|Status\s+Code|'
        r'Mode\s+Table|Operation\s+Mode|Status\s+Table|Run\s+Mode)',
        re.IGNORECASE
    )
    _STATUS_END = re.compile(
        r'(Fault|Alarm|Error|Setting\s+Param|Register\s+List|'
        r'Appendix|Reference|Communication|Device\s+Info)',
        re.IGNORECASE
    )

    lines = full_text.split('\n')
    in_section = False
    sec_lines = []

    for line in lines:
        s = line.strip()
        # Skip TOC entries
        if re.search(r'\.{5,}\s*\d+\s*$', s):
            continue
        if not in_section:
            if _STATUS_HDR.search(s) and len(s) < 100:
                # Avoid matching the Solarize InverterMode Table (handled by original fn)
                if not re.search(r'Inverter\s+Mode\s+Table', s, re.IGNORECASE):
                    in_section = True
                    sec_lines = []
            continue
        if _STATUS_END.search(s) and len(s) < 80:
            in_section = False
            _parse_status_lines(sec_lines, statuses)
            sec_lines = []
            continue
        if len(sec_lines) > 60:  # Safety limit
            in_section = False
            _parse_status_lines(sec_lines, statuses)
            sec_lines = []
            continue
        sec_lines.append(s)

    if sec_lines:
        _parse_status_lines(sec_lines, statuses)

    return statuses


def _parse_status_lines(sec_lines, statuses):
    """Parse status code lines in multiple formats into status dicts."""
    _SKIP = re.compile(
        r'^(No\.?|Code|Value|Description|Status|Mode|State|Index|Name|'
        r'Meaning|Remark|Note|Operating|Running|Device|Inverter)$',
        re.IGNORECASE
    )
    data = [l for l in sec_lines if l and not _SKIP.match(l)]

    seen_codes = {s['code'] for s in statuses}
    i = 0
    while i < len(data):
        line = data[i]

        # Format A: "N : Description" or "N - Description"
        m = re.match(r'^(0x[0-9A-Fa-f]{1,4}|\d{1,3})\s*[-:]\s*(.{3,})', line)
        if m:
            raw = m.group(1)
            code = int(raw, 16) if raw.startswith('0x') else int(raw)
            desc = m.group(2).strip()[:80]
            if code not in seen_codes:
                statuses.append({'code': code, 'name': desc[:40],
                                 'description': desc, 'bit_def': ''})
                seen_codes.add(code)
            i += 1
            continue

        # Format B: bare number then description on next line
        if re.match(r'^(0x[0-9A-Fa-f]{1,4}|\d{1,3})$', line) and i + 1 < len(data):
            nxt = data[i + 1]
            if not re.match(r'^[\d\.]+$', nxt) and len(nxt) > 2:
                raw = line
                code = int(raw, 16) if raw.startswith('0x') else int(raw)
                desc = nxt[:80]
                if code not in seen_codes:
                    statuses.append({'code': code, 'name': desc[:40],
                                     'description': desc, 'bit_def': ''})
                    seen_codes.add(code)
                i += 2
                continue
        i += 1


def _match_status_to_mode(name):
    """Match status name string to InverterMode state name."""
    lower = name.lower()
    for keywords, mode_name, mode_code in _STATUS_KEYWORD_MAP:
        for kw in keywords:
            if kw in lower:
                return mode_name, mode_code
    return 'Unknown', -1


# ---------------------------------------------------------------------------
# Excel styling helpers
# ---------------------------------------------------------------------------

def _make_styles(openpyxl):
    """Return commonly used cell styles."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    hdr_fill  = PatternFill('solid', fgColor='2E75B6')
    hdr_font  = Font(bold=True, color='FFFFFF', size=9)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    sec_fill  = PatternFill('solid', fgColor='D9E1F2')
    sec_font  = Font(bold=True, color='1F3864', size=9)

    auto_fill = PatternFill('solid', fgColor='E2EFDA')   # green  = auto match
    man_fill  = PatternFill('solid', fgColor='FFF2CC')   # yellow = manual
    unm_fill  = PatternFill('solid', fgColor='FCE4D6')   # red    = unmapped

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    body_align = Alignment(vertical='center')
    body_font  = Font(size=9)

    return {
        'hdr_fill': hdr_fill, 'hdr_font': hdr_font, 'hdr_align': hdr_align,
        'sec_fill': sec_fill, 'sec_font': sec_font,
        'auto_fill': auto_fill, 'man_fill': man_fill, 'unm_fill': unm_fill,
        'border': border, 'body_align': body_align, 'body_font': body_font,
    }


def _write_header(ws, headers, styles, row=1):
    """Write a styled header row."""
    for col, title in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=title)
        c.fill   = styles['hdr_fill']
        c.font   = styles['hdr_font']
        c.alignment = styles['hdr_align']
        c.border = styles['border']


def _write_row(ws, row_num, values, styles, fill=None):
    """Write a data row with optional background fill."""
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font   = styles['body_font']
        c.alignment = styles['body_align']
        c.border = styles['border']
        if fill:
            c.fill = fill


# ---------------------------------------------------------------------------
# STAGE 1: PDF (parsed_data) → Excel
# ---------------------------------------------------------------------------

def stage1_extract_to_excel(parsed_data, pdf_path, output_path=None,
                             mode='offline',
                             use_ai=False, api_key=None, model=None,
                             progress_cb=None):
    """
    Stage 1: Export parsed PDF register data to Excel.

    Args:
        parsed_data : dict from parse_modbus_pdf() — may be None
        pdf_path    : Original PDF file path (used for raw text / output path)
        output_path : Destination .xlsx path (auto-generated from pdf_path if None)
        mode        : 'offline' or 'ai' — 'ai' auto-sets use_ai=True if api_key given
        use_ai      : Use Claude API for enhanced extraction (overridden by mode='ai')
        api_key     : Claude API key (required when mode='ai')
        model       : Claude model name
        progress_cb : Callback(msg: str)

    Returns:
        output_path (str)
    """
    openpyxl = _require_openpyxl()

    def _p(msg):
        if progress_cb:
            progress_cb(msg)

    # mode='ai' auto-enables AI extraction when api_key is available
    if mode == 'ai' and api_key and not use_ai:
        use_ai = True

    _p(f"Stage 1: Extracting PDF data to Excel... [mode={mode}]")

    # ── Determine output path ──
    if not output_path:
        base = os.path.splitext(pdf_path or 'output')[0]
        output_path = base + '_registers.xlsx'

    # ── Detect manufacturer ──
    manufacturer = ''
    if parsed_data:
        manufacturer = str(parsed_data.get('manufacturer', '')).lower()
    if not manufacturer and pdf_path:
        pf = os.path.basename(pdf_path).lower()
        if 'sungrow' in pf or 'sg' in pf:
            manufacturer = 'sungrow'
        elif 'kstar' in pf or 'ksg' in pf:
            manufacturer = 'kstar'
        elif 'huawei' in pf or 'sun2000' in pf:
            manufacturer = 'huawei'
        elif 'ekos' in pf or 'ek_' in pf:
            manufacturer = 'ekos'
        elif 'solarize' in pf or 'verterking' in pf or 'vk' in pf:
            manufacturer = 'solarize'

    # ── Collect all registers (offline parser results) ──
    all_regs = []
    if parsed_data:
        all_regs = list(parsed_data.get('all_registers', []))
        _p(f"  Registers from PDF parser: {len(all_regs)}")
    else:
        _p("  No parsed_data supplied — Sheet 1 will be empty.")

    # ── Sungrow dedicated re-parse (v1.4.0) ──
    # Override/supplement offline parser results for Sungrow PDFs because the generic
    # parser captures range values (0-1100) as definition names.
    if 'sungrow' in manufacturer and pdf_path and os.path.isfile(pdf_path):
        _p("  [Sungrow] Running dedicated table parser to fix definition names...")
        try:
            raw_text_sg = _extract_pdf_text(pdf_path)
            sungrow_regs = _extract_sungrow_registers_from_text(raw_text_sg)
            if sungrow_regs:
                _p(f"  [Sungrow] Dedicated parser: {len(sungrow_regs)} registers")
                # Union merge: prefer Sungrow parser (better definitions) over generic
                all_regs, _new, _imp = _union_merge_registers(sungrow_regs, all_regs)
                _p(f"  [Sungrow] After dedicated merge: {len(all_regs)} registers "
                   f"(+{_new} added from generic, {_imp} definitions improved)")
        except Exception as e:
            _p(f"  [Sungrow] Dedicated parser error: {e} — keeping generic results")

    # ── EKOS Korean→English definition translation (v1.4.0) ──
    if 'ekos' in manufacturer:
        _p("  [EKOS] Translating Korean definition names to English...")
        translated = 0
        for r in all_regs:
            orig = r.get('definition', '')
            new_defn = _translate_korean_to_english(orig)
            if new_defn != orig:
                r['definition'] = new_defn
                translated += 1
        _p(f"  [EKOS] Translated {translated} Korean definitions")

    # ── Extract raw text for status/error codes ──
    raw_text = ''
    if pdf_path and os.path.isfile(pdf_path):
        try:
            raw_text = _extract_pdf_text(pdf_path)
        except Exception:
            pass

    # ── Status codes: Solarize-specific + enhanced generic ──
    status_codes = _extract_status_codes_from_text(raw_text)
    if not status_codes or len(status_codes) < 3:
        # Try enhanced generic extraction for non-Solarize brands
        extra_status = _extract_status_codes_enhanced(raw_text, manufacturer)
        if extra_status:
            # Merge: add codes not already in status_codes
            existing_codes = {s['code'] for s in status_codes}
            for sc in extra_status:
                if sc['code'] not in existing_codes:
                    status_codes.append(sc)
                    existing_codes.add(sc['code'])
            _p(f"  Enhanced status extraction: {len(status_codes)} total status codes")

    # ── Error codes: Solarize bit-field + generic multi-brand ──
    error_codes = _extract_error_codes_from_text(raw_text, manufacturer)

    regulation_codes  = _extract_regulation_codes_from_text(raw_text)

    # Additional bit-field tables (Sheet5~Sheet8)
    fuse_open_bits   = _extract_bit_table_from_text(
        raw_text,
        r'^Fuse Open Data Table\s*\(0x1034',
        end_patterns=[r'Inverter Mode Table', r'Fuse open Check Mask',
                      r'Reference document', r'REMS register', r'Additional description'])
    status_bits_03fc = _extract_bit_table_from_text(
        raw_text,
        r'^Additional description\s*\(0x03FC',
        end_patterns=[r'REMS register', r'REMS Error Code', r'Fuse Open',
                      r'Inverter Mode Table', r'Reference document'])
    fuse_mask_bits   = _extract_bit_table_from_text(
        raw_text,
        r'^Fuse open Check Mask Table\s*\(0x513D',
        end_patterns=[r'Reference document', r'Inverter Mode Table',
                      r'REMS register', r'Additional description', r'Fuse Open Data'])
    rems_error_bits  = _extract_bit_table_from_text(
        raw_text,
        r'^REMS Error Code Table$',
        end_patterns=[r'Fuse Open', r'Inverter Mode', r'Additional description',
                      r'Regulation code'])
    rems_regs        = _extract_rems_registers_from_text(raw_text)

    # ── AI independent extraction + union merge (v1.4.0+) ──
    # AI merge runs in BOTH modes when api_key is available:
    # - AI mode: always runs
    # - Offline mode: runs as supplement if api_key provided (improves definitions)
    ai_status_codes = []
    ai_error_codes  = []
    run_ai = (use_ai or (api_key and len(all_regs) < 10)) and pdf_path and os.path.isfile(pdf_path)
    if run_ai and api_key:
        _p("  [AI Stage1] Running independent AI extraction (union merge mode)...")
        try:
            ai_result = _ai_stage1_extract(pdf_path, api_key, model, _p)
            if ai_result and ai_result.get('registers'):
                ai_regs_raw = ai_result['registers']
                all_regs, ai_new, ai_imp = _union_merge_registers(all_regs, ai_regs_raw)
                _p(f"  [AI Stage1] Union merge: {len(all_regs)} registers total "
                   f"(+{ai_new} AI-only, {ai_imp} definitions improved by AI)")
            ai_status_codes = ai_result.get('status_codes', [])
            ai_error_codes  = ai_result.get('error_codes', [])
        except Exception as e:
            _p(f"  [AI Stage1] Extraction failed ({e}) — using offline results only")

    # Merge AI-extracted status/error codes with rule-based results
    if ai_status_codes:
        existing_sc = {s['code'] for s in status_codes}
        for sc in ai_status_codes:
            if sc.get('code') not in existing_sc:
                status_codes.append(sc)
                existing_sc.add(sc.get('code'))
        _p(f"  [AI Stage1] Total status codes after AI merge: {len(status_codes)}")

    if ai_error_codes:
        existing_ec = {(str(e.get('code', '')), str(e.get('description', ''))[:20])
                       for e in error_codes}
        for ec in ai_error_codes:
            key = (str(ec.get('code', '')), str(ec.get('description', ''))[:20])
            if key not in existing_ec:
                error_codes.append(ec)
                existing_ec.add(key)
        _p(f"  [AI Stage1] Total error codes after AI merge: {len(error_codes)}")

    # Fall back to standard InverterMode if nothing found
    if not status_codes:
        status_codes = [{'code': c, 'name': n, 'description': d, 'bit_def': ''}
                        for c, n, d in _INVERTER_MODES]

    # ── Build workbook ──
    _p("  Building Excel workbook...")
    wb     = openpyxl.Workbook()
    styles = _make_styles(openpyxl)

    # ─ Sheet 1: Registers ─
    ws1 = wb.active
    ws1.title = 'Registers'
    ws1.freeze_panes = 'A2'

    s1_headers = ['No', 'Section', 'Address_Hex', 'Address_Dec',
                  'Definition', 'Data_Type', 'FC_Code', 'Registers',
                  'Unit', 'Scale_Factor', 'R/W', 'Description']
    _write_header(ws1, s1_headers, styles)

    # Column widths
    for col, w in zip('ABCDEFGHIJKL', [5, 22, 12, 12, 35, 9, 7, 9, 8, 12, 6, 45]):
        ws1.column_dimensions[col].width = w

    current_section = None
    row_num = 2
    for i, reg in enumerate(all_regs, 1):
        sec  = reg.get('section', '')
        defn = reg.get('definition', '')
        addr = reg.get('address')
        addr_hex = reg.get('address_hex', '')
        if not addr_hex and addr is not None:
            addr_hex = f'0x{int(addr):04X}'
        if addr is None and addr_hex:
            try:
                addr = int(addr_hex, 16)
            except ValueError:
                addr = 0

        regs_n = reg.get('regs', 1)
        dtype  = reg.get('type', 'U16')
        unit   = reg.get('unit', '')
        rw     = reg.get('rw', 'RO')
        comment = reg.get('comment', '')

        # FC code guess
        if rw == 'WO':
            fc = 'FC06'
        elif rw == 'RW':
            fc = 'FC03/FC06'
        else:
            fc = 'FC03'

        # Scale from unit string
        scale = ''
        m = re.match(r'^([0-9]*\.?[0-9]+)\s*[vawkhz%]', unit.lower())
        if m:
            scale = m.group(1)

        # Section divider row
        if sec and sec != current_section:
            current_section = sec
            c = ws1.cell(row=row_num, column=1, value=f'--- {sec} ---')
            c.fill  = styles['sec_fill']
            c.font  = styles['sec_font']
            ws1.merge_cells(start_row=row_num, start_column=1,
                            end_row=row_num, end_column=len(s1_headers))
            row_num += 1

        _write_row(ws1, row_num,
                   [i, sec, addr_hex, addr, defn, dtype, fc,
                    regs_n, unit, scale, rw, comment],
                   styles)
        row_num += 1

    ws1.auto_filter.ref = f'A1:{chr(64+len(s1_headers))}1'

    # ─ Sheet 2: Status Definitions ─
    ws2 = wb.create_sheet('Status_Definitions')
    ws2.freeze_panes = 'A2'
    s2_headers = ['Code', 'Name', 'Description', 'Bit_Definition']
    _write_header(ws2, s2_headers, styles)
    for col, w in zip('ABCD', [8, 28, 50, 40]):
        ws2.column_dimensions[col].width = w

    for i, sc in enumerate(status_codes, 2):
        _write_row(ws2, i,
                   [sc.get('code', ''), sc.get('name', ''),
                    sc.get('description', ''), sc.get('bit_def', '')],
                   styles)

    # Add note row
    note_row = len(status_codes) + 3
    ws2.cell(row=note_row, column=1,
             value='NOTE: Add/edit status codes. Descriptions used in Stage 2 InverterMode mapping.')

    # ─ Sheet 3: Error/Fault Codes (bit-field tables 0x101E/101F/1020) ─
    ws3 = wb.create_sheet('Error_Fault_Codes')
    ws3.freeze_panes = 'A2'
    s3_headers = ['Register', 'Bit', 'Error_No', 'Description', 'Notes']
    _write_header(ws3, s3_headers, styles)
    for col, w in zip('ABCDE', [14, 6, 10, 60, 40]):
        ws3.column_dimensions[col].width = w

    if error_codes:
        for i, ec in enumerate(error_codes, 2):
            code_str = str(ec.get('code', ''))
            # Parse "0x101E Bit3" format
            m_bit = re.match(r'^(0x[0-9A-Fa-f]+)\s+Bit(\d+)$', code_str, re.IGNORECASE)
            if m_bit:
                reg_addr = m_bit.group(1).upper()
                bit_num  = int(m_bit.group(2))
                # Error No. from name field "Bit3_19"
                m_no = re.match(r'^Bit\d+_(\d+)$', ec.get('name', ''))
                err_no = int(m_no.group(1)) if m_no else ''
            else:
                reg_addr = code_str
                bit_num  = ''
                err_no   = ''
            _write_row(ws3, i,
                       [reg_addr, bit_num, err_no,
                        ec.get('description', ''), ec.get('solution', '')],
                       styles)
    else:
        ws3.cell(row=2, column=1, value='(No error codes detected from PDF)')

    # ─ Sheet 4: Grid / Regulation Codes (0x5101) ─
    ws4 = wb.create_sheet('Grid_Codes')
    ws4.freeze_panes = 'A2'
    s4_headers = ['Code_Hex', 'Description']
    _write_header(ws4, s4_headers, styles)
    for col, w in zip('AB', [14, 70]):
        ws4.column_dimensions[col].width = w

    if regulation_codes:
        for i, rc in enumerate(regulation_codes, 2):
            _write_row(ws4, i,
                       [rc.get('code', ''), rc.get('description', '')],
                       styles)
    else:
        ws4.cell(row=2, column=1, value='(No regulation codes detected from PDF)')
    ws4.cell(row=max(len(regulation_codes) + 3, 3), column=1,
             value='NOTE: Register 0x5101 - Grid regulation/country code setting')

    # ─ Sheet 5: Fuse Open Data (0x1034~0x1035) ─
    def _write_bit_sheet(wb, sheet_name, title_note, bit_rows):
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = 'A2'
        _write_header(ws, ['Bit', 'Description', 'Value_Definition'], styles)
        for col, w in zip('ABC', [6, 55, 55]):
            ws.column_dimensions[col].width = w
        if bit_rows:
            for i, r in enumerate(bit_rows, 2):
                _write_row(ws, i, [r.get('bit', ''), r.get('description', ''),
                                   r.get('values', '')], styles)
        else:
            ws.cell(row=2, column=1, value='(Not detected from PDF)')
        ws.cell(row=max(len(bit_rows) + 3, 3), column=1, value=title_note)
        return ws

    _write_bit_sheet(wb, 'Fuse_Open_Data',
                     'NOTE: Registers 0x1034~0x1035 — PV string fuse open status bits',
                     fuse_open_bits)

    # ─ Sheet 6: Status Bits 0x03FC~03FD ─
    _write_bit_sheet(wb, 'Status_Bits_03FC',
                     'NOTE: Registers 0x03FC~03FD — DEA-AVM inverter action/run-state bits',
                     status_bits_03fc)

    # ─ Sheet 7: Fuse Open Check Mask (0x513D~0x513E) ─
    _write_bit_sheet(wb, 'Fuse_Check_Mask',
                     'NOTE: Registers 0x513D~0x513E — PV string fuse check enable/disable mask',
                     fuse_mask_bits)

    # ─ Sheet 8: REMS Error Code Bits ─
    _write_bit_sheet(wb, 'REMS_Error_Bits',
                     'NOTE: REMS Error Code register — bit-field inverter alarm flags',
                     rems_error_bits)

    # ─ Sheet 9: REMS Register Value ─
    ws9 = wb.create_sheet('REMS_Registers')
    ws9.freeze_panes = 'A2'
    _write_header(ws9, ['No', 'Description', 'Length_Bytes', 'Unit'], styles)
    for col, w in zip('ABCD', [5, 50, 14, 14]):
        ws9.column_dimensions[col].width = w
    if rems_regs:
        for i, r in enumerate(rems_regs, 2):
            _write_row(ws9, i, [r.get('no', ''), r.get('description', ''),
                                r.get('length_bytes', ''), r.get('unit', '')], styles)
    else:
        ws9.cell(row=2, column=1, value='(Not detected from PDF)')
    ws9.cell(row=max(len(rems_regs) + 3, 3), column=1,
             value='NOTE: REMS protocol register layout (function code 0x7E)')

    wb.save(output_path)
    _p(f"  Stage 1 complete -> {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# Stage 1 AI-assist helper (Claude API)
# ---------------------------------------------------------------------------

def _ai_stage1_extract(pdf_path, api_key, model, progress_cb):
    """Call Claude API to independently extract register table from PDF text (v1.4.0).

    v1.4.0 changes vs v1.3.0:
    - AI now performs INDEPENDENT extraction (not just gap-filling from the offline parser)
    - Improved prompt: explicit instruction to extract definition NAMES not range values
    - AI also returns status_codes and error_codes arrays for richer extraction
    - Returns dict {registers, status_codes, error_codes} instead of plain list

    Caller (stage1_extract_to_excel) handles the union-merge with offline results.
    """
    try:
        import anthropic
    except ImportError:
        raise ImportError("anthropic package required for AI-enhanced extraction")

    raw_text = _extract_pdf_text(pdf_path)
    if not raw_text:
        return {}

    # Truncate to ~72 KB (leave room for system context)
    if len(raw_text) > 72000:
        raw_text = raw_text[:72000] + '\n...[text truncated for length]'

    prompt = (
        "You are an expert Modbus protocol document parser for solar inverter systems.\n\n"
        "Carefully read the following Modbus documentation and extract THREE types of data:\n\n"
        "## 1. Register Table (JSON array: key \"registers\")\n"
        "Each register entry must have EXACTLY these fields:\n"
        "  - section: section/group name (string)\n"
        "  - definition: the REGISTER NAME/LABEL as written in the document "
        "(NOT the value range like '0-1100' or '0~9999' — those are ranges, not names)\n"
        "  - address: register address as decimal integer\n"
        "  - address_hex: register address as hex string e.g. \"0x1001\"\n"
        "  - regs: number of registers (1 for U16/S16, 2 for U32/S32)\n"
        "  - type: data type — one of U16, S16, U32, S32, ASCII, FLOAT, BOOL\n"
        "  - unit: measurement unit string (V, A, W, kWh, Hz, %, etc.) or empty\n"
        "  - rw: access type — RO, RW, or WO\n"
        "  - comment: any description or note from the document\n\n"
        "CRITICAL RULES:\n"
        "- 'definition' must be the register NAME (e.g. 'Total DC Power', "
        "'L1 Voltage', 'Grid Frequency'), never a numeric range.\n"
        "- Extract ALL registers without exception — do not skip any.\n"
        "- Must include: output power (all bytes), internal temperature, "
        "daily/total energy, all PV channels, all AC phases, frequency, "
        "power factor, status/error codes, control registers.\n"
        "- For U32 registers that use two U16 high/low bytes, list BOTH addresses.\n\n"
        "## 2. Status Codes (JSON array: key \"status_codes\")\n"
        "Extract the inverter operating state/mode table entries:\n"
        "  - code: integer value\n"
        "  - name: state name (e.g. 'Standby', 'Grid-connected', 'Fault')\n"
        "  - description: full description from document\n\n"
        "## 3. Error/Fault Codes (JSON array: key \"error_codes\")\n"
        "Extract fault/alarm code table entries:\n"
        "  - code: fault code or bit reference (e.g. '0x101E Bit3', '32', '0x0001')\n"
        "  - name: fault name\n"
        "  - description: fault description\n"
        "  - cause: root cause if available\n\n"
        "Return ONLY a single JSON object with keys \"registers\", \"status_codes\", "
        "\"error_codes\". No markdown, no explanation.\n\n"
        f"Document:\n{raw_text}"
    )

    client = anthropic.Anthropic(api_key=api_key)
    mdl = model or 'claude-opus-4-6'

    if progress_cb:
        progress_cb(f"  [AI Stage1] Calling {mdl} for independent extraction...")

    try:
        resp = client.messages.create(
            model=mdl,
            max_tokens=8192,
            messages=[{'role': 'user', 'content': prompt}]
        )
    except Exception as e:
        if progress_cb:
            progress_cb(f"  [AI Stage1] API error: {e}")
        return {}

    text = resp.content[0].text.strip()

    # Extract JSON object
    import json
    parsed = None
    # Try object first
    m_obj = re.search(r'\{.*\}', text, re.DOTALL)
    if m_obj:
        try:
            parsed = json.loads(m_obj.group(0))
        except json.JSONDecodeError:
            pass
    # Fallback: try array (old format compatibility)
    if parsed is None:
        m_arr = re.search(r'\[.*\]', text, re.DOTALL)
        if m_arr:
            try:
                arr = json.loads(m_arr.group(0))
                parsed = {'registers': arr, 'status_codes': [], 'error_codes': []}
            except json.JSONDecodeError:
                pass

    if not parsed or not isinstance(parsed, dict):
        if progress_cb:
            progress_cb("  [AI Stage1] Failed to parse JSON response")
        return {}

    # Normalize registers
    _RANGE_PAT = re.compile(r'^[\d\.]+\s*[~\-]\s*[\d\.]+$')
    registers = []
    for item in parsed.get('registers', []):
        if not isinstance(item, dict):
            continue
        addr_hex = str(item.get('address_hex', '') or '')
        addr = item.get('address')
        if addr is None and addr_hex:
            try:
                addr = int(addr_hex, 16)
            except ValueError:
                addr = 0
        defn = str(item.get('definition', '')).strip()
        # Reject if definition looks like a range value
        if _RANGE_PAT.match(defn) or (defn.replace('.', '').isdigit() and len(defn) > 0):
            defn = item.get('comment', '') or defn
        if not defn:
            continue
        registers.append({
            'section':     str(item.get('section', '')),
            'definition':  defn,
            'address':     int(addr) if addr is not None else 0,
            'address_hex': addr_hex or (f'0x{int(addr):04X}' if addr else ''),
            'regs':        int(item.get('regs', 1)),
            'type':        str(item.get('type', 'U16')),
            'unit':        str(item.get('unit', '')),
            'rw':          str(item.get('rw', 'RO')),
            'comment':     str(item.get('comment', '')),
        })

    # Normalize status_codes
    status_codes = []
    for sc in parsed.get('status_codes', []):
        if not isinstance(sc, dict):
            continue
        status_codes.append({
            'code':        sc.get('code', ''),
            'name':        str(sc.get('name', ''))[:40],
            'description': str(sc.get('description', ''))[:80],
            'bit_def':     '',
        })

    # Normalize error_codes
    error_codes = []
    for ec in parsed.get('error_codes', []):
        if not isinstance(ec, dict):
            continue
        error_codes.append({
            'code':        str(ec.get('code', '')),
            'name':        str(ec.get('name', ''))[:40],
            'description': str(ec.get('description', ''))[:80],
            'cause':       str(ec.get('cause', '')),
            'solution':    '',
        })

    if progress_cb:
        progress_cb(f"  [AI Stage1] Extracted: {len(registers)} registers, "
                    f"{len(status_codes)} status codes, {len(error_codes)} error codes")

    return {
        'registers':    registers,
        'status_codes': status_codes,
        'error_codes':  error_codes,
    }


def _union_merge_registers(offline_regs, ai_regs):
    """Union merge offline parser results with AI extraction results (v1.4.0).

    Strategy:
    - For addresses found by both: prefer AI definition if it's longer/better
      (non-empty, not a numeric range); otherwise keep offline definition
    - For addresses only in AI: add as new entries
    - For addresses only in offline: keep as-is
    - Sort by address

    Returns merged list sorted by address.
    """
    _RANGE_PAT = re.compile(r'^[\d\.]+\s*[~\-]\s*[\d\.]+$')

    def _is_better_defn(new_defn, old_defn):
        """Return True if new_defn is a better definition than old_defn."""
        if not new_defn:
            return False
        if not old_defn:
            return True
        # Range value is always worse
        if _RANGE_PAT.match(str(new_defn)):
            return False
        # AI's definition is better if it's longer and more descriptive
        return len(str(new_defn)) > len(str(old_defn))

    # Build map: address → reg dict (from offline)
    merged = {}
    for r in offline_regs:
        addr = r.get('address', 0)
        if addr is not None:
            merged[addr] = dict(r)

    # Merge AI results
    ai_new = 0
    ai_improved = 0
    for r in ai_regs:
        addr = r.get('address', 0)
        if addr is None:
            continue
        if addr not in merged:
            merged[addr] = dict(r)
            ai_new += 1
        else:
            old = merged[addr]
            if _is_better_defn(r.get('definition', ''), old.get('definition', '')):
                merged[addr]['definition'] = r['definition']
                if r.get('comment'):
                    merged[addr]['comment'] = r['comment']
                if r.get('unit') and not old.get('unit'):
                    merged[addr]['unit'] = r['unit']
                ai_improved += 1

    result = sorted(merged.values(), key=lambda x: x.get('address', 0))
    return result, ai_new, ai_improved


# ---------------------------------------------------------------------------
# STAGE 2: Stage 1 Excel → Mapping Excel
# ---------------------------------------------------------------------------

def stage2_create_mapping_excel(stage1_excel_path, output_path=None,
                                 solarize_addr_map=None, solarize_scale=None,
                                 mode='offline',
                                 api_key=None, model=None,
                                 progress_cb=None):
    """
    Stage 2: Create Solarize-standard mapping Excel from Stage 1 register Excel.

    Args:
        stage1_excel_path : Path to Stage 1 .xlsx
        output_path       : Destination .xlsx path (auto-generated if None)
        solarize_addr_map : dict {int_addr: (solarize_name, comment)} — from _SOLARIZE_ADDR_TO_NAME
        solarize_scale    : dict {key: float} — from _SOLARIZE_STANDARD_SCALE
        mode              : 'offline' (rule-based + reference similarity)
                            'ai'      (Claude API few-shot matching for unmapped regs)
        api_key           : Claude API key (mode='ai' only)
        model             : Claude model name (mode='ai' only)
        progress_cb       : Callback(msg: str)

    Returns:
        output_path (str)
    """
    openpyxl = _require_openpyxl()

    def _p(msg):
        if progress_cb:
            progress_cb(msg)

    _p("Stage 2: Creating mapping Excel...")

    if not output_path:
        base = os.path.splitext(stage1_excel_path)[0]
        output_path = base + '_mapping.xlsx'

    if solarize_addr_map is None:
        solarize_addr_map = {}
    if solarize_scale is None:
        solarize_scale = {'voltage': 0.1, 'current': 0.01, 'power': 0.1,
                          'frequency': 0.01, 'power_factor': 0.001}

    # ── Read Stage 1 Excel ──
    wb_src = openpyxl.load_workbook(stage1_excel_path, data_only=True)
    ws_src = wb_src['Registers']

    src_regs = []
    src_status = []
    src_errors = []

    # Read Sheet 1 (Registers) - skip section-divider rows and header
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # Section divider rows have text starting with '---'
        if isinstance(row[0], str) and row[0].startswith('---'):
            continue
        if not isinstance(row[0], (int, float)):
            continue
        src_regs.append({
            'no':       row[0],
            'section':  row[1] or '',
            'addr_hex': str(row[2] or ''),
            'addr_dec': row[3],
            'defn':     row[4] or '',
            'dtype':    row[5] or 'U16',
            'fc':       row[6] or 'FC03',
            'regs':     row[7] or 1,
            'unit':     row[8] or '',
            'scale':    row[9] or '',
            'rw':       row[10] or 'RO',
            'comment':  row[11] or '',
        })

    # Read Sheet 2 (Status Definitions)
    if 'Status_Definitions' in wb_src.sheetnames:
        ws_s2 = wb_src['Status_Definitions']
        for row in ws_s2.iter_rows(min_row=2, values_only=True):
            if row[0] is None or isinstance(row[0], str):
                continue
            src_status.append({
                'code':    row[0],
                'name':    row[1] or '',
                'desc':    row[2] or '',
                'bit_def': row[3] or '',
            })

    # Read Sheet 3 (Error Codes)
    if 'Error_Fault_Codes' in wb_src.sheetnames:
        ws_s3 = wb_src['Error_Fault_Codes']
        for row in ws_s3.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            src_errors.append({
                'code':     str(row[0]),
                'name':     row[1] or '',
                'desc':     row[2] or '',
                'cause':    row[3] or '',
                'solution': row[4] or '',
            })

    _p(f"  Read {len(src_regs)} registers from Stage 1 Excel")

    # ── Load reference manager for offline / AI matching ──
    ref_mgr = _get_ref_manager()
    if ref_mgr:
        _p(f"  Reference library: {ref_mgr.count()} sets loaded")
    else:
        _p("  Reference library: not available")

    # ── Build mapping ──
    wb = openpyxl.Workbook()
    styles = _make_styles(openpyxl)

    # ─ Sheet 1: Register Mapping ─
    ws1 = wb.active
    ws1.title = 'Register_Mapping'
    ws1.freeze_panes = 'A2'

    s1_hdr = ['No', 'Section',
              'Source_Addr_Hex', 'Source_Addr_Dec', 'Source_Name',
              'Source_Type', 'Source_Unit', 'Source_Scale', 'Source_RW', 'Source_Regs',
              '→',
              'Solarize_Name', 'Solarize_Addr_Hex',
              'Match_Type', 'Notes']
    _write_header(ws1, s1_hdr, styles)

    col_widths = [5, 22, 14, 13, 35, 9, 9, 12, 7, 9, 3, 35, 16, 10, 40]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[chr(64 + i)].width = w

    match_stats = {'auto': 0, 'manual': 0, 'unmapped': 0}
    auto_names = {v[0] for v in solarize_addr_map.values()}

    current_sec = None
    row_num = 2
    for reg in src_regs:
        sec = reg['section']
        if sec and sec != current_sec:
            current_sec = sec
            c = ws1.cell(row=row_num, column=1, value=f'--- {sec} ---')
            c.fill = styles['sec_fill']
            c.font = styles['sec_font']
            ws1.merge_cells(start_row=row_num, start_column=1,
                            end_row=row_num, end_column=len(s1_hdr))
            row_num += 1

        # Resolve address
        addr_int = _addr_to_int(reg['addr_hex'], reg['addr_dec'])

        # Try address match in Solarize table
        solarize_name = ''
        solarize_addr_hex = ''
        match_type = 'Unmapped'

        if addr_int is not None and addr_int in solarize_addr_map:
            # Priority 1: exact Solarize address table match
            solarize_name, _ = solarize_addr_map[addr_int]
            solarize_addr_hex = f'0x{addr_int:04X}'
            match_type = 'Auto'
            match_stats['auto'] += 1
        else:
            # Priority 2: exact normalized name match in Solarize table
            src_up = _to_upper_snake(reg['defn'])
            if src_up in auto_names:
                for s_addr, (s_name, _) in solarize_addr_map.items():
                    if s_name == src_up:
                        solarize_name     = src_up
                        solarize_addr_hex = f'0x{s_addr:04X}'
                        match_type = 'Auto'
                        match_stats['auto'] += 1
                        break
            elif addr_int is not None and addr_int in _SUPPLEMENT_ADDR_MAP:
                # Priority 3: manufacturer supplement address map (explicit overrides)
                solarize_name = _SUPPLEMENT_ADDR_MAP[addr_int]
                solarize_addr_hex = f'0x{addr_int:04X}'
                match_type = 'Ref(supp)'
                match_stats['auto'] += 1
            else:
                # Priority 4: reference name similarity (offline mode)
                ref_name = ''
                ref_conf = 0.0
                if ref_mgr:
                    ref_name, ref_conf = ref_mgr.offline_match_name(
                        reg['defn'],
                        source_section=reg.get('section', ''),
                        source_type=reg.get('dtype', ''),
                        source_unit=reg.get('unit', ''),
                        threshold=0.55,
                    )
                if ref_name and ref_conf >= 0.55:
                    solarize_name = ref_name
                    solarize_addr_hex = ''
                    match_type = f'Ref({ref_conf:.0%})'
                    match_stats['auto'] += 1
                else:
                    match_type = 'Unmapped'
                    match_stats['unmapped'] += 1

        # Choose fill color
        if match_type == 'Auto':
            fill = styles['auto_fill']
        elif match_type == 'Manual':
            fill = styles['man_fill']
        else:
            fill = styles['unm_fill']

        _write_row(ws1, row_num,
                   [reg['no'], sec,
                    reg['addr_hex'], reg['addr_dec'], reg['defn'],
                    reg['dtype'], reg['unit'], reg['scale'], reg['rw'], reg['regs'],
                    '→',
                    solarize_name, solarize_addr_hex,
                    match_type, ''],
                   styles, fill=fill)
        row_num += 1

    ws1.auto_filter.ref = f'A1:{chr(64+len(s1_hdr))}1'

    # ── Supplement cluster injection: add missing essential registers ──
    # If ≥3 addresses from a manufacturer cluster are detected in Stage 1,
    # inject any cluster members missing from Stage 1 (e.g. Sungrow 0x13B6).
    src_addrs = {_addr_to_int(r['addr_hex'], r['addr_dec']) for r in src_regs
                 if _addr_to_int(r['addr_hex'], r['addr_dec']) is not None}
    injected = 0
    for cluster in _SUPPLEMENT_CLUSTERS:
        present_in_cluster = cluster & src_addrs
        if len(present_in_cluster) >= 3:
            missing_in_cluster = cluster - src_addrs
            for miss_addr in sorted(missing_in_cluster):
                if miss_addr not in _SUPPLEMENT_ADDR_MAP:
                    continue
                supp_name = _SUPPLEMENT_ADDR_MAP[miss_addr]
                _write_row(ws1, row_num,
                           [f'S{injected+1}', 'supplement',
                            f'0x{miss_addr:04X}', miss_addr, f'[injected] {supp_name}',
                            'U16', '', '', 'RO', 1,
                            '→',
                            supp_name, f'0x{miss_addr:04X}',
                            'Ref(supp)', f'Auto-injected: not in Stage1 PDF extract'],
                           styles, fill=styles['auto_fill'])
                row_num += 1
                injected += 1
    if injected:
        _p(f"  Supplement cluster injection: {injected} essential registers added")

    # ── AI mode: batch-map remaining Unmapped registers via Claude API ──
    if mode == 'ai' and api_key and match_stats['unmapped'] > 0:
        _p(f"  AI mode: mapping {match_stats['unmapped']} unmapped registers via Claude API...")
        ai_results = _ai_stage2_map(ws1, src_regs, solarize_addr_map,
                                     ref_mgr, api_key, model, styles, _p)
        if ai_results:
            match_stats['auto'] += ai_results.get('mapped', 0)
            match_stats['unmapped'] -= ai_results.get('mapped', 0)
            _p(f"  AI mapped: {ai_results.get('mapped', 0)} additional registers")

    # Legend row
    row_num += 1
    ws1.cell(row=row_num, column=1,
             value=f'Mode: {mode.upper()}  |  Match stats: Auto/Ref={match_stats["auto"]}, '
                   f'Unmapped={match_stats["unmapped"]}  '
                   f'| Edit "Solarize_Name" for Unmapped rows (red = needs review).')

    _p(f"  Matched: {match_stats['auto']} (auto+ref), Unmapped={match_stats['unmapped']}")

    # ─ Sheet 2: Status Mapping ─
    ws2 = wb.create_sheet('Status_Mapping')
    ws2.freeze_panes = 'A2'
    s2_hdr = ['Raw_Code', 'Raw_Name', 'InverterMode_State', 'Solarize_Code', 'Notes']
    _write_header(ws2, s2_hdr, styles)
    for col, w in zip('ABCDE', [10, 35, 20, 14, 40]):
        ws2.column_dimensions[col].width = w

    if src_status:
        for i, sc in enumerate(src_status, 2):
            mode_name, mode_code = _match_status_to_mode(sc.get('name', ''))
            note = '' if mode_name != 'Unknown' else 'Review - could not auto-match'
            fill = styles['auto_fill'] if mode_name != 'Unknown' else styles['unm_fill']
            _write_row(ws2, i,
                       [sc['code'], sc['name'], mode_name,
                        mode_code if mode_code >= 0 else '', note],
                       styles, fill=fill)
    else:
        # Use standard modes as template
        for i, (code, name, desc) in enumerate(_INVERTER_MODES, 2):
            _write_row(ws2, i,
                       [code, desc, name, code,
                        'Standard InverterMode - verify against inverter datasheet'],
                       styles, fill=styles['auto_fill'])

    ws2.cell(row=len(src_status or _INVERTER_MODES) + 3, column=1,
             value='NOTE: Edit "InverterMode_State" column: INITIAL/STANDBY/ON_GRID/FAULT/SHUTDOWN')

    # ─ Sheet 3: DER-AVM Control Registers ─
    ws3 = wb.create_sheet('DER_AVM_Control')
    ws3.freeze_panes = 'A2'
    s3_hdr = ['Address_Hex', 'Address_Dec', 'Name', 'Data_Type', 'FC',
              'Registers', 'Unit', 'Scale', 'R/W', 'Description']
    _write_header(ws3, s3_hdr, styles)
    for col, w in zip('ABCDEFGHIJ', [12, 12, 32, 9, 8, 9, 7, 8, 6, 55]):
        ws3.column_dimensions[col].width = w

    for i, reg in enumerate(_DER_AVM_CONTROL_REGS, 2):
        _write_row(ws3, i, list(reg), styles, fill=styles['auto_fill'])

    ws3.cell(row=len(_DER_AVM_CONTROL_REGS) + 3, column=1,
             value='NOTE: DER-AVM control registers (always included). Do not modify addresses.')

    # ─ Sheet 4: DEA-AVM Monitoring Registers ─
    ws4 = wb.create_sheet('DEA_AVM_Monitor')
    ws4.freeze_panes = 'A2'
    _write_header(ws4, s3_hdr, styles)
    for col, w in zip('ABCDEFGHIJ', [12, 12, 32, 9, 8, 9, 7, 8, 6, 55]):
        ws4.column_dimensions[col].width = w

    for i, reg in enumerate(_DEA_AVM_MONITOR_REGS, 2):
        _write_row(ws4, i, list(reg), styles, fill=styles['auto_fill'])

    ws4.cell(row=len(_DEA_AVM_MONITOR_REGS) + 3, column=1,
             value='NOTE: DEA-AVM monitoring registers (always included). Do not modify addresses.')

    wb.save(output_path)
    _p(f"  Stage 2 complete -> {output_path}")
    return output_path


# ---------------------------------------------------------------------------
# AI helper: Stage 2 batch mapping via Claude API
# ---------------------------------------------------------------------------

def _ai_stage2_map(ws1, src_regs, solarize_addr_map, ref_mgr,
                   api_key, model, styles, progress_cb):
    """
    Call Claude API to map remaining Unmapped registers in the mapping Excel.
    Updates ws1 rows in-place where Match_Type == 'Unmapped'.
    Returns {'mapped': int}
    """
    try:
        import anthropic
    except ImportError:
        if progress_cb:
            progress_cb("  AI mapping skipped: anthropic package not installed")
        return {'mapped': 0}

    # Collect all Unmapped rows (col N = col 14 = Match_Type)
    unmapped_rows = []
    for row_idx in range(2, ws1.max_row + 1):
        match_val = ws1.cell(row=row_idx, column=14).value
        if match_val == 'Unmapped':
            unmapped_rows.append((
                row_idx,
                ws1.cell(row=row_idx, column=5).value or '',   # Source_Name
                ws1.cell(row=row_idx, column=3).value or '',   # Source_Addr_Hex
                ws1.cell(row=row_idx, column=2).value or '',   # Section
                ws1.cell(row=row_idx, column=6).value or '',   # Source_Type
                ws1.cell(row=row_idx, column=7).value or '',   # Source_Unit
            ))

    if not unmapped_rows:
        return {'mapped': 0}

    # Few-shot examples from reference manager
    few_shot = ref_mgr.get_few_shot_mapping_examples(n=2) if ref_mgr else ''

    # Build source register list for prompt
    src_list = '\n'.join(
        f'  [{i+1}] name="{r[1]}", addr={r[2]}, section="{r[3]}", type={r[4]}, unit={r[5]}'
        for i, r in enumerate(unmapped_rows)
    )

    # Known Solarize standard names for reference
    sol_names_sample = ', '.join(list({v[0] for v in solarize_addr_map.values()})[:60])

    prompt = (
        "You are an expert at mapping solar inverter Modbus register names to the\n"
        "Solarize standard register naming convention used in RTU firmware.\n\n"
        "Known Solarize standard names (sample):\n"
        f"{sol_names_sample}\n\n"
        "Reference mappings from known inverters:\n"
        f"{few_shot}\n\n"
        "Map these registers to the best matching Solarize standard name.\n"
        "If no good match exists, use null.\n\n"
        "Registers to map:\n"
        f"{src_list}\n\n"
        "Return a JSON array in this exact format (one entry per register, same order):\n"
        '[{"index":1,"solarize_name":"L1_VOLTAGE","confidence":0.9,"notes":"AC output L1"},'
        ' {"index":2,"solarize_name":null,"confidence":0.0,"notes":"unknown"}]\n'
        "Return ONLY the JSON array, no other text."
    )

    if progress_cb:
        progress_cb(f"  Calling Claude API for {len(unmapped_rows)} unmapped registers...")

    try:
        client = anthropic.Anthropic(api_key=api_key)
        mdl = model or 'claude-opus-4-6'
        resp = client.messages.create(
            model=mdl, max_tokens=4096,
            messages=[{'role': 'user', 'content': prompt}]
        )
        text = resp.content[0].text.strip()
    except Exception as e:
        if progress_cb:
            progress_cb(f"  AI API error: {e}")
        return {'mapped': 0}

    # Parse JSON response
    import json as _json
    m = re.search(r'\[.*\]', text, re.DOTALL)
    if not m:
        return {'mapped': 0}

    try:
        items = _json.loads(m.group(0))
    except _json.JSONDecodeError:
        return {'mapped': 0}

    # Apply results to ws1
    mapped_count = 0
    from openpyxl.styles import PatternFill
    ai_fill = PatternFill('solid', fgColor='BDD7EE')   # light blue = AI mapped

    for item in items:
        idx = item.get('index', 0)
        sol_name = item.get('solarize_name') or ''
        conf     = item.get('confidence', 0.0)
        notes    = item.get('notes', '')

        if not sol_name or conf < 0.6:
            continue

        if 1 <= idx <= len(unmapped_rows):
            row_idx = unmapped_rows[idx - 1][0]
            ws1.cell(row=row_idx, column=12).value = sol_name
            ws1.cell(row=row_idx, column=14).value = f'AI({conf:.0%})'
            ws1.cell(row=row_idx, column=15).value = notes

            # Apply AI fill color
            for col in range(1, 16):
                ws1.cell(row=row_idx, column=col).fill = ai_fill

            mapped_count += 1

    return {'mapped': mapped_count}


# ---------------------------------------------------------------------------
# STAGE 3: Mapping Excel → Register .py
# ---------------------------------------------------------------------------

def stage3_generate_py(mapping_excel_path, settings, output_path=None,
                        mode='offline', api_key=None, model=None,
                        progress_cb=None):
    """
    Stage 3: Generate *_registers.py from Stage 2 mapping Excel.

    Args:
        mapping_excel_path : Path to Stage 2 mapping .xlsx
        settings: dict with keys:
            mppt_count    (int, 1-8, default 4)
            string_count  (int, 1-24, default 8)
            iv_scan       (bool, default False)
            der_avm       (bool, default True)
            dea_avm       (bool, default True)
            manufacturer  (str)
            class_name    (str, default 'RegisterMap')
            protocol_name (str)
            fc_code       (str, 'FC03' or 'FC04')
        output_path : Destination .py path (auto-generated if None)
        mode        : 'offline' or 'ai' (ai enables auto-retry x3 on FAIL)
        api_key     : Claude API key (mode='ai' only)
        model       : Claude model name (mode='ai' only)
        progress_cb : Callback(msg: str)

    Returns:
        dict: {
            'code'     : str  (generated Python source),
            'results'  : list of (status, msg) tuples,
            'success'  : bool,
            'output_path': str or None,
        }
    """
    openpyxl = _require_openpyxl()

    def _p(msg):
        if progress_cb:
            progress_cb(msg)

    _p("Stage 3: Generating register .py from mapping Excel...")

    mppt_count   = int(settings.get('mppt_count', 4))
    string_count = int(settings.get('string_count', 8))
    iv_scan      = bool(settings.get('iv_scan', False))
    der_avm      = bool(settings.get('der_avm', True))
    dea_avm      = bool(settings.get('dea_avm', True))
    manufacturer = str(settings.get('manufacturer', 'Unknown')).strip()
    class_name   = str(settings.get('class_name', 'RegisterMap')).strip() or 'RegisterMap'
    protocol     = str(settings.get('protocol_name', 'custom')).strip()
    fc_str       = str(settings.get('fc_code', 'FC03'))
    fc_num       = 3 if fc_str == 'FC03' else 4

    # ── Read mapping Excel ──
    wb = openpyxl.load_workbook(mapping_excel_path, data_only=True)

    # Sheet 1: Register_Mapping
    reg_rows = []
    if 'Register_Mapping' in wb.sheetnames:
        ws = wb['Register_Mapping']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            if isinstance(row[0], str) and row[0].startswith('---'):
                continue
            # Accept numeric row numbers AND supplement rows ('S1', 'S2', ...)
            if not isinstance(row[0], (int, float)):
                if not (isinstance(row[0], str) and row[0].startswith('S')):
                    continue
            reg_rows.append({
                'section':      row[1] or '',
                'src_addr_hex': str(row[2] or ''),
                'src_addr_dec': row[3],
                'src_name':     row[4] or '',
                'src_type':     row[5] or 'U16',
                'src_unit':     row[7] or '',
                'src_scale':    row[7] if len(row) > 7 else '',
                'src_rw':       row[8] or 'RO',
                'src_regs':     row[9] or 1,
                'sol_name':     str(row[11] or '') if len(row) > 11 else '',
                'match_type':   str(row[13] or 'Unmapped') if len(row) > 13 else 'Unmapped',
                'notes':        str(row[14] or '') if len(row) > 14 else '',
            })

    # Sheet 2: Status_Mapping
    status_map = {}   # raw_code -> (mode_name, solarize_code)
    if 'Status_Mapping' in wb.sheetnames:
        ws2 = wb['Status_Mapping']
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                raw_code   = int(row[0])
                mode_name  = str(row[2] or 'Unknown').strip()
                sol_code   = row[3]
                status_map[raw_code] = (mode_name, sol_code)
            except (ValueError, TypeError):
                pass

    _p(f"  Registers: {len(reg_rows)}, Status mappings: {len(status_map)}")

    # ── Determine effective name for each register ──
    # MPPT/String filter regex
    mppt_re   = re.compile(r'^MPPT(\d+)_')
    string_re = re.compile(r'^STRING(\d+)_')

    # Build (name, address, type, unit, notes) list, filtering by mppt/string count
    reg_entries = []
    seen_names  = set()

    for r in reg_rows:
        sol_name = r['sol_name'].strip()
        src_name = r['src_name'].strip()

        # Choose constant name
        if sol_name and r['match_type'] != 'Unmapped':
            const_name = sol_name
        elif src_name:
            const_name = _to_upper_snake(src_name)
        else:
            continue

        if not const_name or const_name in seen_names:
            continue

        # Filter MPPT by count
        m_mppt = mppt_re.match(const_name)
        if m_mppt and int(m_mppt.group(1)) > mppt_count:
            continue

        # Filter STRING by count
        m_str = string_re.match(const_name)
        if m_str and int(m_str.group(1)) > string_count:
            continue

        # Skip IV scan registers if not enabled
        if not iv_scan and const_name.startswith('IV_'):
            continue

        # Resolve address
        addr_int = _addr_to_int(r['src_addr_hex'], r['src_addr_dec'])
        if addr_int is None:
            continue

        seen_names.add(const_name)
        reg_entries.append({
            'section':   r['section'],
            'name':      const_name,
            'addr':      addr_int,
            'addr_hex':  f'0x{addr_int:04X}',
            'dtype':     r['src_type'],
            'rw':        r['src_rw'],
            'regs':      r['src_regs'],
            'unit':      r['src_unit'],
            'notes':     r['notes'] or f"{r['src_type']}",
        })

    # DER-AVM registers (from Sheet 3 or standard table)
    # Always remove standard DER-AVM names from seen_names first, so that the
    # authoritative addresses from the DER-AVM Control sheet are used even when
    # a source PDF register was (incorrectly) mapped to the same name.
    _der_standard_names = {reg[2] for reg in _DER_AVM_CONTROL_REGS}
    seen_names -= _der_standard_names

    der_entries = []
    if der_avm:
        if 'DER_AVM_Control' in wb.sheetnames:
            ws_der = wb['DER_AVM_Control']
            for row in ws_der.iter_rows(min_row=2, values_only=True):
                if row[0] is None or not str(row[0]).startswith('0x'):
                    continue
                name = str(row[2] or '')
                if not name:
                    continue
                addr_int = _addr_to_int(str(row[0]), row[1])
                if addr_int is None:
                    continue
                if name in seen_names:
                    continue
                seen_names.add(name)
                der_entries.append({
                    'section': 'DER-AVM Control',
                    'name': name, 'addr': addr_int,
                    'addr_hex': f'0x{addr_int:04X}',
                    'dtype': str(row[3] or 'U16'),
                    'rw': str(row[8] or 'RW'),
                    'regs': row[5] or 1,
                    'unit': str(row[6] or ''),
                    'notes': str(row[9] or ''),
                })
        else:
            for reg in _DER_AVM_CONTROL_REGS:
                if reg[2] not in seen_names:
                    seen_names.add(reg[2])
                    der_entries.append({
                        'section': 'DER-AVM Control', 'name': reg[2],
                        'addr': reg[1], 'addr_hex': reg[0],
                        'dtype': reg[3], 'rw': reg[8],
                        'regs': reg[5], 'unit': reg[6], 'notes': reg[9],
                    })

    # DEA-AVM monitor registers
    dea_entries = []
    if dea_avm:
        if 'DEA_AVM_Monitor' in wb.sheetnames:
            ws_dea = wb['DEA_AVM_Monitor']
            for row in ws_dea.iter_rows(min_row=2, values_only=True):
                if row[0] is None or not str(row[0]).startswith('0x'):
                    continue
                name = str(row[2] or '')
                if not name or name in seen_names:
                    continue
                addr_int = _addr_to_int(str(row[0]), row[1])
                if addr_int is None:
                    continue
                seen_names.add(name)
                dea_entries.append({
                    'section': 'DEA-AVM Monitor', 'name': name, 'addr': addr_int,
                    'addr_hex': f'0x{addr_int:04X}',
                    'dtype': str(row[3] or 'S32'), 'rw': str(row[8] or 'RO'),
                    'regs': row[5] or 1, 'unit': str(row[6] or ''), 'notes': str(row[9] or ''),
                })
        else:
            for reg in _DEA_AVM_MONITOR_REGS:
                if reg[2] not in seen_names:
                    seen_names.add(reg[2])
                    dea_entries.append({
                        'section': 'DEA-AVM Monitor', 'name': reg[2],
                        'addr': reg[1], 'addr_hex': reg[0],
                        'dtype': reg[3], 'rw': reg[8],
                        'regs': reg[5], 'unit': reg[6], 'notes': reg[9],
                    })

    # ── Generate .py source code ──
    _p("  Generating Python source code...")
    lines = []

    # Header
    lines += [
        '# -*- coding: utf-8 -*-',
        f'"""',
        f'{manufacturer} Inverter Modbus Register Map',
        f'Protocol: {protocol}',
        f'Function Code: FC{fc_num:02d}  |  MPPT: {mppt_count}  |  Strings: {string_count}',
        f'IV Scan: {"Yes" if iv_scan else "No"}  |  DER-AVM: {"Yes" if der_avm else "No"}  |  DEA-AVM Monitor: {"Yes" if dea_avm else "No"}',
        f'',
        f'Auto-generated by UDP RTU Model Maker v1.3.0 - Stage 3 Pipeline',
        f'Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
        f'',
        f'Stage 3 source: {os.path.basename(mapping_excel_path)}',
        f'"""',
        '',
        'import sys',
        'import os',
        'sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), \'..\'))',
        '',
    ]

    # RegisterMap class
    lines.append(f'class {class_name}:')
    lines.append(f'    """{manufacturer} Modbus Register Map"""')
    lines.append('')

    # Group by section
    all_regs_ordered = reg_entries + der_entries + dea_entries
    sections_ordered = {}
    for reg in all_regs_ordered:
        sec = reg['section'] or 'Other'
        sections_ordered.setdefault(sec, []).append(reg)

    for sec_name, sec_regs in sections_ordered.items():
        lines.append(f'    # {"-" * 60}')
        lines.append(f'    # {sec_name}')
        lines.append(f'    # {"-" * 60}')
        for reg in sec_regs:
            addr_hex = reg['addr_hex']
            name     = reg['name']
            comment  = reg['notes'] or reg['dtype']
            lines.append(f'    {name:<50} = {addr_hex}   # {comment}')
        lines.append('')

    # Backward-compat aliases (DER-AVM)
    if der_avm:
        lines += [
            '    # DER-AVM aliases (backward compatibility)',
            '    POWER_FACTOR_SET   = DER_POWER_FACTOR_SET',
            '    OPERATION_MODE     = DER_ACTION_MODE',
            '    REACTIVE_POWER_PCT = DER_REACTIVE_POWER_PCT',
            '    ACTIVE_POWER_PCT   = DER_ACTIVE_POWER_PCT',
            '    ON_OFF_CONTROL     = INVERTER_ON_OFF',
            '',
        ]

    # Firmware alias (only if MASTER_FIRMWARE_VERSION exists in registers)
    has_fw = any(r['name'] == 'MASTER_FIRMWARE_VERSION' for r in all_regs_ordered)
    if has_fw:
        lines += [
            '    # Firmware version alias',
            '    FIRMWARE_VERSION = MASTER_FIRMWARE_VERSION',
            '',
        ]

    # ── Compatibility aliases (auto-generated for validation) ──
    all_names = {r['name'] for r in all_regs_ordered}
    compat_aliases = []

    # L1/L2/L3 phase aliases (L→R/S/T direction)
    phase_map = [
        ('L1_VOLTAGE', ['R_PHASE_VOLTAGE', 'R_VOLTAGE', 'PHASE_A_VOLTAGE',
                        'GRID_R_VOLTAGE', 'U_A', 'PHASE_1_VOLTAGE']),
        ('L2_VOLTAGE', ['S_PHASE_VOLTAGE', 'S_VOLTAGE', 'PHASE_B_VOLTAGE',
                        'GRID_S_VOLTAGE', 'U_B', 'PHASE_2_VOLTAGE']),
        ('L3_VOLTAGE', ['T_PHASE_VOLTAGE', 'T_VOLTAGE', 'PHASE_C_VOLTAGE',
                        'GRID_T_VOLTAGE', 'U_C', 'PHASE_3_VOLTAGE']),
        ('L1_CURRENT', ['R_PHASE_CURRENT', 'R_CURRENT', 'PHASE_A_CURRENT',
                        'INV_R_CURRENT', 'I_A', 'PHASE_1_CURRENT']),
        ('L2_CURRENT', ['S_PHASE_CURRENT', 'S_CURRENT', 'PHASE_B_CURRENT',
                        'INV_S_CURRENT', 'I_B', 'PHASE_2_CURRENT']),
        ('L3_CURRENT', ['T_PHASE_CURRENT', 'T_CURRENT', 'PHASE_C_CURRENT',
                        'INV_T_CURRENT', 'I_C', 'PHASE_3_CURRENT']),
    ]
    for alias, sources in phase_map:
        if alias not in all_names:
            for src in sources:
                if src in all_names:
                    compat_aliases.append(f'    {alias:<40} = {src}')
                    break

    # R/S/T phase aliases (reverse direction, for inverters using L1/L2/L3 naming)
    rphase_map = [
        ('R_PHASE_VOLTAGE', ['L1_VOLTAGE']),
        ('S_PHASE_VOLTAGE', ['L2_VOLTAGE']),
        ('T_PHASE_VOLTAGE', ['L3_VOLTAGE']),
        ('R_PHASE_CURRENT', ['L1_CURRENT']),
        ('S_PHASE_CURRENT', ['L2_CURRENT']),
        ('T_PHASE_CURRENT', ['L3_CURRENT']),
    ]
    for alias, sources in rphase_map:
        if alias not in all_names:
            for src in sources:
                if src in all_names:
                    compat_aliases.append(f'    {alias:<40} = {src}')
                    break

    # Energy aliases (expanded sources)
    energy_map = [
        ('TOTAL_ENERGY_LOW', ['TOTAL_ENERGY', 'ACCUMULATED_ENERGY', 'CUMULATIVE_PRODUCTION_L',
                               'TOTAL_ENERGY_L', 'ACCUMULATED_ENERGY_L', 'TOTAL_ENERGY_KWH',
                               'TOTAL_YIELD_L', 'TOTAL_GENERATION_KWH',
                               # EKOS / generic translated names
                               'TOTAL_GENERATED_ENERGY', 'ACCUMULATED_GENERATED_ENERGY',
                               'ACCUMULATED_GENERATION_ENERGY', 'GENERATION_KWH_TOTAL',
                               'LIFETIME_ENERGY', 'CUMULATIVE_ENERGY']),
        ('TODAY_ENERGY_LOW', ['TODAY_ENERGY', 'TODAY_ENERGY_L', 'DAILY_PRODUCTION',
                               'DAILY_ENERGY', 'DAILY_ENERGY_FEEDIN', 'DAY_ENERGY',
                               'TODAY_GENERATION', 'TODAY_YIELD', 'DAILY_YIELD',
                               # EKOS / generic translated names
                               'TODAY_GENERATED_ENERGY', 'DAILY_GENERATED_ENERGY',
                               'TODAY_GENERATION_ENERGY', 'DAILY_GENERATION_ENERGY',
                               'GENERATED_ENERGY']),
    ]
    for alias, sources in energy_map:
        if alias not in all_names:
            for src in sources:
                if src in all_names:
                    compat_aliases.append(f'    {alias:<40} = {src}')
                    break

    # MPPT aliases (expanded sources)
    mppt_alias_map = [
        ('MPPT1_VOLTAGE', ['PV1_VOLTAGE', 'PV_VOLTAGE', 'MPPT_1_VOLTAGE', 'PV1_VOL',
                            'PV_1_VOLTAGE', 'PV1_INPUT_VOLTAGE', 'DC1_VOLTAGE']),
        ('MPPT1_CURRENT', ['PV1_CURRENT', 'PV_CURRENT', 'MPPT_1_CURRENT', 'PV1_CUR',
                            'PV_1_CURRENT', 'PV1_INPUT_CURRENT', 'DC1_CURRENT']),
    ]
    for alias, sources in mppt_alias_map:
        if alias not in all_names:
            for src in sources:
                if src in all_names:
                    compat_aliases.append(f'    {alias:<40} = {src}')
                    break

    # INVERTER_MODE alias (from status/mode registers when not directly mapped)
    if 'INVERTER_MODE' not in all_names:
        for src in ['RUNNING_STATUS', 'DEVICE_STATUS', 'INVERTER_STATUS',
                    'SYSTEM_STATUS', 'INV_OPERATING_MODE', 'OPERATING_MODE_STATUS']:
            if src in all_names:
                compat_aliases.append(f'    {"INVERTER_MODE":<40} = {src}')
                break

    # ERROR_CODE1/2 aliases (from fault/alarm registers)
    if 'ERROR_CODE1' not in all_names:
        for src in ['FAULT_CODE_1', 'FAULT_CODE1', 'DSP_ALARM_CODE_L', 'ALARM_CODE_1',
                    'ALARM_CODE1', 'FAULT_CODE', 'ALARM_CODE', 'ERROR_CODE',
                    'FAULT_CODE_L', 'DSP_ERROR_CODE_L', 'ERROR_1', 'FAULT_1',
                    # Last-resort fallbacks (Huawei uses _2 naming, some brands use ALARM_n)
                    'FAULT_CODE_2', 'REMOTESIGNAL_ALARM_1', 'ALARM_1', 'FAULT_STATUS']:
            if src in all_names:
                compat_aliases.append(f'    {"ERROR_CODE1":<40} = {src}')
                break
    if 'ERROR_CODE2' not in all_names:
        for src in ['FAULT_CODE_2', 'FAULT_CODE2', 'DSP_ERROR_CODE_H', 'ALARM_CODE_2',
                    'ALARM_CODE2', 'ERROR_2', 'FAULT_2', 'FAULT_CODE_H', 'ALARM_CODE_H',
                    'REMOTESIGNAL_ALARM_2', 'ALARM_2']:
            if src in all_names:
                compat_aliases.append(f'    {"ERROR_CODE2":<40} = {src}')
                break

    # PV_TOTAL_INPUT_POWER_LOW alias (from DC/PV input power registers)
    if 'PV_TOTAL_INPUT_POWER_LOW' not in all_names:
        for src in ['INPUT_POWER', 'PV_POWER', 'DC_INPUT_POWER', 'PV_INPUT_POWER',
                    'TOTAL_INPUT_POWER', 'DC_POWER', 'PV_ACTIVE_POWER',
                    'PV_TOTAL_POWER', 'AC_POWER_L']:
            if src in all_names:
                compat_aliases.append(f'    {"PV_TOTAL_INPUT_POWER_LOW":<40} = {src}')
                break

    # GRID_TOTAL_ACTIVE_POWER_LOW alias (from AC/grid output power registers)
    if 'GRID_TOTAL_ACTIVE_POWER_LOW' not in all_names:
        for src in ['GRID_T_POWER', 'AC_POWER', 'ACTIVE_POWER', 'OUTPUT_POWER',
                    'GRID_POWER', 'GRID_ACTIVE_POWER', 'TOTAL_ACTIVE_POWER',
                    'AC_ACTIVE_POWER', 'GRID_POWER_LOW', 'GRID_WATT',
                    'GRID_TOTAL_POWER', 'AC_POWER_L', 'GRID_R_POWER',
                    'METER_ACTIVE_POWER', 'INVERTER_ACTIVE_POWER']:
            if src in all_names:
                compat_aliases.append(f'    {"GRID_TOTAL_ACTIVE_POWER_LOW":<40} = {src}')
                break

    # POWER_FACTOR alias (when not directly mapped)
    if 'POWER_FACTOR' not in all_names:
        for src in ['POWER_FACTOR_L', 'POWER_FACTOR_VALUE', 'PF', 'POWERFACTOR']:
            if src in all_names:
                compat_aliases.append(f'    {"POWER_FACTOR":<40} = {src}')
                break

    # INVERTER_CONTROL alias
    if 'INVERTER_CONTROL' not in all_names and 'INVERTER_ON_OFF' in all_names:
        compat_aliases.append(f'    {"INVERTER_CONTROL":<40} = INVERTER_ON_OFF')

    if compat_aliases:
        lines.append('    # Compatibility aliases (auto-generated)')
        lines.extend(compat_aliases)
        lines.append('')

    lines.append('')

    # SCALE dict
    lines += [
        'SCALE = {',
        "    'voltage':      0.1,",
        "    'current':      0.01,",
        "    'power':        0.1,",
        "    'frequency':    0.01,",
        "    'power_factor': 0.001,",
    ]
    if dea_avm:
        lines += [
            "    'dea_current':          0.1,",
            "    'dea_voltage':          0.1,",
            "    'dea_active_power':     0.1,",
            "    'dea_reactive_power':   1,",
            "    'dea_frequency':        0.1,",
        ]
    if iv_scan:
        lines += [
            "    'iv_voltage': 0.1,",
            "    'iv_current': 0.1,",
        ]
    lines += ['}', '']

    # DATA_TYPES dict
    lines.append('DATA_TYPES = {')
    for reg in all_regs_ordered:
        if reg['dtype'] in ('U16', 'S16', 'U32', 'S32', 'FLOAT', 'ASCII'):
            lines.append(f"    '{reg['name']}': '{reg['dtype']}',")
    lines += ['}', '']

    # InverterMode class — codes from Solarize Modbus Protocol V1.2.4 Table (0x101D)
    lines += [
        'class InverterMode:',
        '    """Inverter operating mode codes (register 0x101D)."""',
        '    INITIAL  = 0x00  # Initial mode',
        '    STANDBY  = 0x01  # Standby mode',
        '    ON_GRID  = 0x03  # On-Grid mode (grid-connected)',
        '    OFF_GRID = 0x04  # Off-Grid mode (island)',
        '    FAULT    = 0x05  # Fault mode',
        '    SHUTDOWN = 0x09  # Shutdown mode',
        '',
        '    _MAP = {',
        "        0x00: 'Initial',",
        "        0x01: 'Standby',",
        "        0x03: 'On-Grid',",
        "        0x04: 'Off-Grid',",
        "        0x05: 'Fault',",
        "        0x09: 'Shutdown',",
        '    }',
        '',
        '    @classmethod',
        '    def to_string(cls, value):',
        '        return cls._MAP.get(int(value) if value is not None else -1,',
        "                            f'Unknown({value})')",
        '',
    ]

    # StatusConverter
    lines += ['class StatusConverter:', '    """Convert raw inverter status code to InverterMode."""', '']

    if status_map:
        lines.append('    _RAW_TO_MODE = {')
        for raw_code, (mode_name, sol_code) in sorted(status_map.items()):
            target = sol_code if isinstance(sol_code, int) and sol_code >= 0 else -1
            lines.append(f'        {raw_code}: {target},  # {mode_name}')
        lines.append('    }')
        lines.append('')
        lines += [
            '    @classmethod',
            '    def to_inverter_mode(cls, raw_value):',
            '        if raw_value is None:',
            '            return InverterMode.FAULT',
            '        mapped = cls._RAW_TO_MODE.get(int(raw_value), -1)',
            '        if mapped >= 0:',
            '            return mapped',
            '        # Fallback: pass-through for known Solarize mode codes',
            '        v = int(raw_value)',
            '        return v if v in (0x00, 0x01, 0x03, 0x04, 0x05, 0x09) else InverterMode.FAULT',
            '',
        ]
    else:
        lines += [
            '    @classmethod',
            '    def to_inverter_mode(cls, raw_value):',
            '        if raw_value is None:',
            '            return InverterMode.FAULT',
            '        v = int(raw_value)',
            '        return v if v in (0x00, 0x01, 0x03, 0x04, 0x05, 0x09) else InverterMode.FAULT',
            '',
        ]

    # Helper functions
    lines += [
        'def registers_to_u32(low, high):',
        '    """Combine two U16 registers (low-word, high-word) into U32."""',
        '    return ((int(high) & 0xFFFF) << 16) | (int(low) & 0xFFFF)',
        '',
        'def registers_to_s32(low, high):',
        '    """Combine two U16 registers (low-word, high-word) into signed S32."""',
        '    u = registers_to_u32(low, high)',
        '    return u - 0x100000000 if u >= 0x80000000 else u',
        '',
    ]

    # MPPT helper
    lines += [
        'def get_mppt_registers(mppt_index):',
        '    """Return (voltage_addr, current_addr, power_low_addr, power_high_addr) for MPPT channel."""',
        f'    if not (1 <= mppt_index <= {mppt_count}):',
        "        raise ValueError(f'MPPT index {mppt_index} out of range [1, " + str(mppt_count) + "]')",
        f'    r = {class_name}',
        "    v_name  = f'MPPT{mppt_index}_VOLTAGE'",
        "    i_name  = f'MPPT{mppt_index}_CURRENT'",
        "    pl_name = f'MPPT{mppt_index}_POWER_LOW'",
        "    ph_name = f'MPPT{mppt_index}_POWER_HIGH'",
        '    return (',
        '        getattr(r, v_name,  None),',
        '        getattr(r, i_name,  None),',
        '        getattr(r, pl_name, None),',
        '        getattr(r, ph_name, None),',
        '    )',
        '',
    ]

    # String helper
    lines += [
        'def get_string_registers(string_index):',
        '    """Return (voltage_addr, current_addr) for PV string channel."""',
        f'    if not (1 <= string_index <= {string_count}):',
        "        raise ValueError(f'String index {string_index} out of range [1, " + str(string_count) + "]')",
        f'    r = {class_name}',
        "    v_name = f'STRING{string_index}_VOLTAGE'",
        "    i_name = f'STRING{string_index}_CURRENT'",
        '    return (',
        '        getattr(r, v_name, None),',
        '        getattr(r, i_name, None),',
        '    )',
        '',
    ]

    code = '\n'.join(lines)

    # ── Validate generated code ──
    _p("  Running 12-item validation tests...")
    results = _validate_register_code(code, class_name, der_avm)
    fail_count = sum(1 for s, _ in results if s == 'FAIL')
    success = fail_count == 0
    _p(f"  Validation: {sum(1 for s,_ in results if s=='PASS')} PASS, "
       f"{sum(1 for s,_ in results if s=='WARN')} WARN, {fail_count} FAIL")

    # ── AI mode: auto-retry on failure (max 3 attempts) ──
    if not success and mode == 'ai' and api_key:
        _p("  AI mode: auto-retry on validation failure...")
        code, results, success = _ai_stage3_retry(
            code=code, results=results,
            settings=settings, api_key=api_key, model=model,
            class_name=class_name, der_avm=der_avm,
            max_retries=3, progress_cb=_p,
        )

    # ── Save if output_path given and success ──
    if output_path and success:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(code + '\n')
        _p(f"  Stage 3 complete -> {output_path}")
    elif output_path and not success:
        _p("  Validation FAILED — file not saved. Fix mapping Excel and re-generate.")

    return {
        'code':        code,
        'results':     results,
        'success':     success,
        'output_path': output_path if (output_path and success) else None,
        'mode':        mode,
    }


# ---------------------------------------------------------------------------
# AI helper: Stage 3 auto-retry via Claude API
# ---------------------------------------------------------------------------

def _ai_stage3_retry(code, results, settings, api_key, model,
                      class_name, der_avm, max_retries=3, progress_cb=None):
    """
    Retry code generation via Claude API when validation fails.
    Returns (code, results, success).
    """
    try:
        import anthropic
    except ImportError:
        if progress_cb:
            progress_cb("  AI retry skipped: anthropic package not installed")
        return code, results, False

    client = anthropic.Anthropic(api_key=api_key)
    mdl = model or 'claude-opus-4-6'

    for attempt in range(1, max_retries + 1):
        fail_msgs = [msg for s, msg in results if s == 'FAIL']
        if progress_cb:
            progress_cb(f"  AI retry {attempt}/{max_retries}: {len(fail_msgs)} FAIL items")

        prompt = (
            f"Fix the following Python inverter register file to pass all validation tests.\n\n"
            f"Manufacturer: {settings.get('manufacturer', 'Unknown')}\n"
            f"class_name: {class_name}\n"
            f"DER-AVM required: {der_avm}\n\n"
            f"Current validation FAILURES:\n"
            + '\n'.join(f'  - {m}' for m in fail_msgs)
            + '\n\nCurrent code:\n```python\n' + code + '\n```\n\n'
            'Return ONLY the corrected Python code inside a ```python block. '
            'Keep all existing correct parts. Fix only the failing items.'
        )

        try:
            resp = client.messages.create(
                model=mdl, max_tokens=8192,
                messages=[{'role': 'user', 'content': prompt}]
            )
            text = resp.content[0].text.strip()
        except Exception as e:
            if progress_cb:
                progress_cb(f"  AI API error: {e}")
            break

        # Extract code from markdown block
        m = re.search(r'```python\s*(.*?)\s*```', text, re.DOTALL)
        if not m:
            m = re.search(r'```\s*(.*?)\s*```', text, re.DOTALL)
        if m:
            new_code = m.group(1)
        elif text.lstrip().startswith(('# ', '"""', 'class ', 'import ')):
            new_code = text
        else:
            if progress_cb:
                progress_cb(f"  AI returned no extractable code")
            continue

        # Validate new code
        new_results = _validate_register_code(new_code, class_name, der_avm)
        new_fails   = sum(1 for s, _ in new_results if s == 'FAIL')

        if progress_cb:
            new_pass = sum(1 for s, _ in new_results if s == 'PASS')
            new_warn = sum(1 for s, _ in new_results if s == 'WARN')
            progress_cb(f"  Attempt {attempt}: {new_pass}P {new_warn}W {new_fails}F")

        code    = new_code
        results = new_results
        if new_fails == 0:
            if progress_cb:
                progress_cb(f"  AI retry SUCCESS on attempt {attempt}")
            return code, results, True

    fail_count = sum(1 for s, _ in results if s == 'FAIL')
    return code, results, fail_count == 0


# ---------------------------------------------------------------------------
# Save to Reference Library
# ---------------------------------------------------------------------------

def save_to_reference(protocol, py_code, mapping_excel_path=None, meta=None):
    """
    Save a successfully generated register file to the reference library.

    Args:
        protocol           : Reference name key (e.g. 'delta', 'fronius')
        py_code            : Generated Python source code string
        mapping_excel_path : Optional path to Stage 2 mapping Excel
        meta               : dict {manufacturer, mppt_count, string_count, ...}

    Returns:
        (success: bool, message: str)
    """
    ref_mgr = _get_ref_manager()
    if ref_mgr is None:
        return False, "reference_manager module not available"
    try:
        ref_mgr.save_reference(protocol, py_code, meta or {}, mapping_excel_path)
        return True, f"Saved reference: {protocol}"
    except Exception as e:
        return False, str(e)


# ---------------------------------------------------------------------------
# Validation (12-item test suite)
# ---------------------------------------------------------------------------

def _validate_register_code(code, class_name='RegisterMap', der_avm=True):
    """
    Run 12-item validation on generated register code.
    Returns list of (status, message) tuples.
    status: 'PASS', 'WARN', 'FAIL'
    """
    results = []
    ns = {}
    reg_cls = None

    # 1. Syntax
    try:
        compile(code, '<stage3>', 'exec')
        results.append(('PASS', '1. Syntax check'))
    except SyntaxError as e:
        results.append(('FAIL', f'1. Syntax error: {e}'))
        return results

    # 2. Importability / class existence
    try:
        # Inject __file__ so generated sys.path.insert() does not raise NameError
        ns['__file__'] = '<stage3_validation>'
        exec(code, ns)
        reg_cls = ns.get(class_name)
        if reg_cls is None:
            results.append(('FAIL', f'2. Class "{class_name}" not found'))
            return results
        results.append(('PASS', f'2. Class "{class_name}" importable'))
    except Exception as e:
        results.append(('FAIL', f'2. Exec error: {e}'))
        return results

    # 3. Register constants are non-negative integers
    try:
        attrs = {k: v for k, v in vars(reg_cls).items()
                 if not k.startswith('_') and not callable(v)}
        bad = [k for k, v in attrs.items() if not isinstance(v, int) or v < 0]
        if bad:
            results.append(('WARN', f'3. Non-integer/negative constants: {", ".join(bad[:5])}'))
        else:
            results.append(('PASS', f'3. Register constants: {len(attrs)} valid integers'))
    except Exception as e:
        results.append(('WARN', f'3. Could not inspect constants: {e}'))

    # 4. Essential aliases (modbus_handler compatibility)
    essential = ['L1_VOLTAGE', 'L1_CURRENT', 'INVERTER_MODE',
                 'TOTAL_ENERGY_LOW', 'TODAY_ENERGY_LOW',
                 'MPPT1_VOLTAGE', 'MPPT1_CURRENT']
    missing = [a for a in essential if not hasattr(reg_cls, a)]
    if len(missing) > 4:
        results.append(('FAIL', f'4. Missing essential aliases: {", ".join(missing)}'))
    elif missing:
        results.append(('WARN', f'4. Missing optional aliases: {", ".join(missing)}'))
    else:
        results.append(('PASS', '4. Essential aliases all present'))

    # 5. Address uniqueness
    addr_to_names = {}
    for k, v in vars(reg_cls).items():
        if isinstance(v, int) and v >= 0 and not k.startswith('_') and not callable(v):
            addr_to_names.setdefault(v, []).append(k)
    dups = {a: ns_ for a, ns_ in addr_to_names.items() if len(ns_) > 3}
    if dups:
        results.append(('WARN', f'5. Many aliases at same addr ({len(dups)} addrs with >3 names)'))
    else:
        results.append(('PASS', '5. Address uniqueness OK'))

    # 6. SCALE dict
    scale = ns.get('SCALE')
    if scale is None:
        results.append(('FAIL', '6. SCALE dict not found'))
    else:
        required_keys = ['voltage', 'current', 'power', 'frequency', 'power_factor']
        missing_k = [k for k in required_keys if k not in scale]
        if missing_k:
            results.append(('WARN', f'6. SCALE missing standard keys: {", ".join(missing_k)}'))
        else:
            results.append(('PASS', f'6. SCALE dict: {len(scale)} entries, standard keys OK'))

    # 7. InverterMode class
    inv_mode = ns.get('InverterMode')
    if inv_mode is None:
        results.append(('FAIL', '7. InverterMode class not found'))
    else:
        req = ['INITIAL', 'STANDBY', 'ON_GRID', 'OFF_GRID', 'FAULT', 'SHUTDOWN']
        missing_m = [a for a in req if not hasattr(inv_mode, a)]
        if missing_m:
            results.append(('FAIL', f'7. InverterMode missing: {", ".join(missing_m)}'))
        elif not hasattr(inv_mode, 'to_string'):
            results.append(('WARN', '7. InverterMode: to_string() not found'))
        else:
            results.append(('PASS', '7. InverterMode: all states + to_string() OK'))

    # 8. registers_to_u32 / registers_to_s32
    fn_u32 = ns.get('registers_to_u32')
    fn_s32 = ns.get('registers_to_s32')
    if fn_u32 is None or fn_s32 is None:
        missing_fn = []
        if fn_u32 is None: missing_fn.append('registers_to_u32')
        if fn_s32 is None: missing_fn.append('registers_to_s32')
        results.append(('FAIL', f'8. Missing: {", ".join(missing_fn)}'))
    else:
        try:
            u32_ok = (fn_u32(0x1234, 0x5678) == 0x56781234 or
                      fn_u32(0x1234, 0x5678) == 0x12345678)
            s32_ok = (fn_s32(0, 0x8000) < 0 or fn_s32(0x8000, 0) < 0)
            if u32_ok and s32_ok:
                results.append(('PASS', '8. registers_to_u32, registers_to_s32 OK'))
            else:
                results.append(('FAIL', '8. registers_to_u32/s32 wrong return values'))
        except Exception as e:
            results.append(('FAIL', f'8. Helper function error: {e}'))

    # 9. get_mppt_registers
    fn_mppt = ns.get('get_mppt_registers')
    if fn_mppt is None:
        results.append(('FAIL', '9. get_mppt_registers() not found'))
    else:
        try:
            r = fn_mppt(1)
            if isinstance(r, (tuple, list)) and len(r) >= 2:
                results.append(('PASS', '9. get_mppt_registers(1) returns tuple/list'))
            else:
                results.append(('WARN', '9. get_mppt_registers(1) unexpected return type'))
        except Exception as e:
            results.append(('FAIL', f'9. get_mppt_registers error: {e}'))

    # 10. get_string_registers
    fn_str = ns.get('get_string_registers')
    if fn_str is None:
        results.append(('FAIL', '10. get_string_registers() not found'))
    else:
        try:
            r = fn_str(1)
            if isinstance(r, (tuple, list)) and len(r) >= 2:
                results.append(('PASS', '10. get_string_registers(1) returns tuple/list'))
            else:
                results.append(('WARN', '10. get_string_registers(1) unexpected return type'))
        except Exception as e:
            results.append(('FAIL', f'10. get_string_registers error: {e}'))

    # 11. DATA_TYPES dict
    data_types = ns.get('DATA_TYPES')
    if data_types is None:
        results.append(('WARN', '11. DATA_TYPES dict not found'))
    else:
        valid_types = {'U16', 'S16', 'U32', 'S32', 'FLOAT', 'ASCII'}
        bad_t = [k for k, v in data_types.items() if v not in valid_types]
        if bad_t:
            results.append(('WARN', f'11. DATA_TYPES has unknown types: {", ".join(bad_t[:5])}'))
        else:
            results.append(('PASS', f'11. DATA_TYPES: {len(data_types)} entries OK'))

    # 12. DER-AVM attributes
    if der_avm:
        req_der = ['DER_POWER_FACTOR_SET', 'DER_ACTION_MODE',
                   'DER_REACTIVE_POWER_PCT', 'DER_ACTIVE_POWER_PCT',
                   'INVERTER_ON_OFF', 'INVERTER_CONTROL']
        missing_d = [a for a in req_der if not hasattr(reg_cls, a)]
        if missing_d:
            results.append(('FAIL', f'12. DER-AVM attrs missing: {", ".join(missing_d)}'))
        else:
            results.append(('PASS', '12. DER-AVM attributes all present'))
    else:
        results.append(('PASS', '12. DER-AVM: skipped (not enabled)'))

    return results
