# -*- coding: utf-8 -*-
"""
Stage 2 — MPPT/String 필터링 + H01/DER 매칭 검증 → Stage 2 Excel

Stage 1 Excel에서 실제 인버터 용량에 맞게 MPPT/String 채널을 필터링하고,
UDP 프로토콜 H01 바디/DER 필드와의 매칭 상태를 검증한다.
"""
import os
import re
from datetime import datetime
from typing import List, Dict, Optional

from . import (
    PROJECT_ROOT, UDP_PROTOCOL_DIR,
    RegisterRow, CATEGORY_COLORS, MATCH_COLORS,
    load_reference_patterns,
    to_upper_snake, parse_address, match_synonym, match_synonym_fuzzy,
    detect_channel_number,
    load_synonym_db, load_review_history,
    get_openpyxl, ProgressCallback,
)


# ─── UDP 프로토콜 Excel H01 필드 파싱 ────────────────────────────────────────

def _find_udp_excel() -> Optional[str]:
    """UDP 프로토콜 Excel 자동 탐색"""
    import glob
    pattern = os.path.join(UDP_PROTOCOL_DIR, 'Solarize_RTU2Server_UDP_*.xlsx')
    files = sorted(glob.glob(pattern), reverse=True)
    return files[0] if files else None


def parse_udp_h01_fields(device_type: str = 'inverter') -> List[dict]:
    """
    UDP 프로토콜 Excel에서 H01 바디 필드 파싱
    Returns: [{'name': str, 'h01_key': str, 'size': int, 'type': str, 'unit': str}]
    """
    openpyxl = get_openpyxl()
    udp_path = _find_udp_excel()
    if not udp_path:
        return _fallback_h01_fields(device_type)

    wb = openpyxl.load_workbook(udp_path, data_only=True)

    # 설비 타입별 시트 선택
    sheet_map = {
        'inverter': '(P) BODY(인버터)',
        'relay': '(P) BODY(보호계전기)',
        'weather': '(P) BODY(환경센서)',
    }
    target_sheet = sheet_map.get(device_type)
    if not target_sheet or target_sheet not in wb.sheetnames:
        # 유사한 시트명 탐색
        for sn in wb.sheetnames:
            if device_type in sn.lower() or ('인버터' in sn and device_type == 'inverter'):
                target_sheet = sn
                break
        if target_sheet not in wb.sheetnames:
            wb.close()
            return _fallback_h01_fields(device_type)

    ws = wb[target_sheet]
    fields = []
    header_row = None

    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else '' for c in row]
        if not header_row:
            # Name/이름 컬럼이 있는 행을 헤더로
            for i, c in enumerate(cells):
                if any(k in c.lower() for k in ['name', '이름', '항목', 'field']):
                    header_row = cells
                    break
            continue
        if not any(cells):
            continue
        # 필드 파싱
        name = ''
        size = 0
        dtype = ''
        unit = ''
        for i, c in enumerate(cells):
            cl = header_row[i].lower() if i < len(header_row) else ''
            if any(k in cl for k in ['name', '이름', '항목']):
                name = c
            elif any(k in cl for k in ['size', '크기', 'byte', 'length']):
                try:
                    size = int(c)
                except (ValueError, TypeError):
                    pass
            elif any(k in cl for k in ['type', '타입']):
                dtype = c
            elif any(k in cl for k in ['unit', '단위']):
                unit = c

        if name:
            h01_key = _name_to_h01_key(name)
            fields.append({'name': name, 'h01_key': h01_key, 'size': size,
                           'type': dtype, 'unit': unit})

    wb.close()
    return fields


def _name_to_h01_key(name: str) -> str:
    """UDP Excel Name → h01_key 변환"""
    name_lower = name.lower().strip()
    key_map = {
        'pv 전압': 'pv_voltage', 'pv전압': 'pv_voltage',
        'pv 전류': 'pv_current', 'pv전류': 'pv_current',
        'pv 출력': 'pv_power', 'pv출력': 'pv_power',
        'r (rs)상 전압': 'r_voltage', 'r상 전압': 'r_voltage', 'r상전압': 'r_voltage',
        's (st)상 전압': 's_voltage', 's상 전압': 's_voltage', 's상전압': 's_voltage',
        't (tr)상 전압': 't_voltage', 't상 전압': 't_voltage', 't상전압': 't_voltage',
        'r상 전류': 'r_current', 'r상전류': 'r_current',
        's상 전류': 's_current', 's상전류': 's_current',
        't상 전류': 't_current', 't상전류': 't_current',
        '인버터 출력': 'ac_power', '유효전력': 'ac_power', 'ac출력': 'ac_power',
        '역률': 'power_factor', 'power factor': 'power_factor',
        '주파수': 'frequency', 'frequency': 'frequency',
        '무효전력': 'reactive_power', 'reactive power': 'reactive_power',
        '피상전력': 'apparent_power', 'apparent power': 'apparent_power',
        '일일 발전량': 'daily_energy', '금일발전량': 'daily_energy',
        '누적 발전량': 'cumulative_energy', '누적발전량': 'cumulative_energy',
        '상태 정보 1': 'status', '상태정보1': 'status',
        '상태 정보 2': 'alarm1', '상태정보2': 'alarm1',
        '상태 정보 3': 'alarm2', '상태정보3': 'alarm2',
        '상태 정보 4': 'alarm3', '상태정보4': 'alarm3',
        '인버터 온도': 'temperature', '온도': 'temperature',
    }
    for k, v in key_map.items():
        if k in name_lower:
            return v
    # MPPT/String 패턴
    m = re.search(r'mppt(\d+)', name_lower)
    if m:
        n = m.group(1)
        if '전압' in name_lower or 'voltage' in name_lower:
            return f'mppt{n}_voltage'
        if '전류' in name_lower or 'current' in name_lower:
            return f'mppt{n}_current'
    m = re.search(r'string(\d+)', name_lower)
    if m:
        n = m.group(1)
        if '전류' in name_lower or 'current' in name_lower:
            return f'string{n}_current'

    return to_upper_snake(name).lower()


def _fallback_h01_fields(device_type: str) -> List[dict]:
    """UDP Excel 없을 때 기본 H01 필드"""
    if device_type != 'inverter':
        return []
    return [
        {'name': 'PV 전압', 'h01_key': 'pv_voltage', 'size': 2},
        {'name': 'PV 전류', 'h01_key': 'pv_current', 'size': 2},
        {'name': 'PV 출력', 'h01_key': 'pv_power', 'size': 4},
        {'name': 'R상 전압', 'h01_key': 'r_voltage', 'size': 2},
        {'name': 'S상 전압', 'h01_key': 's_voltage', 'size': 2},
        {'name': 'T상 전압', 'h01_key': 't_voltage', 'size': 2},
        {'name': 'R상 전류', 'h01_key': 'r_current', 'size': 2},
        {'name': 'S상 전류', 'h01_key': 's_current', 'size': 2},
        {'name': 'T상 전류', 'h01_key': 't_current', 'size': 2},
        {'name': '인버터 출력', 'h01_key': 'ac_power', 'size': 4},
        {'name': '역률', 'h01_key': 'power_factor', 'size': 2},
        {'name': '주파수', 'h01_key': 'frequency', 'size': 2},
        {'name': '무효전력', 'h01_key': 'reactive_power', 'size': 4},
        {'name': '피상전력', 'h01_key': 'apparent_power', 'size': 4},
        {'name': '일일 발전량', 'h01_key': 'daily_energy', 'size': 4},
        {'name': '누적 발전량', 'h01_key': 'cumulative_energy', 'size': 8},
        {'name': '상태 정보 1', 'h01_key': 'status', 'size': 2},
        {'name': '상태 정보 2', 'h01_key': 'alarm1', 'size': 2},
        {'name': '상태 정보 3', 'h01_key': 'alarm2', 'size': 2},
        {'name': '상태 정보 4', 'h01_key': 'alarm3', 'size': 2},
        {'name': '인버터 온도', 'h01_key': 'temperature', 'size': 2},
    ]


# ─── DER 필드 정의 ──────────────────────────────────────────────────────────

def _get_der_control_fields() -> List[dict]:
    """DER-AVM 제어 필드"""
    return [
        {'name': 'Power Factor Set', 'key': 'power_factor_set'},
        {'name': 'Action Mode', 'key': 'action_mode'},
        {'name': 'Reactive Power %', 'key': 'reactive_power_pct'},
        {'name': 'Active Power %', 'key': 'active_power_pct'},
        {'name': 'ON/OFF', 'key': 'on_off'},
    ]


def _get_der_monitor_fields() -> List[dict]:
    """DER-AVM 모니터링 필드"""
    return [
        {'name': 'DEA L1 Current', 'key': 'dea_l1_current'},
        {'name': 'DEA L2 Current', 'key': 'dea_l2_current'},
        {'name': 'DEA L3 Current', 'key': 'dea_l3_current'},
        {'name': 'DEA L1 Voltage', 'key': 'dea_l1_voltage'},
        {'name': 'DEA L2 Voltage', 'key': 'dea_l2_voltage'},
        {'name': 'DEA L3 Voltage', 'key': 'dea_l3_voltage'},
        {'name': 'DEA Active Power', 'key': 'dea_active_power'},
        {'name': 'DEA Reactive Power', 'key': 'dea_reactive_power'},
        {'name': 'DEA Power Factor', 'key': 'dea_power_factor'},
        {'name': 'DEA Frequency', 'key': 'dea_frequency'},
        {'name': 'DEA Status Flag', 'key': 'dea_status_flag'},
    ]


# ─── Stage 1 Excel 읽기 ─────────────────────────────────────────────────────

def read_stage1_excel(path: str) -> dict:
    """Stage 1 Excel → {meta, registers_by_category}"""
    openpyxl = get_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True)

    # META
    meta = {}
    if 'META' in wb.sheetnames:
        ws = wb['META']
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            if len(cells) >= 2 and cells[0]:
                meta[cells[0]] = cells[1]

    # 카테고리별 레지스터
    categories = {}
    for cat in ['INFO', 'MONITORING', 'STATUS', 'ALARM',
                'DER_CONTROL', 'DER_MONITOR', 'IV_SCAN', 'REVIEW']:
        if cat not in wb.sheetnames:
            continue
        ws = wb[cat]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            categories[cat] = []
            continue

        header = [str(c).strip() if c else '' for c in rows[0]]
        regs = []
        for row in rows[1:]:
            cells = [str(c).strip() if c is not None else '' for c in row]
            if not any(cells):
                continue
            d = {}
            for i, h in enumerate(header):
                if i < len(cells):
                    hl = h.lower()
                    if 'definition' in hl or 'name' in hl:
                        d['definition'] = cells[i]
                    elif 'address' in hl:
                        d['address_hex'] = cells[i]
                        d['address'] = parse_address(cells[i]) or 0
                    elif 'type' in hl:
                        d['data_type'] = cells[i]
                    elif 'unit' in hl or 'scale' in hl:
                        parts = cells[i].split()
                        d['unit'] = parts[0] if parts else ''
                        d['scale'] = parts[1] if len(parts) > 1 else ''
                    elif 'r/w' in hl or 'access' in hl:
                        d['rw'] = cells[i]
                    elif 'comment' in hl or 'remark' in hl:
                        d['comment'] = cells[i]
                    elif 'h01' in hl and 'field' in hl:
                        d['h01_field'] = cells[i]
                    elif 'reg' in hl:
                        d['regs'] = cells[i]
                    elif '사유' in h:
                        d['review_reason'] = cells[i]
                    elif '제안' in h:
                        d['review_suggestion'] = cells[i]
                    elif '판정' in h:
                        d['user_verdict'] = cells[i]
            d['category'] = cat
            regs.append(RegisterRow.from_dict(d))
        categories[cat] = regs

    wb.close()
    return {'meta': meta, 'categories': categories}


# ─── MPPT/String 필터링 ─────────────────────────────────────────────────────

def filter_channels(registers: List[RegisterRow],
                    mppt_count: int,
                    total_strings: int) -> List[RegisterRow]:
    """MPPT/String 채널 수에 따라 필터링"""
    filtered = []
    for reg in registers:
        ch = detect_channel_number(reg.definition)
        if ch:
            prefix, n = ch
            if prefix == 'MPPT' and n > mppt_count:
                continue
            if prefix == 'STRING' and n > total_strings:
                continue
        filtered.append(reg)
    return filtered


# ─── H01/DER 매칭 검증 ──────────────────────────────────────────────────────

def verify_h01_match(all_regs: List[RegisterRow],
                     h01_fields: List[dict],
                     synonym_db: dict) -> List[dict]:
    """
    H01 필드별 매칭 상태 검증
    Returns: [{'h01_field': str, 'h01_name': str, 'matched_reg': str, 'match': 'O'|'X'|'-'}]
    """
    results = []
    # 레지스터의 h01_field 인덱스
    h01_index = {}
    for reg in all_regs:
        if reg.h01_field:
            h01_index.setdefault(reg.h01_field, []).append(reg)

    for field in h01_fields:
        key = field['h01_key']
        name = field['name']

        # MPPT/String 패턴
        m = re.match(r'(mppt|string)(\d+)_(voltage|current|power)', key)
        if m:
            # 채널 데이터는 별도 처리 (ALL에서 확인)
            prefix = m.group(1).upper()
            n = int(m.group(2))
            metric = m.group(3).upper()
            defn_pattern = f'{prefix}{n}_{metric}'
            found = False
            for reg in all_regs:
                if to_upper_snake(reg.definition).endswith(defn_pattern):
                    found = True
                    break
                if reg.h01_field == key:
                    found = True
                    break
            results.append({
                'h01_field': key, 'h01_name': name,
                'matched_reg': defn_pattern if found else '',
                'matched_addr': '',
                'scale': '',
                'match': 'O' if found else 'X',
            })
            continue

        # 일반 필드
        matched = h01_index.get(key, [])
        if matched:
            reg = matched[0]
            results.append({
                'h01_field': key, 'h01_name': name,
                'matched_reg': reg.definition,
                'matched_addr': reg.address_hex,
                'scale': reg.scale,
                'match': 'O',
            })
        else:
            # synonym_db로 재탐색
            found_reg = None
            for reg in all_regs:
                sm = match_synonym(reg.definition, synonym_db)
                if sm and sm.get('h01_field') == key:
                    found_reg = reg
                    break
            if found_reg:
                results.append({
                    'h01_field': key, 'h01_name': name,
                    'matched_reg': found_reg.definition,
                    'matched_addr': found_reg.address_hex,
                    'scale': found_reg.scale,
                    'match': 'O',
                })
            else:
                results.append({
                    'h01_field': key, 'h01_name': name,
                    'matched_reg': '', 'matched_addr': '',
                    'scale': '', 'match': 'X',
                })

    return results


def verify_der_match(all_regs: List[RegisterRow],
                     synonym_db: dict) -> List[dict]:
    """
    DER-AVM 제어/모니터링 매칭 검증
    """
    results = []
    der_regs = [r for r in all_regs if r.category in ('DER_CONTROL', 'DER_MONITOR')]

    # 제어 필드
    for field in _get_der_control_fields():
        found = any(field['key'] in (r.h01_field or '').lower() or
                    field['key'].replace('_', '') in to_upper_snake(r.definition).lower().replace('_', '')
                    for r in der_regs)
        if not found:
            # 주소 기반
            from . import DER_CONTROL_REGS
            for dr in DER_CONTROL_REGS:
                for r in all_regs:
                    a = r.address if isinstance(r.address, int) else parse_address(r.address)
                    if a == dr['addr']:
                        found = True
                        break
                if found:
                    break

        results.append({
            'field': field['name'], 'key': field['key'],
            'type': 'CONTROL', 'match': 'O' if found else 'X',
        })

    # 모니터링 필드
    for field in _get_der_monitor_fields():
        found = any(field['key'] in (r.h01_field or '').lower() or
                    field['key'].replace('_', '') in to_upper_snake(r.definition).lower().replace('_', '')
                    for r in der_regs)
        if not found:
            from . import DER_MONITOR_REGS
            for dr in DER_MONITOR_REGS:
                for r in all_regs:
                    a = r.address if isinstance(r.address, int) else parse_address(r.address)
                    if a == dr['addr']:
                        found = True
                        break
                if found:
                    break

        results.append({
            'field': field['name'], 'key': field['key'],
            'type': 'MONITOR', 'match': 'O' if found else 'X',
        })

    return results


# ─── Stage 2 메인 함수 ───────────────────────────────────────────────────────

def run_stage2(
    stage1_path: str,
    output_dir: str,
    mppt_count: int,
    strings_per_mppt: int,
    capacity: str = '',
    progress: ProgressCallback = None,
) -> dict:
    """
    Stage 2 실행: Stage 1 Excel + 파라미터 → Stage 2 Excel

    Returns:
        {'output_path', 'h01_matched', 'h01_total', 'der_matched', 'der_total',
         'review_count', 'register_count', 'counts'}
    """
    def log(msg, level='info'):
        if progress:
            progress(msg, level)

    openpyxl = get_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    total_strings = mppt_count * strings_per_mppt
    log(f'MPPT: {mppt_count}, Strings/MPPT: {strings_per_mppt}, Total: {total_strings}')

    # ── Step 1: Stage 1 Excel 읽기 ──
    log('Stage 1 Excel 읽기...')
    s1_data = read_stage1_excel(stage1_path)
    meta = s1_data['meta']
    categories = s1_data['categories']

    device_type = meta.get('device_type', 'inverter')
    manufacturer = meta.get('manufacturer', 'Unknown')

    # ── Step 2: REVIEW 자동 처리 ──
    log('REVIEW 항목 처리...')
    review_history = load_review_history()
    synonym_db = load_synonym_db()

    review_regs = categories.get('REVIEW', [])
    auto_processed = 0
    remaining_review = []

    for reg in review_regs:
        # review_history에서 동일 항목 찾기
        auto_applied = False
        for item in review_history.get('approved', []):
            if (item.get('definition', '').upper() == to_upper_snake(reg.definition) or
                    item.get('address') == reg.address_hex):
                verdict = item.get('verdict', '')
                if verdict == 'DELETE':
                    auto_processed += 1
                    auto_applied = True
                    break
                elif verdict.startswith('MOVE:'):
                    target_cat = verdict.replace('MOVE:', '')
                    reg.category = target_cat
                    reg.comment = f'(auto-applied from review_history: {verdict})'
                    categories.setdefault(target_cat, []).append(reg)
                    auto_processed += 1
                    auto_applied = True
                    break
        if not auto_applied:
            remaining_review.append(reg)

    categories['REVIEW'] = remaining_review
    if auto_processed:
        log(f'  REVIEW 자동 처리: {auto_processed}개')
    log(f'  REVIEW 미결: {len(remaining_review)}개')

    # ── Step 3: MPPT/String 필터링 (§8: rules.py) ──
    log('MPPT/String 필터링...')
    from .rules import filter_channels_stage2
    ref_patterns = load_reference_patterns()
    for cat in ['MONITORING', 'IV_SCAN']:
        if cat in categories:
            before = len(categories[cat])
            categories[cat] = filter_channels_stage2(
                categories[cat], mppt_count, total_strings, ref_patterns)
            after = len(categories[cat])
            if before != after:
                log(f'  {cat}: {before} → {after}')

    # ALL 목록 구성
    all_regs = []
    for cat in ['INFO', 'MONITORING', 'STATUS', 'ALARM',
                'DER_CONTROL', 'DER_MONITOR', 'IV_SCAN']:
        all_regs.extend(categories.get(cat, []))

    log(f'필터링 후 총 레지스터: {len(all_regs)}개')

    # ── Step 4: H01 매칭 검증 ──
    log('H01 매칭 검증...')
    h01_fields = parse_udp_h01_fields(device_type)
    # MPPT/String 필드 제거 (동적이라 별도 처리)
    h01_base_fields = [f for f in h01_fields
                       if not re.match(r'mppt\d+|string\d+', f['h01_key'])]
    h01_match_results = verify_h01_match(all_regs, h01_base_fields, synonym_db)
    h01_matched = sum(1 for r in h01_match_results if r['match'] == 'O')
    h01_total = len(h01_match_results)
    log(f'  H01 매칭: {h01_matched}/{h01_total}')

    # ── Step 5: DER 매칭 검증 ──
    der_match_results = []
    der_matched = 0
    der_total = 0
    if device_type == 'inverter':
        log('DER 매칭 검증...')
        der_match_results = verify_der_match(all_regs, synonym_db)
        der_matched = sum(1 for r in der_match_results if r['match'] == 'O')
        der_total = len(der_match_results)
        log(f'  DER 매칭: {der_matched}/{der_total}')

    # ── Step 6: Stage 2 Excel 생성 ──
    basename = os.path.splitext(os.path.basename(stage1_path))[0].replace('_stage1', '')
    cap_str = f'_{capacity}' if capacity else ''
    output_name = f'{basename}{cap_str}_MPPT{mppt_count}_STR{total_strings}_stage2.xlsx'
    output_path = os.path.join(output_dir, output_name)
    log(f'Excel 생성: {output_name}')

    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    # --- SUMMARY 시트 ---
    ws = wb.active
    ws.title = 'SUMMARY'
    ws['A1'] = 'Stage 2 — 필터링 & 매칭 검증 결과'
    ws['A1'].font = Font(bold=True, size=14)
    for i, (k, v) in enumerate([
        ('제조사', manufacturer),
        ('용량', capacity or '-'),
        ('MPPT', f'{mppt_count} (PDF max {meta.get("max_mppt", "?")}개)'),
        ('String', f'{total_strings} ({strings_per_mppt}/MPPT)'),
        ('H01 매칭', f'{h01_matched}/{h01_total}'),
        ('DER 매칭', f'{der_matched}/{der_total}' if device_type == 'inverter' else 'N/A'),
        ('총 레지스터', len(all_regs)),
        ('REVIEW 미결', len(remaining_review)),
    ], start=3):
        ws[f'A{i}'] = k
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = str(v)

    row_n = 12
    ws[f'A{row_n}'] = '카테고리별 수량'
    ws[f'A{row_n}'].font = Font(bold=True, size=12)
    counts = {}
    for cat in ['INFO', 'MONITORING', 'STATUS', 'ALARM',
                'DER_CONTROL', 'DER_MONITOR', 'IV_SCAN', 'REVIEW']:
        cnt = len(categories.get(cat, []))
        counts[cat] = cnt
        if cnt > 0:
            r = row_n + len(counts)
            ws[f'A{r}'] = cat
            ws[f'B{r}'] = cnt
            ws[f'A{r}'].fill = PatternFill('solid', fgColor=CATEGORY_COLORS.get(cat, 'FFFFFF'))

    # --- ALL 시트 ---
    ws_all = wb.create_sheet('ALL')
    all_cols = ['No', 'Category', 'Definition', 'Address', 'Reg', 'Type',
                'Unit/Scale', 'R/W', 'Comment', 'H01 Field', 'H01 Match', 'DER Match']
    for j, col_name in enumerate(all_cols, start=1):
        cell = ws_all.cell(row=1, column=j, value=col_name)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='333333')
        cell.border = thin_border

    for i, reg in enumerate(sorted(all_regs,
                                     key=lambda r: (r.address if isinstance(r.address, int) else 0)),
                            start=1):
        # H01 매칭 상태
        h01_match = '-'
        if reg.h01_field:
            for hm in h01_match_results:
                if hm['h01_field'] == reg.h01_field:
                    h01_match = hm['match']
                    break
            if h01_match == '-':
                h01_match = 'O'  # h01_field가 있으면 매칭된 것

        # DER 매칭 상태
        der_match = '-'
        if reg.category in ('DER_CONTROL', 'DER_MONITOR'):
            der_match = 'O'  # 표준 DER 레지스터는 항상 매칭

        scale_unit = f'{reg.unit} {reg.scale}'.strip() if reg.unit or reg.scale else ''
        row_data = [
            i, reg.category, reg.definition, reg.address_hex, reg.regs,
            reg.data_type, scale_unit, reg.rw, reg.comment, reg.h01_field,
            h01_match, der_match,
        ]
        for j, val in enumerate(row_data, start=1):
            cell = ws_all.cell(row=i + 1, column=j, value=val)
            cell.border = thin_border
            # 카테고리 색상
            if j == 2:
                cell.fill = PatternFill('solid', fgColor=CATEGORY_COLORS.get(reg.category, 'FFFFFF'))
            # 매칭 색상
            if j == 11 and val in MATCH_COLORS:
                cell.fill = PatternFill('solid', fgColor=MATCH_COLORS[val])
            if j == 12 and val in MATCH_COLORS:
                cell.fill = PatternFill('solid', fgColor=MATCH_COLORS[val])

    ws_all.column_dimensions['C'].width = 35
    ws_all.column_dimensions['I'].width = 30

    # --- H01_MATCH 시트 ---
    ws_h01 = wb.create_sheet('H01_MATCH')
    h01_cols = ['No', 'H01 Field', 'H01 Name', 'Matched Register', 'Address', 'Scale', 'Match']
    for j, cn in enumerate(h01_cols, start=1):
        cell = ws_h01.cell(row=1, column=j, value=cn)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='333333')
        cell.border = thin_border

    for i, hm in enumerate(h01_match_results, start=1):
        row_data = [i, hm['h01_field'], hm['h01_name'], hm['matched_reg'],
                    hm.get('matched_addr', ''), hm.get('scale', ''), hm['match']]
        for j, val in enumerate(row_data, start=1):
            cell = ws_h01.cell(row=i + 1, column=j, value=val)
            cell.border = thin_border
            if j == 7 and val in MATCH_COLORS:
                cell.fill = PatternFill('solid', fgColor=MATCH_COLORS[val])

    ws_h01.column_dimensions['B'].width = 20
    ws_h01.column_dimensions['C'].width = 25
    ws_h01.column_dimensions['D'].width = 35

    # --- DER_MATCH 시트 ---
    if device_type == 'inverter' and der_match_results:
        ws_der = wb.create_sheet('DER_MATCH')
        der_cols = ['No', 'Field', 'Key', 'Type', 'Match']
        for j, cn in enumerate(der_cols, start=1):
            cell = ws_der.cell(row=1, column=j, value=cn)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='333333')
            cell.border = thin_border

        for i, dm in enumerate(der_match_results, start=1):
            row_data = [i, dm['field'], dm['key'], dm['type'], dm['match']]
            for j, val in enumerate(row_data, start=1):
                cell = ws_der.cell(row=i + 1, column=j, value=val)
                cell.border = thin_border
                if j == 5 and val in MATCH_COLORS:
                    cell.fill = PatternFill('solid', fgColor=MATCH_COLORS[val])

        ws_der.column_dimensions['B'].width = 25

    # --- REVIEW 시트 (미결 항목) ---
    if remaining_review:
        ws_rv = wb.create_sheet('REVIEW')
        rv_cols = ['No', 'Definition', 'Address', 'Type', 'Unit/Scale', 'R/W',
                   'Comment', 'H01 Field(추정)', '사유', '제안', '사용자 판정']
        for j, cn in enumerate(rv_cols, start=1):
            cell = ws_rv.cell(row=1, column=j, value=cn)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='333333')
            cell.border = thin_border

        for i, reg in enumerate(remaining_review, start=1):
            scale_unit = f'{reg.unit} {reg.scale}'.strip()
            row_data = [i, reg.definition, reg.address_hex, reg.data_type,
                        scale_unit, reg.rw, reg.comment, reg.h01_field,
                        reg.review_reason, reg.review_suggestion, '']
            for j, val in enumerate(row_data, start=1):
                cell = ws_rv.cell(row=i + 1, column=j, value=val)
                cell.border = thin_border
                cell.fill = PatternFill('solid', fgColor=CATEGORY_COLORS['REVIEW'])

        ws_rv.column_dimensions['B'].width = 35
        ws_rv.column_dimensions['I'].width = 40
        ws_rv.column_dimensions['J'].width = 30
        ws_rv.column_dimensions['K'].width = 15

    # 저장
    wb.save(output_path)
    wb.close()
    log(f'Stage 2 완료: {output_name}', 'ok')

    return {
        'output_path': output_path,
        'output_name': output_name,
        'h01_matched': h01_matched,
        'h01_total': h01_total,
        'der_matched': der_matched,
        'der_total': der_total,
        'review_count': len(remaining_review),
        'register_count': len(all_regs),
        'counts': counts,
    }
