# -*- coding: utf-8 -*-
"""
Stage 1 — PDF/Excel → Stage 1 Excel 레지스터맵 추출 (MAPPING_RULES_V2)

V2 핵심 변경:
- H01 DER 겹침 필드 9개는 PDF 매핑 불필요 → DER 고정 주소
- pv_voltage/pv_current는 handler 계산 → PDF 매핑 불필요
- 제어 레지스터 완전 제외 (DER-AVM으로만 제어)
- H01_MATCH 시트 추가: 전체 H01 필드 매칭 상태 표시
"""


class NotRegisterMapError(ValueError):
    """PDF/Excel 파일이 Modbus 레지스터 맵 문서가 아닐 때 발생"""
    pass
import os
import re
import json
from datetime import datetime
from typing import List, Dict, Optional

from . import (
    PROJECT_ROOT, COMMON_DIR, MODEL_MAKER_DIR, UDP_PROTOCOL_DIR,
    RegisterRow, INVERTER_MODES, DER_CONTROL_REGS, DER_MONITOR_REGS,
    CATEGORY_COLORS, MATCH_COLORS,
    H01_DER_OVERLAP_FIELDS, H01_HANDLER_COMPUTED_FIELDS, H01_PDF_REQUIRED_FIELDS,
    to_upper_snake, parse_address, match_synonym, match_synonym_fuzzy,
    detect_channel_number, detect_channel_from_ref,
    get_ref_name_by_addr, get_h01_field_from_ref,
    load_synonym_db, load_review_history, load_reference_patterns,
    get_openpyxl, ProgressCallback,
)
from .rules import (
    classify_register_with_rules, detect_iv_scan_support,
    get_valid_categories, distribute_alarms, is_h01_der_overlap,
)

# ─── PDF 추출 ─────────────────────────────────────────────────────────────────

def extract_pdf_text_and_tables(pdf_path: str) -> List[dict]:
    """PyMuPDF로 PDF 페이지별 텍스트+테이블 추출"""
    import fitz
    import logging
    logging.getLogger('fitz').setLevel(logging.ERROR)
    fitz.TOOLS.mupdf_display_errors(False)
    doc = fitz.open(pdf_path)
    pages = []
    try:
        for i, page in enumerate(doc):
            text = page.get_text()
            tables = []
            for tab in page.find_tables():
                try:
                    tables.append(tab.extract())
                except Exception:
                    pass
            pages.append({'page': i + 1, 'text': text, 'tables': tables})
    finally:
        doc.close()
    return pages


# 제어/설정 레지스터를 STATUS/ALARM 정의 적용 대상에서 제외하기 위한 키워드
_CONTROL_REG_EXCL = [
    'reactive', 'power factor', 'active power', 'voltage limit',
    'current limit', 'frequency limit', 'setting', 'control',
    'threshold', 'setpoint', 'permanen', 'fixed', 'droop',
]


def _is_control_reg(reg) -> bool:
    """제어/설정 레지스터 여부 판단 — STATUS/ALARM 정의 적용 대상 제외용"""
    dl = reg.definition.lower().replace('_', ' ')
    return any(k in dl for k in _CONTROL_REG_EXCL)


def _apply_saved_definitions(categorized: dict, manufacturer: str, log=None):
    """definitions/{manufacturer}_definitions.json에서 정의 로드 (PDF 파싱으로 못 찾은 경우 fallback)"""
    defs_dir = os.path.join(os.path.dirname(__file__), 'definitions')
    fname = f'{manufacturer.lower()}_definitions.json'
    fpath = os.path.join(defs_dir, fname)
    if not os.path.exists(fpath):
        return

    with open(fpath, encoding='utf-8') as f:
        saved = json.load(f)

    status_defs = saved.get('status_definitions', {})
    alarm_codes = saved.get('alarm_codes', {})

    # STATUS: value_definitions 없는 inverter_status 레지스터에 적용 (제어 레지스터 제외)
    if status_defs:
        for reg in categorized.get('STATUS', []):
            if getattr(reg, 'h01_field', '') == 'inverter_status':
                if not getattr(reg, 'value_definitions', None) and not _is_control_reg(reg):
                    reg.value_definitions = status_defs
                    if log:
                        log(f'  정의 파일 적용 (status): {fname} ({len(status_defs)}개)')
                break

    # ALARM: value_definitions 없는 첫 번째 ALARM에 적용 (제어 레지스터 제외)
    if alarm_codes:
        for reg in categorized.get('ALARM', []):
            if not getattr(reg, 'value_definitions', None) and not _is_control_reg(reg):
                reg.value_definitions = alarm_codes
                if log:
                    log(f'  정의 파일 적용 (alarm): {fname} ({len(alarm_codes)}개)')
                break


def _extract_model_from_pdf(pdf_path: str, manufacturer: str) -> str:
    """PDF 메타데이터/파일명/첫 페이지에서 인버터 모델명 추출"""
    model = ''

    # 1) PDF 메타데이터 title
    try:
        import fitz
        doc = fitz.open(pdf_path)
        title = (doc.metadata or {}).get('title', '')
        # title에서 제조사명 제거 후 모델명 추출
        if title:
            # "EG4 18KPV-12LV Modbus Protocol" → "18KPV-12LV"
            # "TRIO-20.0(27.6)-TL-OUTD - Modbus RTU Registers Map" → "TRIO-20.0(27.6)-TL-OUTD"
            cleaned = title
            for rm in ['modbus', 'protocol', 'registers?', 'map', 'rtu', 'rs485',
                        'communication', 'interface', 'guide', 'application', 'note',
                        'version', 'customer', 'user', 'manual', 'definition']:
                cleaned = re.sub(rf'\b{rm}\b', '', cleaned, flags=re.I)
            cleaned = re.sub(rf'\b{re.escape(manufacturer)}\b', '', cleaned, flags=re.I)
            cleaned = re.sub(r'[-–—]+\s*$', '', cleaned)
            cleaned = re.sub(r'\s+', ' ', cleaned).strip(' -–—')
            if cleaned and len(cleaned) >= 3:
                model = cleaned
        doc.close()
    except Exception:
        pass

    # 2) 파일명에서 추출 (메타데이터 없을 때)
    if not model:
        fname = os.path.splitext(os.path.basename(pdf_path))[0]
        # "EG4_18KPV-12LV_Modbus_Protocol" → "18KPV-12LV"
        cleaned = fname
        for rm in ['Modbus', 'Protocol', 'Register', 'Map', 'RTU', 'RS485',
                    'Communication', 'Interface', 'Guide', 'Application', 'Note',
                    'Version', 'Customer', 'User', 'Manual', 'Comm']:
            cleaned = cleaned.replace(rm, '').replace(rm.lower(), '')
        cleaned = re.sub(rf'^{re.escape(manufacturer)}[_\s-]*', '', cleaned, flags=re.I)
        cleaned = re.sub(r'[_\s]+v?\d+$', '', cleaned)  # trailing version
        cleaned = re.sub(r'[_\s]+', ' ', cleaned).strip(' _-')
        if cleaned and len(cleaned) >= 3:
            model = cleaned

    return model


def extract_excel_sheets(excel_path: str) -> Dict[str, List[List[str]]]:
    """openpyxl로 Excel 시트별 행 추출 — 헤더 반복 시 섹션별 분리"""
    openpyxl = get_openpyxl()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    result = {}
    # 헤더 감지 키워드
    _header_keywords = {'fc', 'address', 'addr', 'parameter', 'name', 'definition',
                        'unit', 'r/w', 'type', '속성', '주소', '이름'}
    for name in wb.sheetnames:
        ws = wb[name]
        sections = []
        current_section = []
        for row in ws.iter_rows(values_only=True):
            if not any(c is not None for c in row):
                continue
            cells = [str(c) if c is not None else '' for c in row]
            # 헤더 행 감지 (3개 이상 키워드 매칭)
            cell_lower = [c.lower().strip() for c in cells if c.strip()]
            header_hits = sum(1 for c in cell_lower if any(k in c for k in _header_keywords))
            if header_hits >= 3 and current_section:
                # 이전 섹션 저장, 새 섹션 시작
                sections.append(current_section)
                current_section = []
            current_section.append(cells)
        if current_section:
            sections.append(current_section)
        # 각 섹션을 별도 테이블로
        for i, sec in enumerate(sections):
            key = f'{name}_{i}' if len(sections) > 1 else name
            result[key] = sec
    wb.close()
    return result


# ─── 레지스터 테이블 행 파싱 ──────────────────────────────────────────────────

_ADDR_RE = re.compile(r'(?:0x|0X)?([0-9A-Fa-f]{4})[Hh]?')
_TYPE_RE = re.compile(r'\b(U16|S16|U32|S32|US16|US32|I16|I32|INT16|UINT16|INT32|UINT32|FLOAT32|ASCII|STRING|STR|Bitfield16|Bitfield32)\b', re.I)
_RW_RE   = re.compile(r'\b(R/?W|RO|WO|Read|Write|R/W)\b', re.I)
_SCALE_RE = re.compile(r'(?:scale|factor|×|x)\s*[=:]?\s*([\d.]+)', re.I)
_UNIT_RE = re.compile(r'\b(V|A|W|kW|KW|VA|kVA|KVA|VAr|kVAr|KVar|Hz|°C|℃|Wh|kWh|KWh|Kwh|KWH|MWh|MWH|%)\b')


def _detect_table_columns(header_row: list, data_rows: list = None) -> dict:
    """테이블 헤더에서 컬럼 인덱스 추측"""
    col_map = {}
    for i, cell in enumerate(header_row):
        if not cell:
            continue
        cl = str(cell).lower().strip()
        if cl in ('address', 'addr', '주소', 'offset') or \
           ('addr' in cl and 'register' not in cl and 'reg.' not in cl):
            col_map.setdefault('addr', i)
        elif cl.startswith('reg.') or cl == 'reg.addr':
            # EKOS: reg.addr(30041)가 실제 Modbus 주소, ADDRESS는 오프셋
            # reg.addr을 우선 주소로 사용
            col_map['addr'] = i  # 덮어쓰기 (ADDRESS보다 우선)
        elif any(k in cl for k in ['name', 'definition', '이름', '항목', 'parameter', 'description',
                                    'signal name', 'signalname']):
            col_map.setdefault('name', i)
        elif cl == 'field' or cl == '필드':
            col_map.setdefault('name', i)
        elif any(k in cl for k in ['data type', 'datatype', '데이터', '타입']):
            col_map.setdefault('type', i)
        elif cl == 'type' or cl == 'format':
            col_map.setdefault('type', i)
        elif any(k in cl for k in ['unit', '단위']):
            col_map.setdefault('unit', i)
        elif any(k in cl for k in ['scale', '배율', 'factor', 'gain']):
            col_map.setdefault('scale', i)
        elif any(k in cl for k in ['r/w', 'access', '읽기', 'permission', '속성']):
            col_map.setdefault('rw', i)
        elif any(k in cl for k in ['remark', 'comment', '비고', 'note', '설명']):
            col_map.setdefault('comment', i)
        elif ('register' in cl and ('number' in cl or 'num' in cl or 'count' in cl)) or \
             cl in ('numberofregister', 'numberofreg', 'regs', 'reg count'):
            col_map.setdefault('regs', i)

    if data_rows and 'addr' not in col_map:
        # 1순위: 0x 패턴
        for row in data_rows[:5]:
            for i, cell in enumerate(row):
                if cell and re.match(r'0x[0-9A-Fa-f]{4}', str(cell).strip()):
                    col_map['addr'] = i
                    break
            if 'addr' in col_map:
                break

        # 2순위: 4~5자리 decimal (Sungrow 5000+ 등) — 행번호(1~3자리)와 구분
        if 'addr' not in col_map:
            col_scores = {}  # {col_idx: count_of_4digit_numbers}
            for row in data_rows[:10]:
                for i, cell in enumerate(row):
                    c = str(cell).strip() if cell else ''
                    if re.match(r'^\d{4,5}$', c) and 0 <= int(c) <= 65535:
                        col_scores[i] = col_scores.get(i, 0) + 1
            if col_scores:
                best_col = max(col_scores, key=col_scores.get)
                if col_scores[best_col] >= 2:
                    col_map['addr'] = best_col

    return col_map


def _parse_register_row(row: list, col_map: dict) -> Optional[RegisterRow]:
    """테이블 행 → RegisterRow"""
    if not row:
        return None

    _RANGE_ADDR_RE = re.compile(r'^(\d{4,5})\s*[-–~]\s*(\d{4,5})$')
    _RANGE_HEX_RE = re.compile(r'^(0x[0-9A-Fa-f]{4})\s*[-–~]\s*(0x[0-9A-Fa-f]{4})$', re.I)

    addr = None
    addr_raw = ''

    def _try_parse_addr(raw: str):
        raw = raw.strip().rstrip('.')
        if not raw:
            return None
        if raw.startswith('0x') or raw.startswith('0X'):
            return parse_address(raw)
        m = _RANGE_HEX_RE.match(raw)
        if m:
            return parse_address(m.group(1))
        if re.match(r'^\d{1,5}$', raw) and 0 <= int(raw) <= 65535:
            return int(raw)
        m = _RANGE_ADDR_RE.match(raw)
        if m:
            return int(m.group(1))
        return None

    addr_idx = col_map.get('addr')
    if addr_idx is not None and addr_idx < len(row):
        addr_raw = str(row[addr_idx]).strip() if row[addr_idx] else ''
        addr = _try_parse_addr(addr_raw)

    if addr is None:
        # V2: 4~5자리 숫자 우선 (행번호 1~3자리와 구분)
        candidates = []
        for i, cell in enumerate(row):
            c = str(cell).strip() if cell else ''
            parsed = _try_parse_addr(c)
            if parsed is not None:
                candidates.append((i, c, parsed))
        if candidates:
            # 4~5자리 decimal 우선, 없으면 0x 패턴, 최후에 아무거나
            best = None
            for i, c, parsed in candidates:
                if re.match(r'^\d{4,5}$', c):
                    best = (i, c, parsed)
                    break
            if not best:
                for i, c, parsed in candidates:
                    if c.startswith('0x') or c.startswith('0X'):
                        best = (i, c, parsed)
                        break
            if not best:
                best = candidates[0]
            addr = best[2]
            addr_raw = best[1]
            if 'addr' not in col_map:
                col_map['addr'] = best[0]

    if addr is None:
        return None

    # V2: 주소 컬럼 인덱스 (col_map 또는 fallback에서 설정된 것)
    actual_addr_idx = col_map.get('addr')

    name = ''
    name_idx = col_map.get('name')
    if name_idx is not None and name_idx < len(row):
        name = str(row[name_idx]).strip()
    if not name:
        name = ''
        for i, cell in enumerate(row):
            if i == actual_addr_idx:
                continue
            c = str(cell).strip()
            if not c or (len(c) <= 2 and c.upper() not in ('SN', 'FW', 'PN')):
                continue
            # V2: 숫자만 있는 셀은 행번호 — 이름이 아님
            if re.match(r'^\d{1,5}$', c):
                continue
            # 0x 주소 패턴도 건너뛰기
            if _ADDR_RE.fullmatch(c):
                continue
            name = c
            break
    if not name:
        return None

    # PDF에서 이름+설명이 붙어서 추출된 경우 분리
    # 패턴: "Inverter Model informationIdentifies the TL..." → 소문자 뒤 대문자+소문자2자 이상
    # 80자 초과 + 경계 위치가 20~60자 사이일 때만 분리 (짧은 이름의 PDF 줄바꿈과 구분)
    comment_extra = ''
    if len(name) > 80:
        m_split = re.search(r'([a-z])([A-Z][a-z]{2,})', name)
        if m_split and 15 < m_split.start() < 80:
            comment_extra = name[m_split.start() + 1:]
            name = name[:m_split.start() + 1]

    dtype = ''
    type_idx = col_map.get('type')
    if type_idx is not None and type_idx < len(row):
        m = _TYPE_RE.search(str(row[type_idx]))
        if m:
            dtype = m.group(1).upper()
    if not dtype:
        for cell in row:
            m = _TYPE_RE.search(str(cell))
            if m:
                dtype = m.group(1).upper()
                break
    dtype = dtype.upper()
    for old, new in [('INT16', 'S16'), ('UINT16', 'U16'), ('INT32', 'S32'), ('UINT32', 'U32'),
                     ('I16', 'S16'), ('I32', 'S32'), ('US16', 'U16'), ('US32', 'U32'),
                     ('STR', 'STRING'), ('BITFIELD16', 'U16'), ('BITFIELD32', 'U32')]:
        dtype = dtype.replace(old, new)

    # 단위 + 스케일 추출 — "0.1V", "0.01A", "kWh" 등에서 분리
    _SCALE_UNIT_RE = re.compile(r'^([\d.]+)\s*(V|A|W|kW|KW|VA|kVA|KVA|VAr|kVAr|KVar|Hz|°C|℃|Wh|kWh|KWh|Kwh|KWH|MWh|MWH|%)$')

    unit = ''
    scale = ''

    # 1) scale 컬럼에서 먼저 시도
    scale_idx = col_map.get('scale')
    if scale_idx is not None and scale_idx < len(row):
        s = str(row[scale_idx]).strip()
        if s and s not in ('', 'None', '-'):
            # "0.1V" 형태면 scale+unit 분리
            m = _SCALE_UNIT_RE.match(s)
            if m:
                scale = m.group(1)
                unit = m.group(2)
            else:
                # 숫자만이면 scale
                try:
                    float(s)
                    scale = s
                except ValueError:
                    pass

    # 2) unit 컬럼에서 시도
    unit_idx = col_map.get('unit')
    if unit_idx is not None and unit_idx < len(row):
        u = str(row[unit_idx]).strip()
        if u and u not in ('', 'None', '-'):
            # "0.1V", "0.01A" 형태면 분리
            m = _SCALE_UNIT_RE.match(u)
            if m:
                if not scale:
                    scale = m.group(1)
                if not unit:
                    unit = m.group(2)
            else:
                # 순수 단위
                m2 = _UNIT_RE.search(u)
                if m2 and not unit:
                    unit = m2.group(1)

    # 3) 이름에서 단위 추출 (fallback)
    if not unit:
        m = _UNIT_RE.search(name)
        if m:
            unit = m.group(1)

    # 4) 전체 행에서 "0.1V", "0.01A" 패턴만 탐색 (addr/name 컬럼 제외)
    if not scale:
        skip_cols = {col_map.get('addr'), col_map.get('name'), actual_addr_idx}
        for ci, cell in enumerate(row):
            if ci in skip_cols:
                continue
            c = str(cell).strip()
            m = _SCALE_UNIT_RE.match(c)
            if m and float(m.group(1)) != 1.0:
                scale = m.group(1)
                if not unit:
                    unit = m.group(2)
                break

    rw = ''
    rw_idx = col_map.get('rw')
    if rw_idx is not None and rw_idx < len(row):
        m = _RW_RE.search(str(row[rw_idx]))
        if m:
            rw = m.group(1).upper().replace('READ', 'RO').replace('WRITE', 'WO')
    if not rw:
        for cell in row:
            m = _RW_RE.search(str(cell))
            if m:
                rw = m.group(1).upper().replace('READ', 'RO').replace('WRITE', 'WO')
                break

    comment = ''
    comment_idx = col_map.get('comment')
    if comment_idx is not None and comment_idx < len(row):
        comment = str(row[comment_idx]).strip()
        if comment in ('None', '-'):
            comment = ''

    regs_val = '1'
    regs_idx = col_map.get('regs')
    if regs_idx is not None and regs_idx < len(row):
        r = str(row[regs_idx]).strip()
        if r.isdigit():
            regs_val = r
    if dtype in ('U32', 'S32', 'FLOAT32'):
        regs_val = '2'
    if dtype == 'ASCII':
        regs_val = str(max(1, int(regs_val)))

    return RegisterRow(
        definition=name,
        address=addr,
        address_hex=f'0x{addr:04X}',
        data_type=dtype or 'U16',
        regs=regs_val,
        unit=unit,
        scale=scale,
        rw=rw or 'RO',
        comment=(comment + ' ' + comment_extra).strip() if comment_extra else comment,
    )


def _clean_cell(cell) -> str:
    if cell is None:
        return ''
    return re.sub(r'\s*\n\s*', '', str(cell)).strip()


def _clean_table(table: List[list]) -> List[list]:
    return [[_clean_cell(c) for c in row] for row in table]


# V2: 유효하지 않은 레지스터 이름 필터
_JUNK_NAME_RE = re.compile(r'^\d{1,5}$')  # 숫자만 (105, 220 등)
_MODEL_NAME_RE = re.compile(
    r'^SG\d+|^SH\d+|^SC\d+|^KSG|^SUN\d+|^HU?N?\d+|^GW\d+|'  # Sungrow/Huawei/Goodwe 모델명
    r'^[A-Z]{2,5}\d{2,}[A-Z]*[-_]',  # 일반 모델명 패턴 (SG60KTL-M 등)
    re.I
)
_COUNTRY_NAME_RE = re.compile(
    r'^(?:Mexico|Brazil|Germany|France|Italy|Spain|Australia|India|China|Japan|Korea|'
    r'Thailand|Vietnam|Philippines|South Africa|UK|USA|Canada|Turkey|Holland|'
    r'Belgium|Austria|Denmark|Sweden|Norway|Finland|Poland|Czech|Hungary|Romania|'
    r'Portugal|Greece|Israel|Chile|Colombia|Peru|Argentina|Egypt|Jordan|Saudi|'
    r'UAE|Kuwait|Taiwan|Malaysia|Indonesia|Singapore|New Zealand|'
    r'Oman|Ireland|America|Vorarlberg|AU-WEST|Other\s+\d+Hz|GR\s+IS)',
    re.I
)


def _is_valid_register_name(name: str) -> bool:
    """V2: 유효한 레지스터 이름인지 판단"""
    stripped = name.strip()
    if not stripped or len(stripped) < 2:
        return False
    # 숫자만 (PDF 값이 이름으로 추출됨)
    if _JUNK_NAME_RE.match(stripped):
        return False
    # 인버터 모델명 테이블 (SG60KTL, SG50KTL-M 등)
    if _MODEL_NAME_RE.match(stripped):
        return False
    # 국가명/국가코드 테이블
    if _COUNTRY_NAME_RE.match(stripped):
        return False
    # 숫자 + 짧은 단위만 (예: "220V", "50Hz" — 설정값)
    if re.match(r'^\d+\.?\d*\s*[A-Za-z%°]{0,3}$', stripped):
        return False
    # V2: 무의미한 이름 (Reserved, U16 등 — 데이터 타입이 이름으로 추출된 경우)
    stripped_lower = stripped.lower()
    if stripped_lower in ('reserved', 'u16', 'u32', 's16', 's32', 'n/a', 'none', '-', '--'):
        return False
    # V2: Q(P)/Q(U) 커브 파라미터 (QP P1_, QU V1_, Q U1_, Curve 등)
    if re.match(r'^Q[PU ]\s*[A-Z]\d', stripped, re.I):
        return False
    if re.match(r'^LP\s+P\d', stripped, re.I):  # LP P34KSG_ 등
        return False
    # V2: 상태값 테이블 엔트리 (Initial standby, Starting, Stop, Derating run 등)
    if stripped_lower in ('initial standby', 'standby', 'starting', 'stop',
                           'derating run', 'dispatch run', 'key stop',
                           'curve', 'device abnormal'):
        return False
    # V2: 너무 짧은 이름 (3자 이하인데 키워드가 아닌 것)
    if len(stripped) <= 3 and not any(k in stripped_lower for k in
            ['sn', 'pf', 'pv', 'dc', 'ac', 'bus', 'ia', 'ib', 'ic', 'ua', 'ub', 'uc']):
        return False
    return True


def extract_registers_from_tables(tables: List[List[list]],
                                   fc_list: List[str] = None) -> List[RegisterRow]:
    """모든 테이블에서 레지스터 행 추출
    fc_list: 각 테이블의 FC 값 (예: ['04','04','03','03','',...]). None이면 전부 ''
    """
    registers = []
    seen_addr_fc = set()  # (addr, fc) 튜플로 중복 체크 → FC 다르면 별도 레지스터
    prev_col_map = {}  # 이전 테이블에서 감지한 col_map 유지
    for t_idx, table in enumerate(tables):
        if not table or len(table) < 1:
            continue
        fc = fc_list[t_idx] if fc_list and t_idx < len(fc_list) else ''
        table = _clean_table(table)
        first_has_addr = any(
            str(c).strip().startswith('0x') or str(c).strip().startswith('0X') or
            (re.match(r'^\d{3,5}$', str(c).strip()) and 0 <= int(str(c).strip()) <= 65535)
            for c in table[0] if c)
        if first_has_addr:
            data_rows = table
            col_map = _detect_table_columns([], data_rows)
            # 헤더 없는 테이블 — 이전 col_map이 더 풍부하면 사용
            if prev_col_map and len(col_map) < len(prev_col_map):
                col_map = dict(prev_col_map)
        else:
            data_rows = table[1:]
            col_map = _detect_table_columns(table[0], data_rows)
            if col_map:
                prev_col_map = dict(col_map)
        for row in data_rows:
            reg = _parse_register_row(row, col_map)
            if reg and (reg.address, fc) not in seen_addr_fc:
                # V2: 유효하지 않은 이름 필터 (모델명 테이블, 숫자값 등)
                if not _is_valid_register_name(reg.definition):
                    continue
                reg.fc = fc
                seen_addr_fc.add((reg.address, fc))
                registers.append(reg)
    return registers


# ─── H01 필드 매핑 ──────────────────────────────────────────────────────────

# ─── 정의 테이블 탐색 (Definition-Based Matching) ────────────────────────

_DEF_TABLE_HEADER_RE = re.compile(
    r'(?:Error\s*Code\s*Table\s*(\d)?|Alarm\s*(\d)|Fault\s*Code|'
    r'Inverter\s*Mode\s*Table|Work\s*(?:Mode|State)\s*Table|'
    r'Running\s*Status|Operating\s*Mode|Device\s*Status|'
    r'Table\s*3\.1\.\d|DSP\s*(?:alarm|error)|ARM\s*alarm)\s*'
    r'(?:\(?(?:0x([0-9A-Fa-f]{4}))\)?)?',
    re.I
)
_BIT_DEF_RE = re.compile(r'^(?:Bit\s*)?(\d{1,2})$')
_HEX_VAL_RE = re.compile(r'^0x([0-9A-Fa-f]{2,4})$')
_MODE_VAL_RE = re.compile(r'^(\d{1,2})\s*[:：]\s*(.+)')


def scan_definition_tables(pages: list, excel_sheets: dict = None,
                           registers: list = None) -> dict:
    """
    PDF/Excel에서 STATUS/ALARM 정의 테이블 탐색
    3가지 경로:
    1. Appendix/본문 페이지에서 정의 테이블 탐색
    2. Excel 시트(EKOS Fault MAP, 데이타형식)에서 탐색
    3. 레지스터 comment에서 인라인 정의 추출
    """
    result = {'status_defs': [], 'alarm_defs': []}

    if pages:
        _scan_pdf_definitions(pages, result)
    if excel_sheets:
        _scan_excel_definitions(excel_sheets, result)
    if registers:
        _scan_register_comments(registers, result)

    return result


def _scan_register_comments(registers: list, result: dict):
    """레지스터 이름/comment에서 인라인 정의 추출"""
    for reg in registers:
        comment = (getattr(reg, 'comment', '') or '').strip()
        defn = (reg.definition or '').strip()
        addr = reg.address if isinstance(reg.address, int) else 0
        category = getattr(reg, 'category', '')

        if category not in ('STATUS', 'ALARM') and not any(
                k in defn.lower() for k in ['mode', 'status', 'state', 'error', 'alarm', 'fault']):
            continue

        combined = f'{defn} {comment}'.lower()

        # 1) comment에 인라인 값 정의: "0:cWaitMode1:cNormalMode2:cFaultMode"
        inline_vals = {}
        for m in re.finditer(r'(\d+)\s*[:：]\s*([a-zA-Z가-힣]\w{2,})', comment):
            inline_vals[int(m.group(1))] = m.group(2)
        if len(inline_vals) >= 2:
            def_type = _detect_definition_type(inline_vals)
            if def_type:
                target = 'status_defs' if def_type == 'mode_table' else 'alarm_defs'
                result[target].append({
                    'address': addr, 'name': f'{defn} (comment)',
                    'type': def_type, 'values': inline_vals, 'page': -1,
                })
                continue

        # 1b) comment에 Enum 인라인: "• 0x0002 Reconnecting • 0x0003 Online • 0x0014 Standby"
        if 'enum' in combined or '0x00' in comment:
            enum_vals = {}
            for m in re.finditer(r'0x([0-9A-Fa-f]{2,4})\s+([A-Za-z][\w\s]*?)(?:\n|\(|•|$)', comment):
                val = int(m.group(1), 16)
                desc = m.group(2).strip()[:30]
                if desc:
                    enum_vals[val] = desc
            if len(enum_vals) >= 2:
                target = 'status_defs' if any(k in combined for k in ['mode', 'status', 'state']) else 'alarm_defs'
                result[target].append({
                    'address': addr, 'name': f'{defn} (enum)',
                    'type': 'mode_table', 'values': enum_vals, 'page': -1,
                })
                continue

        # 2) comment에 테이블 참조: "Table 3.1.4", "See Appendix 3", "Error Code Table 1"
        has_table_ref = bool(re.search(r'Table\s+[\d.]+|Appendix\s+\d|Code\s+Table', comment, re.I))
        if has_table_ref:
            # 정의는 Appendix에서 이미 탐색했으므로, 레지스터에 참조 마크만
            # 나중에 run_stage1에서 연결
            pass

        # 3) 레지스터 이름에 "Mode Table", "Error Code" 참조
        if any(k in combined for k in ['mode table', 'error code table', 'fault code table',
                                        'alarm code table']):
            # 이미 Appendix 탐색에서 찾았을 가능성 — 주소 연결용
            pass


def _scan_pdf_definitions(pages: list, result: dict):
    """PDF 페이지에서 STATUS/ALARM 정의 테이블 탐색 (정밀 버전)"""
    for p in pages:
        text = p.get('text', '')
        page_num = p.get('page', 0)
        lines = text.split('\n')

        # ── 1) 헤더 기반: "Error Code Table1 (0x101E)" 등 ──
        for line in lines:
            line_s = line.strip()
            m = _DEF_TABLE_HEADER_RE.search(line_s)
            if not m:
                continue
            addr_hex = m.group(3)
            addr = int(addr_hex, 16) if addr_hex else None
            header_lower = line_s.lower()
            is_mode = any(k in header_lower for k in ['mode', 'status', 'state', 'running', 'operating'])
            is_alarm = any(k in header_lower for k in ['error code', 'alarm', 'fault', 'dsp', 'arm'])

            for tab in p.get('tables', []):
                cleaned = _clean_table(tab)
                values = _extract_value_definitions(cleaned)
                if values and len(values) >= 2:
                    def_type = _detect_definition_type(values)
                    if def_type:
                        target = 'status_defs' if (is_mode or def_type == 'mode_table') else 'alarm_defs'
                        result[target].append({
                            'address': addr, 'name': line_s[:60],
                            'type': def_type, 'values': values, 'page': page_num,
                        })
                        break

        # ── 2) 테이블 기반: hex 모드값 테이블 (Solarize Table: 0x00=Initial) ──
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            hex_modes = {}
            for row in cleaned:
                cells = [str(c).strip() for c in row]
                if not cells:
                    continue
                # "0x00 | (empty) | Initial mode" 또는 "0x00 | Initial mode"
                m_hex = _HEX_VAL_RE.match(cells[0])
                if m_hex:
                    val = int(m_hex.group(1), 16)
                    # 설명은 비어있지 않은 셀에서
                    desc = ''
                    for c in cells[1:]:
                        if c and len(c) > 2 and not c.isdigit() and not c.startswith('0x'):
                            desc = c[:50]
                            break
                    dl = desc.lower()
                    if any(k in dl for k in ['mode', 'standby', 'on-grid', 'fault', 'shutdown',
                                              'initial', 'off-grid', 'waiting', 'normal',
                                              'charge', 'battery', 'running', 'check',
                                              'derating', 'sleep', 'program']):
                        hex_modes[val] = desc
            if len(hex_modes) >= 3:
                if not any(d['page'] == page_num and d['type'] == 'mode_table'
                           and len(d['values']) >= 3 for d in result['status_defs']):
                    result['status_defs'].append({
                        'address': None, 'name': f'Mode Table (P{page_num})',
                        'type': 'mode_table', 'values': hex_modes, 'page': page_num,
                    })

        # ── 2b) 텍스트 기반: hex 상태값 (Huawei: "0x0000 Standby: initializing") ──
        hex_status = {}
        for line in lines:
            m = re.match(r'\s*(0x[0-9A-Fa-f]{2,4})\s+(.+)', line.strip())
            if m:
                val = int(m.group(1), 16)
                desc = m.group(2).strip()[:50]
                dl = desc.lower()
                if any(k in dl for k in ['standby', 'fault', 'shutdown', 'running',
                                          'on-grid', 'off-grid', 'detecting', 'spot',
                                          'charge', 'battery', 'check', 'derating',
                                          'sleep', 'waiting', 'initial', 'program']):
                    hex_status[val] = desc
        if len(hex_status) >= 3:
            if not any(d['page'] == page_num and 'Hex' in d.get('name', '')
                       for d in result['status_defs']):
                result['status_defs'].append({
                    'address': None, 'name': f'Hex Status (P{page_num})',
                    'type': 'mode_table', 'values': hex_status, 'page': page_num,
                })

        # ── 3) 텍스트 기반: "N: description" 패턴 ──
        # Huawei P158: "0: offline", "1: standby", "3: faulty", "4: running"
        # Goodwe: "0:cWaitMode", "1:cNormalMode", "2:cFaultMode"
        num_status = {}
        for line in lines:
            # "N: description" 또는 "NN---description" (ESINV)
            m = re.match(r'\s*(\d{1,2})\s*(?:[:：]|---)\s*(.+)', line.strip())
            if m:
                val = int(m.group(1))
                desc = m.group(2).strip()[:50]
                dl = desc.lower()
                if any(k in dl for k in ['offline', 'standby', 'faulty', 'running', 'power',
                                          'mode', 'wait', 'normal', 'fault', 'stop',
                                          'initializ', 'check', 'mpp', 'search', 'close',
                                          'no response', 'volt', 'reactive', 'fixed',
                                          'disable', 'enable']):
                    num_status[val] = desc
        if len(num_status) >= 3:
            if not any(d['page'] == page_num for d in result['status_defs']):
                result['status_defs'].append({
                    'address': None, 'name': f'Num Status (P{page_num})',
                    'type': 'mode_table', 'values': num_status, 'page': page_num,
                })

        # ── 3b) 테이블 기반: Sungrow "State | Value(0x hex) | Paraphrase" ──
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            if len(cleaned) < 4:
                continue
            # 헤더에 "State" + "Value" 패턴
            header_joined = ' '.join(str(c).lower() for c in cleaned[0] + (cleaned[1] if len(cleaned) > 1 else []))
            if 'state' in header_joined and 'value' in header_joined:
                sg_modes = {}
                for row in cleaned[2:]:
                    cells = [str(c).strip() for c in row]
                    if len(cells) >= 2 and cells[0] and cells[1]:
                        state_name = cells[0]
                        value_str = cells[1]
                        # "0x0" "0x8000" "0x1300" 패턴
                        m_v = re.match(r'^0x([0-9A-Fa-f]+)$', value_str)
                        if m_v:
                            val = int(m_v.group(1), 16)
                            sg_modes[val] = state_name
                if len(sg_modes) >= 4:
                    result['status_defs'].append({
                        'address': None, 'name': f'Work State Table (P{page_num})',
                        'type': 'mode_table', 'values': sg_modes, 'page': page_num,
                    })

        # ── 3c) 테이블 기반: Kstar "SN | Content(00H) | Description" ──
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            if len(cleaned) < 4:
                continue
            header_joined = ' '.join(str(c).lower() for c in cleaned[0])
            if ('sn' in header_joined or 'content' in header_joined) and 'description' in header_joined:
                ks_modes = {}
                for row in cleaned[1:]:
                    cells = [str(c).strip() for c in row]
                    if len(cells) >= 3 and cells[0].isdigit():
                        val = int(cells[0])
                        desc = cells[2] if cells[2] else cells[1]
                        dl = desc.lower()
                        if any(k in dl for k in ['initialization', 'waiting', 'normal', 'error',
                                                  'permanent', 'aging', 'burning', 'detection',
                                                  'checking', 'standby']):
                            ks_modes[val] = desc
                if len(ks_modes) >= 3:
                    result['status_defs'].append({
                        'address': None, 'name': f'Operating Mode (P{page_num})',
                        'type': 'mode_table', 'values': ks_modes, 'page': page_num,
                    })

        # ── 4) 테이블 기반: Sungrow Appendix fault codes ──
        # "002 | 0x0002 | Grid overvoltage | Fault"
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            fault_codes = {}
            for row in cleaned:
                cells = [str(c).strip() for c in row]
                if len(cells) >= 3:
                    # 숫자(코드) + 0x hex + 설명 + 분류(Fault/Alarm/Warning)
                    c0, c1 = cells[0], cells[1]
                    if c0.isdigit() and int(c0) > 0 and c1.startswith('0x'):
                        desc = cells[2][:50] if len(cells) > 2 else ''
                        classification = cells[3].lower() if len(cells) > 3 else ''
                        if any(k in classification for k in ['fault', 'alarm', 'warning', 'protect']):
                            fault_codes[int(c0)] = f'{desc} ({cells[3]})' if len(cells) > 3 else desc
            if len(fault_codes) >= 5:
                if not any(d['page'] == page_num and d['type'] == 'fault_codes'
                           for d in result['alarm_defs']):
                    result['alarm_defs'].append({
                        'address': None, 'name': f'Fault Codes (P{page_num})',
                        'type': 'fault_codes', 'values': fault_codes, 'page': page_num,
                    })

        # ── 4b) 테이블 기반: ABB "숫자 | 설명" 상태/알람 테이블 ──
        # "1 | Checking Grid", "44 | Waiting Start", "45 | MPPT"
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            num_vals = {}
            for row in cleaned:
                cells = [str(c).strip() for c in row]
                if len(cells) >= 2 and cells[0].isdigit() and int(cells[0]) >= 0:
                    val = int(cells[0])
                    desc = ''
                    for c in cells[1:]:
                        if c and len(c) > 2 and not c.isdigit() and not c.startswith('0x'):
                            desc = c[:50]
                            break
                    if desc:
                        num_vals[val] = desc
            if len(num_vals) >= 5:
                # 상태 vs 알람 판별: 상태 키워드 비율
                status_kw = sum(1 for d in num_vals.values()
                               if any(k in d.lower() for k in ['waiting', 'mppt', 'start',
                                                                 'grid', 'off', 'check']))
                is_status = status_kw > len(num_vals) * 0.3
                target = 'status_defs' if is_status else 'alarm_defs'
                def_type = 'mode_table' if is_status else 'fault_codes'
                if not any(d['page'] == page_num and d['type'] == def_type
                           for d in result[target]):
                    result[target].append({
                        'address': None, 'name': f'Num Table (P{page_num})',
                        'type': def_type, 'values': num_vals, 'page': page_num,
                    })

        # ── 4b-2) 테이블 기반: 헤더 "Inverter state" / "Operating mode" 가 있는 Value|Description 테이블 ──
        # ABB: "Inverter state" 헤더 → 0=Stand By, 1=Checking Grid, 2=Run, 45=MPPT...
        # Conext: "Operational Mode State" 인라인 Enum
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            if len(cleaned) < 3:
                continue
            # 첫 행 또는 두 번째 행이 헤더인지 확인
            for hi in range(min(2, len(cleaned))):
                header = ' '.join(str(c).lower() for c in cleaned[hi])
                if any(k in header for k in ['inverter state', 'inverter mode',
                                              'operating mode', 'work state',
                                              'device status']):
                    # 이 테이블의 데이터 행에서 Value→Description 추출
                    state_vals = {}
                    for row in cleaned[hi+1:]:
                        cells = [str(c).strip() for c in row]
                        # 첫 번째 숫자 셀 = Value, 첫 번째 비어있지 않은 텍스트 셀 = Description
                        val_str = ''
                        desc = ''
                        for c in cells:
                            if not val_str and c.isdigit():
                                val_str = c
                            elif val_str and c and len(c) > 1 and not c.isdigit():
                                desc = c[:50]
                                break
                        if val_str and desc:
                            state_vals[int(val_str)] = desc
                    if len(state_vals) >= 3:
                        result['status_defs'].append({
                            'address': None,
                            'name': f'Inverter State (P{page_num})',
                            'type': 'mode_table',
                            'values': state_vals,
                            'page': page_num,
                        })
                    break

        # ── 4b-3) 테이블 셀 기반: Enum 인라인 "• 0x0002 Reconnecting • 0x0003 Online" ──
        # Conext: Data Range 셀에 Enum 값 나열
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            for row in cleaned:
                cells = [str(c).strip() for c in row]
                row_lower = ' '.join(cells).lower()
                if 'operational mode' not in row_lower and 'inverter state' not in row_lower:
                    continue
                # Enum이 있는 셀 찾기
                for cell_text in cells:
                    if '0x00' in cell_text and ('enum' in cell_text.lower() or len(cell_text) > 30):
                        enum_vals = {}
                        for m_e in re.finditer(r'0x([0-9A-Fa-f]{2,4})\s+([A-Za-z][\w]*)', cell_text):
                            val = int(m_e.group(1), 16)
                            desc = m_e.group(2).strip()[:30]
                            if desc:
                                enum_vals[val] = desc
                        if len(enum_vals) >= 2:
                            result['status_defs'].append({
                                'address': None,
                                'name': f'Operational Mode Enum (P{page_num})',
                                'type': 'mode_table',
                                'values': enum_vals,
                                'page': page_num,
                            })

        # ── 4c) 테이블 기반: Growatt 32bit hex fault code ──
        # "0x00000002 | Communication error | 0x0002 | String communication abnormal"
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            hex32_faults = {}
            for row in cleaned:
                cells = [str(c).strip() for c in row]
                if not cells:
                    continue
                m32 = re.match(r'^0x([0-9A-Fa-f]{4,8})$', cells[0])
                if m32:
                    val = int(m32.group(1), 16)
                    desc = ''
                    for c in cells[1:]:
                        if c and len(c) > 2 and not c.startswith('0x') and not c.startswith('\\'):
                            desc = c[:50]
                            break
                    if desc:
                        hex32_faults[val] = desc
            if len(hex32_faults) >= 3:
                if not any(d['page'] == page_num and 'Hex32' in d.get('name', '')
                           for d in result['alarm_defs']):
                    result['alarm_defs'].append({
                        'address': None, 'name': f'Hex32 Fault (P{page_num})',
                        'type': 'fault_codes', 'values': hex32_faults, 'page': page_num,
                    })

        # ── 4d) 테이블 기반: EG4 Bit0~31 Fault/Warning 테이블 ──
        # "0 | Internal comm fault 1 | E000 | Battery comm failure | W000"
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            eg4_faults = {}
            for row in cleaned:
                cells = [str(c).strip() for c in row]
                if len(cells) >= 3 and cells[0].isdigit():
                    bit_num = int(cells[0])
                    # Fault description (col 1) + FaultCode (col 2)
                    desc = cells[1][:40] if cells[1] and cells[1] != 'rsvd' else ''
                    code = cells[2][:10] if len(cells) > 2 else ''
                    if desc and 0 <= bit_num <= 31 and (code.startswith('E') or code.startswith('W')):
                        eg4_faults[bit_num] = f'{desc} ({code})'
            if len(eg4_faults) >= 3:
                if not any(d['page'] == page_num and 'EG4' in d.get('name', '')
                           for d in result['alarm_defs']):
                    result['alarm_defs'].append({
                        'address': None, 'name': f'EG4 Fault/Warn (P{page_num})',
                        'type': 'bitfield', 'values': eg4_faults, 'page': page_num,
                    })

        # ── 5) 테이블 기반: Kstar/Huawei bitfield ──
        # "0 | Bit0 | F00 | Grid Volt Low" — Bit{N} 키워드가 있는 테이블만
        for tab in p.get('tables', []):
            cleaned = _clean_table(tab)
            # 테이블에 "Bit" 키워드가 있는지 먼저 확인
            has_bit_keyword = any('Bit' in str(c) or 'bit' in str(c) or 'BIT' in str(c)
                                  for row in cleaned[:5] for c in row)
            if not has_bit_keyword:
                continue
            bits = {}
            for row in cleaned:
                cells = [str(c).strip() for c in row]
                if len(cells) >= 2:
                    # "Bit0", "Bit15", "BIT00", "BIT01", "Bi12" 등
                    for ci, c in enumerate(cells[:3]):
                        m_bit = re.match(r'^(?:Bit|BIT|Bi)\s*(\d{1,2})$', c)
                        if m_bit:
                            bit_num = int(m_bit.group(1))
                            desc = ''
                            for dc in reversed(cells[ci+1:]):
                                if dc and len(dc) > 2 and not dc.isdigit() and not dc.startswith('Bit'):
                                    desc = dc[:50]
                                    break
                            if desc and 0 <= bit_num <= 31:
                                bits[bit_num] = desc
                            break
            if len(bits) >= 3:
                if not any(d['page'] == page_num and d['type'] == 'bitfield'
                           for d in result['alarm_defs']):
                    result['alarm_defs'].append({
                        'address': None, 'name': f'Bitfield (P{page_num})',
                        'type': 'bitfield', 'values': bits, 'page': page_num,
                    })

        # ── 6) Huawei 텍스트 기반 bitfield ──
        # "Bit00: input overvoltage", "Bit01: input undervoltage"
        bit_defs = {}
        for line in lines:
            m = re.match(r'\s*(?:Bit|BIT|Bi)\s*(\d{1,2})\s*[:：]\s*(.+)', line.strip())
            if m:
                bit_num = int(m.group(1))
                desc = m.group(2).strip()[:50]
                if desc and 0 <= bit_num <= 31:
                    bit_defs[bit_num] = desc
        if len(bit_defs) >= 3:
            if not any(d['page'] == page_num and d['type'] == 'bitfield'
                       for d in result['alarm_defs']):
                result['alarm_defs'].append({
                    'address': None, 'name': f'Bit Defs (P{page_num})',
                    'type': 'bitfield', 'values': bit_defs, 'page': page_num,
                })


def _scan_excel_definitions(sheets: dict, result: dict):
    """Excel 시트에서 정의 테이블 탐색 (EKOS 등)"""
    for key, rows in sheets.items():
        key_lower = key.lower()

        # EKOS: 데이타형식 F007 → status 정의
        if '데이타' in key_lower or 'format' in key_lower:
            in_f007 = False
            values = {}
            for row in rows:
                joined = ' '.join(str(c) for c in row).lower()
                if 'f007' in joined or '인터터 동작상태' in joined or '인버터 동작상태' in joined:
                    in_f007 = True
                    continue
                if in_f007:
                    if 'f008' in joined or 'f009' in joined:
                        break
                    # "0 : Stop", "8 : MPP" 패턴
                    for cell in row:
                        c = str(cell).strip()
                        m = re.match(r'^(\d+)\s*[:：]\s*(.+)', c)
                        if m:
                            values[int(m.group(1))] = m.group(2).strip()
            if len(values) >= 2:
                result['status_defs'].append({
                    'address': None, 'name': 'F007 (EKOS)',
                    'type': 'mode_table', 'values': values, 'page': 0,
                })

        # EKOS: Fault MAP → alarm 정의
        if 'fault' in key_lower:
            current_group = ''
            values = {}
            for row in rows:
                cells = [str(c).strip() if c else '' for c in row]
                # 그룹 헤더 (PV, 인버터, 계통, 컨버터)
                if cells[0] and any(k in cells[0] for k in ['PV', '인버터', '계통', '컨버터']):
                    if values and current_group:
                        result['alarm_defs'].append({
                            'address': None, 'name': f'{current_group} (EKOS)',
                            'type': 'bitfield', 'values': dict(values), 'page': 0,
                        })
                    current_group = cells[0].replace('\n', ' ').strip()[:30]
                    values = {}
                    continue
                # 비트 정의: bit번호 + 설명
                if len(cells) >= 3 and cells[1].isdigit() and cells[2]:
                    bit_num = int(cells[1])
                    desc = cells[2]
                    if desc:
                        values[bit_num] = desc
            if values and current_group:
                result['alarm_defs'].append({
                    'address': None, 'name': f'{current_group} (EKOS)',
                    'type': 'bitfield', 'values': dict(values), 'page': 0,
                })


def _extract_value_definitions(table: list) -> dict:
    """테이블에서 값 → 설명 매핑 추출 (레지스터 데이터 테이블 제외)"""
    values = {}
    for row in table:
        cells = [str(c).strip() for c in row if str(c).strip()]
        if not cells or len(cells) < 2:
            continue
        # 레지스터 데이터 테이블 행 제외 (U16, U32, S16, RO, RW 등 포함)
        joined = ' '.join(cells)
        if re.search(r'\b(U16|U32|S16|S32|FLOAT32|RO|RW|04H|06H)\b', joined):
            continue
        # Bit N 패턴
        m_bit = _BIT_DEF_RE.match(cells[0])
        if m_bit:
            bit_num = int(m_bit.group(1))
            desc = cells[-1] if len(cells) > 1 else ''
            if desc and len(desc) > 2 and not desc.isdigit():
                values[bit_num] = desc[:50]
            continue
        # 0x hex 값 패턴
        m_hex = _HEX_VAL_RE.match(cells[0])
        if m_hex:
            val = int(m_hex.group(1), 16)
            desc = cells[1] if len(cells) > 1 else ''
            if desc and not desc.isdigit():
                values[val] = desc[:50]
            continue
        # 숫자 + 설명
        if cells[0].isdigit() and len(cells) >= 2:
            val = int(cells[0])
            desc = cells[1]
            if desc and len(desc) > 2 and not desc.isdigit():
                values[val] = desc[:50]
    return values


def _detect_definition_type(values: dict) -> Optional[str]:
    """값 패턴으로 정의 타입 추측"""
    if not values:
        return None
    keys = list(values.keys())
    vals_text = ' '.join(str(v).lower() for v in values.values())

    # mode_table: 값이 적고 (< 15), mode/standby/fault 키워드
    if len(keys) <= 15 and any(k in vals_text for k in ['mode', 'standby', 'fault', 'grid',
                                                          'initial', 'shutdown', 'waiting',
                                                          'normal', 'running', 'stop']):
        return 'mode_table'

    # bitfield: 키가 0~15 범위, bit 관련 키워드
    if all(0 <= k <= 31 for k in keys) and any(k in vals_text for k in [
            'over', 'under', 'abnormal', 'fault', 'lock', 'relay', 'igbt',
            'voltage', 'current', 'frequency', 'temperature']):
        return 'bitfield'

    # fault_codes: 키가 크고 (> 15), fault/alarm 분류
    if any(k > 15 for k in keys) and any(k in vals_text for k in ['fault', 'alarm', 'warning']):
        return 'fault_codes'

    return None


def assign_h01_field(reg: RegisterRow, synonym_db: dict,
                     ref_patterns: dict = None) -> str:
    """레지스터에 대응하는 H01 필드명 추정 (V2)"""
    defn_lower = reg.definition.lower().replace('_', ' ')
    category = getattr(reg, 'category', '')

    # V2: INFO/ALARM 카테고리는 H01 모니터링 필드가 아님 — 특정 키워드만 매칭
    comment_lower = (getattr(reg, 'comment', '') or '').lower()

    if category == 'INFO':
        # INFO에서 H01과 겹치는 필드만 매핑
        # cumulative_energy LOW (소수부/Wh): decimal, low byte — comment도 체크
        if any(k in defn_lower for k in ['decimals of total energy', 'decimal of total',
                                          'low byte of total feed', 'low byte oftotal feed']) or \
           any(k in comment_lower for k in ['decimals of total', 'decimal of total']):
            return 'cumulative_energy_low'
        # cumulative_energy HIGH (정수부)
        if any(k in defn_lower for k in ['total energy', 'cumulative energy', 'total power yields',
                                          'total poweryields',
                                          'total energy yield', 'energy yield',
                                          'lifetime energy', 'energy produced',
                                          'ac energy', 'ac_energy',
                                          '누적발전량', '누적 발전량',
                                          'total generation energy',
                                          'high byte of total feed', 'high byte oftotal feed']):
            return 'cumulative_energy'
        # 적산전력량 MWh/Wh 쌍 (EKOS)
        if '적산전력량' in defn_lower:
            unit = getattr(reg, 'unit', '').upper()
            if unit in ('WH', 'wh', 'Wh'):
                return 'cumulative_energy_low'
            return 'cumulative_energy'
        if any(k in defn_lower for k in ['daily energy', 'today energy', 'daily power yields',
                                          '일발전량', '일 발전량', '금일발전량']):
            return 'daily_energy'
        if any(k in defn_lower for k in ['inner_temp', 'inner temp', 'internal temp',
                                          'module temp', 'inverter temp']):
            return 'temperature'
        return ''  # 나머지 INFO는 h01_field 없음

    # V2: STATUS 카테고리 — inverter_status 매핑
    if category == 'STATUS':
        defn_nospace_s = defn_lower.replace(' ', '')
        if any(k in defn_lower for k in ['inverter mode', 'work mode', 'work state',
                                          'operating mode', 'operational mode',
                                          'operation state',
                                          'operating status', 'working mode',
                                          'running status', 'run mode',
                                          'inverter state', 'inverter status',
                                          'inverter current', 'i status',
                                          '인버터 모드', '시스템동작상태', '동작상태',
                                          'device status', 'system status']) or \
           any(k in defn_nospace_s for k in ['workmode', 'invworkmode', 'runningmode',
                                              'workingmodes', 'workingmode',
                                              'sysstatemode', 'currentstatus',
                                              'operatingstatus']) or \
           defn_lower.strip() in ('state', 'running'):
            return 'inverter_status'
        return ''

    # 0) 채널 번호가 있으면 최우선
    ch = detect_channel_number(reg.definition)
    if ch:
        prefix, n = ch
        if prefix == 'MPPT':
            # voltage: Vpv{N}, voltage, 전압
            if any(k in defn_lower for k in ['voltage', '전압']) or \
               re.match(r'^vpv\d', defn_lower):
                return f'mppt{n}_voltage'
            # current: PV{N}Curr, current, 전류
            if any(k in defn_lower for k in ['current', 'curr', '전류']):
                return f'mppt{n}_current'
            # power: Ppv{N}, PV{N}Watt, power, 전력
            if any(k in defn_lower for k in ['power', 'watt', '전력']) or \
               re.match(r'^ppv\d', defn_lower):
                return f'mppt{n}_power'
        elif prefix == 'STRING':
            if any(k in defn_lower for k in ['voltage', '전압']):
                return f'string{n}_voltage'
            if any(k in defn_lower for k in ['current', 'curr', '전류']):
                return f'string{n}_current'

    # 0-1) Central type: 번호 없는 DC/PV/Input voltage/current → mppt1
    defn_ns = defn_lower.replace('_', ' ')
    if re.search(r'\b(i dc|dc)\s*(voltage)', defn_ns) or \
       re.match(r'^(pv|input)\s+(voltage)', defn_ns):
        if 'fault' not in defn_lower and 'high' not in defn_lower and 'low' not in defn_lower:
            return 'mppt1_voltage'
    if re.search(r'\b(i dc|dc)\s*(current)', defn_ns) or \
       re.match(r'^(pv|input)\s+(current)', defn_ns):
        if 'fault' not in defn_lower:
            return 'mppt1_current'

    # 1) V2: pv_power / energy 키워드 (synonym/ref보다 먼저 — 정확한 키워드 우선)
    if any(k in defn_lower for k in ['total dc power', 'total pv power', 'dc power',
                                      'pv total power', 'pv_total_input_power',
                                      'input power', 'output power',
                                      'inverter current output',
                                      'totaldc input', 'totaldcinput',
                                      'i dc power', 'i_dc_power',
                                      '태양전지 전력', '태양전지전력']) or \
       re.search(r'\bpac\b', defn_lower):
        return 'pv_power'
    defn_nospace = defn_lower.replace(' ', '')
    # cumulative_energy LOW (소수부/Wh/Low Byte) — comment도 체크
    if any(k in defn_lower for k in ['decimals of total energy', 'decimal of total',
                                      'low byte of total feed', 'low byte oftotal feed']) or \
       any(k in comment_lower for k in ['decimals of total', 'decimal of total']):
        return 'cumulative_energy_low'
    # 적산전력량 Wh (EKOS LOW)
    if '적산전력량' in defn_lower:
        unit = getattr(reg, 'unit', '').upper()
        if unit in ('WH',):
            return 'cumulative_energy_low'
        return 'cumulative_energy'
    # cumulative_energy HIGH (정수부)
    if any(k in defn_lower for k in ['total energy', 'cumulative energy', 'total power yields',
                                      'total poweryields',
                                      'lifetime energy', 'accumulated energy',
                                      'total power generation', 'total powergeneration',
                                      'total energy yield', 'energy yield',
                                      'energy produced', 'ac energy', 'ac_energy',
                                      'einv all', 'einv_all',
                                      'eac total', 'eac_total',
                                      'energy since', 'energy total',
                                      '누적발전량', '누적 발전량',
                                      'total generation energy',
                                      'high byte of total feed', 'high byte oftotal feed']) or \
       any(k in defn_nospace for k in ['accumulatedpower', 'accumulatedenergy',
                                        'totalpowergeneration', 'totalgenerationenergy',
                                        'totalenergyyield', 'totalpoweryields',
                                        'einvall', 'energysincecommissioning']):
        return 'cumulative_energy'
    if any(k in defn_lower for k in ['daily energy', 'today energy', 'daily power yields',
                                      'daily generation', 'einv day', 'einv_day',
                                      'eac today', 'eac_today',
                                      '일발전량', '일 발전량',
                                      '금일발전량', '금일 발전량']):
        return 'daily_energy'

    # 2) synonym_db 정확 매칭
    syn_match = match_synonym(reg.definition, synonym_db)
    if syn_match and syn_match.get('h01_field'):
        return syn_match['h01_field']

    # 3) 주소 기반 레퍼런스 (같은 제조사만)
    if ref_patterns:
        addr = reg.address if isinstance(reg.address, int) else parse_address(reg.address)
        if addr is not None:
            h01 = get_h01_field_from_ref(addr, ref_patterns, synonym_db)
            if h01:
                return h01

    # 4) 퍼지 매칭
    fuzzy = match_synonym_fuzzy(reg.definition, synonym_db)
    if fuzzy and fuzzy.get('h01_field'):
        return fuzzy['h01_field']

    return ''


# ─── IV Scan 감지 ──────────────────────────────────────────────────────────

_IV_COMMAND_RE = re.compile(r'i-?v\s*(curve\s*)?scan|IV_CURVE_SCAN', re.I)
# Solarize 형식: Tracker N voltage, String N-M current
_TRACKER_VOLTAGE_RE = re.compile(
    r'Tracker\s*(\d+)\s*voltage|IV_TRACKER(\d+)_VOLTAGE(?:_BASE)?|TRACKER_(\d+)_VOLTAGE', re.I)
_IV_STRING_CURRENT_RE = re.compile(
    r'String\s*(\d+)-(\d+)\s*current|IV_STRING(\d+)_(\d+)_CURRENT(?:_BASE)?|'
    r'IV_STRING_(\d+)_(\d+)_CURRENT', re.I)
# Kstar 형식: PV1 Voltage Point 1, PV1 Current Point 1
_PV_VOLTAGE_POINT_RE = re.compile(r'PV(\d+)\s+Voltage\s+Point\s+(\d+)', re.I)
_PV_CURRENT_POINT_RE = re.compile(r'PV(\d+)\s+Current\s+Point\s+(\d+)', re.I)
# "occupying NNNN registers" 패턴
_IV_TOTAL_REGS_RE = re.compile(r'occupying\s+(\d+)\s+registers', re.I)


def detect_iv_from_pdf(registers: List[RegisterRow], pages: list = None) -> dict:
    """
    PDF에서 IV Scan 지원 여부 및 구조 감지
    Solarize 형식: Tracker N voltage (블록), String N-M current (블록)
    Kstar 형식: PV1 Voltage Point 1~100, PV1 Current Point 1~100 (교차)
    """
    result = {'supported': False, 'iv_command_addr': None, 'data_points': 0,
              'trackers': [], 'format': 'unknown', 'total_iv_regs': 0}

    # 1) IV Scan 명령 레지스터 (command만으로는 supported 판정 안 함)
    for reg in registers:
        if _IV_COMMAND_RE.search(reg.definition):
            addr = reg.address if isinstance(reg.address, int) else parse_address(reg.address)
            result['iv_command_addr'] = addr
            # supported는 데이터 레지스터 유무로 판단 (아래에서)
            break

    # 2-A) Solarize 형식: Tracker/String 블록
    tracker_map = {}
    for reg in registers:
        addr = reg.address if isinstance(reg.address, int) else parse_address(reg.address)
        if addr is None:
            continue

        m = _TRACKER_VOLTAGE_RE.search(reg.definition)
        if m:
            tn = int(next(g for g in m.groups() if g))
            tracker_map.setdefault(tn, {'voltage_addr': addr, 'regs': int(reg.regs or '1'), 'strings': {}})
            tracker_map[tn]['voltage_addr'] = addr
            tracker_map[tn]['regs'] = int(reg.regs or '1')
            result['supported'] = True
            continue

        m = _IV_STRING_CURRENT_RE.search(reg.definition)
        if m:
            groups = [g for g in m.groups() if g]
            if len(groups) >= 2:
                tn = int(groups[0])
                sn = int(groups[1])
            tracker_map.setdefault(tn, {'voltage_addr': None, 'regs': 0, 'strings': {}})
            tracker_map[tn]['strings'][sn] = addr
            result['supported'] = True

    if tracker_map:
        result['format'] = 'solarize'
        for tn in sorted(tracker_map):
            t = tracker_map[tn]
            if t.get('regs', 0) > 1:
                result['data_points'] = t['regs']
                break
        # String 주소 간격에서 추정
        if result['data_points'] == 0 and 1 in tracker_map:
            t1 = tracker_map[1]
            if t1['voltage_addr'] and t1['strings'].get(1):
                result['data_points'] = t1['strings'][1] - t1['voltage_addr']
        # Tracker 간 주소 간격에서 추정 (T2 - T1) / (1 + strings_per_tracker)
        if result['data_points'] == 0:
            sorted_trackers = sorted(tracker_map.keys())
            if len(sorted_trackers) >= 2:
                t1_addr = tracker_map[sorted_trackers[0]]['voltage_addr']
                t2_addr = tracker_map[sorted_trackers[1]]['voltage_addr']
                if t1_addr and t2_addr:
                    gap = t2_addr - t1_addr
                    n_blocks = 1 + len(tracker_map[sorted_trackers[0]].get('strings', {}))
                    if n_blocks == 1:
                        n_blocks = 5  # Solarize 기본: 1 tracker + 4 strings
                    result['data_points'] = gap // n_blocks
        for tn in sorted(tracker_map):
            t = tracker_map[tn]
            strings = [{'string_num': sn, 'current_addr': t['strings'][sn]}
                       for sn in sorted(t['strings'])]
            result['trackers'].append({
                'tracker_num': tn, 'voltage_addr': t['voltage_addr'], 'strings': strings,
            })
        return result

    # 2-B) Kstar 형식: PV{n} Voltage/Current Point {m}
    pv_map = {}  # {pv_num: {'voltage_addr': addr, 'max_point': n}}
    for reg in registers:
        addr = reg.address if isinstance(reg.address, int) else parse_address(reg.address)
        if addr is None:
            continue
        defn = reg.definition

        m = _PV_VOLTAGE_POINT_RE.search(defn)
        if m:
            pv_num = int(m.group(1))
            point = int(m.group(2))
            pv_map.setdefault(pv_num, {'voltage_addr': None, 'max_point': 0})
            if point == 1 or pv_map[pv_num]['voltage_addr'] is None:
                pv_map[pv_num]['voltage_addr'] = addr
            pv_map[pv_num]['max_point'] = max(pv_map[pv_num]['max_point'], point)
            result['supported'] = True
            continue

        # "occupying NNNN registers" 패턴에서 총 레지스터 수 추출
        m2 = _IV_TOTAL_REGS_RE.search(reg.comment or '')
        if m2:
            result['total_iv_regs'] = int(m2.group(1))

    if pv_map:
        result['format'] = 'kstar'
        # data_points = max_point (PV1 Voltage Point 100 → 100)
        max_point = max(pv['max_point'] for pv in pv_map.values())
        result['data_points'] = max_point

        # Tracker = PV (Kstar에서는 PV = MPPT = Tracker)
        for pv_num in sorted(pv_map):
            pv = pv_map[pv_num]
            result['trackers'].append({
                'tracker_num': pv_num,
                'voltage_addr': pv['voltage_addr'],
                'strings': [],  # Kstar는 PV 단위 (String 분리 없음)
            })
        return result

    return result


# ─── H01 매칭 상태 표 ──────────────────────────────────────────────────────

# H01 필드별 기본 단위 (RTU UDP 프로토콜 기준)
H01_BASE_UNITS = {
    'r_voltage': 'V', 's_voltage': 'V', 't_voltage': 'V',
    'r_current': 'A', 's_current': 'A', 't_current': 'A',
    'ac_power': 'W', 'power_factor': '', 'frequency': 'Hz',
    'pv_voltage': 'V', 'pv_current': 'A', 'pv_power': 'W',
    'cumulative_energy': 'Wh', 'daily_energy': 'Wh',
}

# 단위 스케일 팩터 (k=1000, M=1000000 등)
_UNIT_SCALE_MAP = {
    ('kWh', 'Wh'): 1000, ('MWh', 'Wh'): 1000000, ('GWh', 'Wh'): 1000000000,
    ('kW', 'W'): 1000, ('MW', 'W'): 1000000,
    ('kVA', 'VA'): 1000, ('kVAr', 'VAr'): 1000,
    ('mA', 'A'): 0.001, ('mV', 'V'): 0.001,
}


def build_h01_match_table(categorized: dict, meta: dict) -> List[dict]:
    """V2: H01 Body 전체 필드 매칭 상태 + 단위/타입 검증"""
    rows = []
    max_mppt = meta.get('max_mppt', 0)
    max_string = meta.get('max_string', 0)

    # 1) DER 겹침 (9개) — 항상 O
    for h01_field, der_info in H01_DER_OVERLAP_FIELDS.items():
        expected_unit = H01_BASE_UNITS.get(h01_field, '')
        rows.append({
            'field': h01_field,
            'source': 'DER',
            'status': 'O',
            'type': 'S32', 'unit': expected_unit, 'scale': '0.1',
            'address': f'0x{der_info["addr_low"]:04X}~0x{der_info["addr_high"]:04X}',
            'definition': der_info['der_name'],
            'note': 'DER-AVM 고정 주소 사용',
        })

    # 2) Handler 계산 (pv_voltage, pv_current) — 항상 O
    for h01_field, rule in H01_HANDLER_COMPUTED_FIELDS.items():
        expected_unit = H01_BASE_UNITS.get(h01_field, '')
        # pv_power: PDF에서 발견되면 PDF 소스, 아니면 handler 계산
        if h01_field == 'pv_power':
            pv_power_reg = _find_matched_reg(categorized, 'pv_power')
            if pv_power_reg:
                rows.append(_make_pdf_match_row('pv_power', pv_power_reg))
            else:
                rows.append({
                    'field': h01_field, 'source': 'HANDLER', 'status': 'O',
                    'address': '-', 'definition': '-',
                    'type': 'U16', 'unit': expected_unit, 'scale': '',
                    'note': f'handler 계산: {rule}',
                })
            continue
        rows.append({
            'field': h01_field,
            'source': 'HANDLER',
            'status': 'O',
            'address': '-',
            'definition': '-',
            'type': 'U16', 'unit': expected_unit, 'scale': '',
            'note': f'handler 계산: {rule}',
        })

    energy_reg = _find_matched_reg(categorized, 'cumulative_energy')
    if not energy_reg:
        energy_reg = _find_matched_reg(categorized, 'total_energy')
    rows.append(_make_pdf_match_row('cumulative_energy', energy_reg, 'Total Energy 미발견'))

    # cumulative_energy_low (소수부/Wh/Low Byte) — 있으면 매핑, 없으면 N/A
    energy_low_reg = _find_matched_reg(categorized, 'cumulative_energy_low')
    if energy_low_reg:
        rows.append(_make_pdf_match_row('cumulative_energy_low', energy_low_reg))
    else:
        rows.append({
            'field': 'cumulative_energy_low', 'source': 'PDF',
            'status': '-', 'address': '-', 'definition': '-',
            'type': '', 'unit': '', 'scale': '',
            'note': '소수부 레지스터 없음 (단일 레지스터)',
        })

    status_reg = _find_matched_reg(categorized, 'inverter_status', cat='STATUS')
    if not status_reg and categorized.get('STATUS'):
        # h01_field가 없는 STATUS 레지스터 중 Work Mode/Inverter Mode 키워드 검색
        for sr in categorized['STATUS']:
            dl = sr.definition.lower().replace('_', ' ')
            if any(k in dl for k in ['inverter mode', 'work mode', 'work state',
                                      'operating mode', 'operation state',
                                      'running status', 'inverter state',
                                      '시스템동작상태', '동작상태']) or \
               dl.strip() == 'state':
                status_reg = sr
                break
    # 상태 정의(값 테이블) 검증 — 통상 Appendix에 별도 정의
    status_note = ''
    rows.append(_make_pdf_match_row('status', status_reg, 'Work State 미발견'))
    if status_note and rows[-1]['status'] == 'O':
        rows[-1]['note'] = status_note

    alarm_regs = categorized.get('ALARM', [])
    alarm_dist = distribute_alarms(alarm_regs)
    for slot in ['alarm1', 'alarm2', 'alarm3']:
        regs = alarm_dist.get(slot, [])
        if regs:
            rows.append(_make_pdf_match_row(slot, regs[0]))
        else:
            rows.append({
                'field': slot, 'source': 'PDF',
                'status': '-' if slot != 'alarm1' else 'X',
                'address': '-', 'definition': '-', 'type': '', 'unit': '', 'scale': '',
                'note': '보조 알람 없음' if slot != 'alarm1' else '알람 미발견',
            })

    for n in range(1, max_mppt + 1):
        for mtype in ['voltage', 'current']:
            field = f'mppt{n}_{mtype}'
            reg = _find_matched_reg(categorized, field)
            if reg:
                rows.append(_make_pdf_match_row(field, reg))
            elif mtype == 'current':
                # current 없으면 voltage + power로 계산 가능한지 체크
                has_v = _find_matched_reg(categorized, f'mppt{n}_voltage')
                has_p = _find_matched_reg(categorized, f'mppt{n}_power')
                if has_v and has_p:
                    rows.append({
                        'field': field, 'source': 'HANDLER', 'status': 'O',
                        'address': '-', 'definition': '-',
                        'type': 'U16', 'unit': 'A', 'scale': '',
                        'note': f'handler 계산: mppt{n}_power / mppt{n}_voltage',
                    })
                else:
                    rows.append(_make_pdf_match_row(field, None))
            else:
                rows.append(_make_pdf_match_row(field, None))

    for n in range(1, max_string + 1):
        field = f'string{n}_current'
        reg = _find_matched_reg(categorized, field)
        if reg:
            rows.append(_make_pdf_match_row(field, reg))
        else:
            # String별 레지스터 없으면: MPPT당 1 String이면 MPPT current = String current
            mppt_n = n  # String N = MPPT N (1:1 매핑, PDF에 String별 레지스터 없을 때)
            if mppt_n <= max_mppt:
                mppt_reg = _find_matched_reg(categorized, f'mppt{mppt_n}_current')
                if mppt_reg:
                    rows.append({
                        'field': field, 'source': 'HANDLER', 'status': 'O',
                        'address': mppt_reg.address_hex, 'definition': mppt_reg.definition,
                        'type': mppt_reg.data_type, 'unit': 'A', 'scale': mppt_reg.scale or '',
                        'note': f'= mppt{mppt_n}_current (MPPT당 1 String)',
                    })
                else:
                    # MPPT current도 handler 계산이면
                    rows.append({
                        'field': field, 'source': 'HANDLER', 'status': 'O',
                        'address': '-', 'definition': '-',
                        'type': 'U16', 'unit': 'A', 'scale': '',
                        'note': f'= mppt{mppt_n}_current (handler 계산)',
                    })
            else:
                rows.append(_make_pdf_match_row(field, None, 'String current 미지원 시 생략'))

    return rows


def _find_matched_reg(categorized: dict, h01_field: str, cat: str = None) -> Optional[RegisterRow]:
    search_cats = [cat] if cat else ['MONITORING', 'INFO', 'STATUS', 'ALARM']
    for c in search_cats:
        for reg in categorized.get(c, []):
            if reg.h01_field == h01_field:
                return reg
    return None


def _get_unit_scale(reg_unit: str, h01_field: str) -> tuple:
    """레지스터 단위 → (H01 기본단위 변환 스케일, 변환 설명)"""
    base = H01_BASE_UNITS.get(h01_field, '')
    ru = (reg_unit or '').strip()
    if not base or not ru or ru == base:
        return (1, '')
    # k/M/G 접두사 변환
    for (src, dst), factor in _UNIT_SCALE_MAP.items():
        if ru.upper() == src.upper() and dst.upper() == base.upper():
            return (factor, f'{ru}→{base} (*{factor})')
    # 스케일 포함된 단위 (예: 0.1V → V)
    return (1, '')


def _make_pdf_match_row(h01_field: str, reg, miss_note: str = '') -> dict:
    """PDF 매칭 행 생성 — scale은 PDF 원문 그대로, 정의 테이블 포함"""
    if reg:
        _, unit_desc = _get_unit_scale(reg.unit, h01_field)
        # 정의 요약 (value_definitions가 있으면)
        val_defs = getattr(reg, 'value_definitions', None)
        val_defs_str = ''
        if val_defs and isinstance(val_defs, dict):
            items = sorted(val_defs.items())[:6]
            val_defs_str = ', '.join(f'{k}={v[:15]}' for k, v in items)
            if len(val_defs) > 6:
                val_defs_str += f'... (+{len(val_defs)-6})'
        note = unit_desc
        if val_defs_str:
            note = f'{val_defs_str}' + (f' | {unit_desc}' if unit_desc else '')
        return {
            'field': h01_field, 'source': 'PDF', 'status': 'O',
            'address': reg.address_hex, 'definition': reg.definition,
            'type': reg.data_type, 'unit': reg.unit or '',
            'scale': reg.scale or '',
            'note': note,
            'value_definitions': val_defs,
        }
    else:
        return {
            'field': h01_field, 'source': 'PDF', 'status': 'X',
            'address': '-', 'definition': '-',
            'type': '', 'unit': '', 'scale': '',
            'note': miss_note,
        }


# ─── Stage 1 메인 ───────────────────────────────────────────────────────────

def run_stage1(
    input_path: str,
    output_dir: str,
    device_type: str = 'inverter',
    progress: ProgressCallback = None,
) -> dict:
    """Stage 1: PDF/Excel → Stage 1 Excel (MAPPING_RULES_V2)"""
    def log(msg, level='info'):
        if progress:
            progress(msg, level)
    openpyxl = get_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    log('레퍼런스 로딩 중...')
    synonym_db = load_synonym_db()
    review_history = load_review_history()
    ref_patterns = load_reference_patterns()
    log(f'  synonym_db: {len(synonym_db.get("fields", {}))}개 필드')
    log(f'  review_history: {len(review_history.get("approved", []))}개 이력')
    log(f'  레퍼런스: {len(ref_patterns)}개 프로토콜')

    ext = os.path.splitext(input_path)[1].lower()
    basename = os.path.splitext(os.path.basename(input_path))[0]

    log(f'입력 파일 읽기: {os.path.basename(input_path)}')
    all_tables = []

    if ext == '.pdf':
        pages = extract_pdf_text_and_tables(input_path)
        log(f'  PDF {len(pages)}페이지 추출')

        # Input Register (FC04) 섹션 키워드
        _INPUT_START = [
            'running information variable address', 'running information',
            'input register', 'input (read only)', 'read only register',
            'input reg',
        ]
        # Holding Register (FC03) 섹션 키워드
        _HOLDING_START = [
            'parameter setting address definition', 'parameter setting',
            'hold register', 'holding register', 'holding reg',
            'parameter register',
        ]

        section_input_start = None   # FC04 Input
        section_holding_start = None  # FC03 Holding

        def _is_section_title(page_text: str, keywords: list) -> bool:
            """줄 단위로 매칭 — 섹션 제목(짧은 줄)에서만 키워드 감지"""
            for line in page_text.split('\n'):
                ll = line.strip().lower()
                if len(ll) > 80:  # 긴 줄은 설명문 → 건너뜀
                    continue
                # "Function N Read/Write ..." 같은 Modbus 함수 설명 제외
                if ll.startswith('function') or 'unsigned integer' in ll:
                    continue
                if any(m in ll for m in keywords):
                    return True
            return False

        for p in pages:
            page_text = p['text']
            pnum = p['page']
            if section_input_start is None and _is_section_title(page_text, _INPUT_START):
                section_input_start = pnum
            if section_holding_start is None and _is_section_title(page_text, _HOLDING_START):
                if pnum != section_input_start:
                    section_holding_start = pnum

        # 페이지별 FC 분류 — 섹션 순서에 무관하게 동작
        # 두 섹션이 모두 감지되면: 먼저 시작하는 섹션 → 나중 섹션 시작 전까지
        tables_input, tables_holding, tables_other = [], [], []
        for p in pages:
            pnum = p['page']
            # 어떤 섹션에 속하는지 판정
            in_input = False
            in_holding = False
            if section_input_start is not None and section_holding_start is not None:
                if section_input_start < section_holding_start:
                    # Input이 먼저: Input → Holding 시작 전까지
                    in_input = (pnum >= section_input_start and pnum < section_holding_start)
                    in_holding = (pnum >= section_holding_start)
                else:
                    # Holding이 먼저: Holding → Input 시작 전까지
                    in_holding = (pnum >= section_holding_start and pnum < section_input_start)
                    in_input = (pnum >= section_input_start)
            elif section_input_start is not None:
                in_input = (pnum >= section_input_start)
            elif section_holding_start is not None:
                in_holding = (pnum >= section_holding_start)

            for tab in p['tables']:
                if in_holding:
                    tables_holding.append(tab)
                elif in_input:
                    tables_input.append(tab)
                else:
                    tables_other.append(tab)

        # FC 태깅: Input=FC04, Holding=FC03, 기타='' (FC 구분 불가)
        all_tables = []
        fc_list = []
        for tab in tables_input:
            all_tables.append(tab); fc_list.append('04')
        for tab in tables_other:
            all_tables.append(tab); fc_list.append('')
        for tab in tables_holding:
            all_tables.append(tab); fc_list.append('03')
        log(f'  Input(FC04): {len(tables_input)}개 (page {section_input_start}~), '
            f'Holding(FC03): {len(tables_holding)}개 (page {section_holding_start}~), '
            f'기타: {len(tables_other)}개')

    elif ext in ('.xlsx', '.xls'):
        sheets = extract_excel_sheets(input_path)
        log(f'  Excel {len(sheets)}시트 추출')
        fc_list = []
        for sname, rows in sheets.items():
            if rows:
                all_tables.append(rows)
                fc_list.append('')  # Excel에서는 FC 구분 불가
    else:
        raise ValueError(f'지원하지 않는 파일 형식: {ext}')

    log('레지스터 테이블 파싱...')
    registers = extract_registers_from_tables(all_tables, fc_list=fc_list)
    log(f'  {len(registers)}개 레지스터 추출 (원본)')

    if not registers:
        raise NotRegisterMapError(
            'Modbus 레지스터(FC03/FC04)를 찾지 못했습니다.\n'
            '이 파일은 Modbus 프로토콜 문서가 아닌 것 같습니다.\n'
            '인버터 Modbus Register Map / Protocol 문서를 업로드해주세요.'
        )

    manufacturer = basename.split('_')[0].split(' ')[0]
    log(f'  제조사 (파일명 기반): {manufacturer}')

    DER_FIXED_ADDRS = set()
    for dr in DER_CONTROL_REGS:
        DER_FIXED_ADDRS.add(dr['addr'])
    for dr in DER_MONITOR_REGS:
        DER_FIXED_ADDRS.add(dr['addr'])

    before = len(registers)
    registers = [r for r in registers
                 if (r.address if isinstance(r.address, int) else parse_address(r.address))
                 not in DER_FIXED_ADDRS]
    if before - len(registers):
        log(f'  DER 고정 주소 제외: {before - len(registers)}개')

    if ref_patterns:
        mfr_lower = manufacturer.lower()
        matched_ref = {k: v for k, v in ref_patterns.items() if mfr_lower in k.lower()}
        if matched_ref:
            enriched = 0
            for reg in registers:
                addr = reg.address if isinstance(reg.address, int) else parse_address(reg.address)
                if addr is None:
                    continue
                ref_name = None
                for proto, addr_map in matched_ref.items():
                    if addr in addr_map:
                        ref_name = addr_map[addr]
                        break
                if ref_name:
                    original = reg.definition
                    reg.definition = ref_name
                    if reg.comment and original != ref_name:
                        reg.comment = f'{original} | {reg.comment}'
                    elif original != ref_name:
                        reg.comment = original
                    enriched += 1
            if enriched:
                log(f'  레퍼런스 enrichment: {enriched}/{len(registers)}개 ({list(matched_ref.keys())})')
        else:
            log(f'  레퍼런스 enrichment: 해당 제조사({manufacturer}) 없음')

    normalized = 0
    _STATUS_ALARM_KW = {'status', 'state', 'mode', 'fault', 'alarm', 'error', 'warning'}
    for reg in registers:
        if detect_channel_number(reg.definition):
            continue
        # STATUS/ALARM 키워드가 있는 레지스터는 이름 변경 방지 (학습 데이터 충돌 방지)
        defn_words = set(reg.definition.lower().replace('_', ' ').split())
        if defn_words & _STATUS_ALARM_KW:
            continue
        syn = match_synonym(reg.definition, synonym_db)
        if syn and syn['field'] != to_upper_snake(reg.definition):
            original = reg.definition
            reg.definition = syn['field']
            if reg.comment and original != syn['field']:
                reg.comment = f'{original} | {reg.comment}'
            elif original != syn['field']:
                reg.comment = original
            normalized += 1
    if normalized:
        log(f'  synonym_db 정규화: {normalized}개')

    seen_names = {}
    for i, reg in enumerate(registers):
        name = to_upper_snake(reg.definition)
        # 이름+단위 조합으로 중복 체크 (같은 이름이라도 단위 다르면 유지)
        # SunSpec: 같은 이름(DCA/DCV/DCW)이 MPPT별 반복 → 주소 다르면 유지
        name_key = f'{name}_{reg.unit}' if reg.unit else name
        if name_key in seen_names:
            # 주소가 다르면 별도 레지스터로 유지
            prev_addr = registers[seen_names[name_key]].address if registers[seen_names[name_key]] else 0
            if reg.address != prev_addr:
                name_key = f'{name}_{reg.address}'  # 주소 포함 키로 변경
            else:
                registers[i] = None
                continue
        seen_names[name_key] = i
    registers = [r for r in registers if r is not None]
    log(f'  중복 제거 후: {len(registers)}개')

    protocol_version = ''
    if ext == '.pdf':
        full_text = ' '.join(p['text'] for p in pages[:3])
        m = re.search(r'[Vv](\d+\.\d+(?:\.\d+)?)', full_text)
        if m:
            protocol_version = f'V{m.group(1)}'

    max_mppt = 0
    max_string = 0
    for reg in registers:
        ch = detect_channel_number(reg.definition)
        if not ch:
            addr = reg.address if isinstance(reg.address, int) else parse_address(reg.address)
            if addr is not None:
                ch = detect_channel_from_ref(addr, ref_patterns)
        if ch:
            prefix, n = ch
            if prefix == 'MPPT':
                max_mppt = max(max_mppt, n)
            elif prefix == 'STRING':
                max_string = max(max_string, n)

    # Central type 감지: MPPT 번호 없지만 DC/PV/Input voltage/current가 있으면 MPPT=1
    if max_mppt == 0:
        for reg in registers:
            dl = reg.definition.lower().replace('_', ' ')
            # DC Voltage/Current (SunSpec I_DC_*, ESINV DC Voltage 1, 일반 DC)
            # PV Voltage/Current (번호 없음)
            # Input Voltage/Current
            if (re.search(r'\b(i dc|dc)\s*(voltage|current)', dl) or
                re.search(r'^(pv|input)\s+(voltage|current)', dl)):
                # "PV VoltageHigh Fault" 같은 오감지 제외
                if 'fault' not in dl and 'high' not in dl and 'low' not in dl:
                    max_mppt = 1
                    # mppt1_voltage/current로 매핑 (assign_h01_field에서 처리)
                    break

    # String 수 보정: String >= MPPT (최소 MPPT당 1 String)
    # PDF에 String별 레지스터가 없어도 실제로는 MPPT에 String이 연결됨
    if max_mppt > 0 and max_string < max_mppt:
        max_string = max_mppt

    # V2: IV Scan 지원 = IV 데이터 레지스터(Tracker/PV point)가 있어야 함
    # IV command(0x600D)만 있고 데이터 레지스터 없으면 미지원 (예: Huawei)
    iv_info = detect_iv_from_pdf(registers)
    iv_scan_supported = iv_info['supported'] and len(iv_info.get('trackers', [])) > 0

    # 확정 제조사 + 데이터 레지스터 없으면 keyword fallback
    if not iv_scan_supported and manufacturer.lower() in ('solarize', 'kstar', 'senergy'):
        iv_scan_supported = True
        iv_info['supported'] = True

    # IV command는 DER 고정 주소(0x600D)에서 제거되므로, 지원 시 고정값 삽입
    if iv_scan_supported and not iv_info.get('iv_command_addr'):
        iv_info['iv_command_addr'] = 0x600D
        iv_info['supported'] = True

    meta = {
        'manufacturer': manufacturer,
        'protocol_version': protocol_version,
        'device_type': device_type,
        'max_mppt': max_mppt,
        'max_string': max_string,
        'string_monitoring': max_string > max_mppt,  # True: String별 전류 모니터링 지원
        'iv_scan': iv_scan_supported and device_type == 'inverter',
        'iv_data_points': iv_info.get('data_points', 0),
        'iv_trackers': len(iv_info.get('trackers', [])),
        'iv_command_addr': f'0x{iv_info["iv_command_addr"]:04X}' if iv_info.get('iv_command_addr') else '-',
        'source_file': os.path.basename(input_path),
        'extracted_date': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'total_extracted': len(registers),
        'mapping_rules': 'V2',
    }

    log(f'  제조사: {manufacturer}, MPPT: {max_mppt}, String: {max_string}')
    if meta['iv_scan']:
        log(f'  IV Scan: Yes (command={meta["iv_command_addr"]}, '
            f'trackers={meta["iv_trackers"]}, data_points={meta["iv_data_points"]})')
    else:
        log(f'  IV Scan: No')

    log('카테고리 분류 중 (MAPPING_RULES_V2)...')
    categorized = {cat: [] for cat in
                   ['INFO', 'MONITORING', 'STATUS', 'ALARM',
                    'DER_CONTROL', 'DER_MONITOR', 'IV_SCAN', 'REVIEW']}
    excluded = []

    for reg in registers:
        cat, reason = classify_register_with_rules(
            reg, synonym_db, review_history, ref_patterns,
            device_type, all_regs=registers)
        if cat == 'EXCLUDE':
            excluded.append(reg)
            continue
        reg.category = cat
        reg.h01_field = assign_h01_field(reg, synonym_db, ref_patterns)

        if cat == 'REVIEW':
            reg.review_reason = reason or '자동 분류 불가'
            fuzzy = match_synonym_fuzzy(reg.definition, synonym_db, threshold=0.4)
            if fuzzy:
                reg.review_suggestion = f'{fuzzy["category"]}/{fuzzy["field"]} (유사도 {fuzzy["score"]:.0%})'
            else:
                reg.review_suggestion = '분류 불가'

        if cat in categorized:
            categorized[cat].append(reg)

    if device_type == 'inverter':
        categorized['DER_CONTROL'] = []
        categorized['DER_MONITOR'] = []
        for dr in DER_CONTROL_REGS:
            categorized['DER_CONTROL'].append(RegisterRow(
                definition=dr['name'], address=dr['addr'],
                data_type=dr['type'], scale=dr.get('scale', ''),
                rw=dr['rw'], comment=dr['desc'], category='DER_CONTROL'))
        for dr in DER_MONITOR_REGS:
            categorized['DER_MONITOR'].append(RegisterRow(
                definition=dr['name'], address=dr['addr'],
                data_type=dr['type'], scale=dr.get('scale', ''),
                unit=dr.get('unit', ''), rw='RO',
                comment=dr.get('desc', ''), category='DER_MONITOR'))

    if device_type != 'inverter':
        categorized['DER_CONTROL'] = []
        categorized['DER_MONITOR'] = []
        categorized['IV_SCAN'] = []

    # ── Model 레지스터 없으면 PDF 메타데이터에서 모델명 추출 ──
    has_model_reg = any(
        any(k in r.definition.lower() for k in ['model', 'device type', 'type code'])
        and not any(k in r.definition.lower() for k in ['working', 'battery', 'pf', 'init fault',
                                                          'output type', 'phase type'])
        for r in categorized.get('INFO', []))
    if not has_model_reg:
        pdf_model = _extract_model_from_pdf(input_path, manufacturer)
        if pdf_model:
            categorized['INFO'].insert(0, RegisterRow(
                definition='DEVICE_MODEL (PDF)',
                address='PDF',
                address_hex='PDF',
                data_type='TEXT',
                rw='RO',
                comment=pdf_model,
                category='INFO',
            ))
            meta['pdf_model'] = pdf_model
            log(f'  Model 레지스터 없음 → PDF에서 추출: {pdf_model}')

    # ── SunSpec: 표준 정의 데이터 적용 (SolarEdge, Fronius 등) ──
    from .sunspec import is_sunspec_pdf, apply_sunspec_definitions, detect_sunspec_mppt
    if is_sunspec_pdf(registers, manufacturer):
        apply_sunspec_definitions(categorized, log)
        # SunSpec MPPT 블록으로 max_mppt 업데이트
        sunspec_mppt = detect_sunspec_mppt(registers)
        log(f'  SunSpec MPPT 블록: {sunspec_mppt["mppt_count"]}개, 현재 max_mppt={max_mppt}')
        if sunspec_mppt['mppt_count'] > max_mppt:
            max_mppt = sunspec_mppt['mppt_count']
            meta['max_mppt'] = max_mppt
            # String 수도 보정
            if max_string < max_mppt:
                max_string = max_mppt
                meta['max_string'] = max_string
            log(f'  SunSpec MPPT → max_mppt={max_mppt}, max_string={max_string}로 업데이트')
        log(f'  SunSpec 표준 인버터 감지 → 정의 데이터 자동 적용')

    counts = {cat: len(regs) for cat, regs in categorized.items()}
    log('분류 결과:')
    for cat, cnt in counts.items():
        if cnt > 0:
            log(f'  {cat:15s}: {cnt}개')
    log(f'  제외: {len(excluded)}개')

    # ── 정의 테이블 탐색 + 레지스터 연결 ──
    if ext == '.pdf':
        def_tables = scan_definition_tables(pages, registers=registers)
        st_count = sum(len(d['values']) for d in def_tables['status_defs'])
        al_count = sum(len(d['values']) for d in def_tables['alarm_defs'])
        if st_count or al_count:
            _link_definitions_to_registers(categorized, def_tables)
            log(f'  정의 테이블: status {len(def_tables["status_defs"])}개({st_count}값), '
                f'alarm {len(def_tables["alarm_defs"])}개({al_count}값)')

    # ── 정의 파일 fallback: definitions/{manufacturer}_definitions.json ──
    _apply_saved_definitions(categorized, manufacturer, log)

    h01_match_table = build_h01_match_table(categorized, meta)

    # ── 5가지 섹션 후보 제안 ──
    all_regs_flat = []
    for cat_name, cat_regs in categorized.items():
        if cat_name not in ('DER_CONTROL', 'DER_MONITOR', 'IV_SCAN'):
            all_regs_flat.extend(cat_regs)
    suggestions = _build_all_suggestions(
        h01_match_table, categorized, all_regs_flat, meta, log)

    h01_matched = sum(1 for r in h01_match_table if r['status'] == 'O')
    h01_total = sum(1 for r in h01_match_table if r['status'] != '-')
    log(f'H01 매칭: {h01_matched}/{h01_total}')

    # ── DER 매칭 테이블 생성 ──
    der_match_table = _build_der_match_table(categorized)
    der_matched = sum(1 for r in der_match_table if r['status'] == 'O')
    der_total = len(der_match_table)
    log(f'DER 매칭: {der_matched}/{der_total}')

    output_name = f'test_{basename}_stage1.xlsx'
    output_path = os.path.join(output_dir, output_name)
    log(f'Excel 생성: {output_name}')

    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='333333')
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color='1F4E79')

    def _write_header(ws, columns, row=1):
        for j, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=j, value=col_name)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = thin_border

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 1: INFO — 인포 매칭 & 메타데이터
    # ═══════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = '1_INFO'

    # 메타데이터 섹션
    ws['A1'] = 'Stage 1 — 레지스터맵 추출 (MAPPING_RULES_V2)'
    ws['A1'].font = title_font
    meta_items = [
        ('제조사', manufacturer), ('프로토콜 버전', protocol_version),
        ('설비 타입', device_type), ('MPPT', max_mppt),
        ('String', max_string),
        ('String 모니터링', 'Yes' if meta['string_monitoring'] else 'No (MPPT당 1 String)'),
        ('IV Scan', 'Yes' if meta['iv_scan'] else 'No'),
        ('IV Data Points', meta.get('iv_data_points', 0) if meta['iv_scan'] else '-'),
        ('IV Trackers', meta.get('iv_trackers', 0) if meta['iv_scan'] else '-'),
        ('원본 파일', os.path.basename(input_path)), ('추출 일시', meta['extracted_date']),
        ('H01 매칭', f'{h01_matched}/{h01_total}'),
        ('DER 매칭', f'{der_matched}/{der_total}'),
        ('추출 레지스터', meta['total_extracted']),
        ('REVIEW', counts.get('REVIEW', 0)),
    ]
    for i, (k, v) in enumerate(meta_items, start=3):
        ws[f'A{i}'] = k
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = str(v)

    # INFO 레지스터 섹션
    info_start = len(meta_items) + 5
    ws.cell(row=info_start, column=1, value='INFO 레지스터').font = section_font
    info_cols = ['No', 'Definition', 'Address', 'FC', 'Type', 'Unit/Scale', 'R/W', 'H01 Field', 'Comment']
    _write_header(ws, info_cols, info_start + 1)

    info_regs = sorted(categorized.get('INFO', []),
                       key=lambda r: (r.address if isinstance(r.address, int) else 0))
    for i, reg in enumerate(info_regs, start=1):
        su = f'{reg.unit} {reg.scale}'.strip() if reg.unit or reg.scale else ''
        vals = [i, reg.definition, reg.address_hex, reg.fc or '', reg.data_type, su,
                reg.rw, reg.h01_field or '', reg.comment]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=info_start + 1 + i, column=j, value=val)
            cell.border = thin_border
            cell.fill = PatternFill('solid', fgColor=CATEGORY_COLORS['INFO'])

    # STATUS 섹션
    status_start = info_start + len(info_regs) + 4
    ws.cell(row=status_start, column=1, value='STATUS 레지스터').font = section_font
    _write_header(ws, info_cols, status_start + 1)

    status_regs = sorted(categorized.get('STATUS', []),
                         key=lambda r: (r.address if isinstance(r.address, int) else 0))
    for i, reg in enumerate(status_regs, start=1):
        su = f'{reg.unit} {reg.scale}'.strip() if reg.unit or reg.scale else ''
        vals = [i, reg.definition, reg.address_hex, reg.fc or '', reg.data_type, su,
                reg.rw, reg.h01_field or '', reg.comment]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=status_start + 1 + i, column=j, value=val)
            cell.border = thin_border
            cell.fill = PatternFill('solid', fgColor=CATEGORY_COLORS['STATUS'])

    # ALARM 섹션
    alarm_start = status_start + len(status_regs) + 4
    ws.cell(row=alarm_start, column=1, value='ALARM 레지스터').font = section_font
    _write_header(ws, info_cols, alarm_start + 1)

    alarm_regs_sorted = sorted(categorized.get('ALARM', []),
                               key=lambda r: (r.address if isinstance(r.address, int) else 0))
    for i, reg in enumerate(alarm_regs_sorted, start=1):
        su = f'{reg.unit} {reg.scale}'.strip() if reg.unit or reg.scale else ''
        vals = [i, reg.definition, reg.address_hex, reg.fc or '', reg.data_type, su,
                reg.rw, reg.h01_field or '', reg.comment]
        for j, val in enumerate(vals, start=1):
            cell = ws.cell(row=alarm_start + 1 + i, column=j, value=val)
            cell.border = thin_border
            cell.fill = PatternFill('solid', fgColor=CATEGORY_COLORS['ALARM'])

    # REVIEW 섹션 (있으면)
    review_regs = categorized.get('REVIEW', [])
    if review_regs:
        review_start = alarm_start + len(alarm_regs_sorted) + 4
        ws.cell(row=review_start, column=1, value=f'REVIEW ({len(review_regs)}개)').font = section_font
        review_cols = ['No', 'Definition', 'Address', 'FC', 'Type', 'Unit/Scale', 'R/W', '사유', '제안']
        _write_header(ws, review_cols, review_start + 1)
        for i, reg in enumerate(sorted(review_regs, key=lambda r: (r.address if isinstance(r.address, int) else 0)),
                                start=1):
            su = f'{reg.unit} {reg.scale}'.strip() if reg.unit or reg.scale else ''
            vals = [i, reg.definition, reg.address_hex, reg.fc or '', reg.data_type, su,
                    reg.rw, reg.review_reason, reg.review_suggestion]
            for j, val in enumerate(vals, start=1):
                cell = ws.cell(row=review_start + 1 + i, column=j, value=val)
                cell.border = thin_border
                cell.fill = PatternFill('solid', fgColor=CATEGORY_COLORS['REVIEW'])

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 35

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 2: H01_MATCH — H01 Body 필드 매칭
    # ═══════════════════════════════════════════════════════════════════
    ws_h01 = wb.create_sheet('2_H01')
    ws_h01['A1'] = f'H01 모니터링 매칭 — {h01_matched}/{h01_total}'
    ws_h01['A1'].font = title_font

    h01_cols = ['No', 'H01 Field', 'Source', 'Status', 'Address', 'Definition', 'Type', 'Unit', 'Scale', 'Note']
    _write_header(ws_h01, h01_cols, 3)

    for i, rd in enumerate(h01_match_table, start=1):
        vals = [i, rd['field'], rd['source'], rd['status'],
                rd['address'], rd['definition'],
                rd.get('type', ''), rd.get('unit', ''), rd.get('scale', ''), rd.get('note', '')]
        sc = MATCH_COLORS.get(rd['status'], 'FFFFFF')
        for j, val in enumerate(vals, start=1):
            cell = ws_h01.cell(row=3 + i, column=j, value=val)
            cell.border = thin_border
            if j == 4:
                cell.fill = PatternFill('solid', fgColor=sc)
            elif rd['source'] == 'DER':
                cell.fill = PatternFill('solid', fgColor='D9D2E9')
            elif rd['source'] == 'HANDLER':
                cell.fill = PatternFill('solid', fgColor='FCE5CD')

    ws_h01.column_dimensions['B'].width = 25
    ws_h01.column_dimensions['E'].width = 20
    ws_h01.column_dimensions['F'].width = 40
    ws_h01.column_dimensions['J'].width = 40

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 3: MPPT_STRING — MPPT/String 채널 매칭 + MONITORING 전체
    # ═══════════════════════════════════════════════════════════════════
    ws_ms = wb.create_sheet('3_MPPT_STRING')
    ws_ms['A1'] = f'MPPT & String 매칭 — MPPT {meta.get("max_mppt", 0)}ch / String {meta.get("max_string", 0)}ch'
    ws_ms['A1'].font = title_font

    # 메타
    ms_meta = [
        ('MPPT 채널 수', meta.get('max_mppt', 0)),
        ('String 채널 수', meta.get('max_string', 0)),
        ('String 모니터링', 'Yes' if meta.get('string_monitoring') else 'No (MPPT당 1 String)'),
    ]
    for i, (k, v) in enumerate(ms_meta, start=3):
        ws_ms.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_ms.cell(row=i, column=2, value=str(v))

    # MPPT 채널 H01 매칭 섹션
    mppt_rows = [r for r in h01_match_table
                 if re.match(r'mppt\d+_(voltage|current|power)', r['field'])]
    mppt_sec = len(ms_meta) + 5
    ws_ms.cell(row=mppt_sec, column=1,
               value=f'MPPT 채널 H01 매칭 ({len(mppt_rows)}개)').font = section_font
    ms_cols = ['No', 'H01 Field', 'Source', 'Status', 'Address', 'Definition', 'Type', 'Unit', 'Scale', 'Note']
    _write_header(ws_ms, ms_cols, mppt_sec + 1)
    for i, rd in enumerate(mppt_rows, start=1):
        sc = MATCH_COLORS.get(rd['status'], 'FFFFFF')
        vals = [i, rd['field'], rd['source'], rd['status'],
                rd['address'], rd['definition'],
                rd.get('type', ''), rd.get('unit', ''), rd.get('scale', ''), rd.get('note', '')]
        for j, val in enumerate(vals, start=1):
            cell = ws_ms.cell(row=mppt_sec + 1 + i, column=j, value=val)
            cell.border = thin_border
            if j == 4:
                cell.fill = PatternFill('solid', fgColor=sc)
            elif rd['source'] == 'HANDLER':
                cell.fill = PatternFill('solid', fgColor='FCE5CD')

    # String 채널 H01 매칭 섹션
    str_rows = [r for r in h01_match_table
                if re.match(r'string\d+_current', r['field'])]
    str_sec = mppt_sec + len(mppt_rows) + 4
    ws_ms.cell(row=str_sec, column=1,
               value=f'String 채널 H01 매칭 ({len(str_rows)}개)').font = section_font
    _write_header(ws_ms, ms_cols, str_sec + 1)
    for i, rd in enumerate(str_rows, start=1):
        sc = MATCH_COLORS.get(rd['status'], 'FFFFFF')
        vals = [i, rd['field'], rd['source'], rd['status'],
                rd['address'], rd['definition'],
                rd.get('type', ''), rd.get('unit', ''), rd.get('scale', ''), rd.get('note', '')]
        for j, val in enumerate(vals, start=1):
            cell = ws_ms.cell(row=str_sec + 1 + i, column=j, value=val)
            cell.border = thin_border
            if j == 4:
                cell.fill = PatternFill('solid', fgColor=sc)
            elif rd['source'] == 'HANDLER':
                cell.fill = PatternFill('solid', fgColor='FCE5CD')

    # MONITORING 전체 목록 (Stage 2 입력용)
    mon_regs = sorted(categorized.get('MONITORING', []),
                      key=lambda r: (r.address if isinstance(r.address, int) else 0))
    mon_sec = str_sec + len(str_rows) + 4
    ws_ms.cell(row=mon_sec, column=1,
               value=f'MONITORING 전체 ({len(mon_regs)}개)').font = section_font
    mon_cols = ['No', 'Definition', 'Address', 'FC', 'Type', 'Unit/Scale', 'R/W', 'H01 Field']
    _write_header(ws_ms, mon_cols, mon_sec + 1)
    for i, reg in enumerate(mon_regs, start=1):
        su = f'{reg.unit} {reg.scale}'.strip() if reg.unit or reg.scale else ''
        vals = [i, reg.definition, reg.address_hex, reg.fc or '', reg.data_type, su,
                reg.rw, reg.h01_field or '']
        for j, val in enumerate(vals, start=1):
            cell = ws_ms.cell(row=mon_sec + 1 + i, column=j, value=val)
            cell.border = thin_border
            cell.fill = PatternFill('solid', fgColor=CATEGORY_COLORS['MONITORING'])

    ws_ms.column_dimensions['B'].width = 25
    ws_ms.column_dimensions['E'].width = 20
    ws_ms.column_dimensions['F'].width = 40
    ws_ms.column_dimensions['J'].width = 40

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 4: DER — DER-AVM 매칭
    # ═══════════════════════════════════════════════════════════════════
    if device_type == 'inverter':
        ws_der = wb.create_sheet('4_DER')
        ws_der['A1'] = f'DER-AVM 매칭 — {der_matched}/{der_total} (Sheet 4)'
        ws_der['A1'].font = title_font

        der_cols = ['No', 'Field', 'Type', 'Status', 'Address', 'Scale', 'R/W', 'Description']
        _write_header(ws_der, der_cols, 3)

        for i, rd in enumerate(der_match_table, start=1):
            vals = [i, rd['name'], rd['type'], rd['status'],
                    rd['address'], rd['scale'], rd['rw'], rd['desc']]
            sc = MATCH_COLORS.get(rd['status'], 'FFFFFF')
            cat_color = CATEGORY_COLORS['DER_CONTROL'] if rd['group'] == 'CONTROL' else CATEGORY_COLORS['DER_MONITOR']
            for j, val in enumerate(vals, start=1):
                cell = ws_der.cell(row=3 + i, column=j, value=val)
                cell.border = thin_border
                if j == 4:
                    cell.fill = PatternFill('solid', fgColor=sc)
                else:
                    cell.fill = PatternFill('solid', fgColor=cat_color)

        ws_der.column_dimensions['B'].width = 35
        ws_der.column_dimensions['E'].width = 15
        ws_der.column_dimensions['H'].width = 45

    # ═══════════════════════════════════════════════════════════════════
    # Sheet 5: IV — IV 스캔 매핑 (지원 시)
    # ═══════════════════════════════════════════════════════════════════
    if meta['iv_scan'] and iv_info.get('supported'):
        ws_iv = wb.create_sheet('5_IV')
        ws_iv['A1'] = f'IV Scan 매핑 — data_points={iv_info["data_points"]}'
        ws_iv['A1'].font = title_font

        iv_fill = PatternFill('solid', fgColor=CATEGORY_COLORS['IV_SCAN'])
        ctrl_fill = PatternFill('solid', fgColor='D9D2E9')

        # IV 메타 정보
        iv_meta = [
            ('IV Scan Command', meta.get('iv_command_addr', '-')),
            ('Data Points', iv_info['data_points']),
            ('Trackers', len(iv_info['trackers'])),
            ('Strings/Tracker', max(len(t['strings']) for t in iv_info['trackers']) if iv_info['trackers'] else 0),
        ]
        for i, (k, v) in enumerate(iv_meta, start=3):
            ws_iv[f'A{i}'] = k
            ws_iv[f'A{i}'].font = Font(bold=True)
            ws_iv[f'B{i}'] = str(v)

        # IV Command 레지스터
        cmd_start = len(iv_meta) + 5
        ws_iv.cell(row=cmd_start, column=1, value='IV Scan Command').font = section_font
        cmd_cols = ['Name', 'Address', 'Type', 'R/W', 'Description']
        _write_header(ws_iv, cmd_cols, cmd_start + 1)
        cmd_addr = iv_info.get('iv_command_addr')
        cmd_vals = ['IV_CURVE_SCAN', f'0x{cmd_addr:04X}' if cmd_addr else '-', 'U16', 'RW',
                    'W: 0=Stop, 1=Start / R: 0=Idle, 1=Running, 2=Finished']
        for j, val in enumerate(cmd_vals, start=1):
            cell = ws_iv.cell(row=cmd_start + 2, column=j, value=val)
            cell.border = thin_border
            cell.fill = ctrl_fill

        # Tracker/String 매핑 테이블
        map_start = cmd_start + 5
        ws_iv.cell(row=map_start, column=1, value='IV Data 레지스터 매핑').font = section_font
        map_cols = ['No', 'Type', 'Name', 'Address', 'Regs', 'Data Type', 'Scale']
        _write_header(ws_iv, map_cols, map_start + 1)

        row_idx = map_start + 2
        reg_no = 1
        for tracker in iv_info['trackers']:
            tn = tracker['tracker_num']
            va = tracker['voltage_addr']
            # Tracker voltage
            vals = [reg_no, 'Tracker Voltage', f'Tracker {tn} voltage',
                    f'0x{va:04X}' if va else '-', iv_info['data_points'], 'U16', '0.1V']
            for j, val in enumerate(vals, start=1):
                cell = ws_iv.cell(row=row_idx, column=j, value=val)
                cell.border = thin_border
                cell.fill = PatternFill('solid', fgColor='D5F5E3')  # 연한 초록
            row_idx += 1
            reg_no += 1

            # String currents
            for sc in tracker['strings']:
                sn = sc['string_num']
                sa = sc['current_addr']
                vals = [reg_no, 'String Current', f'String {tn}-{sn} current',
                        f'0x{sa:04X}', iv_info['data_points'], 'S16', '0.01A']
                for j, val in enumerate(vals, start=1):
                    cell = ws_iv.cell(row=row_idx, column=j, value=val)
                    cell.border = thin_border
                    cell.fill = iv_fill
                row_idx += 1
                reg_no += 1

        ws_iv.column_dimensions['A'].width = 6
        ws_iv.column_dimensions['B'].width = 18
        ws_iv.column_dimensions['C'].width = 25
        ws_iv.column_dimensions['D'].width = 12
        ws_iv.column_dimensions['E'].width = 10

    wb.save(output_path)
    wb.close()

    sheet_count = 4 + (1 if meta['iv_scan'] and iv_info.get('supported') else 0)
    log(f'Stage 1 완료: {output_name} ({sheet_count}시트)', 'ok')

    return {
        'output_path': output_path,
        'output_name': output_name,
        'counts': counts,
        'meta': meta,
        'review_count': counts.get('REVIEW', 0),
        'h01_match': {'matched': h01_matched, 'total': h01_total, 'table': h01_match_table},
        'der_match': {'matched': der_matched, 'total': der_total},
        'iv_info': iv_info,
        'suggestions': suggestions,
    }


def _build_all_suggestions(h01_table: list, categorized: dict,
                           all_regs: list, meta: dict, log=None) -> dict:
    """
    5가지 섹션에 대한 후보 제안 통합
    1. INFO: Model/SN 미매칭
    2. H01: X 필드 (pv_power, cumulative_energy 등)
    3. STATUS/ALARM: 매칭 안 됐거나 정의 없음
    4. MPPT/String: 미감지
    5. IV Scan: 미감지
    """
    suggestions = {}

    # ── 1. INFO: Model/SN 미매칭 ──
    info_regs = categorized.get('INFO', [])
    has_model = any(
        any(k in r.definition.lower() for k in ['model', 'device type', 'type code'])
        and not any(k in r.definition.lower() for k in ['working', 'battery', 'pf', 'init fault'])
        for r in info_regs)
    has_sn = any(
        any(k in r.definition.lower() for k in ['serial_number', 'serial n', 'serialn', 'c_serial',
                                                  'sn[', 'sn0', 'sn1', 'serial no', 'inverter sn',
                                                  'product code', 'c_serialnumber'])
        for r in info_regs)

    if not has_model:
        cands = _suggest_info_field(all_regs, 'model')
        if cands:
            suggestions['info_model'] = cands
            if log: log(f'  제안: INFO Model 후보 {len(cands)}개')

    if not has_sn:
        cands = _suggest_info_field(all_regs, 'sn')
        if cands:
            suggestions['info_sn'] = cands
            if log: log(f'  제안: INFO SN 후보 {len(cands)}개')

    # ── 2. H01: X 필드 ──
    x_fields = [row for row in h01_table if row['status'] == 'X']
    for x_row in x_fields:
        candidates = _suggest_candidates(x_row['field'], all_regs, categorized)
        if candidates:
            suggestions[x_row['field']] = candidates
            if log: log(f'  제안: H01 [{x_row["field"]}] 후보 {len(candidates)}개')

    # ── 3. MPPT/String 미감지 ──
    max_mppt = meta.get('max_mppt', 0)
    max_string = meta.get('max_string', 0)
    if max_mppt == 0:
        cands = _suggest_mppt(all_regs)
        if cands:
            suggestions['mppt_detection'] = cands
            if log: log(f'  제안: MPPT 후보 {len(cands)}개')

    # ── 5. IV Scan ──
    iv_scan = meta.get('iv_scan', False)
    has_iv_regs = bool(categorized.get('IV_SCAN'))
    if not iv_scan and not has_iv_regs:
        cands = _suggest_iv_scan(all_regs)
        if cands:
            suggestions['iv_scan'] = cands
            if log: log(f'  제안: IV Scan 후보 {len(cands)}개')

    return suggestions


def _group_consecutive_regs(matching_regs: list) -> list:
    """
    주소가 연속인 레지스터를 그룹으로 묶어 반환.
    각 그룹: {'regs': [reg, ...], 'start': addr, 'count': N}
    정렬은 주소 기준.
    """
    if not matching_regs:
        return []
    sorted_regs = sorted(matching_regs, key=lambda r: r.address if isinstance(r.address, int) else 0)
    groups = []
    cur = [sorted_regs[0]]
    for reg in sorted_regs[1:]:
        prev_addr = cur[-1].address if isinstance(cur[-1].address, int) else -999
        cur_addr  = reg.address if isinstance(reg.address, int) else -998
        if cur_addr == prev_addr + 1:
            cur.append(reg)
        else:
            groups.append(cur)
            cur = [reg]
    groups.append(cur)
    return [{'regs': g, 'start': g[0].address, 'count': len(g)} for g in groups]


def _suggest_info_field(all_regs: list, field_type: str) -> list:
    """INFO Model/SN 후보 제안.
    - 연속 레지스터 그룹(SN[1]~SN[12] 등)은 하나의 후보로 묶어 표시
    - 2개 미만이면 보조 키워드로 보충
    """
    if field_type == 'model':
        prim_kws  = ['model', 'device type', 'type code', 'product']
        prim_excl = ['working', 'battery', 'pf']
        str_kws   = ['model', 'product']
        sec_kws   = ['type', 'rated', 'name', 'equip', 'identifier', 'kind']
    else:  # sn
        prim_kws  = ['serial', 'sn']
        prim_excl = []
        str_kws   = ['serial', 'sn']
        sec_kws   = ['number', 'id', 'code', 'uid', 'barcode', 'lot']

    def _score_single(reg, kws, excl, str_kws):
        dl = reg.definition.lower()
        if not any(k in dl for k in kws):
            return 0, ''
        if any(k in dl for k in excl):
            return 0, ''
        score = 70
        reason = f'{field_type.upper()} 키워드'
        if (reg.data_type or '').upper() in ('STRING', 'STRINGING', 'ASCII'):
            if any(k in dl for k in str_kws):
                score, reason = 85, f'{field_type.upper()} + STRING 타입'
        return score, reason

    # 1차: 주 키워드 매칭
    primary_matches = []
    for reg in all_regs:
        if not isinstance(reg.address, int):
            continue
        score, reason = _score_single(reg, prim_kws, prim_excl, str_kws)
        if score > 0:
            primary_matches.append((reg, score, reason))

    # 연속 주소 그룹으로 묶기
    groups = _group_consecutive_regs([r for r, _, _ in primary_matches])
    score_map = {id(r): (sc, rs) for r, sc, rs in primary_matches}

    candidates = []
    used_addrs = set()
    for g in groups:
        first = g['regs'][0]
        count = g['count']
        # 그룹 내 최고 점수 사용, 그룹이면 보너스 +5
        best_sc = max(score_map[id(r)][0] for r in g['regs'])
        best_rs = score_map[id(first)][1]
        if count > 1:
            best_sc = min(100, best_sc + 5)
            end_addr = g['start'] + count - 1
            addr_str = f'0x{g["start"]:04X}~0x{end_addr:04X} (×{count}regs)'
            defn_str = f'{first.definition[:30]}...[×{count}]'
            best_rs  = f'{field_type.upper()} 연속 {count}개 레지스터'
        else:
            addr_str = first.address_hex
            defn_str = first.definition[:50]
        candidates.append({
            'addr':      addr_str,
            'definition': defn_str,
            'unit':      first.unit or '',
            'score':     best_sc,
            'reason':    best_rs,
            'source':    'PDF',
            'reg_count': count,
            'reg_start': g['start'],
        })
        for r in g['regs']:
            used_addrs.add(r.address)

    # 2개 미만이면 보조 키워드로 보충
    if len(candidates) < 2:
        for reg in all_regs:
            if not isinstance(reg.address, int) or reg.address in used_addrs:
                continue
            dl = reg.definition.lower()
            if any(k in dl for k in sec_kws) and not any(k in dl for k in prim_excl):
                candidates.append({
                    'addr':       reg.address_hex,
                    'definition': reg.definition[:50],
                    'unit':       reg.unit or '',
                    'score':      40,
                    'reason':     f'{field_type.upper()} 보조 키워드',
                    'source':     'PDF',
                    'reg_count':  1,
                    'reg_start':  reg.address,
                })
                used_addrs.add(reg.address)
                if len(candidates) >= 2:
                    break

    return sorted(candidates, key=lambda c: -c['score'])[:5]


def _suggest_mppt(all_regs: list) -> list:
    """MPPT 후보 제안 (MPPT=0일 때)"""
    candidates = []
    for reg in all_regs:
        dl = reg.definition.lower()
        score = 0
        reason = ''

        # DC/PV/Input voltage/current
        if re.search(r'\b(dc|pv|input)\s*(voltage|current|power)', dl):
            if 'fault' not in dl and 'high' not in dl and 'low' not in dl:
                score = 60
                reason = 'DC/PV/Input V/I/P (Central type MPPT=1)'

        # Numbered PV: PV1, Vpv1, MPPT1 등
        if re.search(r'(pv|mppt|vpv|ppv|ipv)\s*[1-9]', dl):
            score = 80
            reason = 'PV/MPPT 번호 감지'

        if score > 0:
            candidates.append({
                'addr': reg.address_hex,
                'definition': reg.definition[:50],
                'unit': reg.unit or '',
                'score': score,
                'reason': reason,
                'source': 'PDF',
            })

    return sorted(candidates, key=lambda c: -c['score'])[:8]


def _suggest_iv_scan(all_regs: list) -> list:
    """IV Scan 후보 제안"""
    candidates = []
    for reg in all_regs:
        dl = reg.definition.lower()
        if any(k in dl for k in ['iv curve', 'iv scan', 'i-v curve', 'i-v scan',
                                   'iv_curve', 'iv data', 'iv point']):
            candidates.append({
                'addr': reg.address_hex,
                'definition': reg.definition[:50],
                'unit': reg.unit or '',
                'score': 80,
                'reason': 'IV Scan 키워드',
                'source': 'PDF',
            })
    return sorted(candidates, key=lambda c: -c['score'])[:5]


def _suggest_candidates(x_field: str, all_regs: list, categorized: dict) -> list:
    """
    X 필드에 대한 후보 레지스터 제안
    Returns: [{'addr': hex, 'definition': str, 'score': int, 'reason': str, 'source': str}, ...]
    """
    candidates = []

    # 필드별 검색 키워드 + 단위 + 타입 힌트
    _FIELD_HINTS = {
        'pv_power': {
            'keywords': ['dc power', 'pv power', 'input power', 'total power', 'pac',
                         'output power', 'active power', 'i dc power'],
            'unit': 'W', 'type_pref': ['S32', 'U32', 'S16', 'U16'],
            'handler': 'sum(MPPT_N_power) — MPPT power 합산으로 handler 계산',
        },
        'cumulative_energy': {
            'keywords': ['total energy', 'cumulative', 'lifetime energy', 'eac total',
                         'einv all', 'energy total', 'energy since', 'energy yield',
                         'total power yields', 'total poweryields', 'ac energy',
                         'generate energy', 'generated energy', 'accumulated',
                         'production total', 'total production'],
            'unit_keywords': ['kwh', 'wh', 'mwh'],
            'unit': 'Wh', 'alt_units': ['kWh', 'MWh', 'KWH', '0.1kWH'],
            'type_pref': ['U32', 'S32'],
        },
        'status': {
            'keywords': ['inverter mode', 'work mode', 'work state', 'operating mode',
                         'operation state', 'operating status', 'running status',
                         'device status', 'inverter status', 'system status',
                         'inverter state', 'working mode', 'run state', 'operational'],
            'exact': ['state', 'running', 'mode', 'status'],
            'type_pref': ['U16'],
        },
        'alarm1': {
            'keywords': ['fault code', 'error code', 'alarm code', 'faultcode',
                         'warningcode', 'warning code', 'fault status',
                         'fault register', 'alarm register', 'error register',
                         'fault', 'alarm', 'error', 'warning'],
            'type_pref': ['U16', 'U32'],
        },
    }

    # MPPT/String X → 기본 후보 없음 (패턴 감지 문제)
    if 'mppt' in x_field or 'string' in x_field:
        # MPPT/String 번호 추출
        m = re.search(r'(\d+)', x_field)
        n = int(m.group(1)) if m else 0
        is_voltage = 'voltage' in x_field
        is_current = 'current' in x_field

        for reg in all_regs:
            dl = reg.definition.lower()
            score = 0
            reason = ''

            if is_voltage:
                if any(k in dl for k in ['volt', 'vpv', 'v pv', 'dc volt']):
                    if str(n) in dl or f'pv{n}' in dl or f'input {n}' in dl:
                        score = 80
                        reason = f'PV{n} voltage 후보'
            elif is_current:
                if any(k in dl for k in ['curr', 'ipv', 'i pv', 'dc curr']):
                    if str(n) in dl or f'pv{n}' in dl or f'input {n}' in dl:
                        score = 80
                        reason = f'PV{n} current 후보'

            if score > 0:
                candidates.append({
                    'addr': reg.address_hex,
                    'definition': reg.definition[:50],
                    'unit': reg.unit or '',
                    'score': score,
                    'reason': reason,
                    'source': 'PDF',
                })

        # handler 계산 후보 (current = power/voltage)
        if is_current:
            candidates.append({
                'addr': '-',
                'definition': f'mppt{n}_power / mppt{n}_voltage',
                'unit': 'A',
                'score': 60,
                'reason': 'handler 계산 (P/V=I)',
                'source': 'HANDLER',
            })
        return sorted(candidates, key=lambda c: -c['score'])[:5]

    # 일반 필드 (pv_power, cumulative_energy, status, alarm1)
    base_field = x_field.replace('1', '').replace('2', '').replace('3', '')
    hints = _FIELD_HINTS.get(base_field, _FIELD_HINTS.get(x_field, {}))
    if not hints:
        return []

    keywords = hints.get('keywords', [])
    exact_matches = hints.get('exact', [])
    expected_unit = hints.get('unit', '')
    alt_units = hints.get('alt_units', [])

    for reg in all_regs:
        dl = reg.definition.lower().replace('_', ' ')
        dl_nospace = dl.replace(' ', '')
        score = 0
        reason = ''

        # 키워드 매칭
        for kw in keywords:
            if kw in dl or kw.replace(' ', '') in dl_nospace:
                score = max(score, 70)
                reason = f'키워드 "{kw}" 매칭'
                break

        # 정확한 이름 매칭
        for ex in exact_matches:
            if dl.strip() == ex:
                score = max(score, 60)
                reason = f'정확한 이름 "{ex}"'
                break

        # 단위 기반 후보 (cumulative_energy: kWh 단위면 후보)
        ru = (reg.unit or '').strip()
        unit_keywords = hints.get('unit_keywords', [])
        if unit_keywords and ru and ru.lower() in unit_keywords and score == 0:
            if 'energy' in dl or 'total' in dl or 'accum' in dl or 'yield' in dl:
                score = 55
                reason = f'단위 {ru} + 이름 조합'

        # 단위 보너스
        if expected_unit and ru:
            if ru == expected_unit or ru in alt_units:
                score += 15
                reason += f' +단위({ru})'

        # 타입 보너스
        dt = (reg.data_type or '').upper()
        if dt in hints.get('type_pref', []):
            score += 5

        # 이미 h01_field 할당된 레지스터는 감점
        if getattr(reg, 'h01_field', ''):
            score -= 30

        if score >= 40:
            candidates.append({
                'addr': reg.address_hex,
                'definition': reg.definition[:50],
                'unit': ru,
                'score': min(score, 100),
                'reason': reason,
                'source': 'PDF',
            })

    # handler 후보 (pv_power)
    if 'handler' in hints:
        candidates.append({
            'addr': '-',
            'definition': hints['handler'],
            'unit': expected_unit,
            'score': 50,
            'reason': 'handler 계산',
            'source': 'HANDLER',
        })

    # 점수순 정렬, 상위 5개
    return sorted(candidates, key=lambda c: -c['score'])[:5]


def _link_definitions_to_registers(categorized: dict, def_tables: dict):
    """정의 테이블을 STATUS/ALARM 레지스터에 연결"""
    # STATUS: h01_field='inverter_status'인 레지스터에 가장 적합한 mode_table 연결 (제어 레지스터 제외)
    status_reg = None
    for reg in categorized.get('STATUS', []):
        if getattr(reg, 'h01_field', '') == 'inverter_status' and not _is_control_reg(reg):
            status_reg = reg
            break
    if status_reg and not getattr(status_reg, 'value_definitions', None):
        addr = status_reg.address if isinstance(status_reg.address, int) else 0
        best_def = None
        for d in def_tables['status_defs']:
            if d['type'] != 'mode_table':
                continue
            # 1) 주소 매칭 → 최우선
            if d.get('address') and d['address'] == addr:
                best_def = d
                break
            # 2) 가장 적합한 mode_table (status 키워드 포함 + 값 3개 이상)
            vals = d.get('values', {})
            if len(vals) >= 3:
                has_status_kw = any(any(k in str(v).lower() for k in
                                        ['standby', 'fault', 'running', 'wait', 'off',
                                         'sleep', 'check', 'mppt', 'normal', 'initial',
                                         'no response'])
                                    for v in vals.values())
                if has_status_kw:
                    if best_def is None or len(vals) > len(best_def.get('values', {})):
                        best_def = d
        if best_def:
            status_reg.value_definitions = best_def['values']

    # ALARM: alarm1 (첫 번째 ALARM) 레지스터에 가장 큰 bitfield/fault_codes 연결
    alarm_regs = categorized.get('ALARM', [])
    if alarm_regs and def_tables['alarm_defs']:
        # bitfield를 우선, 없으면 fault_codes 중 가장 큰 것
        best_bitfield = None
        best_fault = None
        for d in def_tables['alarm_defs']:
            vals = d.get('values', {})
            if d['type'] == 'bitfield' and len(vals) >= 3:
                if best_bitfield is None or len(vals) > len(best_bitfield.get('values', {})):
                    best_bitfield = d
            elif d['type'] == 'fault_codes' and len(vals) >= 3:
                if best_fault is None or len(vals) > len(best_fault.get('values', {})):
                    best_fault = d

        best_alarm_def = best_bitfield or best_fault
        if best_alarm_def:
            # alarm1에 연결 (제어 레지스터 제외)
            for reg in alarm_regs:
                if not getattr(reg, 'value_definitions', None) and not _is_control_reg(reg):
                    reg.value_definitions = best_alarm_def['values']
                    break  # 첫 번째 ALARM만


def _build_der_match_table(categorized: dict) -> List[dict]:
    """DER-AVM 매칭 테이블 — 고정 주소맵 기반"""
    rows = []
    for dr in DER_CONTROL_REGS:
        rows.append({
            'name': dr['name'], 'type': dr['type'],
            'address': f'0x{dr["addr"]:04X}', 'scale': dr.get('scale', ''),
            'rw': dr['rw'], 'desc': dr['desc'],
            'status': 'O', 'group': 'CONTROL',
        })
    for dr in DER_MONITOR_REGS:
        rows.append({
            'name': dr['name'], 'type': dr['type'],
            'address': f'0x{dr["addr"]:04X}', 'scale': dr.get('scale', ''),
            'rw': 'RO', 'desc': dr.get('desc', ''),
            'status': 'O', 'group': 'MONITOR',
        })
    return rows
