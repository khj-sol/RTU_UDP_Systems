# -*- coding: utf-8 -*-
"""
REST API + WebSocket 라우트 — 파이프라인 연결
"""
from fastapi import APIRouter, UploadFile, File, Form, WebSocket, WebSocketDisconnect, HTTPException
from fastapi.responses import FileResponse
import os
import shutil
import asyncio
import traceback
import json

from .session_store import SessionStore
from .ws_manager import ws_manager

router = APIRouter()

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
COMMON_DIR = os.path.join(PROJECT_ROOT, 'common')


# ─── 비동기 파이프라인 실행 헬퍼 ─────────────────────────────────────────────

async def _run_in_thread(func, *args, **kwargs):
    """동기 함수를 스레드 풀에서 실행"""
    loop = asyncio.get_running_loop()
    return await loop.run_in_executor(None, lambda: func(*args, **kwargs))


def _make_progress_callback(sid: str, stage: str, loop: asyncio.AbstractEventLoop):
    """WebSocket 로그 전송 콜백 생성 — loop을 캡처하여 스레드 안전"""
    def callback(msg: str, level: str = 'info'):
        loop.call_soon_threadsafe(
            asyncio.ensure_future,
            ws_manager.send_json(sid, {
                'stage': stage,
                'text': msg,
                'level': level,
            })
        )
    return callback


# ─── 세션 ────────────────────────────────────────────────────────────────────

@router.post('/session/new')
def new_session():
    sid = SessionStore.create()
    return {'session_id': sid}


@router.get('/session/{sid}')
def get_session(sid: str):
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')
    return s


# ─── Stage 1 ─────────────────────────────────────────────────────────────────

@router.post('/stage1/upload')
async def stage1_upload(
    session_id: str = Form(...),
    file: UploadFile = File(...),
):
    """PDF 또는 Excel 업로드 → 작업 디렉토리에 저장"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    work_dir = SessionStore.get_work_dir(session_id)
    dest = os.path.join(work_dir, file.filename)
    with open(dest, 'wb') as f:
        shutil.copyfileobj(file.file, f)

    # 새 파일 업로드 → 진행 중 태스크 취소 + 세션 상태 완전 리셋
    SessionStore.cancel_running_task(session_id)
    SessionStore.update(
        session_id,
        uploaded_file=dest,
        stage=0,
        stage1_excel=None,
        stage2_excel=None,
        registers_py=None,
        meta={},
    )
    return {'saved': dest, 'filename': file.filename}


@router.post('/stage1/run')
async def stage1_run(body: dict):
    """
    Stage 1 실행
    body: {session_id, filename, device_type}
    """
    sid = body.get('session_id')
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')

    device_type = body.get('device_type', 'inverter')
    uploaded = s.get('uploaded_file')
    if not uploaded or not os.path.exists(uploaded):
        raise HTTPException(400, 'no file uploaded')

    work_dir = SessionStore.get_work_dir(sid)

    # 이전 태스크 취소 후 새 태스크 실행
    SessionStore.cancel_running_task(sid)

    async def _run():
        try:
            from .pipeline.stage1 import run_stage1, NotRegisterMapError
            progress = _make_progress_callback(sid, 's1', asyncio.get_running_loop())
            result = await _run_in_thread(
                run_stage1, uploaded, work_dir, device_type, progress)

            # 태스크 취소로 인해 세션이 리셋된 경우 이벤트 전송 생략
            if SessionStore.get(sid) is None:
                return

            SessionStore.update(sid,
                                stage=1,
                                stage1_excel=result['output_path'],
                                meta=result['meta'],
                                suggestions=result.get('suggestions', {}))

            await ws_manager.send_json(sid, {
                'event': 'stage1_done',
                'stage': 's1',
                'text': f'Stage 1 완료: {result["output_name"]}',
                'level': 'ok',
                'counts': result['counts'],
                'review_count': result['review_count'],
                'iv_scan': result['meta'].get('iv_scan', False),
                'iv_data_points': result['meta'].get('iv_data_points', 0),
                'suggestions': result.get('suggestions', {}),
            })
        except asyncio.CancelledError:
            pass  # 새 파일 업로드로 인해 취소됨 — 정상
        except NotRegisterMapError as e:
            await ws_manager.send_json(sid, {
                'event': 'invalid_file',
                'stage': 's1',
                'text': str(e),
                'level': 'err',
            })
        except Exception as e:
            await ws_manager.send_json(sid, {
                'stage': 's1',
                'text': f'오류: {str(e)}',
                'level': 'err',
            })
            traceback.print_exc()

    task = asyncio.ensure_future(_run())
    SessionStore.update(sid, _running_task=task)
    return {'status': 'running', 'session_id': sid}


@router.post('/stage1/apply-suggestion')
async def apply_suggestion(body: dict):
    """
    Stage 1 suggestion 적용 → synonym_db에 등록 → Stage 1 재실행
    body: {session_id, field, definition, addr, rerun}
    """
    sid = body.get('session_id')
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')

    field = body.get('field', '')
    definition = body.get('definition', '')
    addr = body.get('addr', '')
    rerun = body.get('rerun', True)

    # synonym_db에 (definition → field) 매핑 추가
    if field and definition:
        from .pipeline import load_synonym_db, save_synonym_db
        db = load_synonym_db()
        fields = db.setdefault('fields', {})
        synonyms = fields.setdefault(field, [])
        key = definition.upper().replace(' ', '_')
        if key not in synonyms:
            synonyms.append(key)
            save_synonym_db(db)

    if not rerun:
        return {'status': 'ok', 'field': field, 'added': definition}

    # Stage 1 재실행
    uploaded = s.get('uploaded_file')
    device_type = s.get('meta', {}).get('device_type', 'inverter')
    if not uploaded or not os.path.exists(uploaded):
        raise HTTPException(400, 'no file uploaded')

    work_dir = SessionStore.get_work_dir(sid)

    async def _rerun():
        try:
            from .pipeline.stage1 import run_stage1, NotRegisterMapError
            loop = asyncio.get_running_loop()
            progress = _make_progress_callback(sid, 's1', loop)
            progress(f'[제안 적용] {field} → {definition}', 'info')
            result = await _run_in_thread(run_stage1, uploaded, work_dir, device_type, progress)

            SessionStore.update(sid,
                                stage=1,
                                stage1_excel=result['output_path'],
                                meta=result['meta'],
                                suggestions=result.get('suggestions', {}))

            await ws_manager.send_json(sid, {
                'event': 'stage1_done',
                'stage': 's1',
                'text': f'Stage 1 재실행 완료: {result["output_name"]}',
                'level': 'ok',
                'counts': result['counts'],
                'review_count': result['review_count'],
                'iv_scan': result['meta'].get('iv_scan', False),
                'iv_data_points': result['meta'].get('iv_data_points', 0),
                'suggestions': result.get('suggestions', {}),
            })
        except NotRegisterMapError as e:
            await ws_manager.send_json(sid, {
                'event': 'invalid_file',
                'stage': 's1',
                'text': str(e),
                'level': 'err',
            })
        except Exception as e:
            await ws_manager.send_json(sid, {
                'stage': 's1',
                'text': f'재실행 오류: {str(e)}',
                'level': 'err',
            })
            import traceback; traceback.print_exc()

    asyncio.ensure_future(_rerun())
    return {'status': 'rerunning', 'session_id': sid}


# ─── Stage 2 ─────────────────────────────────────────────────────────────────

@router.post('/stage2/upload')
async def stage2_upload(
    session_id: str = Form(...),
    file: UploadFile = File(...),
):
    """Stage 1 Excel 직접 업로드 → PDF 없이 Stage 2 시작"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    fname = file.filename or ''
    if not fname.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, 'Stage 1 Excel (.xlsx) 파일만 업로드 가능합니다')

    work_dir = SessionStore.get_work_dir(session_id)
    dest = os.path.join(work_dir, fname)
    with open(dest, 'wb') as f:
        shutil.copyfileobj(file.file, f)

    SessionStore.update(session_id,
                        stage1_excel=dest,
                        stage=1,
                        stage2_excel=None,
                        registers_py=None)
    return {'saved': dest, 'filename': fname}


@router.post('/stage2/run')
async def stage2_run(body: dict):
    """
    body: {session_id, mppt, strings_per_mppt, capacity}
    """
    sid = body.get('session_id')
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')

    stage1_excel = s.get('stage1_excel')
    if not stage1_excel or not os.path.exists(stage1_excel):
        raise HTTPException(400, 'Stage 1 not completed')

    mppt = int(body.get('mppt', 4))
    strings_per_mppt = int(body.get('strings_per_mppt', 2))
    capacity = body.get('capacity', '')
    iv_data_points = int(body.get('iv_data_points', 64))
    work_dir = SessionStore.get_work_dir(sid)
    SessionStore.update(sid, iv_data_points=iv_data_points)

    async def _run():
        try:
            from .pipeline.stage2 import run_stage2
            progress = _make_progress_callback(sid, 's2', asyncio.get_running_loop())
            result = await _run_in_thread(
                run_stage2, stage1_excel, work_dir,
                mppt, strings_per_mppt, capacity, progress)

            SessionStore.update(sid,
                                stage=2,
                                stage2_excel=result['output_path'])

            await ws_manager.send_json(sid, {
                'event': 'stage2_done',
                'stage': 's2',
                'text': f'Stage 2 완료: {result["output_name"]}',
                'level': 'ok',
                'h01_matched': result['h01_matched'],
                'h01_total': result['h01_total'],
                'der_matched': result['der_matched'],
                'der_total': result['der_total'],
                'review_count': result['review_count'],
                'review_items': result.get('review_items', []),
            })
        except Exception as e:
            await ws_manager.send_json(sid, {
                'stage': 's2',
                'text': f'오류: {str(e)}',
                'level': 'err',
            })
            traceback.print_exc()

    asyncio.ensure_future(_run())
    return {'status': 'running', 'session_id': sid}


# ─── Stage 3 ─────────────────────────────────────────────────────────────────

@router.post('/stage3/run')
async def stage3_run(body: dict):
    """
    body: {session_id}
    """
    sid = body.get('session_id')
    s = SessionStore.get(sid)
    if not s:
        raise HTTPException(404, 'session not found')

    stage2_excel = s.get('stage2_excel')
    if not stage2_excel or not os.path.exists(stage2_excel):
        raise HTTPException(400, 'Stage 2 not completed')

    work_dir = SessionStore.get_work_dir(sid)
    iv_data_points = s.get('iv_data_points', 64)

    async def _run():
        try:
            from .pipeline.stage3 import run_stage3
            progress = _make_progress_callback(sid, 's3', asyncio.get_running_loop())
            result = await _run_in_thread(
                run_stage3, stage2_excel, work_dir, progress, iv_data_points)

            SessionStore.update(sid,
                                stage=3,
                                registers_py=result['output_path'])

            await ws_manager.send_json(sid, {
                'event': 'stage3_done',
                'stage': 's3',
                'text': f'Stage 3 완료: {result["filename"]}',
                'level': 'ok',
                'filename': result['filename'],
                'validation': result['validation'],
                'synonym_added': result['synonym_added'],
                'review_recorded': result['review_recorded'],
            })
        except Exception as e:
            await ws_manager.send_json(sid, {
                'stage': 's3',
                'text': f'오류: {str(e)}',
                'level': 'err',
            })
            traceback.print_exc()

    asyncio.ensure_future(_run())
    return {'status': 'running', 'session_id': sid}


# ─── 파일 다운로드 ───────────────────────────────────────────────────────────

@router.get('/download/{session_id}/{filename}')
def download_file(session_id: str, filename: str):
    work_dir = SessionStore.get_work_dir(session_id)
    path = os.path.join(work_dir, filename)
    if not os.path.exists(path):
        raise HTTPException(404, 'file not found')
    return FileResponse(path, filename=filename)


@router.get('/download-stage/{session_id}/{stage}')
def download_stage_file(session_id: str, stage: int):
    """Stage별 결과 파일 다운로드"""
    s = SessionStore.get(session_id)
    if not s:
        raise HTTPException(404, 'session not found')

    if stage == 1:
        path = s.get('stage1_excel')
    elif stage == 2:
        path = s.get('stage2_excel')
    elif stage == 3:
        path = s.get('registers_py')
    else:
        raise HTTPException(400, 'invalid stage')

    if not path or not os.path.exists(path):
        raise HTTPException(404, 'file not found')
    return FileResponse(path, filename=os.path.basename(path))


# ─── WebSocket ───────────────────────────────────────────────────────────────

@router.websocket('/ws/{session_id}')
async def websocket_endpoint(ws: WebSocket, session_id: str):
    await ws_manager.connect(session_id, ws)
    try:
        while True:
            data = await ws.receive_text()
            # 클라이언트 메시지 처리 (REVIEW 판정 등)
            try:
                msg = json.loads(data)
                if msg.get('type') == 'review_verdicts':
                    # REVIEW 판정을 Stage 2 Excel에 반영
                    s = SessionStore.get(session_id)
                    if s and s.get('stage2_excel'):
                        _update_review_verdicts(s['stage2_excel'], msg.get('verdicts', []))
                        await ws.send_json({'event': 'verdicts_saved', 'count': len(msg.get('verdicts', []))})
            except (json.JSONDecodeError, KeyError):
                pass
    except WebSocketDisconnect:
        ws_manager.disconnect(session_id)


def _update_review_verdicts(excel_path: str, verdicts: list):
    """Stage 2 Excel REVIEW 시트에 사용자 판정 기록 + review_history 저장"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path)

        # V2: 2_REVIEW 또는 REVIEW 시트
        review_sheet = '2_REVIEW' if '2_REVIEW' in wb.sheetnames else 'REVIEW'
        if review_sheet not in wb.sheetnames:
            wb.close()
            return

        ws = wb[review_sheet]
        header = [str(c.value).strip() if c.value else '' for c in ws[1]]

        # '선택' 또는 '판정' 컬럼 인덱스
        verdict_col = None
        for i, h in enumerate(header):
            if '판정' in h or '선택' in h:
                verdict_col = i + 1
                break
        if verdict_col is None:
            wb.close()
            return

        # Definition 컬럼
        def_col = None
        addr_col = None
        for i, h in enumerate(header):
            if 'definition' in h.lower() or h == 'Definition':
                def_col = i + 1
            if 'address' in h.lower() or h == 'Address':
                addr_col = i + 1

        # 판정 기록 + review_history 수집
        history_entries = []
        for v in verdicts:
            row_idx = v.get('row', 0) + 4  # V2: 데이터는 row 4부터 (title + empty + header + data)
            verdict_val = v.get('verdict', '')
            # ALT1/ALT2 → KEEP으로 매핑 (추천 채택)
            write_val = verdict_val
            if verdict_val in ('ALT1', 'ALT2'):
                write_val = 'KEEP'  # 추천 채택 = 포함
            ws.cell(row=row_idx, column=verdict_col, value=write_val)

            # review_history에 저장할 데이터
            defn = str(ws.cell(row=row_idx, column=def_col).value or '') if def_col else ''
            addr = str(ws.cell(row=row_idx, column=addr_col).value or '') if addr_col else ''
            if defn and verdict_val == 'DELETE':
                history_entries.append({
                    'definition': defn.upper().replace(' ', '_'),
                    'address': addr,
                    'verdict': 'DELETE',
                })

        wb.save(excel_path)
        wb.close()

        # review_history에 DELETE 항목 저장 (다음번 자동 제외)
        if history_entries:
            from .pipeline import load_review_history, save_review_history
            history = load_review_history()
            existing = {item.get('definition', '') for item in history.get('approved', [])}
            added = 0
            for entry in history_entries:
                if entry['definition'] not in existing:
                    history['approved'].append(entry)
                    added += 1
            if added:
                history['stats']['total_reviewed'] = history['stats'].get('total_reviewed', 0) + added
                save_review_history(history)

    except Exception:
        pass
