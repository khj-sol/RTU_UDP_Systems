# -*- coding: utf-8 -*-
"""
API Routes - Model Maker Web v1.0.0

All REST API endpoints for the Model Maker web application.
Reuses stage_pipeline.py, ai_generator.py, reference_manager.py without modification.
"""

import os
import sys
import asyncio
import logging
import threading
import traceback
import json

from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
from typing import Optional

logger = logging.getLogger(__name__)

# Project root on sys.path (set by main.py)
router = APIRouter()

# Injected by main.py
ws_manager = None
store = None

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _project_root():
    """Return the project root (parent of model_maker_web/)."""
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def _import_pipeline():
    """Import stage_pipeline from model_maker/."""
    root = _project_root()
    if root not in sys.path:
        sys.path.insert(0, root)
    from model_maker import stage_pipeline
    return stage_pipeline


def _import_mapper():
    """Import parse_modbus_pdf from model_maker/modbus_to_udp_mapper."""
    root = _project_root()
    if root not in sys.path:
        sys.path.insert(0, root)
    from model_maker.modbus_to_udp_mapper import parse_modbus_pdf
    return parse_modbus_pdf


def _import_ai():
    """Import ai_generator from model_maker/."""
    root = _project_root()
    if root not in sys.path:
        sys.path.insert(0, root)
    from model_maker import ai_generator
    return ai_generator


def _import_ref():
    """Import reference_manager from model_maker/."""
    root = _project_root()
    if root not in sys.path:
        sys.path.insert(0, root)
    from model_maker import reference_manager
    return reference_manager


def _get_session(session_id: str):
    s = store.get(session_id)
    if s is None:
        raise HTTPException(status_code=404, detail=f"Session not found: {session_id}")
    return s


def _excel_to_json_rows(excel_path: str, sheet_name: str) -> list:
    """Read an Excel sheet and return rows as list of dicts."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
        result = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            result.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row)})
        return result
    except Exception as e:
        logger.error(f"Excel read error ({sheet_name}): {e}")
        return []


def _run_in_thread(fn, *args, **kwargs):
    """Run a blocking function in a thread and return a Future."""
    loop = asyncio.get_event_loop()
    future = loop.run_in_executor(None, lambda: fn(*args, **kwargs))
    return future


# ---------------------------------------------------------------------------
# Session Management
# ---------------------------------------------------------------------------

@router.post("/api/session/create")
async def create_session():
    """Create a new session."""
    s = store.create()
    return {"session_id": s.session_id}


# ---------------------------------------------------------------------------
# PDF Upload
# ---------------------------------------------------------------------------

@router.post("/api/upload-pdf")
async def upload_pdf(
    file: UploadFile = File(...),
    session_id: str = Query(...)
):
    """Upload a PDF file. Returns session_id and filename."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="PDF 파일만 업로드 가능합니다.")

    s = store.get_or_create(session_id)
    session_dir = s.session_dir()

    # Save uploaded file
    safe_name = os.path.basename(file.filename)
    pdf_path = os.path.join(session_dir, safe_name)
    contents = await file.read()
    with open(pdf_path, "wb") as f:
        f.write(contents)

    s.pdf_path = pdf_path
    s.pdf_filename = safe_name
    # Reset downstream data
    s.stage1_excel_path = None
    s.stage2_excel_path = None
    s.stage1_rows = []
    s.stage2_rows = []
    s.stage3_code = ""
    s.stage3_results = []

    logger.info(f"PDF uploaded: {safe_name} ({len(contents)} bytes) session={session_id}")
    return {
        "session_id": s.session_id,
        "filename": safe_name,
        "size": len(contents),
    }


# ---------------------------------------------------------------------------
# Stage 1
# ---------------------------------------------------------------------------

class Stage1Request(BaseModel):
    session_id: str
    mode: str = "offline"  # "offline" or "ai"


@router.post("/api/stage1/run")
async def stage1_run(req: Stage1Request):
    """Run Stage 1: PDF → Excel. Progress streamed via WebSocket."""
    s = _get_session(req.session_id)
    if s.pdf_path is None or not os.path.isfile(s.pdf_path):
        raise HTTPException(status_code=400, detail="PDF가 업로드되지 않았습니다.")
    if s.running:
        raise HTTPException(status_code=409, detail="이미 실행 중입니다.")

    s.running = True
    s.stage1_rows = []
    s.stage2_rows = []
    s.stage3_code = ""
    s.stage3_results = []

    async def _do():
        try:
            pipeline = _import_pipeline()
            parse_modbus_pdf = _import_mapper()
            ai_gen = _import_ai()

            session_dir = s.session_dir()
            pdf_path = s.pdf_path
            loop = asyncio.get_event_loop()

            async def progress(msg):
                await ws_manager.send_progress(s.session_id, "stage1", msg)

            await progress(f"PDF 파싱 시작: {s.pdf_filename}")

            # Run blocking parse in thread
            def _parse():
                try:
                    return parse_modbus_pdf(pdf_path), None
                except Exception as e:
                    return None, str(e)

            parsed_data, err = await loop.run_in_executor(None, _parse)
            if err:
                await ws_manager.send_error(s.session_id, "stage1", f"PDF 파싱 오류: {err}")
                s.running = False
                return

            reg_count = len(parsed_data.get("all_registers", [])) if parsed_data else 0
            await progress(f"PDF 파싱 완료: {reg_count}개 레지스터 추출")

            # Stage 1 Excel output path
            base = os.path.splitext(s.pdf_filename)[0]
            stage1_path = os.path.join(session_dir, base + "_registers.xlsx")

            # AI settings
            api_key = ""
            model = "claude-opus-4-6"
            if req.mode == "ai":
                api_key = ai_gen.load_api_key()
                model = ai_gen.load_model_name()
                if not api_key:
                    await ws_manager.send_error(s.session_id, "stage1",
                        "AI 모드를 사용하려면 AI 설정에서 API 키를 먼저 설정하세요.")
                    s.running = False
                    return

            progress_msgs = []

            def _progress_cb(msg):
                progress_msgs.append(msg)
                asyncio.run_coroutine_threadsafe(
                    ws_manager.send_progress(s.session_id, "stage1", msg),
                    loop
                )

            def _run_stage1():
                return pipeline.stage1_extract_to_excel(
                    parsed_data=parsed_data,
                    pdf_path=pdf_path,
                    output_path=stage1_path,
                    mode=req.mode,
                    api_key=api_key if req.mode == "ai" else None,
                    model=model,
                    progress_cb=_progress_cb,
                )

            result_path = await loop.run_in_executor(None, _run_stage1)

            s.stage1_excel_path = result_path
            # Cache rows
            s.stage1_rows = _excel_to_json_rows(result_path, "Registers")

            await ws_manager.send_done(s.session_id, "stage1", True,
                f"Stage 1 완료: {len(s.stage1_rows)}개 레지스터")

        except Exception as e:
            tb = traceback.format_exc()
            logger.error(f"Stage 1 error: {e}\n{tb}")
            await ws_manager.send_error(s.session_id, "stage1", str(e))
        finally:
            s.running = False

    asyncio.create_task(_do())
    return {"status": "started", "session_id": s.session_id}


@router.get("/api/stage1/result")
async def stage1_result(session_id: str = Query(...)):
    """Return Stage 1 result as JSON rows."""
    s = _get_session(session_id)
    if not s.stage1_excel_path:
        raise HTTPException(status_code=404, detail="Stage 1 결과가 없습니다.")
    if not s.stage1_rows:
        s.stage1_rows = _excel_to_json_rows(s.stage1_excel_path, "Registers")
    return {"rows": s.stage1_rows, "count": len(s.stage1_rows)}


@router.get("/api/stage1/download-excel")
async def stage1_download(session_id: str = Query(...)):
    """Download Stage 1 Excel file."""
    s = _get_session(session_id)
    if not s.stage1_excel_path or not os.path.isfile(s.stage1_excel_path):
        raise HTTPException(status_code=404, detail="Stage 1 Excel 파일이 없습니다.")
    return FileResponse(
        s.stage1_excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(s.stage1_excel_path),
    )


# ---------------------------------------------------------------------------
# Stage 2
# ---------------------------------------------------------------------------

class Stage2Request(BaseModel):
    session_id: str
    mode: str = "offline"
    mppt_count: int = 4
    string_count: int = 8


@router.post("/api/stage2/run")
async def stage2_run(req: Stage2Request):
    """Run Stage 2: Excel → Mapping Excel."""
    s = _get_session(req.session_id)
    if not s.stage1_excel_path or not os.path.isfile(s.stage1_excel_path):
        raise HTTPException(status_code=400, detail="Stage 1을 먼저 실행하세요.")
    if s.running:
        raise HTTPException(status_code=409, detail="이미 실행 중입니다.")

    s.running = True
    s.stage2_rows = []
    s.stage3_code = ""
    s.stage3_results = []

    async def _do():
        try:
            pipeline = _import_pipeline()
            ai_gen = _import_ai()

            session_dir = s.session_dir()
            loop = asyncio.get_event_loop()

            base = os.path.splitext(os.path.basename(s.stage1_excel_path))[0]
            stage2_path = os.path.join(session_dir, base + "_mapping.xlsx")

            api_key = ""
            model = "claude-opus-4-6"
            if req.mode == "ai":
                api_key = ai_gen.load_api_key()
                model = ai_gen.load_model_name()
                if not api_key:
                    await ws_manager.send_error(s.session_id, "stage2",
                        "AI 모드를 사용하려면 AI 설정에서 API 키를 먼저 설정하세요.")
                    s.running = False
                    return

            def _progress_cb(msg):
                asyncio.run_coroutine_threadsafe(
                    ws_manager.send_progress(s.session_id, "stage2", msg),
                    loop
                )

            def _run():
                return pipeline.stage2_create_mapping_excel(
                    stage1_excel_path=s.stage1_excel_path,
                    output_path=stage2_path,
                    mode=req.mode,
                    api_key=api_key if req.mode == "ai" else None,
                    model=model,
                    progress_cb=_progress_cb,
                )

            result_path = await loop.run_in_executor(None, _run)
            s.stage2_excel_path = result_path
            s.stage2_rows = _excel_to_json_rows(result_path, "Register_Mapping")

            await ws_manager.send_done(s.session_id, "stage2", True,
                f"Stage 2 완료: {len(s.stage2_rows)}개 매핑")

        except Exception as e:
            tb = traceback.format_exc()
            logger.error(f"Stage 2 error: {e}\n{tb}")
            await ws_manager.send_error(s.session_id, "stage2", str(e))
        finally:
            s.running = False

    asyncio.create_task(_do())
    return {"status": "started", "session_id": s.session_id}


@router.get("/api/stage2/result")
async def stage2_result(session_id: str = Query(...)):
    """Return Stage 2 mapping table as JSON."""
    s = _get_session(session_id)
    if not s.stage2_excel_path:
        raise HTTPException(status_code=404, detail="Stage 2 결과가 없습니다.")
    if not s.stage2_rows:
        s.stage2_rows = _excel_to_json_rows(s.stage2_excel_path, "Register_Mapping")
    return {"rows": s.stage2_rows, "count": len(s.stage2_rows)}


class Stage2UpdateRow(BaseModel):
    session_id: str
    row_index: int
    col_name: str
    value: str


@router.put("/api/stage2/update-row")
async def stage2_update_row(req: Stage2UpdateRow):
    """Inline edit: update a single cell in Stage 2 mapping Excel."""
    s = _get_session(req.session_id)
    if not s.stage2_excel_path or not os.path.isfile(s.stage2_excel_path):
        raise HTTPException(status_code=404, detail="Stage 2 Excel 파일이 없습니다.")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(s.stage2_excel_path)
        ws = wb["Register_Mapping"]
        # Find column index by header name (row 1)
        headers = [cell.value for cell in ws[1]]
        if req.col_name not in headers:
            raise HTTPException(status_code=400, detail=f"컬럼 없음: {req.col_name}")
        col_idx = headers.index(req.col_name) + 1
        excel_row = req.row_index + 2  # +1 for header, +1 for 0-based
        ws.cell(row=excel_row, column=col_idx, value=req.value)
        wb.save(s.stage2_excel_path)

        # Refresh cache
        s.stage2_rows = _excel_to_json_rows(s.stage2_excel_path, "Register_Mapping")
        return {"ok": True, "rows": len(s.stage2_rows)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/stage2/download-excel")
async def stage2_download(session_id: str = Query(...)):
    """Download Stage 2 mapping Excel."""
    s = _get_session(session_id)
    if not s.stage2_excel_path or not os.path.isfile(s.stage2_excel_path):
        raise HTTPException(status_code=404, detail="Stage 2 Excel 파일이 없습니다.")
    return FileResponse(
        s.stage2_excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(s.stage2_excel_path),
    )


# ---------------------------------------------------------------------------
# Stage 3
# ---------------------------------------------------------------------------

class Stage3Request(BaseModel):
    session_id: str
    mode: str = "offline"
    protocol_name: str = "custom"
    manufacturer: str = ""
    mppt_count: int = 4
    string_count: int = 8
    iv_scan: bool = False
    der_avm: bool = True
    dea_avm: bool = True
    class_name: str = "RegisterMap"
    fc_code: str = "FC03"


@router.post("/api/stage3/run")
async def stage3_run(req: Stage3Request):
    """Run Stage 3: Mapping Excel → registers.py with 12-item validation."""
    s = _get_session(req.session_id)
    if not s.stage2_excel_path or not os.path.isfile(s.stage2_excel_path):
        raise HTTPException(status_code=400, detail="Stage 2를 먼저 실행하세요.")
    if s.running:
        raise HTTPException(status_code=409, detail="이미 실행 중입니다.")

    s.running = True
    s.stage3_code = ""
    s.stage3_results = []

    async def _do():
        try:
            pipeline = _import_pipeline()
            ai_gen = _import_ai()

            session_dir = s.session_dir()
            loop = asyncio.get_event_loop()

            out_name = f"{req.protocol_name}_registers.py"
            output_path = os.path.join(session_dir, out_name)

            api_key = ""
            model = "claude-opus-4-6"
            if req.mode == "ai":
                api_key = ai_gen.load_api_key()
                model = ai_gen.load_model_name()
                if not api_key:
                    await ws_manager.send_error(s.session_id, "stage3",
                        "AI 모드를 사용하려면 AI 설정에서 API 키를 먼저 설정하세요.")
                    s.running = False
                    return

            settings = {
                "protocol_name": req.protocol_name,
                "manufacturer":  req.manufacturer,
                "mppt_count":    req.mppt_count,
                "string_count":  req.string_count,
                "iv_scan":       req.iv_scan,
                "der_avm":       req.der_avm,
                "dea_avm":       req.dea_avm,
                "class_name":    req.class_name,
                "fc_code":       req.fc_code,
            }

            def _progress_cb(msg):
                asyncio.run_coroutine_threadsafe(
                    ws_manager.send_progress(s.session_id, "stage3", msg),
                    loop
                )

            def _run():
                return pipeline.stage3_generate_py(
                    mapping_excel_path=s.stage2_excel_path,
                    settings=settings,
                    output_path=output_path,
                    mode=req.mode,
                    api_key=api_key if req.mode == "ai" else None,
                    model=model,
                    progress_cb=_progress_cb,
                )

            result = await loop.run_in_executor(None, _run)

            s.stage3_code = result.get("code", "")
            s.stage3_results = result.get("results", [])
            s.stage3_success = result.get("success", False)
            s.stage3_output_path = result.get("output_path")

            pass_count = sum(1 for r in s.stage3_results if r[0] == "PASS")
            total = len(s.stage3_results)
            await ws_manager.send_done(s.session_id, "stage3", s.stage3_success,
                f"검증 {pass_count}/{total} 통과")

        except Exception as e:
            tb = traceback.format_exc()
            logger.error(f"Stage 3 error: {e}\n{tb}")
            await ws_manager.send_error(s.session_id, "stage3", str(e))
        finally:
            s.running = False

    asyncio.create_task(_do())
    return {"status": "started", "session_id": s.session_id}


@router.get("/api/stage3/result")
async def stage3_result(session_id: str = Query(...)):
    """Return Stage 3 generated code and validation results."""
    s = _get_session(session_id)
    if not s.stage3_code:
        raise HTTPException(status_code=404, detail="Stage 3 결과가 없습니다.")
    return {
        "code": s.stage3_code,
        "results": s.stage3_results,
        "success": s.stage3_success,
    }


class Stage3SaveRequest(BaseModel):
    session_id: str
    protocol_name: str
    save_as_reference: bool = True


@router.post("/api/stage3/save")
async def stage3_save(req: Stage3SaveRequest):
    """Save generated registers.py to common/ directory."""
    s = _get_session(req.session_id)
    if not s.stage3_code:
        raise HTTPException(status_code=400, detail="Stage 3 코드가 없습니다.")

    root = _project_root()

    # Check for protected files
    protected = {
        "solarize_registers.py", "kstar_registers.py", "huawei_registers.py",
        "sungrow_registers.py", "ekos_registers.py", "goodwe_registers.py",
        "senergy_registers.py", "relay_registers.py", "weather_registers.py",
    }
    out_name = f"{req.protocol_name}_registers.py"
    if out_name in protected:
        raise HTTPException(status_code=403,
            detail=f"{out_name}은 보호된 파일입니다. 다른 프로토콜 이름을 사용하세요.")

    out_path = os.path.join(root, "common", out_name)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(s.stage3_code + "\n")

    # Optionally save to reference library
    if req.save_as_reference:
        try:
            pipeline = _import_pipeline()
            pipeline.save_to_reference(
                protocol=req.protocol_name,
                py_code=s.stage3_code,
                mapping_excel_path=s.stage2_excel_path,
                meta={"protocol": req.protocol_name},
            )
        except Exception as e:
            logger.warning(f"Reference save failed: {e}")

    logger.info(f"Saved: {out_path}")
    return {"ok": True, "path": out_path, "filename": out_name}


@router.get("/api/stage3/download-py")
async def stage3_download(session_id: str = Query(...), protocol_name: str = Query("custom")):
    """Download generated registers.py."""
    s = _get_session(session_id)
    if not s.stage3_code:
        raise HTTPException(status_code=404, detail="Stage 3 코드가 없습니다.")

    # Write to temp file if output_path not set
    if s.stage3_output_path and os.path.isfile(s.stage3_output_path):
        path = s.stage3_output_path
    else:
        path = os.path.join(s.session_dir(), f"{protocol_name}_registers.py")
        with open(path, "w", encoding="utf-8") as f:
            f.write(s.stage3_code + "\n")

    return FileResponse(
        path,
        media_type="text/x-python",
        filename=os.path.basename(path),
    )


# ---------------------------------------------------------------------------
# References
# ---------------------------------------------------------------------------

@router.get("/api/references")
async def get_references():
    """List all references (builtin + user-created)."""
    ref_mgr = _import_ref()
    mgr = ref_mgr.get_manager()
    result = []
    for name, meta in mgr._all_index.items():
        result.append({
            "name": name,
            "manufacturer": meta.get("manufacturer", ""),
            "protocol": meta.get("protocol", name),
            "description": meta.get("description", ""),
            "mppt_count": meta.get("mppt_count", 4),
            "string_count": meta.get("string_count", 8),
            "fc_code": meta.get("fc_code", "FC03"),
            "builtin": meta.get("builtin", False),
            "created": meta.get("created", ""),
        })
    return {"references": result}


class SaveReferenceRequest(BaseModel):
    session_id: str
    name: str
    manufacturer: str = ""
    description: str = ""


@router.post("/api/references")
async def save_reference(req: SaveReferenceRequest):
    """Save current Stage 3 result as a new reference."""
    s = _get_session(req.session_id)
    if not s.stage3_code:
        raise HTTPException(status_code=400, detail="Stage 3 코드가 없습니다.")

    ref_mgr = _import_ref()
    mgr = ref_mgr.get_manager()
    meta = {
        "manufacturer": req.manufacturer,
        "protocol": req.name,
        "description": req.description,
    }
    mgr.save_reference(req.name, s.stage3_code, meta, s.stage2_excel_path)
    return {"ok": True, "name": req.name}


@router.delete("/api/references/{name}")
async def delete_reference(name: str):
    """Delete a user-created reference."""
    ref_mgr = _import_ref()
    mgr = ref_mgr.get_manager()
    ok, msg = mgr.delete_reference(name)
    if not ok:
        raise HTTPException(status_code=403, detail=msg)
    return {"ok": True, "message": msg}


# ---------------------------------------------------------------------------
# AI Settings
# ---------------------------------------------------------------------------

@router.get("/api/ai-settings")
async def get_ai_settings():
    """Get current AI settings (API key masked)."""
    ai_gen = _import_ai()
    key = ai_gen.load_api_key()
    model = ai_gen.load_model_name()
    return {
        "api_key_set": bool(key),
        "api_key_preview": (key[:8] + "..." + key[-4:]) if len(key) > 12 else ("***" if key else ""),
        "model": model,
    }


class AISettingsRequest(BaseModel):
    api_key: Optional[str] = None
    model: Optional[str] = None


@router.put("/api/ai-settings")
async def update_ai_settings(req: AISettingsRequest):
    """Update AI settings."""
    ai_gen = _import_ai()
    if req.api_key is not None:
        ai_gen.save_api_key(req.api_key)
    if req.model is not None:
        ai_gen.save_model_name(req.model)
    return {"ok": True}
